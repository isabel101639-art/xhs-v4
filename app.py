#!/usr/bin/env python3
from dotenv import load_dotenv
load_dotenv()
# -*- coding: utf-8 -*-
"""
小红书任务管理系统 v4.0 - 福瑞医科
完全基于需求定制开发
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file, has_request_context
from flask_cors import CORS
from sqlalchemy import or_, text, inspect
from werkzeug.middleware.proxy_fix import ProxyFix
import os
import json
import csv
import io
import base64
import html
from types import SimpleNamespace
from datetime import datetime, timedelta
import random
import re
from collections import Counter, defaultdict
from urllib.parse import urlparse, urlunparse

from automation_hotwords import (
    build_hotword_skeleton_rows,
    build_remote_hotword_request_preview,
    fetch_remote_hotword_items,
    hotword_remote_source_preset_meta,
    hotword_remote_source_presets,
    hotword_source_template_meta,
    hotword_source_template_options,
    normalize_trend_items,
    parse_trend_payload,
    split_keywords as split_hotword_keywords,
)
from automation_accounts import (
    build_creator_sync_request_preview,
    fetch_remote_creator_bundle,
)
from automation_dashboard_routes import register_automation_dashboard_routes
from automation_asset_routes import register_automation_asset_routes
from analytics_routes import register_analytics_routes
from public_routes import register_public_routes
from automation_runtime import (
    AUTOMATION_RUNTIME_CONFIG_DEFAULTS,
    ASSET_STYLE_TYPE_DEFINITIONS,
    PRODUCT_PROFILE_DEFINITIONS,
    VOLCENGINE_MODEL_OPTIONS,
    MEDICAL_SCIENCE_LAYOUT_VARIANTS,
    KNOWLEDGE_CARD_LAYOUT_VARIANTS,
    STYLE_REFERENCE_SIGNATURES,
    _automation_runtime_config,
    _hotword_runtime_settings,
    _resolved_hotword_mode,
    _creator_sync_runtime_settings,
    _resolved_creator_sync_mode,
    _image_provider_options,
    _image_provider_presets,
    _asset_style_type_options,
    _asset_style_meta,
    _image_model_options,
    _image_provider_capabilities,
    _product_category_options,
    _product_visual_role_options,
    _product_profile_options,
    _product_profile_meta,
)
from creator_import import (
    parse_creator_import_bundle,
    parse_creator_import_file,
    preview_creator_import_bundle,
    import_creator_bundle,
)
from creator_tracking import (
    backfill_submission_tracking,
    build_registration_tracking_summary,
    canonicalize_xhs_post_url,
    normalize_tracking_url,
    sync_tracking_for_creator_account,
    sync_tracking_from_submission,
)
from copywriting_skills import (
    COPY_SKILL_OPTIONS,
    IMAGE_SKILL_OPTIONS,
    TITLE_SKILL_OPTIONS,
    build_copy_skill_local_guidance,
    build_copy_skill_prompt_block,
    build_title_skill_local_guidance,
    build_title_skill_prompt_block,
    get_image_skill_presets,
    resolve_copy_skill,
    resolve_title_skill,
)
from models import (
    db,
    Activity,
    ActivitySnapshot,
    BackupRecord,
    AdminUser,
    RolePermission,
    OperationLog,
    Topic,
    Registration,
    Submission,
    Settings,
    SiteTheme,
    SitePageConfig,
    Announcement,
    DataSourceTask,
    DataSourceLog,
    AssetGenerationTask,
    AssetPlanDraft,
    AssetLibrary,
    AutomationSchedule,
    CorpusEntry,
    HotTopicEntry,
    TrendNote,
    TopicIdea,
    CreatorAccount,
    CreatorPost,
    CreatorAccountSnapshot,
    LiverIpProfilePlan,
    PLATFORM_DEFINITIONS,
    TOPIC_IDEA_STATUS_LABELS,
    ACTIVITY_STATUS_LABELS,
    POOL_STATUS_LABELS,
    PRIMARY_PERSONAL_PLATFORMS,
)
from release_manifest import build_release_manifest_payload as _shared_build_release_manifest_payload

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _env_flag(name, default=False):
    raw = os.environ.get(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {'1', 'true', 'yes', 'y', 'on'}


def _build_release_manifest_payload():
    return _shared_build_release_manifest_payload(include_generated_at=True)


def _current_runtime_env():
    return (os.environ.get('APP_ENV') or os.environ.get('FLASK_ENV') or 'local').strip()[:50] or 'local'


def _resolve_database_url():
    database_url = (os.environ.get('DATABASE_URL') or '').strip()
    if not database_url:
        return f'sqlite:///{os.path.join(BASE_DIR, "xhs_system.db")}'
    if database_url.startswith('postgres://'):
        return database_url.replace('postgres://', 'postgresql+psycopg2://', 1)
    if database_url.startswith('postgresql://') and '+psycopg2' not in database_url:
        return database_url.replace('postgresql://', 'postgresql+psycopg2://', 1)
    return database_url


app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'xhs_furui_2026_secret_key')
app.secret_key = app.config['SECRET_KEY']
app.config['SQLALCHEMY_DATABASE_URI'] = _resolve_database_url()
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JSON_AS_ASCII'] = False
try:
    max_content_length_mb = int(os.environ.get('MAX_CONTENT_LENGTH_MB', '16'))
except ValueError:
    max_content_length_mb = 16
app.config['MAX_CONTENT_LENGTH'] = max_content_length_mb * 1024 * 1024
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = os.environ.get('SESSION_COOKIE_SAMESITE', 'Lax')
app.config['SESSION_COOKIE_SECURE'] = _env_flag('SESSION_COOKIE_SECURE', False)
app.config['PREFERRED_URL_SCHEME'] = os.environ.get('PREFERRED_URL_SCHEME', 'https')

if app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'connect_args': {'check_same_thread': False}
    }
else:
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'pool_recycle': 1800,
    }

app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

cors_origins = [item.strip() for item in (os.environ.get('CORS_ORIGINS') or '').split(',') if item.strip()]
if cors_origins:
    CORS(app, resources={r"/api/*": {"origins": cors_origins}})
else:
    CORS(app)

db.init_app(app)

LIVER_KEYWORD_SEEDS = [
    '脂肪肝', '脂肪肝逆转', '脂肪肝减脂', '肝纤维化', '肝硬化', '肝癌', '肝癌预防',
    '乙肝', '丙肝', '酒精肝', '肝功能异常', '转氨酶', '谷丙转氨酶', '谷草转氨酶',
    '肝弹', '肝硬度', 'FibroScan福波看', '健康体检', '体检', '肝病筛查', '肝脏B超', '肝结节',
    '保肝护肝', '养肝', '护肝习惯', '熬夜护肝', '饮酒护肝', '肝区不适', '肝气郁结',
    '肝郁', '肝火旺', '肝胆湿热', '肝血不足', '肝阴不足', '中医养肝', '疏肝理气',
    '情绪与肝', '焦虑失眠', '作息调理', '饮食养肝', '减脂餐', '运动减脂',
    '女性养肝', '更年期养肝', '职场久坐', '慢病管理', '复查管理',
    '肝损伤', '糖尿病', '冠心病', '减肥', '肥胖', '养肝护肝',
]

HOTWORD_SCOPE_PRESETS = [
    {
        'key': 'liver_comorbidity',
        'label': '肝病+共病问题池',
        'description': '围绕肝病本身及其相关慢病/共病，找高搜索问题和爆款笔记。',
        'keywords': [
            '肝纤维化', '肝硬化', '肝癌', '肝损伤', '肝郁', '体检', '转氨酶',
            '肝弹', '糖尿病', '冠心病', '乙肝', '丙肝', '肝功能异常',
        ],
        'preferred_trend_type': 'note_search',
        'preferred_template_key': 'xhs_note_search',
    },
    {
        'key': 'science_qna',
        'label': '科普问题池',
        'description': '围绕“怎么吃/怎么运动/怎么减肥/怎么护肝”这类高搜索问题找爆款笔记。',
        'keywords': [
            '脂肪肝怎么吃', '脂肪肝做什么运动', '如何减肥', '肥胖怎么减',
            '养肝护肝', '保肝护肝', '肝损伤吃什么', '转氨酶高怎么降',
            '脂肪肝怎么减肥', '养肝喝什么',
        ],
        'preferred_trend_type': 'note_search',
        'preferred_template_key': 'xhs_note_search',
    },
    {
        'key': 'xhs_trending',
        'label': '平台热搜/蹭热点池',
        'description': '围绕平台热搜和相关搜索词，抓适合蹭流量的热点话题。',
        'keywords': [
            '养肝护肝', '减肥', '肥胖', '体检', '中医养生', '健康养生',
        ],
        'preferred_trend_type': 'hot_queries',
        'preferred_template_key': 'xhs_hot_queries',
    },
]

HOTWORD_TIME_WINDOW_OPTIONS = [
    {'key': '3d', 'label': '近3天'},
    {'key': '7d', 'label': '近7天'},
    {'key': '30d', 'label': '近30天'},
    {'key': 'custom', 'label': '自定义'},
]

TOPIC_CONTENT_TYPES = [
    '真实经历型', '轻科普问答型', '检查解读型', '避坑清单型',
    '知识卡片型', '门诊答疑型', '复查管理型', '场景种草型',
    '中医调理型', '保肝护肝习惯型', '脂肪肝管理型', '情绪身心型',
    '饮食营养型', '运动减脂型', '家庭照护型', '观点表达型',
]

TOPIC_PERSONAS = [
    '患者本人', '家属视角', '职场应酬族', '体检人群', '门诊答疑视角',
    '病友经验', '健康管理视角', '医学助理视角', '中医调理视角',
    '营养师视角', '运动减脂教练', '职场久坐人群', '女性健康视角',
    '情绪陪伴视角', '慢病管理视角', '陪诊照护者',
]

TOPIC_VISUAL_TYPES = ['医学科普图', '知识卡片', '检查流程图', '误区对照图', '复查清单卡']

COMPLIANCE_BASELINE = (
    '避免绝对化表述、避免功效承诺、避免购买引导；涉及药品时使用“在医生指导下”'
    '“管理方案”“复查评估”这类表达，不写根治、治愈、最有效。'
)

CORPUS_SEED_ENTRIES = [
    {
        'title': '医疗健康内容合规底线',
        'category': '合规表达',
        'source': '系统初始化',
        'tags': '合规,医疗,广告法',
        'content': '涉及药品和治疗方案时，统一使用“在医生指导下”“管理方案”“复查评估”表达，避免根治、治愈、最有效、保证改善、私信购买等话术。'
    },
    {
        'title': '检查解读类爆款结构',
        'category': '爆款拆解',
        'source': '系统初始化',
        'tags': '检查解读,FibroScan,体检',
        'content': '标题先给结果冲突，首屏抛出检查数值焦虑，中段解释指标含义，尾部给下一步动作，如复查、问诊、生活方式管理。'
    },
    {
        'title': '脂肪肝知识卡常用表达',
        'category': '医学科普',
        'source': '系统初始化',
        'tags': '脂肪肝,知识卡片,科普',
        'content': '适合拆成“什么情况要重视”“体检最常见误区”“日常管理三步走”三段结构，画面要清爽、信息块短、标题用问题句。'
    },
    {
        'title': '软植入表达模板',
        'category': '产品卖点',
        'source': '系统初始化',
        'tags': '复方鳖甲软肝片,FibroScan福波看,壳脂胶囊',
        'content': '复方鳖甲软肝片适合放在抗纤维化管理语境，FibroScan福波看适合放在检查评估和复查跟踪语境，壳脂胶囊适合放在脂肪肝管理场景。'
    },
    {
        'title': '封面图常用版式',
        'category': '封面模板',
        'source': '系统初始化',
        'tags': '爆款封面,医学科普图,知识卡片',
        'content': '优先做大标题+一行结论+3个短信息点，颜色用医疗感红橙蓝，减少长段落，突出数字、问句和清单感。'
    },
]

DEFAULT_SITE_NAV_ITEMS = [
    {'label': '话题广场', 'url': '/', 'icon': 'bi-collection', 'target': '_self'},
    {'label': '我的报名', 'url': '/my_registration', 'icon': 'bi-person-check', 'target': '_self'},
    {'label': '热搜话题', 'url': '/hot-topics', 'icon': 'bi-lightning-charge', 'target': '_self'},
    {'label': '数据分析', 'url': '/data_analysis', 'icon': 'bi-bar-chart', 'target': '_self'},
    {'label': 'IP建设', 'url': '/liver-science', 'icon': 'bi-heart-pulse', 'target': '_self'},
    {'label': '自动化中心', 'url': '/automation_center', 'icon': 'bi-lightning-charge', 'target': '_self'},
    {'label': '后台管理', 'url': '/admin', 'icon': 'bi-gear', 'target': '_self'},
]

DEFAULT_SITE_THEME = {
    'theme_key': 'default',
    'name': '福瑞红橙',
    'primary_color': '#ff2442',
    'primary_soft_color': '#ffe5e8',
    'secondary_color': '#ff7a59',
    'secondary_soft_color': '#fff3e8',
    'nav_gradient_start': '#ff2442',
    'nav_gradient_end': '#ff6b6b',
    'hero_gradient_start': '#ff2442',
    'hero_gradient_end': '#ff7a59',
    'background_gradient_start': '#fff7f5',
    'background_gradient_end': '#ffffff',
    'surface_color': '#ffffff',
    'text_color': '#1f2937',
    'muted_text_color': '#6b7280',
    'footer_text': '福瑞医科小红书任务管理系统',
    'font_family': '-apple-system, BlinkMacSystemFont, PingFang SC, sans-serif',
}

DEFAULT_HOME_PAGE_CONFIG = {
    'page_key': 'home',
    'site_name': '福瑞医科',
    'page_title': '福瑞医科内容运营平台',
    'hero_badge': '当前活动期',
    'hero_title': '',
    'hero_subtitle': '',
    'announcement_title': '最新公告',
    'trend_title': '最新热点',
    'primary_section_title': '复方鳖甲软肝片话题',
    'primary_section_icon': 'bi-capsule',
    'secondary_section_title': 'FibroScan体检话题',
    'secondary_section_icon': 'bi-heart-pulse',
    'primary_topic_limit': 18,
    'footer_text': '福瑞医科小红书任务管理系统',
}

DEFAULT_ROLE_PERMISSIONS = {
    'super_admin': {
        'role_name': '超级管理员',
        'permissions': [
            'site_config.manage',
            'activity.manage',
            'topic.manage',
            'snapshot.manage',
            'backup.manage',
            'automation.manage',
            'analytics.view',
            'creator.manage',
            'logs.view',
            'admin_user.manage',
            'role_permission.manage',
            'settings.manage',
        ],
    },
    'operator': {
        'role_name': '内容运营',
        'permissions': [
            'site_config.manage',
            'activity.manage',
            'topic.manage',
            'snapshot.manage',
            'automation.manage',
            'analytics.view',
            'creator.manage',
            'logs.view',
            'settings.manage',
        ],
    },
    'reviewer': {
        'role_name': '审核人员',
        'permissions': [
            'topic.manage',
            'automation.manage',
            'logs.view',
            'analytics.view',
        ],
    },
    'viewer': {
        'role_name': '数据查看者',
        'permissions': [
            'analytics.view',
            'logs.view',
        ],
    },
}


def _default_automation_schedules(default_quota=30):
    return [
        {
            'job_key': 'hotword_sync_daily',
            'name': '热点抓取骨架巡检',
            'task_type': 'hotword_sync',
            'enabled': False,
            'interval_minutes': 360,
            'params_payload': json.dumps({
                'source_platform': '小红书',
                'source_channel': 'Scheduler骨架',
                'keyword_limit': 10,
            }, ensure_ascii=False),
        },
        {
            'job_key': 'topic_ideas_daily',
            'name': '候选话题自动生成',
            'task_type': 'topic_idea_generate',
            'enabled': False,
            'interval_minutes': 1440,
            'params_payload': json.dumps({
                'count': 80,
                'quota': default_quota,
            }, ensure_ascii=False),
        },
        {
            'job_key': 'creator_accounts_sync_half_hourly',
            'name': '报名人账号持续同步',
            'task_type': 'creator_account_sync',
            'enabled': False,
            'interval_minutes': 30,
            'params_payload': json.dumps({
                'source_platform': '小红书',
                'source_channel': 'Crawler服务',
                'batch_limit': 20,
            }, ensure_ascii=False),
        },
    ]


def _admin_json_guard():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'}), 401
    return None


def _default_topic_quota():
    env_value = (os.environ.get('DEFAULT_TOPIC_QUOTA') or '').strip()
    if env_value:
        try:
            value = int(env_value)
            if value > 0:
                return value
        except ValueError:
            pass
    try:
        setting = Settings.query.filter_by(key='default_topic_quota').first()
        if setting and str(setting.value).strip():
            value = int(str(setting.value).strip())
            if value > 0:
                return value
    except Exception:
        pass
    return 30


def _default_activity_id_for_automation():
    activity = Activity.query.filter_by(status='published').order_by(Activity.created_at.desc(), Activity.id.desc()).first()
    if activity:
        return activity.id
    fallback = Activity.query.order_by(Activity.created_at.desc(), Activity.id.desc()).first()
    return fallback.id if fallback else 0


def _automation_keyword_seeds():
    try:
        setting = Settings.query.filter_by(key='automation_keyword_seeds').first()
        parsed = _load_json_value(setting.value if setting else '', [])
        if parsed:
            return [str(item).strip() for item in parsed if str(item).strip()]
    except Exception:
        pass
    return list(LIVER_KEYWORD_SEEDS)


def _hotword_source_template_options():
    return hotword_source_template_options()


def _hotword_source_template_meta(template_key=''):
    return hotword_source_template_meta(template_key)


def _hotword_remote_source_presets():
    return hotword_remote_source_presets()


def _hotword_remote_source_preset_meta(preset_key=''):
    return hotword_remote_source_preset_meta(preset_key)


def _safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _normalize_hex_color(value, default):
    value = (value or '').strip()
    if re.fullmatch(r'#[0-9a-fA-F]{6}', value):
        return value.lower()
    return default


def _split_keywords(text):
    raw = (text or '').replace('，', ',').replace('、', ',').replace('#', ',')
    return [item.strip() for item in raw.split(',') if item and item.strip()]


def _truncate_text(text, limit=80):
    text = (text or '').strip()
    return text if len(text) <= limit else text[:limit] + '...'


def _coerce_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {'1', 'true', 'yes', 'y', 'on'}
    return False


def _normalize_quota(value, default=None, min_value=1, max_value=500):
    fallback = _default_topic_quota() if default is None else default
    quota = _safe_int(value, fallback)
    if quota < min_value:
        quota = min_value
    if quota > max_value:
        quota = max_value
    return quota


def _parse_datetime(value):
    if not value:
        return None
    value = str(value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d']:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), '%Y-%m-%d').date()
    except ValueError:
        return None


def _format_datetime(value):
    return value.strftime('%Y-%m-%d %H:%M:%S') if value else ''


def _format_datetime_local(value):
    return value.strftime('%Y-%m-%dT%H:%M') if value else ''


def _topic_idea_status_label(status):
    return TOPIC_IDEA_STATUS_LABELS.get(status or '', status or '未知')


def _activity_status_label(status):
    return ACTIVITY_STATUS_LABELS.get(status or '', status or '未知')


def _pool_status_label(status):
    return POOL_STATUS_LABELS.get(status or '', status or '未知')


def _positive_int_list(values):
    if not isinstance(values, list):
        values = []
    normalized = []
    seen = set()
    for item in values:
        value = _safe_int(item, 0)
        if value <= 0 or value in seen:
            continue
        normalized.append(value)
        seen.add(value)
    return normalized


def _current_actor():
    if not has_request_context():
        return 'system'
    return session.get('admin_username') or 'system'


def _admin_username():
    session_username = (session.get('admin_username') or '').strip()
    if session_username:
        return session_username
    return os.environ.get('ADMIN_USERNAME', 'furui')


def _admin_password():
    return os.environ.get('ADMIN_PASSWORD', 'wangdandan39')


def _admin_user_record():
    username = (session.get('admin_username') or '').strip()
    if not username:
        return None
    return AdminUser.query.filter_by(username=username).first()


def _current_permissions():
    role_key = (session.get('admin_role_key') or '').strip()
    if role_key:
        role = RolePermission.query.filter_by(role_key=role_key).first()
        if role:
            permissions = _load_json_value(role.permissions, [])
            return permissions if isinstance(permissions, list) else []
    return list(DEFAULT_ROLE_PERMISSIONS['super_admin']['permissions'])


def _log_operation(action, target_type, target_id=None, message='', detail=None):
    detail_text = ''
    if detail is not None:
        if isinstance(detail, str):
            detail_text = detail
        else:
            try:
                detail_text = json.dumps(detail, ensure_ascii=False)
            except TypeError:
                detail_text = str(detail)

    db.session.add(OperationLog(
        actor=_current_actor(),
        action=action,
        target_type=target_type,
        target_id=target_id,
        message=(message or '')[:300],
        detail=detail_text,
    ))


def _serialize_activity(activity):
    return {
        'id': activity.id,
        'name': activity.name,
        'title': activity.title,
        'description': activity.description or '',
        'status': activity.status,
        'status_label': _activity_status_label(activity.status),
        'source_type': activity.source_type or 'manual',
        'source_activity_id': activity.source_activity_id,
        'source_snapshot_id': activity.source_snapshot_id,
        'topic_count': len(activity.topics or []),
        'snapshot_count': len(activity.snapshots or []),
        'archived_at': activity.archived_at.strftime('%Y-%m-%d %H:%M:%S') if getattr(activity, 'archived_at', None) else '',
        'created_at': activity.created_at.strftime('%Y-%m-%d %H:%M:%S') if activity.created_at else '',
    }


def _load_json_value(raw_value, default):
    if not raw_value:
        return default
    try:
        parsed = json.loads(raw_value)
    except Exception:
        return default
    return parsed if isinstance(parsed, type(default)) else default


def _normalize_nav_items(items):
    desired_items = {
        '/': {'label': '话题广场', 'url': '/', 'icon': 'bi-collection', 'target': '_self'},
        '/my_registration': {'label': '我的报名', 'url': '/my_registration', 'icon': 'bi-person-check', 'target': '_self'},
        '/hot-topics': {'label': '热搜话题', 'url': '/hot-topics', 'icon': 'bi-lightning-charge', 'target': '_self'},
        '/data_analysis': {'label': '数据分析', 'url': '/data_analysis', 'icon': 'bi-bar-chart', 'target': '_self'},
        '/liver-science': {'label': 'IP建设', 'url': '/liver-science', 'icon': 'bi-heart-pulse', 'target': '_self'},
        '/automation_center': {'label': '自动化中心', 'url': '/automation_center', 'icon': 'bi-lightning-charge', 'target': '_self'},
        '/admin': {'label': '后台管理', 'url': '/admin', 'icon': 'bi-gear', 'target': '_self'},
    }
    desired_order = [
        '/',
        '/my_registration',
        '/hot-topics',
        '/data_analysis',
        '/liver-science',
        '/automation_center',
        '/admin',
    ]
    normalized = {}
    if not isinstance(items, list):
        items = []
    for raw_item in items[:12]:
        if not isinstance(raw_item, dict):
            continue
        label = (raw_item.get('label') or '').strip()
        url = (raw_item.get('url') or '').strip()
        if not label or not url:
            continue
        if url == '/activity':
            continue
        item = {
            'label': label[:20],
            'url': url[:300],
            'icon': (raw_item.get('icon') or '').strip()[:50],
            'target': '_blank' if (raw_item.get('target') or '').strip() == '_blank' else '_self',
        }
        if item['url'] == '/admin':
            item['label'] = '后台管理'
        elif item['url'] == '/my_registration':
            item['label'] = '我的报名'
        elif item['url'] == '/liver-science':
            item['label'] = 'IP建设'
        elif item['url'] == '/automation_center':
            item['label'] = '自动化中心'
        normalized[item['url']] = item
    for url in desired_order:
        if url not in normalized:
            normalized[url] = dict(desired_items[url])
    return [normalized[url] for url in desired_order]


def _serialize_site_theme(theme):
    return {
        'id': theme.id,
        'theme_key': theme.theme_key,
        'name': theme.name,
        'is_active': bool(theme.is_active),
        'primary_color': theme.primary_color,
        'primary_soft_color': theme.primary_soft_color,
        'secondary_color': theme.secondary_color,
        'secondary_soft_color': theme.secondary_soft_color,
        'nav_gradient_start': theme.nav_gradient_start,
        'nav_gradient_end': theme.nav_gradient_end,
        'hero_gradient_start': theme.hero_gradient_start,
        'hero_gradient_end': theme.hero_gradient_end,
        'background_gradient_start': theme.background_gradient_start,
        'background_gradient_end': theme.background_gradient_end,
        'surface_color': theme.surface_color,
        'text_color': theme.text_color,
        'muted_text_color': theme.muted_text_color,
        'footer_text': theme.footer_text or '',
        'font_family': theme.font_family or DEFAULT_SITE_THEME['font_family'],
        'created_at': _format_datetime(theme.created_at),
        'updated_at': _format_datetime(theme.updated_at),
    }


def _serialize_site_page_config(config):
    return {
        'id': config.id,
        'page_key': config.page_key,
        'site_name': config.site_name or '',
        'page_title': config.page_title or '',
        'hero_badge': config.hero_badge or '',
        'hero_title': config.hero_title or '',
        'hero_subtitle': config.hero_subtitle or '',
        'announcement_title': config.announcement_title or '',
        'trend_title': config.trend_title or '',
        'primary_section_title': config.primary_section_title or '',
        'primary_section_icon': config.primary_section_icon or '',
        'secondary_section_title': config.secondary_section_title or '',
        'secondary_section_icon': config.secondary_section_icon or '',
        'primary_topic_limit': config.primary_topic_limit or DEFAULT_HOME_PAGE_CONFIG['primary_topic_limit'],
        'footer_text': config.footer_text or '',
        'nav_items': _normalize_nav_items(_load_json_value(config.nav_items, [])),
        'created_at': _format_datetime(config.created_at),
        'updated_at': _format_datetime(config.updated_at),
    }


def _serialize_announcement(item):
    return {
        'id': item.id,
        'title': item.title or '',
        'content': item.content or '',
        'link_url': item.link_url or '',
        'button_text': item.button_text or '',
        'priority': item.priority or 0,
        'status': item.status or 'draft',
        'starts_at': _format_datetime(item.starts_at),
        'ends_at': _format_datetime(item.ends_at),
        'starts_at_input': _format_datetime_local(item.starts_at),
        'ends_at_input': _format_datetime_local(item.ends_at),
        'created_at': _format_datetime(item.created_at),
        'updated_at': _format_datetime(item.updated_at),
    }


def _serialize_backup_record(item):
    return {
        'id': item.id,
        'backup_type': item.backup_type or '',
        'target_type': item.target_type or '',
        'target_id': item.target_id,
        'activity_id': item.activity_id,
        'snapshot_id': item.snapshot_id,
        'status': item.status or '',
        'trigger_mode': item.trigger_mode or '',
        'backup_name': item.backup_name or '',
        'storage_path': item.storage_path or '',
        'payload': _load_json_value(item.payload, {}),
        'summary': item.summary or '',
        'restored_activity_id': item.restored_activity_id,
        'created_at': _format_datetime(item.created_at),
    }


def _serialize_admin_user(item):
    return {
        'id': item.id,
        'username': item.username,
        'display_name': item.display_name or item.username,
        'role_key': item.role_key or 'super_admin',
        'status': item.status or 'active',
        'last_login_at': _format_datetime(item.last_login_at),
        'created_at': _format_datetime(item.created_at),
        'updated_at': _format_datetime(item.updated_at),
    }


def _serialize_role_permission(item):
    return {
        'id': item.id,
        'role_key': item.role_key,
        'role_name': item.role_name,
        'permissions': _load_json_value(item.permissions, []),
        'created_at': _format_datetime(item.created_at),
        'updated_at': _format_datetime(item.updated_at),
    }


def _serialize_data_source_log(item):
    return {
        'id': item.id,
        'task_id': item.task_id,
        'level': item.level or 'info',
        'message': item.message or '',
        'detail': item.detail or '',
        'created_at': _format_datetime(item.created_at),
    }


def _serialize_data_source_task(task, detail=False):
    logs_limit = 50 if detail else 5
    logs = DataSourceLog.query.filter_by(task_id=task.id).order_by(DataSourceLog.created_at.desc(), DataSourceLog.id.desc()).limit(logs_limit).all()
    params_json = _load_json_value(task.params_payload, {})
    result_json = _load_json_value(task.result_payload, {})
    raw_result_payload = _load_json_value(task.result_payload, [])
    result_payload = _annotate_generated_asset_results(
        raw_result_payload if isinstance(raw_result_payload, list) else [],
        generation_mode=task.generation_mode or 'smart_bundle',
        cover_style_type=task.cover_style_type or '',
        inner_style_type=task.inner_style_type or '',
        style_preset=task.style_preset or '',
        provider=task.source_provider or 'svg_fallback',
        image_count=task.image_count or (len(raw_result_payload) if isinstance(raw_result_payload, list) else 1),
    )
    result_summary = _summarize_generated_asset_results(result_payload)
    return {
        'id': task.id,
        'task_type': task.task_type,
        'source_platform': task.source_platform or '',
        'source_channel': task.source_channel or '',
        'mode': task.mode or 'skeleton',
        'status': task.status or 'queued',
        'celery_task_id': task.celery_task_id or '',
        'batch_name': task.batch_name or '',
        'keyword_limit': task.keyword_limit or 0,
        'activity_id': task.activity_id,
        'item_count': task.item_count or 0,
        'message': task.message or '',
        'params_payload': task.params_payload or '',
        'params_payload_json': params_json,
        'result_payload': task.result_payload or '',
        'result_payload_json': result_json,
        'started_at': _format_datetime(task.started_at),
        'finished_at': _format_datetime(task.finished_at),
        'created_at': _format_datetime(task.created_at),
        'updated_at': _format_datetime(task.updated_at),
        'recent_logs': [_serialize_data_source_log(item) for item in logs],
    }


def _serialize_asset_generation_task(task, detail=False):
    reg = Registration.query.get(task.registration_id) if task.registration_id else None
    topic = Topic.query.get(task.topic_id) if task.topic_id else None
    product_rows = _resolve_asset_library_rows(task.product_asset_ids or '', limit=20, library_type='product')
    product_asset_ids = [item.id for item in product_rows]
    product_assets = [{
        'id': item.id,
        'title': item.title or '',
        'preview_url': item.preview_url or '',
        'library_type': item.library_type or '',
        'visual_role': item.visual_role or '',
        'product_name': item.product_name or '',
    } for item in product_rows]
    reference_rows = _resolve_reference_asset_rows(task.reference_asset_ids or '', limit=20)
    reference_ids = [item.id for item in reference_rows]
    reference_assets = [{
        'id': item.id,
        'title': item.title or '',
        'preview_url': item.preview_url or '',
        'library_type': item.library_type or '',
        'product_name': item.product_name or '',
        'visual_role': item.visual_role or '',
    } for item in reference_rows]
    return {
        'id': task.id,
        'registration_id': task.registration_id,
        'topic_id': task.topic_id,
        'draft_source_type': task.draft_source_type or '',
        'draft_source_id': task.draft_source_id,
        'draft_plan_id': task.draft_plan_id,
        'registration_name': reg.name if reg else '',
        'registration_phone': reg.phone if reg else '',
        'topic_name': topic.topic_name if topic else '',
        'source_provider': task.source_provider or 'svg_fallback',
        'model_name': task.model_name or '',
        'style_preset': task.style_preset or '小红书图文',
        'generation_mode': task.generation_mode or 'smart_bundle',
        'cover_style_type': task.cover_style_type or '',
        'cover_style_label': _asset_style_meta(task.cover_style_type or '').get('label') if (task.cover_style_type or '').strip() else '',
        'inner_style_type': task.inner_style_type or '',
        'inner_style_label': _asset_style_meta(task.inner_style_type or '').get('label') if (task.inner_style_type or '').strip() else '',
        'product_profile': task.product_profile or '',
        'product_category': task.product_category or '',
        'product_name': task.product_name or '',
        'product_indication': task.product_indication or '',
        'product_asset_ids': product_asset_ids,
        'product_assets': product_assets if detail else product_assets[:3],
        'reference_asset_ids': reference_ids,
        'reference_assets': reference_assets if detail else reference_assets[:3],
        'image_count': task.image_count or 0,
        'status': task.status or 'queued',
        'celery_task_id': task.celery_task_id or '',
        'title_hint': task.title_hint or '',
        'prompt_text': task.prompt_text or '',
        'selected_content': task.selected_content or '',
        'message': task.message or '',
        'result_payload': result_payload,
        'result_summary': result_summary,
        'started_at': _format_datetime(task.started_at),
        'finished_at': _format_datetime(task.finished_at),
        'created_at': _format_datetime(task.created_at),
        'updated_at': _format_datetime(task.updated_at),
        'selected_content_preview': _truncate_text(task.selected_content or '', 120) if not detail else task.selected_content or '',
        **(_asset_generation_quota_payload(task.registration_id) if task.registration_id else {}),
    }


def _count_real_asset_generation_attempts(registration_id):
    registration_id = _safe_int(registration_id, 0)
    if registration_id <= 0:
        return 0
    return db.session.query(db.func.count(AssetGenerationTask.id)).filter(
        AssetGenerationTask.registration_id == registration_id,
        or_(
            AssetGenerationTask.source_provider != 'svg_fallback',
            AssetGenerationTask.model_name.isnot(None) & (AssetGenerationTask.model_name != '')
        )
    ).scalar() or 0


def _asset_generation_quota_payload(registration_id):
    max_attempts = 5
    used_attempts = _count_real_asset_generation_attempts(registration_id)
    remaining_attempts = max(max_attempts - used_attempts, 0)
    return {
        'used_attempts': used_attempts,
        'remaining_attempts': remaining_attempts,
        'max_attempts': max_attempts,
    }


def _serialize_asset_plan_draft(item):
    payload = _load_json_value(item.draft_payload, {})
    plan = payload.get('plan') if isinstance(payload, dict) else {}
    preview_asset = plan.get('preview_asset') if isinstance(plan, dict) else {}
    return {
        'id': item.id,
        'source_type': item.source_type or '',
        'source_id': item.source_id,
        'source_title': item.source_title or '',
        'bucket_label': item.bucket_label or '',
        'template_agent_label': item.template_agent_label or '',
        'image_skill_label': item.image_skill_label or '',
        'style_type': item.style_type or '',
        'generation_mode': item.generation_mode or '',
        'cover_style_type': item.cover_style_type or '',
        'inner_style_type': item.inner_style_type or '',
        'title_hint': item.title_hint or '',
        'selected_content': item.selected_content or '',
        'status': item.status or 'active',
        'plan': plan if isinstance(plan, dict) else {},
        'preview_asset': preview_asset if isinstance(preview_asset, dict) else {},
        'created_at': _format_datetime(item.created_at),
        'updated_at': _format_datetime(item.updated_at),
    }


def _serialize_asset_library_item(item, detail=False):
    reg = Registration.query.get(item.registration_id) if item.registration_id else None
    topic = Topic.query.get(item.topic_id) if item.topic_id else None
    payload_json = _load_json_value(item.raw_payload, {})
    type_label_map = {
        'generated': '生成资产',
        'product': '产品图库',
        'content': '内容素材库',
        'reference': '风格参考库',
    }
    usable_score = _safe_int(payload_json.get('usable_score'), 0)
    recommended_usage = (payload_json.get('recommended_usage') or '').strip()
    usage_rank_map = {
        'cover': 3,
        'inner': 2,
        'general': 1,
    }
    return {
        'id': item.id,
        'asset_generation_task_id': item.asset_generation_task_id,
        'registration_id': item.registration_id,
        'registration_name': reg.name if reg else '',
        'topic_id': item.topic_id,
        'topic_name': topic.topic_name if topic else '',
        'library_type': item.library_type or 'generated',
        'library_type_label': type_label_map.get(item.library_type or 'generated', item.library_type or 'generated'),
        'style_type_key': item.style_type_key or '',
        'style_type_label': _asset_style_meta(item.style_type_key or '').get('label') if (item.style_type_key or '').strip() else '',
        'asset_type': item.asset_type or '',
        'title': item.title or '',
        'subtitle': item.subtitle or '',
        'source_provider': item.source_provider or 'svg_fallback',
        'model_name': item.model_name or '',
        'pool_status': item.pool_status or 'reserve',
        'pool_status_label': _pool_status_label(item.pool_status or 'reserve'),
        'status': item.status or 'active',
        'product_category': item.product_category or '',
        'product_category_label': next((row['label'] for row in _product_category_options() if row['key'] == (item.product_category or '').strip()), item.product_category or ''),
        'product_name': item.product_name or '',
        'product_indication': item.product_indication or '',
        'visual_role': item.visual_role or '',
        'tags': item.tags or '',
        'prompt_text': item.prompt_text or '',
        'preview_url': item.preview_url or '',
        'download_name': item.download_name or '',
        'usable_score': usable_score,
        'usable_label': (payload_json.get('usable_label') or '').strip(),
        'recommended_usage': recommended_usage,
        'recommended_usage_label': (payload_json.get('recommended_usage_label') or '').strip(),
        'usability_note': (payload_json.get('usability_note') or '').strip(),
        'asset_sort_rank': (
            (100000 if (item.library_type or '') == 'generated' else 0)
            + (usable_score * 10)
            + usage_rank_map.get(recommended_usage, 0)
        ),
        'raw_payload': payload_json if detail else {},
        'created_at': _format_datetime(item.created_at),
        'updated_at': _format_datetime(item.updated_at),
    }


def _serialize_automation_schedule(item):
    return {
        'id': item.id,
        'job_key': item.job_key,
        'name': item.name,
        'task_type': item.task_type,
        'enabled': bool(item.enabled),
        'interval_minutes': item.interval_minutes or 0,
        'params_payload': _load_json_value(item.params_payload, {}),
        'next_run_at': _format_datetime(item.next_run_at),
        'last_run_at': _format_datetime(item.last_run_at),
        'last_status': item.last_status or 'idle',
        'last_message': item.last_message or '',
        'last_celery_task_id': item.last_celery_task_id or '',
        'created_at': _format_datetime(item.created_at),
        'updated_at': _format_datetime(item.updated_at),
    }


def _serialize_operation_log(log):
    return {
        'id': log.id,
        'actor': log.actor or 'system',
        'action': log.action,
        'target_type': log.target_type,
        'target_id': log.target_id,
        'message': log.message or '',
        'detail': log.detail or '',
        'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
    }


def _deserialize_operation_detail(detail):
    if not detail:
        return {}
    try:
        parsed = json.loads(detail)
        return parsed if isinstance(parsed, dict) else {'raw': parsed}
    except Exception:
        return {'raw': detail}


def _parse_int_list(value, limit=20):
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = re.split(r'[\s,，;；]+', str(value or ''))
    results = []
    for item in raw_items:
        number = _safe_int(item, 0)
        if number > 0 and number not in results:
            results.append(number)
        if len(results) >= limit:
            break
    return results


def _resolve_asset_library_rows(asset_ids, limit=20, library_type=''):
    normalized_ids = _parse_int_list(asset_ids, limit=limit)
    if not normalized_ids:
        return []
    query = AssetLibrary.query.filter(AssetLibrary.id.in_(normalized_ids))
    if library_type:
        query = query.filter_by(library_type=library_type)
    rows = query.all()
    row_map = {item.id: item for item in rows}
    ordered = []
    for asset_id in normalized_ids:
        item = row_map.get(asset_id)
        if item:
            ordered.append(item)
    return ordered


def _resolve_reference_asset_rows(reference_ids, limit=20):
    return _resolve_asset_library_rows(reference_ids, limit=limit)


INTEGRATION_PING_META = {
    'hotword': {
        'label': '热点接口',
        'success_action': 'hotword_ping_check',
        'failed_action': 'hotword_ping_check_failed',
    },
    'creator_sync': {
        'label': '账号同步 crawler',
        'success_action': 'creator_sync_ping_check',
        'failed_action': 'creator_sync_ping_check_failed',
    },
    'image_provider': {
        'label': '图片接口',
        'success_action': 'image_provider_ping_check',
        'failed_action': 'image_provider_ping_check_failed',
    },
    'copywriter': {
        'label': '文案模型',
        'success_action': 'copywriter_ping_check',
        'failed_action': 'copywriter_ping_check_failed',
    },
}


def _compact_preview_payload(value, max_items=3, max_chars=600):
    if value is None:
        return None
    if isinstance(value, dict):
        items = list(value.items())[:max_items]
        return {key: _compact_preview_payload(item, max_items=max_items, max_chars=max_chars) for key, item in items}
    if isinstance(value, list):
        return [_compact_preview_payload(item, max_items=max_items, max_chars=max_chars) for item in value[:max_items]]
    if isinstance(value, str):
        return value[:max_chars]
    return value


def _log_integration_ping_result(integration_key, result, request_payload=None):
    meta = INTEGRATION_PING_META.get(integration_key, {})
    ok = bool(result.get('ok'))
    action = meta.get('success_action') if ok else meta.get('failed_action')
    if not action:
        return
    response_preview = result.get('normalized_preview')
    if not response_preview:
        response_preview = result.get('response')
    detail = {
        'integration_key': integration_key,
        'label': meta.get('label') or integration_key,
        'status': 'success' if ok else 'failed',
        'checked_at': _format_datetime(datetime.now()),
        'ok': ok,
        'message': result.get('message') or '',
        'status_code': result.get('status_code') or 0,
        'health_url': result.get('health_url') or '',
        'provider': result.get('provider') or '',
        'request_preview': _compact_preview_payload(result.get('request_preview') or request_payload or {}),
        'response_preview': _compact_preview_payload(response_preview),
    }
    _log_operation(
        action,
        'integration_ping',
        message=f'{meta.get("label") or integration_key}联调{"成功" if ok else "失败"}',
        detail=detail,
    )


def _build_integration_ping_history_payload(limit=12):
    safe_limit = min(max(_safe_int(limit, 12), 1), 50)
    action_names = []
    for meta in INTEGRATION_PING_META.values():
        action_names.extend([meta['success_action'], meta['failed_action']])
    logs = OperationLog.query.filter(
        OperationLog.action.in_(action_names)
    ).order_by(OperationLog.created_at.desc(), OperationLog.id.desc()).limit(safe_limit).all()

    items = []
    latest_by_key = {}
    for log in logs:
        detail = _deserialize_operation_detail(log.detail)
        integration_key = detail.get('integration_key') or ''
        status = (detail.get('status') or ('success' if log.action.endswith('_check') else 'failed')).strip() or 'failed'
        item = {
            'id': log.id,
            'integration_key': integration_key,
            'label': detail.get('label') or integration_key or log.target_type,
            'status': status,
            'status_label': '成功' if status == 'success' else '失败',
            'message': detail.get('message') or log.message or '',
            'checked_at': detail.get('checked_at') or _format_datetime(log.created_at),
            'created_at': _format_datetime(log.created_at),
            'actor': log.actor or 'system',
            'status_code': detail.get('status_code') or 0,
            'health_url': detail.get('health_url') or '',
            'provider': detail.get('provider') or '',
            'request_preview': detail.get('request_preview') or {},
            'response_preview': detail.get('response_preview'),
        }
        items.append(item)
        if integration_key and integration_key not in latest_by_key:
            latest_by_key[integration_key] = {
                'integration_key': integration_key,
                'label': item['label'],
                'status': item['status'],
                'status_label': item['status_label'],
                'message': item['message'],
                'checked_at': item['checked_at'],
            }

    summary = {
        'count': len(items),
        'latest_by_key': [latest_by_key[key] for key in ['hotword', 'creator_sync', 'copywriter', 'image_provider'] if key in latest_by_key],
        'success_count': len([item for item in items if item['status'] == 'success']),
        'failed_count': len([item for item in items if item['status'] != 'success']),
    }
    return {
        'success': True,
        'summary': summary,
        'items': items,
    }


def _build_first_run_playbooks_payload():
    hotword_settings = _hotword_runtime_settings()
    creator_sync_settings = _creator_sync_runtime_settings()
    hotword_mode = _resolved_hotword_mode(hotword_settings)
    creator_sync_mode = _resolved_creator_sync_mode(creator_sync_settings)
    image_capabilities = _image_provider_capabilities()
    copywriter_capabilities = _resolve_copywriter_capabilities()
    integration_history = _build_integration_ping_history_payload(limit=30)
    latest_ping_map = {
        item.get('integration_key'): item
        for item in (integration_history.get('summary', {}).get('latest_by_key') or [])
    }

    def step_item(label, status, hint, evidence=''):
        return {
            'label': label,
            'status': status,
            'hint': hint,
            'evidence': evidence,
        }

    def latest_task_status(task):
        if not task:
            return ''
        return (task.status or '').strip() or ''

    hotword_task = DataSourceTask.query.filter_by(task_type='hotword_sync').order_by(
        DataSourceTask.finished_at.desc(), DataSourceTask.updated_at.desc(), DataSourceTask.created_at.desc(), DataSourceTask.id.desc()
    ).first()
    creator_sync_task = DataSourceTask.query.filter_by(task_type='creator_account_sync').order_by(
        DataSourceTask.finished_at.desc(), DataSourceTask.updated_at.desc(), DataSourceTask.created_at.desc(), DataSourceTask.id.desc()
    ).first()
    asset_task = AssetGenerationTask.query.order_by(
        AssetGenerationTask.finished_at.desc(), AssetGenerationTask.updated_at.desc(), AssetGenerationTask.created_at.desc(), AssetGenerationTask.id.desc()
    ).first()

    hotword_ping = latest_ping_map.get('hotword', {})
    creator_ping = latest_ping_map.get('creator_sync', {})
    copywriter_ping = latest_ping_map.get('copywriter', {})
    image_ping = latest_ping_map.get('image_provider', {})
    crawler_probe = _build_crawler_probe_payload()
    crawler_probe_items = {item.get('key'): item for item in (crawler_probe.get('items') or [])}
    bundle_probe = crawler_probe_items.get('bundle_probe') or {}
    login_verify_probe = crawler_probe_items.get('login_verify') or {}
    trend_probe = crawler_probe_items.get('trend_probe_note_search') or {}
    account_probe = crawler_probe_items.get('account_probe') or {}

    hotword_steps = [
        step_item(
            '切到 remote 并填写热点接口 URL',
            'ready' if hotword_mode == 'remote' and bool((hotword_settings.get('hotword_api_url') or '').strip()) else 'blocked',
            '先在自动化中心把热点抓取模式切到 remote，并填好第三方热点接口地址。',
            f"当前模式：{hotword_mode} ｜ URL：{hotword_settings.get('hotword_api_url') or '-'}",
        ),
        step_item(
            '补鉴权头、结果路径和样例 JSON',
            'ready' if hotword_ping.get('status') == 'success' or bool((hotword_settings.get('hotword_api_headers_json') or '').strip()) or bool((hotword_settings.get('hotword_result_path') or '').strip()) else 'pending',
            '如果第三方接口需要会员鉴权或返回层级较深，这一步要一起补齐。',
            f"Headers：{'已配' if (hotword_settings.get('hotword_api_headers_json') or '').strip() else '未配'} ｜ ResultPath：{hotword_settings.get('hotword_result_path') or '-'}",
        ),
        step_item(
            '点击“测试热点接口”直到成功',
            'ready' if hotword_ping.get('status') == 'success' else ('blocked' if hotword_ping else 'pending'),
            '通过标准：接口状态成功，且能看到归一化样例。',
            hotword_ping.get('message') or '还没有热点接口联调记录',
        ),
        step_item(
            '执行首轮热点抓取',
            'ready' if latest_task_status(hotword_task) == 'success' and (hotword_task.item_count or 0) > 0 else ('blocked' if hotword_task else 'pending'),
            '点击“异步抓取热点”，让系统真正落一批 TrendNote。',
            f"最近任务：{latest_task_status(hotword_task) or '无'} ｜ 条数：{hotword_task.item_count if hotword_task else 0}",
        ),
        step_item(
            '验证候选话题自动生成',
            'ready' if TopicIdea.query.count() > 0 else 'pending',
            '如果开启了“抓完热点后自动生题”，这里应该能看到候选话题增长。',
            f"当前候选话题数：{TopicIdea.query.count()}",
        ),
    ]

    creator_steps = [
        step_item(
            '切到 remote 并填写 crawler API URL',
            'ready' if creator_sync_mode == 'remote' and bool((creator_sync_settings.get('creator_sync_api_url') or '').strip()) else 'blocked',
            '先把账号同步模式切到 remote，并填好 crawler 或第三方会员接口地址。',
            f"当前模式：{creator_sync_mode} ｜ URL：{creator_sync_settings.get('creator_sync_api_url') or '-'}",
        ),
        step_item(
            '补登录态 / 会员凭据',
            'ready' if bool((os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH') or '').strip()) or creator_ping.get('status') == 'success' else 'pending',
            '如果是真实抓取，需要 Playwright 登录态或第三方会员凭据。',
            f"PLAYWRIGHT_STORAGE_STATE_PATH：{os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH') or '-'}",
        ),
        step_item(
            '点击“测试 crawler 接口”直到成功',
            'ready' if creator_ping.get('status') == 'success' else ('blocked' if creator_ping else 'pending'),
            '通过标准：/healthz 正常，接口返回账号和笔记结构可解析。',
            creator_ping.get('message') or '还没有 crawler 联调记录',
        ),
        step_item(
            '执行首轮账号同步',
            'ready' if latest_task_status(creator_sync_task) == 'success' and (creator_sync_task.item_count or 0) > 0 else ('blocked' if creator_sync_task else 'pending'),
            '点击“异步同步账号”，让系统真正导入账号和笔记数据。',
            f"最近任务：{latest_task_status(creator_sync_task) or '无'} ｜ 条数：{creator_sync_task.item_count if creator_sync_task else 0}",
        ),
        step_item(
            '验证后续笔记与互动更新',
            'ready' if CreatorAccount.query.count() > 0 and CreatorPost.query.count() > 0 else 'pending',
            '通过标准：账号看板里能看到账号、笔记和互动数据持续累计。',
            f"账号数：{CreatorAccount.query.count()} ｜ 笔记数：{CreatorPost.query.count()}",
        ),
    ]

    image_steps = [
        step_item(
            '选择图片 Provider 并填 API / 模型',
            'ready' if bool(image_capabilities.get('image_provider_configured')) and (image_capabilities.get('image_provider_name') or '') != 'svg_fallback' else 'blocked',
            '先套用火山或兼容接口预设，再补 API URL / Base、模型名和 Key。',
            f"Provider：{image_capabilities.get('image_provider_name') or 'svg_fallback'} ｜ Model：{image_capabilities.get('image_provider_model') or '-'}",
        ),
        step_item(
            '点击“测试图片接口”直到成功',
            'ready' if image_ping.get('status') == 'success' else ('blocked' if image_ping else 'pending'),
            '通过标准：接口状态成功，能看到请求预览和返回样例。',
            image_ping.get('message') or '还没有图片接口联调记录',
        ),
        step_item(
            '执行首轮图片生成任务',
            'ready' if latest_task_status(asset_task) == 'success' and AssetLibrary.query.count() > 0 else ('blocked' if asset_task else 'pending'),
            '等接口联通后，跑一轮真实图片生成任务，把结果落进素材库。',
            f"最近任务：{latest_task_status(asset_task) or '无'} ｜ 素材库：{AssetLibrary.query.count()}",
        ),
        step_item(
            '验证素材库回流和可下载',
            'ready' if AssetLibrary.query.count() > 0 else 'pending',
            '通过标准：素材库能看到预览、来源 provider 和下载名。',
            f"当前素材数：{AssetLibrary.query.count()}",
        ),
    ]

    copywriter_steps = [
        step_item(
            '填写文案模型 API URL / 模型名',
            'ready' if bool(copywriter_capabilities.get('copywriter_api_url')) and bool(copywriter_capabilities.get('copywriter_model')) else 'blocked',
            '先在自动化中心填写文案模型 API URL 和模型名，支持 DeepSeek 或其他 OpenAI 兼容模型。',
            f"Provider：{copywriter_capabilities.get('copywriter_provider') or '-'} ｜ 模型：{copywriter_capabilities.get('copywriter_model') or '-'}",
        ),
        step_item(
            '确认文案模型 Key 已配置',
            'ready' if bool(copywriter_capabilities.get('api_key_configured')) else 'blocked',
            '文案模型调用发生在服务端，用户不需要翻墙，但服务器需要能访问模型接口。',
            f"Key：{'已配置' if copywriter_capabilities.get('api_key_configured') else '未配置'}",
        ),
        step_item(
            '点击“测试文案模型”直到成功',
            'ready' if copywriter_ping.get('status') == 'success' else ('blocked' if copywriter_ping else 'pending'),
            '通过标准：接口状态成功，并能返回一段真实的口语化测试文案。',
            copywriter_ping.get('message') or '还没有文案模型联调记录',
        ),
    ]

    crawler_steps = [
        step_item(
            '保存 Playwright 登录态',
            'ready' if bool((os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH') or '').strip()) else 'blocked',
            '先运行 save_xhs_storage_state.py，把登录态保存到 PLAYWRIGHT_STORAGE_STATE_PATH。',
            f"登录态文件：{os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH') or '-'}",
        ),
        step_item(
            '验证登录态是否有效',
            'ready' if login_verify_probe.get('status') == 'ready' else ('blocked' if login_verify_probe.get('exists') else 'pending'),
            '运行 verify_xhs_login_state.py，确认首页/搜索页不再出现登录提示。',
            login_verify_probe.get('summary') or '还没有登录态验证结果',
        ),
        step_item(
            '执行热点探测脚本',
            'ready' if trend_probe.get('status') == 'ready' else ('blocked' if trend_probe.get('exists') else 'pending'),
            '运行 probe_xhs_trends.py，确认热点/爆款内容能抓到，至少阅读量和热度值有返回。',
            trend_probe.get('summary') or '还没有热点探测结果',
        ),
        step_item(
            '执行账号探测脚本',
            'ready' if account_probe.get('status') == 'ready' else ('blocked' if account_probe.get('exists') else 'pending'),
            '运行 probe_xhs_account_posts.py，确认账号主页、笔记列表和指标抓取正常。',
            account_probe.get('summary') or '还没有账号探测结果',
        ),
        step_item(
            '查看整包联调结论',
            'ready' if bundle_probe.get('status') == 'ready' else ('blocked' if bundle_probe.get('exists') else 'pending'),
            '最后运行 probe_xhs_bundle.py，并在后台诊断页确认联调结论卡为 ready。',
            bundle_probe.get('summary') or crawler_probe.get('summary', {}).get('message') or '还没有整包联调结果',
        ),
    ]

    playbooks = [
        {
            'key': 'hotword',
            'label': '热点抓取首轮运行卡',
            'description': '把第三方热点接口接进来，并跑通“热点 -> 候选话题”闭环。',
            'action_label': '测试热点接口',
            'action_key': 'hotword',
            'steps': hotword_steps,
        },
        {
            'key': 'creator_sync',
            'label': '账号同步首轮运行卡',
            'description': '把报名人账号后续内容和互动数据持续同步回系统。',
            'action_label': '测试 crawler 接口',
            'action_key': 'creator_sync',
            'steps': creator_steps,
        },
        {
            'key': 'image_provider',
            'label': '图片接口首轮运行卡',
            'description': '把图片中心从 SVG fallback 升级为真实图片服务。',
            'action_label': '测试图片接口',
            'action_key': 'image_provider',
            'steps': image_steps,
        },
        {
            'key': 'copywriter',
            'label': '文案模型首轮运行卡',
            'description': '把文案生成从本地兜底升级为真实模型，并启用规划 Agent -> 写作 Agent -> 去模板味改写。',
            'action_label': '测试文案模型',
            'action_key': 'copywriter',
            'steps': copywriter_steps,
        },
        {
            'key': 'crawler_probe',
            'label': 'Playwright 真实联调运行卡',
            'description': '把登录态验证、热点探测、账号探测和整包联调结果串成一条完整检查链。',
            'action_label': '刷新诊断',
            'action_key': 'crawler_probe',
            'steps': crawler_steps,
        },
    ]

    summary = {
        'total_playbooks': len(playbooks),
        'ready_steps': sum(1 for playbook in playbooks for step in playbook['steps'] if step['status'] == 'ready'),
        'blocked_steps': sum(1 for playbook in playbooks for step in playbook['steps'] if step['status'] == 'blocked'),
        'pending_steps': sum(1 for playbook in playbooks for step in playbook['steps'] if step['status'] == 'pending'),
    }
    return {
        'success': True,
        'summary': summary,
        'items': playbooks,
    }


def _sample_creator_sync_targets(limit=2):
    targets = _tracked_creator_sync_targets(limit=limit)
    if targets:
        return targets
    return [{
        'registration_id': 1,
        'submission_id': 1,
        'topic_id': 1,
        'creator_account_id': 0,
        'profile_url': 'https://www.xiaohongshu.com/user/profile/demo_profile',
        'account_handle': 'demo_account',
        'owner_name': '测试账号',
        'owner_phone': '13800000000',
        'last_synced_at': _format_datetime(datetime.now()),
        'note_url': 'https://www.xiaohongshu.com/explore/demo_note',
    }]


def _build_integration_contract_payload():
    hotword_settings = _hotword_runtime_settings()
    creator_sync_settings = _creator_sync_runtime_settings()
    image_capabilities = _image_provider_capabilities()

    hotword_keywords = ['脂肪肝', '肝硬化']
    hotword_request_preview = _build_hotword_remote_preview(
        hotword_settings,
        keywords=hotword_keywords,
        source_platform=hotword_settings.get('hotword_source_platform') or '小红书',
        source_channel=hotword_settings.get('hotword_source_channel') or '会员接口',
        batch_name='contract_demo_hotword',
    )
    creator_targets = _sample_creator_sync_targets(limit=2)
    creator_request_preview = _build_creator_sync_remote_preview(
        creator_sync_settings,
        targets=creator_targets,
        source_channel=creator_sync_settings.get('creator_sync_source_channel') or 'Crawler服务',
        batch_name='contract_demo_creator_sync',
    )
    image_provider = (image_capabilities.get('image_provider_name') or 'generic_json').strip() or 'generic_json'
    if image_provider == 'svg_fallback':
        image_provider = 'generic_json'
    image_request_preview = _build_asset_provider_request_preview(
        image_provider,
        image_capabilities.get('image_provider_model') or 'demo-image-model',
        '生成一张适合小红书医疗科普封面的测试图片，画面干净，标题区留白。',
        image_capabilities.get('image_provider_size') or '1024x1536',
        image_capabilities.get('image_default_style_type') or 'medical_science',
        image_count=1,
    )

    contracts = [
        {
            'key': 'hotword',
            'label': '热点接口合同样例',
            'description': '给第三方热点会员/API 服务商时，要求其至少返回可解析的热点条目列表。',
            'request_preview': hotword_request_preview,
            'response_example': {
                'items': [{
                    'keyword': '脂肪肝',
                    'title': '脂肪肝人群该如何判断风险升级',
                    'link': 'https://example.com/hot/1',
                    'views': 12800,
                    'likes': 356,
                    'favorites': 102,
                    'comments': 41,
                    'author': '示例账号',
                    'summary': '讨论度持续上升，适合延展成医学科普话题。',
                }]
            },
            'required_fields': ['keyword', 'title', 'link', 'views', 'likes', 'favorites', 'comments'],
            'acceptance': [
                '测试接口时能返回 1 条以上热点数据',
                '标题、链接、互动量字段能够被系统识别',
                '点击“测试热点接口”后能看到归一化样例',
            ],
        },
        {
            'key': 'creator_sync',
            'label': '账号同步接口合同样例',
            'description': '给 crawler 或第三方会员服务商时，要求其返回账号、笔记、快照三类数据。',
            'request_preview': creator_request_preview,
            'response_example': {
                'accounts': [{
                    'platform': 'xhs',
                    'account_handle': 'demo_account',
                    'display_name': '测试账号',
                    'profile_url': 'https://www.xiaohongshu.com/user/profile/demo_profile',
                    'follower_count': 1234,
                }],
                'posts': [{
                    'platform': 'xhs',
                    'account_handle': 'demo_account',
                    'profile_url': 'https://www.xiaohongshu.com/user/profile/demo_profile',
                    'post_url': 'https://www.xiaohongshu.com/explore/demo_note',
                    'title': '测试账号新发的一条笔记',
                    'publish_time': '2026-04-15 12:30:00',
                    'views': 980,
                    'likes': 66,
                    'favorites': 18,
                    'comments': 9,
                }],
                'snapshots': [{
                    'platform': 'xhs',
                    'account_handle': 'demo_account',
                    'profile_url': 'https://www.xiaohongshu.com/user/profile/demo_profile',
                    'snapshot_date': '2026-04-15',
                    'follower_count': 1234,
                    'post_count': 18,
                    'total_views': 9800,
                    'total_interactions': 1300,
                }],
            },
            'required_fields': ['accounts', 'posts', 'snapshots', 'post_url', 'profile_url', 'views', 'likes', 'favorites', 'comments'],
            'acceptance': [
                '测试接口时 /healthz 可访问',
                '点击“测试 crawler 接口”后接口状态成功',
                '执行首轮账号同步后，系统里能看到账号和笔记入库',
            ],
        },
        {
            'key': 'image_provider',
            'label': '图片接口合同样例',
            'description': '给火山引擎或其他图片服务商时，要求其至少兼容标准图片生成请求和 URL/Base64 返回。',
            'request_preview': image_request_preview,
            'response_example': {
                'data': [{
                    'url': 'https://example.com/generated/demo-image-1.png',
                }]
            },
            'required_fields': ['model', 'prompt', 'size'],
            'acceptance': [
                '点击“测试图片接口”后接口状态成功',
                '返回结果里至少包含一张图片 URL 或 base64',
                '执行正式图片任务后，素材库能看到预览和下载项',
            ],
        },
    ]

    return {
        'success': True,
        'summary': {
            'count': len(contracts),
            'labels': [item['label'] for item in contracts],
        },
        'items': contracts,
    }


def _latest_data_source_task(task_type):
    return DataSourceTask.query.filter_by(task_type=task_type).order_by(
        DataSourceTask.finished_at.desc(), DataSourceTask.updated_at.desc(), DataSourceTask.created_at.desc(), DataSourceTask.id.desc()
    ).first()


def _latest_asset_generation_task():
    return AssetGenerationTask.query.order_by(
        AssetGenerationTask.finished_at.desc(), AssetGenerationTask.updated_at.desc(), AssetGenerationTask.created_at.desc(), AssetGenerationTask.id.desc()
    ).first()


def _build_integration_acceptance_payload():
    history = _build_integration_ping_history_payload(limit=30)
    latest_ping_map = {
        item.get('integration_key'): item
        for item in (history.get('summary', {}).get('latest_by_key') or [])
    }
    hotword_settings = _hotword_runtime_settings()
    creator_sync_settings = _creator_sync_runtime_settings()
    hotword_mode = _resolved_hotword_mode(hotword_settings)
    creator_sync_mode = _resolved_creator_sync_mode(creator_sync_settings)
    image_capabilities = _image_provider_capabilities()
    copywriter_capabilities = _resolve_copywriter_capabilities()

    hotword_task = _latest_data_source_task('hotword_sync')
    creator_sync_task = _latest_data_source_task('creator_account_sync')
    asset_task = _latest_asset_generation_task()

    def acceptance_item(key, label, status, message, evidence=None, next_action=''):
        return {
            'key': key,
            'label': label,
            'status': status,
            'status_label': {
                'ready': '已通过',
                'blocked': '未通过',
                'pending': '待执行',
                'pending_external': '待外部条件',
            }.get(status, status),
            'message': message,
            'evidence': evidence or [],
            'next_action': next_action,
            'ok': status == 'ready',
        }

    items = []

    hotword_evidence = [
        f"模式：{hotword_mode}",
        f"最近联调：{(latest_ping_map.get('hotword') or {}).get('status_label', '无')}",
        f"最近抓取任务：{(hotword_task.status if hotword_task else '无')}",
        f"热点池条数：{TrendNote.query.count()}",
        f"候选话题数：{TopicIdea.query.count()}",
    ]
    hotword_ping_ok = (latest_ping_map.get('hotword') or {}).get('status') == 'success'
    hotword_task_ok = bool(hotword_task and hotword_task.status == 'success' and (hotword_task.item_count or 0) > 0)
    hotword_data_ok = TrendNote.query.count() > 0
    if hotword_mode != 'remote':
        items.append(acceptance_item(
            'hotword',
            '热点抓取验收',
            'pending_external',
            '当前还没切到真实热点接口，无法做正式验收。',
            evidence=hotword_evidence,
            next_action='先提供第三方热点接口并切到 remote，再跑测试热点接口和首轮抓取。',
        ))
    elif hotword_ping_ok and hotword_task_ok and hotword_data_ok:
        items.append(acceptance_item(
            'hotword',
            '热点抓取验收',
            'ready',
            '热点接口已联通，且首轮抓取与入库已通过。',
            evidence=hotword_evidence,
            next_action='可以继续验证自动生题和审核发布闭环。',
        ))
    elif hotword_ping_ok:
        items.append(acceptance_item(
            'hotword',
            '热点抓取验收',
            'pending',
            '热点接口已经测通，但还没完成正式抓取入库验收。',
            evidence=hotword_evidence,
            next_action='点击“异步抓取热点”，确认 TrendNote 条数增长。',
        ))
    else:
        items.append(acceptance_item(
            'hotword',
            '热点抓取验收',
            'blocked',
            '热点接口还没联通成功，正式验收尚未通过。',
            evidence=hotword_evidence,
            next_action='先反复测试热点接口，直到联调记录显示成功。',
        ))

    creator_evidence = [
        f"模式：{creator_sync_mode}",
        f"最近联调：{(latest_ping_map.get('creator_sync') or {}).get('status_label', '无')}",
        f"最近同步任务：{(creator_sync_task.status if creator_sync_task else '无')}",
        f"账号数：{CreatorAccount.query.count()}",
        f"笔记数：{CreatorPost.query.count()}",
    ]
    creator_ping_ok = (latest_ping_map.get('creator_sync') or {}).get('status') == 'success'
    creator_task_ok = bool(creator_sync_task and creator_sync_task.status == 'success' and (creator_sync_task.item_count or 0) > 0)
    creator_data_ok = CreatorAccount.query.count() > 0 and CreatorPost.query.count() > 0
    if creator_sync_mode != 'remote':
        items.append(acceptance_item(
            'creator_sync',
            '账号同步验收',
            'pending_external',
            '当前还没切到真实 crawler / 第三方账号接口，无法做正式验收。',
            evidence=creator_evidence,
            next_action='先提供 crawler 或会员接口并切到 remote，再跑测试 crawler 接口和首轮同步。',
        ))
    elif creator_ping_ok and creator_task_ok and creator_data_ok:
        items.append(acceptance_item(
            'creator_sync',
            '账号同步验收',
            'ready',
            '账号同步接口已联通，且账号和笔记数据已成功回流。',
            evidence=creator_evidence,
            next_action='可以继续验证后续新笔记累计和互动更新。',
        ))
    elif creator_ping_ok:
        items.append(acceptance_item(
            'creator_sync',
            '账号同步验收',
            'pending',
            'crawler 接口已经测通，但还没完成正式同步入库验收。',
            evidence=creator_evidence,
            next_action='点击“异步同步账号”，确认账号和笔记条数增长。',
        ))
    else:
        items.append(acceptance_item(
            'creator_sync',
            '账号同步验收',
            'blocked',
            '账号同步接口还没联通成功，正式验收尚未通过。',
            evidence=creator_evidence,
            next_action='先把 crawler 接口和登录态调通，直到联调记录显示成功。',
        ))

    image_provider = (image_capabilities.get('image_provider_name') or 'svg_fallback').strip() or 'svg_fallback'
    image_evidence = [
        f"Provider：{image_provider}",
        f"最近联调：{(latest_ping_map.get('image_provider') or {}).get('status_label', '无')}",
        f"最近图片任务：{(asset_task.status if asset_task else '无')}",
        f"素材库条数：{AssetLibrary.query.count()}",
    ]
    image_real_enabled = image_provider != 'svg_fallback' and bool(image_capabilities.get('image_provider_configured'))
    image_ping_ok = (latest_ping_map.get('image_provider') or {}).get('status') == 'success'
    asset_task_ok = bool(asset_task and asset_task.status == 'success' and (asset_task.source_provider or '') != 'svg_fallback')
    asset_data_ok = AssetLibrary.query.filter(AssetLibrary.source_provider != 'svg_fallback').count() > 0
    if not image_real_enabled:
        items.append(acceptance_item(
            'image_provider',
            '图片接口验收',
            'pending_external',
            '当前仍是 SVG fallback，真实图片接口还没接入，无法做正式验收。',
            evidence=image_evidence,
            next_action='先提供火山引擎或其他图片 API，再用图片调试沙盒联调。',
        ))
    elif image_ping_ok and asset_task_ok and asset_data_ok:
        items.append(acceptance_item(
            'image_provider',
            '图片接口验收',
            'ready',
            '图片接口已联通，且真实图片任务和素材库回流已通过。',
            evidence=image_evidence,
            next_action='可以继续切到正式图片工作流。',
        ))
    elif image_ping_ok:
        items.append(acceptance_item(
            'image_provider',
            '图片接口验收',
            'pending',
            '图片接口已经测通，但还没完成正式图片任务验收。',
            evidence=image_evidence,
            next_action='执行一轮正式图片生成任务，确认素材库出现远端生成结果。',
        ))
    else:
        items.append(acceptance_item(
            'image_provider',
            '图片接口验收',
            'blocked',
            '图片接口还没联通成功，正式验收尚未通过。',
            evidence=image_evidence,
            next_action='先把 API URL、模型和 Key 调通，直到测试图片接口成功。',
        ))

    copywriter_evidence = [
        f"Provider：{copywriter_capabilities.get('copywriter_provider') or 'local_fallback'}",
        f"模型：{copywriter_capabilities.get('copywriter_model') or '-'}",
        f"最近联调：{(latest_ping_map.get('copywriter') or {}).get('status_label', '无')}",
    ]
    copywriter_ping_ok = (latest_ping_map.get('copywriter') or {}).get('status') == 'success'
    if not copywriter_capabilities.get('copywriter_configured'):
        items.append(acceptance_item(
            'copywriter',
            '文案模型验收',
            'pending_external',
            '当前还没有可用的文案模型，文案链路仍会回退到本地兜底。',
            evidence=copywriter_evidence,
            next_action='先配置 DeepSeek 或其他 OpenAI 兼容模型，再测试文案模型接口。',
        ))
    elif copywriter_ping_ok:
        items.append(acceptance_item(
            'copywriter',
            '文案模型验收',
            'ready',
            '文案模型已联通，可用于规划 Agent、写作 Agent 和真人化重写。',
            evidence=copywriter_evidence,
            next_action='继续用真实话题验证文案质量和 Agent 差异化表现。',
        ))
    else:
        items.append(acceptance_item(
            'copywriter',
            '文案模型验收',
            'blocked',
            '文案模型配置已存在，但当前还没联通成功。',
            evidence=copywriter_evidence,
            next_action='先在自动化中心点击“测试文案模型”，确认 API URL、模型名和 Key 是否可用。',
        ))

    summary = {
        'total': len(items),
        'ready': len([item for item in items if item['status'] == 'ready']),
        'blocked': len([item for item in items if item['status'] == 'blocked']),
        'pending': len([item for item in items if item['status'] == 'pending']),
        'pending_external': len([item for item in items if item['status'] == 'pending_external']),
    }
    summary['message'] = f"正式验收已通过 {summary['ready']}/{summary['total']} 条链路。"
    return {
        'success': True,
        'summary': summary,
        'items': items,
    }


def _build_trial_readiness_payload():
    launch_milestones = _build_launch_milestones_payload()
    integration_acceptance = _build_integration_acceptance_payload()
    capacity = _build_capacity_readiness_payload()

    milestone_map = {item.get('key'): item for item in (launch_milestones.get('items') or [])}
    acceptance_map = {item.get('key'): item for item in (integration_acceptance.get('items') or [])}

    web_item = milestone_map.get('web_foundation', {})
    async_item = milestone_map.get('async_chain', {})
    hotword_item = acceptance_map.get('hotword', {})
    creator_item = acceptance_map.get('creator_sync', {})
    image_item = acceptance_map.get('image_provider', {})
    copywriter_item = acceptance_map.get('copywriter', {})

    internal_blockers = []
    if web_item.get('status') != 'ready':
        internal_blockers.append('Web 基础服务未完全就绪')
    if async_item.get('status') == 'blocked':
        internal_blockers.append('异步执行链未补齐')
    if not capacity.get('summary', {}).get('capacity_ready'):
        internal_blockers.append('容量准备度未通过')

    acceptance_blockers = [
        item.get('label') for item in [hotword_item, creator_item, image_item, copywriter_item]
        if item.get('status') == 'blocked'
    ]
    external_pending = [
        item.get('label') for item in [hotword_item, creator_item, image_item, copywriter_item]
        if item.get('status') == 'pending_external'
    ]
    execution_pending = [
        item.get('label') for item in [hotword_item, creator_item, image_item, copywriter_item]
        if item.get('status') == 'pending'
    ]

    def phase_item(key, label, status, message, blockers=None, next_action=''):
        return {
            'key': key,
            'label': label,
            'status': status,
            'status_label': {
                'ready': '已就绪',
                'blocked': '未就绪',
                'pending_external': '待外部条件',
                'pending': '待执行',
            }.get(status, status),
            'message': message,
            'blockers': blockers or [],
            'next_action': next_action,
            'ok': status == 'ready',
        }

    phases = []

    if web_item.get('status') == 'ready':
        phases.append(phase_item(
            'demo',
            '系统演示',
            'ready',
            '当前前后台、自动化中心和核心页面已经适合对外演示。',
            next_action='可以继续演示运营流程、自动化中心和账号看板。',
        ))
    else:
        phases.append(phase_item(
            'demo',
            '系统演示',
            'blocked',
            'Web 基础服务还没完全稳定，暂不建议做正式演示。',
            blockers=web_item.get('blockers') or [web_item.get('message') or 'Web 未就绪'],
            next_action='先补齐 Web 缺口，再确认首页、后台登录和自动化中心能正常打开。',
        ))

    if internal_blockers:
        phases.append(phase_item(
            'integration',
            '接口联调准备',
            'blocked',
            '系统内部条件还没完全补齐，接口联调准备度不足。',
            blockers=internal_blockers,
            next_action='先补齐 Worker / Beat / 容量等内部条件，再开始真实接口联调。',
        ))
    elif external_pending:
        phases.append(phase_item(
            'integration',
            '接口联调准备',
            'pending_external',
            '系统内部工具链已经准备好，现在主要等第三方 API/会员接口。',
            blockers=external_pending,
            next_action='你把热点、crawler、图片 API 给我后，就可以直接按运行卡开始联调。',
        ))
    else:
        phases.append(phase_item(
            'integration',
            '接口联调准备',
            'ready',
            '系统内部和外部接口基础条件都已具备，可以连续做真实联调。',
            next_action='依次执行热点、账号同步、图片三条运行卡，并观察联调记录与验收结果。',
        ))

    if internal_blockers or acceptance_blockers:
        phases.append(phase_item(
            'pilot',
            '真实试运行',
            'blocked',
            '当前还不适合进入真实试运行。',
            blockers=internal_blockers + acceptance_blockers,
            next_action='先让三条外部链路至少完成首轮正式验收，再进入试运行。',
        ))
    elif external_pending:
        phases.append(phase_item(
            'pilot',
            '真实试运行',
            'pending_external',
            '系统主干已具备，但还缺第三方真实接口，因此暂不能进入完整试运行。',
            blockers=external_pending,
            next_action='等你把第三方接口给我后，我会按验收面板逐项通过再进入试运行。',
        ))
    elif execution_pending:
        phases.append(phase_item(
            'pilot',
            '真实试运行',
            'pending',
            '接口已经有基础条件，但还差正式抓取 / 同步 / 出图验收。',
            blockers=execution_pending,
            next_action='执行正式热点抓取、正式账号同步、正式图片任务后，再看验收面板是否全部转绿。',
        ))
    else:
        phases.append(phase_item(
            'pilot',
            '真实试运行',
            'ready',
            '当前已经具备进入真实试运行的条件。',
            next_action='可以开始小规模真实试运行，并继续观察任务队列、热点抓取和账号同步稳定性。',
        ))

    summary = {
        'overall_status': phases[-1]['status'] if phases else 'blocked',
        'overall_status_label': phases[-1]['status_label'] if phases else '未就绪',
        'message': phases[-1]['message'] if phases else '暂无判定',
        'phase_count': len(phases),
        'ready': len([item for item in phases if item['status'] == 'ready']),
        'blocked': len([item for item in phases if item['status'] == 'blocked']),
        'pending_external': len([item for item in phases if item['status'] == 'pending_external']),
        'pending': len([item for item in phases if item['status'] == 'pending']),
    }
    return {
        'success': True,
        'summary': summary,
        'items': phases,
    }


def _build_go_live_readiness_payload():
    trial = _build_trial_readiness_payload()
    acceptance = _build_integration_acceptance_payload()
    launch_milestones = _build_launch_milestones_payload()
    capacity = _build_capacity_readiness_payload()

    trial_map = {item.get('key'): item for item in (trial.get('items') or [])}
    acceptance_map = {item.get('key'): item for item in (acceptance.get('items') or [])}
    milestone_map = {item.get('key'): item for item in (launch_milestones.get('items') or [])}

    hotword_mode = _resolved_hotword_mode(_hotword_runtime_settings())
    creator_sync_mode = _resolved_creator_sync_mode(_creator_sync_runtime_settings())
    beat_enabled = _coerce_bool(os.environ.get('ENABLE_AUTOMATION_BEAT', 'true'))
    inline_jobs = _env_flag('INLINE_AUTOMATION_JOBS', False)

    schedules = {
        item.job_key: item
        for item in AutomationSchedule.query.all()
    }

    def phase_item(key, label, status, message, blockers=None, evidence=None, next_action=''):
        return {
            'key': key,
            'label': label,
            'status': status,
            'status_label': {
                'ready': '可上线',
                'blocked': '不可上线',
                'pending_external': '待外部条件',
                'pending': '待执行',
            }.get(status, status),
            'message': message,
            'blockers': blockers or [],
            'evidence': evidence or [],
            'next_action': next_action,
            'ok': status == 'ready',
        }

    foundation_blockers = []
    foundation_evidence = [
        f"Web：{(milestone_map.get('web_foundation') or {}).get('status_label', '未就绪')}",
        f"异步执行链：{(milestone_map.get('async_chain') or {}).get('status_label', '未就绪')}",
        f"容量准备度：{'通过' if capacity.get('summary', {}).get('capacity_ready') else '未通过'}",
    ]
    if (milestone_map.get('web_foundation') or {}).get('status') != 'ready':
        foundation_blockers.append('Web 基础服务未完全就绪')
    if (milestone_map.get('async_chain') or {}).get('status') != 'ready':
        foundation_blockers.append('异步执行链未完全就绪')
    if not capacity.get('summary', {}).get('capacity_ready'):
        foundation_blockers.append('容量准备度未通过')

    phases = []
    if foundation_blockers:
        phases.append(phase_item(
            'foundation',
            '基础上线条件',
            'blocked',
            '当前基础运行条件还不足以支撑正式上线。',
            blockers=foundation_blockers,
            evidence=foundation_evidence,
            next_action='先补齐 Web、Worker/Beat、容量准备度，再考虑正式上线。',
        ))
    else:
        phases.append(phase_item(
            'foundation',
            '基础上线条件',
            'ready',
            '当前基础运行层已经具备正式上线条件。',
            evidence=foundation_evidence,
            next_action='继续检查自动化调度和外部链路验收。',
        ))

    schedule_blockers = []
    schedule_evidence = [
        f"Beat 环境开关：{'启用' if beat_enabled else '关闭'}",
        f"执行模式：{'inline 本地模式' if inline_jobs else 'celery 异步模式'}",
        f"热点调度：{'启用' if bool((schedules.get('hotword_sync_daily') and schedules['hotword_sync_daily'].enabled)) else '关闭'}",
        f"账号同步调度：{'启用' if bool((schedules.get('creator_accounts_sync_half_hourly') and schedules['creator_accounts_sync_half_hourly'].enabled)) else '关闭'}",
    ]
    if inline_jobs:
        schedule_blockers.append('当前仍是 inline 本地模式')
    if beat_enabled and not schedules.get('hotword_sync_daily'):
        schedule_blockers.append('缺少默认热点调度')
    if beat_enabled and hotword_mode == 'remote' and not bool((schedules.get('hotword_sync_daily') and schedules['hotword_sync_daily'].enabled)):
        schedule_blockers.append('热点正式调度未启用')
    if beat_enabled and creator_sync_mode == 'remote' and not bool((schedules.get('creator_accounts_sync_half_hourly') and schedules['creator_accounts_sync_half_hourly'].enabled)):
        schedule_blockers.append('账号同步正式调度未启用')

    if hotword_mode != 'remote' and creator_sync_mode != 'remote':
        phases.append(phase_item(
            'automation',
            '自动化调度条件',
            'pending_external',
            '当前还没有切到真实热点或真实账号同步接口，正式调度暂不需要开启。',
            blockers=['热点和账号同步仍在本地/骨架模式'],
            evidence=schedule_evidence,
            next_action='等真实接口切到 remote 后，再打开对应正式调度。',
        ))
    elif schedule_blockers:
        phases.append(phase_item(
            'automation',
            '自动化调度条件',
            'blocked',
            '自动化调度条件还不满足正式上线要求。',
            blockers=schedule_blockers,
            evidence=schedule_evidence,
            next_action='把执行模式切到 celery，并开启热点/账号同步正式调度。',
        ))
    else:
        phases.append(phase_item(
            'automation',
            '自动化调度条件',
            'ready',
            '正式调度条件已经满足，可以按计划自动执行。',
            evidence=schedule_evidence,
            next_action='继续检查三条关键外部链路的正式验收状态。',
        ))

    acceptance_blockers = []
    acceptance_pending_external = []
    acceptance_pending = []
    acceptance_evidence = []
    for key in ['hotword', 'creator_sync', 'image_provider', 'copywriter']:
        item = acceptance_map.get(key) or {}
        acceptance_evidence.append(f"{item.get('label', key)}：{item.get('status_label', '未检查')}")
        if item.get('status') == 'blocked':
            acceptance_blockers.append(item.get('label') or key)
        elif item.get('status') == 'pending_external':
            acceptance_pending_external.append(item.get('label') or key)
        elif item.get('status') == 'pending':
            acceptance_pending.append(item.get('label') or key)

    if acceptance_blockers:
        phases.append(phase_item(
            'acceptance',
            '关键链路正式验收',
            'blocked',
            '至少有一条关键外部链路还没通过正式验收。',
            blockers=acceptance_blockers,
            evidence=acceptance_evidence,
            next_action='先让热点、账号同步、图片链路的验收面板全部转为“已通过”。',
        ))
    elif acceptance_pending_external:
        phases.append(phase_item(
            'acceptance',
            '关键链路正式验收',
            'pending_external',
            '系统内部条件已经接近完成，但还在等外部接口条件。',
            blockers=acceptance_pending_external,
            evidence=acceptance_evidence,
            next_action='你把第三方接口给我后，我会按运行卡把正式验收补完。',
        ))
    elif acceptance_pending:
        phases.append(phase_item(
            'acceptance',
            '关键链路正式验收',
            'pending',
            '关键链路已经开始联调，但还差正式任务入库 / 同步 / 出图验收。',
            blockers=acceptance_pending,
            evidence=acceptance_evidence,
            next_action='执行正式热点抓取、账号同步和图片任务，再看验收面板是否全部转绿。',
        ))
    else:
        phases.append(phase_item(
            'acceptance',
            '关键链路正式验收',
            'ready',
            '三条关键外部链路都已经通过正式验收。',
            evidence=acceptance_evidence,
            next_action='可以进入正式上线前的最后检查。',
        ))

    overall_blockers = []
    for phase in phases:
        if phase['status'] == 'blocked':
            overall_blockers.extend(phase.get('blockers') or [phase.get('message') or phase.get('label')])

    if overall_blockers:
        overall = phase_item(
            'go_live',
            '正式上线',
            'blocked',
            '当前还不适合进入正式上线运营。',
            blockers=overall_blockers,
            evidence=[item.get('status_label') for item in phases],
            next_action='先补齐所有阻塞项，尤其是调度和正式验收项，再进入正式上线。',
        )
    else:
        pending_external = any(item['status'] == 'pending_external' for item in phases)
        pending_exec = any(item['status'] == 'pending' for item in phases)
        if pending_external:
            overall = phase_item(
                'go_live',
                '正式上线',
                'pending_external',
                '内部主干已经具备，但仍在等待外部接口条件，暂不建议正式上线。',
                blockers=[item['label'] for item in phases if item['status'] == 'pending_external'],
                evidence=[item.get('status_label') for item in phases],
                next_action='等第三方接口到位并完成验收后，再进入正式上线。',
            )
        elif pending_exec:
            overall = phase_item(
                'go_live',
                '正式上线',
                'pending',
                '当前距离正式上线只差最后几项执行型验收。',
                blockers=[item['label'] for item in phases if item['status'] == 'pending'],
                evidence=[item.get('status_label') for item in phases],
                next_action='把最后待执行的正式任务跑完，再复查上线判断。',
            )
        else:
            overall = phase_item(
                'go_live',
                '正式上线',
                'ready',
                '当前已经具备正式上线运营条件。',
                evidence=[item.get('status_label') for item in phases],
                next_action='可以开始正式运营，并持续观察任务成功率和外部接口稳定性。',
            )

    summary = {
        'overall_status': overall['status'],
        'overall_status_label': overall['status_label'],
        'message': overall['message'],
        'phase_count': len(phases) + 1,
        'ready': len([item for item in phases + [overall] if item['status'] == 'ready']),
        'blocked': len([item for item in phases + [overall] if item['status'] == 'blocked']),
        'pending_external': len([item for item in phases + [overall] if item['status'] == 'pending_external']),
        'pending': len([item for item in phases + [overall] if item['status'] == 'pending']),
    }
    return {
        'success': True,
        'summary': summary,
        'items': phases + [overall],
    }


def _build_go_live_checklist_payload():
    acceptance = _build_integration_acceptance_payload()
    go_live = _build_go_live_readiness_payload()
    hotword_mode = _resolved_hotword_mode(_hotword_runtime_settings())
    creator_sync_mode = _resolved_creator_sync_mode(_creator_sync_runtime_settings())
    beat_enabled = _coerce_bool(os.environ.get('ENABLE_AUTOMATION_BEAT', 'true'))
    inline_jobs = _env_flag('INLINE_AUTOMATION_JOBS', False)
    schedules = {item.job_key: item for item in AutomationSchedule.query.all()}

    admin_password_env = (os.environ.get('ADMIN_PASSWORD') or '').strip()
    secret_key_env = (os.environ.get('SECRET_KEY') or '').strip()
    session_secure = _env_flag('SESSION_COOKIE_SECURE', False)
    preferred_scheme = (os.environ.get('PREFERRED_URL_SCHEME') or 'https').strip().lower()
    backup_count = BackupRecord.query.count()
    snapshot_count = ActivitySnapshot.query.count()

    def checklist_item(key, label, status, message, evidence='', next_action=''):
        return {
            'key': key,
            'label': label,
            'status': status,
            'status_label': {
                'ready': '已完成',
                'blocked': '未完成',
                'pending': '待执行',
                'pending_external': '待外部条件',
            }.get(status, status),
            'message': message,
            'evidence': evidence,
            'next_action': next_action,
            'ok': status == 'ready',
        }

    items = []
    security_ready = bool(secret_key_env) and bool(admin_password_env) and admin_password_env != 'wangdandan39' and preferred_scheme == 'https' and session_secure
    items.append(checklist_item(
        'security',
        '安全配置复核',
        'ready' if security_ready else 'blocked',
        '确认 SECRET_KEY、管理员密码、HTTPS 和 Cookie 安全策略都适合正式上线。',
        evidence=f"SECRET_KEY：{'已配' if secret_key_env else '未配'} ｜ ADMIN_PASSWORD：{'已配' if admin_password_env else '未配'} ｜ SESSION_COOKIE_SECURE：{'true' if session_secure else 'false'} ｜ PREFERRED_URL_SCHEME：{preferred_scheme or '-'}",
        next_action='正式上线前建议把 SESSION_COOKIE_SECURE 切到 true，并确认管理员密码不是默认值。',
    ))

    backup_ready = backup_count > 0 or snapshot_count > 0
    items.append(checklist_item(
        'backup',
        '备份 / 快照',
        'ready' if backup_ready else 'pending',
        '上线前至少保留一份活动快照或备份记录，方便异常时快速回滚。',
        evidence=f"备份数：{backup_count} ｜ 快照数：{snapshot_count}",
        next_action='如果还没有备份，先在后台创建活动快照或手动备份。',
    ))

    scheduler_required = hotword_mode == 'remote' or creator_sync_mode == 'remote'
    scheduler_ready = (not beat_enabled) or (not inline_jobs and (
        (hotword_mode != 'remote' or bool((schedules.get('hotword_sync_daily') and schedules['hotword_sync_daily'].enabled))) and
        (creator_sync_mode != 'remote' or bool((schedules.get('creator_accounts_sync_half_hourly') and schedules['creator_accounts_sync_half_hourly'].enabled)))
    ))
    items.append(checklist_item(
        'schedules',
        '正式调度复核',
        'ready' if scheduler_ready else ('pending_external' if not scheduler_required else 'blocked'),
        '确认热点巡检和账号同步的正式调度是否已经按线上策略开启。',
        evidence=f"热点模式：{hotword_mode} ｜ 账号同步模式：{creator_sync_mode} ｜ 热点调度：{'启用' if bool((schedules.get('hotword_sync_daily') and schedules['hotword_sync_daily'].enabled)) else '关闭'} ｜ 账号同步调度：{'启用' if bool((schedules.get('creator_accounts_sync_half_hourly') and schedules['creator_accounts_sync_half_hourly'].enabled)) else '关闭'}",
        next_action='如果要正式依赖自动抓取，先把 Beat 和对应调度项打开。',
    ))

    acceptance_items = acceptance.get('items') or []
    acceptance_ready = all(item.get('status') == 'ready' for item in acceptance_items)
    items.append(checklist_item(
        'acceptance',
        '关键链路正式验收',
        'ready' if acceptance_ready else ('pending_external' if any(item.get('status') == 'pending_external' for item in acceptance_items) else 'pending'),
        '确认热点、账号同步、图片三条关键链路都已经通过正式验收。',
        evidence=' ｜ '.join([f"{item.get('label')}：{item.get('status_label')}" for item in acceptance_items]) or '暂无验收数据',
        next_action='正式上线前建议至少把会用到的外部链路都转成“已通过”。',
    ))

    go_live_summary = go_live.get('summary') or {}
    items.append(checklist_item(
        'final_review',
        '最终上线复核',
        'ready' if go_live_summary.get('overall_status') == 'ready' else 'blocked',
        go_live_summary.get('message') or '暂无最终上线结论',
        evidence=f"上线判断：{go_live_summary.get('overall_status_label') or '-'}",
        next_action='最后再刷新一次“正式上线判断”，确认整体转为“可上线”再开始正式运营。',
    ))

    summary = {
        'count': len(items),
        'ready': len([item for item in items if item['status'] == 'ready']),
        'blocked': len([item for item in items if item['status'] == 'blocked']),
        'pending': len([item for item in items if item['status'] == 'pending']),
        'pending_external': len([item for item in items if item['status'] == 'pending_external']),
    }
    summary['message'] = f"上线前最后动作已完成 {summary['ready']}/{summary['count']} 项。"
    return {
        'success': True,
        'summary': summary,
        'items': items,
    }


def _build_post_launch_watchlist_payload():
    failed_jobs = _build_recent_failed_jobs_payload(limit=20)
    last_worker_ping = _latest_worker_ping_snapshot()
    hotword_task = _latest_data_source_task('hotword_sync')
    creator_sync_task = _latest_data_source_task('creator_account_sync')
    asset_task = _latest_asset_generation_task()

    def watch_item(key, label, status, message, evidence=None, threshold=''):
        return {
            'key': key,
            'label': label,
            'status': status,
            'status_label': {
                'ready': '正常',
                'blocked': '关注',
                'pending': '待观察',
            }.get(status, status),
            'message': message,
            'evidence': evidence or [],
            'threshold': threshold,
            'ok': status == 'ready',
        }

    failed_count = failed_jobs.get('summary', {}).get('count', 0)
    failed_messages = [item.get('message') or item.get('kind_label') for item in (failed_jobs.get('items') or [])[:3]]
    items = [
        watch_item(
            'worker',
            'Worker / 异步任务',
            'ready' if last_worker_ping.get('status') == 'success' and failed_count == 0 else ('pending' if not last_worker_ping.get('has_result') else 'blocked'),
            '上线后优先看 Worker Ping 和失败任务是否持续稳定。',
            evidence=[
                f"最近 Worker Ping：{last_worker_ping.get('status_label') or '未检查'}",
                f"最近失败任务数：{failed_count}",
                *failed_messages,
            ],
            threshold='建议失败任务持续为 0，或出现后能在当班内清零。',
        ),
        watch_item(
            'hotword',
            '热点抓取稳定性',
            'ready' if hotword_task and hotword_task.status == 'success' else ('pending' if not hotword_task else 'blocked'),
            '观察热点抓取是否按调度稳定入库，是否出现长时间断更。',
            evidence=[
                f"最近任务状态：{hotword_task.status if hotword_task else '无'}",
                f"最近入库条数：{hotword_task.item_count if hotword_task else 0}",
                f"热点池总数：{TrendNote.query.count()}",
            ],
            threshold='建议每天都能看到新热点入库，且最近任务状态持续成功。',
        ),
        watch_item(
            'creator_sync',
            '账号同步稳定性',
            'ready' if creator_sync_task and creator_sync_task.status == 'success' else ('pending' if not creator_sync_task else 'blocked'),
            '观察账号同步任务是否能稳定跑完，用户后续笔记和互动是否正常累计。',
            evidence=[
                f"最近同步状态：{creator_sync_task.status if creator_sync_task else '无'}",
                f"最近同步条数：{creator_sync_task.item_count if creator_sync_task else 0}",
                f"账号数：{CreatorAccount.query.count()} ｜ 笔记数：{CreatorPost.query.count()}",
            ],
            threshold='建议最近同步任务持续成功，且账号/笔记数能随业务继续增长。',
        ),
        watch_item(
            'image',
            '图片生成稳定性',
            'ready' if asset_task and asset_task.status == 'success' else ('pending' if not asset_task else 'blocked'),
            '如果正式启用图片接口，上线后前几天要重点看图片任务是否报错、素材库是否有沉淀。',
            evidence=[
                f"最近图片任务：{asset_task.status if asset_task else '无'}",
                f"素材库总数：{AssetLibrary.query.count()}",
                f"真实图片素材数：{AssetLibrary.query.filter(AssetLibrary.source_provider != 'svg_fallback').count()}",
            ],
            threshold='建议真实图片任务成功率稳定，素材库持续新增。',
        ),
        watch_item(
            'capacity',
            '容量与数据增长',
            'ready' if _build_capacity_readiness_payload().get('summary', {}).get('capacity_ready') else 'blocked',
            '关注报名、提报、热点、账号笔记等数据量是否按预期增长，避免突然堆积。',
            evidence=[
                f"报名数：{Registration.query.count()}",
                f"提报数：{Submission.query.count()}",
                f"热点数：{TrendNote.query.count()}",
                f"候选话题数：{TopicIdea.query.count()}",
            ],
            threshold='按你当前目标量级，系统可承接 100+ 人/月、2000 条/月，重点看异步链路是否持续稳定。',
        ),
    ]

    summary = {
        'count': len(items),
        'healthy': len([item for item in items if item['status'] == 'ready']),
        'attention': len([item for item in items if item['status'] == 'blocked']),
        'pending': len([item for item in items if item['status'] == 'pending']),
    }
    summary['message'] = f"上线后首周建议重点盯 {summary['count']} 项，目前正常 {summary['healthy']} 项。"
    return {
        'success': True,
        'summary': summary,
        'items': items,
    }


def _build_integration_handoff_pack_payload(scope='all'):
    hotword_settings = _hotword_runtime_settings()
    creator_sync_settings = _creator_sync_runtime_settings()
    image_capabilities = _image_provider_capabilities()
    checklist = _build_integration_checklist_payload()
    contracts = _build_integration_contract_payload()
    playbooks = _build_first_run_playbooks_payload()
    acceptance = _build_integration_acceptance_payload()
    trial = _build_trial_readiness_payload()
    go_live = _build_go_live_readiness_payload()
    go_live_checklist = _build_go_live_checklist_payload()
    post_launch = _build_post_launch_watchlist_payload()

    safe_scope = (scope or 'all').strip() or 'all'
    scope_meta = {
        'all': {'label': '完整交付包'},
        'hotword': {'label': '热点接口交付包', 'checklist_key': 'hotword_api', 'runtime_keys': ['hotword_mode', 'hotword_api_url']},
        'creator_sync': {'label': '账号同步交付包', 'checklist_key': 'creator_sync', 'runtime_keys': ['creator_sync_mode', 'creator_sync_api_url']},
        'image_provider': {'label': '图片接口交付包', 'checklist_key': 'image_provider', 'runtime_keys': ['image_provider', 'image_api_url', 'image_model']},
    }
    if safe_scope not in scope_meta:
        safe_scope = 'all'

    package = {
        'generated_at': _format_datetime(datetime.now()),
        'scope': safe_scope,
        'scope_label': scope_meta[safe_scope]['label'],
        'runtime_snapshot': {
            'hotword_mode': _resolved_hotword_mode(hotword_settings),
            'hotword_api_url': hotword_settings.get('hotword_api_url') or '',
            'creator_sync_mode': _resolved_creator_sync_mode(creator_sync_settings),
            'creator_sync_api_url': creator_sync_settings.get('creator_sync_api_url') or '',
            'image_provider': image_capabilities.get('image_provider_name') or 'svg_fallback',
            'image_api_url': image_capabilities.get('image_provider_api_url') or '',
            'image_model': image_capabilities.get('image_provider_model') or '',
        },
        'counts': {
            'registrations': Registration.query.count(),
            'submissions': Submission.query.count(),
            'trend_notes': TrendNote.query.count(),
            'topic_ideas': TopicIdea.query.count(),
            'creator_accounts': CreatorAccount.query.count(),
            'creator_posts': CreatorPost.query.count(),
            'asset_library_items': AssetLibrary.query.count(),
            'enabled_schedules': AutomationSchedule.query.filter_by(enabled=True).count(),
        },
        'integration_checklist': checklist,
        'integration_contracts': contracts.get('items') or [],
        'first_run_playbooks': playbooks.get('items') or [],
        'integration_acceptance': acceptance.get('items') or [],
        'trial_readiness': trial.get('items') or [],
        'go_live_readiness': go_live.get('items') or [],
        'go_live_checklist': go_live_checklist.get('items') or [],
        'post_launch_watchlist': post_launch.get('items') or [],
    }
    if safe_scope != 'all':
        package['integration_checklist'] = [
            item for item in package['integration_checklist']
            if item.get('key') == scope_meta[safe_scope].get('checklist_key')
        ]
        package['integration_contracts'] = [
            item for item in package['integration_contracts']
            if item.get('key') == safe_scope
        ]
        package['first_run_playbooks'] = [
            item for item in package['first_run_playbooks']
            if item.get('key') == safe_scope
        ]
        package['integration_acceptance'] = [
            item for item in package['integration_acceptance']
            if item.get('key') == safe_scope
        ]
        package['runtime_snapshot'] = {
            key: value for key, value in package['runtime_snapshot'].items()
            if key in scope_meta[safe_scope].get('runtime_keys', [])
        }
        package['trial_readiness'] = [
            item for item in package['trial_readiness']
            if item.get('key') in {'integration', 'pilot'}
        ]
        package['go_live_readiness'] = [
            item for item in package['go_live_readiness']
            if item.get('key') in {'automation', 'acceptance', 'go_live'}
        ]
        package['go_live_checklist'] = [
            item for item in package['go_live_checklist']
            if item.get('key') in {'schedules', 'acceptance', 'final_review'}
        ]
        package['post_launch_watchlist'] = [
            item for item in package['post_launch_watchlist']
            if item.get('key') in {
                'hotword' if safe_scope == 'hotword' else (
                    'creator_sync' if safe_scope == 'creator_sync' else 'image'
                ),
                'worker',
                'capacity',
            }
        ]
    return {
        'success': True,
        'summary': {
            'generated_at': package['generated_at'],
            'scope': safe_scope,
            'scope_label': package['scope_label'],
            'checklist_sections': len(package['integration_checklist']),
            'contract_count': len(package['integration_contracts']),
            'playbook_count': len(package['first_run_playbooks']),
            'acceptance_count': len(package['integration_acceptance']),
            'go_live_status': go_live.get('summary', {}).get('overall_status_label') or '未就绪',
        },
        'package': package,
    }


def _serialize_topic(topic):
    return {
        'id': topic.id,
        'activity_id': topic.activity_id,
        'topic_name': topic.topic_name,
        'keywords': topic.keywords or '',
        'direction': topic.direction or '',
        'reference_content': topic.reference_content or '',
        'reference_link': topic.reference_link or '',
        'writing_example': topic.writing_example or '',
        'quota': topic.quota or 0,
        'filled': topic.filled or 0,
        'available': max((topic.quota or 0) - (topic.filled or 0), 0),
        'group_num': topic.group_num or '',
        'pool_status': topic.pool_status or 'formal',
        'pool_status_label': _pool_status_label(topic.pool_status or 'formal'),
        'source_type': topic.source_type or 'manual',
        'source_ref_id': topic.source_ref_id,
        'source_snapshot_id': topic.source_snapshot_id,
        'published_at': topic.published_at.strftime('%Y-%m-%d %H:%M:%S') if topic.published_at else '',
        'created_at': topic.created_at.strftime('%Y-%m-%d %H:%M:%S') if topic.created_at else '',
    }


def _serialize_corpus_entry(entry):
    template_meta = _corpus_template_meta(getattr(entry, 'template_type_key', '') or '')
    return {
        'id': entry.id,
        'title': entry.title,
        'category': entry.category,
        'source': entry.source,
        'source_title': getattr(entry, 'source_title', '') or '',
        'reference_url': getattr(entry, 'reference_url', '') or '',
        'template_type_key': getattr(entry, 'template_type_key', '') or '',
        'template_type_label': template_meta.get('label') or '',
        'tags': entry.tags or '',
        'content': entry.content,
        'usage_count': entry.usage_count or 0,
        'status': entry.status,
        'pool_status': entry.pool_status or 'reserve',
        'pool_status_label': _pool_status_label(entry.pool_status or 'reserve'),
        'created_at': entry.created_at.strftime('%Y-%m-%d %H:%M:%S') if entry.created_at else '',
        'updated_at': entry.updated_at.strftime('%Y-%m-%d %H:%M:%S') if entry.updated_at else '',
    }


CORPUS_TEMPLATE_TYPE_DEFINITIONS = [
    {'key': 'checklist', 'label': '清单模板', 'formula': '适合“问题 + 3~5 条建议/动作”的知识清单'},
    {'key': 'myth_busting', 'label': '误区纠正', 'formula': '适合“误区/别再/不是这样”式反认知内容'},
    {'key': 'comparison', 'label': '对比拆解', 'formula': '适合“区别/对比/选哪个”式内容'},
    {'key': 'process', 'label': '流程步骤', 'formula': '适合“先做什么，再做什么，最后如何判断”的流程表达'},
    {'key': 'qna', 'label': '问答答疑', 'formula': '适合“为什么/怎么做/要不要”式答疑内容'},
    {'key': 'case_story', 'label': '案例故事', 'formula': '适合“案例/患者/经历/故事”式情境内容'},
    {'key': 'standard_explain', 'label': '标准说明', 'formula': '适合稳定、规范、知识说明类内容'},
]


def _split_corpus_tags(raw_tags):
    return [item.strip() for item in re.split(r'[\n,，;；、|/]+', raw_tags or '') if item.strip()]


def _split_reference_links(raw_links=''):
    text = str(raw_links or '')
    items = [
        item.strip().rstrip('。；;，,、）)]】》>')
        for item in re.findall(r'https?://[^\s<>"\'，；;、]+', text)
    ]
    deduped = []
    seen = set()
    for link in items:
        if link in seen:
            continue
        deduped.append(link)
        seen.add(link)
    return deduped[:30]


def _compact_reference_links_for_topic(raw_links='', limit=5, max_length=500):
    compacted = []
    for link in _split_reference_links(raw_links):
        normalized = canonicalize_xhs_post_url(link) or normalize_tracking_url(link) or link
        if normalized not in compacted:
            compacted.append(normalized)
        if len(compacted) >= limit:
            break
    selected = []
    current_length = 0
    for link in compacted:
        projected = current_length + len(link) + (1 if selected else 0)
        if projected > max_length:
            break
        selected.append(link)
        current_length = projected
    return '\n'.join(selected)


def _clean_topic_piece(text='', limit=80):
    value = re.sub(r'^\s*\d+[\.、）)]\s*', '', str(text or '').strip())
    value = re.sub(r'\s+', ' ', value)
    return value[:limit]


def _split_idea_topic_title(title=''):
    parts = [part.strip() for part in str(title or '').split('｜') if part and part.strip()]
    topic = _clean_topic_piece(parts[0] if parts else title, 80) or '肝健康话题'
    type_label = _clean_topic_piece(parts[1] if len(parts) > 1 else '', 50)
    return topic, type_label


def _normalize_public_topic_type(type_text=''):
    text = _clean_topic_piece(type_text, 60)
    if '产品实拍' in text:
        return '产品实拍型'
    if '合集' in text or '并发症' in text or '测评' in text:
        return '对比测评型'
    if '大字报' in text:
        return '大字报互动型'
    if '报告解读' in text:
        return '报告解读型'
    if '备忘录' in text or '清单' in text:
        return '清单收藏型'
    if '图表' in text:
        return '图表说明型'
    if '医院' in text or '体检中心' in text or '实拍' in text:
        return '实拍记录型'
    if '医疗行业' in text or '深度科普' in text:
        return '深度科普型'
    return text or '内容执行型'


def _public_topic_keywords(raw_keywords='', topic_name='', limit=6):
    values = []
    for item in re.split(r'[#,\n，、/ ]+', str(raw_keywords or '')):
        word = item.strip()
        if not word or word in {'带话题', '话题', '搜索词'}:
            continue
        if len(word) > 18:
            continue
        if word not in values:
            values.append(word)
        if len(values) >= limit:
            break
    if topic_name and topic_name not in values:
        values.insert(0, topic_name[:18])
    return ','.join(values[:limit])[:500]


def _topic_idea_public_strategy(idea):
    topic_core, title_type = _split_idea_topic_title(idea.topic_title or '')
    type_label = _normalize_public_topic_type(title_type or idea.content_type or '')
    text = ' '.join([
        idea.topic_title or '',
        idea.keywords or '',
        idea.angle or '',
        idea.asset_brief or '',
        idea.soft_insertion or '',
    ])
    product = 'FibroScan福波看' if re.search(r'FibroScan|FS|福波看|肝弹|弹性|体检|检查|报告', text, re.I) else '复方鳖甲软肝片'
    display_title = f'{topic_core}｜{type_label}'[:200]

    if type_label == '产品实拍型':
        direction = '用真实复查/长期管理场景切入，先讲人的经历和变化，再自然承接产品；不要写成药品说明书。'
        key_points = ['真实复查或家庭用药场景', '长期坚持和遵医嘱', '产品只做自然露出']
        angle = '适合患者/家属/资深病友视角，用“为什么愿意坚持”建立信任。'
        hooks = ['复查结果出来后，我才真正踏实', '肝硬化用药这件事，别只看药名', '这几年我才明白：稳定比一时见效更重要']
    elif type_label == '大字报互动型':
        direction = '用一个强问题做封面，正文只回答一个关键判断，适合拉停留和评论互动。'
        key_points = ['封面只放一个冲突问题', '正文先破误区再给判断', '结尾抛一个用户愿意回复的问题']
        angle = '适合把体检焦虑、报告异常、要不要复查这类问题讲得更直观。'
        hooks = ['体检报告出来后，很多人第一眼看错了', '别只盯转氨酶，肝脏筛查还要看这些', '这项异常到底要不要复查？']
    elif type_label == '报告解读型':
        direction = '抓住用户看到报告时的焦虑，先告诉“看哪几项”，再解释趋势和下一步。'
        key_points = ['圈出关键指标', '解释变化趋势', '提醒异常结果问医生']
        angle = '适合报告截图、指标圈重点、检查结果解读。'
        hooks = ['FibroScan报告别只盯一个数字', '肝弹报告不会读？先按这个顺序看', '报告有异常，不代表一定很严重']
    elif type_label == '图表说明型':
        direction = '把复杂项目做成一张选择图，只解决“先查什么、后查什么”的问题。'
        key_points = ['必查/可选分层', '按人群风险选择', '一张图只讲一个判断']
        angle = '适合收藏型图表、体检项目路线图、报告指标说明。'
        hooks = ['体检项目一张图看懂：先查什么，后查什么', '别再乱买套餐，先看这张筛查路线图', '看不懂检查名？先按风险分层']
    elif type_label == '实拍记录型':
        direction = '用真实体检/检查流程降低陌生感，重点拍流程、注意事项和用户最担心的地方。'
        key_points = ['预约到检查流程', '现场注意事项', '隐私信息打码']
        angle = '适合体检中心、医院流程、检查体验实拍。'
        hooks = ['第一次做肝弹检查，我最想提前知道这几步', '体检当天先确认这几个问题，少跑冤枉路', '把检查流程拍清楚，第一次做也不慌']
    elif type_label == '清单收藏型':
        direction = '直接做成用户能截图保存的行动清单，突出下一步怎么做和哪些项目别漏。'
        key_points = ['检查前准备', '核心项目优先级', '复查/咨询医生问题']
        angle = '适合备忘录、攻略、家庭体检清单。'
        hooks = ['这份体检清单先存下来', '不是项目越多越好，先把核心问题查明白', '给家里人安排体检，我会先核对这几项']
    else:
        direction = '把专业信息翻译成用户能执行的判断，一篇只解决一个核心问题。'
        key_points = ['先讲用户痛点', '再讲判断逻辑', '最后给下一步动作']
        angle = '适合专业科普、检查项目解释和长期健康管理。'
        hooks = ['这项检查不是越多越好，关键看适合谁', '把专业判断翻译成人话，用户才会收藏', '先别急着下结论，先看这几个判断点']

    if product == '复方鳖甲软肝片':
        risk = '风险提醒：不承诺疗效，不引导购买，用“遵医嘱、复查变化、长期管理”表达。'
    else:
        risk = '风险提醒：不替用户下诊断结论，检查结果需结合病史、复查和医生判断。'

    reference_content = '\n'.join([
        f'核心关键点：{"；".join(key_points)}',
        f'撰写角度：{angle}',
        f'标题参考：{" / ".join(hooks[:3])}',
        risk,
    ])
    return {
        'topic_name': display_title,
        'keywords': _public_topic_keywords(idea.keywords or '', topic_core),
        'direction': direction,
        'reference_content': reference_content,
        'writing_example': '\n'.join(hooks[:3]),
        'group_num': product,
    }


def _hotword_scope_preset_meta(scope_key=''):
    key = (scope_key or '').strip()
    for item in HOTWORD_SCOPE_PRESETS:
        if item['key'] == key:
            return dict(item)
    return dict(HOTWORD_SCOPE_PRESETS[0])


def _hotword_scope_presets():
    return [dict(item) for item in HOTWORD_SCOPE_PRESETS]


def _hotword_time_window_options():
    return [dict(item) for item in HOTWORD_TIME_WINDOW_OPTIONS]


def _resolve_hotword_scope_keywords(scope_key='', raw_keywords=''):
    custom_keywords = split_hotword_keywords(raw_keywords or '')
    if custom_keywords:
        return custom_keywords[:30]
    preset = _hotword_scope_preset_meta(scope_key)
    return list(preset.get('keywords') or [])[:30]


def _resolve_hotword_date_window(window_key='30d', custom_from='', custom_to=''):
    key = (window_key or '30d').strip().lower() or '30d'
    now = datetime.now()
    date_from = (custom_from or '').strip()
    date_to = (custom_to or '').strip()
    if key == 'custom':
        return {
            'window_key': 'custom',
            'date_from': date_from,
            'date_to': date_to,
            'label': '自定义' if (date_from or date_to) else '未设置',
        }
    days = 30
    if key == '3d':
        days = 3
    elif key == '7d':
        days = 7
    date_to_value = now.strftime('%Y-%m-%d')
    date_from_value = (now - timedelta(days=days - 1)).strftime('%Y-%m-%d')
    label_map = {'3d': '近3天', '7d': '近7天', '30d': '近30天'}
    return {
        'window_key': key if key in {'3d', '7d', '30d'} else '30d',
        'date_from': date_from_value,
        'date_to': date_to_value,
        'label': label_map.get(key, '近30天'),
    }


def _corpus_template_meta(template_key=''):
    raw = (template_key or '').strip()
    for item in CORPUS_TEMPLATE_TYPE_DEFINITIONS:
        if item['key'] == raw:
            return dict(item)
    return dict(CORPUS_TEMPLATE_TYPE_DEFINITIONS[-1])


def _infer_corpus_template_type(title='', content=''):
    text = f"{title or ''}\n{content or ''}".strip()
    line_count = len([line for line in (content or '').splitlines() if line.strip()])
    numbered_lines = len(re.findall(r'(^|\n)\s*(\d+[、\.]|[-*•])', content or ''))

    if any(token in text for token in ['误区', '别再', '不是这样', '避坑', '别把']):
        return _corpus_template_meta('myth_busting')
    if any(token in text for token in ['对比', '区别', 'vs', 'VS', '哪个好', '怎么选']):
        return _corpus_template_meta('comparison')
    if any(token in text for token in ['案例', '患者', '经历', '故事', '复盘']):
        return _corpus_template_meta('case_story')
    if any(token in text for token in ['为什么', '怎么', '如何', '？', '?', '要不要']):
        return _corpus_template_meta('qna')
    if any(token in text for token in ['步骤', '流程', '先', '再', '最后']) and line_count >= 2:
        return _corpus_template_meta('process')
    if numbered_lines >= 2 or any(token in text for token in ['清单', '建议', '做到', '记住', '重点']) or line_count >= 4:
        return _corpus_template_meta('checklist')
    return _corpus_template_meta('standard_explain')


def _extract_reference_note_token(link=''):
    if not link:
        return ''
    parsed = urlparse(link)
    path = (parsed.path or '').rstrip('/')
    if not path:
        return ''
    token = path.split('/')[-1].strip()
    if len(token) < 6:
        return ''
    return token


def _resolve_reference_source_by_link(link=''):
    raw_link = (link or '').strip()
    if not raw_link:
        return {
            'reference_url': '',
            'source_title': '',
            'source_text': '',
        }

    normalized_link = raw_link
    if 'xiaohongshu.com' in raw_link:
        note_token = _extract_reference_note_token(raw_link)
        if note_token and re.fullmatch(r'[0-9a-zA-Z]{8,}', note_token):
            normalized_link = canonicalize_xhs_post_url(raw_link) or raw_link

    note = TrendNote.query.filter(
        TrendNote.link.in_([raw_link, normalized_link])
    ).order_by(TrendNote.created_at.desc(), TrendNote.id.desc()).first()
    if note:
        summary_parts = [note.summary or '', f'标题：{note.title or ""}', f'关键词：{note.keyword or ""}']
        return {
            'reference_url': note.link or normalized_link,
            'source_title': note.title or '',
            'source_text': '\n'.join([part for part in summary_parts if str(part).strip()]).strip(),
        }

    post = CreatorPost.query.filter(
        CreatorPost.post_url.in_([raw_link, normalized_link])
    ).order_by(CreatorPost.created_at.desc(), CreatorPost.id.desc()).first()
    if post:
        post_parts = [
            post.title or '',
            f'话题：{post.topic_title or ""}',
            f'阅读：{post.views or 0} 点赞：{post.likes or 0} 收藏：{post.favorites or 0} 评论：{post.comments or 0}',
        ]
        return {
            'reference_url': post.post_url or normalized_link,
            'source_title': post.title or '',
            'source_text': '\n'.join([part for part in post_parts if str(part).strip()]).strip(),
        }

    token = _extract_reference_note_token(normalized_link)
    if token:
        fuzzy_note = TrendNote.query.filter(TrendNote.link.contains(token)).order_by(
            TrendNote.created_at.desc(),
            TrendNote.id.desc(),
        ).first()
        if fuzzy_note:
            return {
                'reference_url': fuzzy_note.link or normalized_link,
                'source_title': fuzzy_note.title or '',
                'source_text': '\n'.join([part for part in [fuzzy_note.summary or '', fuzzy_note.title or ''] if str(part).strip()]).strip(),
            }
        fuzzy_post = CreatorPost.query.filter(CreatorPost.post_url.contains(token)).order_by(
            CreatorPost.created_at.desc(),
            CreatorPost.id.desc(),
        ).first()
        if fuzzy_post:
            return {
                'reference_url': fuzzy_post.post_url or normalized_link,
                'source_title': fuzzy_post.title or '',
                'source_text': fuzzy_post.title or '',
            }

    return {
        'reference_url': normalized_link,
        'source_title': '',
        'source_text': '',
    }


def _infer_reference_style_signals(title='', content=''):
    text = f"{title or ''}\n{content or ''}".strip()
    title_logic = '结果/结论前置'
    if any(token in title for token in ['？', '?', '怎么', '为什么', '要不要']):
        title_logic = '提问答疑型'
    elif any(token in title for token in ['别再', '误区', '不是', '搞错', '千万']):
        title_logic = '误区纠偏型'
    elif any(token in title for token in ['清单', '步骤', '记住', '攻略', '总结']):
        title_logic = '清单收藏型'
    elif any(token in title for token in ['我', '后来', '经历', '踩坑']):
        title_logic = '经历共鸣型'

    opening_logic = '开头直接给结论或问题'
    if any(token in text for token in ['我当时', '那天', '后来', '一开始', '我以前']):
        opening_logic = '先用真实经历/场景带入'
    elif any(token in text for token in ['先说结论', '先看', '先记住', '第一步']):
        opening_logic = '先给结论或步骤，再展开解释'

    ending_logic = '结尾自然追问或提醒收藏'
    if any(token in text for token in ['你们会怎么做', '你会怎么选', '你们一般']):
        ending_logic = '结尾提问，拉互动'
    elif any(token in text for token in ['建议收藏', '可以先存', '留着备用']):
        ending_logic = '结尾偏收藏提醒'

    return {
        'title_logic': title_logic,
        'opening_logic': opening_logic,
        'ending_logic': ending_logic,
    }


def _build_reference_corpus_entry_payload(reference_url='', source_title='', source_text='', style_hint='', product_anchor='', manual_title='', category='爆款拆解', source='参考链接导入', tags=''):
    clean_title = (manual_title or source_title or '参考链接模板').strip()
    clean_text = (source_text or '').strip()
    clean_hint = (style_hint or '').strip()
    product_anchor = (product_anchor or '').strip()
    template_meta = _infer_corpus_template_type(title=clean_title, content='\n'.join(filter(None, [clean_text, clean_hint])))
    signals = _infer_reference_style_signals(clean_title, clean_text or clean_hint)
    merged_tags = _split_corpus_tags(tags)
    merged_tags.extend([
        '参考链接导入',
        template_meta.get('label') or '',
        product_anchor,
    ])
    merged_tags = [item for item in dict.fromkeys([tag.strip() for tag in merged_tags if tag.strip()])][:12]

    corpus_title = clean_title[:200]
    if not manual_title:
        corpus_title = f'参考模板｜{corpus_title}'[:200]

    body_lines = [
        f'参考链接：{reference_url or "未提供"}',
        f'参考标题：{source_title or clean_title or "未命名"}',
        f'模板类型：{template_meta.get("label") or "标准说明"}',
        f'标题逻辑：{signals["title_logic"]}',
        f'开头方式：{signals["opening_logic"]}',
        f'结尾方式：{signals["ending_logic"]}',
        '仿写要求：只学习结构、节奏、钩子、互动方式和版式感，不照抄原句、不复用原人物故事。',
    ]
    if product_anchor:
        body_lines.append(f'改写锚点：生成时要改写成围绕“{product_anchor}”的产品/服务内容。')
    else:
        body_lines.append('改写锚点：生成时要改写成围绕我们的产品/服务内容。')
    if clean_hint:
        body_lines.append(f'运营备注：{clean_hint}')
    if clean_text:
        body_lines.append('参考文案/拆解：')
        body_lines.append(clean_text[:3000])

    return {
        'title': corpus_title,
        'category': (category or '爆款拆解').strip() or '爆款拆解',
        'source': (source or '参考链接导入').strip() or '参考链接导入',
        'source_title': (source_title or clean_title or '').strip()[:300],
        'reference_url': (reference_url or '').strip()[:500],
        'template_type_key': (template_meta.get('key') or '').strip()[:50],
        'tags': ','.join(merged_tags)[:300],
        'content': '\n'.join([line for line in body_lines if str(line).strip()]).strip(),
        'status': 'active',
    }


def _build_trend_note_corpus_payload(note, *, category='爆款拆解', source='热点爆款转模板'):
    template_meta = _infer_corpus_template_type(
        title=note.title or '',
        content='\n'.join(filter(None, [note.summary or '', note.keyword or '']))
    )
    tags = []
    for item in [note.keyword or '', note.source_platform or '', note.source_channel or '', template_meta.get('label') or '']:
        if str(item or '').strip():
            tags.append(str(item).strip())
    tags = list(dict.fromkeys(tags))
    content_lines = [
        f'参考标题：{note.title or "未命名"}',
        f'参考链接：{note.link or "无"}',
        f'来源平台：{note.source_platform or "小红书"}',
        f'模板类型：{template_meta.get("label") or "标准说明"}',
        f'互动表现：阅读 {note.views or 0} / 点赞 {note.likes or 0} / 收藏 {note.favorites or 0} / 评论 {note.comments or 0}',
        '仿写要求：只学习标题逻辑、结构节奏、开头钩子和互动方式，不照抄原句，不复制原作者经历。',
        '适用方向：优先改写成肝健康IP内容，可按中医调理、脂肪肝管理、检查解读、情绪身心等方向二次生成。',
    ]
    if (note.summary or '').strip():
        content_lines.append(f'摘要拆解：{(note.summary or "").strip()}')
    if (note.keyword or '').strip():
        content_lines.append(f'关键词：{(note.keyword or "").strip()}')
    return {
        'title': f'热点模板｜{(note.title or "未命名模板")[:40]}'[:200],
        'category': category,
        'source': source,
        'source_title': (note.title or '').strip()[:300],
        'reference_url': (note.link or '').strip()[:500],
        'template_type_key': (template_meta.get('key') or '').strip()[:50],
        'tags': ','.join(tags)[:300],
        'content': '\n'.join([row for row in content_lines if str(row).strip()]).strip(),
        'status': 'active',
    }


def _upsert_trend_note_corpus_entries(notes, *, category='爆款拆解', source='热点爆款转模板'):
    created = []
    updated = []
    for note in notes:
        payload = _build_trend_note_corpus_payload(note, category=category, source=source)
        existing = None
        if payload['reference_url']:
            existing = CorpusEntry.query.filter_by(reference_url=payload['reference_url']).first()
        if not existing:
            existing = CorpusEntry.query.filter_by(source=source, source_title=payload['source_title']).first()
        if existing:
            existing.title = payload['title']
            existing.category = payload['category']
            existing.source = payload['source']
            existing.source_title = payload['source_title']
            existing.reference_url = payload['reference_url']
            existing.template_type_key = payload['template_type_key']
            existing.tags = payload['tags']
            existing.content = payload['content']
            existing.status = 'active'
            updated.append(existing)
            continue
        entry = CorpusEntry(
            title=payload['title'],
            category=payload['category'],
            source=payload['source'],
            source_title=payload['source_title'],
            reference_url=payload['reference_url'],
            template_type_key=payload['template_type_key'],
            tags=payload['tags'],
            content=payload['content'],
            status='active',
        )
        db.session.add(entry)
        db.session.flush()
        created.append(entry)
    return {
        'created': created,
        'updated': updated,
    }


def _upsert_reference_corpus_entries(reference_links, *, reference_note_text='', style_hint='', product_anchor='', manual_title='', category='爆款拆解', source='参考链接导入', tags=''):
    created = []
    updated = []
    for link in reference_links:
        source_bundle = _resolve_reference_source_by_link(link)
        entry_payload = _build_reference_corpus_entry_payload(
            reference_url=source_bundle.get('reference_url') or link,
            source_title=source_bundle.get('source_title') or '',
            source_text=reference_note_text or source_bundle.get('source_text') or '',
            style_hint=style_hint,
            product_anchor=product_anchor,
            manual_title=manual_title,
            category=category,
            source=source,
            tags=tags,
        )
        existing = CorpusEntry.query.filter_by(reference_url=entry_payload['reference_url']).first()
        if existing:
            existing.title = entry_payload['title']
            existing.category = entry_payload['category']
            existing.source = entry_payload['source']
            existing.source_title = entry_payload['source_title']
            existing.template_type_key = entry_payload['template_type_key']
            existing.tags = entry_payload['tags']
            existing.content = entry_payload['content']
            existing.status = 'active'
            updated.append(existing)
            continue

        entry = CorpusEntry(
            title=entry_payload['title'],
            category=entry_payload['category'],
            source=entry_payload['source'],
            source_title=entry_payload['source_title'],
            reference_url=entry_payload['reference_url'],
            template_type_key=entry_payload['template_type_key'],
            tags=entry_payload['tags'],
            content=entry_payload['content'],
            status='active',
        )
        db.session.add(entry)
        db.session.flush()
        created.append(entry)
    return {
        'created': created,
        'updated': updated,
    }


def _build_topic_reference_import_payload(topic):
    reference_links = _split_reference_links(topic.reference_link or '')
    topic_keywords = ','.join(_split_keywords(topic.keywords or ''))
    tags = ','.join(filter(None, [
        topic_keywords,
        topic.topic_name or '',
        '话题参考链接',
    ]))
    reference_note_text = '\n'.join(filter(None, [
        (topic.writing_example or '').strip(),
        (topic.reference_content or '').strip(),
    ])).strip()
    style_hint = '\n'.join(filter(None, [
        f'话题名称：{topic.topic_name or ""}',
        f'撰写方向：{topic.direction or ""}',
    ])).strip()
    product_anchor = _detect_soft_insertion(' '.join(filter(None, [topic.topic_name or '', topic.keywords or '', topic.direction or ''])))
    return {
        'reference_links': reference_links,
        'reference_note_text': reference_note_text,
        'style_hint': style_hint,
        'product_anchor': product_anchor,
        'manual_title': '',
        'category': '爆款拆解',
        'source': '话题参考链接导入',
        'tags': tags,
    }


def _build_corpus_insights_payload(pool_status='', category=''):
    expected_categories = ['爆款拆解', '医学科普', '合规表达', '产品卖点', '封面模板']
    query = CorpusEntry.query
    raw_pool = (pool_status or '').strip()
    raw_category = (category or '').strip()
    if raw_pool:
        query = query.filter_by(pool_status=raw_pool)
    if raw_category:
        query = query.filter_by(category=raw_category)
    entries = query.order_by(CorpusEntry.updated_at.desc(), CorpusEntry.id.desc()).all()

    category_counter = Counter()
    source_counter = Counter()
    tag_counter = Counter()
    pool_counter = Counter()
    template_counter = Counter()
    template_samples = defaultdict(list)
    category_stats = defaultdict(lambda: {'count': 0, 'usage': 0, 'length': 0, 'candidate': 0, 'reserve': 0, 'archived': 0})
    template_stats = defaultdict(lambda: {'count': 0, 'usage': 0, 'length': 0})

    reusable_entries = []
    total_length = 0
    for entry in entries:
        content = (entry.content or '').strip()
        title = (entry.title or '').strip()
        content_length = len(content)
        total_length += content_length
        category_key = (entry.category or '未分类').strip() or '未分类'
        source_key = (entry.source or '未标记来源').strip() or '未标记来源'
        pool_key = (entry.pool_status or 'reserve').strip() or 'reserve'
        template_meta = _corpus_template_meta(getattr(entry, 'template_type_key', '') or '') if (getattr(entry, 'template_type_key', '') or '').strip() else _infer_corpus_template_type(title=title, content=content)
        template_key = template_meta['key']

        category_counter[category_key] += 1
        source_counter[source_key] += 1
        pool_counter[pool_key] += 1
        template_counter[template_key] += 1
        for tag in _split_corpus_tags(entry.tags):
            tag_counter[tag] += 1

        stats = category_stats[category_key]
        stats['count'] += 1
        stats['usage'] += entry.usage_count or 0
        stats['length'] += content_length
        stats[pool_key if pool_key in {'candidate', 'reserve', 'archived'} else 'reserve'] += 1

        t_stats = template_stats[template_key]
        t_stats['count'] += 1
        t_stats['usage'] += entry.usage_count or 0
        t_stats['length'] += content_length
        if len(template_samples[template_key]) < 3:
            template_samples[template_key].append(title or f'语料 #{entry.id}')

        reusable_entries.append({
            'id': entry.id,
            'title': title,
            'category': category_key,
            'source': source_key,
            'usage_count': entry.usage_count or 0,
            'content_length': content_length,
            'template_type': template_meta['label'],
            'tags': _split_corpus_tags(entry.tags)[:4],
            'updated_at': _format_datetime(entry.updated_at),
        })

    reusable_entries.sort(key=lambda item: (item['usage_count'], item['content_length']), reverse=True)

    category_rows = []
    for key, stats in category_stats.items():
        category_rows.append({
            'category': key,
            'count': stats['count'],
            'avg_usage': round(stats['usage'] / stats['count'], 2) if stats['count'] else 0,
            'avg_length': round(stats['length'] / stats['count'], 2) if stats['count'] else 0,
            'candidate_count': stats['candidate'],
            'reserve_count': stats['reserve'],
            'archived_count': stats['archived'],
        })
    category_rows.sort(key=lambda item: (item['count'], item['avg_usage']), reverse=True)

    source_rows = [{'source': key, 'count': count} for key, count in source_counter.most_common(8)]
    tag_rows = [{'tag': key, 'count': count} for key, count in tag_counter.most_common(12)]
    pool_rows = [{'pool_status': key, 'count': count, 'label': _pool_status_label(key)} for key, count in pool_counter.items()]

    template_rows = []
    for template_key, count in template_counter.items():
        meta = _corpus_template_meta(template_key)
        stats = template_stats[template_key]
        template_rows.append({
            'template_key': template_key,
            'template_label': meta['label'],
            'formula': meta['formula'],
            'count': count,
            'avg_usage': round(stats['usage'] / count, 2) if count else 0,
            'avg_length': round(stats['length'] / count, 2) if count else 0,
            'sample_titles': template_samples.get(template_key, []),
        })
    template_rows.sort(key=lambda item: (item['count'], item['avg_usage']), reverse=True)

    gaps = []
    for name in expected_categories:
        row = next((item for item in category_rows if item['category'] == name), None)
        if not row:
            gaps.append(f'{name} 暂无语料，建议优先补 3~5 条标准样本。')
        elif row['count'] < 3:
            gaps.append(f'{name} 只有 {row["count"]} 条语料，建议再补充。')
    if not tag_rows:
        gaps.append('当前语料标签较少，建议后续录入时统一补标签，方便检索和复用。')

    summary = {
        'total': len(entries),
        'active': len([item for item in entries if (item.status or 'active') == 'active']),
        'avg_content_length': round(total_length / len(entries), 2) if entries else 0,
        'category_count': len(category_rows),
        'tag_count': len(tag_counter),
        'template_count': len(template_rows),
        'message': (
            f'当前共 {len(entries)} 条语料，已识别 {len(template_rows)} 类模板。'
            if entries else
            '当前语料库还没有数据，先补样本后分析价值更高。'
        ),
    }
    return {
        'success': True,
        'summary': summary,
        'categories': category_rows,
        'sources': source_rows,
        'tags': tag_rows,
        'pools': pool_rows,
        'templates': template_rows,
        'reusable_entries': reusable_entries[:8],
        'gaps': gaps[:8],
        'filters': {
            'pool_status': raw_pool,
            'category': raw_category,
        },
    }


def _image_template_family_meta_map():
    return {
        'medical_science': {
            'label': '报告解读卡',
            'description': '适合体检单、指标说明、检查报告和医学解释类内容。',
        },
        'knowledge_card': {
            'label': '课堂笔记卡',
            'description': '适合知识拆解、误区纠偏、重点提炼和结构化讲解。',
        },
        'poster': {
            'label': '大字封面',
            'description': '适合冲点击、先抓眼球，再把正文和内页接住。',
        },
        'checklist': {
            'label': '清单卡',
            'description': '适合步骤、避坑、对照表、可收藏执行清单。',
        },
        'memo': {
            'label': '备忘录/陪伴卡',
            'description': '适合第一人称经历、陪伴型表达、情绪共鸣和复盘。',
        },
        'custom': {
            'label': '自定义模板',
            'description': '适合你自己指定画面方向，再由系统选近似模板落图。',
        },
    }


def _default_lane_to_image_skill_key(lane_key=''):
    lane_to_image_skill = {
        'report_interpretation': 'report_decode',
        'liver_comorbidity': 'report_decode',
        'fatty_liver_management': 'save_worthy_cards',
        'diet_nutrition': 'save_worthy_cards',
        'exercise_fitness': 'save_worthy_cards',
        'tcm_conditioning': 'story_atmosphere',
        'emotion_mindbody': 'story_atmosphere',
        'family_care': 'story_atmosphere',
        'women_health': 'story_atmosphere',
        'myth_busting': 'high_click_cover',
        'liver_care_habits': 'high_click_cover',
    }
    return lane_to_image_skill.get((lane_key or '').strip(), 'high_click_cover')


def _build_image_template_agent_payload(image_skill_key='high_click_cover'):
    preset_map = get_image_skill_presets() or {}
    preset = preset_map.get((image_skill_key or '').strip()) or preset_map.get('high_click_cover') or {}
    style_meta_map = {
        (item.get('key') or ''): item
        for item in (_asset_style_type_options() or [])
        if isinstance(item, dict)
    }
    family_key = preset.get('family_key') or ''
    cover_style_key = preset.get('cover_style_key') or ''
    inner_style_key = preset.get('inner_style_key') or ''
    family_meta = _image_template_family_meta_map().get(family_key, {})
    return {
        'image_skill': image_skill_key,
        'image_skill_label': preset.get('label') or IMAGE_SKILL_OPTIONS.get(image_skill_key) or image_skill_key,
        'template_agent_label': family_meta.get('label') or '',
        'template_agent_description': family_meta.get('description') or '',
        'cover_style_type': cover_style_key,
        'cover_style_label': (style_meta_map.get(cover_style_key) or {}).get('label') or cover_style_key,
        'inner_style_type': inner_style_key,
        'inner_style_label': (style_meta_map.get(inner_style_key) or {}).get('label') or inner_style_key,
    }


def _recommend_image_template_payload(text='', lane_key=''):
    resolved_lane = (lane_key or '').strip() or _infer_trend_lane_key(text)
    return _build_image_template_agent_payload(_default_lane_to_image_skill_key(resolved_lane))


def _serialize_trend_note(note):
    payload = _load_json_value(note.raw_payload, {})
    metric_sources = payload.get('metric_sources') if isinstance(payload, dict) else {}
    exposures = _safe_int(payload.get('exposures')) if isinstance(payload, dict) else 0
    linked_entry = None
    if note.link:
        linked_entry = CorpusEntry.query.filter_by(reference_url=note.link).order_by(CorpusEntry.updated_at.desc(), CorpusEntry.id.desc()).first()
    if not linked_entry:
        linked_entry = CorpusEntry.query.filter_by(
            source='热点爆款转模板',
            source_title=note.title or '',
        ).order_by(CorpusEntry.updated_at.desc(), CorpusEntry.id.desc()).first()
    recommendation = _recommended_trend_route(note)
    image_template_recommendation = _recommend_image_template_payload(
        text=' '.join([note.keyword or '', note.title or '', note.summary or '']),
        lane_key=recommendation['lane_key'],
    )
    return {
        'id': note.id,
        'source_platform': note.source_platform,
        'source_channel': note.source_channel,
        'source_template_key': note.source_template_key or 'generic_lines',
        'source_template_label': _hotword_source_template_meta(note.source_template_key or 'generic_lines').get('label'),
        'import_batch': note.import_batch or '',
        'keyword': note.keyword or '',
        'title': note.title,
        'author': note.author or '',
        'link': note.link or '',
        'views': note.views or 0,
        'exposures': exposures,
        'likes': note.likes or 0,
        'favorites': note.favorites or 0,
        'comments': note.comments or 0,
        'interactions': (note.likes or 0) + (note.favorites or 0) + (note.comments or 0),
        'hot_score': note.hot_score or _trend_score(note),
        'source_rank': note.source_rank or 0,
        'score': note.hot_score or _trend_score(note),
        'summary': note.summary or '',
        'metric_sources': metric_sources if isinstance(metric_sources, dict) else {},
        'metric_source_summary': _metric_source_summary(metric_sources, preferred_keys=['views', 'exposures', 'hot_value', 'likes', 'favorites', 'comments']),
        'metric_source_summary_text': _metric_source_summary_text(metric_sources, preferred_keys=['views', 'exposures', 'hot_value', 'likes', 'favorites', 'comments']),
        'pool_status': note.pool_status or 'reserve',
        'pool_status_label': _pool_status_label(note.pool_status or 'reserve'),
        'linked_corpus_entry_id': linked_entry.id if linked_entry else None,
        'has_corpus_template': bool(linked_entry),
        'planning_score': recommendation['score'],
        'planning_confidence': recommendation['confidence'],
        'recommended_target': recommendation['target'],
        'recommended_target_label': recommendation['target_label'],
        'recommended_reason': recommendation['reason'],
        'lane_key': recommendation['lane_key'],
        'lane_label': recommendation['lane_label'],
        'persona_key': recommendation['persona_key'],
        'persona_label': recommendation['persona_label'],
        **image_template_recommendation,
        'created_at': note.created_at.strftime('%Y-%m-%d %H:%M:%S') if note.created_at else '',
    }


def _serialize_hot_topic_entry(item):
    return {
        'id': item.id,
        'reference_note_id': item.reference_note_id,
        'title': item.title or '',
        'keyword': item.keyword or '',
        'source_platform': item.source_platform or '',
        'source_channel': item.source_channel or '',
        'reference_url': item.reference_url or '',
        'summary': item.summary or '',
        'hot_score': item.hot_score or 0,
        'lane_key': item.lane_key or '',
        'persona_key': item.persona_key or '',
        'audience_hint': item.audience_hint or '',
        'integration_hint': item.integration_hint or '',
        'usage_tip': item.usage_tip or '',
        'status': item.status or 'active',
        'created_at': item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
        'updated_at': item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else '',
    }


def _infer_trend_lane_key(text=''):
    joined = (text or '').strip()
    if any(token in joined for token in ['中医', '肝气郁结', '肝郁', '肝胆湿热', '疏肝理气', '养肝']):
        return 'tcm_conditioning'
    if any(token in joined for token in ['情绪', '焦虑', '压力', '失眠', '熬夜', '睡眠']):
        return 'emotion_mindbody'
    if any(token in joined for token in ['脂肪肝', '减脂', '体重', '肥胖', '减肥', '饮食', '运动']):
        return 'fatty_liver_management'
    if any(token in joined for token in ['糖尿病', '冠心病', '高血糖', '共病']):
        return 'liver_comorbidity'
    if any(token in joined for token in ['体检', '检查', '指标', '转氨酶', 'FibroScan', '福波看', '肝弹', 'B超', '复查']):
        return 'report_interpretation'
    if any(token in joined for token in ['误区', '别再', '不是这样', '搞错', '避坑', '反常识']):
        return 'myth_busting'
    if any(token in joined for token in ['父母', '家人', '家属', '陪诊', '照护', '老人']):
        return 'family_care'
    if any(token in joined for token in ['女性', '经期', '更年期', '姨妈']):
        return 'women_health'
    if any(token in joined for token in ['护肝', '保肝', '作息', '应酬', '饮酒', '习惯']):
        return 'liver_care_habits'
    return 'report_interpretation'


def _infer_trend_persona_key(text=''):
    joined = (text or '').strip()
    if any(token in joined for token in ['父母', '家人', '家属', '陪诊', '照护', '老人']):
        return 'family'
    if any(token in joined for token in ['中医', '肝气郁结', '肝郁', '疏肝理气', '肝胆湿热', '养肝']):
        return 'tcm'
    if any(token in joined for token in ['饮食', '营养', '减脂餐', '早餐', '食谱', '控糖', '控油']):
        return 'nutrition'
    if any(token in joined for token in ['运动', '健身', '跑步', '力量', '减脂训练']):
        return 'fitness'
    if any(token in joined for token in ['女性', '经期', '更年期', '姨妈']):
        return 'women'
    if any(token in joined for token in ['上班', '职场', '久坐', '加班', '应酬']):
        return 'office'
    if any(token in joined for token in ['体检', '报告', '指标', 'FibroScan', '福波看', '肝弹', '科普', '问答']):
        return 'medical'
    return 'patient'


def _trend_lane_label(lane_key=''):
    label_map = {
        'report_interpretation': '检查解读线',
        'fatty_liver_management': '脂肪肝管理线',
        'liver_care_habits': '保肝护肝习惯线',
        'tcm_conditioning': '中医调理线',
        'emotion_mindbody': '情绪与肝线',
        'family_care': '家庭照护线',
        'women_health': '女性养肝线',
        'myth_busting': '误区纠偏线',
        'liver_comorbidity': '肝病+共病线',
    }
    return label_map.get((lane_key or '').strip(), lane_key or '肝健康方向')


def _trend_persona_label(persona_key=''):
    label_map = {
        'patient': '患者本人',
        'family': '家属/陪诊者',
        'medical': '医学科普型',
        'tcm': '中医调理型',
        'nutrition': '健管/营养型',
        'fitness': '运动减脂型',
        'women': '女性健康型',
        'office': '细分人群博主',
    }
    return label_map.get((persona_key or '').strip(), persona_key or '肝健康创作者')


def _trend_route_target_label(target=''):
    label_map = {
        'current_topic': '发当前话题广场',
        'next_activity': '放下一期储备',
        'ip_column': '放肝健康IP栏目',
        'hot_topic': '放热搜话题',
    }
    return label_map.get((target or '').strip(), target or '待确认')


def _trend_planning_score(note):
    base = note.hot_score or _trend_score(note)
    keyword_text = ' '.join([note.keyword or '', note.title or '', note.summary or ''])
    score = int(base)
    if any(token in keyword_text for token in ['怎么', '如何', '怎么办', '能不能', '吃什么', '做什么', '要不要']):
        score += 40
    if any(token in keyword_text for token in ['体检', '转氨酶', '肝弹', 'FibroScan', '脂肪肝']):
        score += 30
    if any(token in keyword_text for token in ['中医', '肝气郁结', '肝郁', '情绪', '失眠']):
        score += 20
    return score


def _recommended_trend_route(note):
    text = ' '.join([note.keyword or '', note.title or '', note.summary or '']).strip()
    lane_key = _infer_trend_lane_key(text)
    score = _trend_planning_score(note)
    reasons = []
    target = 'next_activity'

    if (note.source_template_key or '') == 'xhs_hot_queries':
        target = 'hot_topic'
        reasons.append('当前是平台热搜/相关搜索词，更适合进热搜话题页')
    elif lane_key in {'tcm_conditioning', 'emotion_mindbody', 'family_care', 'women_health'}:
        target = 'ip_column'
        reasons.append('这类内容更适合沉淀到肝健康IP栏目，做长期人设和栏目内容')
    elif any(token in text for token in ['怎么', '如何', '怎么办', '吃什么', '做什么', '运动', '减肥', '体检', '转氨酶', '肝弹']):
        target = 'next_activity'
        reasons.append('这是高搜索问题型内容，更适合先作为候选话题进入下一期储备')
    if score >= 220 and lane_key in {'report_interpretation', 'fatty_liver_management', 'liver_care_habits'}:
        target = 'current_topic'
        reasons.append('热度和执行价值都较高，适合直接进入当前话题广场')
    if not reasons:
        reasons.append('建议先进入下一期储备，由你确认后再决定是否发布')

    confidence = 'low'
    if score >= 260:
        confidence = 'high'
    elif score >= 180:
        confidence = 'medium'

    return {
        'target': target,
        'target_label': _trend_route_target_label(target),
        'score': score,
        'confidence': confidence,
        'reason': '；'.join(reasons),
        'lane_key': lane_key,
        'lane_label': _trend_lane_label(lane_key),
        'persona_key': _infer_trend_persona_key(text),
        'persona_label': _trend_persona_label(_infer_trend_persona_key(text)),
    }


def _resolve_trend_route_target(note, target=''):
    current_target = (target or '').strip()
    if current_target in {'', 'recommended', 'auto'}:
        return _recommended_trend_route(note).get('target') or 'next_activity'
    return current_target


def _build_trend_topic_payload(note, activity_id, quota=None):
    lane_key = _infer_trend_lane_key(' '.join([note.keyword or '', note.title or '', note.summary or '']))
    persona_key = _infer_trend_persona_key(' '.join([note.keyword or '', note.title or '', note.summary or '']))
    lane_label = _trend_lane_label(lane_key)
    persona_label = _trend_persona_label(persona_key)
    return Topic(
        activity_id=activity_id,
        topic_name=(note.title or note.keyword or '热点话题')[:200],
        keywords=(note.keyword or '').strip()[:500],
        direction='\n'.join(filter(None, [
            f'来源：{note.source_platform or "小红书"} 热点/爆款笔记',
            f'内容方向：{lane_label}',
            f'推荐人设：{persona_label}',
            f'建议先围绕用户高搜索问题切入，再自然带出肝健康管理建议。',
            f'摘要：{note.summary or ""}',
        ])).strip(),
        reference_content=note.summary or '',
        reference_link=note.link or '',
        writing_example=f'参考当前爆款问题的标题逻辑和互动结构，围绕「{note.keyword or note.title or "肝健康"}」重新改写。',
        quota=_normalize_quota(quota),
        group_num='自动化热点',
        pool_status='formal',
        source_type='trend_note',
        source_ref_id=note.id,
        published_at=datetime.now(),
    )


def _build_trend_topic_idea_payload(note, activity_id=None):
    lane_key = _infer_trend_lane_key(' '.join([note.keyword or '', note.title or '', note.summary or '']))
    persona_key = _infer_trend_persona_key(' '.join([note.keyword or '', note.title or '', note.summary or '']))
    lane_label = _trend_lane_label(lane_key)
    persona_label = _trend_persona_label(persona_key)
    return TopicIdea(
        activity_id=activity_id,
        topic_title=(note.title or note.keyword or '热点候选话题')[:200],
        keywords=(note.keyword or '').strip()[:500],
        angle=f'围绕「{note.keyword or note.title or "肝健康"}」这个高搜索问题，按 {lane_label} 切入，内容语气更适合 {persona_label}。',
        content_type='轻科普问答型',
        persona=persona_label[:50],
        soft_insertion=_detect_soft_insertion(' '.join([note.keyword or '', note.title or '', note.summary or ''])),
        hot_value=note.hot_score or _trend_score(note),
        source_note_ids=str(note.id),
        source_links=note.link or '',
        copy_prompt=f'请参考这条爆款问题的标题逻辑和互动方式，围绕「{note.keyword or note.title or "肝健康"}」输出一篇可发小红书的肝健康内容草稿。',
        cover_title=_truncate_text(note.title or note.keyword or '热点候选', 18),
        asset_brief=f'优先做与「{lane_label}」匹配的封面和内页，先抓问题，再给用户可执行答案。',
        compliance_note=COMPLIANCE_BASELINE,
        quota=_default_topic_quota(),
        status='pending_review',
    )


def _build_trend_ip_corpus_payload(note):
    payload = _build_trend_note_corpus_payload(note, category='医学科普', source='热点转IP栏目')
    payload['content'] = (
        f"{payload['content']}\n"
        "归宿：肝健康IP栏目。\n"
        "使用方式：适合做长期栏目内容、知识沉淀和IP表达模板，不直接当任务话题。"
    ).strip()
    return payload


def _build_hot_topic_payload(note):
    lane_key = _infer_trend_lane_key(' '.join([note.keyword or '', note.title or '', note.summary or '']))
    persona_key = _infer_trend_persona_key(' '.join([note.keyword or '', note.title or '', note.summary or '']))
    lane_label = _trend_lane_label(lane_key)
    persona_label = _trend_persona_label(persona_key)
    keyword = (note.keyword or note.title or '').strip()
    return {
        'reference_note_id': note.id,
        'title': (note.title or keyword or '热点话题')[:300],
        'keyword': keyword[:200],
        'source_platform': (note.source_platform or '小红书')[:50],
        'source_channel': (note.source_channel or '热点转热搜话题')[:50],
        'reference_url': (note.link or '').strip()[:500],
        'summary': (note.summary or '').strip(),
        'hot_score': note.hot_score or _trend_score(note),
        'lane_key': lane_key[:50],
        'persona_key': persona_key[:50],
        'audience_hint': f'更适合 {persona_label} 来写，方向偏 {lane_label}。',
        'integration_hint': f'这类热点适合先讲用户最关心的问题，再从 {lane_label} 切入，最后自然带出肝健康管理思路。',
        'usage_tip': '建议先做成候选标题和切入角度，再决定是否进入当前话题广场或下一期储备。',
        'status': 'active',
    }


def _route_trend_note_to_target(note, target='corpus_template', activity_id=0, actor='system'):
    current_target = _resolve_trend_route_target(note, target or 'corpus_template')
    if current_target == 'current_topic':
        next_activity_id = max(_safe_int(activity_id, 0), 0)
        if next_activity_id <= 0:
            raise ValueError('请先选择目标活动，用于发布到当前话题广场')
        topic = _build_trend_topic_payload(note, next_activity_id)
        db.session.add(topic)
        db.session.flush()
        _log_operation('route_trend', 'topic', target_id=topic.id, message='热点内容发布到当前话题广场', detail={
            'trend_note_id': note.id,
            'topic_name': topic.topic_name,
            'activity_id': next_activity_id,
            'actor': actor,
        })
        return {'target': current_target, 'item': _serialize_topic(topic)}
    if current_target == 'next_activity':
        next_activity_id = max(_safe_int(activity_id, 0), 0)
        idea = _build_trend_topic_idea_payload(note, activity_id=next_activity_id or None)
        db.session.add(idea)
        db.session.flush()
        _log_operation('route_trend', 'topic_idea', target_id=idea.id, message='热点内容放入下一期候选池', detail={
            'trend_note_id': note.id,
            'topic_title': idea.topic_title,
            'activity_id': next_activity_id,
            'actor': actor,
        })
        return {'target': current_target, 'item': _serialize_topic_idea(idea)}
    if current_target == 'ip_column':
        payload = _build_trend_ip_corpus_payload(note)
        result = _upsert_reference_corpus_entries(
            [payload['reference_url'] or f'trend-note:{note.id}'],
            reference_note_text=payload['content'],
            style_hint='归宿：肝健康IP栏目',
            product_anchor='',
            manual_title=payload['title'],
            category=payload['category'],
            source=payload['source'],
            tags=payload['tags'],
        )
        created = result['created']
        updated = result['updated']
        item = (created + updated)[0] if (created + updated) else None
        _log_operation('route_trend', 'corpus_entry', target_id=item.id if item else None, message='热点内容放入肝健康IP栏目', detail={
            'trend_note_id': note.id,
            'title': payload['title'],
            'actor': actor,
        })
        return {'target': current_target, 'item': _serialize_corpus_entry(item) if item else {}}
    if current_target == 'hot_topic':
        payload = _build_hot_topic_payload(note)
        entry = HotTopicEntry.query.filter_by(reference_note_id=note.id).first()
        if not entry:
            entry = HotTopicEntry(reference_note_id=note.id)
            db.session.add(entry)
        for key, value in payload.items():
            if key == 'reference_note_id':
                continue
            setattr(entry, key, value)
        db.session.flush()
        _log_operation('route_trend', 'hot_topic_entry', target_id=entry.id, message='热点内容放入热搜话题页', detail={
            'trend_note_id': note.id,
            'title': entry.title,
            'actor': actor,
        })
        return {'target': current_target, 'item': _serialize_hot_topic_entry(entry)}
    raise ValueError(f'不支持的目标去向：{current_target}')


def _build_activity_snapshot_payload(activity):
    topics = Topic.query.filter_by(activity_id=activity.id).order_by(Topic.id.asc()).all()
    topic_ids = [topic.id for topic in topics]
    registrations = Registration.query.filter(Registration.topic_id.in_(topic_ids)).order_by(Registration.id.asc()).all() if topic_ids else []
    submissions = Submission.query.join(Registration).filter(Registration.topic_id.in_(topic_ids)).order_by(Submission.id.asc()).all() if topic_ids else []

    registrations_by_topic = defaultdict(list)
    submissions_by_registration = {}
    for reg in registrations:
        registrations_by_topic[reg.topic_id].append(reg)
    for sub in submissions:
        submissions_by_registration[sub.registration_id] = sub

    return {
        'activity': _serialize_activity(activity),
        'topics': [{
            **_serialize_topic(topic),
            'registrations': [{
                'id': reg.id,
                'group_num': reg.group_num or '',
                'name': reg.name or '',
                'phone': reg.phone or '',
                'xhs_account': reg.xhs_account or '',
                'status': reg.status or '',
                'created_at': reg.created_at.strftime('%Y-%m-%d %H:%M:%S') if reg.created_at else '',
                'submission': None if reg.id not in submissions_by_registration else {
                    'id': submissions_by_registration[reg.id].id,
                    'xhs_link': submissions_by_registration[reg.id].xhs_link or '',
                    'xhs_views': submissions_by_registration[reg.id].xhs_views or 0,
                    'xhs_likes': submissions_by_registration[reg.id].xhs_likes or 0,
                    'xhs_favorites': submissions_by_registration[reg.id].xhs_favorites or 0,
                    'xhs_comments': submissions_by_registration[reg.id].xhs_comments or 0,
                    'douyin_link': submissions_by_registration[reg.id].douyin_link or '',
                    'douyin_views': submissions_by_registration[reg.id].douyin_views or 0,
                    'douyin_likes': submissions_by_registration[reg.id].douyin_likes or 0,
                    'douyin_favorites': submissions_by_registration[reg.id].douyin_favorites or 0,
                    'douyin_comments': submissions_by_registration[reg.id].douyin_comments or 0,
                    'video_link': submissions_by_registration[reg.id].video_link or '',
                    'video_views': submissions_by_registration[reg.id].video_views or 0,
                    'video_likes': submissions_by_registration[reg.id].video_likes or 0,
                    'video_favorites': submissions_by_registration[reg.id].video_favorites or 0,
                    'video_comments': submissions_by_registration[reg.id].video_comments or 0,
                    'weibo_link': submissions_by_registration[reg.id].weibo_link or '',
                    'weibo_views': submissions_by_registration[reg.id].weibo_views or 0,
                    'weibo_likes': submissions_by_registration[reg.id].weibo_likes or 0,
                    'weibo_favorites': submissions_by_registration[reg.id].weibo_favorites or 0,
                    'weibo_comments': submissions_by_registration[reg.id].weibo_comments or 0,
                    'note_title': submissions_by_registration[reg.id].note_title or '',
                    'note_content': submissions_by_registration[reg.id].note_content or '',
                    'content_type': submissions_by_registration[reg.id].content_type or '',
                    'keyword_check': bool(submissions_by_registration[reg.id].keyword_check),
                    'created_at': submissions_by_registration[reg.id].created_at.strftime('%Y-%m-%d %H:%M:%S') if submissions_by_registration[reg.id].created_at else '',
                }
            } for reg in registrations_by_topic.get(topic.id, [])]
        } for topic in topics],
        'summary': {
            'topic_count': len(topics),
            'registration_count': len(registrations),
            'submission_count': len(submissions),
        }
    }


def _create_activity_snapshot(activity, snapshot_name=''):
    payload = _build_activity_snapshot_payload(activity)
    snapshot = ActivitySnapshot(
        activity_id=activity.id,
        snapshot_name=snapshot_name or f'{activity.name} 归档快照',
        source_status=activity.status,
        payload=json.dumps(payload, ensure_ascii=False)
    )
    db.session.add(snapshot)
    return snapshot


def _clone_activity(activity, *, name=None, title=None, description=None):
    cloned = Activity(
        name=name or f'{activity.name}-复制',
        title=title or activity.title,
        description=description if description is not None else activity.description,
        status='draft',
        source_type='clone',
        source_activity_id=activity.id,
    )
    db.session.add(cloned)
    db.session.flush()

    topics = Topic.query.filter_by(activity_id=activity.id).order_by(Topic.id.asc()).all()
    for topic in topics:
        db.session.add(Topic(
            activity_id=cloned.id,
            topic_name=topic.topic_name,
            keywords=topic.keywords,
            direction=topic.direction,
            reference_content=topic.reference_content,
            reference_link=topic.reference_link,
            writing_example=topic.writing_example,
            quota=topic.quota,
            group_num=topic.group_num,
            pool_status='formal',
            source_type=topic.source_type or 'manual',
            source_ref_id=topic.source_ref_id,
            source_snapshot_id=topic.source_snapshot_id,
            published_at=topic.published_at,
            filled=0
        ))
    return cloned


def _restore_activity_from_snapshot(snapshot, *, name=None, title=None, description=None):
    payload, _ = _snapshot_payload_and_summary(snapshot)
    if not isinstance(payload, dict):
        payload = {}
    activity_data = payload.get('activity') or {}
    topics_data = payload.get('topics') or []

    restored = Activity(
        name=name or f"{activity_data.get('name') or '历史活动'}-恢复",
        title=title or activity_data.get('title') or '恢复活动',
        description=description if description is not None else activity_data.get('description', ''),
        status='draft',
        source_type='snapshot_restore',
        source_activity_id=snapshot.activity_id,
        source_snapshot_id=snapshot.id,
    )
    db.session.add(restored)
    db.session.flush()

    for topic_data in topics_data:
        db.session.add(Topic(
            activity_id=restored.id,
            topic_name=topic_data.get('topic_name'),
            keywords=topic_data.get('keywords'),
            direction=topic_data.get('direction'),
            reference_content=topic_data.get('reference_content'),
            reference_link=topic_data.get('reference_link'),
            writing_example=topic_data.get('writing_example'),
            quota=_normalize_quota(topic_data.get('quota'), default=_default_topic_quota()),
            group_num=topic_data.get('group_num'),
            filled=0,
            pool_status='formal',
            source_type='snapshot_restore',
            source_ref_id=topic_data.get('id'),
            source_snapshot_id=snapshot.id,
            published_at=datetime.now(),
        ))
    return restored


def _snapshot_payload_and_summary(snapshot):
    payload = {}
    if snapshot.payload:
        try:
            payload = json.loads(snapshot.payload)
        except json.JSONDecodeError:
            payload = {'raw_payload': snapshot.payload}
    summary = payload.get('summary') if isinstance(payload, dict) else {}
    return payload, (summary if isinstance(summary, dict) else {})


def _is_sqlite_backend():
    try:
        return db.engine.url.get_backend_name() == 'sqlite'
    except Exception:
        return app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite')


def _copywriter_env_ready():
    return bool(
        (os.environ.get('COPYWRITER_API_KEY') or '').strip()
        or (os.environ.get('OPENAI_API_KEY') or '').strip()
        or (os.environ.get('DEEPSEEK_API_KEY') or '').strip()
        or (os.environ.get('DOUBAO_API_KEY') or '').strip()
        or (os.environ.get('YUANBAO_API_KEY') or '').strip()
        or (os.environ.get('HUNYUAN_API_KEY') or '').strip()
    )


def _resolve_copywriter_capabilities(payload=None):
    runtime = _copywriter_runtime_config(payload=payload)
    candidates = _copywriter_runtime_candidates(payload=payload)
    return {
        'copywriter_configured': bool(runtime.get('configured')),
        'copywriter_provider': runtime.get('provider') or 'local_fallback',
        'copywriter_api_url': runtime.get('api_url') or '',
        'copywriter_model': runtime.get('model') or '',
        'copywriter_label': runtime.get('label') or '本地兜底生成',
        'api_key_configured': bool(runtime.get('api_key')),
        'server_side_only': True,
        'end_user_needs_vpn': False,
        'fallback_mode': not bool(runtime.get('configured')),
        'thinking_mode': _copywriter_thinking_enabled(runtime),
        'candidate_count': len(candidates),
        'candidate_chain': [{
            'provider': item.get('provider') or '',
            'model': item.get('model') or '',
            'api_url': item.get('api_url') or '',
            'label': item.get('label') or '',
            'source': item.get('source') or '',
            'thinking_mode': bool(item.get('thinking_mode')),
        } for item in candidates[:5]],
    }


def _copywriter_healthcheck(payload=None, timeout_seconds=20):
    runtime = _copywriter_runtime_config(payload=payload)
    candidates = _copywriter_runtime_candidates(payload=payload)
    request_preview = {
        'api_url': runtime.get('api_url') or '',
        'model': runtime.get('model') or '',
        'provider': runtime.get('provider') or 'local_fallback',
        'thinking_mode': _copywriter_thinking_enabled(runtime),
        'candidate_chain': [{
            'provider': item.get('provider') or '',
            'model': item.get('model') or '',
            'api_url': item.get('api_url') or '',
            'source': item.get('source') or '',
            'thinking_mode': bool(item.get('thinking_mode')),
        } for item in candidates[:5]],
        'prompt': (payload or {}).get('prompt_text') or '请用更像真人的小红书口语风，写一句关于肝健康的开头。',
    }
    if not runtime.get('configured'):
        return {
            'enabled': False,
            'ok': False,
            'message': '当前未配置可用的文案模型，系统会回退到本地兜底。',
            'provider': runtime.get('provider') or 'local_fallback',
            'request_preview': request_preview,
            'response_preview': None,
        }
    try:
        result = _call_copywriter(
            [{'role': 'user', 'content': request_preview['prompt']}],
            temperature=0.7,
            top_p=0.9,
            timeout=max(min(timeout_seconds, 60), 5),
            runtime_override=runtime,
        )
        return {
            'enabled': True,
            'ok': bool(result.get('text')),
            'message': '文案模型接口可用' if result.get('text') else '文案模型返回为空',
            'provider': (result.get('runtime') or {}).get('provider') or runtime.get('provider') or '',
            'request_preview': request_preview,
            'response_preview': {
                'text': (result.get('text') or '')[:800],
                'reasoning_text': (result.get('reasoning_text') or '')[:1200],
                'used_model': (result.get('runtime') or {}).get('model') or '',
                'used_api_url': (result.get('runtime') or {}).get('api_url') or '',
                'attempt_errors': result.get('attempt_errors') or [],
            },
        }
    except Exception as exc:
        return {
            'enabled': True,
            'ok': False,
            'message': f'文案模型联调失败：{exc}',
            'provider': runtime.get('provider') or '',
            'request_preview': request_preview,
            'response_preview': {'attempted_candidates': request_preview['candidate_chain']},
        }


KNOWN_DIRECT_METRIC_SOURCE_SUFFIXES = {
    'view_count',
    'impression_cnt',
    'exposure_count',
    'liked_count',
    'collected_count',
    'comment_count',
    'hot_value',
    'hotscore',
    'score',
    'search_cnt',
}


def _metric_source_mode(source):
    raw = (source or '').strip()
    lowered = raw.lower()
    if not lowered:
        return {'key': 'missing', 'label': '未命中'}
    if 'mock_seed' in lowered or lowered.startswith('mock_') or lowered.startswith('mockseed'):
        return {'key': 'mock', 'label': '模拟'}
    if 'derived' in lowered or 'fallback' in lowered:
        return {'key': 'derived', 'label': '推导'}
    for suffix in KNOWN_DIRECT_METRIC_SOURCE_SUFFIXES:
        if lowered == suffix or lowered.endswith(f'.{suffix}'):
            return {'key': 'direct', 'label': '直取'}
    return {'key': 'fuzzy', 'label': '模糊命中'}


def _metric_source_summary(metric_sources, preferred_keys=None):
    if not isinstance(metric_sources, dict):
        return []
    preferred_keys = preferred_keys or ['views', 'exposures', 'hot_value', 'likes', 'favorites', 'comments']
    rows = []
    seen = set()
    for key in list(preferred_keys) + [item for item in metric_sources.keys() if item not in preferred_keys]:
        if key in seen:
            continue
        seen.add(key)
        mode = _metric_source_mode(metric_sources.get(key))
        rows.append({
            'key': key,
            'source': metric_sources.get(key) or '',
            'mode_key': mode['key'],
            'mode_label': mode['label'],
        })
    return rows


def _metric_source_summary_text(metric_sources, preferred_keys=None):
    rows = _metric_source_summary(metric_sources, preferred_keys=preferred_keys)
    items = []
    for row in rows:
        if row['mode_key'] == 'missing':
            continue
        items.append(f"{row['key']}={row['mode_label']}")
    return ' ｜ '.join(items)


def _serialize_topic_idea(idea):
    source_note_id_list = _split_keywords(idea.source_note_ids or '')
    source_links_list = _split_reference_links(idea.source_links or '')
    topic_stub = SimpleNamespace(
        topic_name=idea.topic_title or '',
        keywords=idea.keywords or '',
        direction=idea.angle or '',
        reference_content=idea.asset_brief or '',
        writing_example=idea.copy_prompt or '',
    )
    heuristic = _build_heuristic_strategy_recommendation(topic_stub)
    decision_profile = _build_strategy_decision_profile(topic_stub)
    recommended = dict(heuristic.get('recommended') or {})
    recommended.update(decision_profile)
    product_label, _ = _resolve_copy_product_selection(decision_profile.get('product_key') or 'auto', ' '.join([
        idea.topic_title or '',
        idea.keywords or '',
        idea.angle or '',
    ]))
    copy_analysis = _build_copy_agent_analysis(
        topic_stub,
        persona_label=COPY_PERSONA_OPTIONS.get(decision_profile.get('persona_key') or 'auto', ''),
        scene_label=COPY_SCENE_OPTIONS.get(decision_profile.get('scene_key') or 'auto', ''),
        direction_label=COPY_DIRECTION_OPTIONS.get(decision_profile.get('direction_key') or 'auto', ''),
        product_label=product_label,
        copy_goal=recommended.get('copy_goal') or 'balanced',
    )
    recommended_copy_route = next(
        (item for item in (copy_analysis.get('copy_routes') or []) if item.get('id') == (copy_analysis.get('recommended_copy_route_id') or '')),
        None,
    ) or ((copy_analysis.get('copy_routes') or [None])[0] or {})
    recommended_image_route = next(
        (item for item in (copy_analysis.get('image_routes') or []) if item.get('id') == (copy_analysis.get('recommended_image_route_id') or '')),
        None,
    ) or ((copy_analysis.get('image_routes') or [None])[0] or {})
    image_template_recommendation = _recommend_image_template_payload(
        text=' '.join([
            idea.topic_title or '',
            idea.keywords or '',
            idea.angle or '',
            idea.content_type or '',
            idea.persona or '',
            idea.asset_brief or '',
        ]),
    )
    return {
        'id': idea.id,
        'activity_id': idea.activity_id,
        'topic_title': idea.topic_title,
        'keywords': idea.keywords or '',
        'angle': idea.angle or '',
        'content_type': idea.content_type or '',
        'persona': idea.persona or '',
        'soft_insertion': idea.soft_insertion or '',
        'hot_value': idea.hot_value or 0,
        'source_note_ids': idea.source_note_ids or '',
        'source_note_id_list': source_note_id_list,
        'source_note_count': len(source_note_id_list),
        'source_links': idea.source_links or '',
        'source_links_list': source_links_list,
        'source_link_count': len(source_links_list),
        'copy_prompt': idea.copy_prompt or '',
        'cover_title': idea.cover_title or '',
        'asset_brief': idea.asset_brief or '',
        'compliance_note': idea.compliance_note or '',
        'quota': idea.quota or _default_topic_quota(),
        'status': idea.status,
        'status_label': _topic_idea_status_label(idea.status),
        'review_note': idea.review_note or '',
        'reviewed_at': idea.reviewed_at.strftime('%Y-%m-%d %H:%M:%S') if idea.reviewed_at else '',
        'published_at': idea.published_at.strftime('%Y-%m-%d %H:%M:%S') if idea.published_at else '',
        'published_topic_id': idea.published_topic_id,
        'recommended_goal_key': recommended.get('copy_goal') or 'balanced',
        'recommended_goal_label': COPY_GOAL_OPTIONS.get(recommended.get('copy_goal') or 'balanced', recommended.get('copy_goal') or 'balanced'),
        'recommended_copy_skill_key': recommended.get('copy_skill') or 'auto',
        'recommended_copy_skill_label': COPY_SKILL_OPTIONS.get(recommended.get('copy_skill') or 'auto', recommended.get('copy_skill') or 'auto'),
        'recommended_title_skill_key': recommended.get('title_skill') or 'auto',
        'recommended_title_skill_label': TITLE_SKILL_OPTIONS.get(recommended.get('title_skill') or 'auto', recommended.get('title_skill') or 'auto'),
        'recommended_image_skill_key': recommended.get('image_skill') or 'auto',
        'recommended_image_skill_label': IMAGE_SKILL_OPTIONS.get(recommended.get('image_skill') or 'auto', recommended.get('image_skill') or 'auto'),
        'recommended_persona_label': COPY_PERSONA_OPTIONS.get(decision_profile.get('persona_key') or 'auto', decision_profile.get('persona_key') or ''),
        'recommended_scene_label': COPY_SCENE_OPTIONS.get(decision_profile.get('scene_key') or 'auto', decision_profile.get('scene_key') or ''),
        'recommended_direction_label': COPY_DIRECTION_OPTIONS.get(decision_profile.get('direction_key') or 'auto', decision_profile.get('direction_key') or ''),
        'recommended_product_label': product_label,
        'recommended_reason': heuristic.get('reason') or '',
        'recommended_copy_route_label': recommended_copy_route.get('label') or '',
        'recommended_copy_route_why': recommended_copy_route.get('why') or '',
        'recommended_image_route_label': recommended_image_route.get('label') or '',
        'recommended_image_route_why': recommended_image_route.get('why') or '',
        'execution_bucket_label': '｜'.join(filter(None, [
            idea.soft_insertion or product_label,
            idea.content_type or '',
            COPY_GOAL_OPTIONS.get(recommended.get('copy_goal') or 'balanced', ''),
        ]))[:180],
        **image_template_recommendation,
        'created_at': idea.created_at.strftime('%Y-%m-%d %H:%M:%S') if idea.created_at else '',
    }


CONTENT_BUNDLE_TYPE = 'xhs_content_bundle'
CONTENT_BUNDLE_SCHEMA_VERSION = '2026-04-28'


def _bundle_trend_note_entry(note):
    return {
        'id': note.id,
        'source_platform': note.source_platform or '小红书',
        'source_channel': note.source_channel or '手动导入',
        'source_template_key': note.source_template_key or 'generic_lines',
        'import_batch': note.import_batch or '',
        'keyword': note.keyword or '',
        'topic_category': note.topic_category or '',
        'title': note.title or '',
        'author': note.author or '',
        'link': note.link or '',
        'views': note.views or 0,
        'likes': note.likes or 0,
        'favorites': note.favorites or 0,
        'comments': note.comments or 0,
        'hot_score': note.hot_score or 0,
        'source_rank': note.source_rank or 0,
        'publish_time': _format_datetime(note.publish_time),
        'summary': note.summary or '',
        'raw_payload': _load_json_value(note.raw_payload, {}),
        'pool_status': note.pool_status or 'reserve',
        'created_at': _format_datetime(note.created_at),
    }


def _bundle_topic_idea_entry(idea):
    return {
        'id': idea.id,
        'activity_id': idea.activity_id,
        'topic_title': idea.topic_title or '',
        'keywords': idea.keywords or '',
        'angle': idea.angle or '',
        'content_type': idea.content_type or '',
        'persona': idea.persona or '',
        'soft_insertion': idea.soft_insertion or '',
        'hot_value': idea.hot_value or 0,
        'source_note_ids': idea.source_note_ids or '',
        'source_links': idea.source_links or '',
        'copy_prompt': idea.copy_prompt or '',
        'cover_title': idea.cover_title or '',
        'asset_brief': idea.asset_brief or '',
        'compliance_note': idea.compliance_note or '',
        'quota': idea.quota or _default_topic_quota(),
        'status': idea.status or 'pending_review',
        'review_note': idea.review_note or '',
        'reviewed_at': _format_datetime(idea.reviewed_at),
        'published_at': _format_datetime(idea.published_at),
        'created_at': _format_datetime(idea.created_at),
    }


def _bundle_topic_entry(topic):
    return {
        'id': topic.id,
        'activity_id': topic.activity_id,
        'topic_name': topic.topic_name or '',
        'keywords': topic.keywords or '',
        'direction': topic.direction or '',
        'reference_content': topic.reference_content or '',
        'reference_link': topic.reference_link or '',
        'writing_example': topic.writing_example or '',
        'quota': topic.quota or _default_topic_quota(),
        'group_num': topic.group_num or '',
        'pool_status': topic.pool_status or 'formal',
        'source_type': topic.source_type or 'manual',
        'source_ref_id': topic.source_ref_id,
        'published_at': _format_datetime(topic.published_at),
        'created_at': _format_datetime(topic.created_at),
    }


def _build_content_bundle_payload(*, trends=None, topic_ideas=None, topics=None, activity=None, note=''):
    trend_items = [_bundle_trend_note_entry(item) for item in (trends or [])]
    topic_idea_items = [_bundle_topic_idea_entry(item) for item in (topic_ideas or [])]
    topic_items = [_bundle_topic_entry(item) for item in (topics or [])]
    return {
        'bundle_type': CONTENT_BUNDLE_TYPE,
        'schema_version': CONTENT_BUNDLE_SCHEMA_VERSION,
        'generated_at': _format_datetime(datetime.now()),
        'source_env': _current_runtime_env(),
        'generated_by': _current_actor(),
        'note': (note or '').strip()[:300],
        'activity': None if not activity else {
            'id': activity.id,
            'name': activity.name or '',
            'title': activity.title or '',
            'status': activity.status or '',
        },
        'summary': {
            'trend_count': len(trend_items),
            'topic_idea_count': len(topic_idea_items),
            'topic_count': len(topic_items),
        },
        'items': {
            'trends': trend_items,
            'topic_ideas': topic_idea_items,
            'topics': topic_items,
        },
    }


def _normalize_content_bundle(bundle):
    if not isinstance(bundle, dict):
        raise ValueError('发布包格式不正确')
    items = bundle.get('items') or {}
    if not isinstance(items, dict):
        raise ValueError('发布包 items 节点格式不正确')
    normalized = {
        'bundle_type': (bundle.get('bundle_type') or '').strip(),
        'schema_version': (bundle.get('schema_version') or '').strip(),
        'generated_at': (bundle.get('generated_at') or '').strip(),
        'source_env': (bundle.get('source_env') or '').strip()[:50],
        'generated_by': (bundle.get('generated_by') or '').strip()[:100],
        'note': (bundle.get('note') or '').strip()[:300],
        'activity': bundle.get('activity') if isinstance(bundle.get('activity'), dict) else None,
        'summary': bundle.get('summary') if isinstance(bundle.get('summary'), dict) else {},
        'items': {
            'trends': items.get('trends') if isinstance(items.get('trends'), list) else [],
            'topic_ideas': items.get('topic_ideas') if isinstance(items.get('topic_ideas'), list) else [],
            'topics': items.get('topics') if isinstance(items.get('topics'), list) else [],
        },
    }
    if normalized['bundle_type'] and normalized['bundle_type'] != CONTENT_BUNDLE_TYPE:
        raise ValueError('不是可识别的 xhs 发布包')
    return normalized


TOPIC_IMPORT_HEADER_ALIASES = {
    'topic_title': 'topic_title',
    '话题': 'topic_title',
    '话题名称': 'topic_title',
    '标题': 'topic_title',
    'topic_name': 'topic_title',
    'keywords': 'keywords',
    '关键词': 'keywords',
    'keyword': 'keywords',
    'direction': 'direction',
    '撰写方向': 'direction',
    '方向': 'direction',
    'angle': 'direction',
    'persona': 'persona',
    '人设': 'persona',
    'content_type': 'content_type',
    '内容类型': 'content_type',
    'copy_prompt': 'copy_prompt',
    '文案提示词': 'copy_prompt',
    '提示词': 'copy_prompt',
    'reference_link': 'reference_link',
    '参考链接': 'reference_link',
    '链接': 'reference_link',
    'reference_content': 'reference_content',
    '参考内容': 'reference_content',
    'asset_brief': 'asset_brief',
    '图片说明': 'asset_brief',
    '素材说明': 'asset_brief',
    'compliance_note': 'compliance_note',
    '合规提醒': 'compliance_note',
    'quota': 'quota',
    '名额': 'quota',
    'group_num': 'group_num',
    '组别': 'group_num',
    'soft_insertion': 'soft_insertion',
    '软植入': 'soft_insertion',
}


def _normalize_topic_import_row(raw_row):
    row = {}
    if isinstance(raw_row, dict):
        for key, value in raw_row.items():
            normalized_key = TOPIC_IMPORT_HEADER_ALIASES.get(str(key or '').strip(), '')
            if not normalized_key:
                continue
            row[normalized_key] = str(value or '').strip()
    topic_title = (row.get('topic_title') or '').strip()[:200]
    if not topic_title:
        return {}
    return {
        'topic_title': topic_title,
        'keywords': (row.get('keywords') or '').strip()[:500],
        'direction': (row.get('direction') or '').strip(),
        'persona': (row.get('persona') or '').strip()[:50],
        'content_type': (row.get('content_type') or '').strip()[:50],
        'copy_prompt': (row.get('copy_prompt') or '').strip(),
        'reference_link': (row.get('reference_link') or '').strip(),
        'reference_content': (row.get('reference_content') or '').strip(),
        'asset_brief': (row.get('asset_brief') or '').strip(),
        'compliance_note': (row.get('compliance_note') or '').strip(),
        'quota': _normalize_quota(row.get('quota'), default=_default_topic_quota()),
        'group_num': (row.get('group_num') or '').strip()[:50],
        'soft_insertion': (row.get('soft_insertion') or '').strip()[:100],
    }


def _parse_topic_import_payload(raw_payload):
    payload = (raw_payload or '').strip()
    if not payload:
        return []
    try:
        parsed = json.loads(payload)
        if isinstance(parsed, dict):
            parsed = parsed.get('items') or parsed.get('topics') or parsed.get('topic_ideas') or []
        if isinstance(parsed, list):
            rows = [_normalize_topic_import_row(item) for item in parsed]
            return [row for row in rows if row]
    except Exception:
        pass

    if '\t' in payload or ',' in payload:
        lines = [line for line in payload.splitlines() if line.strip()]
        if lines:
            sample = '\n'.join(lines[:5])
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t')
            except Exception:
                dialect = csv.excel_tab if '\t' in lines[0] else csv.excel
            reader = csv.DictReader(lines, dialect=dialect)
            rows = [_normalize_topic_import_row(item) for item in reader]
            normalized = [row for row in rows if row]
            if normalized:
                return normalized

    normalized = []
    for line in payload.splitlines():
        raw = line.strip()
        if not raw:
            continue
        parts = [part.strip() for part in raw.split('|')]
        if not parts:
            continue
        normalized.append(_normalize_topic_import_row({
            'topic_title': parts[0] if len(parts) > 0 else '',
            'keywords': parts[1] if len(parts) > 1 else '',
            'direction': parts[2] if len(parts) > 2 else '',
            'persona': parts[3] if len(parts) > 3 else '',
            'content_type': parts[4] if len(parts) > 4 else '',
            'copy_prompt': parts[5] if len(parts) > 5 else '',
            'reference_link': parts[6] if len(parts) > 6 else '',
            'quota': parts[7] if len(parts) > 7 else '',
            'group_num': parts[8] if len(parts) > 8 else '',
        }))
    return [row for row in normalized if row]


def _sheet_topic_group_label(sheet_name=''):
    name = (sheet_name or '').strip()
    if '软肝片' in name:
        return '软肝片话题池'
    if 'FS' in name or 'FibroScan' in name or '福波看' in name:
        return 'FibroScan话题池'
    return '批量导入'


def _sheet_product_hint(sheet_name=''):
    name = (sheet_name or '').strip()
    if '软肝片' in name:
        return '软肝片'
    if 'FS' in name or 'FibroScan' in name or '福波看' in name:
        return 'FibroScan福波看'
    return ''


def _first_meaningful_line(text=''):
    for part in str(text or '').replace('\r', '\n').split('\n'):
        clean = part.strip()
        if clean:
            return clean[:120]
    return ''


def _topic_core_from_search_text(text=''):
    line = _first_meaningful_line(text)
    if not line:
        return ''
    for sep in [' / ', '/', '｜', '|']:
        if sep in line:
            head = line.split(sep)[0].strip()
            if head:
                return head[:80]
    return line[:80]


def _keywords_from_tags_and_topic(topic_core='', tags=''):
    values = []
    if topic_core:
        values.append(topic_core)
    for part in re.split(r'[#\s,，、/]+', tags or ''):
        token = (part or '').strip()
        if not token or len(token) <= 1:
            continue
        values.append(token)
    deduped = []
    for item in values:
        if item not in deduped:
            deduped.append(item)
    return ','.join(deduped[:12])[:500]


def _build_topic_import_rows_from_workbook_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_index = None
    for idx, row in enumerate(rows):
        first_cells = [str(cell or '').strip() for cell in row[:8]]
        joined = ' '.join(first_cells)
        if '笔记类型' in joined and '撰写说明' in joined:
            header_index = idx
            break
    if header_index is None:
        return []

    header_row = [str(cell or '').strip() for cell in rows[header_index]]
    header_aliases = {
        'search_text': ['C端搜索需求词', '小红书C端 / 热门搜索词'],
        'search_keyword': ['小红书C端搜索关键词 / （来自小红书数据平台）', '搜索关键词'],
        'note_type': ['笔记类型', '笔记类型方向'],
        'tag_text': ['话题标签'],
        'writing': ['撰写说明'],
        'links': ['参考笔记链接', '笔记参考链接', '笔记参考链接 / ', '参考笔记链接 / '],
        'sample': ['撰写示例', '笔记撰写示例'],
        'compliance': ['特别注意（合规说明）'],
    }
    index_map = {}
    for idx, header in enumerate(header_row):
        for key, aliases in header_aliases.items():
            if key in index_map:
                continue
            if any(alias in header for alias in aliases):
                index_map[key] = idx
                break

    intro_lines = []
    for row in rows[:header_index]:
        parts = [str(cell or '').strip() for cell in row[:3] if str(cell or '').strip()]
        if parts:
            intro_lines.append(' / '.join(parts))
    shared_compliance = '\n'.join(intro_lines).strip()

    current_search_text = ''
    normalized_rows = []
    for row in rows[header_index + 1:]:
        cells = [str(cell or '').strip() for cell in row]
        if not any(cells):
            continue
        search_text = cells[index_map.get('search_text', 0)] if index_map.get('search_text', 0) < len(cells) else ''
        search_keyword = cells[index_map.get('search_keyword', 1)] if index_map.get('search_keyword', 1) < len(cells) else ''
        note_type = cells[index_map.get('note_type', 1)] if index_map.get('note_type', 1) < len(cells) else ''
        tag_text = cells[index_map.get('tag_text', 3)] if index_map.get('tag_text', 3) < len(cells) else ''
        writing = cells[index_map.get('writing', 4)] if index_map.get('writing', 4) < len(cells) else ''
        links = cells[index_map.get('links', 5)] if index_map.get('links', 5) < len(cells) else ''
        sample = cells[index_map.get('sample', 6)] if index_map.get('sample', 6) < len(cells) else ''
        compliance = cells[index_map.get('compliance', 7)] if index_map.get('compliance', 7) < len(cells) else ''
        if search_text:
            current_search_text = search_text
        if not current_search_text or not any([note_type, writing, links, sample]):
            continue

        topic_core = _topic_core_from_search_text(current_search_text)
        note_label = _topic_core_from_search_text(note_type) or '内容执行'
        topic_title = f'{topic_core}｜{note_label}'.strip('｜')[:200]
        if not topic_title:
            continue
        product_hint = _sheet_product_hint(ws.title)
        combined_direction = '\n'.join(filter(None, [
            current_search_text,
            f'笔记类型：{note_type}' if note_type else '',
            writing,
        ])).strip()
        combined_reference = '\n'.join(filter(None, [
            sample,
            f'产品方向：{product_hint}' if product_hint else '',
        ])).strip()
        combined_compliance = '\n'.join(filter(None, [
            shared_compliance,
            compliance,
        ])).strip()
        normalized_rows.append({
            'topic_title': topic_title,
            'keywords': _keywords_from_tags_and_topic(topic_core, tag_text),
            'direction': '\n'.join(filter(None, [combined_direction, search_keyword])).strip(),
            'persona': '',
            'content_type': note_label[:50],
            'copy_prompt': sample,
            'reference_link': '\n'.join(_split_reference_links(links)) if links else '',
            'reference_content': combined_reference,
            'asset_brief': writing,
            'compliance_note': combined_compliance,
            'quota': _default_topic_quota(),
            'group_num': _sheet_topic_group_label(ws.title),
            'soft_insertion': product_hint,
        })
    return normalized_rows


def _build_topic_import_rows_from_simple_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_index = None
    for idx, row in enumerate(rows):
        cells = [str(cell or '').strip() for cell in row]
        joined = ' '.join(cells)
        if '话题' in joined and '笔记类型' in joined and '爆款笔记链接' in joined:
            header_index = idx
            break
    if header_index is None:
        return []

    header_row = [str(cell or '').strip() for cell in rows[header_index]]
    def find_col(names, default=0):
        for idx, header in enumerate(header_row):
            if any(name in header for name in names):
                return idx
        return default

    topic_col = find_col(['话题'], 1)
    type_col = find_col(['笔记类型'], 2)
    angle_col = find_col(['撰写角度'], 3)
    writing_col = find_col(['撰写说明'], 4)
    sample_col = find_col(['撰写示例'], 5)
    keyword_col = find_col(['关键词'], 6)
    link_col = find_col(['爆款笔记链接', '参考链接', '笔记链接'], 7)

    intro = ''
    for row in rows[:header_index]:
        cells = [str(cell or '').strip() for cell in row if str(cell or '').strip()]
        if cells:
            intro = cells[0]
            break
    product_hint = '软肝片' if '软肝片' in intro or '鳖甲' in intro else _sheet_product_hint(ws.title)
    group_label = intro[:80] if intro else _sheet_topic_group_label(ws.title)
    normalized_rows = []
    current = None

    def cell(cells, idx):
        return cells[idx] if idx < len(cells) else ''

    def flush_current():
        if not current:
            return
        links = _split_reference_links(current.pop('_links_raw', ''))
        current['reference_link'] = '\n'.join(links)
        normalized_rows.append(current)

    for row in rows[header_index + 1:]:
        cells = [str(cell or '').strip() for cell in row]
        if not any(cells):
            continue
        topic = cell(cells, topic_col)
        note_type = cell(cells, type_col)
        raw_links = cell(cells, link_col)
        if topic:
            flush_current()
            topic_core = _topic_core_from_search_text(topic)
            type_label = _topic_core_from_search_text(note_type) or '内容执行'
            keywords = _keywords_from_tags_and_topic(topic_core, cell(cells, keyword_col))
            angle = cell(cells, angle_col)
            writing = cell(cells, writing_col)
            sample = cell(cells, sample_col)
            current = {
                'topic_title': f'{topic_core}｜{type_label}'.strip('｜')[:200],
                'keywords': keywords,
                'direction': '\n'.join(filter(None, [angle, writing])).strip(),
                'persona': angle[:80],
                'content_type': type_label[:50],
                'copy_prompt': sample,
                'reference_link': '',
                'reference_content': sample,
                'asset_brief': writing,
                'compliance_note': COMPLIANCE_BASELINE,
                'quota': _default_topic_quota(),
                'group_num': group_label or '批量导入',
                'soft_insertion': product_hint,
                '_links_raw': raw_links,
            }
        elif current and raw_links:
            current['_links_raw'] = '\n'.join(filter(None, [current.get('_links_raw', ''), raw_links]))
    flush_current()
    return normalized_rows


def _parse_topic_import_file(filename='', file_bytes=b''):
    lower_name = (filename or '').lower()
    if lower_name.endswith('.xlsx') or lower_name.endswith('.xlsm'):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        all_rows = []
        for ws in workbook.worksheets:
            simple_rows = _build_topic_import_rows_from_simple_sheet(ws)
            all_rows.extend(simple_rows or _build_topic_import_rows_from_workbook_sheet(ws))
        return all_rows
    try:
        raw_text = file_bytes.decode('utf-8')
    except UnicodeDecodeError:
        raw_text = file_bytes.decode('utf-8-sig', errors='ignore')
    return _parse_topic_import_payload(raw_text)


def _preview_topic_import_rows(rows, *, activity_id=0, target_type='topic_idea'):
    normalized_target = (target_type or 'topic_idea').strip() or 'topic_idea'
    duplicates = []
    preview_items = []
    for row in rows[:50]:
        existing = None
        if normalized_target == 'topic':
            if activity_id:
                existing = Topic.query.filter_by(activity_id=activity_id, topic_name=row['topic_title']).order_by(Topic.id.desc()).first()
        else:
            existing = TopicIdea.query.filter_by(
                activity_id=activity_id or None,
                topic_title=row['topic_title'],
                keywords=row['keywords'],
            ).order_by(TopicIdea.id.desc()).first()
            if existing is None:
                existing = TopicIdea.query.filter_by(
                    topic_title=row['topic_title'],
                    keywords=row['keywords'],
                ).order_by(TopicIdea.id.desc()).first()
        if existing:
            duplicates.append({
                'topic_title': row['topic_title'],
                'existing_id': existing.id,
                'target_type': normalized_target,
            })
        topic_text = ' '.join(filter(None, [
            row.get('topic_title') or '',
            row.get('keywords') or '',
            row.get('direction') or '',
            row.get('reference_content') or '',
        ]))
        content_type_text = ' '.join(filter(None, [
            row.get('content_type') or '',
            row.get('direction') or '',
            row.get('asset_brief') or '',
        ]))
        topic_stub = SimpleNamespace(
            topic_name=row.get('topic_title') or '',
            keywords=row.get('keywords') or '',
            direction=row.get('direction') or '',
            reference_content=row.get('reference_content') or '',
            writing_example=row.get('copy_prompt') or '',
        )
        heuristic = _build_heuristic_strategy_recommendation(topic_stub)
        decision_profile = _build_strategy_decision_profile(topic_stub)
        recommended = dict(heuristic.get('recommended') or {})
        recommended.update(decision_profile)
        if any(token in content_type_text for token in ['大字报', '互动型']):
            recommended.update({
                'copy_goal': 'comment_engagement',
                'copy_skill': 'discussion_hook',
                'title_skill': 'question_gap',
                'image_skill': 'high_click_cover',
            })
        elif any(token in content_type_text for token in ['深度科普', '医疗行业背景', '图表', '横向测评']):
            recommended.update({
                'copy_goal': 'trust_building',
                'copy_skill': 'report_interpretation',
                'title_skill': 'checklist_collect',
                'image_skill': 'report_decode',
            })
        elif any(token in content_type_text for token in ['备忘录', '清单']):
            recommended.update({
                'copy_goal': 'save_value',
                'copy_skill': 'practical_checklist',
                'title_skill': 'checklist_collect',
                'image_skill': 'save_worthy_cards',
            })
        elif any(token in content_type_text for token in ['产品实拍', '实拍']):
            recommended.update({
                'copy_goal': 'balanced',
                'copy_skill': 'story_empathy',
                'title_skill': 'emotional_diary',
                'image_skill': 'story_atmosphere',
            })
        copy_route_analysis = _build_copy_agent_analysis(
            topic_stub,
            persona_label=COPY_PERSONA_OPTIONS.get(decision_profile.get('persona_key') or 'auto', ''),
            scene_label=COPY_SCENE_OPTIONS.get(decision_profile.get('scene_key') or 'auto', ''),
            direction_label=COPY_DIRECTION_OPTIONS.get(decision_profile.get('direction_key') or 'auto', ''),
            product_label=_resolve_copy_product_selection(decision_profile.get('product_key') or 'auto', topic_text)[0],
            copy_goal=recommended.get('copy_goal') or 'balanced',
        )
        recommended_copy_route = next(
            (item for item in (copy_route_analysis.get('copy_routes') or []) if item.get('id') == (copy_route_analysis.get('recommended_copy_route_id') or '')),
            None,
        ) or ((copy_route_analysis.get('copy_routes') or [None])[0] or {})
        recommended_image_route = next(
            (item for item in (copy_route_analysis.get('image_routes') or []) if item.get('id') == (copy_route_analysis.get('recommended_image_route_id') or '')),
            None,
        ) or ((copy_route_analysis.get('image_routes') or [None])[0] or {})
        image_plan_payload = _build_image_agent_analysis_payload(
            topic_stub,
            selected_content='\n'.join(filter(None, [
                row.get('direction') or '',
                row.get('reference_content') or '',
                row.get('copy_prompt') or '',
            ]))[:4000],
            title_hint=(row.get('topic_title') or '')[:200],
            preferred_route=recommended_image_route,
        )
        recommended_plan = next(
            (item for item in (image_plan_payload.get('plans') or []) if item.get('id') == (image_plan_payload.get('recommended_plan_id') or '')),
            None,
        ) or ((image_plan_payload.get('plans') or [None])[0] or {})
        preview_items.append({
            **row,
            'duplicate': bool(existing),
            'reference_link_count': len(_split_reference_links(row.get('reference_link') or '')),
            'reference_links_preview': _split_reference_links(row.get('reference_link') or '')[:3],
            'agent_summary': heuristic.get('reason') or '',
            'recommended_goal': COPY_GOAL_OPTIONS.get(recommended.get('copy_goal') or 'balanced', recommended.get('copy_goal') or 'balanced'),
            'recommended_copy_skill': COPY_SKILL_OPTIONS.get(recommended.get('copy_skill') or 'auto', recommended.get('copy_skill') or 'auto'),
            'recommended_title_skill': TITLE_SKILL_OPTIONS.get(recommended.get('title_skill') or 'auto', recommended.get('title_skill') or 'auto'),
            'recommended_image_skill': IMAGE_SKILL_OPTIONS.get(recommended.get('image_skill') or 'auto', recommended.get('image_skill') or 'auto'),
            'recommended_persona': COPY_PERSONA_OPTIONS.get(decision_profile.get('persona_key') or 'auto', decision_profile.get('persona_key') or ''),
            'recommended_scene': COPY_SCENE_OPTIONS.get(decision_profile.get('scene_key') or 'auto', decision_profile.get('scene_key') or ''),
            'recommended_direction': COPY_DIRECTION_OPTIONS.get(decision_profile.get('direction_key') or 'auto', decision_profile.get('direction_key') or ''),
            'recommended_copy_route_label': recommended_copy_route.get('label') or '',
            'recommended_copy_route_why': recommended_copy_route.get('why') or '',
            'recommended_image_route_label': recommended_image_route.get('label') or '',
            'recommended_image_route_why': recommended_image_route.get('why') or '',
            'recommended_cover_style': recommended_plan.get('style_label') or recommended_plan.get('label') or '',
            'recommended_cover_fit': recommended_plan.get('cover_fit_label') or '',
        })
    return {
        'count': len(rows),
        'duplicate_count': len(duplicates),
        'duplicates': duplicates[:20],
        'items': preview_items[:20],
    }


def _import_topic_rows(rows, *, activity_id=0, target_type='topic_idea'):
    normalized_target = (target_type or 'topic_idea').strip() or 'topic_idea'
    created = []
    updated = []
    for row in rows:
        if normalized_target == 'topic':
            existing = Topic.query.filter_by(activity_id=activity_id, topic_name=row['topic_title']).order_by(Topic.id.desc()).first()
            if existing is None:
                existing = Topic(activity_id=activity_id)
                db.session.add(existing)
                target_rows = created
            else:
                target_rows = updated
            existing.topic_name = row['topic_title']
            existing.keywords = row['keywords']
            existing.direction = row['direction']
            existing.reference_content = row['reference_content'] or row['asset_brief']
            existing.reference_link = _compact_reference_links_for_topic(row['reference_link'])
            existing.writing_example = row['copy_prompt']
            existing.quota = row['quota']
            existing.group_num = row['group_num'] or '批量导入'
            existing.pool_status = 'formal'
            existing.source_type = 'manual_import'
            existing.published_at = datetime.now()
            db.session.flush()
            target_rows.append({'id': existing.id, 'topic_title': existing.topic_name})
            continue

        existing = TopicIdea.query.filter_by(
            activity_id=activity_id or None,
            topic_title=row['topic_title'],
            keywords=row['keywords'],
        ).order_by(TopicIdea.id.desc()).first()
        if existing is None:
            existing = TopicIdea.query.filter_by(
                topic_title=row['topic_title'],
                keywords=row['keywords'],
            ).order_by(TopicIdea.id.desc()).first()
        if existing is None:
            existing = TopicIdea(activity_id=activity_id or None)
            db.session.add(existing)
            target_rows = created
        else:
            target_rows = updated
        existing.activity_id = activity_id or None
        existing.topic_title = row['topic_title']
        existing.keywords = row['keywords']
        existing.angle = row['direction']
        existing.content_type = row['content_type'] or '轻科普问答型'
        existing.persona = row['persona']
        existing.soft_insertion = row['soft_insertion']
        existing.copy_prompt = row['copy_prompt']
        existing.source_links = row['reference_link']
        existing.asset_brief = row['asset_brief'] or row['reference_content']
        existing.compliance_note = row['compliance_note'] or COMPLIANCE_BASELINE
        existing.quota = row['quota']
        existing.status = 'pending_review'
        db.session.flush()
        target_rows.append({'id': existing.id, 'topic_title': existing.topic_title})
    return {
        'created': created,
        'updated': updated,
    }


def _extract_content_bundle_from_request():
    if request.files:
        file_storage = request.files.get('file')
        if file_storage and file_storage.filename:
            try:
                return _normalize_content_bundle(json.loads(file_storage.read().decode('utf-8')))
            except Exception as exc:
                raise ValueError(f'发布包文件解析失败：{exc}')
    data = request.get_json(silent=True) if request.is_json else request.form
    if not data:
        raise ValueError('未提供发布包内容')
    direct_bundle = data.get('bundle') if isinstance(data, dict) else None
    if isinstance(direct_bundle, dict):
        return _normalize_content_bundle(direct_bundle)
    if isinstance(data, dict) and data.get('bundle_type'):
        return _normalize_content_bundle(data)
    raw_payload = (data.get('raw_payload') or '').strip() if isinstance(data, dict) else ''
    if raw_payload:
        try:
            return _normalize_content_bundle(json.loads(raw_payload))
        except Exception as exc:
            raise ValueError(f'发布包 JSON 解析失败：{exc}')
    raise ValueError('未提供发布包内容')


def _preview_content_bundle_import(bundle, *, target_activity_id=0, import_topics=False):
    preview = {
        'summary': {
            'trend_count': len(bundle['items']['trends']),
            'topic_idea_count': len(bundle['items']['topic_ideas']),
            'topic_count': len(bundle['items']['topics']),
        },
        'duplicates': {
            'trends': [],
            'topic_ideas': [],
            'topics': [],
        },
        'warnings': [],
    }

    for row in bundle['items']['trends'][:100]:
        link = (row.get('link') or '').strip()
        title = (row.get('title') or '').strip()
        existing = None
        if link:
            existing = TrendNote.query.filter_by(link=link).order_by(TrendNote.id.desc()).first()
        if not existing and title:
            existing = TrendNote.query.filter_by(title=title).order_by(TrendNote.id.desc()).first()
        if existing:
            preview['duplicates']['trends'].append({
                'bundle_title': title,
                'existing_id': existing.id,
                'existing_title': existing.title or '',
            })

    for row in bundle['items']['topic_ideas'][:100]:
        topic_title = (row.get('topic_title') or '').strip()
        keywords = (row.get('keywords') or '').strip()
        source_links = (row.get('source_links') or '').strip()
        existing = TopicIdea.query.filter_by(
            topic_title=topic_title,
            keywords=keywords,
            source_links=source_links,
        ).order_by(TopicIdea.id.desc()).first()
        if existing:
            preview['duplicates']['topic_ideas'].append({
                'bundle_title': topic_title,
                'existing_id': existing.id,
                'existing_status': existing.status or '',
            })

    if import_topics and not target_activity_id and bundle['items']['topics']:
        preview['warnings'].append('发布正式话题前需要先选择目标活动期数。')
    if bundle.get('source_env') and bundle.get('source_env') == _current_runtime_env():
        preview['warnings'].append('当前导入包与本环境标记一致，执行前请再次确认是不是同环境回放。')
    if target_activity_id:
        for row in bundle['items']['topics'][:100]:
            topic_name = (row.get('topic_name') or '').strip()
            existing = Topic.query.filter_by(activity_id=target_activity_id, topic_name=topic_name).order_by(Topic.id.desc()).first()
            if existing:
                preview['duplicates']['topics'].append({
                    'bundle_title': topic_name,
                    'existing_id': existing.id,
                    'activity_id': target_activity_id,
                })

    return preview


def _import_content_bundle(bundle, *, target_activity_id=0, import_trends=True, import_topic_ideas=True, import_topics=False, preserve_review_status=True):
    created = {
        'trends': [],
        'topic_ideas': [],
        'topics': [],
    }
    updated = {
        'trends': [],
        'topic_ideas': [],
        'topics': [],
    }

    if import_trends:
        for row in bundle['items']['trends']:
            link = (row.get('link') or '').strip()
            title = (row.get('title') or '').strip()[:300]
            existing = None
            if link:
                existing = TrendNote.query.filter_by(link=link).order_by(TrendNote.id.desc()).first()
            if not existing and title:
                existing = TrendNote.query.filter_by(title=title).order_by(TrendNote.id.desc()).first()
            if existing is None:
                existing = TrendNote()
                db.session.add(existing)
                target_list = created['trends']
            else:
                target_list = updated['trends']
            existing.source_platform = (row.get('source_platform') or '小红书').strip()[:50]
            existing.source_channel = (row.get('source_channel') or '发布包导入').strip()[:50]
            existing.source_template_key = (row.get('source_template_key') or 'generic_lines').strip()[:50]
            existing.import_batch = (row.get('import_batch') or '').strip()[:100]
            existing.keyword = (row.get('keyword') or '').strip()[:200]
            existing.topic_category = (row.get('topic_category') or '').strip()[:100]
            existing.title = title
            existing.author = (row.get('author') or '').strip()[:100]
            existing.link = link[:500]
            existing.views = _safe_int(row.get('views'), 0)
            existing.likes = _safe_int(row.get('likes'), 0)
            existing.favorites = _safe_int(row.get('favorites'), 0)
            existing.comments = _safe_int(row.get('comments'), 0)
            existing.hot_score = _safe_int(row.get('hot_score'), 0)
            existing.source_rank = _safe_int(row.get('source_rank'), 0)
            existing.publish_time = _parse_datetime(row.get('publish_time'))
            existing.summary = (row.get('summary') or '').strip()
            existing.raw_payload = json.dumps(row.get('raw_payload') or {}, ensure_ascii=False)
            existing.pool_status = (row.get('pool_status') or 'reserve').strip()[:20] or 'reserve'
            db.session.flush()
            target_list.append({'id': existing.id, 'title': existing.title or ''})

    if import_topic_ideas:
        for row in bundle['items']['topic_ideas']:
            topic_title = (row.get('topic_title') or '').strip()[:200]
            keywords = (row.get('keywords') or '').strip()[:500]
            source_links = (row.get('source_links') or '').strip()
            existing = TopicIdea.query.filter_by(
                topic_title=topic_title,
                keywords=keywords,
                source_links=source_links,
            ).order_by(TopicIdea.id.desc()).first()
            if existing is None:
                existing = TopicIdea()
                db.session.add(existing)
                target_list = created['topic_ideas']
            else:
                target_list = updated['topic_ideas']
            bundle_status = (row.get('status') or 'pending_review').strip()
            normalized_status = 'pending_review'
            if preserve_review_status:
                if bundle_status in {'approved', 'published'}:
                    normalized_status = 'approved'
                elif bundle_status in {'rejected', 'archived'}:
                    normalized_status = bundle_status
            existing.activity_id = target_activity_id or (_safe_int(row.get('activity_id'), 0) or None)
            existing.topic_title = topic_title
            existing.keywords = keywords
            existing.angle = (row.get('angle') or '').strip()
            existing.content_type = (row.get('content_type') or '').strip()[:50]
            existing.persona = (row.get('persona') or '').strip()[:50]
            existing.soft_insertion = (row.get('soft_insertion') or '').strip()[:100]
            existing.hot_value = _safe_int(row.get('hot_value'), 0)
            existing.source_note_ids = (row.get('source_note_ids') or '').strip()[:200]
            existing.source_links = source_links
            existing.copy_prompt = (row.get('copy_prompt') or '').strip()
            existing.cover_title = (row.get('cover_title') or '').strip()[:120]
            existing.asset_brief = (row.get('asset_brief') or '').strip()
            existing.compliance_note = (row.get('compliance_note') or '').strip()
            existing.quota = _normalize_quota(row.get('quota'), default=_default_topic_quota())
            existing.status = normalized_status
            existing.review_note = (row.get('review_note') or '').strip()
            existing.reviewed_at = _parse_datetime(row.get('reviewed_at'))
            existing.published_at = None
            existing.published_topic_id = None
            db.session.flush()
            target_list.append({'id': existing.id, 'title': existing.topic_title or '', 'status': existing.status or ''})

    if import_topics:
        if not target_activity_id:
            raise ValueError('导入正式话题时必须指定目标活动期数')
        for row in bundle['items']['topics']:
            topic_name = (row.get('topic_name') or '').strip()[:200]
            existing = Topic.query.filter_by(
                activity_id=target_activity_id,
                topic_name=topic_name,
            ).order_by(Topic.id.desc()).first()
            if existing is None:
                existing = Topic(activity_id=target_activity_id)
                db.session.add(existing)
                target_list = created['topics']
            else:
                target_list = updated['topics']
            existing.activity_id = target_activity_id
            existing.topic_name = topic_name
            existing.keywords = (row.get('keywords') or '').strip()[:500]
            existing.direction = (row.get('direction') or '').strip()
            existing.reference_content = (row.get('reference_content') or '').strip()
            existing.reference_link = (row.get('reference_link') or '').strip()[:500]
            existing.writing_example = (row.get('writing_example') or '').strip()
            existing.quota = _normalize_quota(row.get('quota'), default=_default_topic_quota())
            existing.group_num = (row.get('group_num') or '内容发布包').strip()[:50]
            existing.pool_status = 'formal'
            existing.source_type = (row.get('source_type') or 'content_bundle').strip()[:30]
            existing.source_ref_id = _safe_int(row.get('source_ref_id'), 0) or None
            existing.published_at = _parse_datetime(row.get('published_at')) or datetime.now()
            db.session.flush()
            target_list.append({'id': existing.id, 'title': existing.topic_name or '', 'activity_id': existing.activity_id})

    return {
        'created': created,
        'updated': updated,
    }


def _creator_post_interactions(post):
    return (post.likes or 0) + (post.favorites or 0) + (post.comments or 0)


def _creator_post_score(post):
    return (
        post.views or 0,
        _creator_post_interactions(post),
        post.follower_delta or 0,
        post.exposures or 0,
    )


def _infer_viral_post(views=0, likes=0, favorites=0, comments=0, exposures=0):
    interactions = (likes or 0) + (favorites or 0) + (comments or 0)
    return (views or 0) >= 10000 or (exposures or 0) >= 30000 or interactions >= 1000


def _serialize_creator_post(post):
    interactions = _creator_post_interactions(post)
    payload = _load_json_value(post.raw_payload, {})
    metric_sources = payload.get('metric_sources') if isinstance(payload, dict) else {}
    return {
        'id': post.id,
        'creator_account_id': post.creator_account_id,
        'platform_post_id': post.platform_post_id or '',
        'registration_id': post.registration_id,
        'topic_id': post.topic_id,
        'submission_id': post.submission_id,
        'title': post.title,
        'post_url': post.post_url or '',
        'publish_time': post.publish_time.strftime('%Y-%m-%d %H:%M:%S') if post.publish_time else '',
        'topic_title': post.topic_title or '',
        'views': post.views or 0,
        'exposures': post.exposures or 0,
        'likes': post.likes or 0,
        'favorites': post.favorites or 0,
        'comments': post.comments or 0,
        'shares': post.shares or 0,
        'follower_delta': post.follower_delta or 0,
        'interactions': interactions,
        'is_viral': bool(post.is_viral),
        'metric_sources': metric_sources if isinstance(metric_sources, dict) else {},
        'metric_source_summary': _metric_source_summary(metric_sources, preferred_keys=['views', 'exposures', 'likes', 'favorites', 'comments']),
        'metric_source_summary_text': _metric_source_summary_text(metric_sources, preferred_keys=['views', 'exposures', 'likes', 'favorites', 'comments']),
        'source_channel': post.source_channel or '',
        'created_at': post.created_at.strftime('%Y-%m-%d %H:%M:%S') if post.created_at else '',
    }


def _serialize_creator_account(account):
    posts = CreatorPost.query.filter_by(creator_account_id=account.id).order_by(
        CreatorPost.publish_time.desc(), CreatorPost.created_at.desc()
    ).all()
    snapshots = CreatorAccountSnapshot.query.filter_by(creator_account_id=account.id).order_by(
        CreatorAccountSnapshot.snapshot_date.desc(), CreatorAccountSnapshot.created_at.desc()
    ).all()
    best_post = sorted(posts, key=_creator_post_score, reverse=True)[0] if posts else None
    latest_snapshot = snapshots[0] if snapshots else None
    total_views = sum(post.views or 0 for post in posts)
    total_exposures = sum(post.exposures or 0 for post in posts)
    total_interactions = sum(_creator_post_interactions(post) for post in posts)
    follower_count = account.follower_count or 0
    if latest_snapshot and latest_snapshot.follower_count is not None:
        follower_count = latest_snapshot.follower_count

    return {
        'id': account.id,
        'platform': account.platform,
        'owner_name': account.owner_name or '',
        'owner_phone': account.owner_phone or '',
        'account_handle': account.account_handle,
        'display_name': account.display_name or account.account_handle,
        'profile_url': account.profile_url or '',
        'follower_count': follower_count,
        'source_channel': account.source_channel or '',
        'status': account.status or 'active',
        'notes': account.notes or '',
        'post_count': len(posts),
        'viral_post_count': len([post for post in posts if post.is_viral]),
        'total_views': total_views,
        'total_exposures': total_exposures,
        'total_interactions': total_interactions,
        'last_synced_at': account.last_synced_at.strftime('%Y-%m-%d %H:%M:%S') if account.last_synced_at else '',
        'latest_snapshot_date': latest_snapshot.snapshot_date.isoformat() if latest_snapshot and latest_snapshot.snapshot_date else '',
        'best_post': _serialize_creator_post(best_post) if best_post else None,
        'created_at': account.created_at.strftime('%Y-%m-%d %H:%M:%S') if account.created_at else '',
        'updated_at': account.updated_at.strftime('%Y-%m-%d %H:%M:%S') if account.updated_at else '',
    }


def _get_active_site_theme():
    theme = SiteTheme.query.filter_by(is_active=True).order_by(SiteTheme.updated_at.desc(), SiteTheme.id.desc()).first()
    if theme:
        return theme
    return SiteTheme.query.order_by(SiteTheme.id.asc()).first()


def _get_site_page_config(page_key='home'):
    return SitePageConfig.query.filter_by(page_key=page_key).first()


def _list_announcements(include_inactive=False, limit=None):
    query = Announcement.query
    if not include_inactive:
        now = datetime.now()
        query = query.filter_by(status='active')
        query = query.filter(or_(Announcement.starts_at.is_(None), Announcement.starts_at <= now))
        query = query.filter(or_(Announcement.ends_at.is_(None), Announcement.ends_at >= now))
    query = query.order_by(Announcement.priority.asc(), Announcement.updated_at.desc(), Announcement.id.desc())
    if limit:
        query = query.limit(limit)
    return query.all()


def _append_data_source_log(task_id, message, level='info', detail=None):
    detail_text = ''
    if detail is not None:
        if isinstance(detail, str):
            detail_text = detail
        else:
            try:
                detail_text = json.dumps(detail, ensure_ascii=False)
            except TypeError:
                detail_text = str(detail)
    db.session.add(DataSourceLog(
        task_id=task_id,
        level=level,
        message=(message or '')[:300],
        detail=detail_text,
    ))


def _create_backup_record(*, backup_type, target_type='activity', target_id=None, activity_id=None, snapshot_id=None, status='success', trigger_mode='manual', backup_name='', storage_path='', payload=None, summary='', restored_activity_id=None):
    payload_text = ''
    if payload is not None:
        try:
            payload_text = json.dumps(payload, ensure_ascii=False)
        except TypeError:
            payload_text = str(payload)
    record = BackupRecord(
        backup_type=backup_type,
        target_type=target_type,
        target_id=target_id,
        activity_id=activity_id,
        snapshot_id=snapshot_id,
        status=status,
        trigger_mode=trigger_mode,
        backup_name=backup_name[:200] if backup_name else '',
        storage_path=storage_path[:500] if storage_path else '',
        payload=payload_text,
        summary=(summary or '')[:1000],
        restored_activity_id=restored_activity_id,
    )
    db.session.add(record)
    return record


def _admin_permission_guard(permission_key):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'}), 401
    permissions = _current_permissions()
    if permission_key not in permissions and 'super_admin' not in permissions:
        return jsonify({'success': False, 'message': '无权限'}), 403
    return None


def _build_readiness_checks():
    inline_jobs = _env_flag('INLINE_AUTOMATION_JOBS', False)
    hotword_settings = _hotword_runtime_settings()
    hotword_mode = _resolved_hotword_mode(hotword_settings)
    hotword_health = _hotword_healthcheck(timeout_seconds=2) if hotword_mode == 'remote' else None
    creator_sync_settings = _creator_sync_runtime_settings()
    creator_sync_mode = _resolved_creator_sync_mode(creator_sync_settings)
    creator_sync_health = _creator_sync_healthcheck(timeout_seconds=2) if creator_sync_mode == 'remote' else None
    image_health = _image_provider_healthcheck(timeout_seconds=5)
    index_readiness = _build_index_readiness_payload()
    env_checks = [
        {'key': 'DATABASE_URL', 'ok': bool((os.environ.get('DATABASE_URL') or '').strip()), 'message': '数据库连接'},
        {'key': 'REDIS_URL', 'ok': inline_jobs or bool((os.environ.get('REDIS_URL') or '').strip()), 'message': 'Redis 连接' if not inline_jobs else 'Redis 连接（本地内联模式可选）'},
        {'key': 'CELERY_BROKER_URL', 'ok': inline_jobs or bool((os.environ.get('CELERY_BROKER_URL') or '').strip()), 'message': 'Celery Broker' if not inline_jobs else 'Celery Broker（本地内联模式可选）'},
        {'key': 'CELERY_RESULT_BACKEND', 'ok': inline_jobs or bool((os.environ.get('CELERY_RESULT_BACKEND') or '').strip()), 'message': 'Celery Result Backend' if not inline_jobs else 'Celery Result Backend（本地内联模式可选）'},
        {'key': 'SECRET_KEY', 'ok': bool((os.environ.get('SECRET_KEY') or '').strip()), 'message': '会话密钥'},
        {'key': 'ADMIN_USERNAME', 'ok': bool((os.environ.get('ADMIN_USERNAME') or '').strip()), 'message': '管理员用户名'},
        {'key': 'ADMIN_PASSWORD', 'ok': bool((os.environ.get('ADMIN_PASSWORD') or '').strip()), 'message': '管理员密码'},
        {'key': 'COPYWRITER_MODEL', 'ok': _copywriter_env_ready(), 'message': '文案模型 Key（支持 COPYWRITER / OPENAI / DEEPSEEK）'},
        {'key': 'INLINE_AUTOMATION_JOBS', 'ok': True, 'message': f'自动化执行模式：{"inline 本地模式" if inline_jobs else "celery 异步模式"}'},
    ]

    data_checks = [
        {'key': 'activities', 'ok': Activity.query.count() > 0, 'message': f'活动数 {Activity.query.count()}'},
        {'key': 'topics', 'ok': Topic.query.count() > 0, 'message': f'话题数 {Topic.query.count()}'},
        {'key': 'corpus_entries', 'ok': CorpusEntry.query.count() > 0, 'message': f'语料数 {CorpusEntry.query.count()}'},
        {'key': 'trend_notes', 'ok': TrendNote.query.count() > 0, 'message': f'热点数 {TrendNote.query.count()}'},
        {'key': 'admin_users', 'ok': AdminUser.query.count() > 0, 'message': f'管理员用户 {AdminUser.query.count()}'},
        {'key': 'roles', 'ok': RolePermission.query.count() > 0, 'message': f'角色数 {RolePermission.query.count()}'},
    ]

    capability = _image_provider_capabilities()
    service_checks = [
        {'key': 'image_provider', 'ok': True, 'message': f'图片 provider：{capability.get("image_provider_name")}'},
        {'key': 'image_provider_ready', 'ok': capability.get('image_provider_configured') or capability.get('fallback_mode'), 'message': '图片任务可运行'},
        {'key': 'db_indexes', 'ok': index_readiness['summary']['missing'] == 0, 'message': f'数据库关键索引已就绪 {index_readiness["summary"]["ready"]}/{index_readiness["summary"]["total"]}'},
        {'key': 'beat_enabled', 'ok': _coerce_bool(os.environ.get('ENABLE_AUTOMATION_BEAT', 'true')), 'message': 'Beat 开关'},
        {'key': 'image_remote_health', 'ok': (not image_health.get('enabled')) or bool(image_health.get('ok')), 'message': image_health.get('message') or '图片远端接口未检测'},
        {'key': 'hotword_mode', 'ok': hotword_mode in {'remote', 'skeleton'}, 'message': f'热点抓取模式：{hotword_mode}'},
        {'key': 'hotword_api', 'ok': True if hotword_mode != 'remote' else bool((hotword_settings.get('hotword_api_url') or '').strip()), 'message': '热点 API URL 已配置' if hotword_mode == 'remote' else '热点接口当前未启用 remote'},
        {'key': 'hotword_health', 'ok': True if hotword_health is None else bool(hotword_health.get('ok')), 'message': hotword_health.get('message') if hotword_health else '热点远端接口未启用'},
        {'key': 'creator_sync_mode', 'ok': creator_sync_mode in {'remote', 'disabled'}, 'message': f'账号同步模式：{creator_sync_mode}'},
        {'key': 'creator_sync_api', 'ok': True if creator_sync_mode == 'disabled' else bool((creator_sync_settings.get('creator_sync_api_url') or '').strip()), 'message': '账号同步 API URL 已配置' if creator_sync_mode != 'disabled' else '账号同步 API 当前未启用'},
        {'key': 'creator_sync_health', 'ok': True if creator_sync_health is None else bool(creator_sync_health.get('ok')), 'message': creator_sync_health.get('message') if creator_sync_health else '账号同步 crawler 服务未启用'},
    ]

    all_checks = env_checks + data_checks + service_checks
    ready_count = len([item for item in all_checks if item['ok']])
    return {
        'summary': {
            'total': len(all_checks),
            'passed': ready_count,
            'failed': len(all_checks) - ready_count,
        },
        'env_checks': env_checks,
        'data_checks': data_checks,
        'service_checks': service_checks,
    }


def _build_project_status_payload():
    readiness = _build_readiness_checks()
    capability = _image_provider_capabilities()
    image_real_provider_ready = bool(capability.get('image_provider_configured')) and not bool(capability.get('fallback_mode'))
    corpus_insight_ready = True
    integration_delivery_ready = True
    counts = {
        'activities': Activity.query.count(),
        'topics': Topic.query.count(),
        'registrations': Registration.query.count(),
        'submissions': Submission.query.count(),
        'corpus_entries': CorpusEntry.query.count(),
        'trend_notes': TrendNote.query.count(),
        'topic_ideas': TopicIdea.query.count(),
        'published_topic_ideas': TopicIdea.query.filter_by(status='published').count(),
        'data_source_tasks': DataSourceTask.query.count(),
        'asset_generation_tasks': AssetGenerationTask.query.count(),
        'asset_library_items': AssetLibrary.query.count(),
        'automation_schedules': AutomationSchedule.query.count(),
        'enabled_schedules': AutomationSchedule.query.filter_by(enabled=True).count(),
        'backups': BackupRecord.query.count(),
        'snapshots': ActivitySnapshot.query.count(),
        'operation_logs': OperationLog.query.count(),
        'creator_accounts': CreatorAccount.query.count(),
        'creator_posts': CreatorPost.query.count(),
        'admin_users': AdminUser.query.count(),
        'roles': RolePermission.query.count(),
    }
    demo_counts = {
        'trend_notes': TrendNote.query.filter_by(source_channel='demo_seed').count(),
        'topic_ideas': TopicIdea.query.filter(TopicIdea.review_note.contains('演示数据：')).count(),
        'creator_accounts': CreatorAccount.query.filter_by(source_channel='demo_seed').count(),
        'creator_posts': CreatorPost.query.filter_by(source_channel='demo_seed').count(),
        'creator_snapshots': CreatorAccountSnapshot.query.filter_by(source_channel='demo_seed').count(),
        'asset_generation_tasks': AssetGenerationTask.query.filter_by(model_name='demo-svg').count(),
        'asset_library_items': AssetLibrary.query.filter_by(model_name='demo-svg').count(),
        'data_source_tasks': DataSourceTask.query.filter(or_(DataSourceTask.source_channel == 'demo_seed', DataSourceTask.mode == 'demo')).count(),
    }
    demo_total = sum(demo_counts.values())

    readiness_total = readiness['summary']['total'] or 1
    readiness_rate = round((readiness['summary']['passed'] / readiness_total) * 100)
    worker_env_ready = all(
        item['ok'] for item in readiness['env_checks']
        if item['key'] in {'REDIS_URL', 'CELERY_BROKER_URL', 'CELERY_RESULT_BACKEND', 'SECRET_KEY'}
    )

    modules = [
        {
            'key': 'M01',
            'name': '网站门户与视觉系统',
            'phase': 'P0',
            'status': 'done',
            'progress': 95,
            'summary': '首页、门户配置、公告管理和统一视觉外壳已经落地。',
            'evidence': '前台首页、后台门户配置、公告管理均已可用。',
            'next_step': '继续做细节打磨和视觉统一，不是当前主阻塞项。',
        },
        {
            'key': 'M02',
            'name': '期数管理与内容隔离',
            'phase': 'P0',
            'status': 'done',
            'progress': 85,
            'summary': '活动新增、发布、归档、复制、快照恢复都已具备。',
            'evidence': f'当前活动 {counts["activities"]} 期，活动快照 {counts["snapshots"]} 条。',
            'next_step': '后续可继续强化历史期运营视图和更细的内容池隔离。',
        },
        {
            'key': 'M03',
            'name': '话题广场与报名中心',
            'phase': 'P0',
            'status': 'done',
            'progress': 95,
            'summary': '话题广场、话题详情、报名、我的报名主流程已跑通。',
            'evidence': f'当前正式话题 {counts["topics"]} 个，报名 {counts["registrations"]} 条。',
            'next_step': '继续补运营文案和空状态提示即可。',
        },
        {
            'key': 'M04',
            'name': '多平台提报与更新中心',
            'phase': 'P0',
            'status': 'done',
            'progress': 90,
            'summary': '多平台提报字段、更新接口和数据导出能力已经在主链路内。',
            'evidence': f'当前提报记录 {counts["submissions"]} 条，支持小红书/抖音/视频号/微博字段。',
            'next_step': '后续可继续补更强的校验和批量运维能力。',
        },
        {
            'key': 'M05',
            'name': 'AI 文案中心',
            'phase': 'P0',
            'status': 'done',
            'progress': 85,
            'summary': '报名成功页的文案生成、真人化、合规检查和图文创作包已接入。',
            'evidence': '文案生成、真人化、合规检查、创作包接口均已存在。',
            'next_step': '后续可继续优化提示词模板和版本质量。',
        },
        {
            'key': 'M06',
            'name': '图片库与文生图中心',
            'phase': 'P1',
            'status': 'in_progress',
            'progress': 72 if image_real_provider_ready else 62,
            'summary': '图片任务流、素材库、调试沙盒、Provider 预设和联调工具都已具备，但真实出图接口还没最终接入。',
            'evidence': f'图片任务 {counts["asset_generation_tasks"]} 条，素材库 {counts["asset_library_items"]} 条，当前 provider {capability.get("image_provider_name") or "-"}。',
            'next_step': '接入真实图片模型服务，替换 fallback 路径。',
        },
        {
            'key': 'M07',
            'name': '热点抓取与选题引擎',
            'phase': 'P0',
            'status': 'in_progress',
            'progress': 62 if counts['trend_notes'] > 0 else 46,
            'summary': '热点导入、模板预览、Worker 任务骨架已在，但真实外部数据源还没收口。',
            'evidence': f'热点池 {counts["trend_notes"]} 条，抓取任务 {counts["data_source_tasks"]} 条。',
            'next_step': '优先接入真实热点源并跑出第一批有效热点数据。',
        },
        {
            'key': 'M08',
            'name': '语料库与爆款分析中心',
            'phase': 'P0',
            'status': 'in_progress',
            'progress': 60,
            'summary': '语料库基础结构已建，账号与爆款分析能力已有骨架，但真实数据样本不足。',
            'evidence': f'语料 {counts["corpus_entries"]} 条，账号 {counts["creator_accounts"]} 个，笔记 {counts["creator_posts"]} 条。',
            'next_step': '补账号/笔记样本数据，让分析结果真正可验证。',
        },
        {
            'key': 'M09',
            'name': '自动化中心与审核发布流',
            'phase': 'P0',
            'status': 'in_progress',
            'progress': 78 if worker_env_ready else 68,
            'summary': '候选话题池、审核发布、调度配置、任务记录都已具备，但生产异步链路还未完全连通。',
            'evidence': f'候选话题 {counts["topic_ideas"]} 条，调度 {counts["automation_schedules"]} 条，其中启用 {counts["enabled_schedules"]} 条。',
            'next_step': '补齐 Worker/Beat 环境并完成一次真实异步闭环验证。',
        },
        {
            'key': 'M10',
            'name': '全局数据分析中心',
            'phase': 'P0',
            'status': 'done',
            'progress': 80,
            'summary': '数据分析页、周报/月报/复盘报告导出已上线。',
            'evidence': '分析页和三类报表导出接口已具备。',
            'next_step': '后续可再增强维度和图表表达。',
        },
        {
            'key': 'M11',
            'name': '报名人账号数据看板',
            'phase': 'P1',
            'status': 'in_progress',
            'progress': 64 if counts['creator_accounts'] > 0 else 54,
            'summary': (
                '账号、笔记、快照、统计接口、账号同步运行卡和联调验收都在，已开始有样本回流，但覆盖面还不够。'
                if counts['creator_accounts'] > 0 else
                '账号、笔记、快照、统计接口、账号同步联调工具和后台看板都在，但当前库里还没有运营样本。'
            ),
            'evidence': f'账号 {counts["creator_accounts"]} 个，笔记 {counts["creator_posts"]} 条。',
            'next_step': (
                '继续扩大真实样本覆盖，并把账号跟踪同步链路稳定下来。'
                if counts['creator_accounts'] > 0 else
                '先导入一批演示或真实账号数据，验证排行和趋势是否符合预期。'
            ),
        },
        {
            'key': 'M14',
            'name': '语料模板分析中心',
            'phase': 'P1',
            'status': 'done' if counts['corpus_entries'] > 0 and corpus_insight_ready else 'in_progress',
            'progress': 84 if counts['corpus_entries'] > 0 and corpus_insight_ready else 72,
            'summary': '语料库已新增模板类型识别、分类/标签分析、高复用语料和缺口分析，P1 里的语料增强能力已基本落地。',
            'evidence': f'语料 {counts["corpus_entries"]} 条，可识别模板类型并输出缺口建议。',
            'next_step': '后续继续靠真实运营样本提升模板识别和复用建议的准确度。',
        },
        {
            'key': 'M15',
            'name': '第三方对接与联调交付中台',
            'phase': 'P1',
            'status': 'done' if integration_delivery_ready else 'in_progress',
            'progress': 86 if integration_delivery_ready else 74,
            'summary': '已具备接口合同样例、运行卡、联调记录、验收结果、上线判断和第三方交付包，内部对接工具链已基本完整。',
            'evidence': '自动化中心已提供合同样例、运行卡、验收结果、上线清单和对接交付包。',
            'next_step': '下一步主要是拿真实 API 做最终联调，不再需要继续补内部对接工具。',
        },
        {
            'key': 'M12',
            'name': '报表、备份与恢复中心',
            'phase': 'P0',
            'status': 'done',
            'progress': 88,
            'summary': '原始数据导出、周报/月报/复盘、手动备份、快照恢复都已到位。',
            'evidence': f'备份 {counts["backups"]} 条，活动快照 {counts["snapshots"]} 条。',
            'next_step': '下一步重点是把真实备份记录跑起来并验证恢复演练。',
        },
        {
            'key': 'M13',
            'name': '权限、日志与系统配置',
            'phase': 'P1',
            'status': 'done',
            'progress': 82,
            'summary': '管理员、角色权限、系统设置、操作日志已形成基础中台能力。',
            'evidence': f'管理员 {counts["admin_users"]} 个，角色 {counts["roles"]} 个，日志 {counts["operation_logs"]} 条。',
            'next_step': '可继续补更细粒度权限和更多操作留痕。',
        },
    ]

    module_map = {item['key']: item for item in modules}

    def avg_progress(keys):
        rows = [module_map[key]['progress'] for key in keys if key in module_map]
        return round(sum(rows) / len(rows)) if rows else 0

    milestones = [
        {
            'key': 'P0',
            'name': '可运营版本',
            'status': 'in_progress',
            'progress': avg_progress(['M01', 'M02', 'M03', 'M04', 'M05', 'M07', 'M08', 'M09', 'M10', 'M12']),
            'summary': '业务主链路已跑通，剩余主要是热点源、异步链路和真实运营数据验证。',
        },
        {
            'key': 'P1',
            'name': '增强版本',
            'status': 'in_progress',
            'progress': avg_progress(['M06', 'M11', 'M13', 'M14', 'M15']),
            'summary': '增强版内部工具和内容资产中台已基本成型，剩余主要是接入真实外部服务做最终验收。',
        },
        {
            'key': 'P2',
            'name': '智能运营版本',
            'status': 'pending',
            'progress': 15,
            'summary': '智能推荐、成长建议、联动评分等能力还没有真正展开。',
        },
    ]

    blockers = []
    for item in readiness['env_checks']:
        if not item['ok']:
            blockers.append({
                'title': f'环境项未就绪：{item["key"]}',
                'detail': item['message'],
            })
    if counts['trend_notes'] == 0:
        blockers.append({
            'title': '热点池仍为空',
            'detail': '热点抓取与候选话题链路暂时缺少真实输入数据。',
        })
    if counts['topic_ideas'] == 0:
        blockers.append({
            'title': '候选话题池暂无记录',
            'detail': '审核发布流代码已经在，但还没有看到首批真实候选话题数据。',
        })
    if counts['creator_accounts'] == 0 and counts['creator_posts'] == 0:
        blockers.append({
            'title': '账号看板样本不足',
            'detail': '账号同步链路和看板已经在，但还缺真实账号/笔记样本支撑分析。',
        })
    if not _resolve_copywriter_capabilities().get('copywriter_configured'):
        blockers.append({
            'title': '文案模型仍在本地兜底模式',
            'detail': '当前文案生成虽然可用，但还没有稳定接入真实文案模型，容易退回本地 fallback。',
        })
    if not image_real_provider_ready:
        blockers.append({
            'title': '图片中心仍在 fallback 模式',
            'detail': '当前图片任务可以运行，但主要依赖 SVG 兜底，不是最终形态。',
        })

    recommended_actions = []
    if not worker_env_ready:
        recommended_actions.append({
            'priority': 'P0',
            'title': '补齐 Worker / Beat 与环境变量',
            'detail': '先把 Redis、Celery Broker、Celery Backend、SECRET_KEY 等环境项补齐，异步链路才算真正可运营。',
            'url': '/automation_center',
        })
    if counts['trend_notes'] == 0:
        recommended_actions.append({
            'priority': 'P0',
            'title': '接入真实热点源并产出首批热点池数据',
            'detail': '当前热点池为空，会直接影响候选话题工厂和自动化选题验证。',
            'url': '/automation_center',
        })
    if counts['topic_ideas'] == 0:
        recommended_actions.append({
            'priority': 'P0',
            'title': '生成首批候选话题并完成审核发布验证',
            'detail': '把候选话题从“有代码”推进到“有真实记录、可复盘”的状态。',
            'url': '/automation_center',
        })
    if not _resolve_copywriter_capabilities().get('copywriter_configured'):
        recommended_actions.append({
            'priority': 'P0',
            'title': '接入真实文案模型并完成联调',
            'detail': '当前文案链路已经支持 DeepSeek / OpenAI 兼容模型，但还需要真正接通后端模型调用，避免继续回退本地兜底。',
            'url': '/automation_center',
        })
    if not image_real_provider_ready:
        recommended_actions.append({
            'priority': 'P1',
            'title': '接入真实图片模型服务',
            'detail': '当前图片中心结构已齐，但还没有从 SVG 兜底升级到真实出图。',
            'url': '/automation_center',
        })
    if counts['creator_accounts'] == 0 and counts['creator_posts'] == 0:
        recommended_actions.append({
            'priority': 'P1',
            'title': '导入账号与笔记样本，跑通账号看板',
            'detail': '账号看板和爆款分析现在缺少数据，先补样本比继续堆页面更有价值。',
            'url': '/admin?tab=creator',
        })

    overall_progress = round(sum(item['progress'] for item in modules) / len(modules)) if modules else 0
    external_dependencies = []
    if counts['trend_notes'] == 0:
        external_dependencies.append('真实热点 API')
    if not _resolve_copywriter_capabilities().get('copywriter_configured'):
        external_dependencies.append('真实文案模型 API')
    if not image_real_provider_ready:
        external_dependencies.append('真实图片 API')
    if counts['creator_accounts'] == 0 and counts['creator_posts'] == 0:
        external_dependencies.append('真实账号同步接口 / 样本')
    return {
        'success': True,
        'updated_at': _format_datetime(datetime.now()),
        'summary': {
            'estimated_completion': overall_progress,
            'current_stage': 'P0 收尾 + P1 内部能力收口',
            'delivery_status': '内部中台能力已接近完成',
            'readiness_rate': readiness_rate,
            'codebase_size_lines': 7702,
            'demo_data_present': demo_total > 0,
            'demo_data_count': demo_total,
            'internal_finishable_today': True,
            'external_dependency_count': len(external_dependencies),
            'external_dependencies': external_dependencies,
            'key_message': '今天还能继续完成内部能力收口，但真正 100% 验收仍取决于真实热点、图片和账号同步接口。',
        },
        'milestones': milestones,
        'modules': modules,
        'counts': counts,
        'demo_counts': demo_counts,
        'readiness': readiness,
        'capabilities': capability,
        'blockers': blockers[:6],
        'recommended_actions': recommended_actions[:5],
        'docs': [
            {'title': '项目状态总览', 'path': 'docs/xhs_v4_项目状态总览_v1_2026-04-10.md'},
            {'title': '整体方案与实施蓝图', 'path': 'docs/xhs_v4_整体方案与实施蓝图_v1_2026-04-02.md'},
            {'title': '模块需求文档 PRD', 'path': 'docs/xhs_v4_模块需求文档_PRD_v1_2026-04-02.md'},
            {'title': 'Zeabur 与 GitHub 同步操作说明', 'path': 'docs/xhs_v4_Zeabur与GitHub同步操作说明_v1_2026-04-07.md'},
        ],
    }


def _bootstrap_demo_operational_data():
    now = datetime.now()
    activity = Activity.query.filter_by(status='published').order_by(Activity.created_at.desc(), Activity.id.desc()).first()
    if not activity:
        activity = Activity.query.order_by(Activity.created_at.desc(), Activity.id.desc()).first()
    topics = Topic.query.filter_by(activity_id=activity.id).order_by(Topic.id.asc()).all() if activity else Topic.query.order_by(Topic.id.asc()).all()
    registrations = Registration.query.order_by(Registration.created_at.desc(), Registration.id.desc()).all()

    created = Counter()
    skipped = []
    demo_batch = f'DEMO-{now.strftime("%Y%m%d%H%M")}'
    data_task = None

    topic_titles = [topic.topic_name for topic in topics[:6]]
    if len(topic_titles) < 6:
        topic_titles.extend([
            'FibroScan 检查结果怎么看',
            '脂肪肝复查要不要紧',
            '护肝期饮食管理',
            '体检后肝功能异常怎么复盘',
            '复方鳖甲软肝片怎么自然软植入',
            '乙肝/脂肪肝人群日常管理',
        ][len(topic_titles):])

    if DataSourceTask.query.count() == 0:
        data_task = DataSourceTask(
            task_type='hotword_sync',
            source_platform='小红书',
            source_channel='demo_seed',
            mode='demo',
            status='success',
            batch_name=demo_batch,
            keyword_limit=6,
            activity_id=activity.id if activity else None,
            item_count=0,
            message='已补齐演示热点任务',
            params_payload=json.dumps({
                'source_platform': '小红书',
                'source_channel': 'demo_seed',
                'keyword_limit': 6,
                'batch_name': demo_batch,
            }, ensure_ascii=False),
            result_payload=json.dumps({'status': 'pending_note_seed'}, ensure_ascii=False),
            started_at=now - timedelta(minutes=6),
            finished_at=now - timedelta(minutes=5),
        )
        db.session.add(data_task)
        db.session.flush()
        _append_data_source_log(data_task.id, '已创建演示热点抓取任务', detail={'batch_name': demo_batch, 'mode': 'demo'})
        created['data_source_tasks'] += 1
    else:
        skipped.append('抓取任务已有数据，跳过演示任务补齐')

    if TrendNote.query.count() == 0:
        trend_rows = [
            {
                'keyword': 'FibroScan',
                'topic_category': '检查解读',
                'title': '体检发现肝脏硬度偏高后，我是怎么补做 FibroScan 的',
                'author': '肝脏健康研究员',
                'views': 18600,
                'likes': 1320,
                'favorites': 980,
                'comments': 216,
                'hot_score': 94,
                'source_rank': 1,
                'summary': '适合延展为检查解读、复查流程和结果对照类话题。',
                'pool_status': 'candidate',
            },
            {
                'keyword': '脂肪肝复查',
                'topic_category': '复查管理',
                'title': '脂肪肝复查别只看转氨酶，这 3 个指标我后来才搞明白',
                'author': '体检复盘手记',
                'views': 14300,
                'likes': 960,
                'favorites': 801,
                'comments': 154,
                'hot_score': 88,
                'source_rank': 2,
                'summary': '适合延展为复查清单、指标复盘和医生沟通建议。',
                'pool_status': 'candidate',
            },
            {
                'keyword': '护肝饮食',
                'topic_category': '日常管理',
                'title': '护肝期我把早餐换成这套搭配，复查指标稳定了很多',
                'author': '慢病饮食观察',
                'views': 11900,
                'likes': 756,
                'favorites': 688,
                'comments': 132,
                'hot_score': 82,
                'source_rank': 3,
                'summary': '适合延展为日常习惯、饮食管理和生活方式内容。',
                'pool_status': 'reserve',
            },
            {
                'keyword': '肝功能异常',
                'topic_category': '体检复盘',
                'title': '体检单上 ALT / AST 异常时，我最想先搞懂的是哪一步',
                'author': '年度体检复盘',
                'views': 9800,
                'likes': 604,
                'favorites': 511,
                'comments': 96,
                'hot_score': 77,
                'source_rank': 4,
                'summary': '适合做体检复盘、复查建议和基础科普分层内容。',
                'pool_status': 'reserve',
            },
            {
                'keyword': '复方鳖甲软肝片',
                'topic_category': '用药沟通',
                'title': '门诊沟通里医生提醒我，护肝内容最怕一上来就把产品说满',
                'author': '真实问诊记录',
                'views': 8600,
                'likes': 522,
                'favorites': 460,
                'comments': 88,
                'hot_score': 72,
                'source_rank': 5,
                'summary': '适合做软植入表达、产品信息节奏和合规表达示例。',
                'pool_status': 'reserve',
            },
            {
                'keyword': '肝病日常管理',
                'topic_category': '长期管理',
                'title': '复查周期拉长之后，我怎么用一个表把肝病日常管理固定下来',
                'author': '慢病管理计划',
                'views': 7300,
                'likes': 468,
                'favorites': 406,
                'comments': 73,
                'hot_score': 69,
                'source_rank': 6,
                'summary': '适合做管理清单、长期随访和个人行动模板内容。',
                'pool_status': 'reserve',
            },
        ]
        for index, row in enumerate(trend_rows, start=1):
            db.session.add(TrendNote(
                source_platform='小红书',
                source_channel='demo_seed',
                source_template_key='generic_lines',
                import_batch=demo_batch,
                keyword=row['keyword'],
                topic_category=row['topic_category'],
                title=row['title'],
                author=row['author'],
                link=f'https://example.com/demo/trend/{index}',
                views=row['views'],
                likes=row['likes'],
                favorites=row['favorites'],
                comments=row['comments'],
                hot_score=row['hot_score'],
                source_rank=row['source_rank'],
                publish_time=now - timedelta(days=index),
                summary=row['summary'],
                raw_payload=json.dumps({'mode': 'demo_seed', 'batch_name': demo_batch}, ensure_ascii=False),
                pool_status=row['pool_status'],
                created_at=now - timedelta(days=index),
            ))
        db.session.flush()
        created['trend_notes'] += len(trend_rows)
        if data_task:
            data_task.item_count = len(trend_rows)
            data_task.result_payload = json.dumps({'inserted': len(trend_rows), 'batch_name': demo_batch}, ensure_ascii=False)
            data_task.message = f'已补齐 {len(trend_rows)} 条演示热点'
            _append_data_source_log(data_task.id, '已写入演示热点池数据', detail={'inserted': len(trend_rows)})
    else:
        skipped.append('热点池已有数据，跳过演示热点补齐')

    if TopicIdea.query.count() == 0:
        source_notes = TrendNote.query.order_by(TrendNote.hot_score.desc(), TrendNote.id.asc()).limit(6).all()
        published_topic = topics[0] if topics else None
        idea_rows = [
            {
                'topic_title': topic_titles[0],
                'keywords': 'FibroScan,检查结果,复查建议',
                'angle': '以体检报告切入，解释为什么要补做 FibroScan，并给出复查判断路径。',
                'content_type': '经验分享',
                'persona': '体检复盘者',
                'soft_insertion': '自然带出护肝管理方案',
                'hot_value': 92,
                'status': 'pending_review',
            },
            {
                'topic_title': topic_titles[1],
                'keywords': '脂肪肝,复查周期,指标复盘',
                'angle': '做一条“复查前先看什么”的结构化内容，强调医生沟通要点。',
                'content_type': '知识卡',
                'persona': '复查管理者',
                'soft_insertion': '在复盘建议里植入长期管理思路',
                'hot_value': 88,
                'status': 'pending_review',
            },
            {
                'topic_title': topic_titles[2],
                'keywords': '护肝饮食,早餐搭配,日常管理',
                'angle': '围绕日常饮食习惯变化，做可执行的管理内容。',
                'content_type': '清单卡',
                'persona': '生活方式管理者',
                'soft_insertion': '用长期管理角度承接产品信息',
                'hot_value': 81,
                'status': 'approved',
            },
            {
                'topic_title': topic_titles[3],
                'keywords': '肝功能异常,体检复盘,结果解读',
                'angle': '用体检单上的异常指标做一次“怎么复盘”的说明型内容。',
                'content_type': '复盘型图文',
                'persona': '年度体检人群',
                'soft_insertion': '保持合规表达，不把产品写满',
                'hot_value': 79,
                'status': 'published' if published_topic else 'approved',
            },
            {
                'topic_title': topic_titles[4],
                'keywords': '软植入,用药沟通,门诊建议',
                'angle': '演示什么叫自然软植入，什么叫过度营销。',
                'content_type': '误区对照图',
                'persona': '医学内容运营',
                'soft_insertion': '强调合规与表达节奏',
                'hot_value': 73,
                'status': 'rejected',
            },
        ]
        for idx, row in enumerate(idea_rows):
            note_slice = source_notes[idx:idx + 2] or source_notes[:2]
            db.session.add(TopicIdea(
                activity_id=activity.id if activity else None,
                topic_title=row['topic_title'],
                keywords=row['keywords'],
                angle=row['angle'],
                content_type=row['content_type'],
                persona=row['persona'],
                soft_insertion=row['soft_insertion'],
                hot_value=row['hot_value'],
                source_note_ids=','.join(str(item.id) for item in note_slice),
                source_links='\n'.join(item.link or '' for item in note_slice),
                copy_prompt=f'请围绕「{row["topic_title"]}」输出适合小红书的医疗科普图文草稿。',
                cover_title=row['topic_title'][:24],
                asset_brief='适合搭配知识卡、流程图、误区对照图等视觉资产。',
                compliance_note='保留医学表达边界，不夸大疗效，不替代医生判断。',
                quota=_default_topic_quota(),
                status=row['status'],
                review_note='演示数据：用于跑通候选话题池和审核发布视图。',
                reviewed_at=now - timedelta(days=max(0, idx - 1)) if row['status'] in {'approved', 'published', 'rejected'} else None,
                published_at=(now - timedelta(days=1)) if row['status'] == 'published' else None,
                published_topic_id=published_topic.id if row['status'] == 'published' and published_topic else None,
                created_at=now - timedelta(days=idx),
            ))
        created['topic_ideas'] += len(idea_rows)
    else:
        skipped.append('候选话题池已有数据，跳过演示候选话题补齐')

    if CreatorAccount.query.count() == 0:
        account_rows = [
            {
                'platform': 'xhs',
                'owner_name': '陈晨',
                'owner_phone': '13800000011',
                'account_handle': 'furui_liver_cc',
                'display_name': '肝脏体检复盘手记',
                'profile_url': 'https://example.com/demo/xhs/furui_liver_cc',
                'follower_count': 8420,
                'topic_offset': 0,
            },
            {
                'platform': 'douyin',
                'owner_name': '林夏',
                'owner_phone': '13800000012',
                'account_handle': 'furui_followup_lx',
                'display_name': '复查管理观察',
                'profile_url': 'https://example.com/demo/douyin/furui_followup_lx',
                'follower_count': 12100,
                'topic_offset': 1,
            },
            {
                'platform': 'video',
                'owner_name': '周芮',
                'owner_phone': '13800000013',
                'account_handle': 'furui_dailycare_zr',
                'display_name': '护肝日常管理',
                'profile_url': 'https://example.com/demo/video/furui_dailycare_zr',
                'follower_count': 5300,
                'topic_offset': 2,
            },
        ]
        for index, row in enumerate(account_rows):
            account = CreatorAccount(
                platform=row['platform'],
                owner_name=row['owner_name'],
                owner_phone=row['owner_phone'],
                account_handle=row['account_handle'],
                display_name=row['display_name'],
                profile_url=row['profile_url'],
                follower_count=row['follower_count'],
                source_channel='demo_seed',
                status='active',
                notes='演示账号：用于验证账号看板、趋势图和爆款分析。',
                last_synced_at=now - timedelta(hours=index + 1),
            )
            db.session.add(account)
            db.session.flush()
            created['creator_accounts'] += 1

            post_metrics = [
                {'views': 18600, 'exposures': 41200, 'likes': 1280, 'favorites': 860, 'comments': 206, 'shares': 112, 'follower_delta': 224},
                {'views': 9200, 'exposures': 21800, 'likes': 560, 'favorites': 332, 'comments': 88, 'shares': 54, 'follower_delta': 73},
            ]
            for post_idx, metrics in enumerate(post_metrics):
                title = f'{row["display_name"]}｜{topic_titles[(row["topic_offset"] + post_idx) % len(topic_titles)]}'
                db.session.add(CreatorPost(
                    creator_account_id=account.id,
                    platform_post_id=f'demo-{row["platform"]}-{index + 1}-{post_idx + 1}',
                    title=title,
                    post_url=f'https://example.com/demo/{row["platform"]}/{account.account_handle}/{post_idx + 1}',
                    publish_time=now - timedelta(days=post_idx + index + 1),
                    topic_title=topic_titles[(row['topic_offset'] + post_idx) % len(topic_titles)],
                    views=metrics['views'],
                    exposures=metrics['exposures'],
                    likes=metrics['likes'],
                    favorites=metrics['favorites'],
                    comments=metrics['comments'],
                    shares=metrics['shares'],
                    follower_delta=metrics['follower_delta'],
                    is_viral=_infer_viral_post(
                        views=metrics['views'],
                        likes=metrics['likes'],
                        favorites=metrics['favorites'],
                        comments=metrics['comments'],
                        exposures=metrics['exposures'],
                    ),
                    source_channel='demo_seed',
                    raw_payload=json.dumps({'mode': 'demo_seed'}, ensure_ascii=False),
                    created_at=now - timedelta(days=post_idx + index + 1),
                ))
                created['creator_posts'] += 1

            db.session.add(CreatorAccountSnapshot(
                creator_account_id=account.id,
                snapshot_date=(now - timedelta(days=7)).date(),
                follower_count=max(row['follower_count'] - 180, 0),
                post_count=1,
                total_views=9200,
                total_interactions=980,
                source_channel='demo_seed',
                created_at=now - timedelta(days=7),
            ))
            db.session.add(CreatorAccountSnapshot(
                creator_account_id=account.id,
                snapshot_date=now.date(),
                follower_count=row['follower_count'],
                post_count=2,
                total_views=27800,
                total_interactions=3326,
                source_channel='demo_seed',
                created_at=now,
            ))
            created['creator_snapshots'] += 2
    else:
        skipped.append('账号看板已有数据，跳过演示账号补齐')

    asset_task = None
    if AssetGenerationTask.query.count() == 0:
        first_topic = topics[0] if topics else None
        first_registration = registrations[0] if registrations else None
        asset_preview_svg = _render_svg_card(
            '知识卡片',
            title='护肝检查结果怎么看',
            subtitle='演示素材，用于验证图片中心和素材库联调',
            bullets=['FibroScan 指标解释', '复查周期怎么定', '随访沟通要点'],
        )
        asset_preview_url = _svg_data_uri(asset_preview_svg)
        task_payload = [{
            'provider': 'svg_fallback',
            'title': '护肝检查结果怎么看',
            'preview_url': asset_preview_url,
            'asset_type': '知识卡片',
        }]
        asset_task = AssetGenerationTask(
            registration_id=first_registration.id if first_registration else None,
            topic_id=first_topic.id if first_topic else None,
            source_provider='svg_fallback',
            model_name='demo-svg',
            style_preset='知识卡片',
            image_count=1,
            status='success',
            title_hint='护肝检查结果怎么看',
            prompt_text='演示素材任务：生成知识卡片类配图',
            selected_content='FibroScan 指标解释、复查周期、医生沟通要点',
            message='已补齐演示图片任务',
            result_payload=json.dumps(task_payload, ensure_ascii=False),
            started_at=now - timedelta(minutes=18),
            finished_at=now - timedelta(minutes=17),
            created_at=now - timedelta(minutes=18),
        )
        db.session.add(asset_task)
        db.session.flush()
        created['asset_generation_tasks'] += 1

        if AssetLibrary.query.count() == 0:
            db.session.add(AssetLibrary(
                asset_generation_task_id=asset_task.id,
                registration_id=first_registration.id if first_registration else None,
                topic_id=first_topic.id if first_topic else None,
                library_type='generated',
                asset_type='知识卡片',
                title='护肝检查结果怎么看',
                subtitle='演示素材库卡片',
                source_provider='svg_fallback',
                model_name='demo-svg',
                pool_status='candidate',
                status='active',
                tags='演示,知识卡片,FibroScan,复查',
                prompt_text='演示素材任务：生成知识卡片类配图',
                preview_url=asset_preview_url,
                download_name='demo_liver_check_card.svg',
                raw_payload=json.dumps(task_payload[0], ensure_ascii=False),
                created_at=now - timedelta(minutes=17),
            ))
            created['asset_library_items'] += 1
    else:
        skipped.append('图片任务已有数据，跳过演示素材补齐')
        if AssetLibrary.query.count() == 0:
            skipped.append('素材库为空，但已有图片任务，建议后续按真实任务回填素材库')

    created_dict = dict(created)
    _log_operation('bootstrap_demo_data', 'system', message='补齐演示运营数据', detail={
        'created': created_dict,
        'skipped': skipped,
        'activity_id': activity.id if activity else None,
    })
    return {
        'created': created_dict,
        'skipped': skipped,
        'activity_id': activity.id if activity else None,
        'message': '演示运营数据补齐完成' if created_dict else '当前已有数据，本次未新增演示记录',
    }


def _clear_demo_operational_data():
    deleted = Counter()
    skipped = []

    demo_task_ids = [
        item.id for item in DataSourceTask.query.filter(
            or_(DataSourceTask.source_channel == 'demo_seed', DataSourceTask.mode == 'demo')
        ).all()
    ]
    if demo_task_ids:
        deleted['data_source_logs'] += DataSourceLog.query.filter(
            DataSourceLog.task_id.in_(demo_task_ids)
        ).delete(synchronize_session=False)
        deleted['data_source_tasks'] += DataSourceTask.query.filter(
            DataSourceTask.id.in_(demo_task_ids)
        ).delete(synchronize_session=False)
    else:
        skipped.append('暂无演示抓取任务需要清理')

    deleted['trend_notes'] += TrendNote.query.filter_by(
        source_channel='demo_seed'
    ).delete(synchronize_session=False)

    deleted['topic_ideas'] += TopicIdea.query.filter(
        TopicIdea.review_note.contains('演示数据：')
    ).delete(synchronize_session=False)

    deleted['asset_library_items'] += AssetLibrary.query.filter_by(
        model_name='demo-svg'
    ).delete(synchronize_session=False)

    deleted['asset_generation_tasks'] += AssetGenerationTask.query.filter_by(
        model_name='demo-svg'
    ).delete(synchronize_session=False)

    deleted['creator_posts'] += CreatorPost.query.filter_by(
        source_channel='demo_seed'
    ).delete(synchronize_session=False)
    deleted['creator_snapshots'] += CreatorAccountSnapshot.query.filter_by(
        source_channel='demo_seed'
    ).delete(synchronize_session=False)
    deleted['creator_accounts'] += CreatorAccount.query.filter_by(
        source_channel='demo_seed'
    ).delete(synchronize_session=False)

    deleted['operation_logs'] += OperationLog.query.filter_by(
        action='bootstrap_demo_data'
    ).delete(synchronize_session=False)

    deleted_dict = {key: value for key, value in dict(deleted).items() if value}
    _log_operation('clear_demo_data', 'system', message='清理演示运营数据', detail={
        'deleted': deleted_dict,
        'skipped': skipped,
    })
    return {
        'deleted': deleted_dict,
        'skipped': skipped,
        'message': '演示运营数据已清理完成' if deleted_dict else '当前没有可清理的演示数据',
    }


def _latest_worker_ping_snapshot():
    log = OperationLog.query.filter(
        OperationLog.action.in_(['worker_ping_check', 'worker_ping_check_failed'])
    ).order_by(OperationLog.created_at.desc(), OperationLog.id.desc()).first()
    if not log:
        return {
            'has_result': False,
            'status': 'unknown',
            'status_label': '未检查',
            'checked_at': '',
            'message': '尚未执行 Worker 联通检查',
            'task_id': '',
            'state': '',
            'elapsed_ms': 0,
            'error': '',
            'error_type': '',
            'response': None,
        }

    detail = _deserialize_operation_detail(log.detail)
    status = (detail.get('status') or ('success' if log.action == 'worker_ping_check' else 'failed')).strip() or 'failed'
    label_map = {
        'success': '成功',
        'timeout': '超时',
        'dispatch_failed': '派发失败',
        'failed': '失败',
    }
    response = detail.get('response')
    if not isinstance(response, (dict, list, str, int, float, bool, type(None))):
        response = str(response)
    return {
        'has_result': True,
        'status': status,
        'status_label': label_map.get(status, status),
        'checked_at': _format_datetime(log.created_at),
        'message': detail.get('message') or log.message or '',
        'task_id': detail.get('task_id') or '',
        'state': detail.get('state') or '',
        'elapsed_ms': _safe_int(detail.get('elapsed_ms'), 0),
        'error': detail.get('error') or '',
        'error_type': detail.get('error_type') or '',
        'response': response,
    }


def _deployment_config_value(key, runtime_config=None):
    runtime_config = runtime_config or _automation_runtime_config()
    env_value = (os.environ.get(key) or '').strip()
    if env_value:
        return env_value, 'env'

    runtime_key_map = {
        'HOTWORD_FETCH_MODE': 'hotword_fetch_mode',
        'HOTWORD_API_URL': 'hotword_api_url',
        'HOTWORD_API_METHOD': 'hotword_api_method',
        'HOTWORD_API_HEADERS_JSON': 'hotword_api_headers_json',
        'HOTWORD_API_QUERY_JSON': 'hotword_api_query_json',
        'HOTWORD_API_BODY_JSON': 'hotword_api_body_json',
        'HOTWORD_RESULT_PATH': 'hotword_result_path',
        'HOTWORD_KEYWORD_PARAM': 'hotword_keyword_param',
        'HOTWORD_TIMEOUT_SECONDS': 'hotword_timeout_seconds',
        'HOTWORD_TREND_TYPE': 'hotword_trend_type',
        'HOTWORD_PAGE_SIZE': 'hotword_page_size',
        'HOTWORD_MAX_RELATED_QUERIES': 'hotword_max_related_queries',
        'HOTWORD_AUTO_GENERATE_TOPIC_IDEAS': 'hotword_auto_generate_topic_ideas',
        'HOTWORD_AUTO_GENERATE_TOPIC_COUNT': 'hotword_auto_generate_topic_count',
        'HOTWORD_AUTO_GENERATE_TOPIC_ACTIVITY_ID': 'hotword_auto_generate_topic_activity_id',
        'HOTWORD_AUTO_GENERATE_TOPIC_QUOTA': 'hotword_auto_generate_topic_quota',
        'HOTWORD_AUTO_CONVERT_CORPUS_TEMPLATES': 'hotword_auto_convert_corpus_templates',
        'HOTWORD_AUTO_CONVERT_CORPUS_LIMIT': 'hotword_auto_convert_corpus_limit',
        'CREATOR_SYNC_SOURCE_CHANNEL': 'creator_sync_source_channel',
        'CREATOR_SYNC_FETCH_MODE': 'creator_sync_fetch_mode',
        'CREATOR_SYNC_API_URL': 'creator_sync_api_url',
        'CREATOR_SYNC_API_METHOD': 'creator_sync_api_method',
        'CREATOR_SYNC_API_HEADERS_JSON': 'creator_sync_api_headers_json',
        'CREATOR_SYNC_API_QUERY_JSON': 'creator_sync_api_query_json',
        'CREATOR_SYNC_API_BODY_JSON': 'creator_sync_api_body_json',
        'CREATOR_SYNC_RESULT_PATH': 'creator_sync_result_path',
        'CREATOR_SYNC_TIMEOUT_SECONDS': 'creator_sync_timeout_seconds',
        'CREATOR_SYNC_BATCH_LIMIT': 'creator_sync_batch_limit',
        'CREATOR_SYNC_CURRENT_MONTH_ONLY': 'creator_sync_current_month_only',
        'CREATOR_SYNC_DATE_FROM': 'creator_sync_date_from',
        'CREATOR_SYNC_DATE_TO': 'creator_sync_date_to',
        'CREATOR_SYNC_MAX_POSTS_PER_ACCOUNT': 'creator_sync_max_posts_per_account',
        'ASSET_IMAGE_PROVIDER': 'image_provider',
        'ASSET_IMAGE_API_BASE': 'image_api_base',
        'ASSET_IMAGE_API_URL': 'image_api_url',
        'ASSET_IMAGE_MODEL': 'image_model',
        'ASSET_IMAGE_SIZE': 'image_size',
    }
    runtime_key = runtime_key_map.get(key)
    runtime_value = str(runtime_config.get(runtime_key) or '').strip() if runtime_key else ''
    if runtime_value:
        return runtime_value, 'runtime_config'
    return '', ''


def _build_deployment_helper_payload():
    runtime_config = _automation_runtime_config()
    hotword_settings = _hotword_runtime_settings()
    creator_sync_settings = _creator_sync_runtime_settings()
    capabilities = _image_provider_capabilities()
    env_guides = {
        'SECRET_KEY': {
            'label': 'Flask 密钥',
            'purpose': '用于登录会话、表单签名和后台安全校验。',
            'example': '<生成一串随机长密钥>',
            'placeholder': '<required secret>',
        },
        'ADMIN_USERNAME': {
            'label': '后台管理员账号',
            'purpose': '登录管理后台和自动化中心使用。',
            'example': 'furui',
        },
        'ADMIN_PASSWORD': {
            'label': '后台管理员密码',
            'purpose': '登录管理后台和自动化中心使用。',
            'example': '<strong password>',
            'placeholder': '<required password>',
        },
        'COPYWRITER_API_KEY': {
            'label': '文案模型 API Key',
            'purpose': '用于文案规划 Agent、文案生成、真人化重写等 AI 能力，优先支持 OpenAI 兼容接口。',
            'example': '<copywriter api key>',
            'placeholder': '<required api key>',
        },
        'COPYWRITER_API_URL': {
            'label': '文案模型 API URL',
            'purpose': '可填 OpenAI 兼容地址，例如 OpenAI / OpenRouter / DeepSeek 的 chat completions 地址或 base URL。',
            'example': 'https://api.openai.com/v1',
        },
        'COPYWRITER_MODEL': {
            'label': '文案模型名',
            'purpose': '用于文案规划 Agent、文案生成、真人化改写。',
            'example': 'gpt-5.4',
        },
        'DOUBAO_API_KEY': {
            'label': '豆包 API Key',
            'purpose': '如果文案第2或第3模型走豆包兼容接口，可在服务端保存豆包 Key。',
            'example': '<doubao api key>',
            'placeholder': '<optional api key>',
        },
        'DOUBAO_API_URL': {
            'label': '豆包 API URL',
            'purpose': '如果豆包提供 OpenAI 兼容地址，可作为文案第2或第3模型的 API URL。',
            'example': 'https://ark.cn-beijing.volces.com/api/v3',
        },
        'DOUBAO_MODEL': {
            'label': '豆包模型名',
            'purpose': '用于文案第2或第3模型位，例如豆包文案模型名。',
            'example': 'doubao-1.5-pro-32k',
        },
        'YUANBAO_API_KEY': {
            'label': '元宝 API Key',
            'purpose': '如果元宝提供标准/兼容 API，可作为文案第2或第3模型位的密钥。',
            'example': '<yuanbao api key>',
            'placeholder': '<optional api key>',
        },
        'YUANBAO_API_URL': {
            'label': '元宝 API URL',
            'purpose': '如果元宝提供 OpenAI 兼容地址，可配置到文案模型链中。',
            'example': 'https://api.example.com/v1',
        },
        'YUANBAO_MODEL': {
            'label': '元宝模型名',
            'purpose': '用于文案第2或第3模型位。',
            'example': 'yuanbao-chat',
        },
        'DEEPSEEK_API_KEY': {
            'label': 'DeepSeek API Key（兼容旧配置）',
            'purpose': '保留向后兼容；未设置 COPYWRITER_API_KEY 时，系统仍可继续用 DeepSeek。',
            'example': '<deepseek api key>',
            'placeholder': '<optional api key>',
        },
        'DATABASE_URL': {
            'label': 'PostgreSQL 连接串',
            'purpose': '主业务数据库，Web、Worker、Beat 都依赖它。',
            'example': 'postgresql://user:password@postgresql:5432/postgres',
            'placeholder': '<postgres url>',
        },
        'REDIS_URL': {
            'label': 'Redis 地址',
            'purpose': '缓存和异步任务基础依赖，建议与 Broker 保持同一实例。',
            'example': 'redis://redis:6379/0',
        },
        'CELERY_BROKER_URL': {
            'label': 'Celery Broker',
            'purpose': 'Worker/Beat 分发任务时使用的消息队列地址。',
            'example': 'redis://redis:6379/0',
        },
        'CELERY_RESULT_BACKEND': {
            'label': 'Celery Result Backend',
            'purpose': '保存任务执行结果，便于自动化中心查看状态。',
            'example': 'redis://redis:6379/1',
        },
        'ENABLE_AUTOMATION_BEAT': {
            'label': '启用 Beat 调度',
            'purpose': '控制定时任务派发服务是否生效。',
            'example': 'true',
        },
        'CELERY_BEAT_LOG_LEVEL': {
            'label': 'Beat 日志级别',
            'purpose': '建议保留 info，排查调度时更直观。',
            'example': 'info',
        },
        'HOTWORD_FETCH_MODE': {
            'label': '热点抓取模式',
            'purpose': 'remote 时走第三方热点 API，skeleton 时仅本地骨架联调。',
            'example': 'remote',
        },
        'HOTWORD_API_URL': {
            'label': '热点 API URL',
            'purpose': '热点接口入口地址，可填第三方接口，也可直接填 crawler 服务的 /xhs/trends。',
            'example': 'http://crawler:8081/xhs/trends',
        },
        'HOTWORD_RESULT_PATH': {
            'label': '热点结果路径',
            'purpose': '当第三方返回不是根数组时，用于指定 items 所在路径。',
            'example': 'data.items',
        },
        'HOTWORD_API_HEADERS_JSON': {
            'label': '热点请求头 JSON',
            'purpose': '放鉴权 token、会员 key 等头信息。',
            'example': '{"Authorization":"Bearer xxx"}',
        },
        'HOTWORD_API_QUERY_JSON': {
            'label': '热点 Query JSON',
            'purpose': '第三方接口要求 URL 参数时填写。',
            'example': '{"limit":20}',
        },
        'HOTWORD_API_BODY_JSON': {
            'label': '热点 Body JSON',
            'purpose': '第三方接口要求 POST Body 时填写。',
            'example': '{"keyword":"医疗"}',
        },
        'HOTWORD_TREND_TYPE': {
            'label': '热点抓取类型',
            'purpose': '本地 crawler 抓热点时，用 note_search 抓爆款笔记，用 hot_queries 抓相关热搜词。',
            'example': 'note_search',
        },
        'HOTWORD_PAGE_SIZE': {
            'label': '热点搜索结果条数',
            'purpose': '本地 crawler 抓爆款笔记时，限制每轮返回的搜索结果条数。',
            'example': '20',
        },
        'HOTWORD_MAX_RELATED_QUERIES': {
            'label': '热点相关搜索词条数',
            'purpose': '本地 crawler 抓热搜词时，限制每轮返回的相关搜索词条数。',
            'example': '20',
        },
        'CREATOR_SYNC_FETCH_MODE': {
            'label': '账号同步模式',
            'purpose': 'remote 时会调用 crawler 服务抓取用户账号和笔记。',
            'example': 'remote',
        },
        'CREATOR_SYNC_API_URL': {
            'label': 'Crawler API URL',
            'purpose': '报名人账号同步接口地址。',
            'example': 'http://crawler:8081/xhs/account_posts',
        },
        'CREATOR_SYNC_RESULT_PATH': {
            'label': 'Crawler 结果路径',
            'purpose': '当 crawler 返回结构不是根节点时指定路径。',
            'example': 'data',
        },
        'CREATOR_SYNC_API_HEADERS_JSON': {
            'label': 'Crawler 请求头 JSON',
            'purpose': '需要会员鉴权或签名时填写。',
            'example': '{"Authorization":"Bearer xxx"}',
        },
        'ASSET_IMAGE_PROVIDER': {
            'label': '图片 Provider',
            'purpose': '把图片中心从 SVG fallback 切到真实图片服务。',
            'example': 'volcengine_ark',
        },
        'OPENAI_IMAGE_API_KEY': {
            'label': 'OpenAI 图片 API Key',
            'purpose': '当图片 Provider 走 OpenAI 官方或兼容网关时可使用；也可直接复用 OPENAI_API_KEY。',
            'example': '<openai image api key>',
            'placeholder': '<optional api key>',
        },
        'ASSET_IMAGE_API_BASE': {
            'label': '图片 API Base',
            'purpose': 'OpenAI 兼容类接口常用的基础地址。',
            'example': 'https://ark.cn-beijing.volces.com/api/v3',
        },
        'ASSET_IMAGE_API_URL': {
            'label': '图片 API URL',
            'purpose': '直接指定图片生成接口地址。',
            'example': 'https://api.example.com/images/generations',
        },
        'ASSET_IMAGE_MODEL': {
            'label': '图片模型名',
            'purpose': '例如火山引擎模型名或其他供应商模型标识。',
            'example': 'doubao-seedream-3-0-t2i-250415',
        },
        'ASSET_IMAGE_SIZE': {
            'label': '图片尺寸',
            'purpose': '默认生成尺寸，建议与封面/海报规格一致。',
            'example': '1024x1536',
        },
        'CRAWLER_PROVIDER': {
            'label': 'Crawler Provider',
            'purpose': '先可用 mock 跑通，后续切换到 playwright_xhs。',
            'example': 'playwright_xhs',
        },
        'CRAWLER_PORT': {
            'label': 'Crawler 端口',
            'purpose': '独立 crawler 服务监听端口。',
            'example': '8081',
        },
        'PLAYWRIGHT_STORAGE_STATE_PATH': {
            'label': 'Playwright 登录态文件',
            'purpose': '真实抓小红书账号时复用登录态，减少人工扫码。',
            'example': '/app/crawler_service/.state/xhs_storage_state.json',
        },
        'XHS_SEARCH_URL_TEMPLATE': {
            'label': '小红书搜索 URL 模板',
            'purpose': 'crawler 抓热点/爆款笔记时使用的搜索地址模板。',
            'example': 'https://www.xiaohongshu.com/search_result?keyword={keyword}&source=web_explore_feed',
        },
        'XHS_POST_AUTHOR_SELECTOR': {
            'label': '笔记作者选择器',
            'purpose': 'Playwright 抓爆款笔记时提取作者昵称的自定义选择器。',
            'example': '[class*="author"]',
        },
        'XHS_SEARCH_RELATED_QUERY_SELECTOR': {
            'label': '相关搜索词选择器',
            'purpose': 'Playwright 抓热搜词时提取相关搜索词的自定义选择器。',
            'example': '[class*="related"] a',
        },
    }
    sensitive_keys = {
        'SECRET_KEY',
        'ADMIN_PASSWORD',
        'COPYWRITER_API_KEY',
        'DOUBAO_API_KEY',
        'YUANBAO_API_KEY',
        'OPENAI_API_KEY',
        'OPENAI_IMAGE_API_KEY',
        'DEEPSEEK_API_KEY',
        'ASSET_IMAGE_API_KEY',
        'ARK_API_KEY',
        'LAS_API_KEY',
        'POSTGRES_PASSWORD',
        'DATABASE_URL',
    }
    env_defaults = {
        'ADMIN_USERNAME': 'furui',
        'REDIS_URL': 'redis://redis:6379/0',
        'CELERY_BROKER_URL': 'redis://redis:6379/0',
        'CELERY_RESULT_BACKEND': 'redis://redis:6379/1',
        'INLINE_AUTOMATION_JOBS': os.environ.get('INLINE_AUTOMATION_JOBS', 'false') or 'false',
        'DEFAULT_TOPIC_QUOTA': str(_default_topic_quota()),
        'PREFERRED_URL_SCHEME': os.environ.get('PREFERRED_URL_SCHEME', 'https') or 'https',
        'SESSION_COOKIE_SECURE': os.environ.get('SESSION_COOKIE_SECURE', 'false') or 'false',
        'ENABLE_AUTOMATION_BEAT': os.environ.get('ENABLE_AUTOMATION_BEAT', 'true') or 'true',
        'CELERY_BEAT_LOG_LEVEL': os.environ.get('CELERY_BEAT_LOG_LEVEL', 'info') or 'info',
        'HOTWORD_FETCH_MODE': _resolved_hotword_mode(hotword_settings),
        'HOTWORD_API_URL': (hotword_settings.get('hotword_api_url') or ''),
        'HOTWORD_API_METHOD': (hotword_settings.get('hotword_api_method') or 'GET'),
        'HOTWORD_API_HEADERS_JSON': (hotword_settings.get('hotword_api_headers_json') or ''),
        'HOTWORD_API_QUERY_JSON': (hotword_settings.get('hotword_api_query_json') or ''),
        'HOTWORD_API_BODY_JSON': (hotword_settings.get('hotword_api_body_json') or ''),
        'HOTWORD_RESULT_PATH': (hotword_settings.get('hotword_result_path') or ''),
        'HOTWORD_KEYWORD_PARAM': (hotword_settings.get('hotword_keyword_param') or 'keyword'),
        'HOTWORD_TIMEOUT_SECONDS': str(hotword_settings.get('hotword_timeout_seconds') or 30),
        'HOTWORD_TREND_TYPE': (hotword_settings.get('hotword_trend_type') or 'note_search'),
        'HOTWORD_PAGE_SIZE': str(hotword_settings.get('hotword_page_size') or 20),
        'HOTWORD_MAX_RELATED_QUERIES': str(hotword_settings.get('hotword_max_related_queries') or 20),
        'HOTWORD_AUTO_GENERATE_TOPIC_IDEAS': 'true' if hotword_settings.get('hotword_auto_generate_topic_ideas') else 'false',
        'HOTWORD_AUTO_GENERATE_TOPIC_COUNT': str(hotword_settings.get('hotword_auto_generate_topic_count') or 20),
        'HOTWORD_AUTO_GENERATE_TOPIC_ACTIVITY_ID': str(hotword_settings.get('hotword_auto_generate_topic_activity_id') or 0),
        'HOTWORD_AUTO_GENERATE_TOPIC_QUOTA': str(hotword_settings.get('hotword_auto_generate_topic_quota') or _default_topic_quota()),
        'HOTWORD_AUTO_CONVERT_CORPUS_TEMPLATES': 'true' if hotword_settings.get('hotword_auto_convert_corpus_templates') else 'false',
        'HOTWORD_AUTO_CONVERT_CORPUS_LIMIT': str(hotword_settings.get('hotword_auto_convert_corpus_limit') or 10),
        'CREATOR_SYNC_SOURCE_CHANNEL': (creator_sync_settings.get('creator_sync_source_channel') or 'Crawler服务'),
        'CREATOR_SYNC_FETCH_MODE': _resolved_creator_sync_mode(creator_sync_settings),
        'CREATOR_SYNC_API_URL': (creator_sync_settings.get('creator_sync_api_url') or ''),
        'CREATOR_SYNC_API_METHOD': (creator_sync_settings.get('creator_sync_api_method') or 'POST'),
        'CREATOR_SYNC_API_HEADERS_JSON': (creator_sync_settings.get('creator_sync_api_headers_json') or ''),
        'CREATOR_SYNC_API_QUERY_JSON': (creator_sync_settings.get('creator_sync_api_query_json') or ''),
        'CREATOR_SYNC_API_BODY_JSON': (creator_sync_settings.get('creator_sync_api_body_json') or ''),
        'CREATOR_SYNC_RESULT_PATH': (creator_sync_settings.get('creator_sync_result_path') or ''),
        'CREATOR_SYNC_TIMEOUT_SECONDS': str(creator_sync_settings.get('creator_sync_timeout_seconds') or 60),
        'CREATOR_SYNC_BATCH_LIMIT': str(creator_sync_settings.get('creator_sync_batch_limit') or 20),
        'CREATOR_SYNC_CURRENT_MONTH_ONLY': 'true' if creator_sync_settings.get('creator_sync_current_month_only') else 'false',
        'CREATOR_SYNC_DATE_FROM': (creator_sync_settings.get('creator_sync_date_from') or ''),
        'CREATOR_SYNC_DATE_TO': (creator_sync_settings.get('creator_sync_date_to') or ''),
        'CREATOR_SYNC_MAX_POSTS_PER_ACCOUNT': str(creator_sync_settings.get('creator_sync_max_posts_per_account') or 60),
        'CRAWLER_PROVIDER': os.environ.get('CRAWLER_PROVIDER', 'mock') or 'mock',
        'CRAWLER_PORT': os.environ.get('CRAWLER_PORT', '8081') or '8081',
        'CRAWLER_REQUEST_TIMEOUT_SECONDS': os.environ.get('CRAWLER_REQUEST_TIMEOUT_SECONDS', '60') or '60',
        'XHS_PROFILE_URL_TEMPLATE': os.environ.get('XHS_PROFILE_URL_TEMPLATE', 'https://www.xiaohongshu.com/user/profile/{account_handle}') or 'https://www.xiaohongshu.com/user/profile/{account_handle}',
        'PLAYWRIGHT_HEADLESS': os.environ.get('PLAYWRIGHT_HEADLESS', 'true') or 'true',
        'PLAYWRIGHT_NAVIGATION_TIMEOUT_MS': os.environ.get('PLAYWRIGHT_NAVIGATION_TIMEOUT_MS', '30000') or '30000',
        'PLAYWRIGHT_STORAGE_STATE_PATH': os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH', '') or '',
        'PLAYWRIGHT_BROWSER_CHANNEL': os.environ.get('PLAYWRIGHT_BROWSER_CHANNEL', '') or '',
        'XHS_WAIT_AFTER_LOGIN_SECONDS': os.environ.get('XHS_WAIT_AFTER_LOGIN_SECONDS', '90') or '90',
        'XHS_DEBUG_OUTPUT_DIR': os.environ.get('XHS_DEBUG_OUTPUT_DIR', '/tmp/xhs_crawler_debug') or '/tmp/xhs_crawler_debug',
        'XHS_PROFILE_NAME_SELECTOR': os.environ.get('XHS_PROFILE_NAME_SELECTOR', '') or '',
        'XHS_FOLLOWER_COUNT_SELECTOR': os.environ.get('XHS_FOLLOWER_COUNT_SELECTOR', '') or '',
        'XHS_POST_CARD_SELECTOR': os.environ.get('XHS_POST_CARD_SELECTOR', '') or '',
        'XHS_POST_LINK_SELECTOR': os.environ.get('XHS_POST_LINK_SELECTOR', '') or '',
        'XHS_POST_TITLE_SELECTOR': os.environ.get('XHS_POST_TITLE_SELECTOR', '') or '',
        'XHS_POST_LIKES_SELECTOR': os.environ.get('XHS_POST_LIKES_SELECTOR', '') or '',
        'XHS_POST_COMMENTS_SELECTOR': os.environ.get('XHS_POST_COMMENTS_SELECTOR', '') or '',
        'XHS_POST_FAVORITES_SELECTOR': os.environ.get('XHS_POST_FAVORITES_SELECTOR', '') or '',
        'XHS_POST_VIEWS_SELECTOR': os.environ.get('XHS_POST_VIEWS_SELECTOR', '') or '',
        'XHS_POST_TIME_SELECTOR': os.environ.get('XHS_POST_TIME_SELECTOR', '') or '',
        'XHS_MAX_POSTS_PER_ACCOUNT': os.environ.get('XHS_MAX_POSTS_PER_ACCOUNT', '20') or '20',
        'MOCK_POSTS_PER_ACCOUNT': os.environ.get('MOCK_POSTS_PER_ACCOUNT', '2') or '2',
        'ASSET_IMAGE_PROVIDER': capabilities.get('image_provider_name') or 'svg_fallback',
        'ASSET_IMAGE_API_BASE': capabilities.get('image_provider_api_base') or '',
        'ASSET_IMAGE_API_URL': capabilities.get('image_provider_api_url') or '',
        'ASSET_IMAGE_MODEL': capabilities.get('image_provider_model') or '',
        'ASSET_IMAGE_SIZE': capabilities.get('image_provider_size') or '1024x1536',
    }

    def env_item(key, required=True, label=''):
        guide = env_guides.get(key, {})
        current_value, source = _deployment_config_value(key, runtime_config=runtime_config)
        display_value = ''
        if current_value:
            display_value = '<已配置，提交到目标环境时请填真实值>' if key in sensitive_keys else current_value
        elif key in env_defaults and env_defaults[key]:
            display_value = env_defaults[key]
        configured = bool(current_value)
        preview_value = current_value if (configured and key not in sensitive_keys) else (env_defaults.get(key, '') or '')
        if key in sensitive_keys:
            preview_value = guide.get('placeholder') or ('<required>' if required else '<optional>')
        copy_value = preview_value or guide.get('example') or (guide.get('placeholder') if key in sensitive_keys else '')
        return {
            'key': key,
            'label': label or guide.get('label') or key,
            'required': required,
            'configured': configured,
            'display_value': display_value or ('<未配置>' if required else '<可选>'),
            'preview_value': preview_value,
            'copy_value': copy_value,
            'source': source or ('default' if env_defaults.get(key) else ''),
            'sensitive': key in sensitive_keys,
            'purpose': guide.get('purpose') or '',
            'example': guide.get('example') or '',
            'placeholder': guide.get('placeholder') or '',
        }

    services = [
        {
            'key': 'web',
            'name': 'Web',
            'service_name': 'xhs-v4',
            'start_command': './docker/entrypoint-web.sh',
            'summary': '负责前台、后台、自动化中心和健康检查接口。',
            'required_envs': [
                env_item('SECRET_KEY'),
                env_item('ADMIN_USERNAME'),
                env_item('ADMIN_PASSWORD'),
                env_item('COPYWRITER_API_KEY'),
                env_item('DATABASE_URL'),
                env_item('REDIS_URL'),
                env_item('CELERY_BROKER_URL'),
                env_item('CELERY_RESULT_BACKEND'),
            ],
            'optional_envs': [
                env_item('DEFAULT_TOPIC_QUOTA', required=False),
                env_item('COPYWRITER_API_URL', required=False),
                env_item('COPYWRITER_MODEL', required=False),
                env_item('DEEPSEEK_API_KEY', required=False),
                env_item('PREFERRED_URL_SCHEME', required=False),
                env_item('SESSION_COOKIE_SECURE', required=False),
                env_item('HOTWORD_FETCH_MODE', required=False),
                env_item('HOTWORD_API_URL', required=False),
                env_item('HOTWORD_API_METHOD', required=False),
                env_item('HOTWORD_API_HEADERS_JSON', required=False),
                env_item('HOTWORD_API_QUERY_JSON', required=False),
                env_item('HOTWORD_API_BODY_JSON', required=False),
                env_item('HOTWORD_RESULT_PATH', required=False),
                env_item('HOTWORD_KEYWORD_PARAM', required=False),
                env_item('HOTWORD_TIMEOUT_SECONDS', required=False),
                env_item('HOTWORD_TREND_TYPE', required=False),
                env_item('HOTWORD_PAGE_SIZE', required=False),
                env_item('HOTWORD_MAX_RELATED_QUERIES', required=False),
                env_item('ASSET_IMAGE_PROVIDER', required=False),
                env_item('ASSET_IMAGE_API_BASE', required=False),
                env_item('ASSET_IMAGE_API_URL', required=False),
                env_item('ASSET_IMAGE_MODEL', required=False),
                env_item('ASSET_IMAGE_SIZE', required=False),
            ],
        },
        {
            'key': 'worker',
            'name': 'Worker',
            'service_name': 'xhs-v4-worker',
            'start_command': './docker/entrypoint-worker.sh',
            'summary': '负责候选话题生成、热点抓取、图片生成等异步任务。',
            'required_envs': [
                env_item('DATABASE_URL'),
                env_item('REDIS_URL'),
                env_item('CELERY_BROKER_URL'),
                env_item('CELERY_RESULT_BACKEND'),
                env_item('SECRET_KEY'),
            ],
            'optional_envs': [
                env_item('DEFAULT_TOPIC_QUOTA', required=False),
                env_item('COPYWRITER_API_KEY', required=False),
                env_item('COPYWRITER_API_URL', required=False),
                env_item('COPYWRITER_MODEL', required=False),
                env_item('DEEPSEEK_API_KEY', required=False),
                env_item('HOTWORD_FETCH_MODE', required=False),
                env_item('HOTWORD_API_URL', required=False),
                env_item('HOTWORD_API_METHOD', required=False),
                env_item('HOTWORD_API_HEADERS_JSON', required=False),
                env_item('HOTWORD_API_QUERY_JSON', required=False),
                env_item('HOTWORD_API_BODY_JSON', required=False),
                env_item('HOTWORD_RESULT_PATH', required=False),
                env_item('HOTWORD_KEYWORD_PARAM', required=False),
                env_item('HOTWORD_TIMEOUT_SECONDS', required=False),
                env_item('HOTWORD_TREND_TYPE', required=False),
                env_item('HOTWORD_PAGE_SIZE', required=False),
                env_item('HOTWORD_MAX_RELATED_QUERIES', required=False),
                env_item('ASSET_IMAGE_PROVIDER', required=False),
                env_item('ASSET_IMAGE_API_BASE', required=False),
                env_item('ASSET_IMAGE_API_URL', required=False),
                env_item('ASSET_IMAGE_MODEL', required=False),
                env_item('ASSET_IMAGE_SIZE', required=False),
            ],
        },
        {
            'key': 'beat',
            'name': 'Beat',
            'service_name': 'xhs-v4-beat',
            'start_command': './docker/entrypoint-beat.sh',
            'summary': '负责按照后台调度配置定时派发自动化任务。',
            'required_envs': [
                env_item('DATABASE_URL'),
                env_item('REDIS_URL'),
                env_item('CELERY_BROKER_URL'),
                env_item('CELERY_RESULT_BACKEND'),
                env_item('SECRET_KEY'),
                env_item('ENABLE_AUTOMATION_BEAT'),
            ],
            'optional_envs': [
                env_item('CELERY_BEAT_LOG_LEVEL', required=False),
                env_item('DEFAULT_TOPIC_QUOTA', required=False),
            ],
        },
        {
            'key': 'crawler',
            'name': 'Crawler',
            'service_name': 'xhs-v4-crawler',
            'start_command': 'python -m uvicorn crawler_service.main:app --host 0.0.0.0 --port 8081',
            'summary': '负责报名人账号同步接口与小红书页面抓取，可先用 mock，再切 playwright_xhs。',
            'required_envs': [
                env_item('CRAWLER_PROVIDER'),
                env_item('CRAWLER_PORT'),
            ],
            'optional_envs': [
                env_item('CRAWLER_REQUEST_TIMEOUT_SECONDS', required=False),
                env_item('XHS_PROFILE_URL_TEMPLATE', required=False),
                env_item('XHS_SEARCH_URL_TEMPLATE', required=False),
                env_item('PLAYWRIGHT_HEADLESS', required=False),
                env_item('PLAYWRIGHT_NAVIGATION_TIMEOUT_MS', required=False),
                env_item('PLAYWRIGHT_STORAGE_STATE_PATH', required=False),
                env_item('PLAYWRIGHT_BROWSER_CHANNEL', required=False),
                env_item('XHS_WAIT_AFTER_LOGIN_SECONDS', required=False),
                env_item('XHS_DEBUG_OUTPUT_DIR', required=False),
                env_item('XHS_PROFILE_NAME_SELECTOR', required=False),
                env_item('XHS_FOLLOWER_COUNT_SELECTOR', required=False),
                env_item('XHS_POST_CARD_SELECTOR', required=False),
                env_item('XHS_POST_LINK_SELECTOR', required=False),
                env_item('XHS_POST_TITLE_SELECTOR', required=False),
                env_item('XHS_POST_AUTHOR_SELECTOR', required=False),
                env_item('XHS_POST_LIKES_SELECTOR', required=False),
                env_item('XHS_POST_COMMENTS_SELECTOR', required=False),
                env_item('XHS_POST_FAVORITES_SELECTOR', required=False),
                env_item('XHS_POST_VIEWS_SELECTOR', required=False),
                env_item('XHS_POST_TIME_SELECTOR', required=False),
                env_item('XHS_SEARCH_RELATED_QUERY_SELECTOR', required=False),
                env_item('XHS_MAX_POSTS_PER_ACCOUNT', required=False),
                env_item('MOCK_POSTS_PER_ACCOUNT', required=False),
            ],
        },
    ]

    for service in services:
        required_items = service['required_envs']
        missing = [item['key'] for item in required_items if not item['configured']]
        service['ready'] = len(missing) == 0
        service['missing_required'] = missing
        service['missing_required_items'] = [item for item in required_items if not item['configured']]
        preview_lines = []
        for item in required_items + service['optional_envs']:
            if not item['preview_value'] and not item['required']:
                continue
            preview_lines.append(f"{item['key']}={item['preview_value']}")
        service['env_preview'] = '\n'.join(preview_lines)
        missing_preview_lines = []
        for item in service['missing_required_items']:
            missing_preview_lines.append(f"{item['key']}={item['copy_value'] or '<required>'}")
        service['missing_env_preview'] = '\n'.join(missing_preview_lines)

    missing_required_total = sum(len(item['missing_required']) for item in services)
    current_worker_ping = _latest_worker_ping_snapshot()
    return {
        'success': True,
        'summary': {
            'services_ready': len([item for item in services if item['ready']]),
            'services_total': len(services),
            'missing_required_total': missing_required_total,
            'worker_ping_status': current_worker_ping.get('status_label', '未检查'),
            'image_provider_name': capabilities.get('image_provider_name') or 'svg_fallback',
        },
        'services': services,
        'compose': {
            'commands': [
                'docker compose up -d --build',
                'docker compose ps',
                'docker compose logs -f web',
                'docker compose logs -f worker',
                'docker compose logs -f beat',
                'docker compose logs -f crawler',
            ],
            'healthchecks': [
                '/healthz',
                '/admin',
                '/automation_center',
            ],
        },
        'zeabur': {
            'recommended_order': [
                '先部署 web',
                '确认 /healthz 可访问',
                '再新增 worker',
                '再新增 beat',
                '如启用账号同步，再新增 crawler',
                '最后在自动化中心执行检测 Worker、热点接口与 crawler 健康检查',
            ],
            'services': [{
                'name': item['service_name'],
                'start_command': item['start_command'],
            } for item in services],
        },
        'docs': [
            'docs/xhs_v4_生产部署方案_v1_2026-04-02.md',
            'docs/xhs_v4_Zeabur与GitHub同步操作说明_v1_2026-04-07.md',
        ],
    }


def _build_deployment_blockers_payload():
    helper = _build_deployment_helper_payload()
    services = helper.get('services') or []
    blockers = []
    for service in services:
        missing = service.get('missing_required') or []
        if not missing:
            continue
        blockers.append({
            'service_key': service.get('key'),
            'service_name': service.get('name') or service.get('service_name') or service.get('key'),
            'missing_required': missing,
            'missing_items': service.get('missing_required_items') or [],
            'start_command': service.get('start_command') or '',
        })
    return blockers


def _build_launch_milestones_payload(hotword_health=None, creator_sync_health=None, image_health=None):
    helper = _build_deployment_helper_payload()
    services = {item.get('key'): item for item in (helper.get('services') or [])}
    service_matrix = {item.get('key'): item for item in _build_service_matrix_payload()}
    hotword_settings = _hotword_runtime_settings()
    creator_sync_settings = _creator_sync_runtime_settings()
    hotword_mode = _resolved_hotword_mode(hotword_settings)
    creator_sync_mode = _resolved_creator_sync_mode(creator_sync_settings)
    hotword_health = hotword_health or (_hotword_healthcheck(timeout_seconds=3) if hotword_mode == 'remote' else {'enabled': False, 'ok': False, 'message': '当前未启用真实热点接口'})
    creator_sync_health = creator_sync_health or (_creator_sync_healthcheck(timeout_seconds=3) if creator_sync_mode == 'remote' else {'enabled': False, 'ok': False, 'message': '当前未启用账号同步 crawler'})
    image_health = image_health or _image_provider_healthcheck(timeout_seconds=5)
    inline_jobs = _env_flag('INLINE_AUTOMATION_JOBS', False)
    beat_enabled = _coerce_bool(os.environ.get('ENABLE_AUTOMATION_BEAT', 'true'))

    def milestone_item(key, label, status, description, message, blockers=None, next_action=''):
        return {
            'key': key,
            'label': label,
            'status': status,
            'description': description,
            'message': message,
            'blockers': blockers or [],
            'next_action': next_action,
            'ok': status == 'ready',
        }

    def item_labels(items):
        return [item.get('label') or item.get('key') for item in (items or [])]

    milestones = []

    web_service = services.get('web') or {}
    web_missing = item_labels(web_service.get('missing_required_items'))
    if web_missing:
        milestones.append(milestone_item(
            'web_foundation',
            'Web 基础服务',
            'blocked',
            '前后台页面、健康检查和数据库连接的基础运行层。',
            f"Web 还缺少 {len(web_missing)} 项关键配置。",
            blockers=web_missing,
            next_action='先在当前 Web 服务补齐缺失环境变量，再确认 /healthz 和后台登录可访问。',
        ))
    else:
        milestones.append(milestone_item(
            'web_foundation',
            'Web 基础服务',
            'ready',
            '前后台页面、健康检查和数据库连接的基础运行层。',
            '当前 Web 主服务已经具备稳定运行条件。',
            next_action='保持当前 Web 服务为主入口，后续重点补齐异步链路和外部接口。',
        ))

    worker_ready = bool(service_matrix.get('worker', {}).get('ok'))
    beat_ready = bool(service_matrix.get('beat', {}).get('ok'))
    worker_missing = item_labels((services.get('worker') or {}).get('missing_required_items'))
    beat_missing = item_labels((services.get('beat') or {}).get('missing_required_items'))
    async_blockers = worker_missing + beat_missing
    if inline_jobs:
        milestones.append(milestone_item(
            'async_chain',
            '异步执行链',
            'pending_external',
            '负责热点抓取、账号同步、图片生成和定时任务。',
            '当前处于 inline 本地模式，本地联调可用，但生产仍建议单独部署 Worker / Beat。',
            blockers=['尚未切换到生产异步模式'],
            next_action='新增 xhs-v4-worker 和 xhs-v4-beat，并复制部署助手中的缺失模板。',
        ))
    elif worker_ready and ((not beat_enabled) or beat_ready):
        milestones.append(milestone_item(
            'async_chain',
            '异步执行链',
            'ready',
            '负责热点抓取、账号同步、图片生成和定时任务。',
            'Worker 与 Beat 已具备运行条件。',
            next_action='接下来重点验证真实热点源、Crawler 和图片接口。',
        ))
    else:
        if beat_enabled and not beat_ready and not beat_missing:
            async_blockers.append('Beat 服务尚未部署或未检测到可用 Broker')
        if not worker_ready and not worker_missing:
            async_blockers.append('Worker 尚未通过联通检查')
        milestones.append(milestone_item(
            'async_chain',
            '异步执行链',
            'blocked',
            '负责热点抓取、账号同步、图片生成和定时任务。',
            '当前生产异步链路还没有完全到位。',
            blockers=async_blockers,
            next_action='先补齐 Worker / Beat 服务和 Redis/Celery 连接，再执行“检测 Worker”。',
        ))

    if hotword_mode != 'remote':
        milestones.append(milestone_item(
            'hotword_pipeline',
            '热点抓取链路',
            'pending_external',
            '从第三方热点接口抓取内容，并自动生成候选话题。',
            '当前还在 skeleton / 本地模式，真实热点源尚未接入。',
            blockers=['待提供热点 API URL、鉴权和样例响应'],
            next_action='你买好热点会员或 API 后，把接口地址、鉴权头和样例 JSON 给我，我直接联调测试。',
        ))
    elif hotword_health.get('ok'):
        milestones.append(milestone_item(
            'hotword_pipeline',
            '热点抓取链路',
            'ready',
            '从第三方热点接口抓取内容，并自动生成候选话题。',
            hotword_health.get('message') or '热点远端接口可用。',
            next_action='下一步可以直接跑首轮真实 TrendNote 抓取并验证自动生题。',
        ))
    else:
        hotword_blockers = []
        if not (hotword_settings.get('hotword_api_url') or '').strip():
            hotword_blockers.append('热点 API URL')
        if not (hotword_settings.get('hotword_api_headers_json') or '').strip():
            hotword_blockers.append('热点鉴权头 / 会员凭据')
        if not (hotword_settings.get('hotword_result_path') or '').strip():
            hotword_blockers.append('结果路径（如返回非根数组）')
        if not hotword_blockers:
            hotword_blockers.append(hotword_health.get('message') or '热点接口联通失败')
        milestones.append(milestone_item(
            'hotword_pipeline',
            '热点抓取链路',
            'blocked',
            '从第三方热点接口抓取内容，并自动生成候选话题。',
            '热点接口已切到 remote，但还没联通成功。',
            blockers=hotword_blockers,
            next_action='在自动化中心先点“测试热点接口”，根据返回再修请求头、结果路径或超时设置。',
        ))

    if creator_sync_mode != 'remote':
        milestones.append(milestone_item(
            'creator_sync_pipeline',
            '账号同步链路',
            'pending_external',
            '抓取报名人后续发布的笔记，并持续更新互动数据。',
            'Crawler 真实接口尚未启用，当前还不能自动追踪报名人后续内容。',
            blockers=['待提供 crawler API / 第三方会员接口', '待部署独立 crawler 服务'],
            next_action='你买好第三方会员接口后，把 crawler API 结构给我，我直接接通线上账号同步。',
        ))
    elif creator_sync_health.get('ok'):
        milestones.append(milestone_item(
            'creator_sync_pipeline',
            '账号同步链路',
            'ready',
            '抓取报名人后续发布的笔记，并持续更新互动数据。',
            creator_sync_health.get('message') or '账号同步 crawler 服务可用。',
            next_action='下一步可以导入真实账号主页并验证新笔记累计与互动更新。',
        ))
    else:
        creator_blockers = []
        if not (creator_sync_settings.get('creator_sync_api_url') or '').strip():
            creator_blockers.append('Crawler API URL')
        if not (os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH') or '').strip():
            creator_blockers.append('Playwright 登录态 / 会员凭据')
        if not (creator_sync_settings.get('creator_sync_result_path') or '').strip():
            creator_blockers.append('结果路径（如 crawler 返回非根节点）')
        if not creator_blockers:
            creator_blockers.append(creator_sync_health.get('message') or '账号同步接口联通失败')
        milestones.append(milestone_item(
            'creator_sync_pipeline',
            '账号同步链路',
            'blocked',
            '抓取报名人后续发布的笔记，并持续更新互动数据。',
            '账号同步已切到 remote，但 crawler 还没联通成功。',
            blockers=creator_blockers,
            next_action='先在自动化中心点“测试 crawler 接口”，确保 /healthz 和返回结构都正常。',
        ))

    image_provider_name = (_image_provider_capabilities().get('image_provider_name') or 'svg_fallback').strip()
    image_real_enabled = image_provider_name != 'svg_fallback' and bool(_image_provider_capabilities().get('image_provider_configured'))
    if not image_real_enabled:
        milestones.append(milestone_item(
            'image_pipeline',
            '图片生成链路',
            'pending_external',
            '把图片中心从 SVG fallback 升级为真实图片生成服务。',
            '当前仍是 SVG fallback，真实图片 API 还没启用。',
            blockers=['待提供火山引擎或其他图片 API', '待配置模型名 / API Key'],
            next_action='等你把火山引擎 API 给我后，我会直接用图片调试沙盒完成联调。',
        ))
    elif image_health.get('ok'):
        milestones.append(milestone_item(
            'image_pipeline',
            '图片生成链路',
            'ready',
            '把图片中心从 SVG fallback 升级为真实图片生成服务。',
            image_health.get('message') or '图片远端接口可用。',
            next_action='下一步可以把正式图片任务切到远端生成，并验证素材库回流。',
        ))
    else:
        image_blockers = []
        capabilities = _image_provider_capabilities()
        if not capabilities.get('api_key_configured'):
            image_blockers.append('图片 API Key')
        if not capabilities.get('image_provider_api_url') and not capabilities.get('image_provider_api_base'):
            image_blockers.append('图片接口 URL / API Base')
        if not capabilities.get('image_provider_model'):
            image_blockers.append('图片模型名')
        if not image_blockers:
            image_blockers.append(image_health.get('message') or '图片接口联通失败')
        milestones.append(milestone_item(
            'image_pipeline',
            '图片生成链路',
            'blocked',
            '把图片中心从 SVG fallback 升级为真实图片生成服务。',
            '图片接口已配置，但当前还没联通成功。',
            blockers=image_blockers,
            next_action='先套用图片 provider 预设，再用图片调试沙盒逐项排查 API URL、模型和认证信息。',
        ))

    summary = {
        'total': len(milestones),
        'ready': len([item for item in milestones if item['status'] == 'ready']),
        'blocked': len([item for item in milestones if item['status'] == 'blocked']),
        'pending_external': len([item for item in milestones if item['status'] == 'pending_external']),
    }
    summary['message'] = f"已完成 {summary['ready']}/{summary['total']} 个关键里程碑，还差 {summary['blocked'] + summary['pending_external']} 项。"
    return {
        'summary': summary,
        'items': milestones,
    }


def _build_integration_checklist_payload():
    hotword_settings = _hotword_runtime_settings()
    creator_sync_settings = _creator_sync_runtime_settings()
    image_capabilities = _image_provider_capabilities()
    copywriter_capabilities = _resolve_copywriter_capabilities()

    def checklist_item(key, label, configured=False, value=''):
        return {
            'key': key,
            'label': label,
            'configured': bool(configured),
            'value': value if configured else '',
        }

    return [
        {
            'key': 'hotword_api',
            'label': '热点源接口',
            'description': '用于把第三方热点数据接入热点池和自动生题流程。',
            'items': [
                checklist_item('hotword_api_url', '接口 URL', bool(hotword_settings.get('hotword_api_url')), hotword_settings.get('hotword_api_url') or ''),
                checklist_item('hotword_api_method', '请求方法', bool(hotword_settings.get('hotword_api_method')), hotword_settings.get('hotword_api_method') or ''),
                checklist_item('hotword_result_path', '结果路径', bool(hotword_settings.get('hotword_result_path')), hotword_settings.get('hotword_result_path') or ''),
                checklist_item('hotword_api_headers_json', '鉴权请求头', bool(hotword_settings.get('hotword_api_headers_json')), hotword_settings.get('hotword_api_headers_json') or ''),
                checklist_item('sample_response', '样例响应 JSON', False, ''),
            ],
        },
        {
            'key': 'creator_sync',
            'label': '账号同步 / crawler',
            'description': '用于持续抓取报名人账号的新笔记和互动数据。',
            'items': [
                checklist_item('creator_sync_api_url', 'crawler 接口 URL', bool(creator_sync_settings.get('creator_sync_api_url')), creator_sync_settings.get('creator_sync_api_url') or ''),
                checklist_item('creator_sync_api_method', '请求方法', bool(creator_sync_settings.get('creator_sync_api_method')), creator_sync_settings.get('creator_sync_api_method') or ''),
                checklist_item('creator_sync_result_path', '结果路径', bool(creator_sync_settings.get('creator_sync_result_path')), creator_sync_settings.get('creator_sync_result_path') or ''),
                checklist_item('playwright_storage_state', '登录态 / 会员凭据', bool((os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH') or '').strip()), os.environ.get('PLAYWRIGHT_STORAGE_STATE_PATH') or ''),
                checklist_item('sample_profile', '测试账号主页链接', False, ''),
            ],
        },
        {
            'key': 'image_provider',
            'label': '图片接口',
            'description': '用于把图片中心从 SVG fallback 升级到真实图片生成。',
            'items': [
                checklist_item('image_provider', '图片 provider', bool(image_capabilities.get('image_provider_name')), image_capabilities.get('image_provider_name') or ''),
                checklist_item('image_api_url', '图片接口 URL', bool(image_capabilities.get('image_provider_api_url')), image_capabilities.get('image_provider_api_url') or ''),
                checklist_item('image_model', '模型名', bool(image_capabilities.get('image_provider_model')), image_capabilities.get('image_provider_model') or ''),
                checklist_item('image_api_key', 'API Key', bool(image_capabilities.get('api_key_configured')), '<configured>' if image_capabilities.get('api_key_configured') else ''),
                checklist_item('image_prompt_case', '测试提示词案例', True, '生成一张小红书医疗科普封面测试图'),
            ],
        },
        {
            'key': 'copywriter',
            'label': '文案模型',
            'description': '用于文案规划 Agent、文案生成和真人化重写。',
            'items': [
                checklist_item('copywriter_provider', '文案模型 Provider', bool(copywriter_capabilities.get('copywriter_provider')), copywriter_capabilities.get('copywriter_provider') or ''),
                checklist_item('copywriter_api_url', '文案模型 API URL', bool(copywriter_capabilities.get('copywriter_api_url')), copywriter_capabilities.get('copywriter_api_url') or ''),
                checklist_item('copywriter_model', '文案模型名', bool(copywriter_capabilities.get('copywriter_model')), copywriter_capabilities.get('copywriter_model') or ''),
                checklist_item('copywriter_api_key', 'API Key', bool(copywriter_capabilities.get('api_key_configured')), '<configured>' if copywriter_capabilities.get('api_key_configured') else ''),
                checklist_item('copywriter_case', '测试提示词案例', True, '请用更像真人的小红书口语风，写一句关于肝健康的开头。'),
            ],
        },
    ]


def _creator_sync_healthcheck(timeout_seconds=3):
    settings = _creator_sync_runtime_settings()
    api_url = (settings.get('creator_sync_api_url') or '').strip()
    mode = _resolved_creator_sync_mode(settings)
    if mode != 'remote':
        return {
            'enabled': False,
            'ok': False,
            'message': '账号同步模式未启用 remote',
            'health_url': '',
            'status_code': 0,
            'response': None,
        }
    if not api_url:
        return {
            'enabled': True,
            'ok': False,
            'message': '未配置账号同步 API URL',
            'health_url': '',
            'status_code': 0,
            'response': None,
        }

    parsed = urlparse(api_url)
    if parsed.scheme and parsed.netloc:
        health_url = urlunparse((parsed.scheme, parsed.netloc, '/healthz', '', '', ''))
    else:
        health_url = api_url.rstrip('/') + '/healthz'

    try:
        response = requests.get(health_url, timeout=min(max(_safe_int(timeout_seconds, 3), 1), 15))
        payload = None
        try:
            payload = response.json()
        except ValueError:
            payload = response.text[:300]
        return {
            'enabled': True,
            'ok': response.ok,
            'message': '账号同步 crawler 服务可用' if response.ok else f'账号同步 crawler 服务返回 {response.status_code}',
            'health_url': health_url,
            'status_code': response.status_code,
            'response': payload,
        }
    except Exception as exc:
        return {
            'enabled': True,
            'ok': False,
            'message': f'账号同步 crawler 服务不可达：{exc}',
            'health_url': health_url,
            'status_code': 0,
            'response': None,
        }


def _read_crawler_probe_file(path):
    current_path = (path or '').strip()
    if not current_path or not os.path.exists(current_path):
        return {}
    try:
        with open(current_path, 'r', encoding='utf-8') as handle:
            payload = json.load(handle)
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


CRAWLER_HINT_BASELINES = {
    'views': {'view', 'views', 'read', 'reads', 'browse', 'browses', 'pv', 'reading', 'traffic'},
    'exposures': {'impression', 'impressions', 'exposure', 'exposures', 'expo', 'reach', 'show', 'display', 'impr'},
    'hot': {'hot', 'hotvalue', 'score', 'heat', 'trend', 'searchcnt', 'search', 'cnt'},
}

CRAWLER_HINT_STOP_TOKENS = {
    'note', 'notecard', 'card', 'metrics', 'metric', 'interact', 'info', 'item', 'items', 'data',
    'list', 'user', 'profile', 'result', 'results', 'path', 'title', 'desc', 'text', 'value',
    'count', 'counts', 'num', 'nums', 'state', 'page', 'pages', 'feed', 'feeds',
}


def _tokenize_metric_path(path):
    raw = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', str(path or ''))
    tokens = []
    for token in re.split(r'[^a-zA-Z0-9]+', raw.lower()):
        current = token.strip()
        if not current or current.isdigit() or current in CRAWLER_HINT_STOP_TOKENS:
            continue
        if len(current) < 3 and current not in {'pv'}:
            continue
        if current not in tokens:
            tokens.append(current)
    return tokens


def _classify_metric_candidate_path(path):
    lowered = str(path or '').lower()
    if any(token in lowered for token in ['impression', 'exposure', 'expo', 'reach', 'deliver', 'show', 'display', 'impr']):
        return 'exposures'
    if any(token in lowered for token in ['view', 'read', 'browse', 'pv', 'traffic']):
        return 'views'
    if any(token in lowered for token in ['hot', 'score', 'heat', 'trend', 'search_cnt']):
        return 'hot'
    return ''


def _build_metric_hint_suggestions(metric_candidates):
    suggestions = {
        'views': {},
        'exposures': {},
        'hot': {},
    }
    for item in metric_candidates or []:
        if not isinstance(item, dict):
            continue
        bucket = _classify_metric_candidate_path(item.get('path'))
        if not bucket:
            continue
        for token in _tokenize_metric_path(item.get('path')):
            if token in CRAWLER_HINT_BASELINES[bucket]:
                continue
            suggestions[bucket][token] = suggestions[bucket].get(token, 0) + 1

    ranked_tokens = {}
    plain_tokens = {}
    for bucket, token_map in suggestions.items():
        ranked = sorted(token_map.items(), key=lambda item: (-item[1], item[0]))
        ranked_tokens[bucket] = [
            {'token': token, 'count': count}
            for token, count in ranked[:12]
        ]
        plain_tokens[bucket] = [item['token'] for item in ranked_tokens[bucket]]

    env_examples = {}
    if plain_tokens['views']:
        env_examples['XHS_VIEWS_HINT_TOKENS'] = ','.join(plain_tokens['views'][:8])
    if plain_tokens['exposures']:
        env_examples['XHS_EXPOSURE_HINT_TOKENS'] = ','.join(plain_tokens['exposures'][:8])
    if plain_tokens['hot']:
        env_examples['XHS_HOT_HINT_TOKENS'] = ','.join(plain_tokens['hot'][:8])
    env_preview = '\n'.join(f'{key}={value}' for key, value in env_examples.items())
    return {
        'tokens': plain_tokens,
        'ranked_tokens': ranked_tokens,
        'env_examples': env_examples,
        'env_preview': env_preview,
    }


def _build_crawler_probe_payload():
    debug_output_dir = (os.environ.get('XHS_DEBUG_OUTPUT_DIR') or '/tmp/xhs_crawler_debug').strip() or '/tmp/xhs_crawler_debug'
    probe_defs = [
        ('login_verify', '登录态验证', 'xhs_login_verify.json'),
        ('trend_probe_note_search', '热点探测-爆款笔记', 'xhs_trends_probe_note_search.json'),
        ('trend_probe_hot_queries', '热点探测-热搜词', 'xhs_trends_probe_hot_queries.json'),
        ('account_probe', '账号探测', 'xhs_account_probe.json'),
        ('bundle_probe', '整包探测', 'xhs_probe_bundle.json'),
    ]

    items = []
    payload_map = {}
    debug_hints = {
        'search': {},
        'profile': {},
    }

    def resolve_metric_sources(payload):
        metric_sources = payload.get('metric_sources') or {}
        if metric_sources:
            return metric_sources
        for key in ['sample_items', 'sample_posts']:
            rows = payload.get(key) or []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                row_sources = row.get('metric_sources') or {}
                if row_sources:
                    return row_sources
        return {}

    for key, label, filename in probe_defs:
        path = os.path.join(debug_output_dir, filename)
        exists = os.path.exists(path)
        payload = _read_crawler_probe_file(path) if exists else {}
        payload_map[key] = payload
        diagnosis = payload.get('diagnosis') or {}
        summary = diagnosis.get('summary') or ('已生成探测文件' if exists else '暂无探测结果')
        provider = payload.get('provider') or (payload.get('health') or {}).get('provider') or ''
        suggested_actions = diagnosis.get('suggested_actions') or []
        metric_sources = resolve_metric_sources(payload)
        metric_coverage = payload.get('metric_coverage') or {}
        updated_at = _format_datetime(datetime.fromtimestamp(os.path.getmtime(path))) if exists else ''
        items.append({
            'key': key,
            'label': label,
            'exists': exists,
            'path': path,
            'updated_at': updated_at,
            'provider': provider,
            'status': diagnosis.get('status') or ('ready' if exists else 'missing'),
            'summary': summary,
            'suggested_actions': suggested_actions[:5],
            'metric_sources': metric_sources,
            'metric_source_summary_text': _metric_source_summary_text(metric_sources),
            'metric_coverage': metric_coverage if isinstance(metric_coverage, dict) else {},
        })

    def resolve_debug_metric_candidates(payload):
        candidates = payload.get('state_metric_candidates') or []
        return candidates if isinstance(candidates, list) else []

    search_debug_payload = _read_crawler_probe_file(os.path.join(debug_output_dir, 'xhs_search_debug.json'))
    profile_debug_payload = _read_crawler_probe_file(os.path.join(debug_output_dir, 'xhs_profile_debug.json'))
    debug_hints['search'] = {
        'exists': bool(search_debug_payload),
        'path': os.path.join(debug_output_dir, 'xhs_search_debug.json'),
        'metric_candidates': resolve_debug_metric_candidates(search_debug_payload)[:15],
        'state_note_metric_sources': (search_debug_payload.get('state_note_metric_sources') or [])[:5] if isinstance(search_debug_payload, dict) else [],
        'hint_suggestions': _build_metric_hint_suggestions(resolve_debug_metric_candidates(search_debug_payload)[:30]),
    }
    debug_hints['profile'] = {
        'exists': bool(profile_debug_payload),
        'path': os.path.join(debug_output_dir, 'xhs_profile_debug.json'),
        'metric_candidates': resolve_debug_metric_candidates(profile_debug_payload)[:15],
        'state_post_metric_sources': (profile_debug_payload.get('state_post_metric_sources') or [])[:5] if isinstance(profile_debug_payload, dict) else [],
        'hint_suggestions': _build_metric_hint_suggestions(resolve_debug_metric_candidates(profile_debug_payload)[:30]),
    }

    latest_success = next((item for item in items if item['exists']), None)
    existing_items = [item for item in items if item['exists']]
    bundle_payload = payload_map.get('bundle_probe') or {}
    bundle_diagnosis = bundle_payload.get('diagnosis') or {}
    summary_status = 'pending'
    summary_message = '尚未运行 crawler 联调探测，建议先执行 verify/probe 脚本。'
    summary_actions = ['先运行 crawler_service/scripts/verify_xhs_login_state.py 验证登录态']
    metric_highlights = []
    if existing_items:
        if bundle_diagnosis:
            summary_status = bundle_diagnosis.get('status') or 'partial'
            summary_message = bundle_diagnosis.get('summary') or 'crawler 联调已有结果'
            summary_actions = (bundle_diagnosis.get('suggested_actions') or [])[:5]
            for scope_key, scope_label in [('trends', '热点'), ('account_posts', '账号')]:
                coverage = ((bundle_payload.get(scope_key) or {}).get('metric_coverage') or {})
                if not isinstance(coverage, dict):
                    continue
                for metric_key in ['views', 'exposures', 'hot_value']:
                    metric_row = coverage.get(metric_key) or {}
                    if not metric_row:
                        continue
                    metric_highlights.append(
                        f"{scope_label}{metric_key}={metric_row.get('hit_count', 0)}/{metric_row.get('total_count', 0)}"
                    )
        else:
            statuses = {item.get('status') for item in existing_items}
            if 'blocked' in statuses:
                summary_status = 'blocked'
                summary_message = '已有 crawler 联调结果，但至少有一条链路被阻塞。'
            elif 'partial' in statuses:
                summary_status = 'partial'
                summary_message = 'crawler 联调已部分可用，但还需要继续校准。'
            else:
                summary_status = 'ready'
                summary_message = '已有 crawler 联调结果，可以继续在主站里验证。'
            summary_actions = []
            for item in existing_items:
                summary_actions.extend(item.get('suggested_actions') or [])
            summary_actions = _unique(summary_actions)[:5]

    return {
        'debug_output_dir': debug_output_dir,
        'items': items,
        'has_any_result': any(item['exists'] for item in items),
        'latest_result_label': latest_success.get('label') if latest_success else '',
        'debug_hints': debug_hints,
        'summary': {
            'status': summary_status,
            'message': summary_message,
            'suggested_actions': summary_actions,
            'metric_highlights': metric_highlights[:6],
        },
    }


def _build_recent_failed_jobs_payload(limit=10):
    safe_limit = min(max(_safe_int(limit, 10), 1), 50)
    items = []

    failed_data_source_tasks = DataSourceTask.query.filter_by(status='failed').order_by(
        DataSourceTask.finished_at.desc(), DataSourceTask.updated_at.desc(), DataSourceTask.created_at.desc(), DataSourceTask.id.desc()
    ).limit(safe_limit).all()
    for task in failed_data_source_tasks:
        kind_label = '报名人账号同步任务' if task.task_type == 'creator_account_sync' else '热点抓取任务'
        retry_label = '重试账号同步' if task.task_type == 'creator_account_sync' else '重试抓取'
        items.append({
            'kind': 'data_source_task',
            'kind_label': kind_label,
            'id': task.id,
            'title': task.batch_name or task.task_type or kind_label,
            'occurred_at': _format_datetime(task.finished_at or task.updated_at or task.created_at),
            'status': task.status or 'failed',
            'status_label': task.status or 'failed',
            'message': task.message or f'{kind_label}失败',
            'task_id': task.celery_task_id or '',
            'source_label': f'{task.source_platform or "-"} / {task.source_channel or "-"}',
            'detail': {
                'task_type': task.task_type,
                'source_platform': task.source_platform or '',
                'source_channel': task.source_channel or '',
                'batch_name': task.batch_name or '',
                'celery_task_id': task.celery_task_id or '',
                'result_payload': _load_json_value(task.result_payload, {}),
            },
            'retry': {
                'type': 'data_source_task',
                'id': task.id,
                'label': retry_label,
            },
        })

    failed_asset_tasks = AssetGenerationTask.query.filter_by(status='failed').order_by(
        AssetGenerationTask.finished_at.desc(), AssetGenerationTask.updated_at.desc(), AssetGenerationTask.created_at.desc(), AssetGenerationTask.id.desc()
    ).limit(safe_limit).all()
    for task in failed_asset_tasks:
        items.append({
            'kind': 'asset_generation_task',
            'kind_label': '图片生成任务',
            'id': task.id,
            'title': task.title_hint or f'图片任务 #{task.id}',
            'occurred_at': _format_datetime(task.finished_at or task.updated_at or task.created_at),
            'status': task.status or 'failed',
            'status_label': task.status or 'failed',
            'message': task.message or '图片生成失败',
            'task_id': task.celery_task_id or '',
            'source_label': task.source_provider or '-',
            'detail': {
                'topic_id': task.topic_id,
                'registration_id': task.registration_id,
                'source_provider': task.source_provider or '',
                'style_preset': task.style_preset or '',
                'celery_task_id': task.celery_task_id or '',
                'result_payload': _load_json_value(task.result_payload, []),
            },
            'retry': {
                'type': 'asset_generation_task',
                'id': task.id,
                'label': '重试图片',
            },
        })

    failed_schedules = AutomationSchedule.query.filter_by(last_status='failed').order_by(
        AutomationSchedule.last_run_at.desc(), AutomationSchedule.updated_at.desc(), AutomationSchedule.id.desc()
    ).limit(safe_limit).all()
    for schedule in failed_schedules:
        items.append({
            'kind': 'automation_schedule',
            'kind_label': '自动调度',
            'id': schedule.id,
            'title': schedule.name or schedule.job_key or f'调度 #{schedule.id}',
            'occurred_at': _format_datetime(schedule.last_run_at or schedule.updated_at or schedule.created_at),
            'status': schedule.last_status or 'failed',
            'status_label': schedule.last_status or 'failed',
            'message': schedule.last_message or '调度执行失败',
            'task_id': schedule.last_celery_task_id or '',
            'source_label': schedule.task_type or '-',
            'detail': {
                'job_key': schedule.job_key,
                'task_type': schedule.task_type,
                'enabled': bool(schedule.enabled),
                'interval_minutes': schedule.interval_minutes or 0,
                'last_celery_task_id': schedule.last_celery_task_id or '',
                'params_payload': _load_json_value(schedule.params_payload, {}),
            },
            'retry': {
                'type': 'automation_schedule',
                'id': schedule.id,
                'label': '立即执行',
            },
        })

    failed_ping_logs = OperationLog.query.filter_by(action='worker_ping_check_failed').order_by(
        OperationLog.created_at.desc(), OperationLog.id.desc()
    ).limit(safe_limit).all()
    for log in failed_ping_logs:
        detail = _deserialize_operation_detail(log.detail)
        items.append({
            'kind': 'worker_ping',
            'kind_label': 'Worker 联通检查',
            'id': log.id,
            'title': 'Worker Ping',
            'occurred_at': _format_datetime(log.created_at),
            'status': detail.get('status') or 'failed',
            'status_label': detail.get('status') or 'failed',
            'message': detail.get('message') or log.message or 'Worker 联通检查失败',
            'task_id': detail.get('task_id') or '',
            'source_label': detail.get('state') or '-',
            'detail': detail,
            'retry': {
                'type': 'worker_ping',
                'id': 0,
                'label': '重新检测',
            },
        })

    integration_failed_actions = [
        'hotword_ping_check_failed',
        'creator_sync_ping_check_failed',
        'copywriter_ping_check_failed',
        'image_provider_ping_check_failed',
    ]
    failed_integration_logs = OperationLog.query.filter(
        OperationLog.action.in_(integration_failed_actions)
    ).order_by(OperationLog.created_at.desc(), OperationLog.id.desc()).limit(safe_limit).all()
    retry_type_map = {
        'hotword_ping_check_failed': ('hotword_ping', '重试热点接口'),
        'creator_sync_ping_check_failed': ('creator_sync_ping', '重试 crawler 接口'),
        'copywriter_ping_check_failed': ('copywriter_ping', '重试文案模型'),
        'image_provider_ping_check_failed': ('image_ping', '重试图片接口'),
    }
    for log in failed_integration_logs:
        detail = _deserialize_operation_detail(log.detail)
        retry_type, retry_label = retry_type_map.get(log.action, ('', '重试'))
        items.append({
            'kind': 'integration_ping',
            'kind_label': detail.get('label') or '接口联调',
            'id': log.id,
            'title': detail.get('label') or '接口联调',
            'occurred_at': detail.get('checked_at') or _format_datetime(log.created_at),
            'status': detail.get('status') or 'failed',
            'status_label': 'failed',
            'message': detail.get('message') or log.message or '接口联调失败',
            'task_id': detail.get('health_url') or detail.get('provider') or '',
            'source_label': detail.get('integration_key') or '-',
            'detail': detail,
            'retry': {
                'type': retry_type,
                'id': 0,
                'label': retry_label,
            },
        })

    items.sort(key=lambda item: item.get('occurred_at') or '', reverse=True)
    trimmed = items[:safe_limit]
    return {
        'success': True,
        'summary': {
            'count': len(trimmed),
            'limit': safe_limit,
        },
        'items': trimmed,
    }


def _build_service_matrix_payload():
    inline_jobs = _env_flag('INLINE_AUTOMATION_JOBS', False)
    beat_enabled = _coerce_bool(os.environ.get('ENABLE_AUTOMATION_BEAT', 'true'))
    last_worker_ping = _latest_worker_ping_snapshot()
    worker_ok = inline_jobs or last_worker_ping.get('status') == 'success'
    worker_message = (
        '本地 inline 模式已启用，Worker 可选'
        if inline_jobs else
        (last_worker_ping.get('message') or ('最近一次 Worker 联通检查成功' if worker_ok else '尚未检测到可用 Worker'))
    )

    hotword_settings = _hotword_runtime_settings()
    hotword_mode = _resolved_hotword_mode(hotword_settings)
    hotword_health = _hotword_healthcheck(timeout_seconds=2) if hotword_mode == 'remote' else None

    creator_sync_settings = _creator_sync_runtime_settings()
    creator_sync_mode = _resolved_creator_sync_mode(creator_sync_settings)
    creator_sync_health = _creator_sync_healthcheck(timeout_seconds=2) if creator_sync_mode == 'remote' else None

    copywriter_capabilities = _resolve_copywriter_capabilities()
    copywriter_health = _copywriter_healthcheck(timeout_seconds=20) if copywriter_capabilities.get('copywriter_configured') else None
    image_health = _image_provider_healthcheck(timeout_seconds=5)

    return [
        {
            'key': 'web',
            'label': 'Web',
            'ok': True,
            'status': 'ready',
            'message': '当前服务已启动并提供页面与 API',
        },
        {
            'key': 'worker',
            'label': 'Worker',
            'ok': worker_ok,
            'status': 'ready' if worker_ok else 'missing',
            'message': worker_message,
        },
        {
            'key': 'beat',
            'label': 'Beat',
            'ok': (not beat_enabled) or inline_jobs or bool((os.environ.get('CELERY_BROKER_URL') or '').strip()),
            'status': 'ready' if ((not beat_enabled) or inline_jobs or bool((os.environ.get('CELERY_BROKER_URL') or '').strip())) else 'missing',
            'message': (
                '当前未启用 Beat'
                if not beat_enabled else
                ('本地 inline 模式下无需 Beat' if inline_jobs else '需单独部署 beat 服务来执行定时调度')
            ),
        },
        {
            'key': 'crawler',
            'label': 'Crawler',
            'ok': creator_sync_mode == 'disabled' or bool(creator_sync_health and creator_sync_health.get('ok')),
            'status': 'ready' if (creator_sync_mode == 'disabled' or bool(creator_sync_health and creator_sync_health.get('ok'))) else 'missing',
            'message': (
                '账号同步未启用'
                if creator_sync_mode == 'disabled' else
                (creator_sync_health.get('message') if creator_sync_health else '账号同步 crawler 服务未检测')
            ),
        },
        {
            'key': 'copywriter_remote',
            'label': '文案模型',
            'ok': bool(copywriter_health and copywriter_health.get('ok')),
            'status': 'ready' if bool(copywriter_health and copywriter_health.get('ok')) else ('missing' if copywriter_capabilities.get('copywriter_configured') else 'blocked'),
            'message': (
                copywriter_health.get('message')
                if copywriter_health else
                ('文案模型未配置' if not copywriter_capabilities.get('copywriter_configured') else '文案模型未检测')
            ),
        },
        {
            'key': 'image_remote',
            'label': '图片远端接口',
            'ok': provider_ok if (provider_ok := (not image_health.get('enabled') or bool(image_health.get('ok')))) else False,
            'status': 'ready' if provider_ok else 'missing',
            'message': image_health.get('message') or '图片远端接口未检测',
        },
        {
            'key': 'hotword_remote',
            'label': '热点远端源',
            'ok': hotword_mode != 'remote' or bool(hotword_health and hotword_health.get('ok')),
            'status': 'ready' if (hotword_mode != 'remote' or bool(hotword_health and hotword_health.get('ok'))) else 'missing',
            'message': (
                '当前使用 skeleton 模式'
                if hotword_mode != 'remote' else
                (hotword_health.get('message') if hotword_health else '热点远端接口未检测')
            ),
        },
    ]


SCHEMA_REQUIRED_INDEXES = {
    'registration': [
        ('idx_registration_topic_account', ['topic_id', 'xhs_account']),
        ('idx_registration_phone', ['phone']),
        ('idx_registration_group_name', ['group_num', 'name']),
        ('idx_registration_created_at', ['created_at']),
    ],
    'submission': [
        ('idx_submission_registration_id', ['registration_id']),
        ('idx_submission_creator_account', ['xhs_creator_account_id']),
        ('idx_submission_primary_post', ['xhs_primary_post_id']),
        ('idx_submission_tracking_status', ['xhs_tracking_status']),
        ('idx_submission_title_skill', ['selected_title_skill']),
        ('idx_submission_image_skill', ['selected_image_skill']),
        ('idx_submission_strategy_updated', ['strategy_updated_at']),
    ],
    'topic': [
        ('idx_topic_activity_pool', ['activity_id', 'pool_status']),
        ('idx_topic_activity_published', ['activity_id', 'published_at']),
    ],
    'trend_note': [
        ('idx_trend_note_link', ['link']),
        ('idx_trend_note_title_keyword', ['title', 'keyword']),
        ('idx_trend_note_pool_platform', ['pool_status', 'source_platform']),
        ('idx_trend_note_created_at', ['created_at']),
        ('idx_trend_note_hot_score', ['hot_score']),
    ],
    'hot_topic_entry': [
        ('idx_hot_topic_status_created', ['status', 'created_at']),
        ('idx_hot_topic_lane_persona', ['lane_key', 'persona_key']),
        ('idx_hot_topic_reference_note', ['reference_note_id']),
    ],
    'asset_plan_draft': [
        ('idx_asset_plan_draft_status_created', ['status', 'created_at']),
        ('idx_asset_plan_draft_source_ref', ['source_type', 'source_id']),
    ],
    'topic_idea': [
        ('idx_topic_idea_status_activity', ['status', 'activity_id']),
        ('idx_topic_idea_created_at', ['created_at']),
    ],
    'asset_plan_draft': [
        ('idx_asset_plan_draft_status_created', ['status', 'created_at']),
        ('idx_asset_plan_draft_source_ref', ['source_type', 'source_id']),
    ],
    'data_source_task': [
        ('idx_data_source_task_type_status', ['task_type', 'status']),
        ('idx_data_source_task_created_at', ['created_at']),
    ],
    'creator_account': [
        ('idx_creator_account_platform_phone', ['platform', 'owner_phone']),
        ('idx_creator_account_platform_handle', ['platform', 'account_handle']),
        ('idx_creator_account_platform_profile', ['platform', 'profile_url']),
        ('idx_creator_account_last_synced', ['last_synced_at']),
    ],
    'creator_post': [
        ('idx_creator_post_account_postid', ['creator_account_id', 'platform_post_id']),
        ('idx_creator_post_account_posturl', ['creator_account_id', 'post_url']),
        ('idx_creator_post_registration', ['registration_id']),
        ('idx_creator_post_submission', ['submission_id']),
        ('idx_creator_post_publish_time', ['publish_time']),
    ],
    'creator_account_snapshot': [
        ('idx_creator_snapshot_account_date', ['creator_account_id', 'snapshot_date']),
    ],
    'operation_log': [
        ('idx_operation_log_action_created', ['action', 'created_at']),
    ],
    'announcement': [
        ('idx_announcement_status_priority', ['status', 'priority']),
    ],
    'corpus_entry': [
        ('idx_corpus_entry_reference_url', ['reference_url']),
        ('idx_corpus_entry_template_type', ['template_type_key']),
    ],
}


def _existing_index_names(table_name):
    try:
        indexes = inspect(db.engine).get_indexes(table_name)
    except Exception:
        return set()
    return {item.get('name') for item in indexes if item.get('name')}


def _ensure_indexes(conn, table_name, index_specs):
    existing_names = _existing_index_names(table_name)
    created = []
    for index_name, columns in index_specs:
        if index_name in existing_names:
            continue
        columns_sql = ', '.join(columns)
        try:
            conn.execute(text(f'CREATE INDEX IF NOT EXISTS {index_name} ON {table_name} ({columns_sql})'))
            created.append(index_name)
            existing_names.add(index_name)
        except Exception as exc:
            raise RuntimeError(
                f'auto index migration failed for {table_name}.{index_name}: ({columns_sql})'
            ) from exc
    return created


def _build_index_readiness_payload():
    items = []
    total = 0
    ready = 0
    for table_name, specs in SCHEMA_REQUIRED_INDEXES.items():
        existing = _existing_index_names(table_name)
        for index_name, columns in specs:
            total += 1
            ok = index_name in existing
            if ok:
                ready += 1
            items.append({
                'table': table_name,
                'index_name': index_name,
                'columns': columns,
                'ok': ok,
            })
    return {
        'summary': {
            'total': total,
            'ready': ready,
            'missing': total - ready,
            'ready_rate': round((ready / total) * 100, 2) if total else 100,
        },
        'items': items,
    }


def _build_capacity_readiness_payload():
    index_payload = _build_index_readiness_payload()
    counts = {
        'registrations': Registration.query.count(),
        'submissions': Submission.query.count(),
        'trend_notes': TrendNote.query.count(),
        'topic_ideas': TopicIdea.query.count(),
        'creator_accounts': CreatorAccount.query.count(),
        'creator_posts': CreatorPost.query.count(),
    }
    estimated_monthly_targets = {
        'people_per_month_target': 100,
        'posts_per_month_target': 2000,
    }
    risks = []
    if index_payload['summary']['missing'] > 0:
        risks.append('数据库关键索引未完全补齐，后续数据量上来会影响查询和同步效率。')
    if not _env_flag('INLINE_AUTOMATION_JOBS', False) and not bool((os.environ.get('CELERY_BROKER_URL') or '').strip()):
        risks.append('生产异步链路未完全配置，批量任务高峰期可能无法稳定消化。')
    return {
        'summary': {
            'capacity_ready': index_payload['summary']['missing'] == 0,
            'message': '当前数据量级对 100+ 人/月、2000 条/月没有明显压力，关键在于补齐异步服务和索引。'
        },
        'current_counts': counts,
        'target': estimated_monthly_targets,
        'index_readiness': index_payload['summary'],
        'risks': risks,
    }


def _run_worker_ping_check(timeout_seconds=3):
    from time import monotonic
    from celery.result import AsyncResult
    from celery.exceptions import TimeoutError as CeleryTimeoutError
    from celery_app import celery, ping

    safe_timeout = min(max(_safe_int(timeout_seconds, 3), 1), 15)
    started_at = datetime.now()
    started_tick = monotonic()
    task_id = ''
    async_result = None

    def _elapsed_ms():
        return int(round((monotonic() - started_tick) * 1000))

    try:
        task = _enqueue_task(ping)
        task_id = task.id
        async_result = task if _env_flag('INLINE_AUTOMATION_JOBS', False) else AsyncResult(task_id, app=celery)
        _log_operation('dispatch_job', 'worker', message='触发 Worker 联通检查', detail={
            'task_id': task_id,
            'job': 'system.ping',
            'timeout_seconds': safe_timeout,
        })

        payload = async_result.get(timeout=safe_timeout, propagate=False)
        state = async_result.state
        if not isinstance(payload, (dict, list, str, int, float, bool, type(None))):
            payload = str(payload)
        if state == 'SUCCESS':
            detail = {
                'status': 'success',
                'task_id': task_id,
                'state': state,
                'checked_at': _format_datetime(started_at),
                'elapsed_ms': _elapsed_ms(),
                'response': payload,
                'message': 'Worker 返回 pong',
            }
            _log_operation('worker_ping_check', 'worker', message='Worker 联通检查成功', detail=detail)
            return {
                'success': True,
                'message': f'Worker 联通正常，耗时 {detail["elapsed_ms"]}ms',
                'job': 'system.ping',
                'task_id': task_id,
                'state': state,
                'checked_at': detail['checked_at'],
                'elapsed_ms': detail['elapsed_ms'],
                'result': payload,
            }

        detail = {
            'status': 'failed',
            'task_id': task_id,
            'state': state,
            'checked_at': _format_datetime(started_at),
            'elapsed_ms': _elapsed_ms(),
            'response': payload,
            'message': f'Worker 返回状态 {state}',
        }
        _log_operation('worker_ping_check_failed', 'worker', message='Worker 联通检查返回失败状态', detail=detail)
        return {
            'success': False,
            'message': f'Worker 返回状态 {state}',
            'job': 'system.ping',
            'task_id': task_id,
            'state': state,
            'checked_at': detail['checked_at'],
            'elapsed_ms': detail['elapsed_ms'],
            'result': payload,
        }
    except CeleryTimeoutError:
        detail = {
            'status': 'timeout',
            'task_id': task_id,
            'state': async_result.state if async_result else 'PENDING',
            'checked_at': _format_datetime(started_at),
            'elapsed_ms': _elapsed_ms(),
            'message': f'等待 Worker 超时（>{safe_timeout}s）',
        }
        _log_operation('worker_ping_check_failed', 'worker', message='Worker 联通检查超时', detail=detail)
        return {
            'success': False,
            'message': detail['message'],
            'job': 'system.ping',
            'task_id': task_id,
            'state': detail['state'],
            'checked_at': detail['checked_at'],
            'elapsed_ms': detail['elapsed_ms'],
            'result': None,
        }
    except Exception as exc:
        status = 'dispatch_failed' if not task_id else 'failed'
        state = async_result.state if async_result else 'UNKNOWN'
        detail = {
            'status': status,
            'task_id': task_id,
            'state': state,
            'checked_at': _format_datetime(started_at),
            'elapsed_ms': _elapsed_ms(),
            'error': str(exc),
            'error_type': exc.__class__.__name__,
            'message': 'Broker / Worker 链路异常，无法完成联通检查',
        }
        _log_operation('worker_ping_check_failed', 'worker', message='Worker 联通检查失败', detail=detail)
        return {
            'success': False,
            'message': f'Worker 联通检查失败：{exc}',
            'job': 'system.ping',
            'task_id': task_id,
            'state': state,
            'checked_at': detail['checked_at'],
            'elapsed_ms': detail['elapsed_ms'],
            'result': None,
        }


def _next_schedule_time(interval_minutes, base_time=None):
    base = base_time or datetime.now()
    minutes = max(_safe_int(interval_minutes, 60), 1)
    return base + timedelta(minutes=minutes)


def _render_schema_column_sql(column_spec):
    if isinstance(column_spec, str):
        return column_spec
    if not isinstance(column_spec, dict):
        raise ValueError(f'unsupported schema column spec: {column_spec!r}')

    base_type = (column_spec.get('type') or '').strip()
    if not base_type:
        raise ValueError(f'missing column type in schema spec: {column_spec!r}')

    parts = [base_type]
    if 'default' in column_spec:
        default_value = column_spec.get('default')
        if isinstance(default_value, bool):
            parts.append(f"DEFAULT {'TRUE' if default_value else 'FALSE'}")
        elif isinstance(default_value, (int, float)):
            parts.append(f'DEFAULT {default_value}')
        elif default_value is None:
            parts.append('DEFAULT NULL')
        else:
            escaped = str(default_value).replace("'", "''")
            parts.append(f"DEFAULT '{escaped}'")
    if column_spec.get('nullable') is False:
        parts.append('NOT NULL')
    return ' '.join(parts)


def _should_inline_jobs():
    if _env_flag('INLINE_AUTOMATION_JOBS', False):
        return True
    return not bool((os.environ.get('CELERY_BROKER_URL') or '').strip())


def _enqueue_task(task, *args, **kwargs):
    if _should_inline_jobs():
        return task.apply(args=args, kwargs=kwargs)
    return task.delay(*args, **kwargs)


def _apply_creator_post_range(query, args):
    date_from = _parse_date(args.get('date_from')) if hasattr(args, 'get') else None
    date_to = _parse_date(args.get('date_to')) if hasattr(args, 'get') else None
    if date_from:
        start_dt = datetime.combine(date_from, datetime.min.time())
        query = query.filter(or_(CreatorPost.publish_time.is_(None), CreatorPost.publish_time >= start_dt))
    if date_to:
        end_dt = datetime.combine(date_to, datetime.max.time())
        query = query.filter(or_(CreatorPost.publish_time.is_(None), CreatorPost.publish_time <= end_dt))
    return query, date_from, date_to


def _current_month_date_range(now=None):
    current = now or datetime.now()
    start_date = current.replace(day=1).date()
    end_date = current.date()
    return start_date, end_date


def _build_creator_account_query(args):
    phone = (args.get('phone') or '').strip()
    platform = (args.get('platform') or '').strip()
    keyword = (args.get('keyword') or '').strip()
    viral_only = _coerce_bool(args.get('viral_only'))

    query = CreatorAccount.query
    if phone:
        query = query.filter(CreatorAccount.owner_phone.contains(phone))
    if platform:
        query = query.filter_by(platform=platform)
    if keyword:
        query = query.filter(or_(
            CreatorAccount.owner_name.contains(keyword),
            CreatorAccount.owner_phone.contains(keyword),
            CreatorAccount.account_handle.contains(keyword),
            CreatorAccount.display_name.contains(keyword),
        ))
    if viral_only:
        query = query.join(CreatorPost, CreatorPost.creator_account_id == CreatorAccount.id).filter(CreatorPost.is_viral.is_(True)).distinct()
    return query


def _build_creator_analytics_payload(posts, snapshots, date_from=None, date_to=None):
    daily_map = defaultdict(lambda: {
        'post_count': 0,
        'viral_posts': 0,
        'views': 0,
        'exposures': 0,
        'interactions': 0,
        'follower_delta': 0,
    })
    topic_map = defaultdict(lambda: {'post_count': 0, 'views': 0, 'interactions': 0})

    for post in posts:
        base_dt = post.publish_time or post.created_at or datetime.now()
        day_key = base_dt.strftime('%Y-%m-%d')
        interactions = (post.likes or 0) + (post.favorites or 0) + (post.comments or 0)
        row = daily_map[day_key]
        row['post_count'] += 1
        row['viral_posts'] += 1 if post.is_viral else 0
        row['views'] += post.views or 0
        row['exposures'] += post.exposures or 0
        row['interactions'] += interactions
        row['follower_delta'] += post.follower_delta or 0

        topic_key = (post.topic_title or '未关联话题').strip()
        topic_row = topic_map[topic_key]
        topic_row['post_count'] += 1
        topic_row['views'] += post.views or 0
        topic_row['interactions'] += interactions

    daily_rows = [{'date': day_key, **daily_map[day_key]} for day_key in sorted(daily_map.keys())]
    snapshot_rows = [{
        'date': snapshot.snapshot_date.isoformat() if snapshot.snapshot_date else '',
        'follower_count': snapshot.follower_count or 0,
        'post_count': snapshot.post_count or 0,
        'total_views': snapshot.total_views or 0,
        'total_interactions': snapshot.total_interactions or 0,
    } for snapshot in snapshots]

    top_topics = []
    for topic_name, item in topic_map.items():
        top_topics.append({'topic_title': topic_name, **item})
    top_topics.sort(key=lambda row: (row['interactions'], row['views'], row['post_count']), reverse=True)

    total_views = sum(post.views or 0 for post in posts)
    total_exposures = sum(post.exposures or 0 for post in posts)
    total_interactions = sum((post.likes or 0) + (post.favorites or 0) + (post.comments or 0) for post in posts)
    total_follower_delta = sum(post.follower_delta or 0 for post in posts)
    viral_posts = len([post for post in posts if post.is_viral])
    best_post = sorted(
        posts,
        key=lambda item: ((item.views or 0), ((item.likes or 0) + (item.favorites or 0) + (item.comments or 0)), (item.follower_delta or 0)),
        reverse=True
    )[0] if posts else None

    overview = {
        'post_count': len(posts),
        'viral_posts': viral_posts,
        'total_views': total_views,
        'total_exposures': total_exposures,
        'total_interactions': total_interactions,
        'total_follower_delta': total_follower_delta,
        'avg_views': round(total_views / len(posts), 2) if posts else 0,
        'avg_interactions': round(total_interactions / len(posts), 2) if posts else 0,
        'best_post': _serialize_creator_post(best_post) if best_post else None,
        'date_from': date_from.isoformat() if date_from else '',
        'date_to': date_to.isoformat() if date_to else '',
    }

    return {
        'overview': overview,
        'daily_posts': daily_rows,
        'daily_snapshots': snapshot_rows,
        'top_topics': top_topics[:10],
    }


@app.route('/healthz')
def healthz():
    try:
        db.session.execute(text('SELECT 1'))
        return jsonify({
            'success': True,
            'status': 'ok',
            'service': 'xhs_v4',
            'database': 'ok',
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
    except Exception as exc:
        return jsonify({
            'success': False,
            'status': 'degraded',
            'service': 'xhs_v4',
            'database': 'error',
            'message': str(exc),
        }), 503


def _resolve_analysis_range(args):
    today = datetime.now().date()
    range_key = (args.get('range') or 'all').strip()
    start_date = _parse_date(args.get('start_date'))
    end_date = _parse_date(args.get('end_date'))

    if range_key == 'this_week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
        label = '本周'
    elif range_key == 'last_week':
        this_week_start = today - timedelta(days=today.weekday())
        end_date = this_week_start - timedelta(days=1)
        start_date = end_date - timedelta(days=6)
        label = '上周'
    elif range_key == '7d':
        start_date = today - timedelta(days=6)
        end_date = today
        label = '近7天'
    elif range_key == '30d':
        start_date = today - timedelta(days=29)
        end_date = today
        label = '近30天'
    elif range_key == 'custom' and start_date and end_date:
        label = '自定义'
    else:
        range_key = 'all'
        start_date = None
        end_date = None
        label = '全部时间'

    start_dt = datetime.combine(start_date, datetime.min.time()) if start_date else None
    end_dt = datetime.combine(end_date, datetime.max.time()) if end_date else None
    return {
        'key': range_key,
        'label': label,
        'start_date': start_date.isoformat() if start_date else '',
        'end_date': end_date.isoformat() if end_date else '',
        'start_dt': start_dt,
        'end_dt': end_dt,
    }


def _is_in_range(value, start_dt=None, end_dt=None):
    if not start_dt and not end_dt:
        return True
    if not value:
        return False
    if start_dt and value < start_dt:
        return False
    if end_dt and value > end_dt:
        return False
    return True


def _submission_has_platform_link(submission, platform_key):
    return bool((getattr(submission, f'{platform_key}_link', '') or '').strip())


def _submission_has_any_link(submission):
    return any(_submission_has_platform_link(submission, key) for key, _ in PLATFORM_DEFINITIONS)


def _collect_platform_metrics(submission, platform_key):
    keys = [platform_key] if platform_key != 'all' else [key for key, _ in PLATFORM_DEFINITIONS]
    views = likes = favorites = comments = 0
    has_link = False
    for key in keys:
        if _submission_has_platform_link(submission, key):
            has_link = True
        views += getattr(submission, f'{key}_views', 0) or 0
        likes += getattr(submission, f'{key}_likes', 0) or 0
        favorites += getattr(submission, f'{key}_favorites', 0) or 0
        comments += getattr(submission, f'{key}_comments', 0) or 0
    return {
        'has_link': has_link,
        'views': views,
        'likes': likes,
        'favorites': favorites,
        'comments': comments,
        'interactions': likes + favorites + comments,
    }


def _calculate_rate(numerator, denominator):
    if not denominator:
        return None
    return round((numerator / denominator) * 100, 2)


def _format_rate(rate):
    return '-' if rate is None else f'{rate}%'


def _load_activity_scope(activity_id, range_info):
    topics = Topic.query.filter_by(activity_id=activity_id).all()
    topic_ids = [t.id for t in topics]
    if not topic_ids:
        return topics, [], []

    registrations_all = Registration.query.filter(Registration.topic_id.in_(topic_ids)).all()
    submissions_all = Submission.query.join(Registration).filter(
        Registration.topic_id.in_(topic_ids)
    ).all()

    if range_info['key'] == 'all':
        return topics, registrations_all, submissions_all

    participant_ids = set()
    for reg in registrations_all:
        if _is_in_range(reg.created_at, range_info['start_dt'], range_info['end_dt']):
            participant_ids.add(reg.id)

    for sub in submissions_all:
        active_time = sub.created_at or (sub.registration.created_at if sub.registration else None)
        if _is_in_range(active_time, range_info['start_dt'], range_info['end_dt']):
            participant_ids.add(sub.registration_id)

    registrations = [reg for reg in registrations_all if reg.id in participant_ids]
    submissions = []
    for sub in submissions_all:
        active_time = sub.created_at or (sub.registration.created_at if sub.registration else None)
        if sub.registration_id in participant_ids and _is_in_range(active_time, range_info['start_dt'], range_info['end_dt']):
            submissions.append(sub)

    return topics, registrations, submissions


def _build_group_rankings(registrations, submissions, platform_key='all'):
    reg_by_id = {reg.id: reg for reg in registrations}
    participant_counts = Counter((reg.group_num or '未分组') for reg in registrations)
    ranking = {}

    for group, count in participant_counts.items():
        ranking[group] = {
            'group': group,
            'participants': count,
            'published_count': 0,
            'views': 0,
            'likes': 0,
            'favorites': 0,
            'comments': 0,
            'interactions': 0,
        }

    for sub in submissions:
        metrics = _collect_platform_metrics(sub, platform_key)
        if not metrics['has_link']:
            continue
        reg = reg_by_id.get(sub.registration_id) or sub.registration
        group = (reg.group_num if reg else '') or '未分组'
        current = ranking.setdefault(group, {
            'group': group,
            'participants': participant_counts.get(group, 0),
            'published_count': 0,
            'views': 0,
            'likes': 0,
            'favorites': 0,
            'comments': 0,
            'interactions': 0,
        })
        current['published_count'] += 1
        current['views'] += metrics['views']
        current['likes'] += metrics['likes']
        current['favorites'] += metrics['favorites']
        current['comments'] += metrics['comments']
        current['interactions'] += metrics['interactions']

    rows = []
    for row in ranking.values():
        participants = row['participants']
        row['views_per_capita'] = round(row['views'] / participants, 2) if participants else 0
        row['interactions_per_capita'] = round(row['interactions'] / participants, 2) if participants else 0
        row['interaction_rate'] = _calculate_rate(row['interactions'], row['views'])
        row['interaction_rate_display'] = _format_rate(row['interaction_rate'])
        rows.append(row)

    rows.sort(key=lambda item: (item['interactions'], item['views'], item['published_count']), reverse=True)
    return rows


def _build_personal_rankings(registrations, submissions):
    reg_by_id = {reg.id: reg for reg in registrations}
    topic_map = {reg.id: (reg.topic.topic_name if reg.topic else '') for reg in registrations}

    def collect(keys, limit=20):
        rows = []
        for sub in submissions:
            metrics = {
                'has_link': any(_submission_has_platform_link(sub, key) for key in keys),
                'views': 0,
                'likes': 0,
                'favorites': 0,
                'comments': 0,
            }
            for key in keys:
                metrics['views'] += getattr(sub, f'{key}_views', 0) or 0
                metrics['likes'] += getattr(sub, f'{key}_likes', 0) or 0
                metrics['favorites'] += getattr(sub, f'{key}_favorites', 0) or 0
                metrics['comments'] += getattr(sub, f'{key}_comments', 0) or 0

            if not metrics['has_link']:
                continue

            reg = reg_by_id.get(sub.registration_id) or sub.registration
            rows.append({
                'registration_id': sub.registration_id,
                'name': reg.name if reg else '未命名',
                'group': (reg.group_num if reg else '') or '未分组',
                'topic': topic_map.get(sub.registration_id, ''),
                'views': metrics['views'],
                'likes': metrics['likes'],
                'favorites': metrics['favorites'],
                'comments': metrics['comments'],
                'interactions': metrics['likes'] + metrics['favorites'] + metrics['comments'],
            })

        rows.sort(key=lambda item: (item['interactions'], item['views']), reverse=True)
        return rows[:limit]

    return {
        'all_platform': collect(PRIMARY_PERSONAL_PLATFORMS),
        'xhs': collect(['xhs']),
        'douyin': collect(['douyin']),
        'video': collect(['video']),
    }


def _build_task_funnel_payload(registrations, submissions):
    total_tasks = len(registrations or [])
    strategy_ids = set()
    generated_ids = set()
    submitted_ids = set()
    synced_ids = set()

    for sub in submissions or []:
        reg_id = sub.registration_id
        if not reg_id:
            continue
        if any([
            (sub.strategy_payload or '').strip(),
            (sub.selected_copy_goal or '').strip(),
            (sub.selected_copy_skill or '').strip(),
            (sub.selected_title_skill or '').strip(),
            (sub.selected_image_skill or '').strip(),
        ]):
            strategy_ids.add(reg_id)
        if any([
            (sub.selected_copy_text or '').strip(),
            (sub.selected_title or '').strip(),
        ]):
            generated_ids.add(reg_id)
        if _submission_has_any_link(sub):
            submitted_ids.add(reg_id)
        if (
            (sub.xhs_link or '').strip() and (
                (sub.xhs_views or 0) > 0 or
                (sub.xhs_likes or 0) > 0 or
                (sub.xhs_favorites or 0) > 0 or
                (sub.xhs_comments or 0) > 0 or
                (sub.xhs_tracking_status or '').strip() == 'tracking'
            )
        ):
            synced_ids.add(reg_id)

    def build_step(key, label, count, note):
        rate = _calculate_rate(count, total_tasks) or 0
        return {
            'key': key,
            'label': label,
            'count': count,
            'rate': rate,
            'rate_display': _format_rate(rate),
            'note': note,
        }

    return {
        'total_tasks': total_tasks,
        'strategy_selected_count': len(strategy_ids),
        'generated_count': len(generated_ids),
        'submitted_count': len(submitted_ids),
        'synced_count': len(synced_ids),
        'steps': [
            build_step('strategy_selected', '已选策略', len(strategy_ids), '已经完成推荐组合和路线选择'),
            build_step('generated', '已生成内容', len(generated_ids), '已经确认至少一个标题或正文版本'),
            build_step('submitted', '已提交链接', len(submitted_ids), '已经提交至少一个平台链接'),
            build_step('synced', '已同步数据', len(synced_ids), '小红书已有互动数据或持续跟踪结果'),
        ],
    }


def _build_operational_advice_payload(
    *,
    publish_rate=0,
    best_content_type='',
    strategy_insights=None,
    task_funnel=None,
    group_completion=None,
    top_keyword_trends=None,
    platform_stats=None,
    note_improvement_suggestions=None,
    next_topic_suggestions=None,
):
    strategy_insights = strategy_insights or {}
    task_funnel = task_funnel or {}
    group_completion = list(group_completion or [])
    top_keyword_trends = list(top_keyword_trends or [])
    platform_stats = platform_stats or {}
    note_improvement_suggestions = list(note_improvement_suggestions or [])
    next_topic_suggestions = list(next_topic_suggestions or [])

    steps = task_funnel.get('steps') or []
    step_map = {str(item.get('key') or ''): item for item in steps if isinstance(item, dict)}
    title_leader = ((strategy_insights.get('title_skill_rows') or [None])[0]) or {}
    image_leader = ((strategy_insights.get('image_skill_rows') or [None])[0]) or {}
    combo_leader = ((strategy_insights.get('combo_rows') or [None])[0]) or {}

    urgent_actions = []
    if publish_rate < 60:
        urgent_actions.append('先补发布率，优先追“已生成但未提交链接”的任务。')
    if (step_map.get('strategy_selected') or {}).get('rate', 0) < 70:
        urgent_actions.append('策略确认率偏低，建议员工先用系统推荐组合，不要从空白自己配。')
    if (step_map.get('generated') or {}).get('rate', 0) < 60:
        urgent_actions.append('内容生成环节仍有卡点，建议把一键爆款生成作为默认入口。')
    if strategy_insights.get('capture_rate', 0) < 60:
        urgent_actions.append('策略留痕率不足，后续复盘会失真，必须统一从系统标题池和图片路线中选。')
    urgent_actions = urgent_actions[:4] or ['当前没有明显紧急阻塞，先保持节奏并扩大已验证打法。']

    winning_moves = []
    if best_content_type:
        winning_moves.append(f'当前内容类型里更值得放大的是“{best_content_type}”。')
    if title_leader.get('label'):
        winning_moves.append(f'标题打法优先复用“{title_leader["label"]}”，当前平均互动 {title_leader.get("avg_interactions") or 0}。')
    if image_leader.get('label'):
        winning_moves.append(f'图片打法优先复用“{image_leader["label"]}”，当前爆款率 {image_leader.get("viral_rate_display") or "-"}。')
    if combo_leader.get('label'):
        winning_moves.append(f'标题和图片的冠军组合是“{combo_leader["label"]}”，建议先让更多员工直接照这套执行。')
    winning_moves = winning_moves[:4] or ['当前样本还不够，先保证每条任务留痕，再开始放大冠军打法。']

    risk_alerts = []
    low_groups = [row for row in group_completion if (row.get('completion_rate') or 0) < 50]
    if low_groups:
        risk_alerts.append('低完成率小组：' + '、'.join([str(row.get('group') or '') for row in low_groups[:4]]) + '。')
    xhs_stats = platform_stats.get('xhs') or {}
    if xhs_stats.get('interaction_rate_display'):
        risk_alerts.append(f'小红书当前互动率 {xhs_stats.get("interaction_rate_display")} ，后续改动要重点盯标题和封面是否真的抬升互动。')
    if top_keyword_trends and not best_content_type:
        risk_alerts.append('热点词在涨，但内容类型尚未稳定，容易出现跟上热点却拿不到收藏。')
    risk_alerts.extend(note_improvement_suggestions[:2])
    deduped_risks = []
    for item in risk_alerts:
        text = (item or '').strip()
        if text and text not in deduped_risks:
            deduped_risks.append(text)
    risk_alerts = deduped_risks[:4] or ['当前没有新增风险，继续盯提交率和互动率变化。']

    next_week_actions = []
    for row in top_keyword_trends[:3]:
        next_week_actions.append(f'围绕“{row.get("keyword") or ""}”扩成 3-5 个角度，优先做高搜索和高收藏版本。')
    next_week_actions.extend(next_topic_suggestions[:2])
    deduped_next = []
    for item in next_week_actions:
        text = (item or '').strip()
        if text and text not in deduped_next:
            deduped_next.append(text)
    next_week_actions = deduped_next[:5] or ['下周先继续扩大现有冠军题型，同时稳定补充热点储备。']

    headline = urgent_actions[0]
    if publish_rate >= 60 and combo_leader.get('label'):
        headline = f'当前更适合扩大冠军组合“{combo_leader.get("label")}”，而不是继续分散试打法。'

    return {
        'headline': headline,
        'urgent_actions': urgent_actions,
        'winning_moves': winning_moves,
        'risk_alerts': risk_alerts,
        'next_week_actions': next_week_actions,
    }


def _trend_score(note):
    return (
        (note.likes or 0) +
        (note.favorites or 0) * 2 +
        (note.comments or 0) * 3 +
        min((note.views or 0) // 100, 200)
    )


def _extract_keywords_from_note(note):
    keywords = _split_keywords(note.keyword)
    search_text = f"{note.title or ''} {note.summary or ''}"
    for seed in LIVER_KEYWORD_SEEDS:
        if seed in search_text and seed not in keywords:
            keywords.append(seed)
    return keywords[:6]


def _detect_soft_insertion(text):
    text = text or ''
    if any(token in text for token in ['FibroScan', '福波看', '肝弹', '肝硬度', '体检', '检查', '转氨酶']):
        return 'FibroScan福波看'
    if any(token in text for token in ['脂肪肝', '减脂', '肥胖', '代谢']):
        return '壳脂胶囊治疗脂肪肝'
    return '复方鳖甲软肝片'


def _pick_asset_types(keyword):
    text = keyword or ''
    if any(token in text for token in ['FibroScan', '福波看', '肝弹', '体检', '转氨酶']):
        return ['医学科普图', '检查流程图', '知识卡片']
    if '脂肪肝' in text:
        return ['知识卡片', '误区对照图', '复查清单卡']
    return ['医学科普图', '知识卡片', '复查清单卡']


def _matching_corpus_snippets(keyword, limit=3):
    entries = CorpusEntry.query.filter_by(status='active').filter(CorpusEntry.pool_status != 'archived').order_by(CorpusEntry.updated_at.desc()).all()
    if not entries:
        return []

    hits = []
    for entry in entries:
        haystack = ' '.join([
            entry.title or '',
            getattr(entry, 'source_title', '') or '',
            getattr(entry, 'reference_url', '') or '',
            getattr(entry, 'template_type_key', '') or '',
            entry.tags or '',
            entry.content or '',
        ])
        score = 0
        for token in _split_keywords(keyword):
            if token and token in haystack:
                score += 2
        if entry.category in ['合规表达', '爆款拆解', '封面模板']:
            score += 1
        if getattr(entry, 'reference_url', None):
            score += 1
        if score > 0:
            hits.append((score, entry))

    hits.sort(key=lambda item: (item[0], item[1].updated_at or datetime.min), reverse=True)
    return [entry for _, entry in hits[:limit]]


def _build_topic_prompt(title, keyword, persona, content_type, insertion, corpus_entries):
    corpus_text = '\n'.join([f"- {entry.title}：{_truncate_text(entry.content, 80)}" for entry in corpus_entries]) or '- 暂无附加语料'
    return (
        f"请围绕《{title}》写一篇小红书文案。\n"
        f"人设：{persona}\n"
        f"内容类型：{content_type}\n"
        f"重点关键词：{keyword}\n"
        f"软植入方向：{insertion}\n"
        f"语料参考：\n{corpus_text}\n"
        f"合规要求：{COMPLIANCE_BASELINE}\n"
        "风格要求：真实口语化、去AI腔、首屏有钩子、正文200-350字、结尾自然提问。"
    )


def _build_generate_copy_corpus_block(corpus_entries, product_hint=''):
    if not corpus_entries:
        return '无'
    rows = []
    for entry in corpus_entries[:4]:
        template_meta = _corpus_template_meta(getattr(entry, 'template_type_key', '') or '')
        ref_link = getattr(entry, 'reference_url', '') or ''
        ref_text = _truncate_text(entry.content or '', 120)
        rows.append(
            f"- {entry.title or '未命名模板'} ｜ 模板={template_meta.get('label') or '标准说明'}"
            f"{f' ｜ 链接={ref_link}' if ref_link else ''}\n"
            f"  {ref_text}"
        )
    reminder = '只学这些模板的结构、标题逻辑、开头钩子、互动方式和版式感，不照抄原句；最后必须改写成围绕我们产品/服务的内容。'
    if product_hint:
        reminder = f'{reminder} 产品改写主线：{product_hint}'
    return '\n'.join(rows + [reminder]).strip()


def _local_copy_scene_line(scene_text='', lead_keyword=''):
    text = (scene_text or '').strip()
    keyword = lead_keyword or '这件事'
    if any(token in keyword for token in ['解酒', '护肝', '熬夜', '应酬']):
        return f'以前每次熬夜应酬后，我第一反应都是找“{keyword}”，后来才发现最该先补的是判断顺序。'
    if any(token in text for token in ['体检', '异常', '提醒']):
        return f'那次其实也没什么明显不舒服，就是体检单一出来，我对“{keyword}”一下子紧张起来了。'
    if any(token in text for token in ['报告', '解读', '指标', '复查']):
        return f'真正让我难受的不是看到“{keyword}”这几个字，而是拿着报告一时不知道先看什么。'
    if any(token in text for token in ['家属', '陪伴', '照护']):
        return f'陪着家里人处理“{keyword}”这件事时，我才发现很多焦虑都不是病本身，而是顺序没理清。'
    if any(token in text for token in ['饮食', '减脂', '运动']):
        return f'我后来才意识到，“{keyword}”不是一句少吃多动就能带过，真要落地其实很看日常习惯。'
    if any(token in text for token in ['中医', '情绪', '失眠', '压力']):
        return f'那段时间我自己最明显的感觉不是疼，而是整个人一直绷着，连“{keyword}”都越想越慌。'
    return f'我以前对“{keyword}”的理解挺模糊的，真正碰到之后才知道，这事不能只靠想象。'


def _local_copy_persona_line(persona_text='', lead_keyword=''):
    text = (persona_text or '').strip()
    keyword = lead_keyword or '这件事'
    if '患者本人' in text:
        return f'我自己就是那个会把“{keyword}”先往后拖的人，所以这次被提醒时，反而一下子清醒了。'
    if any(token in text for token in ['家属', '陪诊']):
        return f'我是站在家属/陪诊者的角度看“{keyword}”，很多时候不是不重视，而是真的不知道该怎么接下一步。'
    if any(token in text for token in ['医学', '医生助理']):
        return f'从科普角度看，“{keyword}”最怕的不是复杂，而是大家只记住名词，却没记住该怎么判断。'
    if any(token in text for token in ['营养', '健管']):
        return f'从日常管理角度看，“{keyword}”很多问题都不是一顿饭造成的，但确实和长期习惯有关。'
    if any(token in text for token in ['中医']):
        return f'从调理视角看，“{keyword}”常常不是一个孤立点，它和情绪、作息、节律会连在一起。'
    return f'我更想用普通人能听懂的话把“{keyword}”说清楚，而不是只丢一堆概念。'


def _local_copy_action_lines(lead_keyword='', scene_text=''):
    keyword = lead_keyword or '这件事'
    scene = (scene_text or '').strip()
    if any(token in keyword for token in ['FibroScan', '福波看', '肝弹', '检查', '体检', '转氨酶']):
        return [
            '我后来先做的一件事，就是把这次结果和以前的检查放在一起看，不然很容易只看单次数字把自己吓到。',
            '如果医生提到要继续检查，我现在会顺手问清楚：这次到底是想看硬度、看脂肪，还是看变化趋势。',
            '比起到处搜答案，我更建议先把复查时间、指标变化和医生给的重点记下来，这样心里会稳很多。',
        ]
    if any(token in keyword for token in ['脂肪肝', '减肥', '肥胖', '代谢']):
        return [
            '我后来不再只盯体重，而是先把饭点、活动量和复查节奏定下来，这样反而更容易坚持。',
            '很多人一着急就开始极端节食，但我自己的体感是，先把能长期做下去的动作留住更重要。',
            '真正有用的不是今天多狠，而是接下来一个月你能不能把吃饭、运动和复查都接上。',
        ]
    if any(token in keyword for token in ['解酒', '护肝', '熬夜', '应酬']):
        return [
            '我现在会先区分：是偶尔一次应酬后的恢复，还是长期作息和饮酒习惯已经在拖后腿。',
            '别把“护肝”写成万能方法，正文更适合讲清楚少喝酒、补睡眠、看指标和必要时咨询医生这几步。',
            '这类笔记最容易爆的点不是夸一个方法神，而是把“哪些别做、哪些先做”讲得很具体。',
        ]
    if any(token in scene for token in ['家属', '陪伴', '照护']):
        return [
            '我后来会先把医生说的重点翻成家里人听得懂的话，不然大家各自脑补，反而更乱。',
            '很多时候不是不配合，而是不知道先做哪一步，所以把顺序讲清楚比说大道理更有用。',
            '一旦把检查、复查和生活管理拆开来说，家里人的抵触感会小很多。',
        ]
    return [
        f'我后来慢慢发现，处理“{keyword}”最有用的不是吓自己，而是把下一步动作先落下来。',
        '先搞清楚自己现在卡在哪，再决定是去复查、去调整生活习惯，还是继续观察，这样会踏实很多。',
        '把节奏理清之后，整个人反而没那么慌，也更容易把这件事真的做下去。',
    ]


def _local_copy_product_line(product_label=''):
    label = (product_label or '').strip()
    if not label or label in {'自动匹配', '不植入产品'}:
        return '这篇我会先把问题讲明白，不强行塞产品，读起来才更像真实分享。'
    if label == 'FibroScan福波看':
        return '像 FibroScan 福波看这类检查，我现在更愿意放在真实经历里顺手讲清楚，而不是单独拎出来硬推。'
    if label in {'复方鳖甲软肝片', '恩替卡韦联合管理', '壳脂胶囊'}:
        return f'如果要带到 {label}，我更倾向放在真实管理路径里轻轻带一下，而不是一上来就像推荐单品。'
    return f'如果要提到 {label}，我也会更偏向真实经验里的自然带出，不会写成硬广。'


def _local_variant_middle(index, lead_keyword=''):
    keyword = lead_keyword or '这件事'
    variants = [
        f'真正点醒我的，不是“{keyword}”这个词本身，而是我发现自己之前一直把重点放错了。',
        f'后来我才明白，“{keyword}”最怕的不是复杂，而是自己脑补太多、真正该确认的事反而没问。',
        f'如果重新来一次，我不会再对“{keyword}”硬扛着等它自己过去，而是先把顺序理出来。',
    ]
    return variants[index % len(variants)]


def _build_local_copy_body_sections(
    route_key='',
    *,
    lead_keyword='',
    scene_text='',
    persona_text='',
    product_label='',
    prompt_focus='',
    route_body_strategy='',
    reference_hint='',
    index=0,
):
    keyword = lead_keyword or '这件事'
    scene_line = _local_copy_scene_line(scene_text, keyword)
    persona_line = _local_copy_persona_line(persona_text, keyword)
    action_lines = _local_copy_action_lines(keyword, scene_text)
    primary_action = action_lines[index % len(action_lines)]
    secondary_action = action_lines[(index + 1) % len(action_lines)]
    product_line = _local_copy_product_line(product_label)
    focus_line = f'如果只想先抓一个重点，我会先把“{prompt_focus}”这一层讲明白。' if prompt_focus else ''
    route_key = (route_key or '').strip()
    sections = []
    if any(token in keyword for token in ['解酒', '护肝', '熬夜', '应酬']):
        variants = [
            [
                '我以前也会搜各种解酒护肝小方法，后来才发现，最容易误导人的就是把“舒服一点”和“真的护肝”混在一起。',
                '偶尔一次应酬后，我现在会先补水、早点睡、别继续熬；如果是长期喝酒或者指标异常，就不会再指望一个小妙招解决。',
                '真正该做的是把频率降下来，把作息补回来，再看需不需要查肝功能或问医生。',
                '所以这类内容我不会写成万能方法，只会把“哪些别做、哪些先做”讲清楚。',
            ],
            [
                '有段时间我应酬后第二天醒来，第一反应就是找“护肝小妙招”，好像做点什么就能把前一天抵消掉。',
                '后来我才反应过来，身体不是靠临时补救管理的，真正该看的反而是最近喝酒频率、睡眠和体检指标。',
                '如果已经经常不舒服，或者报告里有异常，我会更愿意先去问清楚原因，而不是继续自己试方法。',
                '这也是我现在写这类内容的底线：讲真实恢复顺序，不把偏方包装成答案。',
            ],
            [
                '应酬后我现在会先记住三件事：别继续熬、别乱叠加各种偏方、别把一次不舒服当成可以忽略的小事。',
                '能做的顺序其实很简单：先休息补水，观察身体反应，下一次把饮酒量降下来。',
                '如果最近总是乏、胀、指标不稳，就该把肝功能或相关检查安排上，而不是继续靠经验猜。',
                '这份清单适合直接收藏，真正用的时候按顺序看，比临时到处搜更稳。',
            ],
        ]
        sections = variants[index % len(variants)]
    elif route_key == 'report_emotion':
        sections = [
            scene_line,
            '那一下最容易把人带偏的，不是报告本身，而是看到一个名词就先往严重了想。',
            '我现在会先提醒自己：先看这次检查到底想回答什么，再看前后变化，最后才去查别的。',
            primary_action,
            secondary_action,
        ]
    elif route_key == 'report_decode':
        sections = [
            persona_line,
            f'像“{keyword}”这种检查项，真正要先讲清楚的不是名字，而是它在帮我们确认哪件事。',
            '我现在更习惯先拆“最容易看偏的点”，这样大家一眼就知道该从哪一步开始判断。',
            primary_action,
            '先把最容易看偏的地方说清楚，后面的复查和观察顺序才不容易乱。',
        ]
    elif route_key == 'report_checklist':
        sections = [
            '我后来给自己定的顺序很短：先看前后变化，再问清检查目的，最后定复查时间。',
            '这样处理的好处是，人不会一直慌，和医生沟通时也更容易抓到重点。',
            primary_action,
            secondary_action,
            '把顺序列清楚以后，至少不会看完一堆名词还是不知道下一步做什么。',
        ]
    elif route_key == 'myth_reverse':
        sections = [
            persona_line,
            f'关于“{keyword}”，很多人第一反应都太快了，偏偏最容易偏在第一步判断。',
            '我以前也信过那个最常见的说法，后来才发现，问题不在“懂不懂”，而在判断顺序错了。',
            primary_action,
            '先把最常见的误会拆开，后面的判断反而会简单很多。',
        ]
    elif route_key == 'myth_case':
        sections = [
            scene_line,
            '那次我其实差一点就按自己以前那套理解走下去了，后来才发现真正危险的是“想当然”。',
            primary_action,
            secondary_action,
            '把那次差点踩坑的瞬间讲清楚，比空讲大道理更容易让人记住。',
        ]
    elif route_key == 'myth_checklist':
        sections = [
            f'如果让我重来一次，我会先把关于“{keyword}”最容易搞反的 3 件事写下来。',
            '因为这种内容最怕的不是看不懂，而是半懂不懂地自己下结论。',
            primary_action,
            '把误区和正解摆在一起看，心里会稳很多，也不容易自己吓自己。',
        ]
    elif route_key == 'story_first':
        sections = [
            scene_line,
            '我以前最容易做的，就是觉得这事可以再等等，结果越拖越不敢面对。',
            persona_line,
            primary_action,
            '真正有用的不是把情绪写满，而是把后来怎么改、怎么接住这件事说清楚。',
        ]
    elif route_key == 'qa_first':
        sections = [
            f'每次聊到“{keyword}”，大家最常问的其实就两件事：现在要不要紧，下一步先做什么。',
            '我自己的习惯是先把顺序理清，再决定是复查、观察，还是先把生活管理接上。',
            primary_action,
            secondary_action,
            '先把问题说成人话，后面的动作反而更容易接住。',
        ]
    else:
        sections = [
            persona_line,
            f'把“{keyword}”讲清楚，关键不是堆术语，而是先告诉大家最容易看偏的点。',
            scene_line,
            primary_action,
            '先把判断讲清楚，后面的动作才接得住，不会看完还是一头雾水。',
        ]

    if focus_line:
        sections.append(focus_line)
    if product_line and product_label and product_label not in {'自动匹配', '不植入产品'}:
        sections.append(product_line)
    closing_lines = [
        '我会把这篇写成普通人能直接照着判断的版本，不靠一堆名词撑专业感。',
        '这类内容最重要的不是把话说满，而是让人看完知道下一步先做什么。',
        '我会少讲口号，多写具体场景和判断顺序，这样更像真实笔记。',
    ]
    sections.append(closing_lines[index % len(closing_lines)])

    compact = []
    seen = set()
    for row in sections:
        text = (row or '').strip()
        normalized_text = text.replace('先说结论：', '').replace('说实话：', '').replace('真实经历：', '').strip()
        if not text or text in seen:
            continue
        if normalized_text in seen:
            continue
        seen.add(text)
        seen.add(normalized_text)
        compact.append(text)
        if len(compact) >= 6:
            break
    return compact


def _generate_topic_ideas(count=80, activity_id=None, quota=None):
    recent_notes = TrendNote.query.filter(TrendNote.pool_status != 'archived').order_by(TrendNote.created_at.desc()).limit(300).all()
    topic_quota = _normalize_quota(quota)
    keyword_scores = defaultdict(int)
    keyword_notes = defaultdict(list)

    for note in recent_notes:
        extracted = _extract_keywords_from_note(note)
        if not extracted:
            continue
        score = max(_trend_score(note), 1)
        for keyword in extracted:
            keyword_scores[keyword] += score
            if len(keyword_notes[keyword]) < 5:
                keyword_notes[keyword].append(note)

    for idx, seed in enumerate(LIVER_KEYWORD_SEEDS, start=1):
        keyword_scores[seed] += max(10 - idx // 3, 1)

    sorted_keywords = sorted(keyword_scores.items(), key=lambda item: item[1], reverse=True)
    if not sorted_keywords:
        sorted_keywords = [(seed, len(LIVER_KEYWORD_SEEDS) - idx) for idx, seed in enumerate(LIVER_KEYWORD_SEEDS)]

    angle_templates = [
        '体检发现{keyword}异常，下一步到底先做什么？',
        '{keyword}拖着不管的人，后来最容易后悔哪一步？',
        '关于{keyword}，门诊里最常被问到的3个问题',
        '{keyword}到底要不要复查？什么时间点更关键？',
        '{keyword}相关内容怎么写，既有信任感又不吓人？',
        '{keyword}最容易被误解的地方，很多人第一步就走偏了',
        '{keyword}人群日常管理，哪些动作是高频踩坑点？',
        '{keyword}做成知识卡片时，哪3个信息点最容易被收藏？',
        '{keyword}这类爆款笔记，为什么大家愿意看完并互动？',
        '{keyword}相关复查和管理，真实用户最关心什么？',
    ]

    created = []
    seen_titles = {idea.topic_title for idea in TopicIdea.query.all()}
    max_attempts = max(count * 6, 200)
    attempt = 0

    while len(created) < count and attempt < max_attempts:
        keyword, score = sorted_keywords[attempt % len(sorted_keywords)]
        persona = TOPIC_PERSONAS[attempt % len(TOPIC_PERSONAS)]
        content_type = TOPIC_CONTENT_TYPES[(attempt + len(keyword)) % len(TOPIC_CONTENT_TYPES)]
        angle_template = angle_templates[(attempt + score) % len(angle_templates)]
        title = angle_template.format(keyword=keyword)
        insertion = _detect_soft_insertion(keyword + title)
        asset_types = _pick_asset_types(keyword)
        note_refs = keyword_notes.get(keyword, [])[:3]
        ref_links = [note.link for note in note_refs if note.link]
        ref_ids = [str(note.id) for note in note_refs]
        corpus_entries = _matching_corpus_snippets(keyword)
        source_keywords = [keyword]
        if note_refs:
            source_keywords.extend(_extract_keywords_from_note(note_refs[0]))
        cover_title = _truncate_text(f'{keyword}高频问题', 18)
        angle = (
            f"以{persona}切入，用{content_type}方式拆解“{keyword}”热点话题，"
            f"结合近期热门内容常见的焦虑点、误区和下一步动作，"
            f"自然带出{insertion}，但表达必须符合医疗广告合规。"
        )
        asset_brief = (
            f"建议做{asset_types[0]}、{asset_types[1]}、{asset_types[2]}三套素材；"
            f"主封面标题用“{cover_title}”，画面突出问题句、数字和复查提醒。"
        )

        if title in seen_titles:
            attempt += 1
            continue

        idea = TopicIdea(
            activity_id=activity_id,
            topic_title=title,
            keywords=','.join(dict.fromkeys(source_keywords))[:500],
            angle=angle,
            content_type=content_type,
            persona=persona,
            soft_insertion=insertion,
            hot_value=min(score, 9999),
            source_note_ids=','.join(ref_ids),
            source_links='\n'.join(ref_links),
            copy_prompt=_build_topic_prompt(title, keyword, persona, content_type, insertion, corpus_entries),
            cover_title=cover_title,
            asset_brief=asset_brief,
            compliance_note=COMPLIANCE_BASELINE,
            quota=topic_quota,
            status='pending_review'
        )
        for note in note_refs:
            if note.pool_status != 'archived':
                note.pool_status = 'candidate'
        created.append(idea)
        seen_titles.add(title)
        attempt += 1

    return created


def _wrap_svg_text(text, line_length=11, max_lines=3):
    text = re.sub(r'\s+', '', (text or '').strip())
    if not text:
        return []
    lines = [text[i:i + line_length] for i in range(0, len(text), line_length)]
    return lines[:max_lines]


def _svg_data_uri(svg_text):
    encoded = base64.b64encode(svg_text.encode('utf-8')).decode('ascii')
    return f'data:image/svg+xml;base64,{encoded}'


def _render_svg_card(card_type, title, subtitle, bullets, accent='#ff6b57', bg='#fff7f3'):
    title_lines = _wrap_svg_text(title, line_length=10, max_lines=3)
    subtitle_lines = _wrap_svg_text(subtitle, line_length=16, max_lines=2)
    safe_bullets = bullets[:3]

    title_svg = ''.join(
        f'<text x="48" y="{118 + idx * 44}" font-size="34" font-weight="700" fill="#242424">{html.escape(line)}</text>'
        for idx, line in enumerate(title_lines)
    )
    subtitle_svg = ''.join(
        f'<text x="48" y="{255 + idx * 26}" font-size="20" fill="#4c4c4c">{html.escape(line)}</text>'
        for idx, line in enumerate(subtitle_lines)
    )

    bullet_svg = ''
    base_y = 355
    for idx, bullet in enumerate(safe_bullets):
        y = base_y + idx * 82
        bullet_lines = _wrap_svg_text(bullet, line_length=18, max_lines=2)
        bullet_svg += f'<rect x="44" y="{y - 24}" rx="24" ry="24" width="392" height="58" fill="white" opacity="0.95" />'
        bullet_svg += f'<circle cx="74" cy="{y + 5}" r="14" fill="{accent}" />'
        bullet_svg += f'<text x="69" y="{y + 11}" font-size="16" font-weight="700" fill="white">{idx + 1}</text>'
        for line_idx, line in enumerate(bullet_lines):
            bullet_svg += (
                f'<text x="98" y="{y + line_idx * 22}" font-size="19" fill="#2c2c2c">'
                f'{html.escape(line)}</text>'
            )

    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="1080" height="1440" viewBox="0 0 480 640">
<defs>
<linearGradient id="g" x1="0%" y1="0%" x2="100%" y2="100%">
<stop offset="0%" stop-color="{bg}" />
<stop offset="100%" stop-color="#ffffff" />
</linearGradient>
</defs>
<rect width="480" height="640" rx="36" fill="url(#g)" />
<circle cx="400" cy="70" r="88" fill="{accent}" opacity="0.15" />
<circle cx="100" cy="600" r="72" fill="{accent}" opacity="0.10" />
{title_svg}
{subtitle_svg}
<rect x="42" y="302" rx="28" ry="28" width="396" height="286" fill="{accent}" opacity="0.12" />
{bullet_svg}
</svg>'''


def _render_svg_poster_card(style_key, title, subtitle, bullets, accent='#18e05e', bg='#fffdfd'):
    title_lines = _wrap_svg_text(title, line_length=7, max_lines=4)
    title_svg = ''
    base_y = 180 if len(title_lines) <= 3 else 150
    for idx, line in enumerate(title_lines):
        y = base_y + idx * 86
        highlight = idx == 0 or (style_key == 'poster_handwritten' and idx == 1)
        if highlight:
            width = min(360, 44 + len(line) * 48)
            title_svg += f'<rect x="60" y="{y - 44}" rx="24" ry="24" width="{width}" height="52" fill="{accent}" opacity="0.75" />'
        font_family = '"PingFang SC","Heiti SC","Microsoft YaHei",sans-serif' if style_key == 'poster_bold' else '"Kaiti SC","STKaiti","KaiTi",cursive'
        title_svg += f'<text x="60" y="{y}" font-size="54" font-weight="800" font-family={font_family} fill="#202020">{html.escape(line)}</text>'

    subtitle_svg = ''
    if subtitle:
        subtitle_lines = _wrap_svg_text(subtitle, line_length=12, max_lines=2)
        for idx, line in enumerate(subtitle_lines):
            subtitle_svg += f'<text x="66" y="{base_y + len(title_lines) * 86 + 25 + idx * 26}" font-size="22" fill="#555">{html.escape(line)}</text>'

    bullet_svg = ''
    bullet_y = base_y + len(title_lines) * 86 + 80
    for idx, bullet in enumerate((bullets or [])[:3]):
        bullet_svg += f'<text x="70" y="{bullet_y + idx * 34}" font-size="21" fill="#444">{idx + 1}. {html.escape(bullet[:24])}</text>'

    emoji_svg = ''
    if style_key == 'poster_handwritten':
        emoji_svg = '<text x="360" y="570" font-size="48">😵</text><text x="410" y="570" font-size="48">😭</text>'

    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="1080" height="1440" viewBox="0 0 480 640">
<rect width="480" height="640" fill="{bg}" />
<g opacity="0.16">
  <line x1="40" y1="80" x2="440" y2="80" stroke="#d9d9d9" stroke-width="1" />
  <line x1="40" y1="120" x2="440" y2="120" stroke="#d9d9d9" stroke-width="1" />
  <line x1="40" y1="160" x2="440" y2="160" stroke="#d9d9d9" stroke-width="1" />
</g>
{title_svg}
{subtitle_svg}
{bullet_svg}
{emoji_svg}
</svg>'''


def _render_svg_memo_card(style_key, title, subtitle, bullets, accent='#f4dc62', bg='#fffefe'):
    title_lines = _wrap_svg_text(title, line_length=10, max_lines=3)
    ui_header = '''
<text x="28" y="42" font-size="20" fill="#d6a800">〈 备忘录</text>
<text x="360" y="42" font-size="18" fill="#d6a800">↶</text>
<text x="400" y="42" font-size="18" fill="#d6a800">↷</text>
<text x="440" y="42" font-size="18" fill="#d6a800">⋯</text>
'''
    title_svg = ''.join(
        f'<text x="42" y="{104 + idx * 46}" font-size="38" font-weight="800" fill="#1f1f1f">{html.escape(line)}</text>'
        for idx, line in enumerate(title_lines)
    )
    subtitle_svg = f'<text x="44" y="{126 + len(title_lines) * 46}" font-size="21" fill="#666">{html.escape(subtitle[:36])}</text>' if subtitle else ''

    content_y = 190 + len(title_lines) * 40
    note_lines = ''
    if style_key == 'memo_mobile':
        for idx, bullet in enumerate((bullets or [])[:5], start=1):
            highlight = idx in {1, 3}
            y = content_y + idx * 54
            if highlight:
                note_lines += f'<rect x="72" y="{y - 18}" rx="12" ry="12" width="250" height="26" fill="{accent}" opacity="0.55" />'
            note_lines += f'<rect x="40" y="{y - 23}" rx="10" ry="10" width="30" height="30" fill="#5b84d6" opacity="0.9" />'
            note_lines += f'<text x="50" y="{y - 2}" font-size="16" font-weight="700" fill="white">{idx}</text>'
            note_lines += f'<text x="82" y="{y}" font-size="24" fill="#2b2b2b">{html.escape(bullet[:26])}</text>'
    else:
        for idx, bullet in enumerate((bullets or [])[:5], start=1):
            y = content_y + idx * 60
            note_lines += f'<rect x="36" y="{y - 22}" rx="10" ry="10" width="120" height="24" fill="#eef6c8" />'
            note_lines += f'<text x="44" y="{y - 4}" font-size="18" font-weight="700" fill="#49631f">重点{idx}</text>'
            note_lines += f'<text x="42" y="{y + 24}" font-size="23" fill="#2b2b2b">{html.escape(bullet[:26])}</text>'
            note_lines += f'<line x1="42" y1="{y + 30}" x2="{42 + min(260, len(bullet) * 18)}" y2="{y + 30}" stroke="#b2d66b" stroke-width="4" opacity="0.7" />'

    paper_lines = ''.join(
        f'<line x1="36" y1="{160 + idx * 56}" x2="444" y2="{160 + idx * 56}" stroke="#ececec" stroke-width="1" />'
        for idx in range(8)
    )
    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="1080" height="1440" viewBox="0 0 480 640">
<rect width="480" height="640" fill="{bg}" />
{paper_lines}
{ui_header}
{title_svg}
{subtitle_svg}
{note_lines}
</svg>'''


def _render_svg_checklist_card(style_key, title, subtitle, bullets, accent='#f1c24b', bg='#fff9ea'):
    title_lines = _wrap_svg_text(title, line_length=9, max_lines=2)
    title_svg = ''.join(
        f'<text x="42" y="{84 + idx * 42}" font-size="40" font-weight="800" fill="#242424">{html.escape(line)}</text>'
        for idx, line in enumerate(title_lines)
    )
    subtitle_svg = f'<text x="44" y="{140 + len(title_lines) * 36}" font-size="20" fill="#666">{html.escape(subtitle[:32])}</text>' if subtitle else ''
    body_svg = ''
    if style_key == 'checklist_table':
        cols = ['项目', '图片', '指标', '结论']
        x_positions = [40, 130, 250, 360]
        for idx, col in enumerate(cols):
            body_svg += f'<rect x="{x_positions[idx]}" y="190" width="85" height="42" fill="{accent}" opacity="0.35" stroke="#c9b06b" />'
            body_svg += f'<text x="{x_positions[idx] + 18}" y="217" font-size="18" font-weight="700" fill="#47391d">{col}</text>'
        for row in range(3):
            y = 232 + row * 92
            bullet_text = ((bullets or ['']) + [''] * 3)[row][:8]
            for x in x_positions:
                body_svg += f'<rect x="{x}" y="{y}" width="85" height="92" fill="white" stroke="#d9cfb3" />'
            body_svg += f'<text x="56" y="{y + 38}" font-size="20" fill="#333">{row + 1}</text>'
            body_svg += f'<text x="140" y="{y + 50}" font-size="18" fill="#777">图片</text>'
            body_svg += f'<text x="262" y="{y + 50}" font-size="18" fill="#333">{html.escape(bullet_text)}</text>'
            body_svg += f'<text x="382" y="{y + 50}" font-size="18" fill="#2d8f5a">推荐</text>'
    elif style_key == 'checklist_timeline':
        days = ['Day1', 'Day2', 'Day3']
        for idx, day in enumerate(days):
            body_svg += f'<rect x="{44 + idx * 126}" y="188" rx="18" ry="18" width="112" height="36" fill="{accent if idx == 0 else "#f1f1f1"}" opacity="0.8" />'
            body_svg += f'<text x="{68 + idx * 126}" y="212" font-size="20" font-weight="700" fill="#333">{day}</text>'
        slots = [('早餐', '7:00-8:00'), ('午餐', '12:00-13:00'), ('晚餐', '18:00-20:00')]
        for idx, (slot, tm) in enumerate(slots):
            y = 250 + idx * 112
            body_svg += f'<rect x="40" y="{y}" rx="26" ry="26" width="400" height="92" fill="white" stroke="#e8d9d0" stroke-width="2" />'
            body_svg += f'<text x="56" y="{y + 36}" font-size="28" font-weight="800" fill="#222">{slot}</text>'
            body_svg += f'<rect x="56" y="{y + 48}" rx="10" ry="10" width="110" height="26" fill="#f4f4f4" />'
            body_svg += f'<text x="68" y="{y + 67}" font-size="16" fill="#666">{tm}</text>'
            body_svg += f'<text x="190" y="{y + 42}" font-size="26" fill="#333">{html.escape((bullets or [""])[idx % max(1, len(bullets))][:10] or "主食")}</text>'
            body_svg += f'<text x="305" y="{y + 42}" font-size="30" fill="#999">+</text>'
            body_svg += f'<text x="340" y="{y + 42}" font-size="22" fill="#333">搭配项</text>'
    else:
        for idx, bullet in enumerate((bullets or [])[:4], start=1):
            y = 210 + (idx - 1) * 98
            body_svg += f'<rect x="40" y="{y}" rx="22" ry="22" width="400" height="78" fill="white" stroke="#d7e5ca" stroke-width="2" />'
            body_svg += f'<rect x="54" y="{y + 18}" rx="12" ry="12" width="86" height="24" fill="#e7f2d1" />'
            body_svg += f'<text x="66" y="{y + 35}" font-size="18" font-weight="700" fill="#6a8d37">项目 {idx}</text>'
            body_svg += f'<text x="54" y="{y + 62}" font-size="21" fill="#333">{html.escape(bullet[:24])}</text>'
            body_svg += f'<circle cx="410" cy="{y + 38}" r="14" fill="#f6c24a" opacity="0.75" />'
    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="1080" height="1440" viewBox="0 0 480 640">
<rect width="480" height="640" fill="{bg}" />
<g opacity="0.22">
  <path d="M32 0 L32 640" stroke="#cfd6dc" stroke-width="2" />
  <path d="M0 48 L480 48" stroke="#dce4ea" stroke-width="1" />
  <path d="M0 88 L480 88" stroke="#dce4ea" stroke-width="1" />
</g>
{title_svg}
{subtitle_svg}
{body_svg}
</svg>'''


def _extract_content_points(content):
    text = re.sub(
        r'(标题|开头钩子|钩子|正文|内文|互动结尾|结尾互动|人设|角色|场景|软植入|软植入产品|图片工作流模式|图片主类型|封面样式|内页样式|图片技能包|图片打法|图片提示|自定义图片提示词)\s*[：:]',
        ' ',
        content or '',
    )
    parts = [part.strip() for part in re.split(r'[。！？\n]+', text) if part and part.strip()]
    points = []
    for part in parts:
        if any(token in part for token in ['患者本人', '医生助理', '健管师', '营养师', '系统会先拆封面', '当前模式', '封面主打']):
            continue
        if 6 <= len(part) <= 32:
            points.append(part)
        if len(points) >= 3:
            break
    return points


def _extract_visual_focus_payload(selected_content='', title_hint='', primary_keyword=''):
    card = _parse_generated_copy_card(selected_content or '')
    clean_title = (title_hint or '').strip() or (card.get('title') or '').strip() or primary_keyword or '核心主题'
    hook_text = re.sub(r'[。！？；;，,\s]+$', '', (card.get('hook') or '').strip())
    body_text = (card.get('body') or '').strip()
    support_points = _extract_content_points(selected_content or body_text)
    if hook_text and hook_text not in support_points:
        support_points = [hook_text] + support_points
    cover_headline = clean_title[:22]
    cover_subheadline = ''
    if hook_text:
        cover_subheadline = hook_text[:28]
    elif support_points:
        cover_subheadline = support_points[0][:28]
    body_excerpt = '；'.join([item for item in support_points[:3] if item.strip()]) or f'围绕{primary_keyword or clean_title}做清晰表达'
    return {
        'clean_title': clean_title,
        'hook_text': hook_text,
        'body_text': body_text,
        'support_points': support_points[:4],
        'cover_headline': cover_headline,
        'cover_subheadline': cover_subheadline,
        'body_excerpt': body_excerpt,
    }


def _build_image_workflow_prompt_notes(
    generation_mode='smart_bundle',
    *,
    cover_style_meta=None,
    inner_style_meta=None,
    visual_focus=None,
    image_count=1,
    workflow_role='',
):
    cover_style_meta = cover_style_meta or {}
    inner_style_meta = inner_style_meta or cover_style_meta or {}
    visual_focus = visual_focus or {}
    cover_label = cover_style_meta.get('label') or cover_style_meta.get('asset_type') or cover_style_meta.get('key') or '封面'
    inner_label = inner_style_meta.get('label') or inner_style_meta.get('asset_type') or inner_style_meta.get('key') or '内页'
    cover_headline = visual_focus.get('cover_headline') or ''
    cover_subheadline = visual_focus.get('cover_subheadline') or ''
    body_excerpt = visual_focus.get('body_excerpt') or ''

    if workflow_role == 'cover':
        lines = [
            f'当前只生成封面主图，版式重点按“{cover_label}”来做。',
            f'封面主句优先突出“{cover_headline}”，并给后续叠中文标题预留清晰留白。',
            '首图不要塞满说明文字，不要把内页结构直接搬到封面。',
        ]
        if cover_subheadline:
            lines.append(f'封面副句或短标签可围绕“{cover_subheadline}”来组织。')
        return ' '.join(lines)

    if workflow_role == 'inner':
        return ' '.join([
            f'当前更像图文内页，结构重点按“{inner_label}”来做。',
            f'内页要承接正文重点：{body_excerpt}。',
            '不要再做大字封面，要更像知识卡、报告解读卡、清单卡或课堂笔记页。',
        ])

    if generation_mode == 'cover_only':
        return ' '.join([
            f'这次只做封面主图，统一按“{cover_label}”来表现。',
            f'封面主句围绕“{cover_headline}”，大标题醒目，标题区留白清楚。',
            '不扩展内页，不要把多段正文和太多模块塞进一张图。',
        ])

    if generation_mode == 'inner_only':
        return ' '.join([
            f'这次只做图文内页，统一按“{inner_label}”来表现。',
            f'内页承接重点：{body_excerpt}。',
            '页面要像可收藏的知识卡/清单卡，不要做成封面海报。',
        ])

    lines = [
        f'如果一次生成多张图，请理解成图文套组：第 1 张更像封面，后续更像内页。',
        f'首图按“{cover_label}”做，重点突出“{cover_headline}”，要有大标题留白和缩略图可读性。',
        f'后续图按“{inner_label}”做，重点承接：{body_excerpt}。',
        f'本次预计输出 {max(_safe_int(image_count, 1), 1)} 张图，封面和内页不能长得一模一样。',
    ]
    if cover_subheadline:
        lines.append(f'首图的副句或短标签可以围绕“{cover_subheadline}”展开。')
    return ' '.join(lines)


def _recommended_cover_style_key_for_traits(traits):
    if traits.get('report_like'):
        return 'medical_science'
    if traits.get('story_like') or traits.get('emotion_like'):
        return 'memo_mobile'
    if traits.get('myth_like'):
        return 'poster_bold'
    if traits.get('discussion_like'):
        return 'knowledge_card'
    return 'knowledge_card'


def _score_cover_suitability(
    *,
    style_key='',
    family_key='',
    generation_mode='smart_bundle',
    selected_content='',
    title_hint='',
    workflow_role='',
    reference_guided=False,
):
    merged = ' '.join(filter(None, [title_hint or '', selected_content or '']))
    traits = _detect_topic_strategy_traits(merged, merged, merged)
    style_meta = _asset_style_meta(style_key or family_key or 'knowledge_card')
    resolved_family = (family_key or style_meta.get('family') or style_meta.get('key') or 'knowledge_card').strip()
    score = 72
    reasons = []

    if workflow_role == 'cover' or generation_mode == 'cover_only':
        score += 6
    if workflow_role == 'inner' or generation_mode == 'inner_only':
        score -= 12
        reasons.append('当前更偏内页结构，不是强封面路线')
    if reference_guided:
        score += 8
        reasons.append('已有参考图，封面气质会更稳')

    if resolved_family == 'medical_science':
        if traits.get('report_like'):
            score += 16
            reasons.append('检查/报告类内容非常适合医学解释封面')
        elif traits.get('story_like'):
            score -= 6
            reasons.append('第一人称经历内容不一定适合过强医学说明感')
        else:
            score += 4
    elif resolved_family == 'poster':
        if traits.get('myth_like') or traits.get('discussion_like'):
            score += 14
            reasons.append('结论冲突感强，适合大字封面先抓点击')
        elif traits.get('report_like'):
            score -= 10
            reasons.append('报告类内容更需要信息层级，不适合只靠大字')
        else:
            score += 3
    elif resolved_family == 'memo':
        if traits.get('story_like') or traits.get('emotion_like'):
            score += 14
            reasons.append('经历/情绪类内容很适合备忘录或陪伴感封面')
        elif traits.get('report_like'):
            score -= 8
            reasons.append('报告解读类内容更需要结构化说明')
    elif resolved_family == 'checklist':
        if traits.get('report_like') or traits.get('discussion_like'):
            score += 12
            reasons.append('步骤和判断顺序明确，适合清单型封面')
        elif traits.get('story_like'):
            score -= 4
            reasons.append('强经历类内容直接做清单感会偏硬')
    elif resolved_family == 'knowledge_card':
        score += 8
        reasons.append('知识卡片是相对稳妥的通用封面路线')

    title_text = (title_hint or '').strip()
    if any(token in title_text for token in ['别', '先', '为什么', '很多人', '怎么办', '你会']):
        score += 4
    if 10 <= len(title_text) <= 20:
        score += 3
    elif len(title_text) > 24:
        score -= 3

    score = max(min(score, 96), 38)
    if score >= 86:
        label = '强封面'
    elif score >= 72:
        label = '可直接试'
    else:
        label = '更像内页'

    fallback_style_key = _recommended_cover_style_key_for_traits(traits)
    fallback_style_label = _asset_style_meta(fallback_style_key).get('label') or fallback_style_key
    reason_text = '；'.join(reasons[:2]) or '当前路线可用，但仍建议先看封面是否一眼能读懂。'
    execution_note = (
        f'如果真实出图像内页，不像封面，建议优先切到“{fallback_style_label}”再试。'
        if score < 72 else
        '这条路线适合直接先试封面，再决定是否补内页。'
    )
    return {
        'score': score,
        'label': label,
        'reason': reason_text,
        'fallback_style_key': fallback_style_key,
        'fallback_style_label': fallback_style_label,
        'execution_note': execution_note,
    }


def _default_inner_style_for_cover(cover_style_key=''):
    meta = _asset_style_meta(cover_style_key or 'knowledge_card')
    family = (meta.get('family') or meta.get('key') or 'knowledge_card').strip()
    mapping = {
        'medical_science': 'knowledge_card',
        'knowledge_card': 'knowledge_card',
        'poster': 'knowledge_card',
        'checklist': 'checklist_report',
        'memo': 'memo_classroom',
        'reference_based': 'knowledge_card',
    }
    return mapping.get(family, meta.get('key') or 'knowledge_card')


def _resolve_asset_workflow_decision(
    *,
    style_value='medical_science',
    cover_style_type='',
    inner_style_type='',
    generation_mode='smart_bundle',
    selected_content='',
    title_hint='',
    reference_assets=None,
    prefer_cover_safety=True,
):
    reference_assets = reference_assets or []
    style_meta = _asset_style_meta(style_value)
    original_cover_meta = _asset_style_meta((cover_style_type or style_meta.get('key')))
    cover_meta = original_cover_meta
    inner_meta = _asset_style_meta((inner_style_type or style_meta.get('key')))
    resolved_mode = (generation_mode or style_meta.get('generation_mode') or 'text_to_image').strip()[:50] or 'text_to_image'
    if reference_assets and resolved_mode == 'text_to_image':
        resolved_mode = 'reference_guided'

    cover_fit = _score_cover_suitability(
        style_key=cover_meta.get('key') or style_meta.get('key') or '',
        family_key=cover_meta.get('family') or style_meta.get('family') or '',
        generation_mode=resolved_mode,
        selected_content=selected_content,
        title_hint=title_hint,
        reference_guided=bool(reference_assets),
    )
    auto_adjusted_cover = False
    adjustment_note = ''
    if prefer_cover_safety and resolved_mode in {'smart_bundle', 'cover_only'} and cover_fit['score'] < 72:
        fallback_meta = _asset_style_meta(cover_fit.get('fallback_style_key') or cover_meta.get('key') or style_meta.get('key'))
        if fallback_meta.get('key') and fallback_meta.get('key') != cover_meta.get('key'):
            cover_meta = fallback_meta
            auto_adjusted_cover = True
            adjustment_note = (
                f"原封面样式“{original_cover_meta.get('label') or original_cover_meta.get('key') or '-'}”"
                f"更像内页，已自动切到“{cover_meta.get('label') or cover_meta.get('key') or '-'}”。"
            )
            if not inner_style_type or inner_meta.get('key') == original_cover_meta.get('key'):
                inner_meta = _asset_style_meta(_default_inner_style_for_cover(cover_meta.get('key')))
            cover_fit = _score_cover_suitability(
                style_key=cover_meta.get('key') or style_meta.get('key') or '',
                family_key=cover_meta.get('family') or style_meta.get('family') or '',
                generation_mode=resolved_mode,
                selected_content=selected_content,
                title_hint=title_hint,
                reference_guided=bool(reference_assets),
            )

    prompt_style_meta = inner_meta if resolved_mode == 'inner_only' else cover_meta
    return {
        'style_meta': style_meta,
        'prompt_style_meta': prompt_style_meta,
        'cover_style_meta': cover_meta,
        'inner_style_meta': inner_meta,
        'generation_mode': resolved_mode,
        'cover_fit': cover_fit,
        'auto_adjusted_cover': auto_adjusted_cover,
        'adjustment_note': adjustment_note,
        'original_cover_style_key': original_cover_meta.get('key') or '',
        'original_cover_style_label': original_cover_meta.get('label') or '',
    }


def _normalize_image_prompt_mode(raw_mode='standard'):
    mode = (raw_mode or 'standard').strip().lower()
    return mode if mode in {'standard', 'fast'} else 'standard'


def _infer_asset_layout_variant(style_key, title_text='', body_text='', support_points=None):
    merged = ' '.join(filter(None, [title_text or '', body_text or '', ' '.join(support_points or [])]))
    source = merged.lower()

    if style_key in {'poster_bold', 'poster_handwritten', 'memo_mobile', 'memo_classroom', 'checklist_table', 'checklist_timeline', 'checklist_report'}:
        return style_key

    if style_key == 'medical_science':
        if any(token in source for token in ['对比', 'vs', '真相', '误区', '区别', '熬夜', '伤害', '危害', '全面', '好处', '坏处']):
            return 'impact_compare'
        if any(token in source for token in ['信号', '症状', '警示', '求救', '异常', '发黄', '疼', '不适', '表现', '征兆']):
            return 'symptom_warning'
        if any(token in source for token in ['ct', '核磁', 'b超', '超声', '扫描', '检查', '影像', '成像', '技术', '设备']):
            return 'device_explainer'
        return 'organ_explainer'

    if style_key == 'knowledge_card':
        if any(token in source for token in ['对比', 'vs', '区别', '瘦型', '隐形', '体外瘦', '体内胖', '正常', '异常']):
            return 'comparison_card'
        if any(token in source for token in ['机制', '原理', '通路', '流程', '循环', '导致', '路径', '代谢', '抵抗']):
            return 'mechanism_cycle'
        if any(token in source for token in ['全面', '伤害', '影响', '症状', '信号', '系统', '全身']):
            return 'body_map'
        return 'knowledge_breakdown'

    if style_key == 'poster':
        if any(token in source for token in ['经验', '总结', 'emo', '崩溃', '照顾', '踩坑', '分享', '父亲', '我']):
            return 'poster_handwritten'
        return 'poster_bold'

    if style_key == 'memo':
        if any(token in source for token in ['并发症', '病理', '生理', '检查', '代偿', '失代偿', '表现', '治疗要点', '实验室']):
            return 'memo_classroom'
        return 'memo_mobile'

    if style_key == 'checklist':
        if any(token in source for token in ['早餐', '午餐', '晚餐', 'day', '食谱', '7天', '7 天', '加餐', '几点', '时间']):
            return 'checklist_timeline'
        if any(token in source for token in ['报告', '彩超', '化验', '指标', '一次看懂', '检查结果', 'ast', 'alt', 'ggt', 'alp']):
            return 'checklist_report'
        return 'checklist_table'

    return 'standard'


def _build_variant_alignment_lines(style_key, layout_variant):
    if style_key == 'medical_science':
        if layout_variant == 'impact_compare':
            return [
                '版面尽量做成样例里的强对比封面：左侧偏暗或偏警示色，右侧偏亮或偏健康色，中间可放 VS、箭头或分隔视觉。',
                '左边突出错误行为、受损器官、风险指标；右边突出正确状态、健康节律、改善结果。',
            ]
        if layout_variant == 'symptom_warning':
            return [
                '版面贴近“求救信号/症状警示”样例：人体或器官为主，周围分散多个症状小标签、局部特写和风险提示。',
                '可加入眼睛、手掌、疲惫状态、疼痛位置这类局部说明，使画面像医学警示图而不是普通插画。',
            ]
        if layout_variant == 'device_explainer':
            return [
                '版面贴近“CT/B超/核磁讲解”样例：检查设备要足够大，旁边配成像原理、适合检查、注意事项、优缺点等模块。',
                '信息块要规整，像医学设备知识海报，但依然保留小红书科普卡的易读感。',
            ]
        return [
            '版面贴近“器官科普拆解”样例：一个核心器官占据视觉中心，配剖面、箭头、局部放大和短标签。',
            '适合回答“为什么”“是什么”“怎么判断”这类问题，视觉上要有教材感但不过于严肃。',
        ]

    if style_key == 'knowledge_card':
        if layout_variant == 'comparison_card':
            return [
                '版面贴近“左右对比知识卡”样例：中间主体大图，左右或上下做状态对照，强调外表正常 vs 内部异常、误区 vs 正解。',
                '可加入体型轮廓、器官切面、脂肪分布、两个不同状态的说明框，让读者一眼看懂差异。',
            ]
        if layout_variant == 'mechanism_cycle':
            return [
                '版面贴近“机制图/循环图”样例：中心器官或核心结论居中，环形箭头串起多个环节，形成完整因果链。',
                '每个机制节点配一个小图或局部结构示意，而不是纯文字堆砌。',
            ]
        if layout_variant == 'body_map':
            return [
                '版面贴近“全身影响总览图”样例：一个人体轮廓在中间，左右分布不同系统的影响标签和对应小图标。',
                '重点信息要围绕人体形成放射式布局，像高收藏的全景科普卡。',
            ]
        return [
            '版面贴近“结构拆解卡”样例：大标题下面是核心结构图，四周加简短标签框和局部放大图，底部再补一条结论区。',
            '适合解释定义、成因、检查指标、器官结构和高频误区。',
        ]

    if style_key == 'poster':
        if layout_variant == 'poster_handwritten':
            return [
                '版面贴近“手写经验大字报”样例：大标题竖向堆叠，关键词加荧光涂抹底色，整体留白很多。',
                '适合第一人称经验分享、情绪表达、总结踩坑点这类封面。',
            ]
        return [
            '版面贴近“黑体警示大字报”样例：标题像重磅警示牌，几乎全靠文字冲击力抓住注意力。',
            '适合指南、版本说明、结论型封面，背景保持极简。',
        ]

    if style_key == 'memo':
        if layout_variant == 'memo_classroom':
            return [
                '版面贴近“课堂笔记/复习资料”样例：分点、下划线、圈注、荧光笔和箭头很多，像老师划重点。',
                '适合病理、生理、并发症、检查要点这类需要整理重点的内容。',
            ]
        return [
            '版面贴近“手机备忘录”样例：顶部像原生备忘录 UI，白底、大留白、3-5 条清单式短句。',
            '适合补救指南、生活方式建议、注意事项和高收藏攻略。',
        ]

    if style_key == 'checklist':
        if layout_variant == 'checklist_timeline':
            return [
                '版面贴近“7天食谱/时间轴计划”样例：按 Day 或时间点切块，早餐/午餐/晚餐/加餐等模块清楚。',
                '适合用户照着执行，不只是看概念。',
            ]
        if layout_variant == 'checklist_report':
            return [
                '版面贴近“彩超/化验报告解读”样例：一个项目对应一个解释块，旁边辅以示意图和圈注说明。',
                '适合把专业报告翻译成普通人能看懂的结构化答案。',
            ]
        return [
            '版面贴近“白名单/产品怎么选/参数对比”样例：标准表格或矩阵结构，真实产品或食物图可作为格子里的素材。',
            '每一行或每一列都对应一个维度，最后必须给出推荐语或结论。',
        ]

    return []


def _build_style_specific_prompt(style_meta, clean_title='', primary_keyword='', body_text='', support_points=None, prompt_mode='standard'):
    style_key = style_meta.get('key') or ''
    prompt_mode = _normalize_image_prompt_mode(prompt_mode)
    focus_text = '；'.join((support_points or [])[:3]) or f'围绕{primary_keyword or clean_title or "主题"}做清晰表达'
    layout_variant = _infer_asset_layout_variant(style_key, clean_title, body_text, support_points or [])
    sample_signature = STYLE_REFERENCE_SIGNATURES.get(style_key, {})
    sample_lines = [
        value for value in [
            sample_signature.get('core_style'),
            sample_signature.get('composition'),
            sample_signature.get('palette'),
            sample_signature.get('illustration'),
            sample_signature.get('annotation'),
            sample_signature.get('footer'),
            sample_signature.get('avoid'),
        ] if value
    ]

    shared_lines = [
        '画面比例：竖版 3:4，适合小红书封面或图文首屏。',
        f'版式结构：{style_meta.get("layout_hint") or "顶部标题，中部主体，底部信息条"}。',
        f'视觉气质：{style_meta.get("visual_hint") or "干净、专业、信息层级清楚"}。',
        f'参考方向：{style_meta.get("reference_hint") or "小红书图文信息卡"}。',
        f'文字策略：{style_meta.get("text_policy") or "保留简洁标题和关键词，避免长段落乱码"}。',
        f'避免事项：{style_meta.get("avoid_hint") or "不要杂乱背景和营销感"}。',
    ]

    if style_key == 'medical_science':
        lines = [
            '画面主体优先使用人体轮廓、器官剖面、检查设备、症状标注、箭头标签、对照区块这类医学信息图元素。',
            MEDICAL_SCIENCE_LAYOUT_VARIANTS.get(layout_variant, MEDICAL_SCIENCE_LAYOUT_VARIANTS['organ_explainer']),
            '背景以白底、浅暖底、浅蓝绿医疗色为主，可搭配细边框和少量警示色点缀。',
            f'信息重点：{focus_text}。',
            '整体要像高质量医学科普图，不像普通商业海报，也不要只有一个插画主体没有信息结构。',
        ]
        if prompt_mode != 'fast':
            lines.extend([
                *sample_lines,
                *_build_variant_alignment_lines(style_key, layout_variant),
                '标题区要足够醒目，适合后续叠加中文大标题、副标题和短标签。',
                '如果内容适合，可加入左右对比、局部放大、图标标签、结论提示条，形成一眼看懂的阅读路径。',
                '插画质感以半写实手绘为主，器官结构准确，线条柔和，局部可以轻卡通化提升可读性。',
            ])
        return ' '.join(shared_lines + lines)

    if style_key == 'knowledge_card':
        lines = [
            '画面主体优先使用器官结构图、剖面图、机制箭头、局部放大框、编号提示、总结卡片这些知识卡元素。',
            KNOWLEDGE_CARD_LAYOUT_VARIANTS.get(layout_variant, KNOWLEDGE_CARD_LAYOUT_VARIANTS['knowledge_breakdown']),
            '背景以白底、米白底、浅暖底为主，边框清晰，卡片留白足够，适合收藏截图传播。',
            f'信息重点：{focus_text}。',
            '整体要像医学知识卡片内页，结构完整、说明性强，不能只做成封面海报。',
        ]
        if prompt_mode != 'fast':
            lines.extend([
                *sample_lines,
                *_build_variant_alignment_lines(style_key, layout_variant),
                '可在主体周围安排标签框、解释框、对比箭头、定义区和结论区，让层级一目了然。',
                '插画是医学手绘科普风，不要过度写实，不要廉价卡通，要兼顾专业感和亲和力。',
                '如果模型不擅长精确文字，请用整齐的信息框和标签占位替代密集长文，保留后期排版空间。',
            ])
        return ' '.join(shared_lines + lines)

    if style_key == 'reference_based':
        lines = [
            '这是一张参考图驱动的底图任务：优先吸收参考图的构图、色彩、留白、信息块位置和插画气质。',
            '底图不要直接写中文大段文字，重点是把参考风格转成适合后续叠字的无字或少字底图。',
            f'信息重点：{focus_text}。',
            '如果参考图偏医学科普，就保留器官主体、放射状标注、局部放大和对比结构；如果参考图偏知识卡片，就保留模块化信息岛、粗箭头和柔和配色。',
        ]
        if prompt_mode != 'fast':
            lines.extend([
                '强调“风格靠近而不是照抄”：保留视觉语言，不复制原图中的具体文案、水印或品牌元素。',
                '底图要给后续标题、说明卡片和标签框预留足够留白，方便系统后处理排版。',
            ])
        return ' '.join(shared_lines + lines)

    if style_key == 'poster':
        lines = [
            '核心视觉是大标题本身，文字至少占据画面 60% 以上面积。',
            '关键词要用荧光底块高亮，背景尽量极简，保证缩略图下也能看清。',
            f'信息重点：{focus_text}。',
            '这类图优先走模板排版，不依赖复杂插画。',
        ]
        if prompt_mode != 'fast':
            lines.extend([
                *_build_variant_alignment_lines(style_key, layout_variant),
                '黑体警示版偏重结论和警告；手写经验版偏重第一人称情绪表达和总结感。',
                '文字要有压迫感和冲击力，像流量封面，不像普通知识卡片。',
            ])
        return ' '.join(shared_lines + lines)

    if style_key == 'memo':
        lines = [
            '整体像私人收藏的备忘录或课堂笔记，文字是主体，插画只是少量点缀。',
            '重点内容必须短句化，并用高亮、emoji、圈注或下划线强化阅读路径。',
            f'信息重点：{focus_text}。',
            '这类图也优先走模板排版和后处理叠字。',
        ]
        if prompt_mode != 'fast':
            lines.extend([
                *_build_variant_alignment_lines(style_key, layout_variant),
                '手机备忘录版更生活化、更口语化；课堂笔记版更像老师划重点和复习资料。',
            ])
        return ' '.join(shared_lines + lines)

    if style_key == 'checklist':
        lines = [
            '这类图要帮助用户完成筛选、对比和执行，所以结构要比装饰更重要。',
            '表格、时间轴、模块卡、报告说明框都比纯插画更优先。',
            f'信息重点：{focus_text}。',
            '真实产品图、食物图或示意图可以作为辅助素材，但不应盖过结构本身。',
        ]
        if prompt_mode != 'fast':
            lines.extend([
                *_build_variant_alignment_lines(style_key, layout_variant),
                '表格对比版要明确横纵表头；时间轴版要明确早中晚或 Day1-Day7；报告解读版要做到一项一解释。',
            ])
        return ' '.join(shared_lines + lines)

    extra_lines = [
        f'信息重点：{focus_text}。',
        '整体要求：信息块清楚，适合小红书图文传播。',
    ]
    return ' '.join(shared_lines + extra_lines)


def _build_asset_generation_prompt_from_context(
    topic_name='',
    topic_keywords='',
    selected_content='',
    style_preset='小红书图文',
    title_hint='',
    *,
    cover_style_key='',
    inner_style_key='',
    generation_mode='smart_bundle',
    image_count=1,
    workflow_role='',
):
    runtime_config = _automation_runtime_config()
    prompt_mode = _normalize_image_prompt_mode(runtime_config.get('image_optimize_prompt_mode'))
    style_meta = _asset_style_meta(style_preset or runtime_config.get('image_default_style_type') or 'medical_science')
    cover_style_meta = _asset_style_meta(cover_style_key or style_meta.get('key') or style_preset)
    inner_style_meta = _asset_style_meta(inner_style_key or style_meta.get('key') or style_preset)
    resolved_topic_name = (topic_name or '肝病管理').strip() or '肝病管理'
    keywords = _split_keywords(topic_keywords or resolved_topic_name)
    primary_keyword = keywords[0] if keywords else resolved_topic_name
    visual_focus = _extract_visual_focus_payload(selected_content, title_hint, primary_keyword)
    clean_title = visual_focus['clean_title']
    body = re.sub(r'^(正文|内文)\s*[：:]\s*', '', (visual_focus.get('body_text') or '').strip())
    support_points = visual_focus.get('support_points') or []
    point_text = visual_focus.get('body_excerpt') or f'围绕{primary_keyword}做清晰信息表达'

    prompt_parts = [
        f'为小红书生成 1 张{style_meta["asset_type"]}，主题“{clean_title}”，适合直接做图文配图或封面。',
        f'内容主题：{resolved_topic_name}；核心关键词：{primary_keyword}。',
        f'封面主句：{visual_focus.get("cover_headline") or clean_title}。',
        (f'封面副句：{visual_focus.get("cover_subheadline")}。' if visual_focus.get('cover_subheadline') else ''),
        f'需要表达的重点：{point_text}。',
        _build_style_specific_prompt(
            style_meta,
            clean_title=clean_title,
            primary_keyword=primary_keyword,
            body_text=body,
            support_points=support_points,
            prompt_mode=prompt_mode,
        ),
        _build_image_workflow_prompt_notes(
            generation_mode,
            cover_style_meta=cover_style_meta,
            inner_style_meta=inner_style_meta,
            visual_focus=visual_focus,
            image_count=image_count,
            workflow_role=workflow_role,
        ),
        f'{style_meta["prompt_suffix"]}',
        '整体要求：画面干净、可信、适合截图传播，不过度营销，不出现品牌露出和水印。',
    ]
    prompt = ' '.join(part.strip() for part in prompt_parts if part and str(part).strip())
    prompt_suffix = str(runtime_config.get('image_prompt_suffix') or '').strip()
    if prompt_suffix:
        prompt = f'{prompt} {prompt_suffix}'
    return prompt


def _build_creative_pack(topic, selected_content='', preferred_style='', reference_assets=None):
    topic_name = topic.topic_name or '肝病热点'
    keywords = _split_keywords(topic.keywords or topic_name)
    primary_keyword = keywords[0] if keywords else topic_name
    insertion = _detect_soft_insertion(f'{topic_name} {" ".join(keywords)}')
    content_points = _extract_content_points(selected_content)
    preferred_style = (preferred_style or '').strip()
    reference_assets = reference_assets or []
    reference_titles = [item.title or item.product_name or f'参考图{item.id}' for item in reference_assets[:3]]
    reference_tags = []
    for item in reference_assets[:3]:
        reference_tags.extend([part.strip() for part in (item.tags or '').split(',') if part.strip()])
    reference_tags = reference_tags[:4]

    if preferred_style:
        style_keys = [preferred_style]
        preferred_meta = _asset_style_meta(preferred_style)
        preferred_family = preferred_meta.get('family') or ''
        if preferred_family:
            for item in ASSET_STYLE_TYPE_DEFINITIONS:
                if item.get('family') == preferred_family and item.get('key') not in style_keys:
                    style_keys.append(item.get('key'))
        base_title = (_extract_title_from_version(selected_content) or topic_name or preferred_meta.get('label') or '肝病管理').strip()[:18]
        pack = []
        for idx, style_key in enumerate(style_keys[:3], start=1):
            meta = _asset_style_meta(style_key)
            title = base_title if idx == 1 else f"{primary_keyword[:10]} {idx}"
            subtitle = meta.get('description') or meta.get('label') or '图片方案'
            bullets = content_points[:3] if content_points else list(meta.get('default_bullets') or [])
            if style_key == 'reference_based' and reference_titles:
                subtitle = f"参考 {(' / '.join(reference_titles[:2]))} 的构图和留白"
                bullets = [
                    f"继承{reference_titles[0]}的版心和留白",
                    f"保留“{primary_keyword}”当前内容重点",
                    f"参考标签：{('、'.join(reference_tags[:2]) or '医学科普')}",
                ]
            accent = meta.get('accent') or '#ff7a59'
            bg = meta.get('bg') or '#fff4ee'
            if style_key in {'poster', 'poster_bold', 'poster_handwritten'}:
                render_key = style_key if style_key != 'poster' else _infer_asset_layout_variant(style_key, title, selected_content, bullets)
                svg = _render_svg_poster_card(render_key, title, subtitle, bullets, accent=accent, bg=bg)
            elif style_key in {'memo', 'memo_mobile', 'memo_classroom'}:
                render_key = style_key if style_key != 'memo' else _infer_asset_layout_variant(style_key, title, selected_content, bullets)
                svg = _render_svg_memo_card(render_key, title, subtitle, bullets, accent=accent, bg=bg)
            elif style_key in {'checklist', 'checklist_table', 'checklist_timeline', 'checklist_report'}:
                render_key = style_key if style_key != 'checklist' else _infer_asset_layout_variant(style_key, title, selected_content, bullets)
                svg = _render_svg_checklist_card(render_key, title, subtitle, bullets, accent=accent, bg=bg)
            else:
                svg = _render_svg_card(
                    meta.get('asset_type') or meta.get('label') or '知识卡片',
                    title,
                    subtitle,
                    bullets,
                    accent=accent,
                    bg=bg,
                )
            pack.append({
                'style_type': style_key,
                'type': meta.get('asset_type') or meta.get('label') or '知识卡片',
                'title': title,
                'subtitle': subtitle,
                'bullets': bullets,
                'image_prompt': (
                    _build_asset_generation_prompt(
                        topic,
                        selected_content,
                        style_preset=style_key,
                        title_hint=title,
                        cover_style_key=style_key,
                        inner_style_key=style_key,
                        generation_mode='cover_only' if idx == 1 else 'inner_only',
                        image_count=1,
                        workflow_role='cover' if idx == 1 else 'inner',
                    )
                    + (f" 参考图方向：{' / '.join(reference_titles[:3])}。" if reference_titles else '')
                ),
                'download_name': f'creative_{topic.id}_{style_key}_{idx}.svg',
                'svg_data_uri': _svg_data_uri(svg),
                'reference_titles': reference_titles[:3],
            })
        return pack

    if any(token in topic_name for token in ['体检', 'FibroScan', '福波看', '肝弹', '转氨酶', '肝硬度']):
        configs = [
            ('医学科普图', f'{primary_keyword}怎么判断', '体检后先看这3件事', content_points or ['先看指标变化，不只盯单次结果', '把复查时间和检查方式说明白', f'自然带出{insertion}的检查评估场景'], '#FF7A59', '#FFF4EE'),
            ('检查流程图', f'{primary_keyword}复查流程', '适合做流程型配图', content_points or ['异常发现', '进一步评估', '复查跟踪'], '#F2A43C', '#FFF8EA'),
            ('知识卡片', f'{primary_keyword}高频问题', '评论区最容易被问到的点', content_points or ['为什么会异常', '什么人更要重视', '下一步怎么安排'], '#4C91FF', '#EEF5FF'),
        ]
    elif '脂肪肝' in topic_name:
        configs = [
            ('知识卡片', f'{primary_keyword}别忽视', '适合收藏型封面', content_points or ['先拆误区，再给动作建议', '重点讲体检和复查节奏', f'软植入{insertion}要自然'], '#FF6B6B', '#FFF2F2'),
            ('误区对照图', f'{primary_keyword}常见误区', '左右对照最容易看懂', content_points or ['误区1：没症状就不用管', '误区2：只看一次报告', '误区3：只靠短期节食'], '#D96570', '#FFF2F5'),
            ('复查清单卡', f'{primary_keyword}复查清单', '截图保存型内容', content_points or ['检查项目别漏掉', '饮食运动要记录', '复查前后做对比'], '#61A36B', '#F0F8F1'),
        ]
    else:
        configs = [
            ('医学科普图', f'{primary_keyword}重点提醒', '适合做专业感封面', content_points or ['标题先问问题', '正文给真实场景', f'植入{insertion}时别写成硬广'], '#FF7A59', '#FFF4EE'),
            ('知识卡片', f'{primary_keyword}三件事', '收藏感最强的一种版式', content_points or ['适合拆成三格信息块', '首屏保留数字和问句', '结尾引导经验交流'], '#4C91FF', '#EEF5FF'),
            ('复查清单卡', f'{primary_keyword}管理清单', '适合带复查节奏', content_points or ['什么时候复查', '哪些指标要看', '如何记录趋势'], '#61A36B', '#F0F8F1'),
        ]

    pack = []
    for idx, (card_type, title, subtitle, bullets, accent, bg) in enumerate(configs, start=1):
        svg = _render_svg_card(card_type, title, subtitle, bullets, accent=accent, bg=bg)
        pack.append({
            'type': card_type,
            'title': title,
            'subtitle': subtitle,
            'bullets': bullets,
            'image_prompt': (
                _build_asset_generation_prompt_from_context(
                    topic_name=topic.topic_name or '肝病管理',
                    topic_keywords=topic.keywords or topic.topic_name or primary_keyword,
                    selected_content=selected_content,
                    style_preset=card_type,
                    title_hint=title,
                    cover_style_key=card_type,
                    inner_style_key=card_type,
                    generation_mode='cover_only' if idx == 1 else 'inner_only',
                    image_count=1,
                    workflow_role='cover' if idx == 1 else 'inner',
                ) + f' 软植入重点：{insertion}。'
            ),
            'download_name': f'creative_{topic.id}_{idx}.svg',
            'svg_data_uri': _svg_data_uri(svg),
        })
    return pack


def _build_graphic_article_bundle(topic, selected_content='', cover_style_type='', inner_style_type='', generation_mode='smart_bundle', reference_assets=None):
    cover_style_type = (cover_style_type or '').strip()
    inner_style_type = (inner_style_type or '').strip()
    generation_mode = (generation_mode or 'smart_bundle').strip() or 'smart_bundle'
    reference_assets = reference_assets or []
    if generation_mode == 'cover_only':
        creative_pack = _build_creative_pack(topic, selected_content, preferred_style=cover_style_type or inner_style_type, reference_assets=reference_assets)[:1]
    elif generation_mode == 'inner_only':
        creative_pack = _build_creative_pack(topic, selected_content, preferred_style=inner_style_type or cover_style_type, reference_assets=reference_assets)
    elif cover_style_type and inner_style_type and cover_style_type != inner_style_type:
        cover_assets = _build_creative_pack(topic, selected_content, preferred_style=cover_style_type, reference_assets=reference_assets)[:1]
        inner_assets = _build_creative_pack(topic, selected_content, preferred_style=inner_style_type, reference_assets=reference_assets)
        creative_pack = cover_assets + inner_assets[:2]
    else:
        creative_pack = _build_creative_pack(topic, selected_content, preferred_style=cover_style_type or inner_style_type, reference_assets=reference_assets)
    topic_name = topic.topic_name or '肝病管理'
    keywords = _split_keywords(topic.keywords or topic_name)
    content_title = _extract_title_from_version(selected_content) if selected_content else topic_name
    raw_body = _extract_body_from_version(selected_content) if selected_content else ''
    raw_body = re.sub(r'^(正文|内文)\s*[：:]\s*', '', (raw_body or '').strip())
    if not raw_body:
        raw_body = f'这篇内容围绕“{topic_name}”展开，先讲真实场景，再拆重点，最后给出复查或管理建议。'

    body_sentences = [item.strip() for item in re.split(r'[。！？\n]+', raw_body) if item and item.strip()]
    body_excerpt = '。'.join(body_sentences[:3]).strip()
    if body_excerpt and not body_excerpt.endswith(('。', '！', '？')):
        body_excerpt += '。'

    tag_text = ' '.join([f'#{item}' for item in keywords[:4]])
    bundles = []
    for index, asset in enumerate(creative_pack, start=1):
        bullet_lines = [item.strip() for item in (asset.get('bullets') or []) if item and item.strip()]
        support_text = '；'.join(bullet_lines[:3])
        publish_title = (asset.get('title') or content_title or topic_name).strip()[:18]
        publish_body = body_excerpt
        if support_text:
            publish_body = f'{publish_body}\n\n这版我想重点放在：{support_text}。'
        publish_body = f"{publish_body}\n\n封面我会用“{asset.get('type') or '知识卡片'}”这个版式，方便一眼看懂重点。"
        if asset.get('reference_titles'):
            publish_body = f"{publish_body}\n\n这版封面会优先借参考图的构图和留白，但内容还是按我们现在这篇重新组织。"
        publish_body = f"{publish_body}\n\n你们更想看哪一类延展内容？"

        full_copy = f"标题：{publish_title}\n内文：{publish_body}\n\n推荐标签：{tag_text or '#肝病管理'}"
        bundles.append({
            'id': index,
            'asset': asset,
            'publish_title': publish_title,
            'publish_body': publish_body,
            'full_copy': full_copy,
            'tag_text': tag_text or '#肝病管理',
            'cover_title': asset.get('title') or publish_title,
            'card_type': asset.get('type') or '知识卡片',
            'summary': support_text or asset.get('subtitle') or '',
        })
    return bundles


def _build_asset_generation_prompt(
    topic,
    selected_content='',
    style_preset='小红书图文',
    title_hint='',
    *,
    cover_style_key='',
    inner_style_key='',
    generation_mode='smart_bundle',
    image_count=1,
    workflow_role='',
):
    return _build_asset_generation_prompt_from_context(
        topic_name=topic.topic_name or '肝病管理',
        topic_keywords=topic.keywords or topic.topic_name or '肝病管理',
        selected_content=selected_content,
        style_preset=style_preset,
        title_hint=title_hint,
        cover_style_key=cover_style_key,
        inner_style_key=inner_style_key,
        generation_mode=generation_mode,
        image_count=image_count,
        workflow_role=workflow_role,
    )


def _build_asset_generation_fallback_results(topic, selected_content='', image_count=3, style_preset='', title_hint=''):
    style_meta = _asset_style_meta(style_preset or 'medical_science')
    creative_pack = _build_creative_pack(topic, selected_content)
    if style_preset:
        base_title = (title_hint or _extract_title_from_version(selected_content) or topic.topic_name or style_meta['label']).strip()
        points = _extract_content_points(selected_content) or list(style_meta.get('default_bullets') or [])
        accent = style_meta.get('accent') or '#ff7a59'
        bg = style_meta.get('bg') or '#fff4ee'
        style_key = style_meta.get('key') or ''
        custom_results = []
        for idx in range(1, max(image_count, 1) + 1):
            title = base_title[:18] if idx == 1 else f'{base_title[:14]} {idx}'
            subtitle = style_meta.get('description') or style_meta['label']
            bullets = points[:3] if points else list(style_meta.get('default_bullets') or [])
            if style_key in {'poster', 'poster_bold', 'poster_handwritten'}:
                render_key = style_key if style_key != 'poster' else _infer_asset_layout_variant(style_key, title, selected_content, bullets)
                svg = _render_svg_poster_card(render_key, title, subtitle, bullets, accent=accent, bg=bg)
            elif style_key in {'memo', 'memo_mobile', 'memo_classroom'}:
                render_key = style_key if style_key != 'memo' else _infer_asset_layout_variant(style_key, title, selected_content, bullets)
                svg = _render_svg_memo_card(render_key, title, subtitle, bullets, accent=accent, bg=bg)
            elif style_key in {'checklist', 'checklist_table', 'checklist_timeline', 'checklist_report'}:
                render_key = style_key if style_key != 'checklist' else _infer_asset_layout_variant(style_key, title, selected_content, bullets)
                svg = _render_svg_checklist_card(render_key, title, subtitle, bullets, accent=accent, bg=bg)
            else:
                svg = _render_svg_card(style_meta['label'], title, subtitle, bullets, accent=accent, bg=bg)
            custom_results.append({
                'index': idx,
                'type': style_meta['label'],
                'title': title,
                'subtitle': subtitle,
                'image_prompt': _build_asset_generation_prompt(topic, selected_content, style_preset=style_meta['key'], title_hint=title),
                'preview_url': _svg_data_uri(svg),
                'download_name': f'asset_task_{topic.id}_{style_meta["key"]}_{idx}.svg',
                'provider': 'svg_fallback',
                'format': 'svg',
                'bullets': bullets,
            })
        return custom_results

    results = []
    for idx, asset in enumerate(creative_pack[:max(image_count, 1)], start=1):
        results.append({
            'index': idx,
            'type': asset.get('type') or '知识卡片',
            'title': asset.get('title') or '',
            'subtitle': asset.get('subtitle') or '',
            'image_prompt': asset.get('image_prompt') or '',
            'preview_url': asset.get('svg_data_uri') or '',
            'download_name': asset.get('download_name') or f'asset_{idx}.svg',
            'provider': 'svg_fallback',
            'format': 'svg',
            'bullets': asset.get('bullets') or [],
        })
    return results


def _is_ark_seedream_model(model_name=''):
    return 'seedream' in str(model_name or '').strip().lower()


def _normalize_ark_seedream_size(size=''):
    raw = str(size or '').strip().lower()
    if raw in {'2k', '2048', '1536x2048', '2048x1536', '2048x2048'}:
        return '2K'
    if raw in {'1k', '1024', '1024x1024'}:
        return '2K'
    return size or '2K'


def _build_asset_provider_request_preview(
    provider,
    model_name,
    prompt_text,
    image_size,
    style_preset='小红书图文',
    image_count=3,
    product_assets=None,
    reference_assets=None,
    product_context=None,
):
    safe_provider = (provider or 'svg_fallback').strip() or 'svg_fallback'
    safe_model = (model_name or '').strip()
    safe_prompt = (prompt_text or '').strip()
    safe_size = (image_size or '1024x1536').strip() or '1024x1536'
    safe_style = (style_preset or '小红书图文').strip() or '小红书图文'
    safe_count = min(max(_safe_int(image_count, 3), 1), 4)
    product_assets = product_assets or []
    product_urls = [item.get('preview_url') for item in product_assets if item.get('preview_url')]
    reference_assets = reference_assets or []
    reference_urls = [item.get('preview_url') for item in reference_assets if item.get('preview_url')]
    product_context = product_context or {}

    if safe_provider == 'volcengine_ark':
        if _is_ark_seedream_model(safe_model):
            payload = {
                'model': safe_model or 'doubao-seedream-5-0-lite-260128',
                'prompt': safe_prompt,
                'size': _normalize_ark_seedream_size(safe_size),
                'response_format': 'url',
                'stream': False,
                'watermark': True,
                'sequential_image_generation': 'disabled',
            }
            if safe_count > 1:
                payload['n'] = safe_count
            return payload
        payload = {
            'model': safe_model or 'doubao-seededit-3-0-i2i-250628',
            'prompt': safe_prompt,
            'size': safe_size,
            'response_format': 'url',
            'n': safe_count,
        }
        if product_urls:
            payload['product_images'] = product_urls[:3]
        if reference_urls:
            payload['reference_images'] = reference_urls[:3]
        if product_context:
            payload['product_context'] = product_context
        return payload
    if safe_provider == 'volcengine_las':
        payload = {
            'model': safe_model or 'doubao-seedream-5-0-lite-260128',
            'prompt': safe_prompt,
            'size': safe_size,
            'response_format': 'url',
            'watermark': True,
        }
        if product_urls:
            payload['product_images'] = product_urls[:3]
        if reference_urls:
            payload['reference_images'] = reference_urls[:3]
        if product_context:
            payload['product_context'] = product_context
        return payload
    if safe_provider in {'openai', 'openai_compatible'}:
        payload = {
            'model': safe_model or 'gpt-image-1',
            'prompt': safe_prompt,
            'n': safe_count,
            'size': safe_size,
            'response_format': 'b64_json',
        }
        return payload
    if safe_provider in {'generic_json', 'custom_json'}:
        return {
            'model': safe_model or 'image-default',
            'prompt': safe_prompt,
            'image_count': safe_count,
            'size': safe_size,
            'style': safe_style,
            'response_format': 'b64_json',
            'product_images': product_urls[:5],
            'reference_images': reference_urls[:5],
            'product_context': product_context,
        }
    return {
        'mode': 'svg_fallback',
        'prompt': safe_prompt,
        'image_count': safe_count,
        'style': safe_style,
        'size': safe_size,
        'product_images': product_urls[:5],
        'reference_images': reference_urls[:5],
        'product_context': product_context,
    }


def _resolve_image_provider_api_key(provider=''):
    safe_provider = (provider or '').strip() or 'svg_fallback'
    if safe_provider == 'openai':
        return (
            os.environ.get('ASSET_IMAGE_API_KEY')
            or os.environ.get('OPENAI_IMAGE_API_KEY')
            or os.environ.get('OPENAI_API_KEY')
            or ''
        ).strip()
    if safe_provider == 'openai_compatible':
        return (
            os.environ.get('ASSET_IMAGE_API_KEY')
            or os.environ.get('OPENAI_IMAGE_API_KEY')
            or os.environ.get('OPENAI_API_KEY')
            or os.environ.get('ARK_API_KEY')
            or os.environ.get('LAS_API_KEY')
            or ''
        ).strip()
    if safe_provider == 'volcengine_ark':
        return (
            os.environ.get('ASSET_IMAGE_API_KEY')
            or os.environ.get('ARK_API_KEY')
            or ''
        ).strip()
    if safe_provider == 'volcengine_las':
        return (
            os.environ.get('ASSET_IMAGE_API_KEY')
            or os.environ.get('LAS_API_KEY')
            or ''
        ).strip()
    return (
        os.environ.get('ASSET_IMAGE_API_KEY')
        or os.environ.get('IMAGE_API_KEY')
        or os.environ.get('OPENAI_IMAGE_API_KEY')
        or os.environ.get('OPENAI_API_KEY')
        or ''
    ).strip()


def _resolve_image_provider_capabilities(payload=None):
    runtime_config = _automation_runtime_config()
    merged = dict(runtime_config)
    payload = payload or {}
    override_keys = [
        'image_provider',
        'image_api_base',
        'image_api_url',
        'image_model',
        'image_size',
        'image_timeout_seconds',
        'image_style_preset',
        'image_default_style_type',
        'image_optimize_prompt_mode',
        'image_prompt_suffix',
    ]
    for key in override_keys:
        if payload.get(key) not in [None, '']:
            merged[key] = payload.get(key)

    provider = (os.environ.get('ASSET_IMAGE_PROVIDER') or str(merged.get('image_provider') or 'svg_fallback')).strip() or 'svg_fallback'
    api_base = (os.environ.get('ASSET_IMAGE_API_BASE') or str(merged.get('image_api_base') or '')).strip()
    if provider == 'volcengine_ark' and not api_base:
        api_base = 'https://ark.cn-beijing.volces.com/api/v3'
    if provider == 'volcengine_las' and not api_base:
        api_base = 'https://operator.las.cn-beijing.volces.com/api/v1'
    if provider == 'openai' and not api_base:
        api_base = 'https://api.openai.com/v1'
    api_url = (os.environ.get('ASSET_IMAGE_API_URL') or str(merged.get('image_api_url') or '')).strip()
    if not api_url and api_base:
        api_url = api_base.rstrip('/') + '/images/generations'
    api_key = _resolve_image_provider_api_key(provider)
    model_name = (os.environ.get('ASSET_IMAGE_MODEL') or str(merged.get('image_model') or '')).strip()
    if not model_name and provider in {'volcengine_ark', 'volcengine_las'}:
        model_name = 'doubao-seedream-5-0-lite-260128'
    if not model_name and provider in {'openai', 'openai_compatible'}:
        model_name = 'gpt-image-1'
    image_size = (os.environ.get('ASSET_IMAGE_SIZE') or str(merged.get('image_size') or '1024x1536')).strip()
    timeout_seconds = min(max(_safe_int(merged.get('image_timeout_seconds'), 90), 10), 300)
    configured = bool(api_url and api_key)
    return {
        'image_provider_configured': configured,
        'image_provider_name': provider,
        'image_provider_api_base': api_base,
        'image_provider_api_url': api_url,
        'image_provider_model': model_name,
        'image_provider_size': image_size,
        'image_timeout_seconds': timeout_seconds,
        'image_style_preset': str(merged.get('image_style_preset') or '小红书图文'),
        'image_default_style_type': str(merged.get('image_default_style_type') or 'medical_science'),
        'image_optimize_prompt_mode': str(merged.get('image_optimize_prompt_mode') or 'standard'),
        'image_prompt_suffix': str(merged.get('image_prompt_suffix') or ''),
        'api_key_configured': bool(api_key),
        'fallback_mode': not configured or provider == 'svg_fallback',
        'provider_options': _image_provider_options(),
        'model_options': _image_model_options(provider),
        'style_type_options': _asset_style_type_options(),
    }


def _normalize_asset_provider_results(payload, provider='svg_fallback', image_prompt='', title_hint='测试图片接口'):
    candidates = []
    if isinstance(payload, dict):
        for key in ['data', 'images', 'output', 'results']:
            value = payload.get(key)
            if isinstance(value, list):
                candidates = value
                break
    elif isinstance(payload, list):
        candidates = payload

    normalized = []
    for index, item in enumerate(candidates, start=1):
        if isinstance(item, str):
            normalized.append({
                'index': index,
                'title': title_hint,
                'image_prompt': image_prompt,
                'preview_url': item,
                'provider': provider,
                'format': 'url',
            })
            continue
        if not isinstance(item, dict):
            continue
        if item.get('b64_json'):
            normalized.append({
                'index': index,
                'title': title_hint,
                'image_prompt': image_prompt,
                'preview_url': f"data:image/png;base64,{item.get('b64_json')}",
                'provider': provider,
                'format': 'png',
            })
        elif item.get('image_base64'):
            normalized.append({
                'index': index,
                'title': title_hint,
                'image_prompt': image_prompt,
                'preview_url': f"data:image/png;base64,{item.get('image_base64')}",
                'provider': provider,
                'format': 'png',
            })
        elif item.get('url') or item.get('image_url'):
            normalized.append({
                'index': index,
                'title': title_hint,
                'image_prompt': image_prompt,
                'preview_url': item.get('url') or item.get('image_url'),
                'provider': provider,
                'format': 'url',
            })
    return normalized


def _score_generated_asset_result(
    item,
    *,
    generation_mode='smart_bundle',
    cover_style_type='',
    inner_style_type='',
    style_preset='',
    provider='svg_fallback',
    image_count=1,
):
    index = max(_safe_int(item.get('index'), 1), 1)
    score = 58
    recommended_usage = 'general'
    recommended_usage_label = '通用'

    if generation_mode == 'cover_only':
        recommended_usage = 'cover'
        recommended_usage_label = '封面优先'
        score += 14 if index == 1 else -6
    elif generation_mode == 'inner_only':
        recommended_usage = 'inner'
        recommended_usage_label = '内页优先'
        score += 12
    elif generation_mode == 'smart_bundle':
        if index == 1:
            recommended_usage = 'cover'
            recommended_usage_label = '封面优先'
            score += 14
        else:
            recommended_usage = 'inner'
            recommended_usage_label = '内页优先'
            score += 10

    if provider != 'svg_fallback':
        score += 5
    if (item.get('preview_url') or '').strip():
        score += 5
    if (item.get('format') or '').strip() in {'png', 'url'}:
        score += 4
    if recommended_usage == 'cover' and (cover_style_type or style_preset):
        score += 4
    if recommended_usage == 'inner' and (inner_style_type or style_preset):
        score += 4
    if image_count > 1 and generation_mode == 'smart_bundle' and index > image_count:
        score -= 4

    score = max(min(int(score), 100), 0)
    if score >= 82:
        usable_label = '优先可用'
    elif score >= 72:
        usable_label = '可直接试'
    elif score >= 64:
        usable_label = '可再优化'
    else:
        usable_label = '建议重做'

    if recommended_usage == 'cover':
        usability_note = '更适合先拿来做首图或封面。'
    elif recommended_usage == 'inner':
        usability_note = '更适合放在图文内页承接重点信息。'
    else:
        usability_note = '更适合做通用配图或补充图。'
    if usable_label == '建议重做':
        usability_note = f'{usability_note} 当前可用性偏低，建议重做或切回模板预览。'
    elif usable_label == '可再优化':
        usability_note = f'{usability_note} 建议先人工过一眼再决定是否发布。'

    return {
        'usable_score': score,
        'usable_label': usable_label,
        'recommended_usage': recommended_usage,
        'recommended_usage_label': recommended_usage_label,
        'usability_note': usability_note,
    }


def _annotate_generated_asset_results(
    items,
    *,
    generation_mode='smart_bundle',
    cover_style_type='',
    inner_style_type='',
    style_preset='',
    provider='svg_fallback',
    image_count=1,
):
    annotated = []
    for item in (items or []):
        if not isinstance(item, dict):
            continue
        usability = _score_generated_asset_result(
            item,
            generation_mode=generation_mode,
            cover_style_type=cover_style_type,
            inner_style_type=inner_style_type,
            style_preset=style_preset,
            provider=provider,
            image_count=image_count,
        )
        annotated.append({
            **item,
            **usability,
        })
    return annotated


def _summarize_generated_asset_results(items):
    rows = [item for item in (items or []) if isinstance(item, dict)]
    cover_rows = sorted(
        [item for item in rows if item.get('recommended_usage') == 'cover'],
        key=lambda item: ((item.get('usable_score') or 0), -(item.get('index') or 999)),
        reverse=True,
    )
    inner_rows = sorted(
        [item for item in rows if item.get('recommended_usage') == 'inner'],
        key=lambda item: ((item.get('usable_score') or 0), -(item.get('index') or 999)),
        reverse=True,
    )
    return {
        'result_count': len(rows),
        'best_cover_label': cover_rows[0].get('usable_label') if cover_rows else '',
        'best_cover_score': cover_rows[0].get('usable_score') if cover_rows else 0,
        'best_cover_index': cover_rows[0].get('index') if cover_rows else None,
        'best_inner_label': inner_rows[0].get('usable_label') if inner_rows else '',
        'best_inner_score': inner_rows[0].get('usable_score') if inner_rows else 0,
        'best_inner_index': inner_rows[0].get('index') if inner_rows else None,
    }


def _image_provider_healthcheck(payload=None, timeout_seconds=15):
    capabilities = _resolve_image_provider_capabilities(payload)
    provider = (capabilities.get('image_provider_name') or 'svg_fallback').strip() or 'svg_fallback'
    payload = payload or {}
    if provider == 'svg_fallback':
        return {
            'enabled': False,
            'ok': False,
            'message': '当前图片模式为 SVG 兜底，未启用远端图片接口',
            'provider': provider,
            'request_preview': _build_asset_provider_request_preview(
                provider,
                '',
                (payload.get('prompt_text') or '').strip(),
                capabilities.get('image_provider_size') or '1024x1536',
                capabilities.get('image_default_style_type') or 'medical_science',
                image_count=min(max(_safe_int(payload.get('image_count'), 1), 1), 4),
            ),
            'response': None,
            'normalized_preview': [],
        }

    api_url = (capabilities.get('image_provider_api_url') or '').strip()
    api_key = _resolve_image_provider_api_key(provider)
    if not api_url:
        return {
            'enabled': True,
            'ok': False,
            'message': '未配置图片 API URL',
            'provider': provider,
            'request_preview': {},
            'response': None,
            'normalized_preview': [],
        }
    if not api_key:
        return {
            'enabled': True,
            'ok': False,
            'message': '未配置图片 API Key',
            'provider': provider,
            'request_preview': {},
            'response': None,
            'normalized_preview': [],
        }

    prompt_text = (
        payload.get('prompt_text')
        or '生成一张适合小红书医疗科普封面的测试图片，画面简洁，标题区留白。'
    ).strip()
    title_hint = (payload.get('title_hint') or '测试图片接口').strip()[:200] or '测试图片接口'
    image_count = min(max(_safe_int(payload.get('image_count'), 1), 1), 4)
    request_preview = _build_asset_provider_request_preview(
        provider,
        capabilities.get('image_provider_model'),
        prompt_text,
        capabilities.get('image_provider_size'),
        capabilities.get('image_default_style_type') or 'medical_science',
        image_count=image_count,
    )
    if provider == 'volcengine_ark' and _is_ark_seedream_model(capabilities.get('image_provider_model')):
        request_preview['size'] = '2K'
        request_preview['stream'] = False
        request_preview['watermark'] = True
        request_preview['sequential_image_generation'] = 'disabled'
    base_timeout = min(max(_safe_int(timeout_seconds, 15), 5), 90)
    if provider in {'volcengine_ark', 'volcengine_las'}:
        base_timeout = max(base_timeout, 30)
    try:
        response = None
        last_exc = None
        timeout_attempts = [base_timeout]
        if provider in {'volcengine_ark', 'volcengine_las'}:
            timeout_attempts.append(min(base_timeout + 20, 90))
        for current_timeout in timeout_attempts:
            try:
                response = requests.post(
                    api_url,
                    json=request_preview,
                    headers={
                        'Authorization': f'Bearer {api_key}',
                        'Content-Type': 'application/json',
                    },
                    timeout=current_timeout,
                )
                response.raise_for_status()
                last_exc = None
                break
            except requests.exceptions.ReadTimeout as exc:
                last_exc = exc
                response = None
                continue
        if last_exc:
            raise last_exc
        if response is None:
            raise RuntimeError('图片接口未返回响应')
        payload_json = response.json()
        response_preview = payload_json if not isinstance(payload_json, dict) else dict(payload_json)
        if isinstance(response_preview, dict):
            for key in ['data', 'images', 'output', 'results']:
                if isinstance(response_preview.get(key), list):
                    response_preview[key] = response_preview[key][:3]
        normalized_preview = _normalize_asset_provider_results(payload_json, provider=provider, image_prompt=prompt_text, title_hint=title_hint)[:3]
        return {
            'enabled': True,
            'ok': True,
            'message': f'图片接口可用，provider={provider}',
            'provider': provider,
            'request_preview': request_preview,
            'response': response_preview,
            'normalized_preview': normalized_preview,
        }
    except Exception as exc:
        return {
            'enabled': True,
            'ok': False,
            'message': f'图片接口不可用：{exc}',
            'provider': provider,
            'request_preview': request_preview,
            'response': None,
            'normalized_preview': [],
        }


def _build_dashboard_stats(activity_id, args):
    range_info = _resolve_analysis_range(args)
    topics, registrations, submissions = _load_activity_scope(activity_id, range_info)
    topic_map = {topic.id: topic.topic_name for topic in topics}
    reg_by_id = {reg.id: reg for reg in registrations}

    total_participants = len(registrations)
    total_published_ids = {sub.registration_id for sub in submissions if _submission_has_any_link(sub)}
    total_published = len(total_published_ids)

    total_likes = total_favorites = total_comments = total_views = 0
    platform_published = {}
    platform_stats = {}

    for platform_key, label in PLATFORM_DEFINITIONS:
        platform_submissions = [sub for sub in submissions if _submission_has_platform_link(sub, platform_key)]
        participant_ids = {sub.registration_id for sub in platform_submissions}
        views = sum(getattr(sub, f'{platform_key}_views', 0) or 0 for sub in platform_submissions)
        likes = sum(getattr(sub, f'{platform_key}_likes', 0) or 0 for sub in platform_submissions)
        favorites = sum(getattr(sub, f'{platform_key}_favorites', 0) or 0 for sub in platform_submissions)
        comments = sum(getattr(sub, f'{platform_key}_comments', 0) or 0 for sub in platform_submissions)
        interactions = likes + favorites + comments
        platform_published[platform_key] = len(platform_submissions)
        platform_stats[platform_key] = {
            'label': label,
            'participants': len(participant_ids),
            'published_count': len(platform_submissions),
            'views': views,
            'likes': likes,
            'favorites': favorites,
            'comments': comments,
            'interactions': interactions,
            'interaction_rate': _calculate_rate(interactions, views),
            'interaction_rate_display': _format_rate(_calculate_rate(interactions, views)),
            'penetration_rate': _calculate_rate(len(platform_submissions), total_participants),
            'penetration_rate_display': _format_rate(_calculate_rate(len(platform_submissions), total_participants)),
        }
        total_views += views
        total_likes += likes
        total_favorites += favorites
        total_comments += comments

    total_interactions = total_likes + total_favorites + total_comments
    publish_rate = _calculate_rate(total_published, total_participants)

    group_stats = {}
    for reg in registrations:
        group = reg.group_num or '未分组'
        if group not in group_stats:
            group_stats[group] = {'count': 0, 'published': 0}
        group_stats[group]['count'] += 1
        if reg.id in total_published_ids:
            group_stats[group]['published'] += 1

    group_participants = [
        {'group': group, 'participants': data['count']}
        for group, data in sorted(group_stats.items(), key=lambda item: (-item[1]['count'], item[0]))
    ]

    topic_stats = []
    for topic in topics:
        count = len([reg for reg in registrations if reg.topic_id == topic.id])
        topic_stats.append({'name': topic.topic_name, 'count': count})
    topic_stats.sort(key=lambda item: item['count'], reverse=True)

    content_type_stats = {}
    content_type_perf = {}
    for sub in submissions:
        content_type = (sub.content_type or '').strip() or '未识别'
        metrics = _collect_platform_metrics(sub, 'all')
        content_type_stats[content_type] = content_type_stats.get(content_type, 0) + 1
        current = content_type_perf.setdefault(content_type, {'count': 0, 'views': 0, 'interactions': 0})
        current['count'] += 1
        current['views'] += metrics['views']
        current['interactions'] += metrics['interactions']

    best_content_type = None
    for row in content_type_perf.values():
        row['avg_interactions'] = round(row['interactions'] / row['count'], 2) if row['count'] else 0
        row['interaction_rate'] = _calculate_rate(row['interactions'], row['views'])
        row['interaction_rate_display'] = _format_rate(row['interaction_rate'])

    if content_type_perf:
        best_content_type = sorted(
            content_type_perf.items(),
            key=lambda item: ((item[1]['interaction_rate'] or 0), item[1]['avg_interactions']),
            reverse=True
        )[0][0]

    strategy_insights = _build_strategy_insights(submissions)
    task_funnel = _build_task_funnel_payload(registrations, submissions)

    personal_rankings = _build_personal_rankings(registrations, submissions)
    group_rankings = {
        'all_platform': _build_group_rankings(registrations, submissions, 'all'),
        'xhs': _build_group_rankings(registrations, submissions, 'xhs'),
        'douyin': _build_group_rankings(registrations, submissions, 'douyin'),
        'video': _build_group_rankings(registrations, submissions, 'video'),
        'weibo': _build_group_rankings(registrations, submissions, 'weibo'),
    }

    group_completion = []
    for group, data in group_stats.items():
        completion_rate = _calculate_rate(data['published'], data['count'])
        group_completion.append({
            'group': group,
            'count': data['count'],
            'published': data['published'],
            'completion_rate': completion_rate or 0,
        })
    group_completion.sort(key=lambda item: item['completion_rate'], reverse=True)

    viral_notes = []
    for sub in submissions:
        if not _submission_has_platform_link(sub, 'xhs'):
            continue
        reg = reg_by_id.get(sub.registration_id) or sub.registration
        xhs_interactions = (sub.xhs_likes or 0) + (sub.xhs_favorites or 0) + (sub.xhs_comments or 0)
        viral_notes.append({
            'name': reg.name if reg else '未命名',
            'group': (reg.group_num if reg else '') or '未分组',
            'topic': reg.topic.topic_name if reg and reg.topic else '',
            'content_type': (sub.content_type or '未识别'),
            'xhs_link': sub.xhs_link,
            'xhs_views': sub.xhs_views or 0,
            'xhs_likes': sub.xhs_likes or 0,
            'xhs_favorites': sub.xhs_favorites or 0,
            'xhs_comments': sub.xhs_comments or 0,
            'xhs_interactions': xhs_interactions,
        })
    viral_notes.sort(key=lambda item: (item['xhs_interactions'], item['xhs_views']), reverse=True)

    type_note_recommendations = {}
    for note in viral_notes:
        content_type = note['content_type'] or '未识别'
        type_note_recommendations.setdefault(content_type, [])
        if len(type_note_recommendations[content_type]) < 3:
            type_note_recommendations[content_type].append(note)

    top_keyword_trends = []
    recent_trends = TrendNote.query.order_by(TrendNote.created_at.desc()).limit(200).all()
    trend_keyword_scores = defaultdict(int)
    for note in recent_trends:
        for keyword in _extract_keywords_from_note(note):
            trend_keyword_scores[keyword] += _trend_score(note)
    for keyword, score in sorted(trend_keyword_scores.items(), key=lambda item: item[1], reverse=True)[:5]:
        top_keyword_trends.append({'keyword': keyword, 'score': score})

    note_improvement_suggestions = []
    if publish_rate is not None and publish_rate < 60:
        note_improvement_suggestions.append('先补齐提交率，建议增加组长催更、截止提醒和二次提交入口。')
    if best_content_type:
        note_improvement_suggestions.append(f'当前最佳内容类型是“{best_content_type}”，建议沉淀成固定模板。')
    if strategy_insights['captured_count'] and strategy_insights['title_skill_rows']:
        note_improvement_suggestions.append(
            f"当前标题打法里表现更靠前的是“{strategy_insights['title_skill_rows'][0]['label']}”，可以优先放大。"
        )
    if strategy_insights['capture_rate'] < 60:
        note_improvement_suggestions.append('当前爆款策略留痕率还不高，建议生成后尽量用系统标题池和封面推荐池。')
    low_groups = [row['group'] for row in group_completion if row['completion_rate'] < 50]
    if low_groups:
        note_improvement_suggestions.append('低完成率小组：' + '、'.join(low_groups) + '，建议安排小组共创和复盘。')
    if not note_improvement_suggestions:
        note_improvement_suggestions.append('整体运行平稳，可以继续做标题钩子和封面形式的A/B测试。')

    next_topic_suggestions = []
    if top_keyword_trends:
        next_topic_suggestions.extend([f"围绕“{row['keyword']}”做多角度选题和素材分发" for row in top_keyword_trends[:3]])
    if best_content_type in ['检查解读型', '复查管理型', '轻科普问答型']:
        next_topic_suggestions.append('继续加大检查解读和复查管理类话题占比。')
    if best_content_type in ['真实经历型', '场景种草型']:
        next_topic_suggestions.append('增加患者/家属/体检人群多角色叙事，提高互动回复率。')
    next_topic_suggestions = next_topic_suggestions[:5]

    definitions = [
        '总参与人数 = 报名人数',
        '总发布条数 = 任一平台有链接算1条',
        '平台发布条数 = 该平台有链接',
        '总互动 = 点赞 + 收藏 + 评论',
        '互动率分母为曝光量，曝光量=0时显示“-”',
        '时间筛选按参与活跃时间统计：报名时间或提交更新时间命中筛选区间即纳入统计',
    ]

    operational_advice = _build_operational_advice_payload(
        publish_rate=publish_rate or 0,
        best_content_type=best_content_type,
        strategy_insights=strategy_insights,
        task_funnel=task_funnel,
        group_completion=group_completion,
        top_keyword_trends=top_keyword_trends,
        platform_stats=platform_stats,
        note_improvement_suggestions=note_improvement_suggestions,
        next_topic_suggestions=next_topic_suggestions,
    )

    return {
        'activity_id': activity_id,
        'range': {
            'key': range_info['key'],
            'label': range_info['label'],
            'start_date': range_info['start_date'],
            'end_date': range_info['end_date'],
        },
        'definitions': definitions,
        'overview': {
            'total_participants': total_participants,
            'group_participants': group_participants,
            'total_published': total_published,
            'publish_rate': publish_rate,
            'publish_rate_display': _format_rate(publish_rate),
            'platform_publish_counts': platform_published,
            'platform_penetration': {
                key: {
                    'rate': platform_stats[key]['penetration_rate'],
                    'display': platform_stats[key]['penetration_rate_display'],
                } for key, _ in PLATFORM_DEFINITIONS
            },
        },
        'platforms': platform_stats,
        'group_rankings': group_rankings,
        'personal_rankings': personal_rankings,
        'top_keyword_trends': top_keyword_trends,
        'total_registrations': total_participants,
        'total_published': total_published,
        'total_likes': total_likes,
        'total_favorites': total_favorites,
        'total_comments': total_comments,
        'total_views': total_views,
        'total_interactions': total_interactions,
        'group_stats': group_stats,
        'topic_stats': topic_stats,
        'content_type_stats': content_type_stats,
        'content_type_performance': content_type_perf,
        'best_content_type': best_content_type,
        'strategy_insights': strategy_insights,
        'task_funnel': task_funnel,
        'operational_advice': operational_advice,
        'group_completion': group_completion,
        'viral_notes': viral_notes[:20],
        'type_note_recommendations': type_note_recommendations,
        'note_improvement_suggestions': note_improvement_suggestions,
        'next_topic_suggestions': next_topic_suggestions,
        'top_30': personal_rankings['all_platform'][:30],
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


def _build_report_markdown(activity, stats, report_type='weekly'):
    report_type = (report_type or 'weekly').strip()
    report_title_map = {
        'weekly': '周报',
        'monthly': '月报',
        'review': '活动复盘',
    }
    report_label = report_title_map.get(report_type, '报告')
    group_lines = [
        f"- {row['group']}：参与{row['count']}，已发布{row['published']}，发布率{round(row['completion_rate'], 2)}%"
        for row in stats['group_completion']
    ]
    top_lines = [
        f"{idx}. {row['name']}｜{row['group']}｜互动{row['interactions']}｜{row['topic']}"
        for idx, row in enumerate(stats['personal_rankings']['all_platform'][:20], 1)
    ]
    platform_lines = [
        f"- {platform['label']}：参与人数{platform['participants']}，发布条数{platform['published_count']}，曝光量{platform['views']}，互动率{platform['interaction_rate_display']}"
        for platform in stats['platforms'].values()
    ]
    type_line = '、'.join([
        f"{name}{count}" for name, count in sorted(
            stats['content_type_stats'].items(),
            key=lambda item: item[1],
            reverse=True
        )
    ]) if stats['content_type_stats'] else '暂无'
    task_funnel_lines = [
        f"- {row['label']}：{row['count']}（{row['rate_display']}）｜{row['note']}"
        for row in (stats.get('task_funnel', {}).get('steps') or [])
    ]
    strategy_summary_lines = [
        f"- {line}"
        for line in ((stats.get('strategy_insights') or {}).get('summary_lines') or [])
    ]
    operational_advice = stats.get('operational_advice') or {}
    operational_lines = {
        'urgent': [f"- {line}" for line in (operational_advice.get('urgent_actions') or [])],
        'winning': [f"- {line}" for line in (operational_advice.get('winning_moves') or [])],
        'risk': [f"- {line}" for line in (operational_advice.get('risk_alerts') or [])],
        'next': [f"- {line}" for line in (operational_advice.get('next_week_actions') or [])],
    }
    keyword_lines = [
        f"- {row['keyword']}：热度分 {row['score']}"
        for row in stats.get('top_keyword_trends', [])[:8]
    ]
    viral_lines = [
        f"- {row['name']}｜{row['topic']}｜互动{row['xhs_interactions']}｜阅读{row['xhs_views']}"
        for row in stats.get('viral_notes', [])[:10]
    ]

    sections = [
        f"# {activity.name} {report_label}（{activity.title}）",
        '',
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"统计区间：{stats['range']['label']} {stats['range']['start_date'] or ''} {('~ ' + stats['range']['end_date']) if stats['range']['end_date'] else ''}",
        '',
        '## 一、固定口径',
        *[f"- {line}" for line in stats['definitions']],
        '',
        '## 二、总览区',
        f"- 总参与人数：{stats['overview']['total_participants']}",
        f"- 总发布条数：{stats['overview']['total_published']}",
        f"- 总体发布率：{stats['overview']['publish_rate_display']}",
        f"- 总曝光量：{stats['total_views']}",
        f"- 总互动：{stats['total_interactions']}（点赞{stats['total_likes']} + 收藏{stats['total_favorites']} + 评论{stats['total_comments']}）",
        f"- 当前判断：{operational_advice.get('headline') or '暂无'}",
        '',
        '## 三、任务漏斗',
        *(task_funnel_lines or ['- 暂无任务漏斗数据']),
        '',
        '## 四、策略结论',
        *(strategy_summary_lines or ['- 暂无策略结论']),
        '',
        '## 五、运营建议',
        '### 当前最急',
        *(operational_lines['urgent'] or ['- 暂无']),
        '',
        '### 当前该放大',
        *(operational_lines['winning'] or ['- 暂无']),
        '',
        '### 当前风险',
        *(operational_lines['risk'] or ['- 暂无']),
        '',
        '### 下周动作',
        *(operational_lines['next'] or ['- 暂无']),
        '',
        '## 六、平台分层',
        *(platform_lines or ['- 暂无平台数据']),
        '',
        '## 七、小组排名',
        *(group_lines or ['- 暂无小组数据']),
        '',
        '## 八、优秀个人TOP20（小红书+抖音+视频号）',
        *(top_lines or ['暂无数据']),
        '',
        '## 九、内容类型分布',
        f"- {type_line}",
        f"- 当前最佳内容类型：{stats['best_content_type'] or '暂无'}",
        '',
    ]

    if report_type in {'monthly', 'review'}:
        sections.extend([
            '## 十、热点关键词趋势',
            *(keyword_lines or ['- 暂无热点关键词趋势']),
            '',
            '## 十一、爆款笔记摘要',
            *(viral_lines or ['- 暂无爆款摘要']),
            '',
        ])
    else:
        sections.extend([
            '## 十、优化建议',
            *[f"- {line}" for line in stats['note_improvement_suggestions']],
            '',
            '## 十一、下期选题建议',
            *[f"- {line}" for line in stats['next_topic_suggestions']],
            '',
        ])

    if report_type == 'monthly':
        sections.extend([
            '## 十二、月度结论',
            f"- 本月最佳内容类型：{stats['best_content_type'] or '暂无'}",
            f"- 本月热点趋势重点：{('、'.join([row['keyword'] for row in stats.get('top_keyword_trends', [])[:3]]) or '暂无')}",
            '- 建议围绕高互动内容类型和热点关键词同步优化标题、封面与发布时间。',
            '',
            '## 十三、下月建议',
            *[f"- {line}" for line in stats['next_topic_suggestions']],
            '',
        ])

    if report_type == 'review':
        sections.extend([
            '## 十二、活动复盘亮点',
            f"- 最佳内容类型：{stats['best_content_type'] or '暂无'}",
            f"- 已发布话题数：{stats['total_published']}",
            f"- 累计热点趋势关键词：{('、'.join([row['keyword'] for row in stats.get('top_keyword_trends', [])[:5]]) or '暂无')}",
            '',
            '## 十三、问题与改进',
            *[f"- {line}" for line in stats['note_improvement_suggestions']],
            '',
            '## 十四、下期建议',
            *[f"- {line}" for line in stats['next_topic_suggestions']],
            '',
        ])

    return '\n'.join(sections)


def _build_public_shell_context():
    page_config = _get_site_page_config('home')
    theme = _get_active_site_theme()
    site_config = _serialize_site_page_config(page_config) if page_config else {
        **DEFAULT_HOME_PAGE_CONFIG,
        'nav_items': [dict(item) for item in DEFAULT_SITE_NAV_ITEMS],
    }
    site_theme = _serialize_site_theme(theme) if theme else dict(DEFAULT_SITE_THEME)
    return {
        'site_config': site_config,
        'site_theme': site_theme,
        'release_manifest': _build_release_manifest_payload(),
    }

# ==================== 路由 ====================


@app.route('/api/admin/site-config', methods=['GET', 'POST'])
def admin_site_config():
    guard = _admin_json_guard()
    if guard:
        return guard

    created_defaults = False
    page_config = _get_site_page_config('home')
    if not page_config:
        page_config = SitePageConfig(
            page_key=DEFAULT_HOME_PAGE_CONFIG['page_key'],
            site_name=DEFAULT_HOME_PAGE_CONFIG['site_name'],
            page_title=DEFAULT_HOME_PAGE_CONFIG['page_title'],
            hero_badge=DEFAULT_HOME_PAGE_CONFIG['hero_badge'],
            hero_title=DEFAULT_HOME_PAGE_CONFIG['hero_title'],
            hero_subtitle=DEFAULT_HOME_PAGE_CONFIG['hero_subtitle'],
            announcement_title=DEFAULT_HOME_PAGE_CONFIG['announcement_title'],
            trend_title=DEFAULT_HOME_PAGE_CONFIG['trend_title'],
            primary_section_title=DEFAULT_HOME_PAGE_CONFIG['primary_section_title'],
            primary_section_icon=DEFAULT_HOME_PAGE_CONFIG['primary_section_icon'],
            secondary_section_title=DEFAULT_HOME_PAGE_CONFIG['secondary_section_title'],
            secondary_section_icon=DEFAULT_HOME_PAGE_CONFIG['secondary_section_icon'],
            primary_topic_limit=DEFAULT_HOME_PAGE_CONFIG['primary_topic_limit'],
            footer_text=DEFAULT_HOME_PAGE_CONFIG['footer_text'],
            nav_items=json.dumps(DEFAULT_SITE_NAV_ITEMS, ensure_ascii=False),
        )
        db.session.add(page_config)
        created_defaults = True

    theme = _get_active_site_theme()
    if not theme:
        theme = SiteTheme(theme_key=DEFAULT_SITE_THEME['theme_key'], name=DEFAULT_SITE_THEME['name'], is_active=True)
        for field, default_value in DEFAULT_SITE_THEME.items():
            if field in {'theme_key', 'name'}:
                continue
            setattr(theme, field, default_value)
        db.session.add(theme)
        created_defaults = True
    if created_defaults:
        db.session.commit()
    elif page_config.id is None or theme.id is None:
        db.session.flush()

    if request.method == 'POST':
        payload = request.json or {}
        config_data = payload.get('page_config') if isinstance(payload.get('page_config'), dict) else payload
        theme_data = payload.get('theme') if isinstance(payload.get('theme'), dict) else {}

        page_config.site_name = (config_data.get('site_name') or DEFAULT_HOME_PAGE_CONFIG['site_name']).strip()[:100]
        page_config.page_title = (config_data.get('page_title') or DEFAULT_HOME_PAGE_CONFIG['page_title']).strip()[:200]
        page_config.hero_badge = (config_data.get('hero_badge') or '').strip()[:100]
        page_config.hero_title = (config_data.get('hero_title') or '').strip()[:200]
        page_config.hero_subtitle = (config_data.get('hero_subtitle') or '').strip()
        page_config.announcement_title = (config_data.get('announcement_title') or DEFAULT_HOME_PAGE_CONFIG['announcement_title']).strip()[:100]
        page_config.trend_title = (config_data.get('trend_title') or DEFAULT_HOME_PAGE_CONFIG['trend_title']).strip()[:100]
        page_config.primary_section_title = (config_data.get('primary_section_title') or DEFAULT_HOME_PAGE_CONFIG['primary_section_title']).strip()[:100]
        page_config.primary_section_icon = (config_data.get('primary_section_icon') or DEFAULT_HOME_PAGE_CONFIG['primary_section_icon']).strip()[:50]
        page_config.secondary_section_title = (config_data.get('secondary_section_title') or DEFAULT_HOME_PAGE_CONFIG['secondary_section_title']).strip()[:100]
        page_config.secondary_section_icon = (config_data.get('secondary_section_icon') or DEFAULT_HOME_PAGE_CONFIG['secondary_section_icon']).strip()[:50]
        page_config.primary_topic_limit = _normalize_quota(
            config_data.get('primary_topic_limit'),
            default=DEFAULT_HOME_PAGE_CONFIG['primary_topic_limit'],
            min_value=1,
            max_value=120,
        )
        page_config.footer_text = (config_data.get('footer_text') or DEFAULT_HOME_PAGE_CONFIG['footer_text']).strip()[:200]
        page_config.nav_items = json.dumps(_normalize_nav_items(config_data.get('nav_items')), ensure_ascii=False)

        theme.name = (theme_data.get('name') or theme.name or DEFAULT_SITE_THEME['name']).strip()[:100]
        for field, default_value in DEFAULT_SITE_THEME.items():
            if field in {'theme_key', 'name'}:
                continue
            if field == 'font_family':
                setattr(theme, field, (theme_data.get(field) or theme.font_family or default_value).strip()[:200])
                continue
            if field == 'footer_text':
                setattr(theme, field, (theme_data.get(field) or page_config.footer_text or default_value).strip()[:200])
                continue
            setattr(theme, field, _normalize_hex_color(theme_data.get(field), getattr(theme, field) or default_value))
        theme.is_active = True
        SiteTheme.query.filter(SiteTheme.id != theme.id).update({'is_active': False}, synchronize_session=False)

        db.session.flush()
        _log_operation('save_site_config', 'site_page_config', target_id=page_config.id, message='更新网站门户配置', detail={
            'page_key': page_config.page_key,
            'theme_id': theme.id,
            'site_name': page_config.site_name,
            'nav_count': len(_normalize_nav_items(config_data.get('nav_items'))),
        })
        db.session.commit()

    return jsonify({
        'success': True,
        'page_config': _serialize_site_page_config(page_config),
        'theme': _serialize_site_theme(theme),
    })


@app.route('/api/admin/announcements', methods=['GET', 'POST'])
def admin_announcements():
    guard = _admin_json_guard()
    if guard:
        return guard

    if request.method == 'POST':
        payload = request.json or {}
        announcement_id = _safe_int(payload.get('id'), 0)
        if announcement_id > 0:
            announcement = Announcement.query.get_or_404(announcement_id)
        else:
            announcement = Announcement()
            db.session.add(announcement)

        title = (payload.get('title') or '').strip()
        content = (payload.get('content') or '').strip()
        if not title:
            return jsonify({'success': False, 'message': '公告标题不能为空'})
        if not content:
            return jsonify({'success': False, 'message': '公告内容不能为空'})

        status = (payload.get('status') or 'draft').strip()
        if status not in {'draft', 'active', 'archived'}:
            return jsonify({'success': False, 'message': '不支持的公告状态'})

        announcement.title = title[:200]
        announcement.content = content
        announcement.link_url = (payload.get('link_url') or '').strip()[:500]
        announcement.button_text = (payload.get('button_text') or '').strip()[:50]
        announcement.priority = max(_safe_int(payload.get('priority'), 100), 0)
        announcement.status = status
        announcement.starts_at = _parse_datetime(payload.get('starts_at'))
        announcement.ends_at = _parse_datetime(payload.get('ends_at'))

        db.session.flush()
        _log_operation('save_announcement', 'announcement', target_id=announcement.id, message='保存网站公告', detail={
            'title': announcement.title,
            'status': announcement.status,
            'priority': announcement.priority,
        })
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '公告已保存',
            'item': _serialize_announcement(announcement),
        })

    status = (request.args.get('status') or '').strip()
    query = Announcement.query
    if status:
        query = query.filter_by(status=status)
    items = query.order_by(Announcement.priority.asc(), Announcement.updated_at.desc(), Announcement.id.desc()).all()
    return jsonify({
        'success': True,
        'items': [_serialize_announcement(item) for item in items]
    })


@app.route('/api/admin/announcements/<int:announcement_id>/status', methods=['POST'])
def update_announcement_status(announcement_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    payload = request.json or {}
    status = (payload.get('status') or '').strip()
    if status not in {'draft', 'active', 'archived'}:
        return jsonify({'success': False, 'message': '不支持的公告状态'})

    announcement = Announcement.query.get_or_404(announcement_id)
    announcement.status = status
    _log_operation('update_status', 'announcement', target_id=announcement.id, message='更新公告状态', detail={
        'title': announcement.title,
        'status': status,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': '公告状态已更新',
        'item': _serialize_announcement(announcement),
    })


# AI生成文案API - 支持可切换的 OpenAI 兼容模型
import requests

DEEPSEEK_API_KEY = os.environ.get('DEEPSEEK_API_KEY', '')
DEEPSEEK_API_URL = os.environ.get('DEEPSEEK_API_URL', 'https://api.deepseek.com/chat/completions')
COPYWRITER_API_KEY = os.environ.get('COPYWRITER_API_KEY', '').strip()
COPYWRITER_API_URL = os.environ.get('COPYWRITER_API_URL', '').strip()
COPYWRITER_MODEL = os.environ.get('COPYWRITER_MODEL', '').strip()
OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', '').strip()
OPENAI_API_URL = os.environ.get('OPENAI_API_URL', '').strip()
OPENAI_MODEL = os.environ.get('OPENAI_MODEL', '').strip()
DOUBAO_API_KEY = os.environ.get('DOUBAO_API_KEY', '').strip()
DOUBAO_API_URL = os.environ.get('DOUBAO_API_URL', '').strip()
DOUBAO_MODEL = os.environ.get('DOUBAO_MODEL', '').strip()
YUANBAO_API_KEY = os.environ.get('YUANBAO_API_KEY', '').strip()
YUANBAO_API_URL = os.environ.get('YUANBAO_API_URL', '').strip()
YUANBAO_MODEL = os.environ.get('YUANBAO_MODEL', '').strip()
HUNYUAN_API_KEY = os.environ.get('HUNYUAN_API_KEY', '').strip()
HUNYUAN_API_URL = os.environ.get('HUNYUAN_API_URL', '').strip()
HUNYUAN_MODEL = os.environ.get('HUNYUAN_MODEL', '').strip()


def _normalize_copywriter_api_url(raw_url='', provider='', model=''):
    url = (raw_url or '').strip()
    if not url:
        return ''
    inferred_provider = provider or _infer_copywriter_provider(url, model)
    if '/responses' in url or '/chat/completions' in url:
        return url
    if '/chat/completions' in url:
        return url
    url = url.rstrip('/')
    if inferred_provider == 'deepseek':
        if url.endswith('/v1'):
            return f'{url}/chat/completions'
        return f'{url}/chat/completions'
    if inferred_provider in {'doubao', 'tencent_hunyuan'}:
        if url.endswith('/api/v3'):
            return f'{url}/responses'
        return f'{url}/responses'
    if url.endswith('/v1'):
        return f'{url}/chat/completions'
    if url.endswith('/v1/chat'):
        return f'{url}/completions'
    return f'{url}/v1/chat/completions'


def _infer_copywriter_provider(api_url='', model=''):
    lowered = ' '.join([(api_url or '').strip().lower(), (model or '').strip().lower()])
    if 'deepseek' in lowered:
        return 'deepseek'
    if any(token in lowered for token in ['doubao', 'volces', 'ark.cn-beijing']):
        return 'doubao'
    if any(token in lowered for token in ['hunyuan', 'yuanbao', 'tencent']):
        return 'tencent_hunyuan'
    if 'openai.com' in lowered or any(token in lowered for token in ['gpt-', 'o1', 'o3', 'o4']):
        return 'openai_compatible'
    return 'custom_openai_compatible'


def _is_deepseek_chat_model(model_name=''):
    return str(model_name or '').strip().lower() == 'deepseek-chat'


def _is_deepseek_reasoning_model(model_name=''):
    lowered = str(model_name or '').strip().lower()
    return lowered == 'deepseek-reasoner' or lowered.endswith('-reasoner')


def _is_deepseek_v4_model(model_name=''):
    lowered = str(model_name or '').strip().lower()
    return lowered.startswith('deepseek-v4')


def _default_copywriter_model(provider='', api_url=''):
    inferred_provider = (provider or _infer_copywriter_provider(api_url, '')).strip()
    if inferred_provider == 'deepseek':
        return 'deepseek-v4-pro'
    if inferred_provider == 'openai_compatible':
        return 'gpt-5.4'
    return ''


def _copywriter_thinking_enabled(runtime=None):
    runtime = runtime or {}
    provider = (runtime.get('provider') or '').strip()
    model_name = runtime.get('model') or ''
    return provider == 'deepseek' and (
        _is_deepseek_v4_model(model_name)
        or _is_deepseek_reasoning_model(model_name)
        or _is_deepseek_chat_model(model_name)
    )


def _resolve_copywriter_api_key(api_url='', model=''):
    provider = _infer_copywriter_provider(api_url, model)
    if provider == 'deepseek':
        return (DEEPSEEK_API_KEY or COPYWRITER_API_KEY or '').strip()
    if provider == 'doubao':
        return (DOUBAO_API_KEY or COPYWRITER_API_KEY or '').strip()
    if provider == 'tencent_hunyuan':
        return (YUANBAO_API_KEY or HUNYUAN_API_KEY or COPYWRITER_API_KEY or '').strip()
    if provider == 'openai_compatible':
        return (OPENAI_API_KEY or COPYWRITER_API_KEY or '').strip()
    return (COPYWRITER_API_KEY or OPENAI_API_KEY or DEEPSEEK_API_KEY or DOUBAO_API_KEY or YUANBAO_API_KEY or HUNYUAN_API_KEY or '').strip()


def _build_copywriter_runtime_entry(api_url='', model='', *, source='runtime_config', api_key=''):
    provider = _infer_copywriter_provider(api_url, model)
    normalized_url = _normalize_copywriter_api_url(api_url, provider=provider, model=model)
    normalized_model = (model or '').strip() or _default_copywriter_model(provider=provider, api_url=normalized_url)
    resolved_api_key = (api_key or '').strip() or _resolve_copywriter_api_key(normalized_url, normalized_model)
    if not normalized_url or not normalized_model or not resolved_api_key:
        return {}
    provider_label = {
        'deepseek': f'DeepSeek：{normalized_model}',
        'doubao': f'豆包：{normalized_model}',
        'tencent_hunyuan': f'腾讯混元：{normalized_model}',
        'openai_compatible': f'OpenAI兼容模型：{normalized_model}',
        'custom_openai_compatible': f'可切换模型：{normalized_model}',
    }.get(provider, normalized_model)
    if provider == 'deepseek' and _copywriter_thinking_enabled({'provider': provider, 'model': normalized_model}):
        provider_label = f'DeepSeek：{normalized_model}（思考模式）'
    return {
        'configured': True,
        'provider': provider,
        'api_key': resolved_api_key,
        'api_url': normalized_url,
        'model': normalized_model,
        'label': provider_label,
        'source': source,
        'thinking_mode': _copywriter_thinking_enabled({'provider': provider, 'model': normalized_model}),
    }


def _copywriter_runtime_candidates(payload=None):
    runtime_config = _automation_runtime_config()
    payload = payload or {}
    primary_url = (payload.get('copywriter_api_url') or runtime_config.get('copywriter_api_url') or '').strip()
    primary_key = (payload.get('copywriter_api_key') or runtime_config.get('copywriter_api_key') or '').strip()
    primary_model = (payload.get('copywriter_model') or runtime_config.get('copywriter_model') or '').strip()
    backup_url = (payload.get('copywriter_backup_api_url') or runtime_config.get('copywriter_backup_api_url') or '').strip()
    backup_key = (payload.get('copywriter_backup_api_key') or runtime_config.get('copywriter_backup_api_key') or '').strip()
    backup_model = (payload.get('copywriter_backup_model') or runtime_config.get('copywriter_backup_model') or '').strip()
    third_url = (payload.get('copywriter_third_api_url') or runtime_config.get('copywriter_third_api_url') or '').strip()
    third_key = (payload.get('copywriter_third_api_key') or runtime_config.get('copywriter_third_api_key') or '').strip()
    third_model = (payload.get('copywriter_third_model') or runtime_config.get('copywriter_third_model') or '').strip()

    candidates = []
    seen = set()

    def append_candidate(api_url, model, source, api_key=''):
        entry = _build_copywriter_runtime_entry(api_url, model, source=source, api_key=api_key)
        if not entry:
            return
        dedupe_key = (entry['api_url'], entry['model'])
        if dedupe_key in seen:
            return
        seen.add(dedupe_key)
        candidates.append(entry)

    append_candidate(primary_url, primary_model, 'primary', primary_key)
    append_candidate(backup_url, backup_model, 'backup', backup_key)
    append_candidate(third_url, third_model, 'third', third_key)
    append_candidate(
        COPYWRITER_API_URL,
        COPYWRITER_MODEL or primary_model or _default_copywriter_model(api_url=COPYWRITER_API_URL or primary_url),
        'env_copywriter',
    )
    append_candidate(OPENAI_API_URL or 'https://api.openai.com/v1', OPENAI_MODEL or 'gpt-5.4', 'env_openai')
    append_candidate(DOUBAO_API_URL, DOUBAO_MODEL, 'env_doubao')
    append_candidate(YUANBAO_API_URL or HUNYUAN_API_URL, YUANBAO_MODEL or HUNYUAN_MODEL, 'env_tencent')
    append_candidate(
        DEEPSEEK_API_URL or 'https://api.deepseek.com',
        _default_copywriter_model(provider='deepseek'),
        'env_deepseek',
    )
    return candidates[:3]


def _copywriter_runtime_config(payload=None):
    candidates = _copywriter_runtime_candidates(payload=payload)
    if candidates:
        current = dict(candidates[0])
        current['candidate_count'] = len(candidates)
        current['fallback_mode'] = False
        return current
    runtime_config = _automation_runtime_config()
    payload = payload or {}
    api_url_override = (payload.get('copywriter_api_url') or runtime_config.get('copywriter_api_url') or '').strip()
    model_override = (payload.get('copywriter_model') or runtime_config.get('copywriter_model') or '').strip()
    return {
        'configured': False,
        'provider': 'local_fallback',
        'api_key': '',
        'api_url': _normalize_copywriter_api_url(api_url_override, model=model_override),
        'model': model_override,
        'label': '本地兜底生成',
        'candidate_count': 0,
        'fallback_mode': True,
    }


def _extract_copywriter_text(response_json):
    if not isinstance(response_json, dict):
        return ''
    choices = response_json.get('choices') or []
    if choices:
        message = (choices[0] or {}).get('message') or {}
        content = message.get('content')
        if isinstance(content, str):
            return content.strip()
        if isinstance(content, list):
            parts = []
            for item in content:
                if isinstance(item, dict) and item.get('type') == 'text':
                    parts.append(str(item.get('text') or ''))
            return '\n'.join(parts).strip()
    output = response_json.get('output') or []
    if isinstance(output, list):
        parts = []
        for item in output:
            if isinstance(item, dict):
                for content in (item.get('content') or []):
                    if isinstance(content, dict) and content.get('type') == 'output_text':
                        parts.append(str(content.get('text') or ''))
        return '\n'.join(parts).strip()
    return ''


def _extract_copywriter_reasoning_text(response_json):
    if not isinstance(response_json, dict):
        return ''
    choices = response_json.get('choices') or []
    if choices:
        message = (choices[0] or {}).get('message') or {}
        reasoning_text = message.get('reasoning_content')
        if isinstance(reasoning_text, str):
            return reasoning_text.strip()
    output = response_json.get('output') or []
    if isinstance(output, list):
        parts = []
        for item in output:
            if not isinstance(item, dict):
                continue
            for content in (item.get('content') or []):
                if not isinstance(content, dict):
                    continue
                if content.get('type') in {'reasoning', 'reasoning_text', 'output_reasoning'}:
                    text = str(content.get('text') or content.get('summary') or '').strip()
                    if text:
                        parts.append(text)
        return '\n'.join(parts).strip()
    return ''


def _messages_to_response_input(messages):
    rows = []
    for message in messages or []:
        if not isinstance(message, dict):
            continue
        role = (message.get('role') or 'user').strip() or 'user'
        content = message.get('content')
        if isinstance(content, str):
            text = content.strip()
            if not text:
                continue
            rows.append({
                'role': role,
                'content': [{'type': 'input_text', 'text': text}],
            })
        elif isinstance(content, list):
            parts = []
            for item in content:
                if isinstance(item, str) and item.strip():
                    parts.append({'type': 'input_text', 'text': item.strip()})
                elif isinstance(item, dict) and item.get('type') == 'text':
                    text = str(item.get('text') or '').strip()
                    if text:
                        parts.append({'type': 'input_text', 'text': text})
            if parts:
                rows.append({'role': role, 'content': parts})
    return rows


def _build_copywriter_request_payload(runtime, messages, *, temperature=1.0, top_p=0.9, extra_payload=None):
    api_url = (runtime.get('api_url') or '').strip()
    provider = (runtime.get('provider') or '').strip()
    model_name = (runtime.get('model') or '').strip()
    omit_sampling_controls = provider == 'deepseek' and _is_deepseek_reasoning_model(model_name)
    if api_url.endswith('/responses'):
        payload = {
            'model': runtime['model'],
            'input': _messages_to_response_input(messages),
        }
    else:
        payload = {
            'model': runtime['model'],
            'messages': messages,
        }
    if not omit_sampling_controls:
        payload['temperature'] = temperature
        payload['top_p'] = top_p
    if provider == 'deepseek' and _is_deepseek_chat_model(model_name) and not api_url.endswith('/responses'):
        payload['thinking'] = {'type': 'enabled'}
    if isinstance(extra_payload, dict):
        blocked_keys = set()
        if omit_sampling_controls:
            blocked_keys.update({'temperature', 'top_p', 'presence_penalty', 'frequency_penalty', 'logprobs', 'top_logprobs'})
        for key, value in extra_payload.items():
            if key in blocked_keys:
                continue
            payload[key] = value
    return payload


def _extract_json_object(text):
    content = (text or '').strip()
    if not content:
        return {}
    try:
        data = json.loads(content)
        return data if isinstance(data, dict) else {}
    except Exception:
        pass
    start = content.find('{')
    end = content.rfind('}')
    if start >= 0 and end > start:
        snippet = content[start:end + 1]
        try:
            data = json.loads(snippet)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def _call_copywriter(messages, *, temperature=1.0, top_p=0.9, timeout=40, extra_payload=None, runtime_override=None):
    timeout = max(min(timeout, 12), 3)
    if runtime_override:
        runtime_candidates = runtime_override if isinstance(runtime_override, list) else [runtime_override]
    else:
        runtime_candidates = _copywriter_runtime_candidates()
    if not runtime_candidates:
        raise RuntimeError('copywriter_not_configured')
    attempt_errors = []
    for runtime in runtime_candidates:
        headers = {
            'Authorization': f'Bearer {runtime["api_key"]}',
            'Content-Type': 'application/json',
        }
        payload = _build_copywriter_request_payload(
            runtime,
            messages,
            temperature=temperature,
            top_p=top_p,
            extra_payload=extra_payload,
        )
        try:
            resp = requests.post(runtime['api_url'], json=payload, headers=headers, timeout=timeout)
            resp.raise_for_status()
            response_json = resp.json()
            return {
                'text': _extract_copywriter_text(response_json),
                'reasoning_text': _extract_copywriter_reasoning_text(response_json),
                'runtime': runtime,
                'attempt_errors': attempt_errors,
            }
        except Exception as exc:
            response = getattr(exc, 'response', None)
            error_text = str(exc)[:300]
            status_code = getattr(response, 'status_code', 0) or 0
            if response is not None:
                body_text = re.sub(r'\s+', ' ', (response.text or '').strip())[:300]
                if body_text:
                    error_text = body_text
            attempt_errors.append({
                'provider': runtime.get('provider') or '',
                'model': runtime.get('model') or '',
                'api_url': runtime.get('api_url') or '',
                'status_code': status_code,
                'error': error_text,
            })
            continue
    raise RuntimeError(f'copywriter_all_candidates_failed: {json.dumps(attempt_errors, ensure_ascii=False)}')

def auto_humanize_text(content):
    """自动去AI化重写（保留原意）"""
    if not content or not _copywriter_runtime_config().get('configured'):
        return content
    try:
        prompt = f"""请把下面这段小红书文案做"真人化重写"，并保持原意：

要求：
1) 去AI腔，像真实用户随手发的分享
2) 保留原有结构（标题+钩子+内文）
3) 增加生活细节和情绪波动，但不要夸张
4) 内文保持200-300字
5) 结尾自然提问，不要出现"评论区"
6) 合规：不要绝对化词，不引导购买
7) 不要写“我是从xx身份出发”“我会把xx顺手带出”“围绕xx最容易写空泛”这类模板句

原文：
{content}

只输出重写后的内容。"""
        result = _call_copywriter(
            [{'role': 'user', 'content': prompt}],
            temperature=1.08,
            top_p=0.9,
            timeout=30,
        )
        if result.get('text'):
            return result['text']
    except Exception as e:
        print(f"auto_humanize error: {e}")
    return content

# 小红书API配置
XHS_PROXY = os.environ.get('XHS_PROXY', None)  # 如需代理设置
XHS_WEB_SESSION = os.environ.get('XHS_WEB_SESSION', None)  # 如需登录态

def fetch_xhs_trending(keywords):
    """从获取小红书热门笔记"""
    try:
        import asyncio
        import sys
        sys.path.insert(0, '/home/node/.openclaw/workspace/xiaohongshu/scripts')

        from request.web.xhs_session import create_xhs_session

        async def search():
            xhs = await create_xhs_session(proxy=XHS_PROXY, web_session=XHS_WEB_SESSION)
            # 提取第一个关键词搜索
            kw = keywords.split('#')[0].split(' ')[0].strip() if keywords else '体检'
            res = await xhs.apis.note.search_notes(kw)
            data = await res.json()
            await xhs.close_session()
            return data

        # 同步调用
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result = loop.run_until_complete(search())
            if result.get('code') == -104:
                # 未登录无权限，返回空
                return []
            notes = result.get('data', {}).get('items', [])[:3]
            return notes
        finally:
            loop.close()
    except Exception as e:
        print(f"XHS API error: {e}")
        return []


COPY_PERSONA_OPTIONS = {
    'auto': '系统自动匹配',
    'doctor_assistant': '医生助理',
    'health_manager': '健管师',
    'nutritionist': '营养师',
    'tcm_practitioner': '中医调理视角',
    'fitness_coach': '运动减脂教练',
    'emotional_support': '情绪陪伴者',
    'women_health': '女性健康视角',
    'physical_exam_blogger': '体检复盘博主',
    'chronic_manager': '慢病管理者',
    'patient_self': '患者本人',
    'patient_family': '患者家属',
    'medical_science': '医学科普',
    'patient_friend': '患者朋友',
    'office_worker': '职场久坐人群',
    'caregiver': '陪诊照护者',
    'custom': '自定义身份',
}

COPY_SCENE_OPTIONS = {
    'auto': '系统自动匹配',
    'clinic_consulting': '门诊沟通答疑',
    'physical_exam_alert': '体检异常提醒',
    'followup_review': '复查对比变化',
    'daily_liver_care': '日常护肝管理',
    'drinking_recovery': '熬夜应酬后护肝',
    'family_support': '家属陪伴照护',
    'diet_adjustment': '饮食调整建议',
    'report_interpretation': '检查报告解读',
    'tcm_conditioning': '中医调理养肝',
    'fatty_liver_management': '脂肪肝减脂管理',
    'mood_stress': '情绪压力与肝气',
    'sleep_recovery': '失眠熬夜恢复期',
    'women_dailycare': '女性日常养肝',
    'office_sedentary': '久坐久盯屏场景',
    'exercise_rebuild': '运动减脂重启期',
    'custom': '自定义场景',
}

COPY_DIRECTION_OPTIONS = {
    'auto': '系统自动匹配',
    'report_interpretation': '检查解读线',
    'fatty_liver_management': '脂肪肝管理线',
    'liver_care_habits': '保肝护肝习惯线',
    'tcm_conditioning': '中医调理线',
    'emotion_mindbody': '情绪与肝线',
    'diet_nutrition': '饮食营养线',
    'exercise_fitness': '运动减脂线',
    'family_care': '家庭照护线',
    'women_health': '女性养肝线',
    'myth_busting': '误区纠偏线',
    'custom': '自定义方向',
}

COPY_GOAL_OPTIONS = {
    'balanced': '均衡输出',
    'viral_title': '爆款点击优先',
    'save_value': '收藏种草优先',
    'comment_engagement': '评论互动优先',
    'trust_building': '专业信任优先',
}

COPY_PRODUCT_OPTIONS = {
    'auto': '自动匹配',
    'none': '不植入产品',
    'fibroscan': 'FibroScan福波看',
    'soft_liver_tablet': '复方鳖甲软肝片',
    'entecavir_combo': '恩替卡韦联合管理',
    'qizhi_capsule': '壳脂胶囊',
}


def _resolve_copy_selection(option_map, raw_value, *, custom_value=''):
    key = (raw_value or 'auto').strip() if isinstance(raw_value, str) else 'auto'
    if key == 'custom':
        custom_text = (custom_value or '').strip()
        return custom_text or option_map.get('custom') or ''
    return option_map.get(key, option_map.get('auto') or '')


def _resolve_copy_direction_selection(direction_key, topic_text='', user_prompt='', custom_value=''):
    key = (direction_key or 'auto').strip() if isinstance(direction_key, str) else 'auto'
    if key == 'custom':
        custom_text = (custom_value or '').strip()
        return custom_text or COPY_DIRECTION_OPTIONS.get('custom') or ''
    if key != 'auto':
        return COPY_DIRECTION_OPTIONS.get(key, COPY_DIRECTION_OPTIONS.get('auto') or '')

    text = ' '.join(filter(None, [
        (topic_text or '').strip(),
        (user_prompt or '').strip(),
    ]))
    if any(token in text for token in ['中医', '肝气郁结', '肝郁', '疏肝理气', '肝胆湿热', '养肝']):
        return COPY_DIRECTION_OPTIONS['tcm_conditioning']
    if any(token in text for token in ['脂肪肝', '减脂', '代谢', '体重', '减肥']):
        return COPY_DIRECTION_OPTIONS['fatty_liver_management']
    if any(token in text for token in ['情绪', '焦虑', '压力', '失眠', '睡眠', '熬夜']):
        return COPY_DIRECTION_OPTIONS['emotion_mindbody']
    if any(token in text for token in ['饮食', '营养', '吃什么', '食谱', '控糖', '控油']):
        return COPY_DIRECTION_OPTIONS['diet_nutrition']
    if any(token in text for token in ['运动', '健身', '走路', '跑步', '减脂训练']):
        return COPY_DIRECTION_OPTIONS['exercise_fitness']
    if any(token in text for token in ['父母', '家人', '陪诊', '照护', '家属']):
        return COPY_DIRECTION_OPTIONS['family_care']
    if any(token in text for token in ['女性', '姨妈', '经期', '更年期']):
        return COPY_DIRECTION_OPTIONS['women_health']
    if any(token in text for token in ['误区', '别再', '为什么', '是不是', '搞错']):
        return COPY_DIRECTION_OPTIONS['myth_busting']
    if any(token in text for token in ['体检', '检查', '报告', '指标', 'FibroScan', '福波看', '肝弹', '转氨酶']):
        return COPY_DIRECTION_OPTIONS['report_interpretation']
    return COPY_DIRECTION_OPTIONS['liver_care_habits']


def _derive_topic_product_hint(topic_text=''):
    topic_text = (topic_text or '').strip()
    if any(k in topic_text for k in ['肝弹', 'FibroScan', '福波看', '做检查', '检查']):
        return (
            'FibroScan福波看',
            '主线：FibroScan福波看（检查评估）；强调检查解读、复查趋势和早发现早干预，不要写成硬广。'
        )
    if any(k in topic_text for k in ['肝硬化吃什么药', '肝纤维化吃什么药', '吃什么药', '乙肝肝纤维化', '乙肝肝硬化']):
        return (
            '恩替卡韦联合管理',
            '主线：恩替卡韦（抗病毒）+复方鳖甲软肝片（抗纤维化）联合管理；表达要克制，强调个体化评估和长期管理。'
        )
    if any(k in topic_text for k in ['解酒', '护肝']) and not any(k in topic_text for k in ['纤维化', '肝硬化']):
        return (
            '复方鳖甲软肝片',
            '主线：复方鳖甲软肝片在解酒护肝场景下的自然带出；可以提便携和管理习惯，但不能写成直接促销。'
        )
    if '脂肪肝' in topic_text:
        return (
            '壳脂胶囊',
            '主线：围绕脂肪肝管理场景自然带出壳脂胶囊，可结合减脂、饮食和复查，不要夸大效果。'
        )
    return (
        '自动匹配',
        '按话题自动匹配产品或服务，不强行植入；优先保持内容自然和可信。'
    )


def _resolve_copy_product_selection(product_key, topic_text=''):
    key = (product_key or 'auto').strip() if isinstance(product_key, str) else 'auto'
    topic_label, topic_hint = _derive_topic_product_hint(topic_text)
    if key == 'none':
        return '不植入产品', '本篇只做肝病科普和经验分享，不强行植入产品，但可以保留管理建议和复查提醒。'
    if key == 'fibroscan':
        return 'FibroScan福波看', '请自然带出FibroScan福波看在检查评估、指标解读和复查趋势里的价值，语气像真实经验分享。'
    if key == 'soft_liver_tablet':
        return '复方鳖甲软肝片', '请自然带出复方鳖甲软肝片在抗纤维化管理里的角色，不要写成购买引导。'
    if key == 'entecavir_combo':
        return '恩替卡韦联合管理', '请按“抗病毒+抗纤维化联合管理”主线软植入，强调医生指导和长期管理。'
    if key == 'qizhi_capsule':
        return '壳脂胶囊', '请围绕脂肪肝、代谢管理或饮食调整场景自然带出壳脂胶囊，不要夸大。'
    return topic_label, topic_hint


def _goal_strategy_profile(goal_key='balanced'):
    profiles = {
        'viral_title': {
            'label': COPY_GOAL_OPTIONS['viral_title'],
            'title_rule': '标题优先用结果前置、反差感、提问句或踩坑句式，第一眼要有点击冲动。',
            'variants': ['反常识冲突版', '结果前置版', '问题悬念版'],
        },
        'save_value': {
            'label': COPY_GOAL_OPTIONS['save_value'],
            'title_rule': '标题要突出实用性、清单感和收藏价值，正文尽量给出清楚步骤或判断点。',
            'variants': ['实操清单版', '避坑总结版', '复查提醒版'],
        },
        'comment_engagement': {
            'label': COPY_GOAL_OPTIONS['comment_engagement'],
            'title_rule': '标题要留出讨论空间，正文结尾一定要自然抛问题，引发同类人回复。',
            'variants': ['求经验版', '观点讨论版', '经历求助版'],
        },
        'trust_building': {
            'label': COPY_GOAL_OPTIONS['trust_building'],
            'title_rule': '标题更克制，突出专业感和可信度，正文要像真实沟通，不要像科普课件。',
            'variants': ['门诊解释版', '复查分析版', '陪诊总结版'],
        },
        'balanced': {
            'label': COPY_GOAL_OPTIONS['balanced'],
            'title_rule': '标题兼顾点击、可信和互动，不要太硬广，也不要太平。',
            'variants': ['故事共鸣版', '轻科普拆解版', '互动讨论版'],
        },
    }
    return profiles.get(goal_key, profiles['balanced'])


def _extract_labeled_block(text, labels, all_labels):
    import re
    pattern = '|'.join(re.escape(label) for label in labels)
    next_pattern = '|'.join(re.escape(label) for label in all_labels)
    match = re.search(
        rf'(?ms)(?:^|\n)\s*(?:{pattern})\s*[：:]\s*(.*?)(?=\n\s*(?:{next_pattern})\s*[：:]|$)',
        text or '',
    )
    return (match.group(1) if match else '').strip()


def _render_generated_copy_card(card):
    item = dict(card or {})
    lines = []
    for label, key in [
        ('人设', 'persona'),
        ('场景', 'scene'),
        ('软植入', 'insertion'),
        ('标题', 'title'),
        ('开头钩子', 'hook'),
        ('正文', 'body'),
        ('互动结尾', 'ending'),
    ]:
        value = (item.get(key) or '').strip()
        if value:
            lines.append(f'{label}：{value}')
    return '\n'.join(lines).strip()


def _parse_generated_copy_card(version_text: str, defaults=None):
    import re

    defaults = defaults or {}
    text = re.sub(r'===+', '', (version_text or '')).strip()
    label_aliases = {
        'persona': ['人设', '角色'],
        'scene': ['场景'],
        'insertion': ['软植入', '软植入产品'],
        'title': ['标题'],
        'hook': ['开头钩子', '钩子'],
        'body': ['正文', '内文'],
        'ending': ['互动结尾', '结尾互动'],
    }
    all_labels = []
    for rows in label_aliases.values():
        all_labels.extend(rows)

    card = {}
    for key, labels in label_aliases.items():
        card[key] = _extract_labeled_block(text, labels, all_labels)

    if not card.get('title') or not card.get('body'):
        fallback_title, fallback_body = _parse_model_output(text)
        card['title'] = card.get('title') or fallback_title
        card['body'] = card.get('body') or fallback_body

    for key in ['persona', 'scene', 'insertion']:
        if not (card.get(key) or '').strip():
            card[key] = (defaults.get(key) or '').strip()

    body = re.sub(r'(?m)^\s*(人设|角色|场景|软植入|软植入产品|标题|开头钩子|钩子|互动结尾|结尾互动)\s*[：:].*$', '', card.get('body') or '')
    body = re.sub(r'\n{3,}', '\n\n', body).strip()
    hook_text = (card.get('hook') or '').strip()
    if hook_text and body:
        body_lines = [line.strip() for line in body.splitlines() if line.strip()]
        if body_lines and body_lines[0] == hook_text:
            body = '\n'.join(body_lines[1:]).strip()
    card['body'] = body

    if not (card.get('hook') or '').strip() and body:
        first_line = next((line.strip() for line in body.splitlines() if line.strip()), '')
        card['hook'] = first_line[:40]
    if not (card.get('ending') or '').strip():
        card['ending'] = (defaults.get('ending') or '你们会怎么做？').strip()

    title = re.sub(r'^(标题\s*[：:]\s*)+', '', (card.get('title') or '').strip()).strip('：: =-')
    if not title:
        title = '分享笔记'
    card['title'] = title[:30]
    card['copy_text'] = _render_generated_copy_card(card)
    return card


def _copy_card_usable(card):
    title = (card.get('title') or '').strip()
    body = (card.get('body') or '').strip()
    bad_title_tokens = ['带话题', '关键词', '版本', '#']
    if not title or len(title) > 24:
        return False
    if any(token in title for token in bad_title_tokens):
        return False
    if len(body) < 80:
        return False
    return True


def _load_strategy_payload(raw_payload):
    if not raw_payload:
        return {}
    try:
        data = json.loads(raw_payload)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _serialize_submission_strategy(submission):
    payload = _load_strategy_payload(submission.strategy_payload)
    return {
        'selected_persona_key': payload.get('selected_persona_key') or '',
        'selected_scene_key': payload.get('selected_scene_key') or '',
        'selected_direction_key': payload.get('selected_direction_key') or '',
        'selected_product_key': payload.get('selected_product_key') or '',
        'selected_agent_copy_route_id': payload.get('selected_agent_copy_route_id') or '',
        'selected_agent_image_route_id': payload.get('selected_agent_image_route_id') or '',
        'selected_image_agent_plan_id': payload.get('selected_image_agent_plan_id') or '',
        'selected_reference_plan_id': payload.get('selected_reference_plan_id') or '',
        'selected_reference_asset_ids': payload.get('selected_reference_asset_ids') or [],
        'selected_title': submission.selected_title or payload.get('selected_title') or '',
        'selected_title_source': submission.selected_title_source or payload.get('selected_title_source') or '',
        'selected_title_index': submission.selected_title_index if submission.selected_title_index is not None else (payload.get('selected_title_index') or 0),
        'selected_copy_version_index': submission.selected_copy_version_index if submission.selected_copy_version_index is not None else (payload.get('selected_copy_version_index') or 0),
        'selected_copy_goal': submission.selected_copy_goal or payload.get('selected_copy_goal') or '',
        'selected_copy_skill': submission.selected_copy_skill or payload.get('selected_copy_skill') or '',
        'selected_title_skill': submission.selected_title_skill or payload.get('selected_title_skill') or '',
        'selected_image_skill': submission.selected_image_skill or payload.get('selected_image_skill') or '',
        'selected_cover_style_type': submission.selected_cover_style_type or payload.get('selected_cover_style_type') or '',
        'selected_inner_style_type': submission.selected_inner_style_type or payload.get('selected_inner_style_type') or '',
        'selected_generation_mode': submission.selected_generation_mode or payload.get('selected_generation_mode') or '',
        'selected_copy_text': submission.selected_copy_text or payload.get('selected_copy_text') or '',
        'strategy_updated_at': submission.strategy_updated_at.strftime('%Y-%m-%d %H:%M:%S') if submission.strategy_updated_at else '',
        'generator_context': payload.get('generator_context') or {},
        'title_options': payload.get('title_options') or [],
        'reason': payload.get('reason') or '',
    }


def _apply_submission_strategy_snapshot(submission, payload, registration=None):
    registration = registration or getattr(submission, 'registration', None)
    title_options = payload.get('title_options') or []
    normalized_title_options = []
    if isinstance(title_options, list):
        for item in title_options[:8]:
            if isinstance(item, str):
                title_text = _normalize_title_candidate(item)
                if title_text:
                    normalized_title_options.append({'title': title_text, 'source': '系统推荐'})
                continue
            if not isinstance(item, dict):
                continue
            title_text = _normalize_title_candidate(item.get('title') or '')
            if not title_text:
                continue
            normalized_title_options.append({
                'title': title_text,
                'source': (str(item.get('source') or '系统推荐').strip()[:100] or '系统推荐'),
            })

    selected_title = _normalize_title_candidate(payload.get('selected_title') or '')
    selected_title_source = (payload.get('selected_title_source') or '').strip()[:100]
    selected_title_index = max(_safe_int(payload.get('selected_title_index'), 0), 0)
    if not selected_title and normalized_title_options and selected_title_index < len(normalized_title_options):
        selected_title = normalized_title_options[selected_title_index]['title']
    if not selected_title_source and normalized_title_options and selected_title_index < len(normalized_title_options):
        selected_title_source = normalized_title_options[selected_title_index].get('source') or '系统推荐'

    selected_copy_text = (payload.get('selected_copy_text') or '').strip()
    if len(selected_copy_text) > 4000:
        selected_copy_text = selected_copy_text[:4000]

    generator_context = payload.get('generator_context') if isinstance(payload.get('generator_context'), dict) else {}
    strategy_snapshot = {
        'reason': (payload.get('reason') or '').strip()[:80],
        'registration_id': registration.id if registration else submission.registration_id,
        'topic_id': registration.topic_id if registration else None,
        'topic_name': registration.topic.topic_name if registration and registration.topic else '',
        'selected_persona_key': (payload.get('selected_persona_key') or '').strip()[:50],
        'selected_scene_key': (payload.get('selected_scene_key') or '').strip()[:50],
        'selected_direction_key': (payload.get('selected_direction_key') or '').strip()[:50],
        'selected_product_key': (payload.get('selected_product_key') or '').strip()[:50],
        'selected_agent_copy_route_id': (payload.get('selected_agent_copy_route_id') or '').strip()[:80],
        'selected_agent_image_route_id': (payload.get('selected_agent_image_route_id') or '').strip()[:80],
        'selected_image_agent_plan_id': (payload.get('selected_image_agent_plan_id') or '').strip()[:120],
        'selected_reference_plan_id': (payload.get('selected_reference_plan_id') or '').strip()[:120],
        'selected_reference_asset_ids': [
            str(item).strip()[:20]
            for item in (payload.get('selected_reference_asset_ids') or [])
            if str(item).strip()
        ][:6],
        'selected_title': selected_title,
        'selected_title_source': selected_title_source,
        'selected_title_index': selected_title_index,
        'selected_copy_version_index': max(_safe_int(payload.get('selected_copy_version_index'), 0), 0),
        'selected_copy_goal': (payload.get('selected_copy_goal') or '').strip()[:50],
        'selected_copy_skill': (payload.get('selected_copy_skill') or '').strip()[:50],
        'selected_title_skill': (payload.get('selected_title_skill') or '').strip()[:50],
        'selected_image_skill': (payload.get('selected_image_skill') or '').strip()[:50],
        'selected_cover_style_type': (payload.get('selected_cover_style_type') or '').strip()[:50],
        'selected_inner_style_type': (payload.get('selected_inner_style_type') or '').strip()[:50],
        'selected_generation_mode': (payload.get('selected_generation_mode') or '').strip()[:50],
        'selected_copy_text': selected_copy_text,
        'generator_context': {str(key)[:50]: str(value)[:120] for key, value in generator_context.items()},
        'title_options': normalized_title_options,
    }

    submission.selected_title = strategy_snapshot['selected_title'] or submission.selected_title
    submission.selected_title_source = strategy_snapshot['selected_title_source'] or submission.selected_title_source
    submission.selected_title_index = strategy_snapshot['selected_title_index']
    submission.selected_copy_version_index = strategy_snapshot['selected_copy_version_index']
    submission.selected_copy_goal = strategy_snapshot['selected_copy_goal'] or submission.selected_copy_goal
    submission.selected_copy_skill = strategy_snapshot['selected_copy_skill'] or submission.selected_copy_skill
    submission.selected_title_skill = strategy_snapshot['selected_title_skill'] or submission.selected_title_skill
    submission.selected_image_skill = strategy_snapshot['selected_image_skill'] or submission.selected_image_skill
    submission.selected_cover_style_type = strategy_snapshot['selected_cover_style_type'] or submission.selected_cover_style_type
    submission.selected_inner_style_type = strategy_snapshot['selected_inner_style_type'] or submission.selected_inner_style_type
    submission.selected_generation_mode = strategy_snapshot['selected_generation_mode'] or submission.selected_generation_mode
    submission.selected_copy_text = strategy_snapshot['selected_copy_text'] or submission.selected_copy_text
    submission.strategy_payload = json.dumps(strategy_snapshot, ensure_ascii=False)
    submission.strategy_updated_at = datetime.now()
    return strategy_snapshot


def _build_strategy_insights(submissions):
    scoped_submissions = [sub for sub in submissions if sub.selected_title_skill or sub.selected_image_skill or sub.strategy_payload]
    captured_count = len(scoped_submissions)
    title_skill_label_map = {
        str(key).strip(): str(value).strip()
        for key, value in (TITLE_SKILL_OPTIONS or {}).items()
    }
    image_skill_label_map = {
        str(key).strip(): str(value).strip()
        for key, value in (IMAGE_SKILL_OPTIONS or {}).items()
    }
    if not scoped_submissions:
        return {
            'captured_count': 0,
            'capture_rate': 0,
            'capture_rate_display': '0%',
            'title_skill_rows': [],
            'image_skill_rows': [],
            'combo_rows': [],
            'latest_submissions': [],
            'summary_lines': [],
            'best_title_skill': {},
            'best_image_skill': {},
            'best_combo': {},
        }

    def build_metric_rows(rows):
        items = []
        for key, item in rows.items():
            avg_views = round(item['views'] / item['count'], 2) if item['count'] else 0
            avg_interactions = round(item['interactions'] / item['count'], 2) if item['count'] else 0
            avg_interaction_rate = _calculate_rate(item['interactions'], item['views'])
            viral_rate = _calculate_rate(item['viral_count'], item['count'])
            items.append({
                'key': key,
                'label': item['label'],
                'count': item['count'],
                'views': item['views'],
                'interactions': item['interactions'],
                'avg_views': avg_views,
                'avg_interactions': avg_interactions,
                'avg_interaction_rate': avg_interaction_rate,
                'avg_interaction_rate_display': _format_rate(avg_interaction_rate),
                'viral_count': item['viral_count'],
                'viral_rate': viral_rate,
                'viral_rate_display': _format_rate(viral_rate),
            })
        items.sort(key=lambda row: ((row['avg_interactions'] or 0), (row['viral_rate'] or 0), row['count']), reverse=True)
        return items

    def build_compare_rows(items, field, label_map):
        ordered = sorted(
            items,
            key=lambda sub: (
                sub.strategy_updated_at or sub.updated_at or sub.created_at or datetime.min,
                sub.id or 0,
            ),
            reverse=True,
        )
        recent_items = ordered[:12]
        previous_items = ordered[12:24]
        buckets = {}

        def ingest(rows, bucket_key):
            for sub in rows:
                raw_key = (getattr(sub, field, '') or '未记录').strip() or '未记录'
                row = buckets.setdefault(raw_key, {
                    'label': label_map.get(raw_key, raw_key),
                    'recent_count': 0,
                    'previous_count': 0,
                    'recent_interactions': 0,
                    'previous_interactions': 0,
                })
                interactions = (sub.xhs_likes or 0) + (sub.xhs_favorites or 0) + (sub.xhs_comments or 0)
                if bucket_key == 'recent':
                    row['recent_count'] += 1
                    row['recent_interactions'] += interactions
                else:
                    row['previous_count'] += 1
                    row['previous_interactions'] += interactions

        ingest(recent_items, 'recent')
        ingest(previous_items, 'previous')

        results = []
        for key, row in buckets.items():
            recent_avg = round(row['recent_interactions'] / row['recent_count'], 2) if row['recent_count'] else 0
            previous_avg = round(row['previous_interactions'] / row['previous_count'], 2) if row['previous_count'] else 0
            delta = round(recent_avg - previous_avg, 2)
            if row['recent_count'] and not row['previous_count']:
                trend_label = '新上升'
            elif delta >= 8:
                trend_label = '上升'
            elif delta <= -8:
                trend_label = '回落'
            else:
                trend_label = '持平'
            results.append({
                'key': key,
                'label': row['label'],
                'recent_count': row['recent_count'],
                'previous_count': row['previous_count'],
                'recent_avg_interactions': recent_avg,
                'previous_avg_interactions': previous_avg,
                'delta_avg_interactions': delta,
                'trend_label': trend_label,
            })
        results.sort(
            key=lambda item: (
                item['recent_count'],
                item['recent_avg_interactions'],
                item['delta_avg_interactions'],
            ),
            reverse=True,
        )
        return results[:5]

    title_skill_rows = {}
    image_skill_rows = {}
    combo_rows = {}
    latest_submissions = []

    for sub in scoped_submissions:
        xhs_interactions = (sub.xhs_likes or 0) + (sub.xhs_favorites or 0) + (sub.xhs_comments or 0)
        xhs_views = sub.xhs_views or 0
        is_viral = _infer_viral_post(
            views=sub.xhs_views or 0,
            likes=sub.xhs_likes or 0,
            favorites=sub.xhs_favorites or 0,
            comments=sub.xhs_comments or 0,
        ) if _submission_has_platform_link(sub, 'xhs') else False
        title_skill_key = (sub.selected_title_skill or '未记录').strip() or '未记录'
        image_skill_key = (sub.selected_image_skill or '未记录').strip() or '未记录'
        title_skill_label = title_skill_label_map.get(title_skill_key, title_skill_key)
        image_skill_label = image_skill_label_map.get(image_skill_key, image_skill_key)
        combo_key = f'{title_skill_key} × {image_skill_key}'
        combo_label = f'{title_skill_label} × {image_skill_label}'

        for container, key, label in [
            (title_skill_rows, title_skill_key, title_skill_label),
            (image_skill_rows, image_skill_key, image_skill_label),
            (combo_rows, combo_key, combo_label),
        ]:
            row = container.setdefault(key, {
                'label': label,
                'count': 0,
                'views': 0,
                'interactions': 0,
                'viral_count': 0,
            })
            row['count'] += 1
            row['views'] += xhs_views
            row['interactions'] += xhs_interactions
            row['viral_count'] += 1 if is_viral else 0

        if len(latest_submissions) < 8:
            latest_submissions.append({
                'registration_id': sub.registration_id,
                'title': sub.selected_title or sub.note_title or '未命名标题',
                'title_skill': title_skill_label,
                'image_skill': image_skill_label,
                'views': xhs_views,
                'interactions': xhs_interactions,
                'strategy_updated_at': sub.strategy_updated_at.strftime('%Y-%m-%d %H:%M:%S') if sub.strategy_updated_at else '',
            })

    title_skill_result = build_metric_rows(title_skill_rows)[:8]
    image_skill_result = build_metric_rows(image_skill_rows)[:8]
    combo_result = build_metric_rows(combo_rows)[:8]
    title_trend_rows = build_compare_rows(scoped_submissions, 'selected_title_skill', title_skill_label_map)
    image_trend_rows = build_compare_rows(scoped_submissions, 'selected_image_skill', image_skill_label_map)

    best_title_skill = title_skill_result[0] if title_skill_result else {}
    best_image_skill = image_skill_result[0] if image_skill_result else {}
    best_combo = combo_result[0] if combo_result else {}

    summary_lines = []
    if best_title_skill:
        summary_lines.append(
            f"当前更值得放大的标题打法是“{best_title_skill['label']}”，样本 {best_title_skill['count']} 条，平均互动 {best_title_skill['avg_interactions']}，爆款率 {best_title_skill['viral_rate_display']}。"
        )
    if best_image_skill:
        summary_lines.append(
            f"当前更稳的图片打法是“{best_image_skill['label']}”，样本 {best_image_skill['count']} 条，平均互动 {best_image_skill['avg_interactions']}，爆款率 {best_image_skill['viral_rate_display']}。"
        )
    if best_combo:
        summary_lines.append(
            f"当前最该优先复用的组合是“{best_combo['label']}”，样本 {best_combo['count']} 条，平均互动 {best_combo['avg_interactions']}。"
        )
    if captured_count and (captured_count / max(len(submissions), 1)) < 0.6:
        summary_lines.append('当前策略留痕样本还不够，后续建议统一从系统标题池和图片打法里选，避免分析结论失真。')

    return {
        'captured_count': captured_count,
        'capture_rate': _calculate_rate(captured_count, len(submissions)) or 0,
        'capture_rate_display': _format_rate(_calculate_rate(captured_count, len(submissions))),
        'title_skill_rows': title_skill_result,
        'image_skill_rows': image_skill_result,
        'combo_rows': combo_result,
        'title_trend_rows': title_trend_rows,
        'image_trend_rows': image_trend_rows,
        'latest_submissions': latest_submissions,
        'summary_lines': summary_lines[:4],
        'best_title_skill': best_title_skill,
        'best_image_skill': best_image_skill,
        'best_combo': best_combo,
    }


def _detect_topic_strategy_traits(topic_name='', keywords='', direction=''):
    text = ' '.join([
        str(topic_name or ''),
        str(keywords or ''),
        str(direction or ''),
    ]).strip()
    lowered = text.lower()
    return {
        'report_like': bool(re.search(r'体检|检查|报告|指标|fibroscan|福波看|复查', text, re.I)),
        'discussion_like': bool(re.search(r'怎么办|怎么选|会怎么做|你会|求助|有人知道|能不能', text, re.I)),
        'myth_like': bool(re.search(r'误区|为什么|是不是|别再|千万别|想当然|搞错', text, re.I)),
        'story_like': bool(re.search(r'经历|经验|踩坑|复盘|后来|家属|陪诊|我妈|我爸|我自己', text, re.I)),
        'emotion_like': bool(re.search(r'焦虑|害怕|崩溃|担心|慌|拖着', text, re.I)),
        'raw_text': text,
        'lowered': lowered,
    }


def _detect_topic_format_traits(text=''):
    joined = str(text or '').strip()
    return {
        'poster_like': bool(re.search(r'大字报|互动型|超大字号|一句话核心|强烈视觉冲击', joined, re.I)),
        'checklist_like': bool(re.search(r'备忘录|清单|checklist|攻略|必做项目清单|抄作业|省钱清单', joined, re.I)),
        'chart_like': bool(re.search(r'图表|对比图|流程图|数据可视化|风险分层|表格', joined, re.I)),
        'realshot_like': bool(re.search(r'实拍|医院|体检中心|药盒|背影|等待区|门口|报告实拍', joined, re.I)),
        'professional_like': bool(re.search(r'深度科普|医疗行业背景|医学生|药学|营养师|文献|研报|科研', joined, re.I)),
        'report_photo_like': bool(re.search(r'报告解读|报告单|圈出|关键指标|结果解读|报告结果', joined, re.I)),
        'comparison_like': bool(re.search(r'横向测评|测评|对比|红黑榜|vs|区别', joined, re.I)),
    }


def _build_heuristic_strategy_recommendation(topic):
    topic_name = topic.topic_name if topic else ''
    keywords = topic.keywords if topic else ''
    direction = topic.direction if topic else ''
    traits = _detect_topic_strategy_traits(topic_name, keywords, direction)
    format_traits = _detect_topic_format_traits(' '.join([topic_name or '', keywords or '', direction or '']))
    if format_traits['professional_like'] or format_traits['chart_like'] or format_traits['report_photo_like']:
        recommended = {
            'copy_goal': 'trust_building',
            'copy_skill': 'report_interpretation',
            'title_skill': 'checklist_collect',
            'image_skill': 'report_decode',
            'cover_style_type': 'medical_science',
            'inner_style_type': 'checklist_report' if format_traits['report_photo_like'] else 'knowledge_card',
            'generation_mode': 'smart_bundle',
        }
        reason = '当前更适合走专业解释、图表拆解或报告翻译路线，先建立信任和收藏价值。'
    elif format_traits['checklist_like']:
        recommended = {
            'copy_goal': 'save_value',
            'copy_skill': 'practical_checklist',
            'title_skill': 'checklist_collect',
            'image_skill': 'save_worthy_cards',
            'cover_style_type': 'checklist',
            'inner_style_type': 'checklist_report',
            'generation_mode': 'smart_bundle',
        }
        reason = '当前内容类型更像备忘录/清单，优先做可收藏、可截图、可执行的版本。'
    elif format_traits['poster_like'] or traits['discussion_like']:
        recommended = {
            'copy_goal': 'comment_engagement',
            'copy_skill': 'discussion_hook',
            'title_skill': 'question_gap',
            'image_skill': 'high_click_cover',
            'cover_style_type': 'poster_bold',
            'inner_style_type': 'knowledge_card',
            'generation_mode': 'smart_bundle',
        }
        reason = '当前更适合先冲点击和评论区互动，优先用提问型标题和高点击封面。'
    elif format_traits['realshot_like']:
        recommended = {
            'copy_goal': 'balanced',
            'copy_skill': 'story_empathy',
            'title_skill': 'emotional_diary',
            'image_skill': 'story_atmosphere',
            'cover_style_type': 'memo_mobile',
            'inner_style_type': 'memo_classroom',
            'generation_mode': 'smart_bundle',
        }
        reason = '当前更适合走“真实陪同 / 真实体验 / 真实复查”表达，先做可信的经历感。'
    elif traits['report_like']:
        recommended = {
            'copy_goal': 'save_value',
            'copy_skill': 'report_interpretation',
            'title_skill': 'checklist_collect',
            'image_skill': 'report_decode',
            'cover_style_type': 'medical_science',
            'inner_style_type': 'checklist_report',
            'generation_mode': 'smart_bundle',
        }
        reason = '当前是检查/报告类话题，更容易靠“看懂指标 + 收藏备用”起量。'
    elif traits['myth_like']:
        recommended = {
            'copy_goal': 'viral_title',
            'copy_skill': 'myth_busting',
            'title_skill': 'conflict_reverse',
            'image_skill': 'high_click_cover',
            'cover_style_type': 'poster_bold',
            'inner_style_type': 'knowledge_card',
            'generation_mode': 'smart_bundle',
        }
        reason = '当前话题自带纠偏和反差感，更适合先冲点击。'
    elif traits['story_like'] or traits['emotion_like']:
        recommended = {
            'copy_goal': 'balanced',
            'copy_skill': 'story_empathy',
            'title_skill': 'emotional_diary',
            'image_skill': 'story_atmosphere',
            'cover_style_type': 'memo_mobile',
            'inner_style_type': 'memo_classroom',
            'generation_mode': 'smart_bundle',
        }
        reason = '当前更适合走真实经历和情绪代入，先做共鸣。'
    else:
        recommended = {
            'copy_goal': 'viral_title',
            'copy_skill': 'story_empathy',
            'title_skill': 'result_first',
            'image_skill': 'high_click_cover',
            'cover_style_type': 'poster_bold',
            'inner_style_type': 'knowledge_card',
            'generation_mode': 'smart_bundle',
        }
        reason = '默认先走“高点击封面 + 结果前置标题 + 共鸣正文”的通用爆款组合。'
    return {
        'recommended': recommended,
        'reason': reason,
        'traits': traits,
    }


def _recommended_direction_key(topic_text=''):
    text = str(topic_text or '')
    if any(token in text for token in ['中医', '肝气郁结', '肝郁', '疏肝理气', '肝胆湿热', '养肝']):
        return 'tcm_conditioning'
    if any(token in text for token in ['脂肪肝', '减脂', '代谢', '体重', '减肥']):
        return 'fatty_liver_management'
    if any(token in text for token in ['情绪', '焦虑', '压力', '失眠', '睡眠', '熬夜']):
        return 'emotion_mindbody'
    if any(token in text for token in ['饮食', '营养', '吃什么', '食谱', '控糖', '控油']):
        return 'diet_nutrition'
    if any(token in text for token in ['运动', '健身', '走路', '跑步', '减脂训练']):
        return 'exercise_fitness'
    if any(token in text for token in ['父母', '家人', '陪诊', '照护', '家属']):
        return 'family_care'
    if any(token in text for token in ['女性', '姨妈', '经期', '更年期']):
        return 'women_health'
    if any(token in text for token in ['误区', '别再', '为什么', '是不是', '搞错']):
        return 'myth_busting'
    if any(token in text for token in ['体检', '检查', '报告', '指标', 'FibroScan', '福波看', '肝弹', '转氨酶']):
        return 'report_interpretation'
    return 'liver_care_habits'


def _recommended_persona_key(traits, topic_text=''):
    text = str(topic_text or '')
    if any(token in text for token in ['父母', '家人', '陪诊', '照护', '家属']):
        return 'caregiver'
    if traits.get('report_like'):
        return 'physical_exam_blogger'
    if traits.get('myth_like'):
        return 'medical_science'
    if any(token in text for token in ['中医', '肝气郁结', '肝郁', '疏肝理气', '肝胆湿热']):
        return 'tcm_practitioner'
    if any(token in text for token in ['女性', '姨妈', '经期', '更年期']):
        return 'women_health'
    if any(token in text for token in ['脂肪肝', '减脂', '代谢', '体重', '减肥']):
        return 'health_manager'
    if any(token in text for token in ['运动', '健身', '走路', '跑步', '减脂训练']):
        return 'fitness_coach'
    if traits.get('emotion_like'):
        return 'emotional_support'
    if traits.get('discussion_like'):
        return 'patient_friend'
    if traits.get('story_like'):
        return 'patient_self'
    return 'doctor_assistant'


def _recommended_scene_key(traits, topic_text=''):
    text = str(topic_text or '')
    if any(token in text for token in ['父母', '家人', '陪诊', '照护', '家属']):
        return 'family_support'
    if traits.get('report_like'):
        return 'report_interpretation'
    if any(token in text for token in ['脂肪肝', '减脂', '代谢', '体重', '减肥']):
        return 'fatty_liver_management'
    if any(token in text for token in ['中医', '肝气郁结', '肝郁', '疏肝理气', '肝胆湿热']):
        return 'tcm_conditioning'
    if any(token in text for token in ['情绪', '焦虑', '压力']):
        return 'mood_stress'
    if any(token in text for token in ['失眠', '睡眠', '熬夜']):
        return 'sleep_recovery'
    if any(token in text for token in ['女性', '姨妈', '经期', '更年期']):
        return 'women_dailycare'
    if any(token in text for token in ['久坐', '盯屏', '办公室', '职场']):
        return 'office_sedentary'
    if any(token in text for token in ['运动', '健身', '跑步', '重启']):
        return 'exercise_rebuild'
    if traits.get('discussion_like'):
        return 'clinic_consulting'
    return 'daily_liver_care'


def _recommended_product_key(topic_text=''):
    topic_label, _ = _derive_topic_product_hint(topic_text)
    label_map = {
        '自动匹配': 'auto',
        '不植入产品': 'none',
        'FibroScan福波看': 'fibroscan',
        '复方鳖甲软肝片': 'soft_liver_tablet',
        '恩替卡韦联合管理': 'entecavir_combo',
        '壳脂胶囊': 'qizhi_capsule',
    }
    return label_map.get(topic_label, 'auto')


def _build_strategy_decision_profile(topic):
    topic_text = ' '.join(filter(None, [
        getattr(topic, 'topic_name', '') or '',
        getattr(topic, 'keywords', '') or '',
        getattr(topic, 'direction', '') or '',
    ]))
    traits = _detect_topic_strategy_traits(
        getattr(topic, 'topic_name', '') or '',
        getattr(topic, 'keywords', '') or '',
        getattr(topic, 'direction', '') or '',
    )
    direction_key = _recommended_direction_key(topic_text)
    persona_key = _recommended_persona_key(traits, topic_text)
    scene_key = _recommended_scene_key(traits, topic_text)
    product_key = _recommended_product_key(topic_text)
    product_label, _ = _resolve_copy_product_selection(product_key, topic_text)
    return {
        'persona_key': persona_key,
        'persona_label': COPY_PERSONA_OPTIONS.get(persona_key, COPY_PERSONA_OPTIONS.get('auto') or ''),
        'scene_key': scene_key,
        'scene_label': COPY_SCENE_OPTIONS.get(scene_key, COPY_SCENE_OPTIONS.get('auto') or ''),
        'direction_key': direction_key,
        'direction_label': COPY_DIRECTION_OPTIONS.get(direction_key, COPY_DIRECTION_OPTIONS.get('auto') or ''),
        'product_key': product_key,
        'product_label': product_label,
    }


def _strategy_row_relevance(row, traits):
    score = 0
    if traits['report_like']:
        if row.get('copy_goal') == 'save_value':
            score += 1
        if row.get('copy_skill') in {'report_interpretation', 'practical_checklist'}:
            score += 2
        if row.get('title_skill') == 'checklist_collect':
            score += 3
        if row.get('image_skill') == 'report_decode':
            score += 3
    if traits['discussion_like']:
        if row.get('copy_goal') == 'comment_engagement':
            score += 2
        if row.get('copy_skill') == 'discussion_hook':
            score += 2
        if row.get('title_skill') == 'question_gap':
            score += 3
        if row.get('image_skill') == 'story_atmosphere':
            score += 1
    if traits['myth_like']:
        if row.get('copy_goal') == 'viral_title':
            score += 2
        if row.get('copy_skill') == 'myth_busting':
            score += 2
        if row.get('title_skill') == 'conflict_reverse':
            score += 3
        if row.get('image_skill') == 'high_click_cover':
            score += 2
    if traits['story_like'] or traits['emotion_like']:
        if row.get('copy_skill') == 'story_empathy':
            score += 2
        if row.get('title_skill') == 'emotional_diary':
            score += 3
        if row.get('image_skill') == 'story_atmosphere':
            score += 2
    return score


def _build_strategy_skill_leaders(historical_rows, field, label_map):
    buckets = {}
    for row in historical_rows:
        raw_key = (row.get(field) or '').strip()
        if not raw_key:
            continue
        bucket = buckets.setdefault(raw_key, {
            'skill_key': raw_key,
            'skill_label': label_map.get(raw_key, raw_key),
            'count': 0,
            'views': 0,
            'interactions': 0,
            'viral_count': 0,
            'relevance': 0,
            'sample_titles': [],
        })
        bucket['count'] += row.get('count') or 0
        bucket['views'] += row.get('views') or 0
        bucket['interactions'] += row.get('interactions') or 0
        bucket['viral_count'] += row.get('viral_count') or 0
        bucket['relevance'] = max(bucket['relevance'], row.get('relevance') or 0)
        for sample in (row.get('sample_titles') or []):
            sample_text = str(sample or '').strip()
            if sample_text and sample_text not in bucket['sample_titles'] and len(bucket['sample_titles']) < 3:
                bucket['sample_titles'].append(sample_text)

    leaders = []
    for bucket in buckets.values():
        avg_interactions = round(bucket['interactions'] / bucket['count'], 2) if bucket['count'] else 0
        viral_rate = _calculate_rate(bucket['viral_count'], bucket['count']) or 0
        avg_interaction_rate = _calculate_rate(bucket['interactions'], bucket['views']) or 0
        ranking_score = (bucket['relevance'] * 800) + (avg_interactions * 10) + (viral_rate * 5) + (avg_interaction_rate * 2) + bucket['count']
        leaders.append({
            **bucket,
            'avg_interactions': avg_interactions,
            'viral_rate': viral_rate,
            'viral_rate_display': _format_rate(viral_rate),
            'avg_interaction_rate': avg_interaction_rate,
            'avg_interaction_rate_display': _format_rate(avg_interaction_rate),
            'ranking_score': round(ranking_score, 2),
        })
    leaders.sort(
        key=lambda item: (item['ranking_score'], item['count'], item['avg_interactions'], item['viral_rate']),
        reverse=True,
    )
    return leaders[:3]


def _build_recent_strategy_leader_snapshot(registration):
    payload = _build_strategy_recommendation_payload(registration)
    return {
        'source': payload.get('source') or 'heuristic',
        'confidence': payload.get('confidence') or 'low',
        'title_skill_leader': ((payload.get('title_skill_leaders') or [None])[0]) or {},
        'image_skill_leader': ((payload.get('image_skill_leaders') or [None])[0]) or {},
    }


def _build_strategy_recommendation_payload(registration):
    if not registration or not registration.topic:
        return {
            'success': False,
            'message': '报名信息不存在',
        }

    heuristic = _build_heuristic_strategy_recommendation(registration.topic)
    traits = heuristic['traits']
    base_recommended = dict(heuristic['recommended'])
    decision_profile = _build_strategy_decision_profile(registration.topic)
    base_recommended.update(decision_profile)
    submissions = Submission.query.filter(
        Submission.strategy_updated_at.isnot(None)
    ).order_by(
        Submission.strategy_updated_at.desc(),
        Submission.id.desc(),
    ).limit(500).all()

    strategy_rows = {}
    for sub in submissions:
        title_skill = (sub.selected_title_skill or '').strip()
        image_skill = (sub.selected_image_skill or '').strip()
        copy_skill = (sub.selected_copy_skill or '').strip()
        copy_goal = (sub.selected_copy_goal or '').strip()
        if not any([title_skill, image_skill, copy_skill, copy_goal]):
            continue
        key = '|'.join([
            copy_goal or '-',
            copy_skill or '-',
            title_skill or '-',
            image_skill or '-',
            (sub.selected_cover_style_type or '').strip() or '-',
            (sub.selected_inner_style_type or '').strip() or '-',
            (sub.selected_generation_mode or '').strip() or '-',
        ])
        row = strategy_rows.setdefault(key, {
            'copy_goal': copy_goal or base_recommended['copy_goal'],
            'copy_skill': copy_skill or base_recommended['copy_skill'],
            'title_skill': title_skill or base_recommended['title_skill'],
            'image_skill': image_skill or base_recommended['image_skill'],
            'cover_style_type': (sub.selected_cover_style_type or '').strip() or base_recommended['cover_style_type'],
            'inner_style_type': (sub.selected_inner_style_type or '').strip() or base_recommended['inner_style_type'],
            'generation_mode': (sub.selected_generation_mode or '').strip() or base_recommended['generation_mode'],
            'count': 0,
            'views': 0,
            'interactions': 0,
            'viral_count': 0,
            'sample_titles': [],
        })
        xhs_views = sub.xhs_views or 0
        xhs_interactions = (sub.xhs_likes or 0) + (sub.xhs_favorites or 0) + (sub.xhs_comments or 0)
        row['count'] += 1
        row['views'] += xhs_views
        row['interactions'] += xhs_interactions
        row['viral_count'] += 1 if (_submission_has_platform_link(sub, 'xhs') and _infer_viral_post(
            views=sub.xhs_views or 0,
            likes=sub.xhs_likes or 0,
            favorites=sub.xhs_favorites or 0,
            comments=sub.xhs_comments or 0,
        )) else 0
        if len(row['sample_titles']) < 3:
            row['sample_titles'].append(sub.selected_title or sub.note_title or '未命名标题')

    historical_rows = []
    for key, row in strategy_rows.items():
        avg_interactions = round(row['interactions'] / row['count'], 2) if row['count'] else 0
        viral_rate = _calculate_rate(row['viral_count'], row['count']) or 0
        avg_interaction_rate = _calculate_rate(row['interactions'], row['views']) or 0
        relevance = _strategy_row_relevance(row, traits)
        ranking_score = (relevance * 1000) + (avg_interactions * 10) + (viral_rate * 5) + (avg_interaction_rate * 2) + row['count']
        historical_rows.append({
            **row,
            'key': key,
            'avg_interactions': avg_interactions,
            'viral_rate': viral_rate,
            'viral_rate_display': _format_rate(viral_rate),
            'avg_interaction_rate': avg_interaction_rate,
            'avg_interaction_rate_display': _format_rate(avg_interaction_rate),
            'relevance': relevance,
            'ranking_score': round(ranking_score, 2),
        })

    historical_rows.sort(
        key=lambda item: (item['ranking_score'], item['count'], item['avg_interactions'], item['viral_rate']),
        reverse=True,
    )
    top_rows = historical_rows[:3]
    title_skill_leaders = _build_strategy_skill_leaders(historical_rows, 'title_skill', TITLE_SKILL_OPTIONS)
    image_skill_leaders = _build_strategy_skill_leaders(historical_rows, 'image_skill', IMAGE_SKILL_OPTIONS)

    source = 'heuristic'
    confidence = 'low'
    reason = heuristic['reason']
    recommended = dict(base_recommended)
    applied_historical_combo = False
    if top_rows:
        best_row = top_rows[0]
        if best_row['count'] >= 2 or best_row['relevance'] >= 3:
            applied_historical_combo = True
            source = 'historical'
            confidence = 'high' if best_row['count'] >= 3 and best_row['relevance'] >= 3 else 'medium'
            recommended.update({
                'copy_goal': best_row['copy_goal'] or recommended['copy_goal'],
                'copy_skill': best_row['copy_skill'] or recommended['copy_skill'],
                'title_skill': best_row['title_skill'] or recommended['title_skill'],
                'image_skill': best_row['image_skill'] or recommended['image_skill'],
                'cover_style_type': best_row['cover_style_type'] or recommended['cover_style_type'],
                'inner_style_type': best_row['inner_style_type'] or recommended['inner_style_type'],
                'generation_mode': best_row['generation_mode'] or recommended['generation_mode'],
            })
            reason = (
                f"参考近期待留痕样本，和当前话题更接近的组合里，“{best_row['title_skill']} × {best_row['image_skill']}”"
                f"共出现 {best_row['count']} 次，平均互动 {best_row['avg_interactions']}，爆款率 {best_row['viral_rate_display']}。"
            )
    if (not applied_historical_combo) and (title_skill_leaders or image_skill_leaders):
        hints = []
        title_leader = title_skill_leaders[0] if title_skill_leaders else None
        image_leader = image_skill_leaders[0] if image_skill_leaders else None
        if title_leader and title_leader['count'] >= 2 and title_leader['relevance'] >= 2:
            recommended['title_skill'] = title_leader['skill_key'] or recommended['title_skill']
            hints.append(f"标题更稳的是“{title_leader['skill_label']}”")
        if image_leader and image_leader['count'] >= 2 and image_leader['relevance'] >= 2:
            recommended['image_skill'] = image_leader['skill_key'] or recommended['image_skill']
            hints.append(f"图片更稳的是“{image_leader['skill_label']}”")
        if hints:
            source = 'hybrid'
            confidence = 'medium'
            reason = f"{heuristic['reason']} 另外参考近期单项表现，{'，'.join(hints)}。"

    return {
        'success': True,
        'source': source,
        'confidence': confidence,
        'reason': reason,
        'recommended': recommended,
        'heuristic': heuristic['recommended'],
        'decision_profile': decision_profile,
        'traits': {key: value for key, value in traits.items() if key not in {'raw_text', 'lowered'}},
        'historical_rows': [{
            'copy_goal': row['copy_goal'],
            'copy_skill': row['copy_skill'],
            'title_skill': row['title_skill'],
            'image_skill': row['image_skill'],
            'cover_style_type': row['cover_style_type'],
            'inner_style_type': row['inner_style_type'],
            'generation_mode': row['generation_mode'],
            'count': row['count'],
            'avg_interactions': row['avg_interactions'],
            'viral_rate_display': row['viral_rate_display'],
            'relevance': row['relevance'],
            'sample_titles': row['sample_titles'],
        } for row in top_rows],
        'title_skill_leaders': [{
            'skill_key': row['skill_key'],
            'skill_label': row['skill_label'],
            'count': row['count'],
            'avg_interactions': row['avg_interactions'],
            'viral_rate_display': row['viral_rate_display'],
            'relevance': row['relevance'],
            'sample_titles': row['sample_titles'],
        } for row in title_skill_leaders],
        'image_skill_leaders': [{
            'skill_key': row['skill_key'],
            'skill_label': row['skill_label'],
            'count': row['count'],
            'avg_interactions': row['avg_interactions'],
            'viral_rate_display': row['viral_rate_display'],
            'relevance': row['relevance'],
            'sample_titles': row['sample_titles'],
        } for row in image_skill_leaders],
    }


def _normalize_title_candidate(title):
    import re

    text = re.sub(r'^(标题\s*[：:]\s*)+', '', (title or '').strip()).strip('：: =-')
    text = re.sub(r'[#]+', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    if not text:
        return ''
    bad_title_tokens = ['带话题', '关键词', '版本', '封面', '互动型', '实拍', '笔记类型', '类型']
    if any(token in text for token in bad_title_tokens):
        return ''
    return text[:24]


def _score_title_clickability(title='', route_key='', copy_goal='balanced'):
    text = (title or '').strip()
    if not text:
        return 0
    score = 50
    length = len(text)
    if 10 <= length <= 18:
        score += 8
    elif 7 <= length <= 22:
        score += 4
    else:
        score -= 4

    bonus_tokens = ['先', '别', '很多人', '为什么', '我', '你', '报告', '体检', 'FibroScan', '这3', '三件事', '第一反应']
    score += sum(3 for token in bonus_tokens if token in text)
    if any(token in text for token in ['？', '会怎么', '怎么办', '你会']):
        score += 5
    if any(token in text for token in ['先别', '别再', '别只', '别急']):
        score += 5
    if any(token in text for token in ['清单', '收好', '先存', '只看这3步', '确认这3件事']):
        score += 4
    if text.startswith('关于'):
        score -= 8
    if '这件事' in text:
        score -= 5
    if '我后来' in text and route_key not in {'story_first', 'report_emotion'}:
        score -= 3

    if copy_goal == 'viral_title':
        if any(token in text for token in ['别', '很多人', '第一反应', '看偏了']):
            score += 6
    elif copy_goal == 'save_value':
        if any(token in text for token in ['清单', '收好', '先存', '确认这3件事']):
            score += 6
    elif copy_goal == 'comment_engagement':
        if any(token in text for token in ['你会', '怎么办', '怎么选', '？']):
            score += 6
    return score


def _build_title_option_pool(cards, title_skill_profile, topic, keywords='', copy_goal='balanced', selected_copy_route=None):
    import re

    keyword_source = re.sub(r'带话题|#', ' ', keywords or '')
    keyword_items = [item.strip() for item in re.split(r'[\s,，、/]+', keyword_source) if item.strip()]
    lead_keyword = next((item for item in keyword_items if 1 < len(item) <= 10), '') or (topic.topic_name or '护肝管理')
    if len(lead_keyword) > 12:
        lead_keyword = (topic.topic_name or lead_keyword)[:12]

    title_guidance = build_title_skill_local_guidance(
        title_skill_profile,
        lead_keyword=lead_keyword,
        topic_name=(topic.topic_name or lead_keyword),
    )
    guidance_titles = title_guidance.get('titles') or []

    goal_title_templates = []
    topic_text = ' '.join([(topic.topic_name or '').strip(), (keywords or '').strip()])
    report_like = bool(re.search(r'体检|检查|报告|指标|FibroScan|福波看|肝弹|转氨酶', topic_text, re.I))
    fibroscan_like = bool(re.search(r'FibroScan|福波看|肝弹', topic_text, re.I))
    hangover_like = bool(re.search(r'解酒|护肝|熬夜|应酬|喝酒', topic_text, re.I))
    route_key = (selected_copy_route or {}).get('id', '') if isinstance(selected_copy_route, dict) else ''
    if copy_goal == 'viral_title':
        goal_title_templates = [
            f'{lead_keyword}别再拖了',
            f'很多人把{lead_keyword}想简单了',
        ]
    elif copy_goal == 'save_value':
        goal_title_templates = [
            f'{lead_keyword}这几步别漏',
            f'关于{lead_keyword}，先存这份',
        ]
    elif copy_goal == 'comment_engagement':
        goal_title_templates = [
            f'{lead_keyword}这一步，你会怎么选',
            f'碰到{lead_keyword}，你会先做什么',
        ]
    elif copy_goal == 'trust_building':
        goal_title_templates = [
            f'{lead_keyword}先把逻辑看懂',
            f'关于{lead_keyword}，门诊更关注这几点',
        ]
    else:
        goal_title_templates = [
            f'{lead_keyword}先看这几点',
            f'关于{lead_keyword}，我后来才想明白',
        ]
    if report_like:
        goal_title_templates = [
            f'体检单里有{lead_keyword}，先别慌',
            f'{lead_keyword}别只盯这一项',
            f'看到{lead_keyword}，我先确认这3件事',
        ] + goal_title_templates
    if fibroscan_like:
        goal_title_templates = [
            '报告里有FibroScan，先别自己吓自己',
            'FibroScan这项检查，很多人第一眼就看偏了',
            '看到FibroScan，我现在会先看这几点',
            '体检单上这项，我以前也总看不懂',
            '看到FibroScan先别百度，我现在只看这3步',
            '报告里写了FibroScan，我后来先问了这3句',
        ] + goal_title_templates
    if hangover_like:
        goal_title_templates = [
            '别再把解酒当护肝了',
            '应酬后护肝，我现在先做这3件事',
            '喝完酒第二天，先别乱补',
            '护肝小妙招，最怕用错方向',
            '熬夜应酬后，我不再只找偏方',
        ] + goal_title_templates
    if route_key == 'report_emotion':
        goal_title_templates = [
            f'拿到{lead_keyword}这项时，我第一反应真是慌',
            f'{lead_keyword}写在报告上时，我后来先做了这件事',
            f'体检单出现{lead_keyword}，先别一个人吓自己',
        ] + goal_title_templates
    elif route_key == 'report_decode':
        goal_title_templates = [
            f'{lead_keyword}不是看见就严重，先看这几点',
            f'报告里这项很多人第一眼就看偏了',
            f'{lead_keyword}结果怎么读，我后来先改了这个习惯',
        ] + goal_title_templates
    elif route_key == 'report_checklist':
        goal_title_templates = [
            f'看到{lead_keyword}，我现在会先确认这3件事',
            f'体检单里出现{lead_keyword}，这份清单先收好',
            f'{lead_keyword}复查前后，我只盯这3个变化',
        ] + goal_title_templates
    elif route_key == 'story_first':
        goal_title_templates = [
            f'{lead_keyword}这件事，我是真的拖过',
            f'关于{lead_keyword}，我是后来才醒过来的',
            f'{lead_keyword}别再一个人硬扛了',
        ] + goal_title_templates
    elif route_key == 'qa_first':
        goal_title_templates = [
            f'碰到{lead_keyword}，你第一步会做什么',
            f'如果换成你，{lead_keyword}会怎么处理',
            f'{lead_keyword}这种情况，你会先怎么选',
        ] + goal_title_templates

    candidate_rows = []
    for index, card in enumerate(cards or []):
        candidate_rows.append({
            'title': card.get('title') or '',
            'source': f'正文版本 {index + 1}',
        })
    for item in guidance_titles:
        candidate_rows.append({
            'title': item,
            'source': f'{title_skill_profile.get("label") or "标题技能包"}',
        })
    for item in goal_title_templates:
        candidate_rows.append({
            'title': item,
            'source': '目标扩展',
        })
    if (topic.topic_name or '').strip():
        candidate_rows.append({
            'title': (topic.topic_name or '').strip()[:24],
            'source': '原始话题',
        })

    options = []
    seen = set()
    tone_label = {
        'report_emotion': '情绪共鸣',
        'report_decode': '解读拆解',
        'report_checklist': '收藏清单',
        'story_first': '经历分享',
        'qa_first': '问题互动',
    }.get(route_key, '')
    goal_reason = {
        'viral_title': '更偏点击',
        'save_value': '更偏收藏',
        'comment_engagement': '更偏互动',
        'trust_building': '更偏信任',
        'balanced': '更偏均衡',
    }.get(copy_goal, '更偏均衡')
    for row in candidate_rows:
        normalized_title = _normalize_title_candidate(row.get('title') or '')
        if not normalized_title:
            continue
        if normalized_title in seen:
            continue
        seen.add(normalized_title)
        source_label = row.get('source') or '系统推荐'
        reason = ''
        if source_label == '目标扩展':
            reason = f'当前{goal_reason}'
        elif '正文版本' in source_label:
            reason = '和当前正文更贴'
        elif '标题技能包' in source_label or title_skill_profile.get('label') == source_label:
            reason = '更贴当前标题打法'
        elif source_label == '原始话题':
            reason = '保留原话题表达'
        score = 78
        if source_label == '目标扩展':
            score = 90 if copy_goal == 'viral_title' else 86
        elif '正文版本' in source_label:
            score = 88
        elif '标题技能包' in source_label:
            score = 84
        score += _score_title_clickability(normalized_title, route_key=route_key, copy_goal=copy_goal)
        options.append({
            'title': normalized_title,
            'source': source_label,
            'tone': tone_label,
            'reason': reason,
            'score': score,
        })
    if not options:
        options.append({
            'title': '分享笔记',
            'source': '系统兜底',
            'tone': '',
            'reason': '兜底标题',
            'score': 60,
        })
    options.sort(key=lambda item: (item.get('score') or 0, item.get('source') == '正文版本 1'), reverse=True)
    return options[:5]


def _build_seed_plan_versions(route_plan, default_cards, output_count=3):
    route_plan = route_plan or {}
    output_count = min(max(_safe_int(output_count, 3), 1), 3)
    title_examples = list(route_plan.get('title_examples') or [])
    hook_example = (route_plan.get('hook_example') or '').strip()
    body_strategy = (route_plan.get('body_strategy') or '').strip()
    ending_direction = (route_plan.get('ending_direction') or '').strip()
    route_label = (route_plan.get('label') or '').strip()
    route_why = (route_plan.get('why') or '').strip()

    seeds = []
    for index in range(output_count):
        default_card = default_cards[min(index, len(default_cards) - 1)] if default_cards else {}
        title_hint = (title_examples[index % len(title_examples)] if title_examples else '').strip()
        seeds.append({
            'persona': (default_card.get('persona') or '').strip(),
            'scene': (default_card.get('scene') or '').strip(),
            'angle': title_hint or route_label or '按当前话题自然切入',
            'hook_focus': hook_example or '先用一个真实场景把人带进去',
            'body_focus': body_strategy or route_why or '先讲清判断逻辑，再给具体动作',
            'insertion_strategy': (default_card.get('insertion') or '').strip(),
            'ending_direction': ending_direction or (default_card.get('ending') or '').strip(),
        })
    return seeds


def _repair_copy_cards(
    cards,
    *,
    topic,
    keywords='',
    copy_goal='balanced',
    title_skill_profile=None,
    selected_copy_route=None,
    selected_product_label='',
    user_prompt='',
    default_cards=None,
):
    cards = list(cards or [])
    default_cards = default_cards or []
    if not cards:
        return cards

    keyword_source = re.sub(r'带话题|#', ' ', keywords or '')
    keyword_items = [item.strip() for item in re.split(r'[\s,，、/]+', keyword_source) if item.strip()]
    lead_keyword = next((item for item in keyword_items if 1 < len(item) <= 10), '') or (topic.topic_name or '护肝管理')
    if len(lead_keyword) > 12:
        lead_keyword = (topic.topic_name or lead_keyword)[:12]
    prompt_terms = _extract_prompt_terms(user_prompt)
    prompt_focus = prompt_terms[0] if prompt_terms else ''
    route_key = (selected_copy_route or {}).get('id', '') if isinstance(selected_copy_route, dict) else ''
    route_titles = list((selected_copy_route or {}).get('title_examples') or []) if isinstance(selected_copy_route, dict) else []
    route_hook_example = (selected_copy_route or {}).get('hook_example', '') if isinstance(selected_copy_route, dict) else ''
    route_body_strategy = (selected_copy_route or {}).get('body_strategy', '') if isinstance(selected_copy_route, dict) else ''

    title_pool = _build_title_option_pool(
        cards,
        title_skill_profile or {'label': '系统标题池'},
        topic,
        keywords=keywords,
        copy_goal=copy_goal,
        selected_copy_route=selected_copy_route,
    )
    candidate_titles = [item.get('title') for item in title_pool if (item.get('title') or '').strip()]
    used_titles = set()
    repaired = []

    for index, raw_card in enumerate(cards):
        defaults = default_cards[min(index, len(default_cards) - 1)] if default_cards else {}
        card = _parse_generated_copy_card(_render_generated_copy_card(raw_card), defaults=defaults)
        body = (card.get('body') or '').strip()
        hook = (card.get('hook') or '').strip()
        title = (card.get('title') or '').strip()

        if not hook and route_hook_example:
            hook = route_hook_example

        if len(body) < 140:
            supplemental_sections = _build_local_copy_body_sections(
                route_key,
                lead_keyword=lead_keyword,
                scene_text=card.get('scene') or defaults.get('scene') or '',
                persona_text=card.get('persona') or defaults.get('persona') or '',
                product_label=selected_product_label or card.get('insertion') or defaults.get('insertion') or '',
                prompt_focus=prompt_focus,
                route_body_strategy=route_body_strategy,
                reference_hint='',
                index=index,
            )
            existing_lines = [line.strip() for line in body.splitlines() if line.strip()]
            for section in supplemental_sections:
                if section not in existing_lines:
                    existing_lines.append(section)
                if len('\n'.join(existing_lines)) >= 180:
                    break
            body = '\n'.join(existing_lines).strip()

        if not title or len(title) < 6 or title in used_titles:
            replacement = ''
            for candidate in route_titles + candidate_titles:
                normalized = _normalize_title_candidate(candidate or '')
                if normalized and normalized not in used_titles:
                    replacement = normalized
                    break
            if replacement:
                title = replacement

        if title in used_titles:
            title = f'{title} {index + 1}'.strip()
        title = _normalize_title_candidate(title) or f'分享笔记 {index + 1}'
        if len(title) > 24:
            title = title[:24]
        used_titles.add(title)

        card['title'] = title
        card['hook'] = hook or (body.splitlines()[0].strip()[:36] if body else '')
        card['body'] = body
        if not (card.get('ending') or '').strip():
            card['ending'] = (defaults.get('ending') or '你们会怎么做？').strip()
        card['copy_text'] = _render_generated_copy_card(card)
        repaired.append(card)

    for index in range(1, len(repaired)):
        current = repaired[index]
        previous = repaired[index - 1]
        if _text_similarity(current.get('body') or '', previous.get('body') or '') > 0.84:
            defaults = default_cards[min(index, len(default_cards) - 1)] if default_cards else {}
            sections = _build_local_copy_body_sections(
                route_key,
                lead_keyword=lead_keyword,
                scene_text=current.get('scene') or defaults.get('scene') or '',
                persona_text=current.get('persona') or defaults.get('persona') or '',
                product_label=selected_product_label or current.get('insertion') or defaults.get('insertion') or '',
                prompt_focus=prompt_focus,
                route_body_strategy=route_body_strategy,
                reference_hint='',
                index=index + 2,
            )
            current['body'] = '\n'.join(sections).strip()
            if route_hook_example:
                current['hook'] = route_hook_example
            current['copy_text'] = _render_generated_copy_card(current)

    return repaired


def _copy_card_ai_tone_penalty(text=''):
    content = str(text or '')
    penalty_tokens = [
        '首先', '其次', '综上', '建议大家', '值得注意的是', '围绕', '维度', '逻辑上',
        '从xx身份出发', '顺手带出', '最容易写空泛', '展开说明', '以下几点',
    ]
    penalty = 0
    for token in penalty_tokens:
        if token in content:
            penalty += 4
    return penalty


def _score_copy_card_quality(card, *, route_key='', copy_goal='balanced', sibling_cards=None):
    title = (card.get('title') or '').strip()
    hook = (card.get('hook') or '').strip()
    body = (card.get('body') or '').strip()
    score = 60
    reasons = []

    title_score = _score_title_clickability(title, route_key=route_key, copy_goal=copy_goal)
    score += max(min(title_score - 50, 28), -12)
    if title_score >= 72:
        reasons.append('标题抓眼')
    elif title_score < 56:
        reasons.append('标题偏平')

    body_length = len(body)
    if 170 <= body_length <= 320:
        score += 10
        reasons.append('正文长度合适')
    elif 130 <= body_length <= 360:
        score += 4
    else:
        score -= 10
        reasons.append('正文过短或过长')

    if hook and hook in body[:80]:
        score += 4
        reasons.append('开头进入场景快')
    if any(token in body for token in ['我后来', '那次', '当时', '先看', '先问清', '先把', '我现在']):
        score += 5
        reasons.append('动作感更强')

    ai_penalty = _copy_card_ai_tone_penalty(f'{hook}\n{body}')
    score -= ai_penalty
    if ai_penalty >= 6:
        reasons.append('AI腔偏重')

    sibling_cards = sibling_cards or []
    if sibling_cards:
        max_similarity = max(
            (_text_similarity(body, (item.get('body') or '').strip()) for item in sibling_cards if item is not card),
            default=0.0,
        )
        if max_similarity > 0.84:
            score -= 10
            reasons.append('和其他版本太像')
        elif max_similarity < 0.65:
            score += 4
            reasons.append('差异度更好')

    return {
        'score': max(min(int(score), 99), 35),
        'reasons': reasons[:3],
    }


def _rerank_copy_cards(cards, *, route_key='', copy_goal='balanced'):
    cards = list(cards or [])
    if not cards:
        return cards

    scored = []
    for index, card in enumerate(cards):
        quality = _score_copy_card_quality(card, route_key=route_key, copy_goal=copy_goal, sibling_cards=cards)
        enriched = dict(card)
        enriched['quality_score'] = quality['score']
        enriched['quality_reasons'] = quality['reasons']
        enriched['copy_text'] = _render_generated_copy_card(enriched)
        scored.append((index, enriched))

    scored.sort(
        key=lambda item: (
            item[1].get('quality_score') or 0,
            _score_title_clickability(item[1].get('title') or '', route_key=route_key, copy_goal=copy_goal),
            -item[0],
        ),
        reverse=True,
    )
    return [item[1] for item in scored]


def _fallback_copy_agent_routes(topic, topic_text='', keywords='', persona_label='', scene_label='', direction_label='', product_label='', copy_goal='balanced'):
    traits = _detect_topic_strategy_traits(topic.topic_name or topic_text, keywords, direction_label)
    format_traits = _detect_topic_format_traits(' '.join([topic.topic_name or topic_text, keywords or '', direction_label or '']))
    lead_keyword = _split_keywords(keywords or topic_text or topic.topic_name or '')
    lead_keyword = (lead_keyword[0] if lead_keyword else (topic.topic_name or '这件事'))[:12]
    persona = persona_label or '患者本人'
    scene = scene_label or '体检异常提醒'

    def route_item(key, label, why, title_examples, hook_example, body_strategy, ending_direction, image_hint):
        return {
            'id': key,
            'label': label,
            'why': why,
            'title_examples': title_examples[:3],
            'hook_example': hook_example,
            'body_strategy': body_strategy,
            'ending_direction': ending_direction,
            'image_hint': image_hint,
            'persona_hint': persona,
            'scene_hint': scene,
        }

    if format_traits['professional_like'] or format_traits['chart_like'] or format_traits['report_photo_like']:
        return [
            route_item(
                'professional_decode',
                '先把专业判断翻译成人话',
                '适合专业背景、文献解读、图表说明和报告翻译，重点是把难点讲明白。',
                [f'{lead_keyword}先把逻辑看懂', f'关于{lead_keyword}，我会先拆这3个指标', f'{lead_keyword}这份图表先看这里'],
                '如果只看结论，很容易把这件事想简单。真正有用的是先把判断顺序理清。',
                '正文像门诊解释或科普笔记：先拆核心概念，再讲判断顺序，最后给具体建议。',
                '结尾更适合问“你们最容易卡在哪个指标或概念？”',
                '图片适合图表卡、报告解读卡、知识卡片。',
            ),
            route_item(
                'professional_checklist',
                '先给用户一份可收藏清单',
                '适合把复杂信息压缩成 3-5 个关键点，方便收藏和复用。',
                [f'{lead_keyword}这份清单先收好', f'碰到{lead_keyword}，我会先看这几项', f'{lead_keyword}别只看表面，先确认这3点'],
                '我会先把真正影响判断的几个点列出来，不然很多人看完还是会慌。',
                '正文按清单推进：先看什么、怎么理解、接下来做什么。',
                '结尾更适合问“哪一项是你最想让医生解释清楚的？”',
                '图片适合清单卡、对照表、报告逐项解读。',
            ),
            route_item(
                'professional_compare',
                '先做横向对比或红黑榜',
                '适合测评、对比、红黑榜和“不同方案怎么选”的内容。',
                [f'{lead_keyword}怎么选，我会先对比这几项', f'{lead_keyword}别乱选，这张对比表先看', f'关于{lead_keyword}，红黑榜我只看这几点'],
                '如果只讲一个方案，用户很难判断。先把差异摆出来，阅读效率会更高。',
                '正文优先用对比、优缺点和适用场景讲清楚，再给结论。',
                '结尾更适合问“如果是你，你更在意哪一项差异？”',
                '图片适合图表卡、对比卡、知识卡片。',
            ),
        ]

    if format_traits['checklist_like']:
        return [
            route_item(
                'checklist_first',
                '先给一份能直接保存的清单',
                '适合备忘录、攻略、体检项目清单和“照着做”的内容。',
                [f'{lead_keyword}先存这份清单', f'{lead_keyword}这几步别漏', f'关于{lead_keyword}，我会先按这个顺序来'],
                '这篇我不想讲虚的，先把顺序和重点列出来，方便你直接照着做。',
                '正文按步骤和清单推进，每条都尽量给动作和判断标准。',
                '结尾更适合问“你们还会补哪一项？”',
                '图片适合备忘录、清单卡、对照表。',
            ),
            route_item(
                'checklist_budget',
                '先从预算和优先级切入',
                '适合“基础版 / 升级版 / 全面版”这类方案拆分。',
                [f'{lead_keyword}预算不同，我会这么排', f'{lead_keyword}怎么取舍，这张表先收好', f'{lead_keyword}想少花冤枉钱，先看优先级'],
                '很多人不是不愿意做，而是不知道预算该放在哪。先帮他把优先级排出来。',
                '正文用“先必做、再可选、最后升级项”讲清楚。',
                '结尾更适合问“如果预算有限，你最先保留哪一项？”',
                '图片适合价格对照表、预算清单卡。',
            ),
            route_item(
                'checklist_story',
                '先讲一次真实准备过程',
                '适合女儿/家属/陪诊者视角，把清单做得更有代入感。',
                [f'给爸妈准备{lead_keyword}时，我先做了这份清单', f'{lead_keyword}那天，我最怕漏掉这几项', f'如果再来一次，{lead_keyword}我会先这么排'],
                '先讲那次准备或体检时的慌乱，再把最后用到的清单交出来。',
                '正文先经历后清单，兼顾真实感和实用性。',
                '结尾更适合问“你们带父母体检最怕漏什么？”',
                '图片适合备忘录 / 清单卡 / 课堂笔记。',
            ),
        ]

    if format_traits['poster_like']:
        return [
            route_item(
                'poster_hook',
                '先打一条能让人停下来的大问题',
                '适合大字报封面和互动型内容，先让人停下来再展开。',
                [f'{lead_keyword}这件事，你会怎么选', f'{lead_keyword}先别急着下结论', f'很多人对{lead_keyword}第一反应都错了'],
                '先把用户最纠结的那个问题打出来，不急着给标准答案。',
                '正文像讨论贴：先抛问题，再给经历或判断，最后把评论区打开。',
                '结尾更适合问“如果是你，你会怎么做？”',
                '图片适合大字封面，内页再换知识卡。',
            ),
            route_item(
                'poster_story',
                '先讲一个会共鸣的瞬间',
                '适合实拍+大字报混合，用情绪和经历把评论区拉起来。',
                [f'那次因为{lead_keyword}，我真的慌了', f'{lead_keyword}这件事，我差点也拖过去', f'关于{lead_keyword}，我后来不敢再赌'],
                '那一下的真实反应，比空讲道理更能让人代入。',
                '正文先讲瞬间，再讲为什么，最后给动作。',
                '结尾更适合问“你以前也有过这种反应吗？”',
                '图片适合大字封面 + 备忘录内页。',
            ),
            route_item(
                'poster_reverse',
                '先纠偏，再抛互动问题',
                '适合误区、大众认知偏差和讨论型场景。',
                [f'{lead_keyword}很多人第一步就想错了', f'关于{lead_keyword}，最容易误会的是这里', f'{lead_keyword}别再按老办法想了'],
                '如果先讲经历太慢，这条就直接先纠偏，再给评论区留空间。',
                '正文先说误区，再讲正确理解，最后抛一个问题收尾。',
                '结尾更适合问“你以前是不是也这么想？”',
                '图片适合大字报和知识卡片。',
            ),
        ]

    if traits['report_like']:
        return [
            route_item(
                'report_emotion',
                '先接住“看到报告那一下慌”',
                '检查/报告类内容，如果先讲结论，容易像说教；先把那一刻的情绪写出来，更像真人分享。',
                ['报告里写了FibroScan，我第一反应是慌', '看到FibroScan，我后来先问了这3句', 'FibroScan这项检查，先别自己吓自己'],
                '拿到报告那一刻，我真正慌的不是这几个字，而是一时不知道先看什么。',
                '先写拿到报告的停顿和误判，再拆“不能只盯一个词/一项数字”的原因，最后给下一步动作。',
                '结尾更适合问“你们复查时最先看哪一项？”',
                '图片更适合报告解读卡或知识卡，不适合硬做大字报。',
            ),
            route_item(
                'report_decode',
                '先拆最容易看偏的点',
                '这条路线更适合做收藏和转发，重点不是情绪，而是把判断顺序讲清楚。',
                ['FibroScan不是看见就严重，先看这几点', '报告里有FibroScan，很多人第一眼就看偏了', 'FibroScan结果怎么读，我后来先改了这个习惯'],
                '我后来才发现，很多人一看到报告里有这项检查，第一反应就已经跑偏了。',
                '正文像“解释给朋友听”：先说最常见误解，再说真正该确认的点，最后落到复查动作。',
                '结尾更适合问“如果是你，会先问医生什么？”',
                '图片适合报告解读卡、检查流程卡、知识卡片。',
            ),
            route_item(
                'report_checklist',
                '直接给下一步清单',
                '适合冲收藏，重点是让人看完知道下一步该干什么，而不是只记住名词。',
                ['看到FibroScan，我现在会先确认这3件事', '体检单里出现FibroScan，这份清单先收好', 'FibroScan复查前后，我只盯这3个变化'],
                '我后来给自己定的第一件事，不是继续搜，而是先把下一步要确认的问题列出来。',
                '正文按步骤推进：先看前后变化，再问清检查目的，再定复查时间，最后再补个人经验。',
                '结尾更适合问“你们一般是先复查，还是先继续观察？”',
                '图片适合清单卡、时间轴清单、报告解读卡。',
            ),
        ]

    if traits['myth_like']:
        return [
            route_item(
                'myth_reverse',
                '先纠正常见误解',
                '适合冲点击，但要像真人提醒，不像标题党。',
                [f'很多人把{lead_keyword}看偏了', f'{lead_keyword}这件事，别再想当然了', f'关于{lead_keyword}，最容易误会的其实是前面这一步'],
                f'我后来才发现，大家对“{lead_keyword}”最常见的误解，恰好就是我自己以前也信过的。',
                '先讲自己以前怎么误会，再拆为什么错，最后给正确动作。',
                '结尾更适合问“你以前是不是也这么想？”',
                '图片适合大字封面 + 知识卡片，不适合太多复杂结构。',
            ),
            route_item(
                'myth_case',
                '先讲一个差点踩坑的瞬间',
                '这条更有代入感，适合让用户先停下来。',
                [f'{lead_keyword}这件事，我差点也理解错了', f'那次看到{lead_keyword}，我一下警惕了', f'关于{lead_keyword}，我是差点踩坑才懂'],
                '那次我其实差一点就按自己以前那套理解走下去了，后来才发现问题不在表面。',
                '先讲踩坑瞬间，再讲真正应该怎么看，最后给用户一个更稳的判断方法。',
                '结尾更适合问“如果是你，当时会怎么理解？”',
                '图片适合备忘录/陪伴卡或大字封面。',
            ),
            route_item(
                'myth_checklist',
                '先给“别再这样理解”的清单',
                '更适合收藏型误区内容，不用情绪太满。',
                [f'{lead_keyword}这几种理解最容易出错', f'碰到{lead_keyword}，先别这么想', f'关于{lead_keyword}，这几件事最容易搞反'],
                '如果重新来一次，我会先把最容易搞反的几件事写下来，不然真的很容易越看越乱。',
                '用 3 个误区 + 3 个正解来组织，信息感更强。',
                '结尾更适合问“你最容易卡在哪个误区？”',
                '图片适合误区对照卡、知识卡片。',
            ),
        ]

    return [
        route_item(
            'story_first',
            '先讲真实经历和感受',
            '更像真人说话，适合先建立代入感。',
            [f'{lead_keyword}这件事，我是真的拖过', f'关于{lead_keyword}，我是后来才醒过来的', f'{lead_keyword}别再一个人硬扛了'],
            f'我以前对“{lead_keyword}”的理解其实挺模糊的，真碰到的时候才知道这事不能硬拖。',
            '先讲自己当时怎么想、怎么拖、后来怎么改，正文重点放具体动作。',
            '结尾更适合问“如果是你，会先从哪一步开始？”',
            '图片适合备忘录/陪伴卡或课堂笔记卡。',
        ),
        route_item(
            'decode_first',
            '先把关键点拆清楚',
            '适合做信任和收藏，不容易显得像硬广。',
            [f'{lead_keyword}先别急，先看这几点', f'关于{lead_keyword}，我现在会先改这一点', f'{lead_keyword}这几步真的别漏'],
            f'后来我才知道，“{lead_keyword}”最怕的不是复杂，而是自己脑补太多。',
            '正文先解释关键点，再给行动顺序，最后轻轻带出产品或服务。',
            '结尾更适合问“你更想先解决哪一步？”',
            '图片适合知识卡片或清单卡。',
        ),
        route_item(
            'qa_first',
            '先从用户最常问的问题切入',
            '更适合把互动打开，也更容易贴近搜索场景。',
            [f'碰到{lead_keyword}，你第一步会做什么', f'{lead_keyword}这种情况，你会先怎么选', f'如果换成你，{lead_keyword}会怎么处理'],
            f'每次聊到“{lead_keyword}”，我都发现大家最纠结的其实不是答案，而是先后顺序。',
            '正文先写问题，再给两个判断口径，最后落到你自己的做法。',
            '结尾更适合问“换成你，你会怎么选？”',
            '图片适合课堂笔记卡、清单卡。',
        ),
    ]


def _fallback_image_agent_routes(topic_text='', copy_routes=None):
    traits = _detect_topic_strategy_traits(topic_text, topic_text, topic_text)
    format_traits = _detect_topic_format_traits(topic_text)
    routes = []

    def add_route(key, label, why, family_key, image_skill, cover_style_type, inner_style_type, preview_focus):
        routes.append({
            'id': key,
            'label': label,
            'why': why,
            'family_key': family_key,
            'image_skill': image_skill,
            'mode_key': 'smart_bundle',
            'cover_style_type': cover_style_type,
            'inner_style_type': inner_style_type,
            'preview_focus': preview_focus,
        })

    if format_traits['professional_like'] or format_traits['chart_like'] or format_traits['report_photo_like']:
        add_route('report_decode', '报告解读卡', '适合深度科普、图表说明、报告翻译和指标拆解。', 'medical_science', 'report_decode', 'medical_science', 'checklist_report', '封面先讲“这张图要看哪”，内页拆指标、对照和下一步动作。')
        add_route('knowledge_decode', '知识图表卡', '适合横向测评、概念解释和科研/文献转人话。', 'knowledge_card', 'classroom_focus', 'knowledge_card', 'knowledge_card', '封面一句话点题，内页拆模块或对比。')
        add_route('checklist_table', '对照清单卡', '适合预算对比、项目清单、红黑榜和流程图。', 'checklist', 'save_worthy_cards', 'checklist_table', 'checklist_table', '封面做对照句，内页用表格或步骤承接。')
        return routes
    if format_traits['checklist_like']:
        add_route('checklist_cards', '清单卡', '适合备忘录、项目清单、攻略和截图收藏型内容。', 'checklist', 'save_worthy_cards', 'checklist', 'checklist_table', '封面先做问题句，内页用步骤、表格或优先级列表。')
        add_route('memo_story', '备忘录/课堂笔记', '适合把清单写得更像真实陪同经验或复盘。', 'memo', 'story_atmosphere', 'memo_mobile', 'memo_classroom', '封面像手机备忘录，内页像手写笔记。')
        add_route('medical_science', '医学解释图', '适合把清单背后的原因和指标关系讲清楚。', 'medical_science', 'report_decode', 'medical_science', 'knowledge_card', '封面突出重点项目，内页再解释原因。')
        return routes
    if format_traits['poster_like']:
        add_route('poster_click', '大字封面', '适合先冲点击和互动，封面用一句强问题或强提醒。', 'poster', 'high_click_cover', 'poster_bold', 'knowledge_card', '封面只打一条问题句或纠偏句，内页换成知识卡。')
        add_route('memo_story', '陪伴感卡片', '适合封面冲点击后，内页继续用真实经历接住。', 'memo', 'story_atmosphere', 'poster_handwritten', 'memo_classroom', '封面情绪更强，内页更像复盘。')
        add_route('knowledge_card', '知识卡片', '适合把互动型问题往可收藏解释型内容过渡。', 'knowledge_card', 'classroom_focus', 'knowledge_card', 'knowledge_card', '封面一句话，内页拆 3 个判断点。')
        return routes
    if traits['report_like']:
        add_route('report_decode', '报告解读卡', '适合看懂报告、指标和复查顺序，收藏和转发都会更稳。', 'medical_science', 'report_decode', 'medical_science', 'checklist_report', '封面突出“先别慌/先看哪项”，内页拆指标、趋势和下一步动作。')
        add_route('knowledge_decode', '知识卡片', '适合把最容易看偏的点讲清楚，比大字报更有信息密度。', 'knowledge_card', 'classroom_focus', 'knowledge_card', 'knowledge_card', '封面放一个核心判断，内页拆 3 个关键点。')
        add_route('checklist_action', '清单卡', '适合直接给“接下来怎么做”的执行清单。', 'checklist', 'save_worthy_cards', 'checklist', 'checklist_report', '封面做问题句，内页做步骤或报告解读表。')
        return routes
    if traits['story_like'] or traits['emotion_like']:
        add_route('memo_story', '备忘录/陪伴卡', '适合第一人称经历和情绪转折，更像真人分享。', 'memo', 'story_atmosphere', 'memo_mobile', 'memo_classroom', '封面像手机备忘录，内页像课堂笔记或复盘。')
        add_route('poster_story', '大字封面', '适合把那一下情绪和提醒先打出来，但内页仍要讲清楚。', 'poster', 'high_click_cover', 'poster_handwritten', 'knowledge_card', '封面抓情绪，内页拆原因和动作。')
        add_route('knowledge_story', '课堂笔记卡', '适合边讲经历边讲判断逻辑，信息更稳。', 'knowledge_card', 'classroom_focus', 'knowledge_card', 'memo_classroom', '封面不要太满，内页更像老师划重点。')
        return routes
    add_route('poster_click', '大字封面', '适合冲点击，但只适合做封面，不适合整套都走大字报。', 'poster', 'high_click_cover', 'poster_bold', 'knowledge_card', '封面只打一条强结论，内页必须换成知识卡。')
    add_route('knowledge_card', '知识卡片', '适合大多数解释型内容，稳定、不容易翻车。', 'knowledge_card', 'classroom_focus', 'knowledge_card', 'knowledge_card', '封面一句话，内页拆结构。')
    add_route('checklist_cards', '清单卡', '适合“3点/步骤/不要漏”的收藏型内容。', 'checklist', 'save_worthy_cards', 'checklist', 'checklist_table', '封面做问题句，内页做清单/对照表。')
    return routes


def _build_copy_agent_analysis(topic, *, user_prompt='', persona_label='', scene_label='', direction_label='', product_label='', copy_goal='balanced'):
    topic_text = ' '.join(filter(None, [topic.topic_name or '', topic.keywords or '', direction_label or '']))
    routes = _fallback_copy_agent_routes(
        topic,
        topic_text=topic_text,
        keywords=topic.keywords or '',
        persona_label=persona_label,
        scene_label=scene_label,
        direction_label=direction_label,
        product_label=product_label,
        copy_goal=copy_goal,
    )
    image_routes = _fallback_image_agent_routes(topic_text=topic_text, copy_routes=routes)
    return {
        'summary': 'Agent 已先把内容写法和图片路线拆开，你先选路线，再生成文案和图片，会比直接一键乱出更稳。',
        'recommended_copy_route_id': routes[0]['id'] if routes else '',
        'recommended_image_route_id': image_routes[0]['id'] if image_routes else '',
        'copy_routes': routes,
        'image_routes': image_routes,
    }


def _image_agent_template_label(style_key=''):
    key = (style_key or '').strip()
    if key in {'medical_science', 'flowchart'}:
        return '医学解释图'
    if key in {'knowledge_card', 'myth_compare'}:
        return '知识卡片'
    if key in {'poster', 'poster_bold', 'poster_handwritten'}:
        return '大字封面'
    if key in {'checklist', 'checklist_table', 'checklist_timeline', 'checklist_report'}:
        return '清单卡'
    if key in {'memo', 'memo_mobile', 'memo_classroom'}:
        return '备忘录/课堂笔记'
    return '图文方案'


def _build_image_agent_analysis_payload(topic, *, selected_content='', title_hint='', preferred_route=None):
    recommendation_payload = _build_asset_style_recommendation_payload({
        'title_hint': title_hint,
        'selected_content': selected_content,
    })
    recommendation_items = recommendation_payload.get('items') or []
    preferred_route = preferred_route or {}
    plan_rows = []
    seen = set()

    def append_plan(style_key, reason='', image_skill='', route_label='', preview_focus=''):
        style_meta = _asset_style_meta(style_key)
        family_key = style_meta.get('family') or 'knowledge_card'
        family_meta = {
            'medical_science': '更适合检查解读、器官说明和“先看哪项”这类视觉说明书',
            'knowledge_card': '更适合机制拆解、误区纠偏和重点卡片',
            'poster': '更适合封面先抓眼球，但只建议封面强冲击，内页还是要讲清楚',
            'checklist': '更适合步骤、对照、报告解读和收藏型清单',
            'memo': '更适合经历分享、陪伴感和课堂笔记',
        }.get(family_key, '更适合做成小红书图文卡片')
        cover_style_type = style_key
        inner_style_type = preferred_route.get('inner_style_type') if preferred_route.get('cover_style_type') == style_key else ''
        if not inner_style_type:
            if family_key == 'checklist':
                inner_style_type = 'checklist_report'
            elif family_key == 'poster':
                inner_style_type = 'knowledge_card'
            elif family_key == 'memo':
                inner_style_type = 'memo_classroom'
            elif family_key == 'medical_science':
                inner_style_type = 'knowledge_card'
            else:
                inner_style_type = style_key
        dedupe_key = (cover_style_type, inner_style_type)
        if dedupe_key in seen:
            return
        seen.add(dedupe_key)
        cover_focus = {
            'medical_science': '封面重点是“先看什么/为什么重要”，不要做成广告牌。',
            'knowledge_card': '封面一句话讲清主题，不要堆太多字。',
            'poster': '封面只打一条强提醒，大字抓眼球，别把整篇内容都塞上去。',
            'checklist': '封面先抛问题，内页再用步骤/对照表接住。',
            'memo': '封面更像真实备忘录或课堂笔记，不要太像海报。',
        }.get(family_key, '封面先抓住一个核心问题。')
        inner_focus = {
            'medical_science': '内页适合拆器官示意、误区对照、检查解释。',
            'knowledge_card': '内页适合拆 3 个判断点或 3 个误区。',
            'poster': '内页一定要切成知识卡，不要整套都走大字报。',
            'checklist': '内页适合清单、对照、时间轴、报告逐项解读。',
            'memo': '内页适合笔记感、复盘感、陪伴式解释。',
        }.get(family_key, '内页要接住信息，不要只做装饰。')
        cover_fit = _score_cover_suitability(
            style_key=cover_style_type,
            family_key=family_key,
            generation_mode=preferred_route.get('mode_key') or 'smart_bundle',
            selected_content=selected_content,
            title_hint=title_hint,
        )
        plan_rows.append({
            'id': f'{cover_style_type}__{inner_style_type}',
            'label': _image_agent_template_label(cover_style_type),
            'style_label': style_meta.get('label') or cover_style_type,
            'family_key': family_key,
            'why': reason or family_meta,
            'image_skill': image_skill or (preferred_route.get('image_skill') or ''),
            'generation_mode': preferred_route.get('mode_key') or 'smart_bundle',
            'cover_style_type': cover_style_type,
            'inner_style_type': inner_style_type,
            'preview_focus': preview_focus or family_meta,
            'cover_focus': cover_focus,
            'inner_focus': inner_focus,
            'route_label': route_label or '',
            'cover_fit_score': cover_fit['score'],
            'cover_fit_label': cover_fit['label'],
            'cover_fit_reason': cover_fit['reason'],
            'fallback_cover_style_key': cover_fit['fallback_style_key'],
            'fallback_cover_style_label': cover_fit['fallback_style_label'],
            'execution_note': cover_fit['execution_note'],
        })

    if preferred_route.get('cover_style_type'):
        append_plan(
            preferred_route.get('cover_style_type'),
            reason=preferred_route.get('why') or preferred_route.get('preview_focus') or '',
            image_skill=preferred_route.get('image_skill') or '',
            route_label=preferred_route.get('label') or '',
            preview_focus=preferred_route.get('preview_focus') or '',
        )

    for item in recommendation_items[:4]:
        append_plan(
            item.get('style_key') or '',
            reason=item.get('reason') or '',
            image_skill=preferred_route.get('image_skill') or '',
            route_label=preferred_route.get('label') or '',
        )

    recommended_plan = max(
        plan_rows,
        key=lambda item: (item.get('cover_fit_score') or 0, item.get('cover_style_type') == (preferred_route.get('cover_style_type') or ''))
    ) if plan_rows else {}

    return {
        'success': True,
        'summary': '图片 Agent 已先根据当前正文和标题，拆出几条可执行的出图路线。先选路线，再预览素材或图文套组，会比直接乱出更稳。',
        'plans': plan_rows[:4],
        'recommended_plan_id': (recommended_plan.get('id') or ''),
    }


def _task_agent_metric_focus(copy_goal='balanced'):
    key = (copy_goal or 'balanced').strip() or 'balanced'
    focus_map = {
        'viral_title': ['先看阅读量', '再看前 2 小时互动速度', '标题和封面不要同时大改'],
        'save_value': ['重点看收藏量', '同时看收藏率和评论里的“先存了”反馈', '清单型内容尽量保留信息密度'],
        'comment_engagement': ['重点看评论量', '优先观察用户是不是愿意接话', '评论区问题要及时回复'],
        'trust_building': ['重点看收藏和高质量评论', '观察是否有人追问细节', '专业表达稳定比冲点击更重要'],
        'balanced': ['同步看阅读、收藏、评论三项', '不要只盯单个爆点指标', '先复用有效组合，再细调变量'],
    }
    return focus_map.get(key, focus_map['balanced'])


def _task_agent_compliance_hint(topic, traits):
    text = ' '.join(filter(None, [
        getattr(topic, 'topic_name', '') or '',
        getattr(topic, 'keywords', '') or '',
        getattr(topic, 'direction', '') or '',
    ]))
    if traits.get('report_like'):
        return '检查解读类内容不要把单个指标直接等同严重程度，建议写成“先看什么、再确认什么”。'
    if any(token in text for token in ['药', '复方鳖甲', '恩替卡韦', '胶囊', '软肝片']):
        return '涉及药品时避免绝对化疗效和替代诊疗表达，产品只做弱植入和经验型带出。'
    if traits.get('myth_like'):
        return '误区纠偏类内容要避免“千万”“一定”“绝对”这类过强判断，结论尽量留给检查和医生建议。'
    return '默认按医疗健康内容处理：少下结论、多讲判断顺序和下一步动作。'


def _build_task_agent_brief_payload(registration):
    if not registration or not registration.topic:
        return {
            'success': False,
            'message': '报名信息不存在',
        }

    topic = registration.topic
    submission = registration.submission
    strategy_payload = _build_strategy_recommendation_payload(registration)
    recommended = dict(strategy_payload.get('recommended') or strategy_payload.get('heuristic') or {})
    decision_profile = strategy_payload.get('decision_profile') or {}
    traits = strategy_payload.get('traits') or {}

    persona_key = (decision_profile.get('persona_key') or recommended.get('persona_key') or 'auto').strip() or 'auto'
    scene_key = (decision_profile.get('scene_key') or recommended.get('scene_key') or 'auto').strip() or 'auto'
    direction_key = (decision_profile.get('direction_key') or recommended.get('direction_key') or 'auto').strip() or 'auto'
    product_key = (decision_profile.get('product_key') or recommended.get('product_key') or 'auto').strip() or 'auto'
    copy_goal = (recommended.get('copy_goal') or 'balanced').strip() or 'balanced'
    copy_skill = (recommended.get('copy_skill') or 'auto').strip() or 'auto'
    title_skill = (recommended.get('title_skill') or 'auto').strip() or 'auto'
    image_skill = (recommended.get('image_skill') or 'auto').strip() or 'auto'
    generation_mode = (recommended.get('generation_mode') or 'smart_bundle').strip() or 'smart_bundle'

    persona_label = COPY_PERSONA_OPTIONS.get(persona_key, COPY_PERSONA_OPTIONS.get('auto') or '')
    scene_label = COPY_SCENE_OPTIONS.get(scene_key, COPY_SCENE_OPTIONS.get('auto') or '')
    direction_label = COPY_DIRECTION_OPTIONS.get(direction_key, COPY_DIRECTION_OPTIONS.get('auto') or '')
    product_label, _ = _resolve_copy_product_selection(product_key, ' '.join(filter(None, [
        topic.topic_name or '',
        topic.keywords or '',
        topic.direction or '',
    ])))

    copy_analysis = _build_copy_agent_analysis(
        topic,
        persona_label=persona_label,
        scene_label=scene_label,
        direction_label=direction_label,
        product_label=product_label,
        copy_goal=copy_goal,
    )
    selected_copy_route = next(
        (item for item in (copy_analysis.get('copy_routes') or []) if item.get('id') == (copy_analysis.get('recommended_copy_route_id') or '')),
        None,
    ) or ((copy_analysis.get('copy_routes') or [None])[0] or {})
    selected_image_route = next(
        (item for item in (copy_analysis.get('image_routes') or []) if item.get('id') == (copy_analysis.get('recommended_image_route_id') or '')),
        None,
    ) or ((copy_analysis.get('image_routes') or [None])[0] or {})

    brief_selected_content = '\n'.join(filter(None, [
        topic.direction or '',
        topic.reference_content or '',
        topic.writing_example or '',
        f"推荐目标：{COPY_GOAL_OPTIONS.get(copy_goal, copy_goal)}",
        f"文案打法：{COPY_SKILL_OPTIONS.get(copy_skill, copy_skill)}",
        f"图片打法：{IMAGE_SKILL_OPTIONS.get(image_skill, image_skill)}",
    ])).strip()[:4000]
    brief_title_hint = ((selected_copy_route.get('title_examples') or [topic.topic_name or ''])[0] or topic.topic_name or '').strip()[:200]
    image_analysis = _build_image_agent_analysis_payload(
        topic,
        selected_content=brief_selected_content,
        title_hint=brief_title_hint,
        preferred_route=selected_image_route,
    )
    selected_image_plan = next(
        (item for item in (image_analysis.get('plans') or []) if item.get('id') == (image_analysis.get('recommended_plan_id') or '')),
        None,
    ) or ((image_analysis.get('plans') or [None])[0] or {})

    has_strategy = bool(submission and (submission.strategy_payload or '').strip())
    has_copy = bool(submission and (submission.selected_copy_text or '').strip())
    has_link = bool(submission and any([
        (submission.xhs_link or '').strip(),
        (submission.douyin_link or '').strip(),
        (submission.video_link or '').strip(),
        (submission.weibo_link or '').strip(),
    ]))

    action_steps = [
        {'key': 'strategy', 'label': '应用任务建议', 'done': has_strategy},
        {'key': 'copy', 'label': '生成文案', 'done': has_copy},
        {'key': 'image', 'label': '确认图片方案', 'done': bool(has_strategy and (submission.selected_image_skill or '').strip())},
        {'key': 'publish', 'label': '发布并提交链接', 'done': has_link},
        {'key': 'review', 'label': '回看数据再优化', 'done': bool(submission and (submission.xhs_views or 0) > 0)},
    ]

    professional_advice = [
        f"这条任务先按“{COPY_GOAL_OPTIONS.get(copy_goal, copy_goal)} + {COPY_SKILL_OPTIONS.get(copy_skill, copy_skill)} + {IMAGE_SKILL_OPTIONS.get(image_skill, image_skill)}”执行，先别同时改太多变量。",
        f"当前更推荐的文案路线是“{selected_copy_route.get('label') or '默认路线'}”，图片先走“{selected_image_route.get('label') or '默认路线'}”，这样更容易拿到第一版可发内容。",
        _task_agent_compliance_hint(topic, traits),
    ]

    return {
        'success': True,
        'summary': f"系统已经先帮你定好一版推荐打法：{COPY_GOAL_OPTIONS.get(copy_goal, copy_goal)}，优先减少犹豫和来回试错。",
        'source': strategy_payload.get('source') or 'heuristic',
        'confidence': strategy_payload.get('confidence') or 'low',
        'reason': strategy_payload.get('reason') or '',
        'preset': {
            'personaKey': persona_key,
            'sceneKey': scene_key,
            'directionKey': direction_key,
            'productKey': product_key,
            'copyGoal': copy_goal,
            'copySkill': copy_skill,
            'titleSkill': title_skill,
            'imageSkill': image_skill,
            'generation_mode': generation_mode,
            'cover_style_type': recommended.get('cover_style_type') or '',
            'inner_style_type': recommended.get('inner_style_type') or '',
        },
        'labels': {
            'persona': persona_label,
            'scene': scene_label,
            'direction': direction_label,
            'product': product_label,
            'copy_goal': COPY_GOAL_OPTIONS.get(copy_goal, copy_goal),
            'copy_skill': COPY_SKILL_OPTIONS.get(copy_skill, copy_skill),
            'title_skill': TITLE_SKILL_OPTIONS.get(title_skill, title_skill),
            'image_skill': IMAGE_SKILL_OPTIONS.get(image_skill, image_skill),
        },
        'copy_route': {
            'id': selected_copy_route.get('id') or '',
            'label': selected_copy_route.get('label') or '',
            'why': selected_copy_route.get('why') or '',
            'hook_example': selected_copy_route.get('hook_example') or '',
        },
        'image_route': {
            'id': selected_image_route.get('id') or '',
            'label': selected_image_route.get('label') or '',
            'why': selected_image_route.get('why') or '',
        },
        'image_plan': {
            'id': selected_image_plan.get('id') or '',
            'label': selected_image_plan.get('label') or '',
            'cover_fit_label': selected_image_plan.get('cover_fit_label') or '',
            'cover_fit_score': selected_image_plan.get('cover_fit_score') or 0,
            'execution_note': selected_image_plan.get('execution_note') or '',
        },
        'metric_focus': _task_agent_metric_focus(copy_goal),
        'professional_advice': professional_advice,
        'action_steps': action_steps,
    }


def _build_reference_image_analysis_payload(topic, *, selected_content='', title_hint='', reference_asset_ids=None):
    reference_rows = _resolve_reference_asset_rows(reference_asset_ids or '', limit=6)
    if not reference_rows:
        return {
            'success': False,
            'message': '请先选 1-3 张参考图，再让 Agent 分析。',
            'plans': [],
            'references': [],
        }

    merged = ' '.join(filter(None, [topic.topic_name or '', topic.keywords or '', selected_content or '', title_hint or '']))
    traits = _detect_topic_strategy_traits(merged, merged, merged)
    references = []
    reference_titles = []
    families = []
    for item in reference_rows:
        style_meta = _asset_style_meta(item.style_type_key or '')
        reference_titles.append(item.title or style_meta.get('label') or '参考图')
        families.append(style_meta.get('family') or 'reference_based')
        references.append({
            'id': item.id,
            'title': item.title or '未命名参考图',
            'subtitle': item.subtitle or '',
            'preview_url': item.preview_url or '',
            'style_type_key': item.style_type_key or '',
            'style_type_label': style_meta.get('label') or item.style_type_key or '',
            'library_type': item.library_type or '',
            'tags': [part.strip() for part in (item.tags or '').split(',') if part.strip()][:6],
        })

    dominant_family = families[0] if families else 'reference_based'
    if traits.get('report_like'):
        inner_style = 'checklist_report'
        summary_hint = '这批参考图会更适合走“参考图气质 + 报告解读结构”'
    elif traits.get('story_like') or traits.get('emotion_like'):
        inner_style = 'memo_classroom'
        summary_hint = '这批参考图会更适合走“参考图气质 + 课堂笔记/陪伴感内页”'
    else:
        inner_style = 'knowledge_card'
        summary_hint = '这批参考图会更适合走“参考图气质 + 知识卡片内页”'

    reference_follow_fit = _score_cover_suitability(
        style_key='reference_based',
        family_key='custom',
        generation_mode='smart_bundle',
        selected_content=selected_content,
        title_hint=title_hint,
        reference_guided=True,
    )
    reference_science_fit = _score_cover_suitability(
        style_key='reference_based',
        family_key='medical_science',
        generation_mode='smart_bundle',
        selected_content=selected_content,
        title_hint=title_hint,
        reference_guided=True,
    )
    reference_collect_fit = _score_cover_suitability(
        style_key='reference_based',
        family_key='checklist',
        generation_mode='smart_bundle',
        selected_content=selected_content,
        title_hint=title_hint,
        reference_guided=True,
    )

    plan_rows = [
        {
            'id': 'reference_follow',
            'label': '贴近参考图气质',
            'why': '优先继承你选中参考图的构图、留白和色彩气质，但不会直接照抄原图。',
            'family_key': 'custom',
            'cover_style_type': 'reference_based',
            'inner_style_type': inner_style,
            'cover_focus': '封面优先保留参考图的版心结构、视觉节奏和留白位置。',
            'inner_focus': '内页按当前内容改写成更适合阅读的知识/报告结构，不会生搬硬套原图。',
            'reference_logic': summary_hint,
            'inherit_points': ['构图和留白', '色彩气质', '信息块节奏'],
            'avoid_points': ['不要照抄原图文案', '不要保留原图水印/品牌', '不要把原图直接拼贴上去'],
            'cover_fit_score': reference_follow_fit['score'],
            'cover_fit_label': reference_follow_fit['label'],
            'cover_fit_reason': reference_follow_fit['reason'],
            'fallback_cover_style_key': reference_follow_fit['fallback_style_key'],
            'fallback_cover_style_label': reference_follow_fit['fallback_style_label'],
            'execution_note': reference_follow_fit['execution_note'],
        },
        {
            'id': 'reference_plus_science',
            'label': '参考图 + 医学解释卡',
            'why': '保留参考图气质，同时把信息结构做得更适合医学科普和检查解读。',
            'family_key': 'custom',
            'cover_style_type': 'reference_based',
            'inner_style_type': 'medical_science' if dominant_family == 'medical_science' else inner_style,
            'cover_focus': '封面继续贴近参考图，但中部主体更强调解释关系和结构层级。',
            'inner_focus': '内页更像说明书/拆解卡，适合把“是什么、怎么看、下一步做什么”讲清楚。',
            'reference_logic': '适合你给的是医学科普封面、知识图解、器官示意这类参考图。',
            'inherit_points': ['主体构图', '医学信息图语气', '局部放大/箭头关系'],
            'avoid_points': ['不要做成纯商业海报', '不要只剩插画没有信息层级', '不要把标签堆太满'],
            'cover_fit_score': reference_science_fit['score'],
            'cover_fit_label': reference_science_fit['label'],
            'cover_fit_reason': reference_science_fit['reason'],
            'fallback_cover_style_key': reference_science_fit['fallback_style_key'],
            'fallback_cover_style_label': reference_science_fit['fallback_style_label'],
            'execution_note': reference_science_fit['execution_note'],
        },
        {
            'id': 'reference_plus_collect',
            'label': '参考图 + 收藏型清单',
            'why': '如果你更想要小红书收藏感，这条会保留参考图气质，同时把信息收成步骤/清单。',
            'family_key': 'custom',
            'cover_style_type': 'reference_based',
            'inner_style_type': 'checklist_report' if traits.get('report_like') else 'checklist_table',
            'cover_focus': '封面继续借参考图的感觉，但标题和问题句会更直接。',
            'inner_focus': '内页会更偏清单、对照或步骤，适合用户截图保存。',
            'reference_logic': '适合你给的是风格参考，但最终仍希望内容更像可收藏的知识卡或报告解读卡。',
            'inherit_points': ['留白和版心', '配色和质感', '封面氛围'],
            'avoid_points': ['不要把封面做得太满', '不要把清单做成大段文字', '不要把参考图原样照搬'],
            'cover_fit_score': reference_collect_fit['score'],
            'cover_fit_label': reference_collect_fit['label'],
            'cover_fit_reason': reference_collect_fit['reason'],
            'fallback_cover_style_key': reference_collect_fit['fallback_style_key'],
            'fallback_cover_style_label': reference_collect_fit['fallback_style_label'],
            'execution_note': reference_collect_fit['execution_note'],
        },
    ]

    return {
        'success': True,
        'summary': f'已分析 {len(references)} 张参考图：{(" / ".join(reference_titles[:3]))}。系统会先学它们的构图、留白和色彩气质，再按当前内容改写成更适合的小红书图文。',
        'references': references,
        'plans': plan_rows,
        'recommended_plan_id': 'reference_follow',
    }


@app.route('/api/copy_agent_analysis', methods=['POST'])
def copy_agent_analysis():
    data = request.json or {}
    registration_id = data.get('registration_id')
    reg = Registration.query.get(registration_id)
    if not reg:
        return jsonify({'success': False, 'message': '报名信息不存在'})

    topic = reg.topic
    persona_key = (data.get('persona_key') or 'auto').strip()
    custom_persona = (data.get('custom_persona') or '').strip()
    scene_key = (data.get('scene_key') or 'auto').strip()
    custom_scene = (data.get('custom_scene') or '').strip()
    direction_key = (data.get('direction_key') or 'auto').strip()
    custom_direction = (data.get('custom_direction') or '').strip()
    product_key = (data.get('product_key') or 'auto').strip()
    copy_goal = (data.get('copy_goal') or 'balanced').strip()
    user_prompt = (data.get('user_prompt') or '').strip()

    selected_persona_label = _resolve_copy_selection(COPY_PERSONA_OPTIONS, persona_key, custom_value=custom_persona)
    selected_scene_label = _resolve_copy_selection(COPY_SCENE_OPTIONS, scene_key, custom_value=custom_scene)
    selected_direction_label = _resolve_copy_direction_selection(direction_key, topic_text=(topic.topic_name or ''), user_prompt=user_prompt, custom_value=custom_direction)
    selected_product_label, _ = _resolve_copy_product_selection(product_key, topic.topic_name or '')

    analysis = _build_copy_agent_analysis(
        topic,
        user_prompt=user_prompt,
        persona_label=selected_persona_label,
        scene_label=selected_scene_label,
        direction_label=selected_direction_label,
        product_label=selected_product_label,
        copy_goal=copy_goal,
    )
    return jsonify({
        'success': True,
        'analysis': analysis,
    })


@app.route('/api/image_agent_analysis', methods=['POST'])
def image_agent_analysis():
    data = request.json or {}
    registration_id = data.get('registration_id')
    reg = Registration.query.get(registration_id)
    if not reg:
        return jsonify({'success': False, 'message': '报名信息不存在'})

    topic = reg.topic
    preferred_route = {
        'id': (data.get('image_route_id') or '').strip(),
        'label': (data.get('image_route_label') or '').strip(),
        'why': (data.get('image_route_why') or '').strip(),
        'image_skill': (data.get('image_skill') or '').strip(),
        'mode_key': (data.get('mode_key') or '').strip(),
        'cover_style_type': (data.get('cover_style_type') or '').strip(),
        'inner_style_type': (data.get('inner_style_type') or '').strip(),
        'preview_focus': (data.get('preview_focus') or '').strip(),
    }
    payload = _build_image_agent_analysis_payload(
        topic,
        selected_content=(data.get('selected_content') or '').strip(),
        title_hint=(data.get('title_hint') or '').strip(),
        preferred_route=preferred_route,
    )
    return jsonify(payload)


@app.route('/api/reference_image_analysis', methods=['POST'])
def reference_image_analysis():
    data = request.json or {}
    registration_id = data.get('registration_id')
    reg = Registration.query.get(registration_id)
    if not reg:
        return jsonify({'success': False, 'message': '报名信息不存在'})

    topic = reg.topic
    payload = _build_reference_image_analysis_payload(
        topic,
        selected_content=(data.get('selected_content') or '').strip(),
        title_hint=(data.get('title_hint') or '').strip(),
        reference_asset_ids=data.get('reference_asset_ids') or '',
    )
    return jsonify(payload)

@app.route('/api/generate_copy', methods=['POST'])
def generate_copy():
    data = request.json or {}
    registration_id = data.get('registration_id')
    user_prompt = (data.get('user_prompt') or '').strip()
    fast_mode = bool(data.get('fast_mode', True))
    persona_key = (data.get('persona_key') or 'auto').strip()
    custom_persona = (data.get('custom_persona') or '').strip()
    scene_key = (data.get('scene_key') or 'auto').strip()
    custom_scene = (data.get('custom_scene') or '').strip()
    direction_key = (data.get('direction_key') or 'auto').strip()
    custom_direction = (data.get('custom_direction') or '').strip()
    product_key = (data.get('product_key') or 'auto').strip()
    copy_goal = (data.get('copy_goal') or 'balanced').strip()
    copy_skill = (data.get('copy_skill') or 'auto').strip()
    title_skill = (data.get('title_skill') or 'auto').strip()
    agent_copy_route_id = (data.get('agent_copy_route_id') or '').strip()
    agent_image_route_id = (data.get('agent_image_route_id') or '').strip()
    local_only = bool(data.get('local_only') or data.get('force_local_agent'))

    reg = Registration.query.get(registration_id)
    if not reg:
        return jsonify({'success': False, 'message': '报名信息不存在'})

    topic = reg.topic
    keywords = topic.keywords or ''
    direction = topic.direction or ''
    direction_clean = direction.strip()
    if (not direction_clean) or direction_clean in ['同上', '同上。', '同上 ', '同上（同前）']:
        direction_clean = ''

    import random
    import re
    import time
    started_at = time.time()

    output_count = 3
    topic_text = (topic.topic_name or '').strip()
    generation_id = f"G{int(time.time()*1000)}-{random.randint(1000,9999)}"

    recent_snippets = []
    try:
        setting_item = Settings.query.filter_by(key='recent_copy_snippets').first()
        if setting_item and setting_item.value:
            recent_snippets = json.loads(setting_item.value)
            if not isinstance(recent_snippets, list):
                recent_snippets = []
    except Exception:
        recent_snippets = []

    auto_persona_pool = [
        '医生助理', '健管师', '营养师', '中医调理视角', '运动减脂教练',
        '患者本人', '患者家属', '医学科普', '患者朋友', '职场久坐人群',
        '女性健康视角', '情绪陪伴者', '慢病管理者', '体检复盘博主'
    ]
    auto_scene_pool = [
        '门诊沟通答疑', '体检异常提醒', '复查对比变化', '日常护肝管理',
        '熬夜应酬后护肝', '家属陪伴照护', '饮食调整建议', '检查报告解读',
        '中医调理养肝', '脂肪肝减脂管理', '情绪压力与肝气', '失眠熬夜恢复期',
        '女性日常养肝', '久坐久盯屏场景', '运动减脂重启期'
    ]
    default_endings = [
        '你们会怎么做？',
        '有类似经历的人可以说说吗？',
        '如果是你，你会先从哪一步开始？',
        '你们复查时最在意哪个指标？',
        '这种情况你们一般会怎么跟家里人解释？',
    ]

    selected_persona_label = _resolve_copy_selection(COPY_PERSONA_OPTIONS, persona_key, custom_value=custom_persona)
    selected_scene_label = _resolve_copy_selection(COPY_SCENE_OPTIONS, scene_key, custom_value=custom_scene)
    selected_direction_label = _resolve_copy_direction_selection(direction_key, topic_text=topic_text, user_prompt=user_prompt, custom_value=custom_direction)
    selected_product_label, product_hint = _resolve_copy_product_selection(product_key, topic_text)
    goal_profile = _goal_strategy_profile(copy_goal)
    skill_profile = resolve_copy_skill(copy_skill, topic_text=topic_text, copy_goal=copy_goal)
    skill_prompt_block = build_copy_skill_prompt_block(skill_profile)
    title_skill_profile = resolve_title_skill(title_skill, topic_text=topic_text, copy_goal=copy_goal, copy_skill_key=skill_profile.get('key') or '')
    title_skill_prompt_block = build_title_skill_prompt_block(title_skill_profile)
    agent_analysis = _build_copy_agent_analysis(
        topic,
        user_prompt=user_prompt,
        persona_label=selected_persona_label,
        scene_label=selected_scene_label,
        direction_label=selected_direction_label,
        product_label=selected_product_label,
        copy_goal=copy_goal,
    )
    selected_copy_route = next((item for item in (agent_analysis.get('copy_routes') or []) if item.get('id') == agent_copy_route_id), None)
    if not selected_copy_route:
        selected_copy_route = next(
            (item for item in (agent_analysis.get('copy_routes') or []) if item.get('id') == (agent_analysis.get('recommended_copy_route_id') or '')),
            None,
        ) or ((agent_analysis.get('copy_routes') or [None])[0])
    selected_image_route = next((item for item in (agent_analysis.get('image_routes') or []) if item.get('id') == agent_image_route_id), None)
    if not selected_image_route:
        selected_image_route = next(
            (item for item in (agent_analysis.get('image_routes') or []) if item.get('id') == (agent_analysis.get('recommended_image_route_id') or '')),
            None,
        ) or ((agent_analysis.get('image_routes') or [None])[0])
    strategy_recommendation_payload = _build_strategy_recommendation_payload(reg)
    title_skill_leader = ((strategy_recommendation_payload.get('title_skill_leaders') or [None])[0]) or {}
    image_skill_leader = ((strategy_recommendation_payload.get('image_skill_leaders') or [None])[0]) or {}
    recommendation_hint_lines = []
    if strategy_recommendation_payload.get('source') in {'historical', 'hybrid'}:
        recommendation_hint_lines.append(f"当前推荐来源：{strategy_recommendation_payload.get('source')} ｜ 置信度：{strategy_recommendation_payload.get('confidence') or 'low'}。")
    if title_skill_leader.get('skill_label'):
        recommendation_hint_lines.append(
            f"近期标题冠军：{title_skill_leader.get('skill_label')}（出现 {title_skill_leader.get('count') or 0} 次，平均互动 {title_skill_leader.get('avg_interactions') or 0}，爆款率 {title_skill_leader.get('viral_rate_display') or '-'}）。"
        )
    if image_skill_leader.get('skill_label'):
        recommendation_hint_lines.append(
            f"近期图片冠军：{image_skill_leader.get('skill_label')}（出现 {image_skill_leader.get('count') or 0} 次，平均互动 {image_skill_leader.get('avg_interactions') or 0}，爆款率 {image_skill_leader.get('viral_rate_display') or '-'}）。"
        )
    recommendation_hint_block = '\n'.join(recommendation_hint_lines) if recommendation_hint_lines else '暂无可用的近期冠军数据，按当前规则推荐生成。'
    reference_corpus_entries = _matching_corpus_snippets(','.join([topic_text, keywords]), limit=4)
    reference_corpus_block = _build_generate_copy_corpus_block(reference_corpus_entries, product_hint=product_hint)

    version_personas = random.sample(auto_persona_pool, output_count) if persona_key == 'auto' else [selected_persona_label] * output_count
    version_scenes = random.sample(auto_scene_pool, output_count) if scene_key == 'auto' else [selected_scene_label] * output_count
    skill_default_endings = build_copy_skill_local_guidance(skill_profile, lead_keyword=topic_text or '护肝管理').get('endings') or []
    ending_pool = skill_default_endings or default_endings
    if len(ending_pool) >= output_count:
        version_endings = random.sample(ending_pool, output_count)
    else:
        version_endings = [ending_pool[i % len(ending_pool)] for i in range(output_count)]
    goal_variants = goal_profile.get('variants') or ['故事共鸣版', '轻科普拆解版', '互动讨论版']
    skill_variants = skill_profile.get('variants') or ['技能版']
    version_styles = [
        f"{skill_variants[i % len(skill_variants)]} / {goal_variants[i % len(goal_variants)]}"
        for i in range(output_count)
    ]

    role_scene_block = '\n'.join([
        f"- 版本{i+1}：人设={version_personas[i]} ｜ 场景={version_scenes[i]} ｜ 风格={version_styles[i % len(version_styles)]}"
        for i in range(output_count)
    ])

    knowledge_hint = ''
    try:
        with open('/home/node/.openclaw/workspace/knowledge/xhs_viral_templates.md', 'r', encoding='utf-8') as f:
            knowledge_hint = f.read()[-2500:] if fast_mode else f.read()[-5000:]
    except Exception:
        knowledge_hint = ''

    prompt = f"""你是一个非常懂小红书内容运营的医疗健康创作者教练，现在要为报名人生成 3 篇“千人千面”的小红书文案。

【当前话题】
{topic.topic_name}

【关键词】
{keywords or '无'}

【撰写说明】
{direction_clean or '无，按话题自动理解'}

【用户自定义提词器】
{user_prompt or '无，按系统默认策略生成'}

【本次内容方向】
{selected_direction_label}

【软植入要求】
{product_hint}

【本轮目标】
{goal_profile['label']}。{goal_profile['title_rule']}

【写作技能包】
{skill_prompt_block}

【标题技能包】
{title_skill_prompt_block}

【近期策略冠军提示】
{recommendation_hint_block}

【可仿写模板语料】（只学结构和节奏，不得照抄）
{reference_corpus_block}

【每个版本的人设/场景/风格】
{role_scene_block}

【Agent已选写作路线】
{json.dumps(selected_copy_route, ensure_ascii=False) if selected_copy_route else '未显式选择，按默认推荐路线生成'}

【参考语料】（学习结构和语气，不得照抄）
{knowledge_hint or '无'}

【必须做到】
1. 每一篇都要有爆款标题，标题 8-16 字，不能太像广告。
2. 每一篇都要有“开头钩子”，第一句话就要把用户带进来。
3. 正文必须有真实细节、情绪、场景和科普信息，不要空话套话。
4. 软植入要自然，只能像真实经验里顺手带出，不能像硬广。
5. 结尾要自然引导互动，不能出现“评论区”等生硬引导词。
6. 三个版本必须明显不同，不能只换几个词；要在人设、场景、切入角度、标题逻辑上拉开差异。
7. 医疗健康内容必须合规，不说绝对化词，不承诺疗效，不诱导购买。
8. 语言像真人，不要“首先/其次/综上/建议大家”等 AI 腔。

【输出格式】
===版本1===
人设：
场景：
软植入：
标题：
开头钩子：
正文：
互动结尾：

===版本2===
人设：
场景：
软植入：
标题：
开头钩子：
正文：
互动结尾：

===版本3===
人设：
场景：
软植入：
标题：
开头钩子：
正文：
互动结尾：
"""

    default_cards = [
        {
            'persona': version_personas[i],
            'scene': version_scenes[i],
            'insertion': selected_product_label,
            'ending': version_endings[i],
        }
        for i in range(output_count)
    ]

    cards = []
    generation_engine = 'local_fallback'
    generation_runtime = None
    agent_mode = 'rule_based'
    model_attempted = False
    model_error_message = ''
    try:
        runtime = _copywriter_runtime_config()
        if (not local_only) and runtime.get('configured'):
            generation_engine = runtime['provider']
            model_attempted = True
            if fast_mode:
                single_pass_result = _call_copywriter(
                    [
                        {'role': 'system', 'content': f'你先严格执行当前写作技能包，再保证贴题、真实口语化和差异化。禁止跑题，禁止硬广。当前技能包：{skill_profile["label"]}。'},
                        {'role': 'user', 'content': prompt}
                    ],
                    temperature=1.08,
                    top_p=0.88,
                    timeout=8,
                    extra_payload={
                        'presence_penalty': 0.65,
                        'frequency_penalty': 0.25,
                    }
                )
                generation_runtime = single_pass_result.get('runtime') or generation_runtime
                generation_engine = (generation_runtime or {}).get('provider') or generation_engine
                agent_mode = 'single_pass'
                content = single_pass_result.get('text') or ''
                parts = re.split(r'===版本\s*\d+\s*===', content)
                for index, part in enumerate(parts):
                    if len(cards) >= output_count:
                        break
                    version_text = (part or '').strip()
                    if not version_text:
                        continue
                    card = _parse_generated_copy_card(version_text, defaults=default_cards[min(index, output_count - 1)])
                    if not (card.get('body') or '').strip():
                        continue
                    cards.append(card)
            else:
                seed_plan_versions = _build_seed_plan_versions(selected_copy_route, default_cards, output_count=output_count)
                planning_prompt = f"""你现在是小红书内容规划 Agent，不直接写正文，先给出 3 条明确不同的写作方案。

请只输出 JSON，不要输出解释。结构如下：
{{
  "versions": [
    {{
      "persona": "",
      "scene": "",
      "angle": "",
      "hook_focus": "",
      "body_focus": "",
      "insertion_strategy": "",
      "ending_direction": ""
    }}
  ]
}}

要求：
1. 必须给 3 个版本
2. 三个版本差异要明显，不是换同义词
3. 版本要贴合当前话题、人设、场景、内容目标
4. 不要写官话，不要写空泛描述
5. 优先参考下面的系统种子方案，如果你有更好的写法可以覆盖，但不能把 3 个版本写成同一路线。

当前话题：{topic.topic_name}
关键词：{keywords or '无'}
内容方向：{selected_direction_label}
内容目标：{goal_profile['label']}
写作技能包：{skill_profile['label']}
标题技能包：{title_skill_profile['label']}
用户补充要求：{user_prompt or '无'}
软植入要求：{product_hint}
近期策略冠军提示：
{recommendation_hint_block}
版本默认人设/场景：
{role_scene_block}

系统种子方案：
{json.dumps({'versions': seed_plan_versions}, ensure_ascii=False, indent=2)}
"""
                planning_result = _call_copywriter(
                    [
                        {'role': 'system', 'content': '你是一个会先做内容规划，再动笔写作的小红书内容策划 Agent。'},
                        {'role': 'user', 'content': planning_prompt},
                    ],
                    temperature=0.72,
                    top_p=0.82,
                    timeout=8,
                )
                generation_runtime = planning_result.get('runtime') or generation_runtime
                generation_engine = (generation_runtime or {}).get('provider') or generation_engine
                plan_json = _extract_json_object(planning_result.get('text') or '')
                plan_versions = (plan_json.get('versions') or []) if isinstance(plan_json, dict) else []
                normalized_plans = []
                for index in range(output_count):
                    seed_plan = seed_plan_versions[index] if index < len(seed_plan_versions) else {}
                    current_plan = plan_versions[index] if index < len(plan_versions) and isinstance(plan_versions[index], dict) else {}
                    normalized_plans.append({
                        'persona': (current_plan.get('persona') or seed_plan.get('persona') or default_cards[index]['persona']).strip(),
                        'scene': (current_plan.get('scene') or seed_plan.get('scene') or default_cards[index]['scene']).strip(),
                        'angle': (current_plan.get('angle') or seed_plan.get('angle') or '').strip(),
                        'hook_focus': (current_plan.get('hook_focus') or seed_plan.get('hook_focus') or '').strip(),
                        'body_focus': (current_plan.get('body_focus') or seed_plan.get('body_focus') or '').strip(),
                        'insertion_strategy': (current_plan.get('insertion_strategy') or seed_plan.get('insertion_strategy') or selected_product_label).strip(),
                        'ending_direction': (current_plan.get('ending_direction') or seed_plan.get('ending_direction') or default_cards[index]['ending']).strip(),
                    })
                writing_plan_block = '\n'.join([
                    (
                        f"版本{i+1}：人设={row['persona']} ｜ 场景={row['scene']} ｜ 切入={row['angle'] or '按当前话题自然切入'} ｜ "
                        f"钩子重点={row['hook_focus'] or '先接住用户情绪'} ｜ 正文重点={row['body_focus'] or '讲清具体动作'} ｜ "
                        f"软植入方式={row['insertion_strategy'] or selected_product_label} ｜ 互动方向={row['ending_direction'] or default_cards[i]['ending']}"
                    )
                    for i, row in enumerate(normalized_plans)
                ])
                writing_prompt = f"""你现在是小红书写作 Agent。请严格按下面 3 条规划分别写出 3 个版本。

写作规划：
{writing_plan_block}

共通要求：
1. 每个版本都要像真人在说话
2. 不要写“我是从xx身份出发”“围绕xx最容易写空泛”“我会把xx顺手带出”这类模板句
3. 一开头就进入具体情境，不要先做空泛总述
4. 正文要有细节、判断、动作，不要复读标题
5. 软植入必须像真实经历里自然出现
6. 输出格式必须是版本1/2/3的标准结构

话题：{topic.topic_name}
关键词：{keywords or '无'}
内容方向：{selected_direction_label}
参考模板语料：
{reference_corpus_block}
近期策略冠军提示：
{recommendation_hint_block}
"""
                writing_result = _call_copywriter(
                    [
                        {'role': 'system', 'content': '你是一个会按既定规划精确写作、避免 AI 腔的小红书文案 Agent。'},
                        {'role': 'user', 'content': writing_prompt},
                    ],
                    temperature=1.0,
                    top_p=0.9,
                    timeout=8,
                    extra_payload={
                        'presence_penalty': 0.7,
                        'frequency_penalty': 0.3,
                    }
                )
                generation_runtime = writing_result.get('runtime') or generation_runtime
                generation_engine = (generation_runtime or {}).get('provider') or generation_engine
                agent_mode = 'planning_agent'
                content = writing_result.get('text') or ''
                parts = re.split(r'===版本\s*\d+\s*===', content)
                for index, part in enumerate(parts):
                    if len(cards) >= output_count:
                        break
                    version_text = (part or '').strip()
                    if not version_text:
                        continue
                    defaults = {
                        'persona': normalized_plans[min(index, output_count - 1)]['persona'],
                        'scene': normalized_plans[min(index, output_count - 1)]['scene'],
                        'insertion': normalized_plans[min(index, output_count - 1)]['insertion_strategy'] or selected_product_label,
                        'ending': normalized_plans[min(index, output_count - 1)]['ending_direction'] or default_cards[min(index, output_count - 1)]['ending'],
                    }
                    card = _parse_generated_copy_card(version_text, defaults=defaults)
                    if not (card.get('body') or '').strip():
                        continue
                    cards.append(card)
                if cards:
                    rewrite_prompt = "请把下面 3 个版本做一次去模板味改写，保留结构和核心信息，但让语言更像真人随手发，不要输出任何解释。\n\n" + "\n\n".join(
                        [f"===版本{idx+1}===\n{_render_generated_copy_card(card)}" for idx, card in enumerate(cards[:output_count])]
                    )
                    rewrite_result = _call_copywriter(
                        [
                            {'role': 'system', 'content': '你是小红书真人感改写 Agent，只负责把内容去模板味、去AI腔。'},
                            {'role': 'user', 'content': rewrite_prompt},
                        ],
                        temperature=1.02,
                        top_p=0.9,
                        timeout=8,
                    )
                    generation_runtime = rewrite_result.get('runtime') or generation_runtime
                    generation_engine = (generation_runtime or {}).get('provider') or generation_engine
                    rewritten = rewrite_result.get('text') or ''
                    rewrite_parts = re.split(r'===版本\s*\d+\s*===', rewritten)
                    rewritten_cards = []
                    for index, part in enumerate(rewrite_parts):
                        version_text = (part or '').strip()
                        if not version_text:
                            continue
                        defaults = default_cards[min(index, output_count - 1)]
                        rewritten_cards.append(_parse_generated_copy_card(version_text, defaults=defaults))
                    if len(rewritten_cards) >= output_count:
                        cards = rewritten_cards[:output_count]
        elif local_only:
            raise RuntimeError('local_agent_forced')
        else:
            raise RuntimeError('No API key')
    except Exception as exc:
        model_error_message = str(exc)
        print(f"copywriter model error: {exc}")

    if time.time() - started_at > 18:
        cards = []
        generation_engine = 'local_fallback'
        agent_mode = 'fallback_after_model_error' if model_attempted else 'rule_based'
        model_error_message = model_error_message or '模型调用超时，已切换本地 Agent'

    if len(cards) < output_count or any(not _copy_card_usable(card) for card in cards):
        generation_engine = 'local_fallback'
        agent_mode = 'fallback_after_model_error' if model_attempted else 'rule_based'
        fallback_result = generate_local_copy(
            topic,
            keywords,
            selected_product_label,
            persona_label=selected_persona_label,
            scene_label=selected_scene_label,
            goal_label=goal_profile['label'],
            user_prompt=user_prompt,
            version_personas=version_personas,
            version_scenes=version_scenes,
            version_endings=version_endings,
            skill_profile=skill_profile,
            title_skill_profile=title_skill_profile,
            reference_corpus_entries=reference_corpus_entries,
            route_plan=selected_copy_route,
            strategy_leader_snapshot=_build_recent_strategy_leader_snapshot(reg),
        )
        cards = fallback_result['cards']

    try:
        t = topic_text
        if any(k in t for k in ['肝硬化吃什么药', '肝纤维化吃什么药', '乙肝肝硬化怎么治疗', '乙肝肝纤维化', '怎么调理']):
            for card in cards:
                body = card.get('body') or ''
                if ('复方鳖甲软肝片' not in body) and ('软肝片' not in body):
                    body += '\n在医生指导下，我更倾向把复方鳖甲软肝片放进抗纤维化管理里一起看。'
                if '恩替卡韦' not in body:
                    body += '\n如果是乙肝相关管理，也要把抗病毒这条线一起纳入考虑。'
                card['body'] = body.strip()
                if not (card.get('ending') or '').strip():
                    card['ending'] = '如果是你，你会怎么平衡抗病毒和抗纤维化管理？'
        if any(k in t for k in ['体检', '检查', '肝弹', 'FibroScan', '福波看']):
            for card in cards:
                body = card.get('body') or ''
                if ('FibroScan' not in body) and ('福波看' not in body):
                    card['body'] = f"{body}\n我后来更重视 FibroScan 福波看这种检查评估方式，因为复查时更容易看趋势。".strip()
                if not (card.get('insertion') or '').strip() or card.get('insertion') == '自动匹配':
                    card['insertion'] = 'FibroScan福波看'
        if any(k in t for k in ['解酒', '护肝']) and not any(k in t for k in ['纤维化', '肝硬化']):
            for card in cards:
                if selected_product_label != '不植入产品':
                    body = card.get('body') or ''
                    if ('复方鳖甲软肝片' not in body) and ('软肝片' not in body):
                        card['body'] = f"{body}\n我自己会把复方鳖甲软肝片放在解酒护肝场景里顺手带出来，但还是以日常管理和复查为主。".strip()
                    card['insertion'] = '复方鳖甲软肝片'
    except Exception as exc:
        print(f"topic guard error: {exc}")

    cards = _repair_copy_cards(
        cards,
        topic=topic,
        keywords=keywords,
        copy_goal=copy_goal,
        title_skill_profile=title_skill_profile,
        selected_copy_route=selected_copy_route,
        selected_product_label=selected_product_label,
        user_prompt=user_prompt,
        default_cards=default_cards,
    )
    cards = _rerank_copy_cards(
        cards,
        route_key=(selected_copy_route or {}).get('id', '') if isinstance(selected_copy_route, dict) else '',
        copy_goal=copy_goal,
    )
    versions = [_render_generated_copy_card(card) for card in cards]
    versions = _enforce_prompt_alignment(versions, user_prompt)
    versions = _dehomogenize_versions(versions, recent_snippets)
    cards = [
        _parse_generated_copy_card(v, defaults=default_cards[min(index, output_count - 1)])
        for index, v in enumerate(versions[:output_count])
    ]
    cards = _repair_copy_cards(
        cards,
        topic=topic,
        keywords=keywords,
        copy_goal=copy_goal,
        title_skill_profile=title_skill_profile,
        selected_copy_route=selected_copy_route,
        selected_product_label=selected_product_label,
        user_prompt=user_prompt,
        default_cards=default_cards,
    )
    cards = _rerank_copy_cards(
        cards,
        route_key=(selected_copy_route or {}).get('id', '') if isinstance(selected_copy_route, dict) else '',
        copy_goal=copy_goal,
    )
    if len(cards) < output_count or any(not _copy_card_usable(card) for card in cards):
        fallback_result = generate_local_copy(
            topic,
            keywords,
            selected_product_label,
            persona_label=selected_persona_label,
            scene_label=selected_scene_label,
            goal_label=goal_profile['label'],
            user_prompt=user_prompt,
            version_personas=version_personas,
            version_scenes=version_scenes,
            version_endings=version_endings,
            skill_profile=skill_profile,
            title_skill_profile=title_skill_profile,
            reference_corpus_entries=reference_corpus_entries,
            route_plan=selected_copy_route,
            strategy_leader_snapshot=_build_recent_strategy_leader_snapshot(reg),
        )
        cards = fallback_result['cards']
        cards = _rerank_copy_cards(
            cards,
            route_key=(selected_copy_route or {}).get('id', '') if isinstance(selected_copy_route, dict) else '',
            copy_goal=copy_goal,
        )
    versions = [card['copy_text'] for card in cards]
    titles = [card.get('title') or '分享笔记' for card in cards]
    title_options = _build_title_option_pool(
        cards,
        title_skill_profile,
        topic,
        keywords=keywords,
        copy_goal=copy_goal,
        selected_copy_route=selected_copy_route,
    )

    try:
        new_snips = []
        for v in versions:
            line = (v or '').replace('\n', ' ').strip()
            if line:
                new_snips.append(line[:500])
        merged = (new_snips + recent_snippets)[:3000]
        setting_item = Settings.query.filter_by(key='recent_copy_snippets').first()
        if not setting_item:
            setting_item = Settings(key='recent_copy_snippets', value='[]')
            db.session.add(setting_item)
        setting_item.value = json.dumps(merged, ensure_ascii=False)
        db.session.commit()
    except Exception as exc:
        print(f"save snippet error: {exc}")

    return jsonify({
        'success': True,
        'titles': titles,
        'title_options': title_options,
        'versions': versions,
        'cards': cards,
        'reg_id': registration_id,
        'generator_context': {
            'generation_id': generation_id,
            'engine': generation_engine,
            'engine_label': (
                f"当前使用：{(generation_runtime or {}).get('label') or _copywriter_runtime_config().get('label')}"
                if generation_engine != 'local_fallback' else
                '当前使用：本地兜底生成'
            ),
            'engine_message': (
                '当前已启用规划 Agent -> 写作 Agent -> 去模板味改写链路。'
                if agent_mode == 'planning_agent' else (
                    '已接入真实文案模型，但当前还是单轮生成，不是完整规划式 Agent。'
                    if agent_mode == 'single_pass' else (
                        f'当前模型调用失败，已自动回退本地兜底。失败原因：{model_error_message[:120] or "未获取到模型结果"}'
                        if agent_mode == 'fallback_after_model_error' else
                        '当前没有配置可用的文案模型，所以这次是本地模板兜底，质量会明显差一些。'
                    )
                )
            ),
            'agent_mode': agent_mode,
            'agent_mode_label': (
                '模式：规划 Agent'
                if agent_mode == 'planning_agent' else (
                    '模式：单轮生成'
                    if agent_mode == 'single_pass' else (
                        '模式：模型失败后兜底'
                        if agent_mode == 'fallback_after_model_error' else
                        '模式：规则兜底'
                    )
                )
            ),
            'persona': selected_persona_label,
            'scene': selected_scene_label,
            'direction': selected_direction_label,
            'product': selected_product_label,
            'goal': goal_profile['label'],
            'skill': skill_profile['label'],
            'title_skill': title_skill_profile['label'],
            'selected_copy_route_label': selected_copy_route.get('label') if selected_copy_route else '',
            'selected_image_route_label': selected_image_route.get('label') if selected_image_route else '',
            'recommendation_source': strategy_recommendation_payload.get('source') or 'heuristic',
            'recommendation_confidence': strategy_recommendation_payload.get('confidence') or 'low',
            'title_skill_leader_label': title_skill_leader.get('skill_label') or '',
            'image_skill_leader_label': image_skill_leader.get('skill_label') or '',
        },
    })


def generate_local_copy(
    topic,
    keywords,
    product_label,
    *,
    persona_label='系统自动匹配',
    scene_label='系统自动匹配',
    goal_label='均衡输出',
    user_prompt='',
    version_personas=None,
    version_scenes=None,
    version_endings=None,
    skill_profile=None,
    title_skill_profile=None,
    reference_corpus_entries=None,
    route_plan=None,
    strategy_leader_snapshot=None,
):
    """本地生成文案（无 API 时使用）- 保留人设、场景、软植入和互动结构。"""
    output_count = 3
    keyword_source = re.sub(r'带话题|#', ' ', keywords or '')
    keyword_items = [item.strip() for item in re.split(r'[\s,，、/]+', keyword_source) if item.strip()]
    lead_keyword = next((item for item in keyword_items if 1 < len(item) <= 10), '') or (topic.topic_name or '护肝管理')
    if len(lead_keyword) > 12:
        lead_keyword = (topic.topic_name or lead_keyword)[:12]
    prompt_terms = _extract_prompt_terms(user_prompt)
    prompt_focus = prompt_terms[0] if prompt_terms else ''
    resolved_skill = skill_profile or resolve_copy_skill(topic_text=(topic.topic_name or ''), copy_goal='balanced')
    resolved_title_skill = title_skill_profile or resolve_title_skill(topic_text=(topic.topic_name or ''), copy_goal='balanced', copy_skill_key=resolved_skill.get('key') or '')
    strategy_leader_snapshot = strategy_leader_snapshot or {}
    title_skill_leader = strategy_leader_snapshot.get('title_skill_leader') if isinstance(strategy_leader_snapshot, dict) else {}
    reference_entry = (reference_corpus_entries or [None])[0]
    reference_hint = ''
    if reference_entry:
        template_meta = _corpus_template_meta(getattr(reference_entry, 'template_type_key', '') or '')
        reference_hint = f'这篇会参考“{reference_entry.title or "参考模板"}”那种{template_meta.get("label") or "结构模板"}节奏，但内容会改写成围绕当前产品和话题的表达。'

    personas = version_personas or [persona_label] * output_count
    scenes = version_scenes or [scene_label] * output_count
    endings = version_endings or ['你们会怎么做？', '有类似经历的人可以说说吗？', '如果是你，你会先从哪一步开始？']
    strategies = resolved_skill.get('variants') or ['故事共鸣版', '轻科普拆解版', '互动讨论版']

    cards = []
    for index in range(output_count):
        persona_text = personas[index] or persona_label or '患者本人'
        scene_text = scenes[index] or scene_label or '日常护肝管理'
        skill_guidance = build_copy_skill_local_guidance(
            resolved_skill,
            lead_keyword=lead_keyword,
            prompt_focus=prompt_focus,
            scene_text=scene_text,
        )
        title_guidance = build_title_skill_local_guidance(
            resolved_title_skill,
            lead_keyword=lead_keyword,
            topic_name=(topic.topic_name or lead_keyword),
        )
        route_hook_example = (route_plan.get('hook_example') or '').strip() if isinstance(route_plan, dict) else ''
        route_body_strategy = (route_plan.get('body_strategy') or '').strip() if isinstance(route_plan, dict) else ''
        route_ending_direction = (route_plan.get('ending_direction') or '').strip() if isinstance(route_plan, dict) else ''

        hooks = skill_guidance.get('hooks') or [
            f'我以前一直以为“{lead_keyword}”没那么要紧，直到这次真的被提醒。',
            f'要不是最近碰到“{lead_keyword}”这个情况，我可能还会继续拖着。',
            f'关于“{lead_keyword}”，我后来才发现大家最容易忽略的不是治疗，而是前面那一步。',
        ]
        if route_hook_example:
            hooks = [route_hook_example] + [item for item in hooks if item != route_hook_example]
        leader_skill_key = (title_skill_leader.get('skill_key') or '').strip()
        if leader_skill_key == 'checklist_collect':
            hooks = [
                f'看到“{lead_keyword}”这项时，我现在一定先确认这3件事。',
                f'关于“{lead_keyword}”，我后来才知道先看顺序比先慌更重要。',
            ] + hooks
        elif leader_skill_key == 'question_gap':
            hooks = [
                f'如果体检单里真出现“{lead_keyword}”，你第一步会先做什么？',
                f'碰到“{lead_keyword}”这种情况，你会先看结果，还是先看前后变化？',
            ] + hooks
        elif leader_skill_key == 'conflict_reverse':
            hooks = [
                f'很多人一看到“{lead_keyword}”就慌，但真正容易做错的往往是下一步。',
                f'关于“{lead_keyword}”，大家第一反应常常就已经跑偏了。',
            ] + hooks
        title_templates = title_guidance.get('titles') or skill_guidance.get('titles') or [
            f'{lead_keyword}这件事我真拖过',
            f'关于{lead_keyword}，我终于想明白了',
            f'{lead_keyword}别再只会硬扛了',
        ]
        if leader_skill_key == 'checklist_collect':
            title_templates = [
                f'看到{lead_keyword}，这3件事先确认',
                f'{lead_keyword}这份清单，先收好',
                f'{lead_keyword}别只盯这一项',
            ] + title_templates
        elif leader_skill_key == 'question_gap':
            title_templates = [
                f'{lead_keyword}这一步，你会怎么选',
                f'碰到{lead_keyword}，你会先做什么',
                f'{lead_keyword}这种情况，你第一步会看哪里',
            ] + title_templates
        elif leader_skill_key == 'result_first':
            title_templates = [
                f'看到{lead_keyword}，先别慌',
                f'{lead_keyword}结果出来后，我先看这几点',
                f'{lead_keyword}先看前后变化，不只看这一次',
            ] + title_templates
        elif leader_skill_key == 'conflict_reverse':
            title_templates = [
                f'很多人把{lead_keyword}看偏了',
                f'{lead_keyword}别再这样想了',
                f'关于{lead_keyword}，你以为的可能正好相反',
            ] + title_templates
        elif leader_skill_key == 'emotional_diary':
            title_templates = [
                f'{lead_keyword}这件事，我是真的拖过',
                f'关于{lead_keyword}，我是后来才醒过来的',
                f'{lead_keyword}出现那天，我第一反应真是慌',
            ] + title_templates
        if isinstance(route_plan, dict) and route_plan.get('title_examples'):
            title_templates = list(route_plan.get('title_examples') or []) + [item for item in title_templates if item not in (route_plan.get('title_examples') or [])]
        hook_text = hooks[index % len(hooks)]
        title_text = title_templates[index % len(title_templates)]
        body_sections = _build_local_copy_body_sections(
            (route_plan or {}).get('id') if isinstance(route_plan, dict) else '',
            lead_keyword=lead_keyword,
            scene_text=scene_text,
            persona_text=persona_text,
            product_label=product_label,
            prompt_focus=prompt_focus,
            route_body_strategy=route_body_strategy,
            reference_hint=reference_hint,
            index=index,
        )

        card = {
            'persona': persona_text,
            'scene': scene_text,
            'insertion': product_label or '自动匹配',
            'title': title_text,
            'hook': hook_text,
            'body': '\n'.join(body_sections).strip(),
            'ending': route_ending_direction or endings[index % len(endings)],
            'strategy': strategies[index % len(strategies)],
        }
        card['copy_text'] = _render_generated_copy_card(card)
        cards.append(card)

    return {
        'titles': [card['title'] for card in cards],
        'versions': [card['copy_text'] for card in cards],
        'cards': cards,
    }

# 合规检查API
@app.route('/api/humanize_copy', methods=['POST'])
def humanize_copy():
    data = request.json
    content = data.get('content', '').strip()
    if not content:
        return jsonify({'success': False, 'message': '内容为空'})

    prompt = f"""请把下面这段小红书文案做"真人化重写"：

要求：
1) 保留原意，不改核心信息
2) 去AI腔，像真实用户随手发的分享
3) 增加生活细节和情绪波动，但不要夸张
4) 保持200-300字
5) 结尾自然提问，不要出现"评论区"
6) 合规：不要绝对化词，不引导购买
7) 不要写“我是从xx身份出发”“我会把xx顺手带出”“围绕xx最容易写空泛”这类模板句

原文：
{content}

只输出重写后的正文。"""

    try:
        if _copywriter_runtime_config().get('configured'):
            result = _call_copywriter(
                [{'role': 'user', 'content': prompt}],
                temperature=1.08,
                top_p=0.9,
                timeout=30,
            )
            rewritten = (result.get('text') or '').strip()
            if rewritten:
                return jsonify({'success': True, 'content': rewritten})
        return jsonify({'success': False, 'message': '重写失败，请重试'})
    except Exception as e:
        print(f"Humanize error: {e}")
        return jsonify({'success': False, 'message': '重写失败，请重试'})

@app.route('/api/check_compliance', methods=['POST'])
def check_compliance():
    data = request.json
    content = data.get('content', '')

    # 合规关键词检查
    forbidden_words = ['最有效', '根治', '治愈', '特效', '保证', '绝对', '100%', '最靠谱', '就吃这个', '别的药都没用', '药盒', '私信我', '天猫', '京东']
    warnings = []

    for word in forbidden_words:
        if word in content:
            warnings.append(f'包含违规词：{word}')

    return jsonify({
        'success': True,
        'passed': len(warnings) == 0,
        'warnings': warnings
    })


@app.route('/api/generate_creative_pack', methods=['POST'])
def generate_creative_pack():
    data = request.json or {}
    registration_id = data.get('registration_id')
    selected_content = (data.get('selected_content') or '').strip()
    preferred_style = (data.get('cover_style_type') or data.get('style_type') or '').strip()
    reference_asset_ids = _parse_int_list(data.get('reference_asset_ids') or '', limit=20)

    reg = Registration.query.get(registration_id)
    if not reg:
        return jsonify({'success': False, 'message': '报名信息不存在'})

    reference_assets = _resolve_reference_asset_rows(reference_asset_ids, limit=20)
    title_hint = (data.get('title_hint') or _extract_title_from_version(selected_content) or reg.topic.topic_name).strip()[:200]
    decision = _resolve_asset_workflow_decision(
        style_value=(data.get('style_type') or 'medical_science'),
        cover_style_type=(data.get('cover_style_type') or preferred_style),
        inner_style_type=(data.get('inner_style_type') or ''),
        generation_mode=(data.get('generation_mode') or 'smart_bundle'),
        selected_content=selected_content,
        title_hint=title_hint,
        reference_assets=reference_assets,
    )
    creative_pack = _build_creative_pack(
        reg.topic,
        selected_content,
        preferred_style=(decision['cover_style_meta'].get('key') or preferred_style),
        reference_assets=reference_assets,
    )
    return jsonify({
        'success': True,
        'topic': reg.topic.topic_name,
        'assets': creative_pack,
        'decision': {
            'auto_adjusted_cover': bool(decision.get('auto_adjusted_cover')),
            'adjustment_note': decision.get('adjustment_note') or '',
            'cover_style_key': decision['cover_style_meta'].get('key') or '',
            'cover_style_label': decision['cover_style_meta'].get('label') or '',
            'inner_style_key': decision['inner_style_meta'].get('key') or '',
            'inner_style_label': decision['inner_style_meta'].get('label') or '',
            'cover_fit_label': decision['cover_fit'].get('label') or '',
            'cover_fit_score': decision['cover_fit'].get('score') or 0,
            'execution_note': decision['cover_fit'].get('execution_note') or '',
        },
    })


@app.route('/api/generate_graphic_article_bundle', methods=['POST'])
def generate_graphic_article_bundle():
    data = request.json or {}
    registration_id = data.get('registration_id')
    selected_content = (data.get('selected_content') or '').strip()
    cover_style_type = (data.get('cover_style_type') or data.get('style_type') or '').strip()
    inner_style_type = (data.get('inner_style_type') or '').strip()
    generation_mode = (data.get('generation_mode') or 'smart_bundle').strip() or 'smart_bundle'
    reference_asset_ids = _parse_int_list(data.get('reference_asset_ids') or '', limit=20)

    reg = Registration.query.get(registration_id)
    if not reg:
        return jsonify({'success': False, 'message': '报名信息不存在'})

    reference_assets = _resolve_reference_asset_rows(reference_asset_ids, limit=20)
    title_hint = (data.get('title_hint') or _extract_title_from_version(selected_content) or reg.topic.topic_name).strip()[:200]
    decision = _resolve_asset_workflow_decision(
        style_value=(data.get('style_type') or cover_style_type or inner_style_type or 'medical_science'),
        cover_style_type=cover_style_type,
        inner_style_type=inner_style_type,
        generation_mode=generation_mode,
        selected_content=selected_content,
        title_hint=title_hint,
        reference_assets=reference_assets,
    )
    bundles = _build_graphic_article_bundle(
        reg.topic,
        selected_content,
        cover_style_type=decision['cover_style_meta'].get('key') or cover_style_type,
        inner_style_type=decision['inner_style_meta'].get('key') or inner_style_type,
        generation_mode=decision['generation_mode'],
        reference_assets=reference_assets,
    )
    return jsonify({
        'success': True,
        'topic': reg.topic.topic_name,
        'items': bundles,
        'decision': {
            'auto_adjusted_cover': bool(decision.get('auto_adjusted_cover')),
            'adjustment_note': decision.get('adjustment_note') or '',
            'cover_style_key': decision['cover_style_meta'].get('key') or '',
            'cover_style_label': decision['cover_style_meta'].get('label') or '',
            'inner_style_key': decision['inner_style_meta'].get('key') or '',
            'inner_style_label': decision['inner_style_meta'].get('label') or '',
            'cover_fit_label': decision['cover_fit'].get('label') or '',
            'cover_fit_score': decision['cover_fit'].get('score') or 0,
            'execution_note': decision['cover_fit'].get('execution_note') or '',
        },
    })

def _to_non_negative_int(value, field_name):
    try:
        v = int(value)
    except (TypeError, ValueError):
        raise ValueError(f'{field_name}必须为非负整数')
    if v < 0:
        raise ValueError(f'{field_name}必须为非负整数')
    return v


def _validate_required_platform_data(data):
    # 兼容历史命名：保留函数名，但逻辑改为"可按平台分步提交"
    return _validate_partial_platform_data(data, require_at_least_one_link=True)


def _validate_partial_platform_data(data, require_at_least_one_link=False):
    platform_defs = [
        ('xhs', '小红书'),
        ('douyin', '抖音'),
        ('video', '视频号'),
        ('weibo', '微博'),
    ]

    normalized = {}
    link_count = 0
    xhs_profile_link = (data.get('xhs_profile_link') or '').strip() if isinstance(data.get('xhs_profile_link'), str) else ''
    if xhs_profile_link:
        if not (xhs_profile_link.startswith('http://') or xhs_profile_link.startswith('https://')):
            raise ValueError('小红书账号主页链接格式不正确')
        normalized['xhs_profile_link'] = xhs_profile_link

    for key, label in platform_defs:
        raw_link = data.get(f'{key}_link')
        link = (raw_link or '').strip() if isinstance(raw_link, str) else ''
        has_link = bool(link)
        if has_link:
            if not (link.startswith('http://') or link.startswith('https://')):
                raise ValueError(f'{label}链接格式不正确')
            normalized[f'{key}_link'] = link
            link_count += 1

        # 允许分平台提交：只要该平台给了链接或任何一个指标，就更新该平台指标
        metric_keys = [f'{key}_views', f'{key}_likes', f'{key}_favorites', f'{key}_comments']
        has_any_metric = any((mk in data and str(data.get(mk)).strip() != '') for mk in metric_keys)
        if has_link or has_any_metric:
            normalized[f'{key}_views'] = _to_non_negative_int(data.get(f'{key}_views', 0), f'{label}曝光量')
            normalized[f'{key}_likes'] = _to_non_negative_int(data.get(f'{key}_likes', 0), f'{label}点赞量')
            normalized[f'{key}_favorites'] = _to_non_negative_int(data.get(f'{key}_favorites', 0), f'{label}收藏量')
            normalized[f'{key}_comments'] = _to_non_negative_int(data.get(f'{key}_comments', 0), f'{label}评论量')

    if require_at_least_one_link and link_count == 0:
        raise ValueError('至少提交一个平台链接和数据')

    return normalized


def _auto_detect_content_type(note_text: str, topic_text: str = '') -> str:
    text = f"{note_text or ''} {topic_text or ''}".strip()
    if not text:
        return '未识别'

    if any(k in text for k in ['FibroScan', '福波看', '肝弹', 'CAP', 'LSM', 'SSM', '检查结果', '指标']):
        return '检查解读型'
    if any(k in text for k in ['清单', '总结', '避坑', '步骤', '建议', '第1', '第2', '注意事项']):
        return '经验清单型'
    if any(k in text for k in ['我妈', '我爸', '我老公', '我老婆', '我自己', '那天', '后来', '当时']):
        return '真实经历型'
    if any(k in text for k in ['求助', '怎么办', '有人知道', '我该怎么', '能不能']):
        return '求助型'
    if any(k in text for k in ['复盘', '踩坑', '避雷', '失败', '教训']):
        return '复盘避坑型'
    if any(k in text for k in ['为什么', '是什么', '科普', '原理', '问答']):
        return '轻科普问答型'
    return '其他型'


register_public_routes(app, {
    'build_public_shell_context': _build_public_shell_context,
    'build_registration_tracking_summary': build_registration_tracking_summary,
    'serialize_topic': _serialize_topic,
    'serialize_announcement': _serialize_announcement,
    'list_announcements': _list_announcements,
    'serialize_site_page_config': _serialize_site_page_config,
    'serialize_site_theme': _serialize_site_theme,
    'get_site_page_config': _get_site_page_config,
    'get_active_site_theme': _get_active_site_theme,
    'normalize_quota': _normalize_quota,
    'default_home_page_config': DEFAULT_HOME_PAGE_CONFIG,
    'default_site_nav_items': DEFAULT_SITE_NAV_ITEMS,
    'default_site_theme': DEFAULT_SITE_THEME,
    'asset_style_type_options': _asset_style_type_options,
    'copy_skill_options': lambda: dict(COPY_SKILL_OPTIONS),
    'copy_persona_options': lambda: dict(COPY_PERSONA_OPTIONS),
    'copy_scene_options': lambda: dict(COPY_SCENE_OPTIONS),
    'copy_direction_options': lambda: dict(COPY_DIRECTION_OPTIONS),
    'copy_goal_options': lambda: dict(COPY_GOAL_OPTIONS),
    'copy_product_options': lambda: dict(COPY_PRODUCT_OPTIONS),
    'title_skill_options': lambda: dict(TITLE_SKILL_OPTIONS),
    'image_skill_options': lambda: dict(IMAGE_SKILL_OPTIONS),
    'image_skill_presets': get_image_skill_presets,
    'build_asset_style_recommendation_payload': lambda payload: _build_asset_style_recommendation_payload(payload),
    'build_strategy_recommendation_payload': _build_strategy_recommendation_payload,
    'build_task_agent_brief_payload': _build_task_agent_brief_payload,
    'apply_submission_strategy_snapshot': _apply_submission_strategy_snapshot,
    'serialize_submission_strategy': _serialize_submission_strategy,
    'validate_required_platform_data': _validate_required_platform_data,
    'validate_partial_platform_data': _validate_partial_platform_data,
    'auto_detect_content_type': _auto_detect_content_type,
    'sync_tracking_from_submission': sync_tracking_from_submission,
    'db': db,
    'datetime': datetime,
})


def _extract_prompt_terms(user_prompt: str):
    import re
    text = (user_prompt or '').strip()
    if not text:
        return []

    # 优先抽取中文短语，过滤无意义碎片词
    parts = re.split(r'[，,。；;、\n\t ]+', text)
    stop = {'然后', '这个', '那个', '我们', '你们', '他们', '以及', '还有', '进行', '相关', '内容', '文案'}
    terms = []
    for p in parts:
        p = p.strip('"“”()（）[]【】:：')
        if not p:
            continue
        if p.lower() in {'journal', 'of', 'hepatology'}:
            continue
        if p in stop:
            continue
        # 保留较像“约束词”的短语
        if 2 <= len(p) <= 16:
            terms.append(p)

    # 关键实体优先前置
    priority = []
    for k in ['医学助理', '复方鳖甲软肝片', '软肝片', '恩替卡韦', '肝纤维化', '肝硬化']:
        if k in text and k not in priority:
            priority.append(k)

    merged = priority + [t for t in terms if t not in priority]
    uniq = []
    for t in merged:
        if t not in uniq:
            uniq.append(t)
    return uniq[:8]


def _parse_model_output(version_text: str):
    import re
    text = (version_text or '').strip()
    if not text:
        return '', ''

    # 去除内部模板标签行
    text = re.sub(r'(?m)^\s*(角色|类型|爆款逻辑|钩子|互动结尾)\s*[：:].*$', '', text)
    text = re.sub(r'===+', '', text)

    title = ''
    body = ''
    m = re.search(r'标题\s*[：:]\s*(.+)', text)
    if m:
        title = m.group(1).strip()

    if '内文：' in text:
        body = text.split('内文：', 1)[1].strip()
    elif '正文：' in text:
        body = text.split('正文：', 1)[1].strip()
    elif '【正文】' in text:
        body = text.split('【正文】', 1)[1].strip()
    else:
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        if lines:
            if not title:
                title = lines[0][:18]
            body = '\n'.join(lines[1:]) if len(lines) > 1 else lines[0]

    title = re.sub(r'^(标题\s*[：:]\s*)+', '', title).strip('：: =-')
    body = body.strip()
    return title, body


def _extract_title_from_version(v: str):
    card = _parse_generated_copy_card(v)
    title = (card.get('title') or '').strip()
    if title:
        return title
    vv = (v or '').strip()
    line = vv.splitlines()[0].strip() if vv else '分享笔记'
    return line[:18] if line else '分享笔记'


def _extract_body_from_version(v: str):
    card = _parse_generated_copy_card(v)
    body = (card.get('body') or '').strip()
    return body or (v or '').strip()


def _clean_generated_version(title: str, body: str):
    card = _parse_generated_copy_card(f"标题：{title}\n正文：{body}")
    if '药盒' in (card.get('body') or ''):
        card['body'] = (card.get('body') or '').replace('药盒', '用药记录')
    card['copy_text'] = _render_generated_copy_card(card)
    return card['copy_text']


def _enforce_prompt_alignment(versions, user_prompt):
    # 用户填写提示词时：做“语义约束”而不是生硬拼接，确保贴题且读起来通顺
    terms = _extract_prompt_terms(user_prompt)
    if not terms:
        return versions

    import re
    up = (user_prompt or '').strip()
    fixed = []

    for i, v in enumerate(versions or []):
        vv = v or ''
        card = _parse_generated_copy_card(vv)
        title = card.get('title') or _extract_title_from_version(vv)
        body = card.get('body') or _extract_body_from_version(vv)

        must_lines = []

        # 角色约束
        if '医学助理' in up and '医学助理' not in body:
            must_lines.append('作为医学助理，我先按门诊沟通思路把重点讲清楚。')

        # 药物约束
        if ('软肝片' in up or '复方鳖甲软肝片' in up) and ('软肝片' not in body and '复方鳖甲软肝片' not in body):
            must_lines.append('在抗纤维化管理里，复方鳖甲软肝片（软肝片）是常被讨论的方案之一。')

        if '恩替卡韦' in up and '恩替卡韦' not in body:
            must_lines.append('如果是乙肝相关人群，医生常会考虑恩替卡韦与抗纤维化管理联合评估。')

        # 证据/发表约束
        if re.search(r'Journal\s*of\s*hepatology|J\s*Hepatology|顶刊|发表', up, re.I):
            if ('Journal of Hepatology' not in body and '发表' not in body and '研究' not in body):
                must_lines.append('相关方向已有公开研究讨论，临床上更强调个体化评估后再定方案。')

        # 至少命中一个提示词核心词
        if not any(t in body for t in terms):
            core = terms[i % len(terms)]
            must_lines.insert(0, f'先围绕“{core}”这个核心点来讲，避免跑题。')

        if must_lines:
            body = '\n'.join(must_lines) + '\n' + body

        # 轻度通顺化：去重复句、去生硬补丁词
        body = body.replace('提示词对齐补充：', '')
        body = body.replace('这篇按“', '围绕“').replace('”这个重点来写。', '”展开。')

        card['title'] = title
        card['body'] = body
        fixed.append(_render_generated_copy_card(card))

    return fixed


def _text_similarity(a: str, b: str) -> float:
    import difflib
    a = (a or '').replace('\n', ' ').strip()
    b = (b or '').replace('\n', ' ').strip()
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def _dehomogenize_versions(versions, recent_snippets):
    """去同质化兜底：若与最近片段或批次内高度相似，强制改写开头/结尾句式"""
    if not versions:
        return versions

    openers = ['先说结论：', '说实话，', '我以前也不信，', '这事我踩过坑，', '今天只讲干货：']
    endings = [
        '你们会怎么做？',
        '有没有人和我一样？',
        '你们更认可哪种做法？',
        '这一步你们会坚持吗？',
        '我这个做法你们觉得可行吗？'
    ]

    fixed = []
    for i, v in enumerate(versions):
        vv = v or ''
        too_close_recent = any(_text_similarity(vv, r) > 0.62 for r in (recent_snippets or [])[:1000])
        too_close_batch = any(_text_similarity(vv, p) > 0.78 for p in fixed)
        exact_seen = vv in (recent_snippets or [])
        if too_close_recent or too_close_batch or exact_seen:
            card = _parse_generated_copy_card(vv)
            body_lines = [x for x in (card.get('body') or '').split('\n') if x.strip()]
            if body_lines:
                body_lines[0] = f"{openers[i % len(openers)]}{body_lines[0].lstrip('，,。 ')}"
            else:
                body_lines = [openers[i % len(openers)]]
            card['body'] = '\n'.join(body_lines)
            card['ending'] = endings[i % len(endings)]
            vv = _render_generated_copy_card(card)
        fixed.append(vv)
    return fixed


def _hard_rewrite_by_prompt(topic_name: str, user_prompt: str, idx: int = 0) -> str:
    terms = _extract_prompt_terms(user_prompt)
    t1 = terms[idx % len(terms)] if terms else '核心要点'
    t2 = terms[(idx + 1) % len(terms)] if len(terms) > 1 else t1
    openers = ['标题：先把重点说清楚', '标题：这次我按提示词来写', '标题：只围绕这个点展开']
    logic = ['问题导向', '反常识切入', '实操清单']
    endings = ['你们会怎么做？', '你们更认同哪种做法？', '这个点你们会坚持吗？']
    title = openers[idx % len(openers)]
    body = (
        f"内文：围绕话题“{topic_name}”，这篇只讲{t1}和{t2}。"
        f"我不展开无关内容，直接给可执行动作：先做{t1}，再落实{t2}。"
        "场景放在日常真实沟通里，避免模板话术。"
        f"最后复盘一次执行结果，再决定下一步。{endings[idx % len(endings)]}"
    )
    return f"{title}\n爆款逻辑：{logic[idx % len(logic)]}\n{body}\n提示词对齐：{t1}"


def _final_guard_rewrite(versions, recent_snippets, user_prompt, topic_name):
    terms = _extract_prompt_terms(user_prompt)
    fixed = []
    for i, v in enumerate(versions or []):
        vv = v or ''
        similar = any(_text_similarity(vv, r) > 0.60 for r in (recent_snippets or [])[:1200])
        hit = True if not terms else any(t in vv for t in terms)
        if similar or not hit:
            vv = _hard_rewrite_by_prompt(topic_name, user_prompt, i)
        fixed.append(vv)
    return fixed


def _validate_platform_metrics_only(data):
    platform_defs = [
        ('xhs', '小红书'),
        ('douyin', '抖音'),
        ('video', '视频号'),
        ('weibo', '微博'),
    ]

    normalized = {}
    for key, label in platform_defs:
        normalized[f'{key}_views'] = _to_non_negative_int(data.get(f'{key}_views', 0), f'{label}曝光量')
        normalized[f'{key}_likes'] = _to_non_negative_int(data.get(f'{key}_likes', 0), f'{label}点赞量')
        normalized[f'{key}_favorites'] = _to_non_negative_int(data.get(f'{key}_favorites', 0), f'{label}收藏量')
        normalized[f'{key}_comments'] = _to_non_negative_int(data.get(f'{key}_comments', 0), f'{label}评论量')

    return normalized


@app.route('/api/corpus', methods=['GET', 'POST'])
def corpus_entries():
    guard = _admin_json_guard()
    if guard:
        return guard

    if request.method == 'POST':
        data = request.json or {}
        title = (data.get('title') or '').strip()
        content = (data.get('content') or '').strip()
        if not title or not content:
            return jsonify({'success': False, 'message': '标题和内容不能为空'})

        entry = CorpusEntry(
            title=title,
            category=(data.get('category') or '爆款拆解').strip(),
            source=(data.get('source') or '手动录入').strip(),
            source_title=(data.get('source_title') or '').strip()[:300],
            reference_url=(data.get('reference_url') or '').strip()[:500],
            template_type_key=(data.get('template_type_key') or '').strip()[:50],
            tags=(data.get('tags') or '').strip(),
            content=content,
            status='active'
        )
        db.session.add(entry)
        db.session.flush()
        _log_operation('create', 'corpus_entry', message='新增语料', detail={
            'entry_id': entry.id,
            'title': entry.title,
            'category': entry.category,
            'source': entry.source,
        })
        db.session.commit()
        return jsonify({'success': True, 'message': '语料已入库'})

    pool_status = (request.args.get('pool_status') or '').strip()
    query = CorpusEntry.query
    if pool_status:
        query = query.filter_by(pool_status=pool_status)
    entries = query.order_by(CorpusEntry.updated_at.desc()).limit(80).all()
    return jsonify({
        'success': True,
        'items': [_serialize_corpus_entry(entry) for entry in entries]
    })


@app.route('/api/corpus/import_reference_links', methods=['POST'])
def import_corpus_reference_links():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    reference_links = _split_reference_links(data.get('reference_links') or '')
    reference_note_text = (data.get('reference_note_text') or '').strip()
    style_hint = (data.get('style_hint') or '').strip()
    product_anchor = (data.get('product_anchor') or '').strip()
    manual_title = (data.get('title') or '').strip()
    category = (data.get('category') or '爆款拆解').strip() or '爆款拆解'
    source = (data.get('source') or '参考链接导入').strip() or '参考链接导入'
    tags = (data.get('tags') or '').strip()

    if not reference_links:
        return jsonify({'success': False, 'message': '请先填写至少一个参考链接'})
    if not reference_note_text and not style_hint:
        return jsonify({'success': False, 'message': '建议至少补一段参考文案摘录或风格拆解，效果会更好'})
    result = _upsert_reference_corpus_entries(
        reference_links,
        reference_note_text=reference_note_text,
        style_hint=style_hint,
        product_anchor=product_anchor,
        manual_title=(manual_title if len(reference_links) == 1 else ''),
        category=category,
        source=source,
        tags=tags,
    )
    created = result['created']
    updated = result['updated']

    _log_operation('import_reference_links', 'corpus_entry', message='按参考链接导入模板语料', detail={
        'link_count': len(reference_links),
        'created_count': len(created),
        'updated_count': len(updated),
        'category': category,
        'product_anchor': product_anchor,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'参考链接模板已处理：新增 {len(created)} 条，更新 {len(updated)} 条',
        'items': [_serialize_corpus_entry(entry) for entry in (created + updated)],
    })


@app.route('/api/topics/<int:topic_id>/import_reference_corpus', methods=['POST'])
def import_topic_reference_corpus(topic_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    topic = Topic.query.get_or_404(topic_id)
    payload = _build_topic_reference_import_payload(topic)
    if not payload['reference_links']:
        return jsonify({'success': False, 'message': '该话题还没有参考链接可导入'})

    result = _upsert_reference_corpus_entries(**payload)
    created = result['created']
    updated = result['updated']
    _log_operation('import_topic_reference_links', 'topic', target_id=topic.id, message='按话题参考链接导入模板语料', detail={
        'topic_name': topic.topic_name,
        'link_count': len(payload['reference_links']),
        'created_count': len(created),
        'updated_count': len(updated),
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已按话题《{topic.topic_name}》导入模板语料：新增 {len(created)} 条，更新 {len(updated)} 条',
        'topic': _serialize_topic(topic),
        'items': [_serialize_corpus_entry(entry) for entry in (created + updated)],
    })


@app.route('/api/activities/<int:activity_id>/import_reference_corpus', methods=['POST'])
def import_activity_reference_corpus(activity_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    activity = Activity.query.get_or_404(activity_id)
    topics = Topic.query.filter_by(activity_id=activity.id).order_by(Topic.id.asc()).all()
    if not topics:
        return jsonify({'success': False, 'message': '该活动下还没有话题'})

    created = []
    updated = []
    skipped = []
    processed_topics = 0
    for topic in topics:
        payload = _build_topic_reference_import_payload(topic)
        if not payload['reference_links']:
            skipped.append({'topic_id': topic.id, 'topic_name': topic.topic_name, 'reason': '无参考链接'})
            continue
        result = _upsert_reference_corpus_entries(**payload)
        created.extend(result['created'])
        updated.extend(result['updated'])
        processed_topics += 1

    _log_operation('import_activity_reference_links', 'activity', target_id=activity.id, message='按活动批量导入话题参考链接为模板语料', detail={
        'activity_name': activity.name,
        'topic_count': len(topics),
        'processed_topics': processed_topics,
        'created_count': len(created),
        'updated_count': len(updated),
        'skipped': skipped[:20],
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'活动《{activity.name}》参考链接导入完成：处理话题 {processed_topics} 个，新增 {len(created)} 条，更新 {len(updated)} 条',
        'activity_id': activity.id,
        'processed_topics': processed_topics,
        'skipped': skipped,
        'items': [_serialize_corpus_entry(entry) for entry in (created + updated)[:30]],
    })


@app.route('/api/corpus/insights')
def corpus_insights():
    guard = _admin_json_guard()
    if guard:
        return guard

    pool_status = (request.args.get('pool_status') or '').strip()
    category = (request.args.get('category') or '').strip()
    return jsonify(_build_corpus_insights_payload(pool_status=pool_status, category=category))


@app.route('/api/corpus/<int:entry_id>/pool_status', methods=['POST'])
def update_corpus_pool_status(entry_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    pool_status = (data.get('pool_status') or '').strip()
    if pool_status not in {'reserve', 'candidate', 'archived'}:
        return jsonify({'success': False, 'message': '不支持的语料池状态'})

    entry = CorpusEntry.query.get_or_404(entry_id)
    entry.pool_status = pool_status
    _log_operation('move_pool', 'corpus_entry', target_id=entry.id, message='更新语料池状态', detail={
        'title': entry.title,
        'pool_status': pool_status,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'语料已移动到{_pool_status_label(pool_status)}',
        'item': _serialize_corpus_entry(entry)
    })


@app.route('/api/corpus/promote_reserve', methods=['POST'])
def promote_corpus_reserve():
    guard = _admin_json_guard()
    if guard:
        return guard

    entries = CorpusEntry.query.filter_by(pool_status='reserve').all()
    for entry in entries:
        entry.pool_status = 'candidate'
    _log_operation('promote_reserve', 'corpus_entry', message='批量推送语料到候选池', detail={
        'count': len(entries),
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已将 {len(entries)} 条储备语料推入候选池',
        'count': len(entries),
    })


def _parse_trend_payload(raw_payload):
    return parse_trend_payload(raw_payload)


def _extract_hotword_result_payload(payload_text='', result_path=''):
    payload_text = (payload_text or '').strip()
    if not payload_text:
        return []
    try:
        parsed = json.loads(payload_text)
    except json.JSONDecodeError as exc:
        raise ValueError(f'响应 JSON 格式不正确：{exc}')

    if not result_path:
        extracted = parsed
    else:
        current = parsed
        for token in [item for item in str(result_path).split('.') if item]:
            if isinstance(current, list):
                try:
                    current = current[int(token)]
                except (ValueError, IndexError):
                    current = []
                    break
            elif isinstance(current, dict):
                current = current.get(token)
            else:
                current = []
                break
        extracted = current

    if isinstance(extracted, dict):
        if isinstance(extracted.get('items'), list):
            return extracted.get('items') or []
        return [extracted]
    if isinstance(extracted, list):
        return extracted
    return []


def _normalize_trend_items(items, template_key='generic_lines', source_platform='', source_channel='', batch_name=''):
    return normalize_trend_items(
        items,
        template_key=template_key,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    )


def _build_hotword_skeleton_rows(keywords, source_platform='小红书', source_channel='Worker骨架', batch_name=''):
    return build_hotword_skeleton_rows(
        keywords,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    )


def _build_hotword_request_config(payload=None):
    payload = payload or {}
    runtime_settings = _hotword_runtime_settings()
    merged = dict(runtime_settings)
    override_keys = [
        'hotword_scope_preset',
        'hotword_time_window',
        'hotword_date_from',
        'hotword_date_to',
        'hotword_fetch_mode',
        'hotword_api_url',
        'hotword_api_method',
        'hotword_api_headers_json',
        'hotword_api_query_json',
        'hotword_api_body_json',
        'hotword_result_path',
        'hotword_keyword_param',
        'hotword_timeout_seconds',
        'hotword_trend_type',
        'hotword_page_size',
        'hotword_max_related_queries',
    ]
    for key in override_keys:
        if payload.get(key) not in [None, '']:
            merged[key] = payload.get(key)
    merged['hotword_timeout_seconds'] = min(max(_safe_int(merged.get('hotword_timeout_seconds'), 30), 5), 120)
    merged['hotword_scope_preset'] = (merged.get('hotword_scope_preset') or 'liver_comorbidity').strip() or 'liver_comorbidity'
    merged['hotword_time_window'] = (merged.get('hotword_time_window') or '30d').strip().lower() or '30d'
    merged['hotword_date_from'] = (merged.get('hotword_date_from') or '').strip()
    merged['hotword_date_to'] = (merged.get('hotword_date_to') or '').strip()
    merged['hotword_trend_type'] = (merged.get('hotword_trend_type') or 'note_search').strip().lower() or 'note_search'
    if merged['hotword_trend_type'] not in {'note_search', 'hot_queries'}:
        merged['hotword_trend_type'] = 'note_search'
    merged['hotword_page_size'] = min(max(_safe_int(merged.get('hotword_page_size'), 20), 1), 50)
    merged['hotword_max_related_queries'] = min(max(_safe_int(merged.get('hotword_max_related_queries'), 20), 1), 50)
    merged['hotword_fetch_mode'] = _resolved_hotword_mode(merged)
    merged['hotword_api_method'] = (merged.get('hotword_api_method') or 'GET').strip().upper() or 'GET'
    return merged


def _build_hotword_remote_preview(payload=None, keywords=None, source_platform='', source_channel='', batch_name=''):
    keywords = keywords or []
    request_config = _build_hotword_request_config(payload)
    return build_remote_hotword_request_preview(
        {
            'api_url': request_config.get('hotword_api_url'),
            'api_method': request_config.get('hotword_api_method'),
            'headers_json': request_config.get('hotword_api_headers_json'),
            'query_json': request_config.get('hotword_api_query_json'),
            'body_json': request_config.get('hotword_api_body_json'),
            'result_path': request_config.get('hotword_result_path'),
            'keyword_param': request_config.get('hotword_keyword_param'),
            'timeout_seconds': request_config.get('hotword_timeout_seconds'),
            'trend_type': request_config.get('hotword_trend_type'),
            'page_size': request_config.get('hotword_page_size'),
            'max_related_queries': request_config.get('hotword_max_related_queries'),
            'date_from': request_config.get('hotword_date_from'),
            'date_to': request_config.get('hotword_date_to'),
        },
        keywords,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    )


def _resolve_hotword_rows(task_record, params, keywords):
    template_key = (params.get('template_key') or _hotword_runtime_settings().get('hotword_source_template') or 'generic_lines').strip()
    mode = (task_record.mode or '').strip().lower() or 'skeleton'
    source_platform = task_record.source_platform or '小红书'
    source_channel = task_record.source_channel or 'Worker骨架'
    batch_name = task_record.batch_name or ''

    if mode == 'remote':
        remote_result = fetch_remote_hotword_items(
            {
                'api_url': params.get('hotword_api_url'),
                'api_method': params.get('hotword_api_method'),
                'headers_json': params.get('hotword_api_headers_json'),
                'query_json': params.get('hotword_api_query_json'),
                'body_json': params.get('hotword_api_body_json'),
                'result_path': params.get('hotword_result_path'),
                'keyword_param': params.get('hotword_keyword_param'),
                'timeout_seconds': params.get('hotword_timeout_seconds'),
                'trend_type': params.get('hotword_trend_type'),
                'page_size': params.get('hotword_page_size'),
                'max_related_queries': params.get('hotword_max_related_queries'),
                'date_from': params.get('hotword_date_from'),
                'date_to': params.get('hotword_date_to'),
            },
            keywords,
            source_platform=source_platform,
            source_channel=source_channel,
            batch_name=batch_name,
        )
        response_preview = remote_result.get('response_preview') or {}
        if isinstance(response_preview, list):
            response_preview = response_preview[:5]
        elif isinstance(response_preview, dict):
            preview_copy = dict(response_preview)
            for key in ['items', 'data', 'results', 'list']:
                if isinstance(preview_copy.get(key), list):
                    preview_copy[key] = preview_copy[key][:5]
            response_preview = preview_copy
        rows = _normalize_trend_items(
            remote_result.get('items') or [],
            template_key=template_key,
            source_platform=source_platform,
            source_channel=source_channel,
            batch_name=batch_name,
        )
        return {
            'mode': 'remote',
            'template_key': template_key,
            'rows': rows,
            'request_preview': remote_result.get('request_preview') or {},
            'response_preview': response_preview,
        }

    return {
        'mode': 'skeleton',
        'template_key': template_key,
        'rows': _build_hotword_skeleton_rows(
            keywords,
            source_platform=source_platform,
            source_channel=source_channel,
            batch_name=batch_name,
        ),
        'request_preview': {},
        'response_preview': {},
    }


def _tracked_creator_sync_targets(limit=20, registration_id=0, creator_account_id=0):
    limit = min(max(_safe_int(limit, 20), 1), 200)
    registration_id = _safe_int(registration_id, 0)
    creator_account_id = _safe_int(creator_account_id, 0)

    rows = []
    for submission in Submission.query.order_by(Submission.id.asc()).all():
        registration = submission.registration
        if not registration:
            continue
        if registration_id and registration.id != registration_id:
            continue
        if creator_account_id and _safe_int(submission.xhs_creator_account_id, 0) != creator_account_id:
            continue

        profile_url = (submission.xhs_profile_link or '').strip()
        account = CreatorAccount.query.get(submission.xhs_creator_account_id) if submission.xhs_creator_account_id else None
        account_handle = (
            (account.account_handle if account else '') or
            (registration.xhs_account or '')
        ).strip()
        owner_phone = ((account.owner_phone if account else '') or (registration.phone or '')).strip()
        if not profile_url and account and account.profile_url:
            profile_url = (account.profile_url or '').strip()

        if not (profile_url or account_handle or owner_phone):
            continue

        rows.append({
            'registration_id': registration.id,
            'submission_id': submission.id,
            'topic_id': registration.topic_id,
            'creator_account_id': _safe_int(submission.xhs_creator_account_id, 0) or (account.id if account else 0),
            'profile_url': profile_url,
            'account_handle': account_handle,
            'owner_name': ((account.owner_name if account else '') or (registration.name or '')).strip(),
            'owner_phone': owner_phone,
            'last_synced_at': submission.xhs_last_synced_at or (account.last_synced_at if account else None) or submission.created_at,
            'note_url': (submission.xhs_link or '').strip(),
        })

    rows.sort(key=lambda item: item.get('last_synced_at') or datetime(1970, 1, 1))

    deduped = []
    dedupe_map = {}
    for item in rows:
        dedupe_key = (
            item.get('creator_account_id') or 0,
            item.get('profile_url') or '',
            item.get('account_handle') or '',
            item.get('owner_phone') or '',
        )
        existing = dedupe_map.get(dedupe_key)
        if not existing:
            item['registration_ids'] = [item['registration_id']]
            item['submission_ids'] = [item['submission_id']]
            dedupe_map[dedupe_key] = item
            deduped.append(item)
            continue
        if item['registration_id'] not in existing['registration_ids']:
            existing['registration_ids'].append(item['registration_id'])
        if item['submission_id'] not in existing['submission_ids']:
            existing['submission_ids'].append(item['submission_id'])
        if item.get('last_synced_at') and (not existing.get('last_synced_at') or item['last_synced_at'] < existing['last_synced_at']):
            existing['last_synced_at'] = item['last_synced_at']

    result = []
    for item in deduped[:limit]:
        result.append({
            **item,
            'last_synced_at': _format_datetime(item.get('last_synced_at')),
        })
    return result


def _build_creator_sync_request_config(payload=None):
    payload = payload or {}
    runtime_settings = _creator_sync_runtime_settings()
    merged = dict(runtime_settings)
    override_keys = [
        'creator_sync_source_channel',
        'creator_sync_fetch_mode',
        'creator_sync_api_url',
        'creator_sync_api_method',
        'creator_sync_api_headers_json',
        'creator_sync_api_query_json',
        'creator_sync_api_body_json',
        'creator_sync_result_path',
        'creator_sync_timeout_seconds',
        'creator_sync_batch_limit',
        'creator_sync_current_month_only',
        'creator_sync_date_from',
        'creator_sync_date_to',
        'creator_sync_max_posts_per_account',
    ]
    allow_blank_override_keys = {'creator_sync_date_from', 'creator_sync_date_to'}
    for key in override_keys:
        if key not in payload:
            continue
        value = payload.get(key)
        if value is None:
            continue
        if value == '' and key not in allow_blank_override_keys:
            continue
        merged[key] = value
    merged['creator_sync_source_channel'] = (merged.get('creator_sync_source_channel') or 'Crawler服务').strip()[:50] or 'Crawler服务'
    merged['creator_sync_timeout_seconds'] = min(max(_safe_int(merged.get('creator_sync_timeout_seconds'), 60), 5), 300)
    merged['creator_sync_batch_limit'] = min(max(_safe_int(merged.get('creator_sync_batch_limit'), 20), 1), 200)
    merged['creator_sync_current_month_only'] = _coerce_bool(merged.get('creator_sync_current_month_only')) if 'creator_sync_current_month_only' in merged else True
    merged['creator_sync_date_from'] = (merged.get('creator_sync_date_from') or '').strip()
    merged['creator_sync_date_to'] = (merged.get('creator_sync_date_to') or '').strip()
    merged['creator_sync_max_posts_per_account'] = min(max(_safe_int(merged.get('creator_sync_max_posts_per_account'), 60), 1), 100)
    merged['creator_sync_fetch_mode'] = _resolved_creator_sync_mode(merged)
    merged['creator_sync_api_method'] = (merged.get('creator_sync_api_method') or 'POST').strip().upper() or 'POST'
    return merged


def _build_creator_sync_remote_preview(payload=None, targets=None, source_channel='', batch_name=''):
    targets = targets or []
    request_config = _build_creator_sync_request_config(payload)
    return build_creator_sync_request_preview(
        {
            'api_url': request_config.get('creator_sync_api_url'),
            'api_method': request_config.get('creator_sync_api_method'),
            'headers_json': request_config.get('creator_sync_api_headers_json'),
            'query_json': request_config.get('creator_sync_api_query_json'),
            'body_json': request_config.get('creator_sync_api_body_json'),
            'result_path': request_config.get('creator_sync_result_path'),
            'timeout_seconds': request_config.get('creator_sync_timeout_seconds'),
            'current_month_only': request_config.get('creator_sync_current_month_only', True),
            'date_from': request_config.get('creator_sync_date_from', ''),
            'date_to': request_config.get('creator_sync_date_to', ''),
            'max_posts_per_account': request_config.get('creator_sync_max_posts_per_account', 60),
        },
        targets,
        source_channel=source_channel or request_config.get('creator_sync_source_channel') or 'Crawler服务',
        batch_name=batch_name,
    )


def _preview_hotword_rows(payload=None, sample_keywords=None, source_platform='', source_channel='', batch_name=''):
    settings = _build_hotword_request_config(payload)
    mode = _resolved_hotword_mode(settings)
    keywords = sample_keywords or _automation_keyword_seeds()[:min(max(_safe_int(settings.get('hotword_keyword_limit'), 3), 1), 3)]
    template_key = (settings.get('hotword_source_template') or _hotword_runtime_settings().get('hotword_source_template') or 'generic_lines').strip()
    source_platform = (source_platform or settings.get('hotword_source_platform') or _hotword_runtime_settings().get('hotword_source_platform') or '小红书').strip()
    source_channel = (source_channel or settings.get('hotword_source_channel') or _hotword_runtime_settings().get('hotword_source_channel') or 'Worker骨架').strip()
    batch_name = (batch_name or settings.get('batch_name') or 'preview_hotword').strip()

    if mode == 'remote':
        remote_result = fetch_remote_hotword_items(
            {
                'api_url': settings.get('hotword_api_url'),
                'api_method': settings.get('hotword_api_method'),
                'headers_json': settings.get('hotword_api_headers_json'),
                'query_json': settings.get('hotword_api_query_json'),
                'body_json': settings.get('hotword_api_body_json'),
                'result_path': settings.get('hotword_result_path'),
                'keyword_param': settings.get('hotword_keyword_param'),
                'timeout_seconds': settings.get('hotword_timeout_seconds'),
            },
            keywords,
            source_platform=source_platform,
            source_channel=source_channel,
            batch_name=batch_name,
        )
        response_preview = remote_result.get('response_preview') or {}
        if isinstance(response_preview, list):
            response_preview = response_preview[:5]
        elif isinstance(response_preview, dict):
            preview_copy = dict(response_preview)
            for key in ['items', 'data', 'results', 'list']:
                if isinstance(preview_copy.get(key), list):
                    preview_copy[key] = preview_copy[key][:5]
            response_preview = preview_copy
        rows = _normalize_trend_items(
            remote_result.get('items') or [],
            template_key=template_key,
            source_platform=source_platform,
            source_channel=source_channel,
            batch_name=batch_name,
        )
        return {
            'mode': mode,
            'template_key': template_key,
            'keywords': keywords,
            'rows': rows,
            'request_preview': remote_result.get('request_preview') or {},
            'response_preview': response_preview,
        }

    rows = _build_hotword_skeleton_rows(
        keywords,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    )
    return {
        'mode': 'skeleton',
        'template_key': template_key,
        'keywords': keywords,
        'rows': rows,
        'request_preview': {},
        'response_preview': {},
    }


def _hotword_healthcheck(payload=None, timeout_seconds=3, sample_keywords=None, include_rows=False):
    settings = _build_hotword_request_config(payload)
    mode = _resolved_hotword_mode(settings)
    api_url = (settings.get('hotword_api_url') or '').strip()
    if mode != 'remote':
        return {
            'enabled': False,
            'ok': False,
            'message': '热点抓取模式未启用 remote',
            'health_url': '',
            'status_code': 0,
            'response': None,
            'request_preview': {},
            'normalized_preview': [],
        }
    if not api_url:
        return {
            'enabled': True,
            'ok': False,
            'message': '未配置热点 API URL',
            'health_url': '',
            'status_code': 0,
            'response': None,
            'request_preview': {},
            'normalized_preview': [],
        }

    keywords = sample_keywords or _automation_keyword_seeds()[:min(max(_safe_int(settings.get('hotword_keyword_limit'), 3), 1), 3)]
    preview = _build_hotword_remote_preview(
        settings,
        keywords,
        source_platform=settings.get('hotword_source_platform') or '小红书',
        source_channel=settings.get('hotword_source_channel') or 'Worker骨架',
        batch_name='healthcheck_hotword',
    )
    health_url = api_url
    health_response = None
    parsed = urlparse(api_url)
    if parsed.scheme and parsed.netloc:
        health_url = urlunparse((parsed.scheme, parsed.netloc, '/healthz', '', '', ''))
        try:
            response = requests.get(health_url, timeout=min(max(_safe_int(timeout_seconds, 3), 1), 15))
            try:
                health_response = response.json()
            except ValueError:
                health_response = response.text[:300]
        except Exception:
            health_response = None

    try:
        preview_data = _preview_hotword_rows(
            settings,
            sample_keywords=keywords,
            source_platform=settings.get('hotword_source_platform') or '小红书',
            source_channel=settings.get('hotword_source_channel') or 'Worker骨架',
            batch_name='healthcheck_hotword',
        )
        response_preview = preview_data.get('response_preview')
        if isinstance(health_response, dict):
            response_preview = {
                'health': health_response,
                'sample_response': response_preview,
            }
        return {
            'enabled': True,
            'ok': True,
            'message': '热点源接口可用',
            'health_url': health_url,
            'status_code': 200,
            'response': response_preview,
            'request_preview': preview,
            'normalized_preview': preview_data.get('rows', [])[:5] if include_rows else [],
        }
    except Exception as exc:
        status_code = getattr(getattr(exc, 'response', None), 'status_code', 0) or 0
        return {
            'enabled': True,
            'ok': False,
            'message': f'热点源接口不可达：{exc}',
            'health_url': health_url,
            'status_code': status_code,
            'response': {'health': health_response} if isinstance(health_response, dict) else None,
            'request_preview': preview,
            'normalized_preview': [],
        }


def _dispatch_hotword_sync(payload, actor='system'):
    runtime_config = _automation_runtime_config()
    request_config = _build_hotword_request_config(payload)
    source_platform = (payload.get('source_platform') or str(runtime_config.get('hotword_source_platform') or '小红书')).strip()
    source_channel = (payload.get('source_channel') or str(runtime_config.get('hotword_source_channel') or 'Worker骨架')).strip()
    mode = request_config.get('hotword_fetch_mode') or 'skeleton'
    keyword_limit = min(max(_safe_int(payload.get('keyword_limit'), runtime_config.get('hotword_keyword_limit') or 10), 1), 30)
    batch_name = (payload.get('batch_name') or datetime.now().strftime('hotword_%Y%m%d_%H%M%S')).strip()[:120]
    scope_preset = (payload.get('scope_preset') or str(runtime_config.get('hotword_scope_preset') or 'liver_comorbidity')).strip() or 'liver_comorbidity'
    scope_meta = _hotword_scope_preset_meta(scope_preset)
    resolved_window = _resolve_hotword_date_window(
        payload.get('time_window') or request_config.get('hotword_time_window') or '30d',
        custom_from=(payload.get('date_from') if 'date_from' in payload else request_config.get('hotword_date_from')) or '',
        custom_to=(payload.get('date_to') if 'date_to' in payload else request_config.get('hotword_date_to')) or '',
    )
    auto_generate_topic_ideas = _coerce_bool(payload.get('hotword_auto_generate_topic_ideas')) if 'hotword_auto_generate_topic_ideas' in payload else _coerce_bool(runtime_config.get('hotword_auto_generate_topic_ideas'))
    auto_generate_topic_count = min(max(_safe_int(payload.get('hotword_auto_generate_topic_count'), runtime_config.get('hotword_auto_generate_topic_count') or 20), 1), 120)
    auto_generate_topic_activity_id = max(_safe_int(payload.get('hotword_auto_generate_topic_activity_id'), runtime_config.get('hotword_auto_generate_topic_activity_id') or 0), 0)
    if auto_generate_topic_ideas and auto_generate_topic_activity_id <= 0:
        auto_generate_topic_activity_id = _default_activity_id_for_automation()
    auto_generate_topic_quota = min(max(_safe_int(payload.get('hotword_auto_generate_topic_quota'), runtime_config.get('hotword_auto_generate_topic_quota') or _default_topic_quota()), 1), 300)
    auto_convert_corpus_templates = _coerce_bool(payload.get('hotword_auto_convert_corpus_templates')) if 'hotword_auto_convert_corpus_templates' in payload else _coerce_bool(runtime_config.get('hotword_auto_convert_corpus_templates'))
    auto_convert_corpus_limit = min(max(_safe_int(payload.get('hotword_auto_convert_corpus_limit'), runtime_config.get('hotword_auto_convert_corpus_limit') or 10), 1), 50)
    raw_keywords = payload.get('keywords')
    if isinstance(raw_keywords, list):
        keyword_items = [str(item).strip() for item in raw_keywords if str(item).strip()]
    else:
        keyword_items = split_hotword_keywords(raw_keywords or '')
    resolved_scope_keywords = _resolve_hotword_scope_keywords(scope_preset, raw_keywords or '')
    keywords = keyword_items[:keyword_limit] if keyword_items else (resolved_scope_keywords[:keyword_limit] if resolved_scope_keywords else _automation_keyword_seeds()[:keyword_limit])
    template_key = (payload.get('template_key') or scope_meta.get('preferred_template_key') or runtime_config.get('hotword_source_template') or 'generic_lines').strip()
    if not payload.get('template_key') and scope_meta.get('preferred_trend_type'):
        request_config['hotword_trend_type'] = payload.get('hotword_trend_type') or scope_meta.get('preferred_trend_type')
    request_config['hotword_date_from'] = resolved_window['date_from']
    request_config['hotword_date_to'] = resolved_window['date_to']
    remote_preview = _build_hotword_remote_preview(
        request_config,
        keywords,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    ) if mode == 'remote' else {}

    task_record = DataSourceTask(
        task_type='hotword_sync',
        source_platform=source_platform,
        source_channel=source_channel,
        mode=mode,
        status='queued',
        batch_name=batch_name,
        keyword_limit=len(keywords),
        activity_id=_safe_int(payload.get('activity_id'), 0) or None,
        message='等待 Worker 执行远端热点抓取' if mode == 'remote' else '等待 Worker 执行热点抓取骨架',
        params_payload=json.dumps({
            'keywords': keywords,
            'keyword_limit': keyword_limit,
            'source_platform': source_platform,
            'source_channel': source_channel,
            'mode': mode,
            'template_key': template_key,
            'batch_name': batch_name,
            'scope_preset': scope_preset,
            'scope_label': scope_meta.get('label') or scope_preset,
            'time_window': resolved_window['window_key'],
            'date_from': resolved_window['date_from'],
            'date_to': resolved_window['date_to'],
            'hotword_api_url': request_config.get('hotword_api_url'),
            'hotword_api_method': request_config.get('hotword_api_method'),
            'hotword_api_headers_json': request_config.get('hotword_api_headers_json'),
            'hotword_api_query_json': request_config.get('hotword_api_query_json'),
            'hotword_api_body_json': request_config.get('hotword_api_body_json'),
            'hotword_result_path': request_config.get('hotword_result_path'),
            'hotword_keyword_param': request_config.get('hotword_keyword_param'),
            'hotword_timeout_seconds': request_config.get('hotword_timeout_seconds'),
            'hotword_trend_type': request_config.get('hotword_trend_type'),
            'hotword_page_size': request_config.get('hotword_page_size'),
            'hotword_max_related_queries': request_config.get('hotword_max_related_queries'),
            'hotword_auto_generate_topic_ideas': auto_generate_topic_ideas,
            'hotword_auto_generate_topic_count': auto_generate_topic_count,
            'hotword_auto_generate_topic_activity_id': auto_generate_topic_activity_id,
            'hotword_auto_generate_topic_quota': auto_generate_topic_quota,
            'hotword_auto_convert_corpus_templates': auto_convert_corpus_templates,
            'hotword_auto_convert_corpus_limit': auto_convert_corpus_limit,
            'request_preview': remote_preview,
        }, ensure_ascii=False),
    )
    db.session.add(task_record)
    db.session.flush()
    _append_data_source_log(task_record.id, '已创建热点抓取任务，等待 Worker 处理', detail={
        'keywords': keywords,
        'source_platform': source_platform,
        'source_channel': source_channel,
        'mode': mode,
        'template_key': template_key,
        'batch_name': batch_name,
        'scope_preset': scope_preset,
        'scope_label': scope_meta.get('label') or scope_preset,
        'time_window': resolved_window['window_key'],
        'date_from': resolved_window['date_from'],
        'date_to': resolved_window['date_to'],
        'request_preview': remote_preview,
        'auto_generate_topic_ideas': auto_generate_topic_ideas,
        'auto_generate_topic_count': auto_generate_topic_count,
        'auto_generate_topic_activity_id': auto_generate_topic_activity_id,
        'auto_generate_topic_quota': auto_generate_topic_quota,
        'auto_convert_corpus_templates': auto_convert_corpus_templates,
        'auto_convert_corpus_limit': auto_convert_corpus_limit,
    })

    from celery_app import sync_hotwords_job

    inline_mode = _env_flag('INLINE_AUTOMATION_JOBS', False)
    task_record_id = task_record.id
    if inline_mode:
        db.session.commit()
        async_task = _enqueue_task(sync_hotwords_job, task_record_id)
        task_record = DataSourceTask.query.get(task_record_id)
    else:
        async_task = _enqueue_task(sync_hotwords_job, task_record_id)
    task_record.celery_task_id = async_task.id
    task_record.updated_at = datetime.now()
    _log_operation('dispatch_job', 'data_source_task', target_id=task_record.id, message='触发热点抓取 Worker 任务', detail={
        'task_id': async_task.id,
        'job': 'jobs.hotwords.sync',
        'source_platform': source_platform,
        'batch_name': batch_name,
        'keyword_count': len(keywords),
        'mode': mode,
        'template_key': template_key,
        'scope_preset': scope_preset,
        'time_window': resolved_window['window_key'],
        'date_from': resolved_window['date_from'],
        'date_to': resolved_window['date_to'],
        'auto_convert_corpus_templates': auto_convert_corpus_templates,
        'auto_convert_corpus_limit': auto_convert_corpus_limit,
        'actor': actor,
    })
    db.session.commit()
    return {
        'task_record': task_record,
        'task_id': async_task.id,
        'keyword_count': len(keywords),
    }


def _dispatch_hotword_planning_bundle(payload, actor='system'):
    scope_keys = ['liver_comorbidity', 'science_qna', 'xhs_trending']
    base_time_window = (payload.get('time_window') or '').strip()
    base_date_from = (payload.get('date_from') or '').strip()
    base_date_to = (payload.get('date_to') or '').strip()
    dispatched_items = []
    base_batch_prefix = datetime.now().strftime('planning_%Y%m%d_%H%M%S')
    for index, scope_key in enumerate(scope_keys, start=1):
        scope_meta = _hotword_scope_preset_meta(scope_key)
        scope_payload = {
            **dict(payload or {}),
            'scope_preset': scope_key,
            'batch_name': f'{base_batch_prefix}_{scope_key}',
            'template_key': scope_meta.get('preferred_template_key') or payload.get('template_key') or '',
            'hotword_trend_type': scope_meta.get('preferred_trend_type') or payload.get('hotword_trend_type') or '',
            'keywords': ','.join(scope_meta.get('keywords') or []),
        }
        if base_time_window:
            scope_payload['time_window'] = base_time_window
        if base_date_from:
            scope_payload['date_from'] = base_date_from
        if base_date_to:
            scope_payload['date_to'] = base_date_to
        dispatched = _dispatch_hotword_sync(scope_payload, actor=actor)
        dispatched_items.append({
            'scope_preset': scope_key,
            'scope_label': scope_meta.get('label') or scope_key,
            'task_id': dispatched.get('task_id'),
            'task_record_id': dispatched.get('task_record').id if dispatched.get('task_record') else None,
            'keyword_count': dispatched.get('keyword_count') or 0,
        })
    return {
        'batch_name': base_batch_prefix,
        'items': dispatched_items,
    }


def _dispatch_creator_account_sync(payload, actor='system'):
    request_config = _build_creator_sync_request_config(payload)
    mode = request_config.get('creator_sync_fetch_mode') or 'disabled'
    if mode != 'remote':
        raise ValueError('账号同步接口尚未配置，请先在自动化中心填写账号同步 API URL')

    source_platform = (payload.get('source_platform') or '小红书').strip() or '小红书'
    source_channel = (payload.get('source_channel') or request_config.get('creator_sync_source_channel') or 'Crawler服务').strip() or 'Crawler服务'
    batch_limit = min(max(_safe_int(payload.get('batch_limit'), request_config.get('creator_sync_batch_limit') or 20), 1), 200)
    max_posts_per_account = min(max(_safe_int(payload.get('max_posts_per_account'), request_config.get('creator_sync_max_posts_per_account') or 60), 1), 100)
    now = datetime.now()
    default_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    current_month_only = (
        _coerce_bool(payload.get('current_month_only'))
        if 'current_month_only' in payload
        else _coerce_bool(request_config.get('creator_sync_current_month_only'))
    )
    date_from = (
        (payload.get('date_from') if 'date_from' in payload else request_config.get('creator_sync_date_from'))
        or ''
    ).strip()[:19]
    date_to = (
        (payload.get('date_to') if 'date_to' in payload else request_config.get('creator_sync_date_to'))
        or ''
    ).strip()[:19]
    if current_month_only:
        if not date_from:
            date_from = default_month_start.strftime('%Y-%m-%d')
        if not date_to:
            date_to = now.strftime('%Y-%m-%d')
    batch_name = (payload.get('batch_name') or datetime.now().strftime('creator_sync_%Y%m%d_%H%M%S')).strip()[:120]
    registration_id = _safe_int(payload.get('registration_id'), 0)
    creator_account_id = _safe_int(payload.get('creator_account_id'), 0)
    targets = _tracked_creator_sync_targets(
        limit=batch_limit,
        registration_id=registration_id,
        creator_account_id=creator_account_id,
    )
    if not targets:
        raise ValueError('当前没有可同步的报名人账号，请先填写账号主页链接并提交一次笔记')

    remote_preview = _build_creator_sync_remote_preview(
        {
            **request_config,
            'creator_sync_current_month_only': current_month_only,
            'creator_sync_date_from': date_from,
            'creator_sync_date_to': date_to,
            'creator_sync_max_posts_per_account': max_posts_per_account,
        },
        targets,
        source_channel=source_channel,
        batch_name=batch_name,
    )
    task_record = DataSourceTask(
        task_type='creator_account_sync',
        source_platform=source_platform,
        source_channel=source_channel,
        mode=mode,
        status='queued',
        batch_name=batch_name,
        keyword_limit=len(targets),
        message='等待 Worker 执行报名人账号同步',
        params_payload=json.dumps({
            'targets': targets,
            'batch_limit': batch_limit,
            'source_platform': source_platform,
            'source_channel': source_channel,
            'mode': mode,
            'batch_name': batch_name,
            'registration_id': registration_id,
            'creator_account_id': creator_account_id,
            'current_month_only': current_month_only,
            'date_from': date_from,
            'date_to': date_to,
            'max_posts_per_account': max_posts_per_account,
            'creator_sync_api_url': request_config.get('creator_sync_api_url'),
            'creator_sync_api_method': request_config.get('creator_sync_api_method'),
            'creator_sync_api_headers_json': request_config.get('creator_sync_api_headers_json'),
            'creator_sync_api_query_json': request_config.get('creator_sync_api_query_json'),
            'creator_sync_api_body_json': request_config.get('creator_sync_api_body_json'),
            'creator_sync_result_path': request_config.get('creator_sync_result_path'),
            'creator_sync_timeout_seconds': request_config.get('creator_sync_timeout_seconds'),
            'request_preview': remote_preview,
        }, ensure_ascii=False),
    )
    db.session.add(task_record)
    db.session.flush()
    _append_data_source_log(task_record.id, '已创建报名人账号同步任务，等待 Worker 处理', detail={
        'target_count': len(targets),
        'source_platform': source_platform,
        'source_channel': source_channel,
        'batch_name': batch_name,
        'registration_id': registration_id,
        'creator_account_id': creator_account_id,
        'current_month_only': current_month_only,
        'date_from': date_from,
        'date_to': date_to,
        'max_posts_per_account': max_posts_per_account,
        'request_preview': remote_preview,
    })

    from celery_app import sync_creator_accounts_job

    inline_mode = _env_flag('INLINE_AUTOMATION_JOBS', False)
    task_record_id = task_record.id
    if inline_mode:
        db.session.commit()
        async_task = _enqueue_task(sync_creator_accounts_job, task_record_id)
        task_record = DataSourceTask.query.get(task_record_id)
    else:
        async_task = _enqueue_task(sync_creator_accounts_job, task_record_id)
    task_record.celery_task_id = async_task.id
    task_record.updated_at = datetime.now()
    _log_operation('dispatch_job', 'data_source_task', target_id=task_record.id, message='触发报名人账号同步 Worker 任务', detail={
        'task_id': async_task.id,
        'job': 'jobs.creator_accounts.sync',
        'source_platform': source_platform,
        'batch_name': batch_name,
        'target_count': len(targets),
        'mode': mode,
        'current_month_only': current_month_only,
        'date_from': date_from,
        'date_to': date_to,
        'max_posts_per_account': max_posts_per_account,
        'actor': actor,
    })
    db.session.commit()
    return {
        'task_record': task_record,
        'task_id': async_task.id,
        'target_count': len(targets),
    }


def _dispatch_topic_idea_generation(payload, actor='system'):
    count = min(max(_safe_int(payload.get('count'), 80), 1), 120)
    activity_id = _safe_int(payload.get('activity_id'), 0) or None
    quota = _normalize_quota(payload.get('quota'))

    from celery_app import generate_topic_ideas_job

    async_task = _enqueue_task(generate_topic_ideas_job, count=count, activity_id=activity_id, quota=quota)
    _log_operation('dispatch_job', 'topic_idea', message='触发异步生成候选话题', detail={
        'task_id': async_task.id,
        'count': count,
        'activity_id': activity_id,
        'quota': quota,
        'actor': actor,
    })
    db.session.commit()
    return {
        'task_id': async_task.id,
        'count': count,
        'activity_id': activity_id,
        'quota': quota,
    }


def _dispatch_asset_generation(payload, actor='system'):
    max_remote_generation_attempts = 5
    runtime_config = _automation_runtime_config()
    registration_id = _safe_int(payload.get('registration_id'), 0)
    reg = Registration.query.get(registration_id) if registration_id else None
    if not reg:
        raise ValueError('报名信息不存在')

    configured_provider = (os.environ.get('ASSET_IMAGE_PROVIDER') or str(runtime_config.get('image_provider') or 'svg_fallback')).strip() or 'svg_fallback'
    if configured_provider != 'svg_fallback':
        used_attempts = _count_real_asset_generation_attempts(reg.id)
        remaining_attempts = max(max_remote_generation_attempts - used_attempts, 0)
        if remaining_attempts <= 0:
            raise ValueError(f'该报名记录的真实图片生成次数已用完（最多 {max_remote_generation_attempts} 次），请先复用现有图片或改用模板预览。')
    else:
        used_attempts = 0
        remaining_attempts = max_remote_generation_attempts

    selected_content = (payload.get('selected_content') or '').strip()
    product_meta = _product_profile_meta(payload.get('product_profile') or '')
    product_profile = (payload.get('product_profile') or '').strip()[:80]
    product_category = (payload.get('product_category') or product_meta.get('product_category') or '').strip()[:30]
    product_name = (payload.get('product_name') or product_meta.get('product_name') or '').strip()[:200]
    product_indication = (payload.get('product_indication') or product_meta.get('product_indication') or '').strip()[:200]
    product_asset_ids = _parse_int_list(payload.get('product_asset_ids') or '', limit=20)
    product_assets = _resolve_asset_library_rows(product_asset_ids, limit=20, library_type='product')
    reference_asset_ids = _parse_int_list(payload.get('reference_asset_ids') or '', limit=20)
    reference_assets = _resolve_reference_asset_rows(reference_asset_ids, limit=20)
    reference_text = ''
    product_asset_text = ''
    if product_assets:
        product_titles = [item.title or item.product_name or f'产品资产{item.id}' for item in product_assets[:3]]
        product_asset_text = f" 产品图素材：{(' / '.join(product_titles))}。"
    if reference_assets:
        ref_titles = [item.title or item.product_name or f'资产{item.id}' for item in reference_assets[:3]]
        reference_text = f" 参考图方向：{(' / '.join(ref_titles))}。"
    raw_style_type = (payload.get('style_type') or runtime_config.get('image_default_style_type') or 'medical_science')
    style_meta = _asset_style_meta(raw_style_type)
    custom_prompt = (payload.get('custom_prompt') or '').strip()
    image_count = min(max(_safe_int(payload.get('image_count'), 3), 1), 4)
    title_hint = (payload.get('title_hint') or _extract_title_from_version(selected_content) or reg.topic.topic_name).strip()[:200]
    decision = _resolve_asset_workflow_decision(
        style_value=raw_style_type,
        cover_style_type=(payload.get('cover_style_type') or ''),
        inner_style_type=(payload.get('inner_style_type') or ''),
        generation_mode=(payload.get('generation_mode') or 'smart_bundle'),
        selected_content=selected_content,
        title_hint=title_hint,
        reference_assets=reference_assets,
    )
    style_meta = decision['style_meta']
    style_preset = (decision['prompt_style_meta']['label'] or style_meta['label'])[:50]
    cover_style_meta = decision['cover_style_meta']
    inner_style_meta = decision['inner_style_meta']
    generation_mode = decision['generation_mode']
    prompt_text = _build_asset_generation_prompt(
        reg.topic,
        selected_content=selected_content,
        style_preset=decision['prompt_style_meta']['key'],
        title_hint=title_hint,
        cover_style_key=cover_style_meta.get('key') or style_meta['key'],
        inner_style_key=inner_style_meta.get('key') or style_meta['key'],
        generation_mode=generation_mode,
        image_count=image_count,
    )
    workflow_notes = [
        f'图片工作流模式：{generation_mode}。',
        f'封面样式：{cover_style_meta.get("label") or cover_style_meta.get("key") or "-"}。',
    ]
    if generation_mode == 'smart_bundle':
        workflow_notes.append(f'内页样式：{inner_style_meta.get("label") or inner_style_meta.get("key") or "-"}。')
        workflow_notes.append('如果生成多张图，请优先理解为图文套组：首张更像封面，后续更像内页。')
    elif generation_mode == 'cover_only':
        workflow_notes.append('这次只做封面或主图，不扩展内页。')
    elif generation_mode == 'inner_only':
        workflow_notes.append('这次只做内页风格，重点放结构化信息。')
    if decision.get('adjustment_note'):
        workflow_notes.append(decision['adjustment_note'])
    if custom_prompt:
        workflow_notes.append(f'自定义视觉要求：{custom_prompt}')
    prompt_text = f"{prompt_text} {' '.join(workflow_notes)}"
    if product_name:
        prompt_text = f"{prompt_text} 产品信息：{product_name}；适应方向：{product_indication or '未标记'}。"
    if product_asset_text:
        prompt_text = f"{prompt_text}{product_asset_text}"
    if reference_text:
        prompt_text = f"{prompt_text}{reference_text}"

    task = AssetGenerationTask(
        registration_id=reg.id,
        topic_id=reg.topic_id,
        source_provider=configured_provider,
        model_name=(os.environ.get('ASSET_IMAGE_MODEL') or str(runtime_config.get('image_model') or '')).strip()[:100],
        style_preset=style_preset,
        generation_mode=generation_mode,
        cover_style_type=cover_style_meta.get('key') or '',
        inner_style_type=inner_style_meta.get('key') or '',
        product_profile=product_profile,
        product_category=product_category,
        product_name=product_name,
        product_indication=product_indication,
        product_asset_ids=','.join(str(item) for item in product_asset_ids),
        reference_asset_ids=','.join(str(item) for item in reference_asset_ids),
        image_count=image_count,
        status='queued',
        title_hint=title_hint,
        prompt_text=prompt_text,
        selected_content=selected_content,
        message=('等待 Worker 生成图片任务（已自动调整封面样式）' if decision.get('auto_adjusted_cover') else '等待 Worker 生成图片任务'),
    )
    db.session.add(task)
    db.session.flush()
    _log_operation('dispatch_job', 'asset_generation_task', target_id=task.id, message='触发图片生成任务', detail={
        'registration_id': reg.id,
        'topic_id': reg.topic_id,
        'style_preset': style_preset,
        'style_type': style_meta['key'],
        'image_count': image_count,
        'source_provider': task.source_provider,
        'product_name': product_name,
        'product_category': product_category,
        'product_asset_ids': product_asset_ids,
        'reference_asset_ids': reference_asset_ids,
        'auto_adjusted_cover': bool(decision.get('auto_adjusted_cover')),
        'adjustment_note': decision.get('adjustment_note') or '',
        'actor': actor,
    })

    from celery_app import generate_asset_images_job

    inline_mode = _env_flag('INLINE_AUTOMATION_JOBS', False)
    task_id = task.id
    if inline_mode:
        db.session.commit()
        async_task = _enqueue_task(generate_asset_images_job, task_id)
        task = AssetGenerationTask.query.get(task_id)
    else:
        async_task = _enqueue_task(generate_asset_images_job, task_id)
    task.celery_task_id = async_task.id
    db.session.commit()
    return {
        'task_record': task,
        'task_id': async_task.id,
        'image_count': image_count,
        'used_attempts': used_attempts + (1 if configured_provider != 'svg_fallback' else 0),
        'remaining_attempts': max(remaining_attempts - (1 if configured_provider != 'svg_fallback' else 0), 0),
        'max_attempts': max_remote_generation_attempts,
        'decision': {
            'auto_adjusted_cover': bool(decision.get('auto_adjusted_cover')),
            'adjustment_note': decision.get('adjustment_note') or '',
            'cover_style_key': cover_style_meta.get('key') or '',
            'cover_style_label': cover_style_meta.get('label') or '',
            'inner_style_key': inner_style_meta.get('key') or '',
            'inner_style_label': inner_style_meta.get('label') or '',
        },
    }


def _build_asset_generation_plan_payload(payload):
    runtime_config = _automation_runtime_config()
    registration_id = _safe_int(payload.get('registration_id'), 0)
    reg = Registration.query.get(registration_id) if registration_id else None
    style_value = (payload.get('style_type') or runtime_config.get('image_default_style_type') or 'medical_science')
    style_meta = _asset_style_meta(style_value)
    selected_content = (payload.get('selected_content') or '').strip()
    title_hint = (payload.get('title_hint') or _extract_title_from_version(selected_content) or (reg.topic.topic_name if reg and reg.topic else '') or style_meta.get('label') or '图片方案').strip()[:200]
    product_meta = _product_profile_meta(payload.get('product_profile') or '')
    product_profile = (payload.get('product_profile') or '').strip()[:80]
    product_category = (payload.get('product_category') or product_meta.get('product_category') or '').strip()[:30]
    product_name = (payload.get('product_name') or product_meta.get('product_name') or '').strip()[:200]
    product_indication = (payload.get('product_indication') or product_meta.get('product_indication') or '').strip()[:200]
    product_asset_ids = _parse_int_list(payload.get('product_asset_ids') or '', limit=20)
    product_rows = _resolve_asset_library_rows(product_asset_ids, limit=20, library_type='product')
    product_assets = [{
        'id': item.id,
        'title': item.title or '',
        'preview_url': item.preview_url or '',
        'visual_role': item.visual_role or '',
        'product_name': item.product_name or '',
    } for item in product_rows]
    reference_asset_ids = _parse_int_list(payload.get('reference_asset_ids') or '', limit=20)
    reference_rows = _resolve_reference_asset_rows(reference_asset_ids, limit=20)
    reference_assets = [{
        'id': item.id,
        'title': item.title or '',
        'preview_url': item.preview_url or '',
        'product_name': item.product_name or '',
        'library_type': item.library_type or '',
    } for item in reference_rows]
    topic = reg.topic if reg else None
    decision = _resolve_asset_workflow_decision(
        style_value=style_value,
        cover_style_type=(payload.get('cover_style_type') or ''),
        inner_style_type=(payload.get('inner_style_type') or ''),
        generation_mode=(payload.get('generation_mode') or ''),
        selected_content=selected_content,
        title_hint=title_hint,
        reference_assets=reference_assets,
    )
    style_meta = decision['style_meta']
    cover_style_meta = decision['cover_style_meta']
    inner_style_meta = decision['inner_style_meta']
    generation_mode = decision['generation_mode']
    cover_fit = decision['cover_fit']

    prompt_text = _build_asset_generation_prompt(
        topic or type('TopicLike', (), {
            'topic_name': product_indication or '肝病管理',
            'keywords': product_indication or '肝病管理',
        })(),
        selected_content=selected_content,
        style_preset=decision['prompt_style_meta']['key'],
        title_hint=title_hint,
        cover_style_key=cover_style_meta.get('key') or style_meta['key'],
        inner_style_key=inner_style_meta.get('key') or style_meta['key'],
        generation_mode=generation_mode,
        image_count=min(max(_safe_int(payload.get('image_count'), 1), 1), 4),
    )
    if product_name:
        prompt_text = f"{prompt_text} 产品信息：{product_name}；适应方向：{product_indication or '未标记'}。"
    if product_assets:
        product_titles = []
        for item in product_assets[:3]:
            product_titles.append(item.get('title') or item.get('product_name') or f"产品资产{item.get('id')}")
        prompt_text = f"{prompt_text} 产品图素材：{' / '.join(product_titles)}。"
    if reference_assets:
        reference_titles = []
        for item in reference_assets[:3]:
            reference_titles.append(item.get('title') or item.get('product_name') or f"资产{item.get('id')}")
        prompt_text = f"{prompt_text} 参考图方向：{' / '.join(reference_titles)}。"

    product_context = {
        'product_profile': product_profile,
        'product_category': product_category,
        'product_name': product_name,
        'product_indication': product_indication,
    }
    request_preview = _build_asset_provider_request_preview(
        (os.environ.get('ASSET_IMAGE_PROVIDER') or str(runtime_config.get('image_provider') or 'svg_fallback')).strip() or 'svg_fallback',
        (os.environ.get('ASSET_IMAGE_MODEL') or str(runtime_config.get('image_model') or '')).strip()[:100],
        prompt_text,
        (os.environ.get('ASSET_IMAGE_SIZE') or str(runtime_config.get('image_size') or '1024x1536')).strip(),
        style_preset=style_meta['key'],
        image_count=min(max(_safe_int(payload.get('image_count'), 1), 1), 4),
        product_assets=product_assets,
        reference_assets=reference_assets,
        product_context=product_context,
    )

    points = _extract_content_points(selected_content)
    strategy_reason = ''
    if generation_mode == 'template_first':
        strategy_reason = '当前类型更适合模板直出，重点在版式和文字层级，不必优先依赖图片模型。'
    elif reference_assets:
        strategy_reason = '当前已带参考图，更适合走参考图驱动或 img2img 路线。'
    elif product_assets:
        strategy_reason = '当前已带真实产品图，适合做产品图合成或产品主视觉辅助生成。'
    else:
        strategy_reason = '当前更适合先用文案驱动生成底图，后续再通过产品图或参考图增强。'
    if decision.get('adjustment_note'):
        strategy_reason = f"{strategy_reason} {decision.get('adjustment_note')}"
    if cover_fit['score'] < 72:
        strategy_reason = f"{strategy_reason} 当前封面适配度偏低，建议优先切到“{cover_fit['fallback_cover_style_label']}”或先预览图文套组。"

    overlay_plan = {
        'headline': title_hint,
        'subheadline': product_name or (style_meta.get('description') or ''),
        'bullet_points': points[:3] if points else list(style_meta.get('default_bullets') or []),
        'postprocess': '先出无字底图，再由系统叠加中文标题、说明卡片和标签。',
    }
    preview_topic = topic or type('TopicLike', (), {
        'topic_name': product_indication or '肝病管理',
        'keywords': product_indication or '肝病管理',
        'id': 0,
    })()
    preview_assets = _build_asset_generation_fallback_results(
        preview_topic,
        selected_content=selected_content,
        image_count=1,
        style_preset=decision['prompt_style_meta'].get('key') or style_meta.get('key') or '',
        title_hint=title_hint,
    )
    preview_asset = (preview_assets or [{}])[0] if preview_assets else {}
    return {
        'success': True,
        'plan': {
            'registration_id': registration_id,
            'registration_name': reg.name if reg else '',
            'topic_name': topic.topic_name if topic else '',
            'style_type': style_meta.get('key') or '',
            'style_label': style_meta.get('label') or '',
            'generation_mode': generation_mode,
            'cover_style_type': cover_style_meta.get('key') or '',
            'cover_style_label': cover_style_meta.get('label') or '',
            'inner_style_type': inner_style_meta.get('key') or '',
            'inner_style_label': inner_style_meta.get('label') or '',
            'strategy_reason': strategy_reason,
            'cover_fit_score': cover_fit['score'],
            'cover_fit_label': cover_fit['label'],
            'cover_fit_reason': cover_fit['reason'],
            'fallback_cover_style_key': cover_fit['fallback_style_key'],
            'fallback_cover_style_label': cover_fit['fallback_style_label'],
            'execution_note': cover_fit['execution_note'],
            'auto_adjusted_cover': bool(decision.get('auto_adjusted_cover')),
            'adjustment_note': decision.get('adjustment_note') or '',
            'original_cover_style_key': decision.get('original_cover_style_key') or '',
            'original_cover_style_label': decision.get('original_cover_style_label') or '',
            'title_hint': title_hint,
            'product_context': product_context,
            'product_assets': product_assets,
            'reference_assets': reference_assets,
            'content_points': points[:5],
            'prompt_text': prompt_text,
            'request_preview': request_preview,
            'preview_asset': preview_asset,
            'overlay_plan': overlay_plan,
        }
    }


def _build_asset_style_recommendation_payload(payload):
    selected_content = (payload.get('selected_content') or '').strip()
    title_hint = (payload.get('title_hint') or '').strip()
    merged = ' '.join(filter(None, [title_hint, selected_content]))
    lowered = merged.lower()
    traits = _detect_topic_strategy_traits(merged, merged, merged)
    product_meta = _product_profile_meta(payload.get('product_profile') or '')
    product_name = (payload.get('product_name') or product_meta.get('product_name') or '').strip()
    product_category = (payload.get('product_category') or product_meta.get('product_category') or '').strip()
    product_asset_ids = _parse_int_list(payload.get('product_asset_ids') or '', limit=20)
    selected_product_assets = _resolve_asset_library_rows(product_asset_ids, limit=20, library_type='product')
    all_product_assets = AssetLibrary.query.filter_by(library_type='product', product_name=product_name).all() if product_name else []
    product_asset_count = len(all_product_assets)
    coverage_hint = ''
    if product_name:
        if product_asset_count == 0:
            coverage_hint = f'当前产品“{product_name}”还没有任何真实产品图，优先补主图和说明图。'
        elif product_asset_count < 3:
            coverage_hint = f'当前产品“{product_name}”真实产品图还不够多，适合先走模板和产品图轻合成。'
        else:
            coverage_hint = f'当前产品“{product_name}”真实产品图覆盖较好，可以更放心做产品图合成。'
    suggestions = []

    def add_style(style_key, reason):
        meta = _asset_style_meta(style_key)
        family_key = meta.get('family') or meta.get('key') or style_key
        cover_fit = _score_cover_suitability(
            style_key=meta.get('key') or style_key,
            family_key=family_key,
            generation_mode=meta.get('generation_mode') or 'text_to_image',
            selected_content=selected_content,
            title_hint=title_hint,
            reference_guided=bool(selected_product_assets),
        )
        ranking_bonus = 0
        if traits.get('report_like') and family_key in {'medical_science', 'checklist'}:
            ranking_bonus += 8 if family_key == 'medical_science' else 4
        if traits.get('story_like') and family_key in {'memo', 'knowledge_card'}:
            ranking_bonus += 8 if family_key == 'memo' else 3
        if traits.get('myth_like') and family_key in {'poster', 'knowledge_card'}:
            ranking_bonus += 7 if family_key == 'poster' else 4
        if traits.get('discussion_like') and family_key in {'knowledge_card', 'checklist'}:
            ranking_bonus += 5
        if product_category == 'device' and family_key == 'medical_science':
            ranking_bonus += 4
        if product_category == 'medicine' and family_key == 'checklist':
            ranking_bonus += 4
        suggestions.append({
            'style_key': meta.get('key') or style_key,
            'style_label': meta.get('label') or style_key,
            'family_key': family_key,
            'reason': reason,
            'generation_mode': meta.get('generation_mode') or 'text_to_image',
            'asset_type': meta.get('asset_type') or '',
            'cover_fit_score': cover_fit['score'],
            'cover_fit_label': cover_fit['label'],
            'cover_fit_reason': cover_fit['reason'],
            'fallback_style_key': cover_fit['fallback_style_key'],
            'fallback_style_label': cover_fit['fallback_style_label'],
            'execution_note': cover_fit['execution_note'],
            'ranking_score': (cover_fit['score'] or 0) + ranking_bonus,
        })

    if any(token in lowered for token in ['指南', '版本', '必须看', '警惕', '经验', '总结', 'emo', '崩溃']):
        add_style('poster_bold' if '经验' not in lowered and '总结' not in lowered else 'poster_handwritten', '文案更像封面结论或经验表达，优先大字报路线。')
    if any(token in lowered for token in ['备忘录', '提醒', '技巧', '注意事项', '攻略', '补救', '计划']):
        add_style('memo_mobile', '文案适合做成可收藏的手机备忘录风格。')
    if any(token in lowered for token in ['病理', '并发症', '生理', '实验室', '检查要点', '代偿', '失代偿']):
        add_style('memo_classroom', '文案更像课堂重点整理，适合课堂笔记风格。')
    if any(token in lowered for token in ['怎么选', '白名单', '对比', '参数', '品牌', '哪个好']):
        add_style('checklist_table', '内容偏产品选择或参数对比，适合表格清单图。')
    if any(token in lowered for token in ['day', '早餐', '午餐', '晚餐', '7天', '7 天', '食谱', '几点']):
        add_style('checklist_timeline', '内容有明显时间轴和执行顺序，适合时间轴清单图。')
    if any(token in lowered for token in ['报告', '彩超', '化验', '指标', 'alt', 'ast', 'ggt', 'alp', '一次看懂']):
        add_style('checklist_report', '内容偏报告解读和指标说明，适合报告解读图。')
    if any(token in lowered for token in ['机制', '原理', '对比', '影响', '全身', '伤害']):
        add_style('knowledge_card', '内容适合做成收藏型知识卡片。')
    if any(token in lowered for token in ['症状', '信号', '警示', '求救', '发黄', '体检', '检查', '器官']):
        add_style('medical_science', '内容适合做医学科普信息图。')
    if product_category == 'device':
        add_style('medical_science', f'当前产品“{product_name or product_category}”更适合搭配医学科普或设备解读型画面。')
        add_style('knowledge_card', '器械产品也适合知识卡片式介绍和设备对比。')
    if product_category == 'medicine':
        add_style('checklist_table', f'当前产品“{product_name or product_category}”很适合做产品选择/参数清单或白名单风格。')
        add_style('poster_bold', '药品内容如果做封面，通常适合用大字报先抓注意力。')
    if selected_product_assets:
        add_style('reference_based', '你已经手动选中了产品图素材，适合尝试产品图合成或参考图生成路线。')
    if not suggestions:
        add_style('medical_science', '默认推荐医学科普类，适合作为通用科普配图。')
        add_style('knowledge_card', '也可尝试知识卡片类，适合收藏传播。')

    unique = []
    seen = set()
    for item in suggestions:
        if item['style_key'] in seen:
            continue
        if coverage_hint:
            item['coverage_hint'] = coverage_hint
        unique.append(item)
        seen.add(item['style_key'])
    unique.sort(
        key=lambda item: (
            item.get('ranking_score') or 0,
            item.get('cover_fit_score') or 0,
            item.get('style_key') == _recommended_cover_style_key_for_traits(traits),
        ),
        reverse=True,
    )
    if unique and (unique[0].get('cover_fit_score') or 0) < 72:
        fallback_key = unique[0].get('fallback_style_key') or _recommended_cover_style_key_for_traits(traits)
        if fallback_key and fallback_key not in {item.get('style_key') for item in unique}:
            fallback_meta = _asset_style_meta(fallback_key)
            fallback_fit = _score_cover_suitability(
                style_key=fallback_meta.get('key') or fallback_key,
                family_key=fallback_meta.get('family') or fallback_meta.get('key') or fallback_key,
                generation_mode=fallback_meta.get('generation_mode') or 'text_to_image',
                selected_content=selected_content,
                title_hint=title_hint,
                reference_guided=bool(selected_product_assets),
            )
            unique.insert(0, {
                'style_key': fallback_meta.get('key') or fallback_key,
                'style_label': fallback_meta.get('label') or fallback_key,
                'family_key': fallback_meta.get('family') or fallback_key,
                'reason': '当前内容更适合先切到这条封面路线，减少出图后“不像封面”的风险。',
                'generation_mode': fallback_meta.get('generation_mode') or 'text_to_image',
                'asset_type': fallback_meta.get('asset_type') or '',
                'cover_fit_score': fallback_fit['score'],
                'cover_fit_label': fallback_fit['label'],
                'cover_fit_reason': fallback_fit['reason'],
                'fallback_style_key': fallback_fit['fallback_style_key'],
                'fallback_style_label': fallback_fit['fallback_style_label'],
                'execution_note': fallback_fit['execution_note'],
                'ranking_score': (fallback_fit['score'] or 0) + 12,
                'coverage_hint': coverage_hint,
            })
    return {
        'success': True,
        'items': unique[:4],
    }


def _build_batch_asset_plan_drafts(payload):
    source_type = (payload.get('source_type') or 'trend').strip()
    raw_ids = payload.get('item_ids') or []
    item_ids = []
    for item in raw_ids:
        value = _safe_int(item, 0)
        if value > 0:
            item_ids.append(value)
    item_ids = list(dict.fromkeys(item_ids))[:12]
    if not item_ids:
        return {
            'success': False,
            'message': '请先选择要生成草案的内容',
            'items': [],
        }

    if source_type == 'trend':
        rows = TrendNote.query.filter(TrendNote.id.in_(item_ids)).all()
        serialized_rows = [_serialize_trend_note(item) for item in rows]
    elif source_type == 'idea':
        rows = TopicIdea.query.filter(TopicIdea.id.in_(item_ids)).all()
        serialized_rows = [_serialize_topic_idea(item) for item in rows]
    else:
        return {
            'success': False,
            'message': '不支持的草案来源类型',
            'items': [],
        }

    drafts = []
    bucket_counter = Counter()
    for item in serialized_rows:
        if source_type == 'trend':
            selected_content = '\n'.join(filter(None, [
                item.get('summary') or '',
                f"推荐理由：{item.get('recommended_reason')}" if item.get('recommended_reason') else '',
                f"图片模板：{item.get('template_agent_label')}" if item.get('template_agent_label') else '',
                f"封面样式：{item.get('cover_style_label')}" if item.get('cover_style_label') else '',
                f"内页样式：{item.get('inner_style_label')}" if item.get('inner_style_label') else '',
            ]))
            title_hint = (item.get('title') or item.get('keyword') or '图片草案').strip()
        else:
            selected_content = '\n'.join(filter(None, [
                item.get('angle') or '',
                item.get('asset_brief') or '',
                f"文案提示词：{item.get('copy_prompt')}" if item.get('copy_prompt') else '',
                f"图片模板：{item.get('template_agent_label')}" if item.get('template_agent_label') else '',
                f"封面样式：{item.get('cover_style_label')}" if item.get('cover_style_label') else '',
                f"内页样式：{item.get('inner_style_label')}" if item.get('inner_style_label') else '',
            ]))
            title_hint = (item.get('cover_title') or item.get('topic_title') or '图片草案').strip()

        plan_payload = {
            'style_type': item.get('cover_style_type') or item.get('inner_style_type') or 'medical_science',
            'generation_mode': 'smart_bundle',
            'cover_style_type': item.get('cover_style_type') or '',
            'inner_style_type': item.get('inner_style_type') or '',
            'title_hint': title_hint[:200],
            'selected_content': selected_content[:4000],
            'image_count': 1,
        }
        plan_result = _build_asset_generation_plan_payload(plan_payload)
        plan = (plan_result or {}).get('plan') or {}
        bucket_label = item.get('template_agent_label') or item.get('image_skill_label') or '未分类'
        bucket_counter[bucket_label] += 1
        drafts.append({
            'source_type': source_type,
            'source_id': item.get('id'),
            'source_title': item.get('title') or item.get('topic_title') or title_hint,
            'template_agent_label': item.get('template_agent_label') or '',
            'image_skill_label': item.get('image_skill_label') or '',
            'cover_style_label': item.get('cover_style_label') or '',
            'inner_style_label': item.get('inner_style_label') or '',
            'plan': {
                'style_type': plan.get('style_type') or '',
                'style_label': plan.get('style_label') or '',
                'generation_mode': plan.get('generation_mode') or '',
                'cover_style_type': plan.get('cover_style_type') or '',
                'cover_style_label': plan.get('cover_style_label') or '',
                'inner_style_type': plan.get('inner_style_type') or '',
                'inner_style_label': plan.get('inner_style_label') or '',
                'strategy_reason': plan.get('strategy_reason') or '',
                'title_hint': plan.get('title_hint') or '',
                'content_points': (plan.get('content_points') or [])[:3],
                'preview_asset': plan.get('preview_asset') or {},
            },
        })

    return {
        'success': True,
        'message': f'已生成 {len(drafts)} 条图片草案，先确认后再决定是否发起真实图片任务',
        'bucket_summary': [{'label': key, 'count': count} for key, count in bucket_counter.most_common()],
        'items': drafts,
    }


def _dispatch_automation_schedule(schedule, actor='system'):
    params = _load_json_value(schedule.params_payload, {})
    schedule.last_run_at = datetime.now()
    schedule.next_run_at = _next_schedule_time(schedule.interval_minutes, schedule.last_run_at)

    if schedule.task_type == 'hotword_sync':
        dispatched = _dispatch_hotword_sync(params, actor=actor)
        schedule.last_celery_task_id = dispatched['task_id']
        schedule.last_status = 'queued'
        schedule.last_message = f'已派发热点抓取任务 {dispatched["task_record"].id}'
    elif schedule.task_type == 'creator_account_sync':
        dispatched = _dispatch_creator_account_sync(params, actor=actor)
        schedule.last_celery_task_id = dispatched['task_id']
        schedule.last_status = 'queued'
        schedule.last_message = f'已派发账号同步任务 {dispatched["task_record"].id}'
    elif schedule.task_type == 'topic_idea_generate':
        dispatched = _dispatch_topic_idea_generation(params, actor=actor)
        schedule.last_celery_task_id = dispatched['task_id']
        schedule.last_status = 'queued'
        schedule.last_message = f'已派发候选话题任务，生成 {dispatched["count"]} 个'
    else:
        schedule.last_status = 'failed'
        schedule.last_message = f'未知任务类型：{schedule.task_type}'
        db.session.commit()
        raise ValueError(f'未知任务类型：{schedule.task_type}')

    db.session.commit()
    return dispatched


register_automation_dashboard_routes(app, {
    'admin_json_guard': _admin_json_guard,
    'default_topic_quota': _default_topic_quota,
    'asset_style_type_options': _asset_style_type_options,
    'image_provider_options': _image_provider_options,
    'image_provider_presets': _image_provider_presets,
    'image_model_options': _image_model_options,
    'product_category_options': _product_category_options,
    'product_visual_role_options': _product_visual_role_options,
    'product_profile_options': _product_profile_options,
    'safe_int': _safe_int,
    'build_readiness_checks': _build_readiness_checks,
    'build_project_status_payload': _build_project_status_payload,
    'default_activity_id_for_automation': _default_activity_id_for_automation,
    'bootstrap_demo_operational_data': _bootstrap_demo_operational_data,
    'clear_demo_operational_data': _clear_demo_operational_data,
    'build_deployment_helper_payload': _build_deployment_helper_payload,
    'build_deployment_blockers_payload': _build_deployment_blockers_payload,
    'build_launch_milestones_payload': _build_launch_milestones_payload,
    'build_integration_checklist_payload': _build_integration_checklist_payload,
    'build_integration_ping_history_payload': _build_integration_ping_history_payload,
    'build_first_run_playbooks_payload': _build_first_run_playbooks_payload,
    'build_integration_contract_payload': _build_integration_contract_payload,
    'build_integration_acceptance_payload': _build_integration_acceptance_payload,
    'build_release_manifest_payload': _build_release_manifest_payload,
    'build_trial_readiness_payload': _build_trial_readiness_payload,
    'build_go_live_readiness_payload': _build_go_live_readiness_payload,
    'build_go_live_checklist_payload': _build_go_live_checklist_payload,
    'build_post_launch_watchlist_payload': _build_post_launch_watchlist_payload,
    'build_integration_handoff_pack_payload': _build_integration_handoff_pack_payload,
    'build_capacity_readiness_payload': _build_capacity_readiness_payload,
    'build_recent_failed_jobs_payload': _build_recent_failed_jobs_payload,
    'build_service_matrix_payload': _build_service_matrix_payload,
    'build_crawler_probe_payload': _build_crawler_probe_payload,
    'automation_runtime_config': _automation_runtime_config,
    'hotword_runtime_settings': _hotword_runtime_settings,
    'hotword_scope_preset_meta': _hotword_scope_preset_meta,
    'resolve_hotword_scope_keywords': _resolve_hotword_scope_keywords,
    'resolve_hotword_date_window': _resolve_hotword_date_window,
    'creator_sync_runtime_settings': _creator_sync_runtime_settings,
    'image_provider_capabilities': _image_provider_capabilities,
    'build_asset_provider_request_preview': _build_asset_provider_request_preview,
    'build_asset_generation_prompt_from_context': _build_asset_generation_prompt_from_context,
    'asset_style_meta': _asset_style_meta,
    'hotword_source_template_meta': _hotword_source_template_meta,
    'hotword_source_template_options': _hotword_source_template_options,
    'hotword_remote_source_presets': _hotword_remote_source_presets,
    'automation_keyword_seeds': _automation_keyword_seeds,
    'build_hotword_remote_preview': _build_hotword_remote_preview,
    'resolved_hotword_mode': _resolved_hotword_mode,
    'hotword_healthcheck': _hotword_healthcheck,
    'image_provider_healthcheck': _image_provider_healthcheck,
    'build_creator_sync_remote_preview': _build_creator_sync_remote_preview,
    'resolved_creator_sync_mode': _resolved_creator_sync_mode,
    'tracked_creator_sync_targets': _tracked_creator_sync_targets,
    'copywriter_capabilities': _resolve_copywriter_capabilities,
    'copywriter_healthcheck': _copywriter_healthcheck,
    'creator_sync_healthcheck': _creator_sync_healthcheck,
    'log_operation': _log_operation,
    'serialize_data_source_task': _serialize_data_source_task,
    'latest_worker_ping_snapshot': _latest_worker_ping_snapshot,
    'format_datetime': _format_datetime,
    'operation_log_model': OperationLog,
    'is_sqlite_backend': _is_sqlite_backend,
    'os': os,
    'env_flag': _env_flag,
    'coerce_bool': _coerce_bool,
    'topic_model': Topic,
    'registration_model': Registration,
    'submission_model': Submission,
    'serialize_operation_log': _serialize_operation_log,
    'json': json,
})

register_automation_asset_routes(app, {
    'admin_json_guard': _admin_json_guard,
    'safe_int': _safe_int,
    'serialize_asset_generation_task': _serialize_asset_generation_task,
    'serialize_asset_plan_draft': _serialize_asset_plan_draft,
    'serialize_asset_library_item': _serialize_asset_library_item,
    'serialize_automation_schedule': _serialize_automation_schedule,
    'pool_status_label': _pool_status_label,
    'current_actor': _current_actor,
    'load_json_value': _load_json_value,
    'dispatch_asset_generation': _dispatch_asset_generation,
    'dispatch_hotword_sync': _dispatch_hotword_sync,
    'dispatch_creator_account_sync': _dispatch_creator_account_sync,
    'dispatch_automation_schedule': _dispatch_automation_schedule,
    'build_asset_generation_plan_payload': _build_asset_generation_plan_payload,
    'build_batch_asset_plan_drafts': _build_batch_asset_plan_drafts,
    'build_asset_style_recommendation_payload': _build_asset_style_recommendation_payload,
    'log_operation': _log_operation,
    'db': db,
    'datetime': datetime,
    'normalize_quota': _normalize_quota,
    'product_profile_meta': _product_profile_meta,
    'product_profile_options': _product_profile_options,
    'coerce_bool': _coerce_bool,
    'next_schedule_time': _next_schedule_time,
})

register_analytics_routes(app, {
    'build_dashboard_stats': _build_dashboard_stats,
    'build_report_markdown': _build_report_markdown,
    'build_release_manifest_payload': _build_release_manifest_payload,
})


@app.route('/api/trends/import', methods=['POST'])
def import_trends():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    template_key = (data.get('template_key') or 'generic_lines').strip()
    items = _parse_trend_payload(data.get('payload'))
    if not items:
        return jsonify({'success': False, 'message': '没有识别到可导入的热点数据'})

    batch_name = (data.get('batch_name') or datetime.now().strftime('batch_%Y%m%d_%H%M%S')).strip()
    source_platform = (data.get('source_platform') or '小红书').strip()
    source_channel = (data.get('source_channel') or '手动导入').strip()
    normalized_items = _normalize_trend_items(
        items,
        template_key=template_key,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    )
    if not normalized_items:
        return jsonify({'success': False, 'message': '已识别到数据，但未标准化出可入库热点，请检查模板类型'})

    inserted = 0
    skipped = 0
    for item in normalized_items:
        title = (item.get('title') or '').strip()
        link = (item.get('link') or '').strip()
        if not title:
            skipped += 1
            continue

        duplicate = None
        if link:
            duplicate = TrendNote.query.filter_by(link=link).first()
        if not duplicate:
            duplicate = TrendNote.query.filter_by(title=title, keyword=(item.get('keyword') or '').strip()).first()
        if duplicate:
            skipped += 1
            continue

        note = TrendNote(
            source_platform=(item.get('source_platform') or source_platform).strip(),
            source_channel=(item.get('source_channel') or source_channel).strip(),
            source_template_key=template_key,
            import_batch=batch_name,
            keyword=(item.get('keyword') or '').strip(),
            topic_category=(item.get('topic_category') or '').strip(),
            title=title,
            author=(item.get('author') or '').strip(),
            link=link,
            views=_safe_int(item.get('views')),
            likes=_safe_int(item.get('likes')),
            favorites=_safe_int(item.get('favorites')),
            comments=_safe_int(item.get('comments')),
            hot_score=_safe_int(item.get('hot_score')),
            source_rank=_safe_int(item.get('normalized_rank')),
            publish_time=_parse_datetime(item.get('publish_time')),
            summary=(item.get('summary') or '').strip(),
            raw_payload=json.dumps(item, ensure_ascii=False)
        )
        db.session.add(note)
        inserted += 1

    _log_operation('import', 'trend_note', message='导入热点数据', detail={
        'template_key': template_key,
        'source_platform': source_platform,
        'source_channel': source_channel,
        'batch_name': batch_name,
        'inserted': inserted,
        'skipped': skipped,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'导入完成：新增{inserted}条，跳过{skipped}条',
        'inserted': inserted,
        'skipped': skipped
    })


@app.route('/api/trends/import_preview', methods=['POST'])
def preview_trends_import():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    template_key = (data.get('template_key') or 'generic_lines').strip()
    items = _parse_trend_payload(data.get('payload'))
    if not items:
        return jsonify({'success': False, 'message': '没有识别到可预览的数据'})

    batch_name = (data.get('batch_name') or datetime.now().strftime('preview_%Y%m%d_%H%M%S')).strip()
    source_platform = (data.get('source_platform') or '小红书').strip()
    source_channel = (data.get('source_channel') or '手动导入').strip()
    normalized_items = _normalize_trend_items(
        items,
        template_key=template_key,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    )
    preview_items = normalized_items[:10]
    duplicate_count = 0
    for item in normalized_items:
        link = (item.get('link') or '').strip()
        title = (item.get('title') or '').strip()
        keyword = (item.get('keyword') or '').strip()
        duplicate = None
        if link:
            duplicate = TrendNote.query.filter_by(link=link).first()
        if not duplicate and title:
            duplicate = TrendNote.query.filter_by(title=title, keyword=keyword).first()
        if duplicate:
            duplicate_count += 1
    return jsonify({
        'success': True,
        'template': _hotword_source_template_meta(template_key),
        'raw_count': len(items),
        'normalized_count': len(normalized_items),
        'estimated_duplicate_count': duplicate_count,
        'items': preview_items,
    })


@app.route('/api/content_bundle/export', methods=['POST'])
def export_content_bundle():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    activity_id = _safe_int(data.get('activity_id'), 0)
    trend_note_ids = _positive_int_list(data.get('trend_note_ids') or [])
    topic_idea_ids = _positive_int_list(data.get('topic_idea_ids') or [])
    topic_ids = _positive_int_list(data.get('topic_ids') or [])
    include_trends = _coerce_bool(data.get('include_trends', True))
    include_topic_ideas = _coerce_bool(data.get('include_topic_ideas', True))
    include_topics = _coerce_bool(data.get('include_topics', True))

    activity = Activity.query.get(activity_id) if activity_id else None
    trends = []
    topic_ideas = []
    topics = []

    if include_trends and trend_note_ids:
        trends = TrendNote.query.filter(TrendNote.id.in_(trend_note_ids)).order_by(TrendNote.hot_score.desc(), TrendNote.id.desc()).all()

    if include_topic_ideas:
        if topic_idea_ids:
            topic_ideas = TopicIdea.query.filter(TopicIdea.id.in_(topic_idea_ids)).order_by(TopicIdea.created_at.desc(), TopicIdea.id.desc()).all()
        elif activity_id:
            topic_ideas = TopicIdea.query.filter_by(activity_id=activity_id).order_by(TopicIdea.created_at.desc(), TopicIdea.id.desc()).all()

    if include_topics:
        if topic_ids:
            topics = Topic.query.filter(Topic.id.in_(topic_ids)).order_by(Topic.created_at.desc(), Topic.id.desc()).all()
        elif activity_id:
            topics = Topic.query.filter_by(activity_id=activity_id).order_by(Topic.created_at.desc(), Topic.id.desc()).all()

    bundle = _build_content_bundle_payload(
        trends=trends,
        topic_ideas=topic_ideas,
        topics=topics,
        activity=activity,
        note=data.get('note') or '',
    )
    _log_operation('export_bundle', 'content_bundle', message='导出内容发布包', detail={
        'activity_id': activity_id or None,
        'trend_count': len(trends),
        'topic_idea_count': len(topic_ideas),
        'topic_count': len(topics),
        'source_env': bundle['source_env'],
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已生成内容发布包：热点 {len(trends)} 条，候选话题 {len(topic_ideas)} 条，正式话题 {len(topics)} 条',
        'bundle': bundle,
    })


@app.route('/api/content_bundle/import_preview', methods=['POST'])
def preview_content_bundle_import():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = (request.get_json(silent=True) or request.form or {})
    target_activity_id = _safe_int(data.get('target_activity_id'), 0)
    import_topics = _coerce_bool(data.get('import_topics', False))
    try:
        bundle = _extract_content_bundle_from_request()
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})

    preview = _preview_content_bundle_import(
        bundle,
        target_activity_id=target_activity_id,
        import_topics=import_topics,
    )
    return jsonify({
        'success': True,
        'message': '已生成发布包导入预览',
        'bundle_meta': {
            'source_env': bundle.get('source_env') or '',
            'generated_at': bundle.get('generated_at') or '',
            'generated_by': bundle.get('generated_by') or '',
            'schema_version': bundle.get('schema_version') or '',
        },
        'preview': preview,
    })


@app.route('/api/content_bundle/import', methods=['POST'])
def import_content_bundle():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = (request.get_json(silent=True) or request.form or {})
    target_activity_id = _safe_int(data.get('target_activity_id'), 0)
    import_trends = _coerce_bool(data.get('import_trends', True))
    import_topic_ideas = _coerce_bool(data.get('import_topic_ideas', True))
    import_topics = _coerce_bool(data.get('import_topics', False))
    preserve_review_status = _coerce_bool(data.get('preserve_review_status', True))

    try:
        bundle = _extract_content_bundle_from_request()
        result = _import_content_bundle(
            bundle,
            target_activity_id=target_activity_id,
            import_trends=import_trends,
            import_topic_ideas=import_topic_ideas,
            import_topics=import_topics,
            preserve_review_status=preserve_review_status,
        )
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})

    _log_operation('import_bundle', 'content_bundle', message='导入内容发布包', detail={
        'source_env': bundle.get('source_env') or '',
        'target_activity_id': target_activity_id or None,
        'import_trends': import_trends,
        'import_topic_ideas': import_topic_ideas,
        'import_topics': import_topics,
        'created_counts': {key: len(value) for key, value in result['created'].items()},
        'updated_counts': {key: len(value) for key, value in result['updated'].items()},
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': (
            f"发布包已导入：新增热点 {len(result['created']['trends'])}、候选话题 {len(result['created']['topic_ideas'])}、正式话题 {len(result['created']['topics'])}；"
            f"更新热点 {len(result['updated']['trends'])}、候选话题 {len(result['updated']['topic_ideas'])}、正式话题 {len(result['updated']['topics'])}"
        ),
        'result': result,
    })


@app.route('/api/trends/parse_remote_preview', methods=['POST'])
def preview_remote_trends_parse():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    template_key = (data.get('template_key') or 'generic_json').strip()
    source_platform = (data.get('source_platform') or '小红书').strip()
    source_channel = (data.get('source_channel') or '接口响应沙盒').strip()
    batch_name = (data.get('batch_name') or datetime.now().strftime('remote_preview_%Y%m%d_%H%M%S')).strip()
    result_path = (data.get('result_path') or '').strip()
    raw_response = data.get('response_payload') or ''

    try:
        items = _extract_hotword_result_payload(raw_response, result_path=result_path)
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})

    if not items:
        return jsonify({'success': False, 'message': '未从响应中解析出可预览的数据，请检查结果路径或返回结构'})

    normalized_items = _normalize_trend_items(
        items,
        template_key=template_key,
        source_platform=source_platform,
        source_channel=source_channel,
        batch_name=batch_name,
    )
    if not normalized_items:
        return jsonify({'success': False, 'message': '已解析出响应数据，但未标准化出可用热点，请调整模板或结果路径'})

    return jsonify({
        'success': True,
        'template': _hotword_source_template_meta(template_key),
        'message': f'解析成功，共识别 {len(items)} 条原始记录，标准化 {len(normalized_items)} 条',
        'raw_count': len(items),
        'normalized_count': len(normalized_items),
        'items': normalized_items[:10],
        'result_path': result_path,
    })


@app.route('/api/trends')
def list_trends():
    guard = _admin_json_guard()
    if guard:
        return guard

    keyword = (request.args.get('keyword') or '').strip()
    source_platform = (request.args.get('source_platform') or '').strip()
    source_template_key = (request.args.get('source_template_key') or '').strip()
    pool_status = (request.args.get('pool_status') or '').strip()
    query = TrendNote.query
    if keyword:
        query = query.filter(or_(
            TrendNote.keyword.contains(keyword),
            TrendNote.title.contains(keyword),
            TrendNote.summary.contains(keyword),
        ))
    if source_platform:
        query = query.filter_by(source_platform=source_platform)
    if source_template_key:
        query = query.filter_by(source_template_key=source_template_key)
    if pool_status:
        query = query.filter_by(pool_status=pool_status)

    notes = query.order_by(TrendNote.hot_score.desc(), TrendNote.created_at.desc()).limit(120).all()
    return jsonify({
        'success': True,
        'items': [_serialize_trend_note(note) for note in notes]
    })


@app.route('/api/trends/<int:note_id>/pool_status', methods=['POST'])
def update_trend_pool_status(note_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    pool_status = (data.get('pool_status') or '').strip()
    if pool_status not in {'reserve', 'candidate', 'archived'}:
        return jsonify({'success': False, 'message': '不支持的热点池状态'})

    note = TrendNote.query.get_or_404(note_id)
    note.pool_status = pool_status
    _log_operation('move_pool', 'trend_note', target_id=note.id, message='更新热点池状态', detail={
        'title': note.title,
        'pool_status': pool_status,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'热点已移动到{_pool_status_label(pool_status)}',
        'item': _serialize_trend_note(note)
    })


@app.route('/api/trends/<int:note_id>/to_corpus', methods=['POST'])
def convert_trend_to_corpus(note_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    note = TrendNote.query.get_or_404(note_id)
    data = request.json or {}
    category = (data.get('category') or '爆款拆解').strip() or '爆款拆解'
    result = _upsert_trend_note_corpus_entries([note], category=category)
    created = result['created']
    updated = result['updated']
    _log_operation('trend_to_corpus', 'trend_note', target_id=note.id, message='将热点爆款笔记转为模板语料', detail={
        'title': note.title,
        'reference_url': note.link or '',
        'created_count': len(created),
        'updated_count': len(updated),
        'category': category,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已把《{note.title}》转成模板语料',
        'trend': _serialize_trend_note(note),
        'items': [_serialize_corpus_entry(item) for item in (created + updated)],
    })


@app.route('/api/trends/<int:note_id>/route_target', methods=['POST'])
def route_trend_to_target(note_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    note = TrendNote.query.get_or_404(note_id)
    data = request.json or {}
    target = (data.get('target') or '').strip()
    activity_id = _safe_int(data.get('activity_id'), 0)
    try:
        routed = _route_trend_note_to_target(note, target=target, activity_id=activity_id, actor=_current_actor())
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})
    db.session.commit()
    return jsonify({
        'success': True,
        'message': '热点内容已完成分流',
        **routed,
    })


@app.route('/api/trends/route_target_batch', methods=['POST'])
def route_trends_to_target_batch():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    requested_target = (data.get('target') or 'recommended').strip() or 'recommended'
    activity_id = _safe_int(data.get('activity_id'), 0)
    limit = min(max(_safe_int(data.get('limit'), 20), 1), 50)

    note_ids = []
    for item in (data.get('note_ids') or []):
        value = _safe_int(item, 0)
        if value > 0:
            note_ids.append(value)
    note_ids = list(dict.fromkeys(note_ids))

    query = TrendNote.query
    if note_ids:
        query = query.filter(TrendNote.id.in_(note_ids))
    else:
        keyword = (data.get('keyword') or '').strip()
        source_platform = (data.get('source_platform') or '').strip()
        source_template_key = (data.get('source_template_key') or '').strip()
        pool_status = (data.get('pool_status') or '').strip()
        if keyword:
            query = query.filter(or_(
                TrendNote.keyword.contains(keyword),
                TrendNote.title.contains(keyword),
                TrendNote.summary.contains(keyword),
            ))
        if source_platform:
            query = query.filter_by(source_platform=source_platform)
        if source_template_key:
            query = query.filter_by(source_template_key=source_template_key)
        if pool_status:
            query = query.filter_by(pool_status=pool_status)
        query = query.order_by(TrendNote.hot_score.desc(), TrendNote.created_at.desc()).limit(limit)

    notes = query.all()
    if not notes:
        return jsonify({'success': False, 'message': '当前筛选下没有可分流的热点内容'})

    routed_items = []
    errors = []
    target_counts = defaultdict(int)
    actor = _current_actor()
    for note in notes:
        resolved_target = _resolve_trend_route_target(note, requested_target)
        try:
            routed = _route_trend_note_to_target(note, target=resolved_target, activity_id=activity_id, actor=actor)
            target_counts[resolved_target] += 1
            routed_items.append({
                'note_id': note.id,
                'title': note.title or note.keyword or '热点内容',
                'target': resolved_target,
                'target_label': _trend_route_target_label(resolved_target),
            })
        except ValueError as exc:
            errors.append({
                'note_id': note.id,
                'title': note.title or note.keyword or '热点内容',
                'message': str(exc),
            })

    if routed_items:
        _log_operation('route_trend_batch', 'trend_note', message='批量分流热点内容', detail={
            'selected_count': len(notes),
            'requested_target': requested_target,
            'activity_id': activity_id,
            'target_counts': dict(target_counts),
            'success_count': len(routed_items),
            'error_count': len(errors),
            'actor': actor,
        })

    db.session.commit()

    if not routed_items:
        message = errors[0]['message'] if errors else '批量分流失败'
        return jsonify({
            'success': False,
            'message': message,
            'errors': errors[:20],
        })

    count_summary = '，'.join(
        f'{_trend_route_target_label(target)} {count} 条'
        for target, count in target_counts.items()
    ) or f'已处理 {len(routed_items)} 条'
    if errors:
        count_summary = f'{count_summary}，跳过 {len(errors)} 条'

    return jsonify({
        'success': True,
        'message': f'批量分流完成：{count_summary}',
        'items': routed_items[:20],
        'target_counts': dict(target_counts),
        'errors': errors[:20],
    })


@app.route('/api/trends/to_corpus_batch', methods=['POST'])
def convert_trends_to_corpus_batch():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    category = (data.get('category') or '爆款拆解').strip() or '爆款拆解'
    note_ids = []
    for item in (data.get('note_ids') or []):
        value = _safe_int(item, 0)
        if value > 0:
            note_ids.append(value)
    note_ids = list(dict.fromkeys(note_ids))

    query = TrendNote.query
    if note_ids:
        query = query.filter(TrendNote.id.in_(note_ids))
    else:
        keyword = (data.get('keyword') or '').strip()
        source_platform = (data.get('source_platform') or '').strip()
        source_template_key = (data.get('source_template_key') or '').strip()
        pool_status = (data.get('pool_status') or '').strip()
        limit = min(max(_safe_int(data.get('limit'), 20), 1), 50)
        if keyword:
            query = query.filter(or_(
                TrendNote.keyword.contains(keyword),
                TrendNote.title.contains(keyword),
                TrendNote.summary.contains(keyword),
            ))
        if source_platform:
            query = query.filter_by(source_platform=source_platform)
        if source_template_key:
            query = query.filter_by(source_template_key=source_template_key)
        if pool_status:
            query = query.filter_by(pool_status=pool_status)
        query = query.order_by(TrendNote.hot_score.desc(), TrendNote.created_at.desc()).limit(limit)

    notes = query.all()
    if not notes:
        return jsonify({'success': False, 'message': '当前没有可转换的热点笔记'})

    result = _upsert_trend_note_corpus_entries(notes, category=category)
    created = result['created']
    updated = result['updated']
    _log_operation('trend_to_corpus_batch', 'trend_note', message='批量将热点爆款笔记转为模板语料', detail={
        'selected_count': len(notes),
        'created_count': len(created),
        'updated_count': len(updated),
        'category': category,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已批量转换 {len(notes)} 条热点笔记：新增 {len(created)} 条，更新 {len(updated)} 条',
        'count': len(notes),
        'items': [_serialize_corpus_entry(item) for item in (created + updated)[:30]],
    })


@app.route('/api/trends/promote_reserve', methods=['POST'])
def promote_trends_reserve():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    source_platform = (data.get('source_platform') or '').strip()
    query = TrendNote.query.filter_by(pool_status='reserve')
    if source_platform:
        query = query.filter_by(source_platform=source_platform)

    notes = query.all()
    for note in notes:
        note.pool_status = 'candidate'
    _log_operation('promote_reserve', 'trend_note', message='批量推送热点到候选池', detail={
        'source_platform': source_platform,
        'count': len(notes),
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已将 {len(notes)} 条储备热点推入候选池',
        'count': len(notes),
    })


@app.route('/api/topic_ideas', methods=['GET'])
def list_topic_ideas():
    guard = _admin_json_guard()
    if guard:
        return guard

    activity_id = request.args.get('activity_id', type=int)
    status = (request.args.get('status') or '').strip()
    query = TopicIdea.query
    if activity_id:
        query = query.filter_by(activity_id=activity_id)
    if status:
        query = query.filter_by(status=status)

    ideas = query.order_by(TopicIdea.created_at.desc()).limit(200).all()
    return jsonify({
        'success': True,
        'items': [_serialize_topic_idea(idea) for idea in ideas]
    })


@app.route('/api/topics/import_preview', methods=['POST'])
def preview_topics_import():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = (request.get_json(silent=True) or request.form or {})
    if request.files:
        file_storage = request.files.get('file')
        rows = _parse_topic_import_file(file_storage.filename or '', file_storage.read()) if file_storage and file_storage.filename else []
    else:
        rows = _parse_topic_import_payload(data.get('raw_payload') or '')
    if not rows:
        return jsonify({'success': False, 'message': '没有识别到可导入的话题数据'})

    activity_id = _safe_int(data.get('activity_id'), 0)
    target_type = (data.get('target_type') or 'topic_idea').strip() or 'topic_idea'
    preview = _preview_topic_import_rows(rows, activity_id=activity_id, target_type=target_type)
    return jsonify({
        'success': True,
        'message': f'已识别 {preview["count"]} 条话题数据，疑似重复 {preview["duplicate_count"]} 条',
        'preview': preview,
    })


@app.route('/api/topics/import', methods=['POST'])
def import_topics():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = (request.get_json(silent=True) or request.form or {})
    if request.files:
        file_storage = request.files.get('file')
        rows = _parse_topic_import_file(file_storage.filename or '', file_storage.read()) if file_storage and file_storage.filename else []
    else:
        rows = _parse_topic_import_payload(data.get('raw_payload') or '')
    if not rows:
        return jsonify({'success': False, 'message': '没有识别到可导入的话题数据'})

    activity_id = _safe_int(data.get('activity_id'), 0)
    target_type = (data.get('target_type') or 'topic_idea').strip() or 'topic_idea'
    if target_type == 'topic' and not activity_id:
        return jsonify({'success': False, 'message': '导入正式话题时必须选择活动期数'})

    result = _import_topic_rows(rows, activity_id=activity_id, target_type=target_type)
    _log_operation('import', 'topic_batch', message='批量导入话题', detail={
        'activity_id': activity_id or None,
        'target_type': target_type,
        'count': len(rows),
        'created_count': len(result['created']),
        'updated_count': len(result['updated']),
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'批量导入完成：新增 {len(result["created"])} 条，更新 {len(result["updated"])} 条',
        'result': result,
    })


@app.route('/api/topic_ideas/generate', methods=['POST'])
def generate_topic_ideas():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    count = min(max(_safe_int(data.get('count'), 80), 1), 120)
    activity_id = data.get('activity_id')
    quota = _normalize_quota(data.get('quota'))

    ideas = _generate_topic_ideas(count=count, activity_id=activity_id, quota=quota)
    for idea in ideas:
        db.session.add(idea)

    if ideas:
        matching_ids = {entry.id for entry in _matching_corpus_snippets(','.join(LIVER_KEYWORD_SEEDS[:5]), limit=5)}
        matched_entries = CorpusEntry.query.filter(CorpusEntry.id.in_(matching_ids)).all() if matching_ids else []
        for entry in matched_entries:
            entry.usage_count = (entry.usage_count or 0) + 1

    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已生成{len(ideas)}个候选话题，默认名额为{quota}，并进入待审核状态',
        'count': len(ideas)
    })


@app.route('/api/topic_ideas/<int:idea_id>/review', methods=['POST'])
def review_topic_idea(idea_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    action = (data.get('action') or '').strip()
    note = (data.get('note') or '').strip()
    idea = TopicIdea.query.get_or_404(idea_id)

    if idea.status == 'published' and action != 'archive':
        return jsonify({'success': False, 'message': '已发布的话题请先归档后再调整'})

    if action == 'approve':
        idea.status = 'approved'
        message = '候选话题已审核通过'
    elif action == 'reject':
        idea.status = 'rejected'
        message = '候选话题已驳回'
    elif action == 'reset':
        idea.status = 'pending_review'
        message = '候选话题已退回待审核'
    elif action == 'archive':
        idea.status = 'archived'
        message = '候选话题已归档'
    else:
        return jsonify({'success': False, 'message': '不支持的审核动作'})

    idea.review_note = note
    idea.reviewed_at = datetime.now()
    _log_operation('review', 'topic_idea', target_id=idea.id, message='审核候选话题', detail={
        'action': action,
        'topic_title': idea.topic_title,
        'status': idea.status,
        'note': note,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': message,
        'item': _serialize_topic_idea(idea)
    })


@app.route('/api/topic_ideas/review_batch', methods=['POST'])
def review_topic_ideas_batch():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    action = (data.get('action') or '').strip()
    note = (data.get('note') or '').strip()
    raw_ids = data.get('idea_ids') or []
    idea_ids = []
    for item in raw_ids:
        value = _safe_int(item, 0)
        if value > 0:
            idea_ids.append(value)
    idea_ids = list(dict.fromkeys(idea_ids))
    if not idea_ids:
        return jsonify({'success': False, 'message': '请先选择要批量处理的话题'})
    if action not in {'approve', 'reject', 'reset', 'archive'}:
        return jsonify({'success': False, 'message': '不支持的批量审核动作'})

    ideas = TopicIdea.query.filter(TopicIdea.id.in_(idea_ids)).all()
    if not ideas:
        return jsonify({'success': False, 'message': '未找到可处理的话题'})

    status_map = {
        'approve': ('approved', '已审核通过'),
        'reject': ('rejected', '已驳回'),
        'reset': ('pending_review', '已退回待审核'),
        'archive': ('archived', '已归档'),
    }
    target_status, message = status_map[action]
    updated = []
    skipped = []
    for idea in ideas:
        if idea.status == 'published' and action != 'archive':
            skipped.append({'id': idea.id, 'title': idea.topic_title, 'reason': '已发布的话题仅支持归档'})
            continue
        idea.status = target_status
        idea.review_note = note
        idea.reviewed_at = datetime.now()
        updated.append({'id': idea.id, 'title': idea.topic_title, 'status': idea.status})

    _log_operation('review_batch', 'topic_idea', message='批量审核候选话题', detail={
        'action': action,
        'note': note,
        'selected_count': len(idea_ids),
        'updated_count': len(updated),
        'skipped_count': len(skipped),
        'updated': updated,
        'skipped': skipped,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'批量处理完成：{message} {len(updated)} 条，跳过 {len(skipped)} 条',
        'updated': updated,
        'skipped': skipped,
    })


@app.route('/api/topic_ideas/<int:idea_id>/publish', methods=['POST'])
def publish_topic_idea(idea_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    idea = TopicIdea.query.get_or_404(idea_id)
    public_payload = _topic_idea_public_strategy(idea)
    if idea.status == 'published' and idea.published_topic_id:
        topic = Topic.query.get(idea.published_topic_id)
        if topic:
            topic.topic_name = public_payload['topic_name']
            topic.keywords = public_payload['keywords']
            topic.direction = public_payload['direction']
            topic.reference_content = public_payload['reference_content']
            topic.reference_link = _compact_reference_links_for_topic(idea.source_links)
            topic.writing_example = public_payload['writing_example']
            topic.group_num = public_payload['group_num']
            topic.quota = _normalize_quota(data.get('quota'), default=idea.quota or topic.quota or _default_topic_quota())
            db.session.commit()
            return jsonify({'success': True, 'message': '已同步优化后的活动话题'})
        return jsonify({'success': False, 'message': '该候选话题已发布过，但未找到正式话题'})
    if idea.status not in {'approved', 'published'}:
        return jsonify({'success': False, 'message': '请先审核通过再发布到活动'})

    activity_id = _safe_int(data.get('activity_id') or idea.activity_id)
    if not activity_id:
        current_activity = Activity.query.filter_by(status='published').order_by(Activity.created_at.desc()).first()
        activity_id = current_activity.id if current_activity else None
    if not activity_id:
        return jsonify({'success': False, 'message': '请先选择目标活动期数'})
    quota = _normalize_quota(data.get('quota'), default=idea.quota or _default_topic_quota())

    topic = Topic(
        activity_id=activity_id,
        topic_name=public_payload['topic_name'],
        keywords=public_payload['keywords'],
        direction=public_payload['direction'],
        reference_content=public_payload['reference_content'],
        reference_link=_compact_reference_links_for_topic(idea.source_links),
        writing_example=public_payload['writing_example'],
        quota=quota,
        group_num=(data.get('group_num') or public_payload['group_num'] or '自动化选题').strip(),
        pool_status='formal',
        source_type='topic_idea',
        source_ref_id=idea.id,
        published_at=datetime.now()
    )
    db.session.add(topic)
    db.session.flush()
    idea.quota = quota
    idea.status = 'published'
    idea.published_topic_id = topic.id
    idea.published_at = datetime.now()
    _log_operation('publish', 'topic_idea', target_id=idea.id, message='候选话题发布到正式活动', detail={
        'topic_title': idea.topic_title,
        'activity_id': activity_id,
        'published_topic_id': topic.id,
        'quota': quota,
    })
    db.session.commit()
    return jsonify({'success': True, 'message': '已发布到活动题库'})


@app.route('/api/topic_ideas/publish_batch', methods=['POST'])
def publish_topic_ideas_batch():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    raw_ids = data.get('idea_ids') or []
    idea_ids = []
    for item in raw_ids:
        value = _safe_int(item, 0)
        if value > 0:
            idea_ids.append(value)
    idea_ids = list(dict.fromkeys(idea_ids))
    if not idea_ids:
        return jsonify({'success': False, 'message': '请先选择要批量发布的话题'})

    activity_id = _safe_int(data.get('activity_id'))
    if not activity_id:
        return jsonify({'success': False, 'message': '请先选择目标活动期数'})

    ideas = TopicIdea.query.filter(TopicIdea.id.in_(idea_ids)).all()
    if not ideas:
        return jsonify({'success': False, 'message': '未找到可发布的话题'})

    published = []
    skipped = []
    for idea in ideas:
        if idea.status != 'approved':
            skipped.append({'id': idea.id, 'title': idea.topic_title, 'reason': '未审核通过'})
            continue
        if idea.published_topic_id:
            skipped.append({'id': idea.id, 'title': idea.topic_title, 'reason': '已发布过'})
            continue

        public_payload = _topic_idea_public_strategy(idea)
        topic = Topic(
            activity_id=activity_id,
            topic_name=public_payload['topic_name'],
            keywords=public_payload['keywords'],
            direction=public_payload['direction'],
            reference_content=public_payload['reference_content'],
            reference_link=_compact_reference_links_for_topic(idea.source_links),
            writing_example=public_payload['writing_example'],
            quota=_normalize_quota(idea.quota, default=_default_topic_quota()),
            group_num=public_payload['group_num'] or '自动化选题',
            pool_status='formal',
            source_type='topic_idea',
            source_ref_id=idea.id,
            published_at=datetime.now()
        )
        db.session.add(topic)
        db.session.flush()
        idea.status = 'published'
        idea.published_topic_id = topic.id
        idea.published_at = datetime.now()
        published.append({'id': idea.id, 'title': idea.topic_title, 'topic_id': topic.id})

    _log_operation('publish_batch', 'topic_idea', message='批量发布候选话题', detail={
        'activity_id': activity_id,
        'selected_count': len(idea_ids),
        'published_count': len(published),
        'skipped_count': len(skipped),
        'published': published,
        'skipped': skipped,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'批量发布完成：成功 {len(published)} 条，跳过 {len(skipped)} 条',
        'published': published,
        'skipped': skipped,
    })


@app.route('/api/topic_ideas/<int:idea_id>/quota', methods=['POST'])
def update_topic_idea_quota(idea_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    idea = TopicIdea.query.get_or_404(idea_id)
    idea.quota = _normalize_quota(data.get('quota'), default=idea.quota or _default_topic_quota())
    _log_operation('update_quota', 'topic_idea', target_id=idea.id, message='调整候选话题名额', detail={
        'topic_title': idea.topic_title,
        'quota': idea.quota,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'候选话题名额已调整为 {idea.quota}',
        'item': _serialize_topic_idea(idea)
    })


@app.route('/api/creator_accounts', methods=['GET', 'POST'])
def creator_accounts():
    guard = _admin_json_guard()
    if guard:
        return guard

    if request.method == 'POST':
        data = request.json or {}
        account_id = data.get('id')
        if account_id:
            account = CreatorAccount.query.get_or_404(account_id)
        else:
            account = CreatorAccount()
            db.session.add(account)

        platform = (data.get('platform') or 'xhs').strip()
        account_handle = (data.get('account_handle') or '').strip()
        display_name = (data.get('display_name') or '').strip()
        if not account_handle and not display_name:
            return jsonify({'success': False, 'message': '账号标识或昵称至少填写一个'})

        account.platform = platform
        account.owner_name = (data.get('owner_name') or '').strip()
        account.owner_phone = (data.get('owner_phone') or '').strip()
        account.account_handle = account_handle or display_name
        account.display_name = display_name or account.account_handle
        account.profile_url = normalize_tracking_url(data.get('profile_url') or '')
        account.follower_count = _safe_int(data.get('follower_count'), account.follower_count or 0)
        account.source_channel = (data.get('source_channel') or account.source_channel or 'manual').strip()
        account.status = (data.get('status') or account.status or 'active').strip()
        account.notes = (data.get('notes') or '').strip()
        account.last_synced_at = datetime.now()
        db.session.flush()
        sync_tracking_for_creator_account(account, refresh_snapshot=False)
        _log_operation('save', 'creator_account', target_id=account.id, message='保存账号信息', detail={
            'platform': account.platform,
            'owner_name': account.owner_name,
            'account_handle': account.account_handle,
        })
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '账号已保存',
            'item': _serialize_creator_account(account)
        })

    query = _build_creator_account_query(request.args)
    accounts = query.order_by(CreatorAccount.updated_at.desc(), CreatorAccount.created_at.desc()).limit(200).all()
    return jsonify({
        'success': True,
        'items': [_serialize_creator_account(account) for account in accounts]
    })


@app.route('/api/creator_accounts/import_preview', methods=['POST'])
def creator_accounts_import_preview():
    guard = _admin_json_guard()
    if guard:
        return guard

    try:
        if request.files.get('file'):
            file_storage = request.files['file']
            bundle = parse_creator_import_file(file_storage.filename or '', file_storage.read())
        else:
            data = request.json or {}
            bundle = parse_creator_import_bundle(data.get('raw_payload') or '')
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})

    preview = preview_creator_import_bundle(bundle)
    return jsonify({
        'success': True,
        'message': '导入预览已生成',
        **preview,
    })


@app.route('/api/creator_accounts/import', methods=['POST'])
def creator_accounts_import():
    guard = _admin_json_guard()
    if guard:
        return guard

    try:
        if request.files.get('file'):
            file_storage = request.files['file']
            bundle = parse_creator_import_file(file_storage.filename or '', file_storage.read())
        else:
            data = request.json or {}
            bundle = parse_creator_import_bundle(data.get('raw_payload') or '')
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})

    result = import_creator_bundle(bundle, log_operation=_log_operation)
    return jsonify({
        'success': True,
        'message': '账号看板数据已导入',
        **result,
    })


@app.route('/api/creator_accounts/analytics')
def creator_accounts_analytics():
    guard = _admin_json_guard()
    if guard:
        return guard

    accounts = _build_creator_account_query(request.args).all()
    account_ids = [account.id for account in accounts]
    date_from = _parse_date(request.args.get('date_from'))
    date_to = _parse_date(request.args.get('date_to'))

    posts = []
    snapshots = []
    if account_ids:
        post_query = CreatorPost.query.filter(CreatorPost.creator_account_id.in_(account_ids))
        post_query, _, _ = _apply_creator_post_range(post_query, request.args)
        posts = post_query.order_by(CreatorPost.publish_time.asc(), CreatorPost.created_at.asc()).all()

        snapshot_query = CreatorAccountSnapshot.query.filter(CreatorAccountSnapshot.creator_account_id.in_(account_ids))
        if date_from:
            snapshot_query = snapshot_query.filter(CreatorAccountSnapshot.snapshot_date >= date_from)
        if date_to:
            snapshot_query = snapshot_query.filter(CreatorAccountSnapshot.snapshot_date <= date_to)
        snapshots = snapshot_query.order_by(CreatorAccountSnapshot.snapshot_date.asc(), CreatorAccountSnapshot.created_at.asc()).all()

    analytics = _build_creator_analytics_payload(posts, snapshots, date_from=date_from, date_to=date_to)
    month_from, month_to = _current_month_date_range()
    current_month_posts = [
        post for post in posts
        if post.publish_time and month_from <= post.publish_time.date() <= month_to
    ]
    current_month_account_ids = {post.creator_account_id for post in current_month_posts if post.creator_account_id}

    platform_map = defaultdict(lambda: {
        'account_count': 0,
        'post_count': 0,
        'viral_posts': 0,
        'total_views': 0,
        'total_interactions': 0,
        'follower_count': 0,
    })
    account_rows = []
    posts_by_account = defaultdict(list)
    for post in posts:
        posts_by_account[post.creator_account_id].append(post)

    for account in accounts:
        platform_key = account.platform or 'unknown'
        account_posts = posts_by_account.get(account.id, [])
        post_count = len(account_posts)
        viral_posts = len([post for post in account_posts if post.is_viral])
        total_views = sum(post.views or 0 for post in account_posts)
        total_interactions = sum((post.likes or 0) + (post.favorites or 0) + (post.comments or 0) for post in account_posts)
        platform_row = platform_map[platform_key]
        platform_row['account_count'] += 1
        platform_row['post_count'] += post_count
        platform_row['viral_posts'] += viral_posts
        platform_row['total_views'] += total_views
        platform_row['total_interactions'] += total_interactions
        platform_row['follower_count'] += account.follower_count or 0

        account_rows.append({
            'id': account.id,
            'owner_name': account.owner_name or '',
            'display_name': account.display_name or account.account_handle or '',
            'platform': platform_key,
            'post_count': post_count,
            'viral_posts': viral_posts,
            'total_views': total_views,
            'total_interactions': total_interactions,
        })

    platform_rows = [{'platform': platform, **stats} for platform, stats in platform_map.items()]
    platform_rows.sort(key=lambda item: (item['total_interactions'], item['total_views'], item['account_count']), reverse=True)
    account_rows.sort(key=lambda item: (item['total_interactions'], item['total_views'], item['post_count']), reverse=True)

    analytics['overview']['account_count'] = len(accounts)
    analytics['overview']['current_month_label'] = month_from.strftime('%Y-%m')
    analytics['overview']['current_month_account_count'] = len(current_month_account_ids)
    analytics['overview']['current_month_post_count'] = len(current_month_posts)
    analytics['overview']['current_month_total_views'] = sum(post.views or 0 for post in current_month_posts)
    analytics['overview']['current_month_total_interactions'] = sum((post.likes or 0) + (post.favorites or 0) + (post.comments or 0) for post in current_month_posts)
    analytics['platforms'] = platform_rows
    analytics['top_accounts'] = account_rows[:10]
    analytics['current_month_top_posts'] = [
        _serialize_creator_post(post) for post in sorted(
            current_month_posts,
            key=lambda item: ((item.views or 0), ((item.likes or 0) + (item.favorites or 0) + (item.comments or 0)), (item.follower_delta or 0)),
            reverse=True,
        )[:10]
    ]
    return jsonify({
        'success': True,
        **analytics,
    })


@app.route('/api/creator_accounts/export')
def export_creator_accounts():
    guard = _admin_json_guard()
    if guard:
        return guard

    accounts = _build_creator_account_query(request.args).order_by(
        CreatorAccount.updated_at.desc(),
        CreatorAccount.created_at.desc()
    ).all()
    rows = ['平台,姓名,手机号,账号标识,显示昵称,当前粉丝,发文数,爆款数,总阅读,总互动,最佳笔记,最佳笔记阅读,上次同步']
    for account in accounts:
        item = _serialize_creator_account(account)
        best_post = item.get('best_post') or {}
        row = [
            item.get('platform', ''),
            (item.get('owner_name') or '').replace(',', ' '),
            (item.get('owner_phone') or '').replace(',', ' '),
            (item.get('account_handle') or '').replace(',', ' '),
            (item.get('display_name') or '').replace(',', ' '),
            str(item.get('follower_count') or 0),
            str(item.get('post_count') or 0),
            str(item.get('viral_post_count') or 0),
            str(item.get('total_views') or 0),
            str(item.get('total_interactions') or 0),
            (best_post.get('title') or '').replace(',', ' '),
            str(best_post.get('views') or 0),
            (item.get('last_synced_at') or '').replace(',', ' '),
        ]
        rows.append(','.join(row))

    _log_operation('export', 'creator_account_list', message='导出账号列表', detail={
        'platform': request.args.get('platform', ''),
        'phone': request.args.get('phone', ''),
        'keyword': request.args.get('keyword', ''),
        'viral_only': request.args.get('viral_only', ''),
        'count': len(accounts),
    })
    db.session.commit()
    content = '\n'.join(rows)
    return content, 200, {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': 'attachment; filename=creator_accounts.csv'
    }


@app.route('/api/creator_accounts/<int:account_id>/export')
def export_creator_account_detail(account_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    account = CreatorAccount.query.get_or_404(account_id)
    post_query = CreatorPost.query.filter_by(creator_account_id=account.id)
    post_query, date_from, date_to = _apply_creator_post_range(post_query, request.args)
    posts = post_query.order_by(CreatorPost.publish_time.desc(), CreatorPost.created_at.desc()).all()

    rows = [
        '账号维度',
        '平台,姓名,手机号,账号标识,显示昵称,当前粉丝,发文数,爆款数,总阅读,总互动,统计起始,统计截止',
        ','.join([
            account.platform or '',
            (account.owner_name or '').replace(',', ' '),
            (account.owner_phone or '').replace(',', ' '),
            (account.account_handle or '').replace(',', ' '),
            (account.display_name or '').replace(',', ' '),
            str(account.follower_count or 0),
            str(len(posts)),
            str(len([post for post in posts if post.is_viral])),
            str(sum(post.views or 0 for post in posts)),
            str(sum((post.likes or 0) + (post.favorites or 0) + (post.comments or 0) for post in posts)),
            date_from.isoformat() if date_from else '',
            date_to.isoformat() if date_to else '',
        ]),
        '',
        '笔记维度',
        '标题,话题,发布时间,阅读,曝光,点赞,收藏,评论,涨粉,是否爆款,链接',
    ]
    for post in posts:
        rows.append(','.join([
            (post.title or '').replace(',', ' '),
            (post.topic_title or '').replace(',', ' '),
            post.publish_time.strftime('%Y-%m-%d %H:%M:%S') if post.publish_time else '',
            str(post.views or 0),
            str(post.exposures or 0),
            str(post.likes or 0),
            str(post.favorites or 0),
            str(post.comments or 0),
            str(post.follower_delta or 0),
            '是' if post.is_viral else '否',
            (post.post_url or '').replace(',', ' '),
        ]))

    _log_operation('export', 'creator_account', target_id=account.id, message='导出单账号数据', detail={
        'date_from': date_from.isoformat() if date_from else '',
        'date_to': date_to.isoformat() if date_to else '',
        'post_count': len(posts),
    })
    db.session.commit()
    content = '\n'.join(rows)
    return content, 200, {
        'Content-Type': 'text/csv; charset=utf-8',
        'Content-Disposition': f'attachment; filename=creator_account_{account.id}.csv'
    }


@app.route('/api/creator_accounts/<int:account_id>/posts', methods=['GET', 'POST'])
def creator_account_posts(account_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    account = CreatorAccount.query.get_or_404(account_id)
    if request.method == 'POST':
        data = request.json or {}
        post_id = data.get('id')
        if post_id:
            post = CreatorPost.query.filter_by(id=post_id, creator_account_id=account.id).first()
            if not post:
                return jsonify({'success': False, 'message': '笔记不存在'})
        else:
            post = CreatorPost(creator_account_id=account.id)
            db.session.add(post)

        title = (data.get('title') or '').strip()
        if not title:
            return jsonify({'success': False, 'message': '笔记标题不能为空'})

        post.platform_post_id = (data.get('platform_post_id') or '').strip()
        post.registration_id = _safe_int(data.get('registration_id'), post.registration_id)
        post.topic_id = _safe_int(data.get('topic_id'), post.topic_id)
        post.submission_id = _safe_int(data.get('submission_id'), post.submission_id)
        post.title = title
        raw_post_url = (data.get('post_url') or '').strip()
        if account.platform == 'xhs':
            post.post_url = canonicalize_xhs_post_url(raw_post_url)
            if post.post_url and not post.platform_post_id:
                post.platform_post_id = post.post_url.rstrip('/').split('/')[-1]
        else:
            post.post_url = normalize_tracking_url(raw_post_url)
        post.publish_time = _parse_datetime(data.get('publish_time'))
        post.topic_title = (data.get('topic_title') or '').strip()
        post.views = _safe_int(data.get('views'))
        post.exposures = _safe_int(data.get('exposures'))
        post.likes = _safe_int(data.get('likes'))
        post.favorites = _safe_int(data.get('favorites'))
        post.comments = _safe_int(data.get('comments'))
        post.shares = _safe_int(data.get('shares'))
        post.follower_delta = _safe_int(data.get('follower_delta'))
        raw_is_viral = data.get('is_viral')
        if raw_is_viral is None:
            post.is_viral = _infer_viral_post(
                views=post.views,
                likes=post.likes,
                favorites=post.favorites,
                comments=post.comments,
                exposures=post.exposures
            )
        else:
            post.is_viral = _coerce_bool(raw_is_viral)
        post.source_channel = (data.get('source_channel') or post.source_channel or 'manual').strip()
        post.raw_payload = json.dumps(data, ensure_ascii=False)
        account.last_synced_at = datetime.now()
        db.session.flush()
        sync_tracking_for_creator_account(account)
        _log_operation('save', 'creator_post', target_id=post.id, message='保存账号笔记表现', detail={
            'creator_account_id': account.id,
            'title': post.title,
            'views': post.views,
            'is_viral': bool(post.is_viral),
        })
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '账号笔记已保存',
            'item': _serialize_creator_post(post)
        })

    post_query = CreatorPost.query.filter_by(creator_account_id=account.id)
    post_query, _, _ = _apply_creator_post_range(post_query, request.args)
    posts = post_query.order_by(
        CreatorPost.publish_time.desc(), CreatorPost.created_at.desc()
    ).limit(200).all()
    return jsonify({
        'success': True,
        'account': _serialize_creator_account(account),
        'items': [_serialize_creator_post(post) for post in posts]
    })


@app.route('/api/creator_accounts/<int:account_id>/posts/<int:post_id>/viral', methods=['POST'])
def update_creator_post_viral(account_id, post_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    account = CreatorAccount.query.get_or_404(account_id)
    post = CreatorPost.query.filter_by(id=post_id, creator_account_id=account.id).first()
    if not post:
        return jsonify({'success': False, 'message': '笔记不存在'})

    data = request.json or {}
    is_viral = _coerce_bool(data.get('is_viral'))
    post.is_viral = is_viral
    _log_operation('mark_viral', 'creator_post', target_id=post.id, message='人工修正爆款标记', detail={
        'creator_account_id': account.id,
        'title': post.title,
        'is_viral': is_viral,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已将笔记标记为{"爆款" if is_viral else "普通"}',
        'item': _serialize_creator_post(post)
    })


@app.route('/api/creator_accounts/<int:account_id>/snapshots', methods=['GET', 'POST'])
def creator_account_snapshots(account_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    account = CreatorAccount.query.get_or_404(account_id)
    if request.method == 'POST':
        data = request.json or {}
        snapshot_date = _parse_date(data.get('snapshot_date')) or datetime.now().date()
        snapshot = CreatorAccountSnapshot(
            creator_account_id=account.id,
            snapshot_date=snapshot_date,
            follower_count=_safe_int(data.get('follower_count'), account.follower_count or 0),
            post_count=_safe_int(data.get('post_count')),
            total_views=_safe_int(data.get('total_views')),
            total_interactions=_safe_int(data.get('total_interactions')),
            source_channel=(data.get('source_channel') or 'manual').strip()
        )
        account.follower_count = snapshot.follower_count
        account.last_synced_at = datetime.now()
        db.session.add(snapshot)
        db.session.flush()
        _log_operation('save', 'creator_snapshot', message='保存账号快照', detail={
            'snapshot_id': snapshot.id,
            'creator_account_id': account.id,
            'snapshot_date': snapshot_date.isoformat() if snapshot_date else '',
            'follower_count': snapshot.follower_count,
        })
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '账号快照已保存',
            'item': {
                'id': snapshot.id,
                'snapshot_date': snapshot.snapshot_date.isoformat() if snapshot.snapshot_date else '',
                'follower_count': snapshot.follower_count or 0,
                'post_count': snapshot.post_count or 0,
                'total_views': snapshot.total_views or 0,
                'total_interactions': snapshot.total_interactions or 0,
                'source_channel': snapshot.source_channel or '',
                'created_at': snapshot.created_at.strftime('%Y-%m-%d %H:%M:%S') if snapshot.created_at else '',
            }
        })

    snapshot_query = CreatorAccountSnapshot.query.filter_by(creator_account_id=account.id)
    date_from = _parse_date(request.args.get('date_from'))
    date_to = _parse_date(request.args.get('date_to'))
    if date_from:
        snapshot_query = snapshot_query.filter(CreatorAccountSnapshot.snapshot_date >= date_from)
    if date_to:
        snapshot_query = snapshot_query.filter(CreatorAccountSnapshot.snapshot_date <= date_to)

    snapshots = snapshot_query.order_by(
        CreatorAccountSnapshot.snapshot_date.desc(), CreatorAccountSnapshot.created_at.desc()
    ).limit(120).all()
    return jsonify({
        'success': True,
        'account': _serialize_creator_account(account),
        'items': [{
            'id': snapshot.id,
            'snapshot_date': snapshot.snapshot_date.isoformat() if snapshot.snapshot_date else '',
            'follower_count': snapshot.follower_count or 0,
            'post_count': snapshot.post_count or 0,
            'total_views': snapshot.total_views or 0,
            'total_interactions': snapshot.total_interactions or 0,
            'source_channel': snapshot.source_channel or '',
            'created_at': snapshot.created_at.strftime('%Y-%m-%d %H:%M:%S') if snapshot.created_at else '',
        } for snapshot in snapshots]
    })


@app.route('/api/creator_accounts/<int:account_id>/analytics')
def creator_account_analytics(account_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    account = CreatorAccount.query.get_or_404(account_id)
    post_query = CreatorPost.query.filter_by(creator_account_id=account.id)
    post_query, date_from, date_to = _apply_creator_post_range(post_query, request.args)
    posts = post_query.order_by(CreatorPost.publish_time.asc(), CreatorPost.created_at.asc()).all()

    snapshot_query = CreatorAccountSnapshot.query.filter_by(creator_account_id=account.id)
    if date_from:
        snapshot_query = snapshot_query.filter(CreatorAccountSnapshot.snapshot_date >= date_from)
    if date_to:
        snapshot_query = snapshot_query.filter(CreatorAccountSnapshot.snapshot_date <= date_to)
    snapshots = snapshot_query.order_by(CreatorAccountSnapshot.snapshot_date.asc(), CreatorAccountSnapshot.created_at.asc()).all()

    daily_map = defaultdict(lambda: {
        'post_count': 0,
        'viral_posts': 0,
        'views': 0,
        'exposures': 0,
        'interactions': 0,
        'follower_delta': 0,
    })
    topic_map = defaultdict(lambda: {'post_count': 0, 'views': 0, 'interactions': 0})
    for post in posts:
        base_dt = post.publish_time or post.created_at or datetime.now()
        day_key = base_dt.strftime('%Y-%m-%d')
        interactions = (post.likes or 0) + (post.favorites or 0) + (post.comments or 0)
        row = daily_map[day_key]
        row['post_count'] += 1
        row['viral_posts'] += 1 if post.is_viral else 0
        row['views'] += post.views or 0
        row['exposures'] += post.exposures or 0
        row['interactions'] += interactions
        row['follower_delta'] += post.follower_delta or 0

        topic_key = (post.topic_title or '未关联话题').strip()
        topic_row = topic_map[topic_key]
        topic_row['post_count'] += 1
        topic_row['views'] += post.views or 0
        topic_row['interactions'] += interactions

    daily_rows = []
    for day_key in sorted(daily_map.keys()):
        item = daily_map[day_key]
        daily_rows.append({
            'date': day_key,
            **item,
        })

    snapshot_rows = [{
        'date': snapshot.snapshot_date.isoformat() if snapshot.snapshot_date else '',
        'follower_count': snapshot.follower_count or 0,
        'post_count': snapshot.post_count or 0,
        'total_views': snapshot.total_views or 0,
        'total_interactions': snapshot.total_interactions or 0,
    } for snapshot in snapshots]

    top_topics = []
    for topic_name, item in topic_map.items():
        top_topics.append({
            'topic_title': topic_name,
            **item,
        })
    top_topics.sort(key=lambda row: (row['interactions'], row['views'], row['post_count']), reverse=True)

    total_views = sum(post.views or 0 for post in posts)
    total_exposures = sum(post.exposures or 0 for post in posts)
    total_interactions = sum((post.likes or 0) + (post.favorites or 0) + (post.comments or 0) for post in posts)
    total_follower_delta = sum(post.follower_delta or 0 for post in posts)
    viral_posts = len([post for post in posts if post.is_viral])
    month_from, month_to = _current_month_date_range()
    current_month_posts = [
        post for post in posts
        if post.publish_time and month_from <= post.publish_time.date() <= month_to
    ]
    best_post = sorted(
        posts,
        key=lambda item: ((item.views or 0), ((item.likes or 0) + (item.favorites or 0) + (item.comments or 0)), (item.follower_delta or 0)),
        reverse=True
    )[0] if posts else None

    overview = {
        'post_count': len(posts),
        'viral_posts': viral_posts,
        'total_views': total_views,
        'total_exposures': total_exposures,
        'total_interactions': total_interactions,
        'total_follower_delta': total_follower_delta,
        'avg_views': round(total_views / len(posts), 2) if posts else 0,
        'avg_interactions': round(total_interactions / len(posts), 2) if posts else 0,
        'best_post': _serialize_creator_post(best_post) if best_post else None,
        'date_from': date_from.isoformat() if date_from else '',
        'date_to': date_to.isoformat() if date_to else '',
        'current_month_label': month_from.strftime('%Y-%m'),
        'current_month_post_count': len(current_month_posts),
        'current_month_total_views': sum(post.views or 0 for post in current_month_posts),
        'current_month_total_interactions': sum((post.likes or 0) + (post.favorites or 0) + (post.comments or 0) for post in current_month_posts),
    }

    return jsonify({
        'success': True,
        'account': _serialize_creator_account(account),
        'overview': overview,
        'daily_posts': daily_rows,
        'daily_snapshots': snapshot_rows,
        'top_topics': top_topics[:10],
        'current_month_posts': [_serialize_creator_post(post) for post in current_month_posts[:20]],
    })

# 活动管理
@app.route('/activity')
def activity_list():
    activities = Activity.query.order_by(Activity.created_at.desc()).all()
    return render_template('activity_list.html', activities=activities)

@app.route('/activity/<int:activity_id>')
def activity_detail(activity_id):
    activity = Activity.query.get_or_404(activity_id)
    return render_template('activity_detail.html', activity=activity)

@app.route('/api/activity/<int:activity_id>/snapshots')
def list_activity_snapshots(activity_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    snapshots = ActivitySnapshot.query.filter_by(activity_id=activity_id).order_by(ActivitySnapshot.created_at.desc()).all()
    items = []
    for snapshot in snapshots:
        _, summary = _snapshot_payload_and_summary(snapshot)
        items.append({
            'id': snapshot.id,
            'activity_id': snapshot.activity_id,
            'snapshot_name': snapshot.snapshot_name or '',
            'source_status': snapshot.source_status or '',
            'topic_count': summary.get('topic_count', 0),
            'registration_count': summary.get('registration_count', 0),
            'submission_count': summary.get('submission_count', 0),
            'created_at': snapshot.created_at.strftime('%Y-%m-%d %H:%M:%S') if snapshot.created_at else '',
        })
    return jsonify({'success': True, 'items': items})


@app.route('/api/activity_snapshots/<int:snapshot_id>')
def activity_snapshot_detail(snapshot_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    snapshot = ActivitySnapshot.query.get_or_404(snapshot_id)
    payload, _ = _snapshot_payload_and_summary(snapshot)
    return jsonify({
        'success': True,
        'item': {
            'id': snapshot.id,
            'activity_id': snapshot.activity_id,
            'snapshot_name': snapshot.snapshot_name or '',
            'source_status': snapshot.source_status or '',
            'created_at': snapshot.created_at.strftime('%Y-%m-%d %H:%M:%S') if snapshot.created_at else '',
            'payload': payload,
        }
    })


@app.route('/api/operation_logs')
def operation_logs():
    guard = _admin_json_guard()
    if guard:
        return guard

    action = (request.args.get('action') or '').strip()
    target_type = (request.args.get('target_type') or '').strip()
    keyword = (request.args.get('keyword') or '').strip()
    limit = min(max(_safe_int(request.args.get('limit'), 50), 1), 500)

    query = OperationLog.query
    if action:
        query = query.filter_by(action=action)
    if target_type:
        query = query.filter_by(target_type=target_type)
    if keyword:
        query = query.filter(or_(
            OperationLog.actor.contains(keyword),
            OperationLog.message.contains(keyword),
            OperationLog.detail.contains(keyword),
        ))

    logs = query.order_by(OperationLog.created_at.desc(), OperationLog.id.desc()).limit(limit).all()
    return jsonify({
        'success': True,
        'items': [_serialize_operation_log(log) for log in logs]
    })


@app.route('/api/admin/backups')
def list_backup_records():
    guard = _admin_json_guard()
    if guard:
        return guard

    backup_type = (request.args.get('backup_type') or '').strip()
    activity_id = _safe_int(request.args.get('activity_id'), 0)
    limit = min(max(_safe_int(request.args.get('limit'), 50), 1), 200)

    query = BackupRecord.query
    if backup_type:
        query = query.filter_by(backup_type=backup_type)
    if activity_id:
        query = query.filter_by(activity_id=activity_id)

    items = query.order_by(BackupRecord.created_at.desc(), BackupRecord.id.desc()).limit(limit).all()
    return jsonify({
        'success': True,
        'items': [_serialize_backup_record(item) for item in items]
    })


@app.route('/api/admin/backups/<int:record_id>')
def backup_record_detail(record_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    item = BackupRecord.query.get_or_404(record_id)
    return jsonify({
        'success': True,
        'item': _serialize_backup_record(item)
    })


@app.route('/api/admin/backups/create', methods=['POST'])
def create_manual_backup():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    activity_id = _safe_int(data.get('activity_id'), 0)
    if not activity_id:
        return jsonify({'success': False, 'message': '请选择要备份的活动'})
    activity = Activity.query.get_or_404(activity_id)
    backup_name = (data.get('backup_name') or '').strip()

    snapshot = _create_activity_snapshot(activity, snapshot_name=backup_name)
    db.session.flush()
    _, summary = _snapshot_payload_and_summary(snapshot)
    record = _create_backup_record(
        backup_type='manual_backup',
        target_type='activity',
        target_id=activity.id,
        activity_id=activity.id,
        snapshot_id=snapshot.id,
        status='success',
        trigger_mode='manual',
        backup_name=snapshot.snapshot_name or backup_name or f'{activity.name} 手动备份',
        payload=summary,
        summary=f'手动备份：话题 {summary.get("topic_count", 0)} 个，报名 {summary.get("registration_count", 0)} 条，提报 {summary.get("submission_count", 0)} 条',
    )
    _log_operation('backup', 'activity', target_id=activity.id, message='创建手动备份', detail={
        'snapshot_id': snapshot.id,
        'backup_record_id': record.id,
        'backup_name': record.backup_name,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已创建手动备份：{record.backup_name}',
        'item': _serialize_backup_record(record),
    })


@app.route('/api/admin/backups/<int:record_id>/restore', methods=['POST'])
def restore_from_backup_record(record_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    record = BackupRecord.query.get_or_404(record_id)
    snapshot_id = record.snapshot_id
    if not snapshot_id:
        return jsonify({'success': False, 'message': '该备份记录没有可恢复的快照'})
    snapshot = ActivitySnapshot.query.get(snapshot_id)
    if not snapshot:
        return jsonify({'success': False, 'message': '快照不存在，无法恢复'})

    data = request.json or {}
    restored = _restore_activity_from_snapshot(
        snapshot,
        name=(data.get('name') or '').strip(),
        title=(data.get('title') or '').strip(),
        description=(data.get('description') or '').strip() if data.get('description') is not None else None,
    )
    db.session.flush()
    record.restored_activity_id = restored.id
    _log_operation('restore_backup', 'backup_record', target_id=record.id, message='从备份记录恢复新活动', detail={
        'backup_record_id': record.id,
        'snapshot_id': snapshot.id,
        'restored_activity_id': restored.id,
        'restored_name': restored.name,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已从备份恢复为新活动：{restored.name}',
        'backup': _serialize_backup_record(record),
        'activity': _serialize_activity(restored),
    })


@app.route('/api/admin/users', methods=['GET', 'POST'])
def admin_users():
    guard = _admin_permission_guard('admin_user.manage')
    if guard:
        return guard

    if request.method == 'POST':
        data = request.json or {}
        user_id = _safe_int(data.get('id'), 0)
        if user_id:
            user = AdminUser.query.get_or_404(user_id)
        else:
            user = AdminUser()
            db.session.add(user)

        username = (data.get('username') or '').strip()
        password = (data.get('password') or '').strip()
        if not username:
            return jsonify({'success': False, 'message': '用户名不能为空'})
        if not user_id and not password:
            return jsonify({'success': False, 'message': '新用户必须设置密码'})

        duplicate = AdminUser.query.filter(AdminUser.username == username, AdminUser.id != user.id).first()
        if duplicate:
            return jsonify({'success': False, 'message': '用户名已存在'})

        user.username = username
        user.display_name = (data.get('display_name') or username).strip()[:100]
        user.role_key = (data.get('role_key') or 'super_admin').strip()[:50]
        user.status = (data.get('status') or 'active').strip()[:20]
        if password:
            user.password = password[:200]
        db.session.flush()
        _log_operation('save', 'admin_user', target_id=user.id, message='保存管理员用户', detail={
            'username': user.username,
            'role_key': user.role_key,
            'status': user.status,
        })
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '管理员用户已保存',
            'item': _serialize_admin_user(user),
        })

    items = AdminUser.query.order_by(AdminUser.created_at.asc(), AdminUser.id.asc()).all()
    return jsonify({
        'success': True,
        'items': [_serialize_admin_user(item) for item in items]
    })


@app.route('/api/admin/roles', methods=['GET', 'POST'])
def admin_roles():
    guard = _admin_permission_guard('role_permission.manage')
    if guard:
        return guard

    if request.method == 'POST':
        data = request.json or {}
        role_id = _safe_int(data.get('id'), 0)
        if role_id:
            role = RolePermission.query.get_or_404(role_id)
        else:
            role = RolePermission()
            db.session.add(role)

        role_key = (data.get('role_key') or '').strip()
        role_name = (data.get('role_name') or '').strip()
        permissions = data.get('permissions') or []
        if not role_key or not role_name:
            return jsonify({'success': False, 'message': '角色编码和角色名称不能为空'})

        duplicate = RolePermission.query.filter(RolePermission.role_key == role_key, RolePermission.id != role.id).first()
        if duplicate:
            return jsonify({'success': False, 'message': '角色编码已存在'})

        normalized_permissions = [str(item).strip() for item in permissions if str(item).strip()]
        role.role_key = role_key[:50]
        role.role_name = role_name[:100]
        role.permissions = json.dumps(sorted(set(normalized_permissions)), ensure_ascii=False)
        db.session.flush()
        _log_operation('save', 'role_permission', target_id=role.id, message='保存角色权限', detail={
            'role_key': role.role_key,
            'permission_count': len(normalized_permissions),
        })
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '角色权限已保存',
            'item': _serialize_role_permission(role),
        })

    items = RolePermission.query.order_by(RolePermission.created_at.asc(), RolePermission.id.asc()).all()
    return jsonify({
        'success': True,
        'items': [_serialize_role_permission(item) for item in items]
    })


@app.route('/api/admin/settings', methods=['GET', 'POST'])
def admin_settings():
    guard = _admin_permission_guard('settings.manage')
    if guard:
        return guard

    managed_keys = ['default_topic_quota', 'automation_keyword_seeds', 'automation_runtime_config']

    if request.method == 'POST':
        data = request.json or {}
        updates = {}
        for key in managed_keys:
            if key not in data:
                continue
            value = data.get(key)
            if isinstance(value, (dict, list)):
                value_text = json.dumps(value, ensure_ascii=False)
            else:
                value_text = str(value)
            setting = Settings.query.filter_by(key=key).first()
            if not setting:
                setting = Settings(key=key, value='')
                db.session.add(setting)
            setting.value = value_text
            updates[key] = value

        _log_operation('save', 'settings', message='更新系统配置', detail=updates)
        db.session.commit()

    items = Settings.query.filter(Settings.key.in_(managed_keys)).all()
    result = {}
    for item in items:
        if item.key in {'automation_keyword_seeds', 'automation_runtime_config'}:
            result[item.key] = _load_json_value(item.value, [] if item.key == 'automation_keyword_seeds' else {})
        else:
            result[item.key] = item.value
    return jsonify({
        'success': True,
        'items': result,
        'available_permissions': sorted({
            permission
            for role in DEFAULT_ROLE_PERMISSIONS.values()
            for permission in role['permissions']
        }),
    })


@app.route('/api/jobs/ping', methods=['POST'])
def trigger_worker_ping():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    timeout_seconds = min(max(_safe_int(data.get('wait_seconds'), 3), 1), 15)
    result = _run_worker_ping_check(timeout_seconds=timeout_seconds)
    db.session.commit()
    return jsonify(result)


@app.route('/api/jobs/hotwords/run', methods=['POST'])
def trigger_hotword_sync_job():
    guard = _admin_json_guard()
    if guard:
        return guard

    payload = request.json or {}
    dispatched = _dispatch_hotword_sync(payload, actor=_current_actor())
    return jsonify({
        'success': True,
        'message': f'已触发热点抓取任务，关键词 {dispatched["keyword_count"]} 个',
        'task_id': dispatched['task_id'],
        'job': 'jobs.hotwords.sync',
        'data_source_task_id': dispatched['task_record'].id,
    })


@app.route('/api/jobs/hotwords/run_planning_bundle', methods=['POST'])
def trigger_hotword_planning_bundle():
    guard = _admin_json_guard()
    if guard:
        return guard

    payload = request.json or {}
    bundle = _dispatch_hotword_planning_bundle(payload, actor=_current_actor())
    return jsonify({
        'success': True,
        'message': f'已触发三块内容规划搜索任务，共 {len(bundle["items"])} 条',
        'batch_name': bundle['batch_name'],
        'items': bundle['items'],
    })


@app.route('/api/jobs/hotwords/ping', methods=['POST'])
def trigger_hotword_sync_ping():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    timeout_seconds = min(max(_safe_int(data.get('wait_seconds'), 3), 1), 15)
    keywords = split_hotword_keywords(data.get('keywords') or '')
    result = _hotword_healthcheck(payload=data, timeout_seconds=timeout_seconds, sample_keywords=keywords, include_rows=True)
    _log_integration_ping_result('hotword', result, request_payload=data)
    db.session.commit()
    return jsonify({
        'success': bool(result.get('ok')),
        **result,
    })


@app.route('/api/jobs/creator-accounts/run', methods=['POST'])
def trigger_creator_account_sync_job():
    guard = _admin_json_guard()
    if guard:
        return guard

    payload = request.json or {}
    try:
        dispatched = _dispatch_creator_account_sync(payload, actor=_current_actor())
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})
    return jsonify({
        'success': True,
        'message': f'已触发报名人账号同步任务，本轮目标 {dispatched["target_count"]} 个账号',
        'task_id': dispatched['task_id'],
        'job': 'jobs.creator_accounts.sync',
        'data_source_task_id': dispatched['task_record'].id,
    })


@app.route('/api/jobs/creator-accounts/ping', methods=['POST'])
def trigger_creator_account_sync_ping():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    timeout_seconds = min(max(_safe_int(data.get('wait_seconds'), 3), 1), 15)
    result = _creator_sync_healthcheck(timeout_seconds=timeout_seconds)
    _log_integration_ping_result('creator_sync', result, request_payload=data)
    db.session.commit()
    return jsonify({
        'success': bool(result.get('ok')),
        **result,
    })


@app.route('/api/jobs/assets/ping', methods=['POST'])
def trigger_asset_provider_ping():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    timeout_seconds = min(max(_safe_int(data.get('wait_seconds'), 10), 1), 60)
    result = _image_provider_healthcheck(payload=data, timeout_seconds=timeout_seconds)
    _log_integration_ping_result('image_provider', result, request_payload=data)
    db.session.commit()
    return jsonify({
        'success': bool(result.get('ok')),
        **result,
    })


@app.route('/api/jobs/copywriter/ping', methods=['POST'])
def trigger_copywriter_ping():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    timeout_seconds = min(max(_safe_int(data.get('wait_seconds'), 10), 1), 60)
    result = _copywriter_healthcheck(payload=data, timeout_seconds=timeout_seconds)
    _log_integration_ping_result('copywriter', result, request_payload=data)
    db.session.commit()
    return jsonify({
        'success': bool(result.get('ok')),
        **result,
    })


@app.route('/api/jobs/assets/generate', methods=['POST'])
def trigger_asset_generation_job():
    data = request.json or {}
    try:
        dispatched = _dispatch_asset_generation(data, actor=_current_actor())
    except ValueError as exc:
        return jsonify({'success': False, 'message': str(exc)})
    return jsonify({
        'success': True,
        'message': f'已创建图片生成任务，预计输出 {dispatched["image_count"]} 张。真实图片次数剩余 {dispatched["remaining_attempts"]}/{dispatched["max_attempts"]}',
        'task_id': dispatched['task_id'],
        'asset_task_id': dispatched['task_record'].id,
        'remaining_attempts': dispatched['remaining_attempts'],
        'used_attempts': dispatched['used_attempts'],
        'max_attempts': dispatched['max_attempts'],
        'decision': dispatched.get('decision') or {},
        'job': 'jobs.assets.generate',
    })


@app.route('/api/asset_generation_quota/<int:registration_id>')
def asset_generation_quota(registration_id):
    reg = Registration.query.get_or_404(registration_id)
    return jsonify({
        'success': True,
        'registration_id': reg.id,
        **_asset_generation_quota_payload(reg.id),
    })


@app.route('/api/jobs/topic_ideas/generate', methods=['POST'])
def trigger_generate_topic_ideas_job():
    guard = _admin_json_guard()
    if guard:
        return guard

    data = request.json or {}
    dispatched = _dispatch_topic_idea_generation(data, actor=_current_actor())
    return jsonify({
        'success': True,
        'message': '已触发异步生成候选话题任务',
        'task_id': dispatched['task_id'],
        'job': 'jobs.generate_topic_ideas',
    })


@app.route('/api/jobs/<task_id>')
def get_job_status(task_id):
    guard = _admin_json_guard()
    if guard:
        return guard

    from celery.result import AsyncResult
    from celery_app import celery

    result = AsyncResult(task_id, app=celery)
    payload = result.result
    if not isinstance(payload, (dict, list, str, int, float, bool, type(None))):
        payload = str(payload)
    return jsonify({
        'success': True,
        'task_id': task_id,
        'state': result.state,
        'result': payload,
    })


@app.route('/api/jobs/history')
def get_job_history():
    guard = _admin_json_guard()
    if guard:
        return guard

    from celery.result import AsyncResult
    from celery_app import celery

    limit = min(max(_safe_int(request.args.get('limit'), 20), 1), 200)
    logs = OperationLog.query.filter(OperationLog.action.in_([
        'dispatch_job', 'worker_generate', 'worker_sync', 'worker_generate_asset', 'worker_ping_check', 'worker_ping_check_failed'
    ])).order_by(
        OperationLog.created_at.desc(),
        OperationLog.id.desc()
    ).limit(limit).all()
    items = []
    for log in logs:
        detail_obj = _deserialize_operation_detail(log.detail)
        items.append({
            'id': log.id,
            'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
            'actor': log.actor or 'system',
            'action': log.action,
            'target_type': log.target_type,
            'target_id': log.target_id,
            'message': log.message or '',
            'detail': detail_obj,
            'task_id': detail_obj.get('task_id') or '',
            'state': '',
        })
        task_id = detail_obj.get('task_id')
        if task_id:
            try:
                async_result = AsyncResult(task_id, app=celery)
                items[-1]['state'] = async_result.state
            except Exception:
                items[-1]['state'] = 'UNKNOWN'
    return jsonify({
        'success': True,
        'items': items,
    })


@app.route('/admin/activity_snapshots/<int:snapshot_id>/restore', methods=['POST'])
def restore_activity_snapshot(snapshot_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'})

    snapshot = ActivitySnapshot.query.get(snapshot_id)
    if not snapshot:
        return jsonify({'success': False, 'message': '快照不存在'})

    data = request.json if request.is_json else request.form
    restored = _restore_activity_from_snapshot(
        snapshot,
        name=(data.get('name') or '').strip() if data else '',
        title=(data.get('title') or '').strip() if data else '',
        description=(data.get('description') or '').strip() if data else None,
    )
    db.session.flush()
    _, summary = _snapshot_payload_and_summary(snapshot)
    _create_backup_record(
        backup_type='restore_from_snapshot',
        target_type='activity_snapshot',
        target_id=snapshot.id,
        activity_id=snapshot.activity_id,
        snapshot_id=snapshot.id,
        status='success',
        trigger_mode='manual',
        backup_name=f'恢复 {snapshot.snapshot_name or snapshot.id}',
        payload=summary,
        summary=f'从快照恢复新活动：{restored.name}',
        restored_activity_id=restored.id,
    )
    _log_operation('restore_snapshot', 'activity', target_id=restored.id, message='从活动快照恢复新活动', detail={
        'snapshot_id': snapshot.id,
        'restored_name': restored.name,
        'source_activity_id': snapshot.activity_id,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'快照已恢复为新活动：{restored.name}',
        'activity': _serialize_activity(restored),
    })

@app.route('/admin')
def admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    allowed_tabs = {'status', 'activities', 'topics', 'portal', 'data', 'snapshots', 'backups', 'creator', 'logs', 'system'}
    initial_admin_tab = (request.args.get('tab') or 'activities').strip()
    if initial_admin_tab not in allowed_tabs:
        initial_admin_tab = 'activities'
    activities = Activity.query.order_by(Activity.created_at.desc()).all()
    return render_template(
        'admin.html',
        activities=activities,
        default_topic_quota=_default_topic_quota(),
        initial_admin_tab=initial_admin_tab,
    )


@app.route('/admin/project-status')
def admin_project_status():
    return redirect(url_for('admin', tab='status'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = (request.form.get('password') or '').strip()

        matched_user = AdminUser.query.filter_by(username=username, status='active').first()
        password_ok = False
        if matched_user and matched_user.password == password:
            password_ok = True
        elif username == os.environ.get('ADMIN_USERNAME', 'furui') and password == _admin_password():
            password_ok = True

        if password_ok:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            if matched_user:
                matched_user.last_login_at = datetime.now()
                session['admin_role_key'] = matched_user.role_key or 'super_admin'
                db.session.commit()
            else:
                session['admin_role_key'] = 'super_admin'
            return redirect(url_for('admin'))
        else:
            return render_template('admin_login.html', error='用户名或密码错误')

    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    session.pop('admin_role_key', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/activity/add', methods=['POST'])
def add_activity():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'})

    activity = Activity(
        name=request.form.get('name'),
        title=request.form.get('title'),
        description=request.form.get('description'),
        status='draft'
    )
    db.session.add(activity)
    db.session.flush()
    _log_operation('create', 'activity', target_id=activity.id, message='新增活动', detail={
        'name': activity.name,
        'title': activity.title,
    })
    db.session.commit()

    return jsonify({'success': True, 'message': '活动创建成功'})

@app.route('/admin/topic/add', methods=['POST'])
def add_topic():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'})

    quota = _normalize_quota(request.form.get('quota'))
    topic = Topic(
        activity_id=request.form.get('activity_id'),
        topic_name=request.form.get('topic_name'),
        keywords=request.form.get('keywords'),
        direction=request.form.get('direction'),
        reference_content=request.form.get('reference_content'),
        reference_link=request.form.get('reference_link'),
        quota=quota,
        group_num=request.form.get('group_num'),
        pool_status='formal',
        source_type='manual',
        published_at=datetime.now()
    )
    db.session.add(topic)
    db.session.flush()
    _log_operation('create', 'topic', target_id=topic.id, message='手工新增正式话题', detail={
        'activity_id': topic.activity_id,
        'topic_name': topic.topic_name,
        'quota': topic.quota,
    })
    db.session.commit()

    return jsonify({'success': True, 'message': '话题创建成功'})

@app.route('/admin/activity/<int:activity_id>/publish')
def publish_activity(activity_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'})

    activity = Activity.query.get(activity_id)
    if not activity:
        return jsonify({'success': False, 'message': '活动不存在'})
    activity.status = 'published'
    activity.archived_at = None
    _log_operation('publish', 'activity', target_id=activity.id, message='发布活动', detail={
        'name': activity.name,
        'title': activity.title,
    })
    db.session.commit()

    return jsonify({'success': True, 'message': '发布成功'})


@app.route('/admin/activity/<int:activity_id>/archive', methods=['POST'])
def archive_activity(activity_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'})

    activity = Activity.query.get(activity_id)
    if not activity:
        return jsonify({'success': False, 'message': '活动不存在'})
    if activity.status == 'archived':
        return jsonify({'success': False, 'message': '该活动已归档'})

    snapshot_name = (request.json or {}).get('snapshot_name', '') if request.is_json else ''
    snapshot = _create_activity_snapshot(activity, snapshot_name=snapshot_name)
    activity.status = 'archived'
    activity.archived_at = datetime.now()
    db.session.flush()
    _, summary = _snapshot_payload_and_summary(snapshot)
    _create_backup_record(
        backup_type='activity_snapshot',
        target_type='activity',
        target_id=activity.id,
        activity_id=activity.id,
        snapshot_id=snapshot.id,
        status='success',
        trigger_mode='manual',
        backup_name=snapshot.snapshot_name or f'{activity.name} 归档快照',
        payload=summary,
        summary=f'活动归档快照：话题 {summary.get("topic_count", 0)} 个，报名 {summary.get("registration_count", 0)} 条，提报 {summary.get("submission_count", 0)} 条',
    )
    _log_operation('archive', 'activity', target_id=activity.id, message='归档活动并生成快照', detail={
        'name': activity.name,
        'snapshot_id': snapshot.id,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': '活动已归档并生成快照',
        'activity': _serialize_activity(activity),
    })


@app.route('/admin/activity/<int:activity_id>/clone', methods=['POST'])
def clone_activity(activity_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': '未登录'})

    activity = Activity.query.get(activity_id)
    if not activity:
        return jsonify({'success': False, 'message': '活动不存在'})

    data = request.json if request.is_json else request.form
    cloned = _clone_activity(
        activity,
        name=(data.get('name') or '').strip() if data else '',
        title=(data.get('title') or '').strip() if data else '',
        description=(data.get('description') or '').strip() if data else None,
    )
    db.session.flush()
    _log_operation('clone', 'activity', target_id=cloned.id, message='复制活动为新一期', detail={
        'source_activity_id': activity.id,
        'cloned_name': cloned.name,
    })
    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'已复制为新活动：{cloned.name}',
        'activity': _serialize_activity(cloned),
    })

@app.route('/admin/export/<int:activity_id>')
def export_data(activity_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    topics = Topic.query.filter_by(activity_id=activity_id).all()
    topic_ids = [t.id for t in topics]

    registrations = Registration.query.filter(Registration.topic_id.in_(topic_ids)).all()

    # 生成CSV（多平台）
    csv_content = "姓名,小组号,小红书账号,联系方式,话题,小红书链接,小红书曝光量,小红书点赞量,小红书收藏量,小红书评论量,抖音链接,抖音曝光量,抖音点赞量,抖音收藏量,抖音评论量,视频号链接,视频号曝光量,视频号点赞量,视频号收藏量,视频号评论量,微博链接,微博曝光量,微博点赞量,微博收藏量,微博评论量\n"
    for reg in registrations:
        topic = reg.topic
        sub = reg.submission
        csv_content += f"{reg.name},{reg.group_num},{reg.xhs_account},{reg.phone},{topic.topic_name},{sub.xhs_link if sub else ''},{sub.xhs_views if sub else 0},{sub.xhs_likes if sub else 0},{sub.xhs_favorites if sub else 0},{sub.xhs_comments if sub else 0},{sub.douyin_link if sub else ''},{sub.douyin_views if sub else 0},{sub.douyin_likes if sub else 0},{sub.douyin_favorites if sub else 0},{sub.douyin_comments if sub else 0},{sub.video_link if sub else ''},{sub.video_views if sub else 0},{sub.video_likes if sub else 0},{sub.video_favorites if sub else 0},{sub.video_comments if sub else 0},{sub.weibo_link if sub else ''},{sub.weibo_views if sub else 0},{sub.weibo_likes if sub else 0},{sub.weibo_favorites if sub else 0},{sub.weibo_comments if sub else 0}\n"

    return csv_content, 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename=activity_{activity_id}_data.csv'
    }

# ==================== 初始化 ====================

_INIT_DB_DONE = False

def init_db():
    global _INIT_DB_DONE
    if _INIT_DB_DONE:
        return
    with app.app_context():
        db.create_all()

        timestamp_type = 'TIMESTAMP'
        schema_required_columns = {
            'activity': {
                'archived_at': timestamp_type,
                'source_type': {'type': 'VARCHAR(30)', 'default': 'manual'},
                'source_activity_id': 'INTEGER',
                'source_snapshot_id': 'INTEGER',
            },
            'topic': {
                'pool_status': {'type': 'VARCHAR(20)', 'default': 'formal'},
                'source_type': {'type': 'VARCHAR(30)', 'default': 'manual'},
                'source_ref_id': 'INTEGER',
                'source_snapshot_id': 'INTEGER',
                'published_at': timestamp_type,
            },
            'corpus_entry': {
                'pool_status': {'type': 'VARCHAR(20)', 'default': 'reserve'},
                'source_title': 'VARCHAR(300)',
                'reference_url': 'VARCHAR(500)',
                'template_type_key': 'VARCHAR(50)',
            },
            'trend_note': {
                'pool_status': {'type': 'VARCHAR(20)', 'default': 'reserve'},
                'source_template_key': {'type': 'VARCHAR(50)', 'default': 'generic_lines'},
                'hot_score': {'type': 'INTEGER', 'default': 0},
                'source_rank': {'type': 'INTEGER', 'default': 0},
            },
            'submission': {
                'xhs_profile_link': 'VARCHAR(500)',
                'xhs_views': {'type': 'INTEGER', 'default': 0},
                'xhs_likes': {'type': 'INTEGER', 'default': 0},
                'xhs_favorites': {'type': 'INTEGER', 'default': 0},
                'xhs_comments': {'type': 'INTEGER', 'default': 0},
                'xhs_creator_account_id': 'INTEGER',
                'xhs_primary_post_id': 'INTEGER',
                'xhs_tracking_enabled': {'type': 'BOOLEAN', 'default': False},
                'xhs_tracking_status': {'type': 'VARCHAR(30)', 'default': 'empty'},
                'xhs_tracking_message': 'VARCHAR(300)',
                'xhs_last_synced_at': timestamp_type,
                'douyin_link': 'VARCHAR(500)',
                'douyin_views': {'type': 'INTEGER', 'default': 0},
                'douyin_likes': {'type': 'INTEGER', 'default': 0},
                'douyin_favorites': {'type': 'INTEGER', 'default': 0},
                'douyin_comments': {'type': 'INTEGER', 'default': 0},
                'video_link': 'VARCHAR(500)',
                'video_views': {'type': 'INTEGER', 'default': 0},
                'video_likes': {'type': 'INTEGER', 'default': 0},
                'video_favorites': {'type': 'INTEGER', 'default': 0},
                'video_comments': {'type': 'INTEGER', 'default': 0},
                'weibo_link': 'VARCHAR(500)',
                'weibo_views': {'type': 'INTEGER', 'default': 0},
                'weibo_likes': {'type': 'INTEGER', 'default': 0},
                'weibo_favorites': {'type': 'INTEGER', 'default': 0},
                'weibo_comments': {'type': 'INTEGER', 'default': 0},
                'content_type': {'type': 'VARCHAR(30)', 'default': '未识别'},
                'note_title': 'VARCHAR(300)',
                'note_content': 'TEXT',
                'selected_title': 'VARCHAR(200)',
                'selected_title_source': 'VARCHAR(100)',
                'selected_title_index': {'type': 'INTEGER', 'default': 0},
                'selected_copy_version_index': {'type': 'INTEGER', 'default': 0},
                'selected_copy_goal': 'VARCHAR(50)',
                'selected_copy_skill': 'VARCHAR(50)',
                'selected_title_skill': 'VARCHAR(50)',
                'selected_image_skill': 'VARCHAR(50)',
                'selected_cover_style_type': 'VARCHAR(50)',
                'selected_inner_style_type': 'VARCHAR(50)',
                'selected_generation_mode': 'VARCHAR(50)',
                'selected_copy_text': 'TEXT',
                'strategy_payload': 'TEXT',
                'strategy_updated_at': timestamp_type,
            },
            'topic_idea': {
                'quota': {'type': 'INTEGER', 'default': 30},
                'review_note': 'TEXT',
                'reviewed_at': timestamp_type,
                'published_at': timestamp_type,
                'published_topic_id': 'INTEGER',
            },
            'creator_post': {
                'registration_id': 'INTEGER',
                'topic_id': 'INTEGER',
                'submission_id': 'INTEGER',
            },
            'asset_generation_task': {
                'draft_source_type': 'VARCHAR(20)',
                'draft_source_id': 'INTEGER',
                'draft_plan_id': 'INTEGER',
                'generation_mode': {'type': 'VARCHAR(50)', 'default': 'smart_bundle'},
                'cover_style_type': 'VARCHAR(50)',
                'inner_style_type': 'VARCHAR(50)',
                'product_profile': 'VARCHAR(80)',
                'product_category': 'VARCHAR(30)',
                'product_name': 'VARCHAR(200)',
                'product_indication': 'VARCHAR(200)',
                'product_asset_ids': 'VARCHAR(500)',
                'reference_asset_ids': 'VARCHAR(500)',
            },
            'asset_plan_draft': {
                'source_type': {'type': 'VARCHAR(20)', 'default': 'trend'},
                'source_id': 'INTEGER',
                'source_title': 'VARCHAR(200)',
                'bucket_label': 'VARCHAR(100)',
                'template_agent_label': 'VARCHAR(100)',
                'image_skill_label': 'VARCHAR(100)',
                'style_type': 'VARCHAR(50)',
                'generation_mode': {'type': 'VARCHAR(50)', 'default': 'smart_bundle'},
                'cover_style_type': 'VARCHAR(50)',
                'inner_style_type': 'VARCHAR(50)',
                'title_hint': 'VARCHAR(200)',
                'selected_content': 'TEXT',
                'draft_payload': 'TEXT',
            },
            'asset_library': {
                'style_type_key': 'VARCHAR(50)',
                'product_category': 'VARCHAR(30)',
                'product_name': 'VARCHAR(200)',
                'product_indication': 'VARCHAR(200)',
                'visual_role': 'VARCHAR(50)',
            },
        }

        def current_columns(table_name):
            return {column['name'] for column in inspect(db.engine).get_columns(table_name)}

        def ensure_columns(conn, table_name, required_columns):
            existing = current_columns(table_name)
            for column_name, column_type in required_columns.items():
                if column_name in existing:
                    continue
                rendered_column_sql = _render_schema_column_sql(column_type)
                try:
                    conn.execute(text(f'ALTER TABLE {table_name} ADD COLUMN {column_name} {rendered_column_sql}'))
                except Exception as exc:
                    message = str(exc).lower()
                    if 'duplicate column name' in message or 'already exists' in message:
                        existing.add(column_name)
                        continue
                    raise RuntimeError(
                        f'auto schema migration failed for {table_name}.{column_name}: {rendered_column_sql}'
                    ) from exc
                existing.add(column_name)
            return existing

        with db.engine.begin() as conn:
            existing_columns = {}
            for table_name, required_columns in schema_required_columns.items():
                existing_columns[table_name] = ensure_columns(conn, table_name, required_columns)
            for table_name, index_specs in SCHEMA_REQUIRED_INDEXES.items():
                _ensure_indexes(conn, table_name, index_specs)

            submission_columns = existing_columns.get('submission', set())
            if {'likes', 'favorites', 'comments'}.issubset(submission_columns):
                conn.execute(text(
                    "UPDATE submission SET "
                    "xhs_likes = COALESCE(xhs_likes, likes), "
                    "xhs_favorites = COALESCE(xhs_favorites, favorites), "
                    "xhs_comments = COALESCE(xhs_comments, comments)"
                ))

            conn.execute(text(
                "UPDATE topic_idea SET status = 'pending_review' "
                "WHERE status IS NULL OR status = '' OR status = 'draft'"
            ))
            conn.execute(text("UPDATE topic_idea SET quota = COALESCE(quota, 30)"))
            conn.execute(text("UPDATE topic SET pool_status = COALESCE(pool_status, 'formal')"))
            conn.execute(text("UPDATE topic SET source_type = COALESCE(source_type, 'manual')"))
            conn.execute(text("UPDATE topic SET published_at = COALESCE(published_at, created_at)"))
            conn.execute(text("UPDATE activity SET source_type = COALESCE(source_type, 'manual')"))
            conn.execute(text("UPDATE corpus_entry SET pool_status = COALESCE(pool_status, 'reserve')"))
            conn.execute(text("UPDATE trend_note SET pool_status = COALESCE(pool_status, 'reserve')"))
            conn.execute(text("UPDATE trend_note SET source_template_key = COALESCE(source_template_key, 'generic_lines')"))
            conn.execute(text("UPDATE trend_note SET hot_score = COALESCE(hot_score, 0)"))
            conn.execute(text("UPDATE trend_note SET source_rank = COALESCE(source_rank, 0)"))

        # 通用回填：PostgreSQL/SQLite 都可安全执行
        Activity.query.filter(Activity.source_type.is_(None)).update(
            {'source_type': 'manual'},
            synchronize_session=False
        )
        Topic.query.filter(Topic.pool_status.is_(None)).update(
            {'pool_status': 'formal'},
            synchronize_session=False
        )
        Topic.query.filter(Topic.source_type.is_(None)).update(
            {'source_type': 'manual'},
            synchronize_session=False
        )
        CorpusEntry.query.filter(CorpusEntry.pool_status.is_(None)).update(
            {'pool_status': 'reserve'},
            synchronize_session=False
        )
        TrendNote.query.filter(TrendNote.pool_status.is_(None)).update(
            {'pool_status': 'reserve'},
            synchronize_session=False
        )
        TrendNote.query.filter(TrendNote.source_template_key.is_(None)).update(
            {'source_template_key': 'generic_lines'},
            synchronize_session=False
        )
        TrendNote.query.filter(TrendNote.hot_score.is_(None)).update(
            {'hot_score': 0},
            synchronize_session=False
        )
        TrendNote.query.filter(TrendNote.source_rank.is_(None)).update(
            {'source_rank': 0},
            synchronize_session=False
        )
        TopicIdea.query.filter(or_(TopicIdea.status.is_(None), TopicIdea.status == '', TopicIdea.status == 'draft')).update(
            {'status': 'pending_review'},
            synchronize_session=False
        )
        db.session.commit()

        Submission.query.filter(Submission.xhs_tracking_status.is_(None)).update(
            {'xhs_tracking_status': 'empty'},
            synchronize_session=False
        )
        Submission.query.filter(Submission.xhs_tracking_enabled.is_(None)).update(
            {'xhs_tracking_enabled': False},
            synchronize_session=False
        )
        db.session.commit()

        if CorpusEntry.query.count() == 0:
            for seed in CORPUS_SEED_ENTRIES:
                db.session.add(CorpusEntry(**seed))
            db.session.commit()

        keyword_setting = Settings.query.filter_by(key='automation_keyword_seeds').first()
        if not keyword_setting:
            keyword_setting = Settings(key='automation_keyword_seeds', value=json.dumps(LIVER_KEYWORD_SEEDS, ensure_ascii=False))
            db.session.add(keyword_setting)
            db.session.commit()

        default_quota_setting = Settings.query.filter_by(key='default_topic_quota').first()
        if not default_quota_setting:
            default_quota_setting = Settings(key='default_topic_quota', value='30')
            db.session.add(default_quota_setting)
            db.session.commit()
        automation_runtime_setting = Settings.query.filter_by(key='automation_runtime_config').first()
        if not automation_runtime_setting:
            automation_runtime_setting = Settings(
                key='automation_runtime_config',
                value=json.dumps(AUTOMATION_RUNTIME_CONFIG_DEFAULTS, ensure_ascii=False)
            )
            db.session.add(automation_runtime_setting)
            db.session.commit()

        existing_roles = {item.role_key for item in RolePermission.query.all()}
        for role_key, row in DEFAULT_ROLE_PERMISSIONS.items():
            if role_key in existing_roles:
                continue
            db.session.add(RolePermission(
                role_key=role_key,
                role_name=row['role_name'],
                permissions=json.dumps(row['permissions'], ensure_ascii=False),
            ))
        db.session.commit()

        env_admin_username = os.environ.get('ADMIN_USERNAME', 'furui').strip() or 'furui'
        env_admin_password = _admin_password()
        admin_user = AdminUser.query.filter_by(username=env_admin_username).first()
        if not admin_user:
            admin_user = AdminUser(
                username=env_admin_username,
                password=env_admin_password,
                display_name=env_admin_username,
                role_key='super_admin',
                status='active',
            )
            db.session.add(admin_user)
            db.session.commit()
        default_quota = _default_topic_quota()

        Topic.query.filter(or_(Topic.quota.is_(None), Topic.quota <= 0)).update(
            {'quota': default_quota},
            synchronize_session=False
        )
        TopicIdea.query.filter(or_(TopicIdea.quota.is_(None), TopicIdea.quota <= 0)).update(
            {'quota': default_quota},
            synchronize_session=False
        )
        db.session.commit()

        if SiteTheme.query.count() == 0:
            site_theme = SiteTheme(theme_key=DEFAULT_SITE_THEME['theme_key'], name=DEFAULT_SITE_THEME['name'], is_active=True)
            for field, value in DEFAULT_SITE_THEME.items():
                if field in {'theme_key', 'name'}:
                    continue
                setattr(site_theme, field, value)
            db.session.add(site_theme)
            db.session.commit()
        elif not SiteTheme.query.filter_by(is_active=True).first():
            first_theme = SiteTheme.query.order_by(SiteTheme.id.asc()).first()
            if first_theme:
                first_theme.is_active = True
                db.session.commit()

        home_page_config = SitePageConfig.query.filter_by(page_key=DEFAULT_HOME_PAGE_CONFIG['page_key']).first()
        if not home_page_config:
            home_page_config = SitePageConfig(
                page_key=DEFAULT_HOME_PAGE_CONFIG['page_key'],
                site_name=DEFAULT_HOME_PAGE_CONFIG['site_name'],
                page_title=DEFAULT_HOME_PAGE_CONFIG['page_title'],
                hero_badge=DEFAULT_HOME_PAGE_CONFIG['hero_badge'],
                hero_title=DEFAULT_HOME_PAGE_CONFIG['hero_title'],
                hero_subtitle=DEFAULT_HOME_PAGE_CONFIG['hero_subtitle'],
                announcement_title=DEFAULT_HOME_PAGE_CONFIG['announcement_title'],
                trend_title=DEFAULT_HOME_PAGE_CONFIG['trend_title'],
                primary_section_title=DEFAULT_HOME_PAGE_CONFIG['primary_section_title'],
                primary_section_icon=DEFAULT_HOME_PAGE_CONFIG['primary_section_icon'],
                secondary_section_title=DEFAULT_HOME_PAGE_CONFIG['secondary_section_title'],
                secondary_section_icon=DEFAULT_HOME_PAGE_CONFIG['secondary_section_icon'],
                primary_topic_limit=DEFAULT_HOME_PAGE_CONFIG['primary_topic_limit'],
                footer_text=DEFAULT_HOME_PAGE_CONFIG['footer_text'],
                nav_items=json.dumps(DEFAULT_SITE_NAV_ITEMS, ensure_ascii=False),
            )
            db.session.add(home_page_config)
            db.session.commit()
        elif not home_page_config.nav_items:
            home_page_config.nav_items = json.dumps(DEFAULT_SITE_NAV_ITEMS, ensure_ascii=False)
            db.session.commit()

        existing_schedule_keys = {item.job_key for item in AutomationSchedule.query.all()}
        for row in _default_automation_schedules(default_quota):
            if row['job_key'] in existing_schedule_keys:
                continue
            schedule = AutomationSchedule(
                job_key=row['job_key'],
                name=row['name'],
                task_type=row['task_type'],
                enabled=row['enabled'],
                interval_minutes=row['interval_minutes'],
                params_payload=row['params_payload'],
                next_run_at=_next_schedule_time(row['interval_minutes']),
                last_status='paused' if not row['enabled'] else 'idle',
                last_message='等待首次执行' if row['enabled'] else '默认暂停，需手动开启',
            )
            db.session.add(schedule)
        db.session.commit()

        # 如果没有活动，创建默认活动
        if Activity.query.count() == 0:
            activity = Activity(
                name='第1期',
                title='福瑞医科小红书任务第一期',
                description='邀请员工参与小红书推广任务',
                status='published'
            )
            db.session.add(activity)
            db.session.commit()

            # 添加示例话题
            topics = [
                {'topic_name': 'FibroScan肝纤维化检测', 'keywords': 'FibroScan,肝纤维化,肝脏健康', 'direction': '分享FibroScan检测体验', 'quota': default_quota, 'group_num': '第一组'},
                {'topic_name': '脂肪肝科普', 'keywords': '脂肪肝,肝脏健康,体检', 'direction': '科普脂肪肝知识', 'quota': default_quota, 'group_num': '第一组'},
                {'topic_name': '护肝片推荐', 'keywords': '护肝,肝脏保健,复方鳖甲', 'direction': '分享护肝产品', 'quota': default_quota, 'group_num': '第二组'},
            ]
            for t in topics:
                topic = Topic(activity_id=activity.id, **t)
                db.session.add(topic)
            db.session.commit()
            print("数据库初始化完成")

        backfill_submission_tracking()
        db.session.commit()
        _INIT_DB_DONE = True


if __name__ != '__main__':
    init_db()


if __name__ == '__main__':
    init_db()
    app.run(
        host='0.0.0.0',
        port=int(os.environ.get('PORT', '5000')),
        debug=_env_flag('FLASK_DEBUG', False)
    )
