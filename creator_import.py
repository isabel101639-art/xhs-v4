import csv
import json
import os
from collections import Counter
from datetime import datetime
from io import BytesIO, StringIO

from creator_tracking import canonicalize_xhs_post_url, normalize_tracking_url, sync_tracking_for_creator_account
from models import (
    db,
    CreatorAccount,
    CreatorPost,
    CreatorAccountSnapshot,
    PLATFORM_DEFINITIONS,
)


def _safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _coerce_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {'1', 'true', 'yes', 'y', 'on'}
    return False


def _parse_datetime(value):
    if not value:
        return None
    value = str(value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d']:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), '%Y-%m-%d').date()
    except ValueError:
        return None


def _infer_viral_post(views=0, likes=0, favorites=0, comments=0, exposures=0):
    interactions = (likes or 0) + (favorites or 0) + (comments or 0)
    return (views or 0) >= 10000 or (exposures or 0) >= 30000 or interactions >= 1000


def normalize_creator_platform(value=''):
    raw = (value or '').strip().lower()
    alias_map = {
        'xhs': 'xhs',
        'xiaohongshu': 'xhs',
        '小红书': 'xhs',
        'red': 'xhs',
        'douyin': 'douyin',
        '抖音': 'douyin',
        'dy': 'douyin',
        'video': 'video',
        'shipinhao': 'video',
        'wechat_video': 'video',
        '视频号': 'video',
        'weibo': 'weibo',
        '微博': 'weibo',
    }
    valid_keys = {key for key, _ in PLATFORM_DEFINITIONS}
    return alias_map.get(raw, raw if raw in valid_keys else 'xhs')


CREATOR_IMPORT_HEADER_ALIASES = {
    'platform': {'platform', '平台'},
    'account_handle': {'account_handle', 'creator_account_handle', 'xhs_id', 'xhs_account', '小红书id', '小红书账号', '账号', '账号id', '小红书id/账号'},
    'display_name': {'display_name', 'creator_display_name', 'nickname', '昵称', '用户昵称', '账号昵称', '名称'},
    'owner_name': {'owner_name', '姓名', '达人姓名', '博主姓名'},
    'owner_phone': {'owner_phone', 'phone', '手机号', '联系方式', '电话', '手机号/联系方式'},
    'profile_url': {'profile_url', 'xhs_profile_link', '主页链接', '账号链接', '个人主页链接', '小红书主页链接'},
    'post_url': {'post_url', 'xhs_link', '笔记链接', '笔记url', '小红书笔记链接', '小红书链接', '链接', '作品链接'},
    'title': {'title', '笔记标题', '标题', '作品标题'},
    'publish_time': {'publish_time', '发布时间', '发布日期', '时间'},
    'views': {'views', 'view_count', '阅读量', '浏览量', '播放量', '传播量', '曝光量', '小红书传播量', '小红书浏览量', '小红书阅读量'},
    'exposures': {'exposures', '曝光', '曝光数'},
    'likes': {'likes', '点赞', '点赞量', '赞数', '小红书点赞量'},
    'favorites': {'favorites', '收藏', '收藏量', '小红书收藏量'},
    'comments': {'comments', '评论', '评论量', '小红书评论量'},
    'shares': {'shares', '分享', '分享量'},
    'follower_count': {'follower_count', '粉丝', '粉丝量', '粉丝数'},
    'post_count': {'post_count', '发文数', '笔记数'},
    'total_views': {'total_views', '总阅读', '总浏览', '总传播量'},
    'total_interactions': {'total_interactions', '总互动', '总互动量'},
    'source_channel': {'source_channel', '来源', '来源渠道', '数据来源'},
    'topic_title': {'topic_title', '话题', '话题标题'},
    'snapshot_date': {'snapshot_date', '快照日期', '统计日期', '日期'},
}


def _normalize_import_header(value=''):
    text = str(value or '').strip().lower()
    text = text.replace('（', '(').replace('）', ')')
    text = text.replace('：', ':').replace('_', '').replace('-', '').replace(' ', '')
    return text


def _resolve_import_header(value=''):
    normalized = _normalize_import_header(value)
    if not normalized:
        return ''
    for canonical, aliases in CREATOR_IMPORT_HEADER_ALIASES.items():
        normalized_aliases = {_normalize_import_header(alias) for alias in aliases}
        if normalized in normalized_aliases:
            return canonical
    return normalized


def _guess_post_title(row):
    title = (row.get('title') or '').strip()
    if title:
        return title
    post_url = canonicalize_xhs_post_url((row.get('post_url') or '').strip())
    if post_url:
        post_id = post_url.rstrip('/').split('/')[-1]
        return f'小红书笔记 {post_id}'
    account_handle = (row.get('account_handle') or '').strip()
    publish_time = (row.get('publish_time') or '').strip()
    if account_handle or publish_time:
        return f'{account_handle or "账号"} 笔记 {publish_time or ""}'.strip()
    return '小红书笔记'


def _rows_to_creator_bundle(rows):
    if not rows or len(rows[0]) < 2:
        raise ValueError('未识别到有效表头，请确认是 Excel 复制内容或标准 CSV')

    raw_headers = rows[0]
    headers = [_resolve_import_header(item) for item in raw_headers]
    accounts = []
    posts = []
    snapshots = []

    seen_account_keys = set()
    for raw_row in rows[1:]:
        if not any(str(item or '').strip() for item in raw_row):
            continue
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = str(raw_row[idx]).strip() if idx < len(raw_row) and raw_row[idx] is not None else ''

        platform = normalize_creator_platform(row.get('platform') or 'xhs')
        account_handle = (row.get('account_handle') or '').strip()
        display_name = (row.get('display_name') or '').strip()
        owner_phone = (row.get('owner_phone') or '').strip()
        profile_url = normalize_tracking_url(row.get('profile_url') or '')
        owner_name = (row.get('owner_name') or '').strip()
        if not (account_handle or display_name or profile_url or owner_phone):
            continue

        account_key = (platform, account_handle or display_name or profile_url or owner_phone)
        if account_key not in seen_account_keys:
            accounts.append({
                'platform': platform,
                'account_handle': account_handle,
                'display_name': display_name or account_handle,
                'owner_name': owner_name,
                'owner_phone': owner_phone,
                'profile_url': profile_url,
                'follower_count': row.get('follower_count') or '',
                'source_channel': row.get('source_channel') or 'spreadsheet_import',
            })
            seen_account_keys.add(account_key)

        post_url = canonicalize_xhs_post_url((row.get('post_url') or '').strip())
        if post_url:
            posts.append({
                'platform': platform,
                'account_handle': account_handle,
                'display_name': display_name,
                'owner_name': owner_name,
                'owner_phone': owner_phone,
                'profile_url': profile_url,
                'post_url': post_url,
                'title': _guess_post_title(row),
                'publish_time': row.get('publish_time') or '',
                'views': row.get('views') or '',
                'exposures': row.get('exposures') or '',
                'likes': row.get('likes') or '',
                'favorites': row.get('favorites') or '',
                'comments': row.get('comments') or '',
                'shares': row.get('shares') or '',
                'topic_title': row.get('topic_title') or '',
                'source_channel': row.get('source_channel') or 'spreadsheet_import',
            })

        if any((row.get(key) or '').strip() for key in ['follower_count', 'post_count', 'total_views', 'total_interactions']):
            snapshot_date = row.get('snapshot_date') or row.get('publish_time') or datetime.now().strftime('%Y-%m-%d')
            snapshots.append({
                'platform': platform,
                'account_handle': account_handle,
                'display_name': display_name,
                'owner_name': owner_name,
                'owner_phone': owner_phone,
                'profile_url': profile_url,
                'snapshot_date': snapshot_date,
                'follower_count': row.get('follower_count') or '',
                'post_count': row.get('post_count') or '',
                'total_views': row.get('total_views') or '',
                'total_interactions': row.get('total_interactions') or '',
                'source_channel': row.get('source_channel') or 'spreadsheet_import',
            })

    return {
        'accounts': accounts,
        'posts': posts,
        'snapshots': snapshots,
    }


def _parse_creator_import_table(raw_payload=''):
    payload = (raw_payload or '').strip()
    if not payload:
        return {
            'accounts': [],
            'posts': [],
            'snapshots': [],
        }

    lines = [line for line in payload.splitlines() if line.strip()]
    if len(lines) < 2:
        raise ValueError('表格导入至少需要表头和一行数据')

    delimiter = '\t' if '\t' in lines[0] else ','
    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)
    return _rows_to_creator_bundle(rows)


def parse_creator_import_file(filename='', content_bytes=b''):
    name = (filename or '').strip()
    if not name:
        raise ValueError('未识别到文件名')
    ext = os.path.splitext(name.lower())[1]
    if ext in {'.csv', '.tsv', '.txt'}:
        try:
            text = content_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = content_bytes.decode('gb18030')
        return _parse_creator_import_table(text)
    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError('当前环境未安装 openpyxl，暂时无法解析 xlsx 文件') from exc
        workbook = load_workbook(filename=BytesIO(content_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell).strip() if cell is not None else '' for cell in row])
        return _rows_to_creator_bundle(rows)
    raise ValueError('仅支持 csv/tsv/txt/xlsx 文件导入')


def parse_creator_import_bundle(raw_payload=''):
    payload = (raw_payload or '').strip()
    if not payload:
        return {
            'accounts': [],
            'posts': [],
            'snapshots': [],
        }

    if not payload.startswith('{') and not payload.startswith('['):
        return _parse_creator_import_table(payload)

    try:
        parsed = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise ValueError(f'导入 JSON 格式不正确：{exc}')

    if isinstance(parsed, list):
        return {
            'accounts': parsed,
            'posts': [],
            'snapshots': [],
        }

    if not isinstance(parsed, dict):
        raise ValueError('导入内容必须是 JSON 对象或数组')

    def _ensure_list(value):
        return value if isinstance(value, list) else []

    return {
        'accounts': _ensure_list(parsed.get('accounts') or parsed.get('creator_accounts')),
        'posts': _ensure_list(parsed.get('posts') or parsed.get('creator_posts')),
        'snapshots': _ensure_list(parsed.get('snapshots') or parsed.get('creator_snapshots')),
    }


def _creator_account_identity_keys(platform='', account_handle='', display_name='', owner_phone='', profile_url=''):
    keys = []
    platform = normalize_creator_platform(platform)
    handle = (account_handle or '').strip().lower()
    name = (display_name or '').strip().lower()
    phone = (owner_phone or '').strip()
    normalized_profile = normalize_tracking_url(profile_url)
    if handle:
        keys.append(f'{platform}:handle:{handle}')
    if name:
        keys.append(f'{platform}:name:{name}')
    if phone:
        keys.append(f'phone:{phone}')
    if normalized_profile:
        keys.append(f'{platform}:profile:{normalized_profile}')
    return keys


def _build_creator_account_cache():
    cache = {}
    for account in CreatorAccount.query.all():
        keys = _creator_account_identity_keys(
            account.platform,
            account.account_handle,
            account.display_name,
            account.owner_phone,
            account.profile_url,
        )
        for key in keys:
            cache[key] = account
    return cache


def _cache_creator_account(account, cache):
    keys = _creator_account_identity_keys(
        account.platform,
        account.account_handle,
        account.display_name,
        account.owner_phone,
        account.profile_url,
    )
    for key in keys:
        cache[key] = account


def _find_creator_account_for_import(row, cache):
    if not isinstance(row, dict):
        return None
    account_id = _safe_int(row.get('id') or row.get('creator_account_id'), 0)
    if account_id > 0:
        account = CreatorAccount.query.get(account_id)
        if account:
            return account
    keys = _creator_account_identity_keys(
        row.get('platform'),
        row.get('account_handle') or row.get('creator_account_handle'),
        row.get('display_name') or row.get('creator_display_name'),
        row.get('owner_phone') or row.get('phone'),
        row.get('profile_url') or row.get('xhs_profile_link'),
    )
    for key in keys:
        if key in cache:
            return cache[key]
    return None


def _upsert_creator_account_row(row, cache):
    if not isinstance(row, dict):
        return None, 'skip'
    account = _find_creator_account_for_import(row, cache)
    action = 'update' if account else 'create'
    if not account:
        account = CreatorAccount()
        db.session.add(account)

    platform = normalize_creator_platform(row.get('platform'))
    account_handle = (row.get('account_handle') or row.get('creator_account_handle') or '').strip()
    display_name = (row.get('display_name') or row.get('creator_display_name') or '').strip()
    owner_name = (row.get('owner_name') or '').strip()
    owner_phone = (row.get('owner_phone') or row.get('phone') or '').strip()
    profile_url = normalize_tracking_url(row.get('profile_url') or row.get('xhs_profile_link') or '')
    if not account_handle and not display_name:
        return None, 'skip'

    account.platform = platform
    account.owner_name = owner_name
    account.owner_phone = owner_phone
    account.account_handle = account_handle or display_name
    account.display_name = display_name or account.account_handle
    if profile_url:
        account.profile_url = profile_url
    account.follower_count = _safe_int(row.get('follower_count'), account.follower_count or 0)
    account.source_channel = (row.get('source_channel') or account.source_channel or 'import').strip()
    account.status = (row.get('status') or account.status or 'active').strip()
    account.notes = (row.get('notes') or '').strip()
    account.last_synced_at = _parse_datetime(row.get('last_synced_at')) or datetime.now()
    db.session.flush()
    _cache_creator_account(account, cache)
    return account, action


def _resolve_or_create_import_account(row, cache, created_counter):
    account = _find_creator_account_for_import(row, cache)
    if account:
        return account

    fallback_row = {
        'platform': row.get('platform'),
        'account_handle': row.get('account_handle') or row.get('creator_account_handle'),
        'display_name': row.get('display_name') or row.get('creator_display_name') or row.get('account_handle') or row.get('creator_account_handle'),
        'owner_name': row.get('owner_name') or '',
        'owner_phone': row.get('owner_phone') or row.get('phone') or '',
        'source_channel': row.get('source_channel') or 'import_placeholder',
        'notes': '导入时自动创建的占位账号',
    }
    account, action = _upsert_creator_account_row(fallback_row, cache)
    if account and action == 'create':
        created_counter['placeholder_accounts'] += 1
    return account


def _match_creator_post(account, row):
    platform_post_id = (row.get('platform_post_id') or '').strip()
    if platform_post_id:
        post = CreatorPost.query.filter_by(creator_account_id=account.id, platform_post_id=platform_post_id).first()
        if post:
            return post
    post_url = (row.get('post_url') or '').strip()
    if post_url:
        post = CreatorPost.query.filter_by(creator_account_id=account.id, post_url=post_url).first()
        if post:
            return post
    title = (row.get('title') or '').strip()
    publish_time = _parse_datetime(row.get('publish_time'))
    if title and publish_time:
        return CreatorPost.query.filter_by(creator_account_id=account.id, title=title, publish_time=publish_time).first()
    return None


def preview_creator_import_bundle(bundle):
    account_cache = _build_creator_account_cache()
    summary = Counter()
    samples = {
        'accounts': [],
        'posts': [],
        'snapshots': [],
    }

    for row in bundle.get('accounts', []):
        if not isinstance(row, dict):
            summary['accounts_skipped'] += 1
            continue
        platform = normalize_creator_platform(row.get('platform'))
        account_handle = (row.get('account_handle') or row.get('creator_account_handle') or '').strip()
        display_name = (row.get('display_name') or row.get('creator_display_name') or '').strip()
        if not account_handle and not display_name:
            summary['accounts_skipped'] += 1
            continue
        existing = _find_creator_account_for_import(row, account_cache)
        action = 'update' if existing else 'create'
        summary[f'accounts_{action}'] += 1
        if len(samples['accounts']) < 3:
            samples['accounts'].append({
                'platform': platform,
                'account_handle': account_handle or display_name,
                'display_name': display_name or account_handle,
            })

    imported_account_keys = set()
    for row in bundle.get('accounts', []):
        keys = _creator_account_identity_keys(
            row.get('platform'),
            row.get('account_handle') or row.get('creator_account_handle'),
            row.get('display_name') or row.get('creator_display_name'),
            row.get('owner_phone') or row.get('phone'),
            row.get('profile_url') or row.get('xhs_profile_link'),
        )
        for key in keys:
            imported_account_keys.add(key)

    def preview_account_exists(row):
        return _find_creator_account_for_import(row, account_cache) is not None or bool(
            imported_account_keys.intersection(_creator_account_identity_keys(
                row.get('platform'),
                row.get('account_handle') or row.get('creator_account_handle'),
                row.get('display_name') or row.get('creator_display_name'),
                row.get('owner_phone') or row.get('phone'),
                row.get('profile_url') or row.get('xhs_profile_link'),
            ))
        )

    for row in bundle.get('posts', []):
        title = (row.get('title') or '').strip()
        if not title:
            summary['posts_skipped'] += 1
            continue
        if preview_account_exists(row):
            summary['posts_ready'] += 1
        else:
            summary['posts_need_placeholder'] += 1
        if len(samples['posts']) < 3:
            samples['posts'].append({
                'title': title,
                'platform': normalize_creator_platform(row.get('platform')),
                'account_handle': (row.get('account_handle') or row.get('creator_account_handle') or '').strip(),
            })

    for row in bundle.get('snapshots', []):
        snapshot_date = _parse_date(row.get('snapshot_date'))
        if not snapshot_date:
            summary['snapshots_skipped'] += 1
            continue
        if preview_account_exists(row):
            summary['snapshots_ready'] += 1
        else:
            summary['snapshots_need_placeholder'] += 1
        if len(samples['snapshots']) < 3:
            samples['snapshots'].append({
                'snapshot_date': snapshot_date.isoformat(),
                'platform': normalize_creator_platform(row.get('platform')),
                'account_handle': (row.get('account_handle') or row.get('creator_account_handle') or '').strip(),
            })

    return {
        'summary': {
            'accounts_total': len(bundle.get('accounts', [])),
            'posts_total': len(bundle.get('posts', [])),
            'snapshots_total': len(bundle.get('snapshots', [])),
            **dict(summary),
        },
        'samples': samples,
    }


def import_creator_bundle(bundle, log_operation):
    account_cache = _build_creator_account_cache()
    summary = Counter()
    touched_account_ids = set()

    for row in bundle.get('accounts', []):
        account, action = _upsert_creator_account_row(row, account_cache)
        if not account:
            summary['accounts_skipped'] += 1
            continue
        summary[f'accounts_{action}'] += 1
        touched_account_ids.add(account.id)

    for row in bundle.get('posts', []):
        title = (row.get('title') or '').strip()
        if not title:
            summary['posts_skipped'] += 1
            continue
        account = _resolve_or_create_import_account(row, account_cache, summary)
        if not account:
            summary['posts_skipped'] += 1
            continue
        post = _match_creator_post(account, row)
        action = 'update' if post else 'create'
        if not post:
            post = CreatorPost(creator_account_id=account.id)
            db.session.add(post)

        post.platform_post_id = (row.get('platform_post_id') or '').strip()
        post.registration_id = _safe_int(row.get('registration_id'), post.registration_id)
        post.topic_id = _safe_int(row.get('topic_id'), post.topic_id)
        post.submission_id = _safe_int(row.get('submission_id'), post.submission_id)
        post.title = title
        raw_post_url = (row.get('post_url') or '').strip()
        post.post_url = canonicalize_xhs_post_url(raw_post_url) if account.platform == 'xhs' else normalize_tracking_url(raw_post_url)
        if account.platform == 'xhs' and post.post_url and not post.platform_post_id:
            post.platform_post_id = post.post_url.rstrip('/').split('/')[-1]
        post.publish_time = _parse_datetime(row.get('publish_time'))
        post.topic_title = (row.get('topic_title') or '').strip()
        post.views = _safe_int(row.get('views'))
        post.exposures = _safe_int(row.get('exposures'))
        post.likes = _safe_int(row.get('likes'))
        post.favorites = _safe_int(row.get('favorites'))
        post.comments = _safe_int(row.get('comments'))
        post.shares = _safe_int(row.get('shares'))
        post.follower_delta = _safe_int(row.get('follower_delta'))
        raw_is_viral = row.get('is_viral')
        if raw_is_viral is None:
            post.is_viral = _infer_viral_post(
                views=post.views,
                likes=post.likes,
                favorites=post.favorites,
                comments=post.comments,
                exposures=post.exposures,
            )
        else:
            post.is_viral = _coerce_bool(raw_is_viral)
        post.source_channel = (row.get('source_channel') or post.source_channel or 'import').strip()
        post.raw_payload = json.dumps(row, ensure_ascii=False)
        account.last_synced_at = datetime.now()
        db.session.flush()
        summary[f'posts_{action}'] += 1
        touched_account_ids.add(account.id)

    for row in bundle.get('snapshots', []):
        snapshot_date = _parse_date(row.get('snapshot_date'))
        if not snapshot_date:
            summary['snapshots_skipped'] += 1
            continue
        account = _resolve_or_create_import_account(row, account_cache, summary)
        if not account:
            summary['snapshots_skipped'] += 1
            continue
        snapshot = CreatorAccountSnapshot.query.filter_by(
            creator_account_id=account.id,
            snapshot_date=snapshot_date,
        ).first()
        action = 'update' if snapshot else 'create'
        if not snapshot:
            snapshot = CreatorAccountSnapshot(
                creator_account_id=account.id,
                snapshot_date=snapshot_date,
            )
            db.session.add(snapshot)

        snapshot.follower_count = _safe_int(row.get('follower_count'), account.follower_count or 0)
        snapshot.post_count = _safe_int(row.get('post_count'))
        snapshot.total_views = _safe_int(row.get('total_views'))
        snapshot.total_interactions = _safe_int(row.get('total_interactions'))
        snapshot.source_channel = (row.get('source_channel') or snapshot.source_channel or 'import').strip()
        account.follower_count = snapshot.follower_count
        account.last_synced_at = datetime.now()
        db.session.flush()
        summary[f'snapshots_{action}'] += 1
        touched_account_ids.add(account.id)

    for account_id in touched_account_ids:
        account = CreatorAccount.query.get(account_id)
        if account:
            sync_tracking_for_creator_account(account)

    log_operation('import', 'creator_bundle', message='批量导入账号看板数据', detail=dict(summary))
    db.session.commit()
    return {
        'summary': dict(summary),
    }
